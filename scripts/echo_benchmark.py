#!/usr/bin/env python3
"""Run the Echo Autosurvey blind benchmark with agent-authored decisions."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import shutil
import statistics
import subprocess
import sys
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


BLIND = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/109-2601 Echo BH.xlsx")
ANNOTATED = Path("/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260300_ECHO.xlsx")
RUN_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/agent-native-echo-2026-06-24")
INTERNAL = RUN_DIR / ".autosurvey-internal"
PUBLIC = RUN_DIR / "public"
SKILL_VERSION = "echo-authenticity-agent-native-v2026-06-24.2"

EXCLUDED_FIELDS = {
    "qc",
    "TERMFLAGS",
    "qc5",
    "qc5_Pasted",
    "LangAssessReadLevel",
    "LangAssessReadEase",
    "LangAssessNumSen",
    "LangAssessNumWords",
    "LangAssessNumSyl",
    "url",
    "session",
    "camp",
    "bhf",
    "sfh",
    "intcode",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions")

IMPORTANT_PREFIXES = (
    "record",
    "uuid",
    "date",
    "qtime",
    "ipAddress",
    "SUPNAME",
    "qGender",
    "qager1",
    "age",
    "qEthnic",
    "qEd",
    "qStateVer",
    "qEmploy",
    "qUSHHI",
    "q10",
    "q11",
    "q14",
    "q15",
    "q16",
    "q19",
    "q20",
    "q21",
    "q23",
    "q24",
    "q25",
    "q26",
    "q28",
    "q29",
    "q30",
    "q31",
    "q32",
    "q38",
    "q39",
    "q40",
    "q41",
    "q42",
    "q43",
    "q44",
    "q45",
    "POSSIBLEBRANDS",
    "TOPBRANDS",
    "AWARE",
    "CLASSIFY",
    "REGION",
)

AGENT_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "required": ["batch_id", "skill_version", "decisions"],
    "properties": {
        "batch_id": {"type": "string"},
        "skill_version": {"type": "string"},
        "decisions": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "required": [
                    "respondent_id",
                    "source_excel_row",
                    "agent_decision",
                    "binary_prediction",
                    "authenticity_risk",
                    "client_rejection_probability",
                    "confidence",
                    "semantic_decision_criteria",
                    "signals",
                    "signal_evidence",
                    "protective_evidence",
                    "key_relationships",
                    "adjudication",
                ],
                "properties": {
                    "respondent_id": {"type": "string"},
                    "source_excel_row": {"type": "integer"},
                    "agent_decision": {
                        "type": "string",
                        "enum": [
                            "Clean keep",
                            "Keep with note",
                            "Light review",
                            "Review closely",
                            "Exclude candidate",
                        ],
                    },
                    "binary_prediction": {"type": "boolean"},
                    "authenticity_risk": {
                        "type": "string",
                        "enum": ["low", "moderate", "high", "severe"],
                    },
                    "client_rejection_probability": {"type": "number", "minimum": 0, "maximum": 1},
                    "confidence": {"type": "number", "minimum": 0, "maximum": 1},
                    "semantic_decision_criteria": {"type": "string"},
                    "signals": {"type": "array", "items": {"type": "string"}, "minItems": 1},
                    "signal_evidence": {"type": "array", "items": {"type": "string"}, "minItems": 1},
                    "protective_evidence": {"type": "string"},
                    "key_relationships": {"type": "string"},
                    "adjudication": {"type": "string"},
                },
            },
        },
    },
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def norm(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def clean_text(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def is_excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(prefix) for prefix in EXCLUDED_PREFIXES)


def load_datamap(wb: Any) -> dict[str, list[str]]:
    ws = wb["Datamap"]
    out: dict[str, list[str]] = {}
    current: str | None = None
    for a, b, c in ws.iter_rows(values_only=True):
        a_s = clean_text(a)
        if a_s.startswith("[") and "]" in a_s:
            current = a_s.split("]", 1)[0][1:]
            out[current] = [a_s]
            continue
        if current and (a_s or b is not None or c is not None):
            if a_s:
                out[current].append(a_s)
            if b is not None or c is not None:
                out[current].append(f"{b} | {c}")
    return out


def value_label(field: str, value: Any, qmap: dict[str, list[str]]) -> str:
    if value is None or value == "":
        return ""
    lines = qmap.get(field, [])
    svalue = str(value)
    for line in lines:
        if " | " not in line:
            continue
        left, right = line.split(" | ", 1)
        if clean_text(left) == svalue:
            return f"{value}={right}"
    return str(value)


def field_label(field: str, qmap: dict[str, list[str]]) -> str:
    lines = qmap.get(field, [])
    if not lines:
        return field
    first = lines[0]
    if "] | " in first:
        return first.split("] | ", 1)[1]
    if "]: " in first:
        return first.split("]: ", 1)[1]
    return first


def pct_rank(values: list[float], value: float | None) -> float | None:
    if value is None or not values:
        return None
    return sum(1 for v in values if v <= value) / len(values)


def matrix_stats(vals: list[Any]) -> dict[str, Any]:
    xs = [v for v in vals if v not in (None, "")]
    if not xs:
        return {"answered": 0}
    counts = Counter(map(str, xs))
    longest = 1
    run = 1
    prev = None
    for x in map(str, xs):
        if x == prev:
            run += 1
        else:
            run = 1
            prev = x
        longest = max(longest, run)
    entropy = 0.0
    for n in counts.values():
        p = n / len(xs)
        entropy -= p * math.log2(p)
    return {
        "answered": len(xs),
        "unique": len(counts),
        "modal_value": counts.most_common(1)[0][0],
        "modal_share": round(counts.most_common(1)[0][1] / len(xs), 3),
        "longest_run": longest,
        "entropy": round(entropy, 3),
    }


def compact_group(row: dict[str, Any], qmap: dict[str, list[str]], prefix: str) -> dict[str, str]:
    out = {}
    for k, v in row.items():
        if not k.startswith(prefix) or v in (None, ""):
            continue
        lab = field_label(k, qmap)
        val = value_label(k, v, qmap)
        if val:
            out[k] = f"{lab}: {val}"
    return out


def selected_fields(row: dict[str, Any], qmap: dict[str, list[str]]) -> dict[str, str]:
    out = {}
    for k, v in row.items():
        if v in (None, "") or is_excluded(k):
            continue
        if not k.startswith(IMPORTANT_PREFIXES):
            continue
        out[k] = f"{field_label(k, qmap)}: {value_label(k, v, qmap)}"
    return out


def joined_values(row: dict[str, Any], prefix: str, limit: int = 260) -> str:
    vals = []
    for k in sorted(row):
        if k.startswith(prefix):
            v = row.get(k)
            vals.append("." if v in (None, "") else clean_text(v))
    text = ",".join(vals)
    return text if len(text) <= limit else text[:limit] + "...[truncated]"


def short_chain(row: dict[str, Any], qmap: dict[str, list[str]]) -> dict[str, Any]:
    direct = [
        "CLASSIFY",
        "CONAGE",
        "PROAGE",
        "REGION",
        "qGender",
        "qager1",
        "age",
        "qStateVer",
        "qEd",
        "qEmploy",
        "qUSHHI",
        "q9",
        "q21c_2026",
        "q25",
        "q29",
        "q30_2026",
        "q38_2026",
        "q39_2026",
        "q40_2026",
        "q41",
        "q42",
        "q43",
        "q44",
        "q45",
    ]
    out: dict[str, Any] = {}
    for k in direct:
        if row.get(k) not in (None, ""):
            out[k] = f"{field_label(k, qmap)}: {value_label(k, row.get(k), qmap)}"
    families = {
        "screening_q10": "q10",
        "equipment_owned_q11a": "q11a",
        "equipment_used_q11b": "q11b",
        "equipment_other_q11o": "q11",
        "retail_channels_q14": "q14",
        "importance_q15": "q15_2026",
        "top_brands_q16": "q16_2026",
        "brand_share_q19": "q19_2026",
        "consideration_q20": "q20",
        "supplier_used_q21b": "q21b_2026",
        "supplier_reasons_q21d": "q21d_2026",
        "possible_brands": "POSSIBLEBRANDS",
        "brand_nps_q24": "q24_2026",
        "ad_seen_q28": "q28",
        "switch_reasons_q31": "q31_2026",
        "loyalty_reasons_q32": "q32_2026",
    }
    for label, prefix in families.items():
        text = joined_values(row, prefix)
        if text and set(text) != {"."}:
            out[label] = text
    return out


def make_prepare() -> None:
    INTERNAL.mkdir(parents=True, exist_ok=True)
    PUBLIC.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(BLIND, read_only=True, data_only=True)
    ws = wb["A1"]
    qmap = load_datamap(wb)
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    qtimes: list[float] = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean_text(raw.get("uuid")) or clean_text(raw.get("record")) or f"row_{excel_row}"
        raw["_respondent_id"] = rid
        raw["_source_excel_row"] = excel_row
        rows.append(raw)
        try:
            qtimes.append(float(raw.get("qtime")))
        except (TypeError, ValueError):
            pass

    open_counts: Counter[str] = Counter()
    response_vectors: defaultdict[str, list[str]] = defaultdict(list)
    for row in rows:
        open_bits = []
        vec_bits = []
        for k, v in row.items():
            if k.startswith("_") or is_excluded(k) or v in (None, ""):
                continue
            sv = clean_text(v)
            if re.search(r"(oe$|oth|q29$|q30_2026r6oe|q40_2026r6oe|q42r6oe)", k):
                key = re.sub(r"[^a-z0-9 ]+", "", sv.lower())
                if len(key) >= 3:
                    open_counts[key] += 1
                    open_bits.append(f"{k}:{sv}")
            if k.startswith(("q10", "q11", "q14", "q15", "q16", "q19", "q20", "q23", "q24", "q25", "q26", "q28", "q31", "q32", "q41")):
                vec_bits.append(f"{k}={sv}")
        digest = hashlib.sha1("|".join(vec_bits).encode("utf-8")).hexdigest()[:16]
        response_vectors[digest].append(row["_respondent_id"])

    packets = []
    evidence_rows = []
    for row in rows:
        rid = row["_respondent_id"]
        qtime = None
        try:
            qtime = float(row.get("qtime"))
        except (TypeError, ValueError):
            pass
        opens = {}
        open_words = 0
        duplicate_open = []
        for k, v in row.items():
            if v in (None, ""):
                continue
            sv = clean_text(v)
            if re.search(r"(oe$|oth|q29$|q30_2026r6oe|q40_2026r6oe|q42r6oe)", k):
                opens[k] = f"{field_label(k, qmap)}: {sv}"
                open_words += len(re.findall(r"\w+", sv))
                key = re.sub(r"[^a-z0-9 ]+", "", sv.lower())
                if len(key) >= 3 and open_counts[key] > 1:
                    duplicate_open.append(f"{k} repeats {open_counts[key]}x: {sv}")

        matrix_prefixes = ["q11a", "q11b", "q15_2026", "q23_2026", "q24_2026", "q25_2026", "q28", "q31_2026", "q32_2026", "q41_2026"]
        matrix = {}
        for pfx in matrix_prefixes:
            vals = [row.get(k) for k in sorted(row) if k.startswith(pfx)]
            stat = matrix_stats(vals)
            if stat.get("answered"):
                matrix[pfx] = stat
        vec_bits = []
        for k, v in row.items():
            if k.startswith(("q10", "q11", "q14", "q15", "q16", "q19", "q20", "q23", "q24", "q25", "q26", "q28", "q31", "q32", "q41")) and v not in (None, ""):
                vec_bits.append(f"{k}={clean_text(v)}")
        digest = hashlib.sha1("|".join(vec_bits).encode("utf-8")).hexdigest()[:16]
        exact_vector_peers = [x for x in response_vectors[digest] if x != rid][:8]
        packet = {
            "respondent_id": rid,
            "source_excel_row": row["_source_excel_row"],
            "raw_pointer": "blind_raw_row_packets.ndjson",
            "timing": {
                "qtime_seconds": qtime,
                "qtime_percentile": None if qtime is None else round(pct_rank(qtimes, qtime), 3),
                "date": row.get("date"),
            },
            "technical": {
                "supplier": row.get("SUPNAME"),
                "ip": row.get("ipAddress"),
                "user_agent": row.get("userAgent"),
            },
            "question_context": "Echo outdoor power equipment survey. Review the persona, equipment ownership/use, brand funnel, retailer/supplier chain, advertising recall, switching/loyalty, demographics, timing, matrices, and open ends together.",
            "important_answer_chain": short_chain(row, qmap),
            "open_ends_verbatim": opens,
            "blind_evidence": {
                "open_end_word_count": open_words,
                "duplicate_open_end_candidates": duplicate_open[:10],
                "exact_response_vector_peers": exact_vector_peers,
                "matrix_statistics": matrix,
            },
        }
        packets.append(packet)
        evidence_rows.append({
            "respondent_id": rid,
            "source_excel_row": row["_source_excel_row"],
            "qtime_seconds": qtime,
            "qtime_percentile": packet["timing"]["qtime_percentile"],
            "open_end_word_count": open_words,
            "duplicate_open_end_candidate_count": len(duplicate_open),
            "exact_response_vector_peer_count": len(exact_vector_peers),
        })

    contract = {
        "source_workbook": str(BLIND),
        "source_sha256": sha256(BLIND),
        "respondent_count": len(rows),
        "skill_version": SKILL_VERSION,
        "field_count": len(headers),
        "question_contract": {
            k: {
                "question_id": k,
                "text_and_labels": v[:60],
                "response_type": "open" if any("Open " in line for line in v[:3]) else "coded_or_matrix",
                "legitimate_uniform_or_missing_reasons": "May be legitimate when routed out, when brand not known, or when a respondent has a single-brand ownership story; must be judged against the full chain.",
            }
            for k, v in qmap.items()
            if not is_excluded(k)
        },
        "relation_graph": [
            "Qualification and persona answers must be coherent with ownership, usage, property, and employment paths.",
            "Brand awareness, top brand, share, consideration, choice, NPS, association, ad exposure, switching, and loyalty are a connected funnel.",
            "Open ends explain closed answers; generic text is weaker than grounded details, but concise grounded text can be valid.",
            "Matrix uniformity is suspicious only when opposed, unrelated, or high-burden questions receive mechanical answers without protective content.",
            "Timing is judged relative to effort, complexity, open-end detail, matrix burden, and route length.",
            "Duplicate language or response vectors are population-level evidence, not automatic exclusion without semantic convergence.",
        ],
    }

    (INTERNAL / "frozen_question_contract.json").write_text(json.dumps(contract, indent=2, ensure_ascii=False))
    with (INTERNAL / "blind_corpus_evidence.csv").open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(evidence_rows[0]))
        writer.writeheader()
        writer.writerows(evidence_rows)
    with (INTERNAL / "semantic_packets.ndjson").open("w") as f:
        for packet in packets:
            f.write(json.dumps(packet, ensure_ascii=False, separators=(",", ":")) + "\n")
    (INTERNAL / "agent_output_schema.json").write_text(json.dumps(AGENT_SCHEMA, indent=2))
    prompt = f"""You are Autosurvey's agent-native blind authenticity reviewer for the Echo survey. Labels are sealed and unavailable. You receive compact packets built only from the blind workbook plus blind corpus evidence. You must personally decide whether each respondent is likely authentic or should be excluded as fabricated, bot-like, non-authentic, or survey-gaming behavior.

Skill version: {SKILL_VERSION}

Allowed decisions: Clean keep, Keep with note, Light review, Review closely, Exclude candidate. Only Exclude candidate maps to binary_prediction true.

Use three perspectives for every row: forensic investigator, human advocate, and evidence judge. Consider question responsiveness, open-end grounding, prompt drift, survey-meta substitution, persona fit, category knowledge, brand-funnel consistency, routing coherence, matrix behavior, timing relative to effort, duplicate/templated behavior, cross-question contradictions, and protective human explanations. A single weak cue cannot produce Exclude candidate. Concise human answers, spelling errors, non-native wording, a legitimate one-brand story, or valid don't-know behavior must be protected when the full chain supports them.

Return exactly one schema-valid batch object. Keep rationales concise, specific, and cited to actual packet evidence. Do not mention client labels.
"""
    (INTERNAL / "agent_prompt_prefix.txt").write_text(prompt)
    make_batches()
    manifest = {
        "state": "BLIND_CONTEXT_FROZEN",
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "annotated_counterpart_opened": False,
        "source_sha256": sha256(BLIND),
        "question_contract_sha256": sha256(INTERNAL / "frozen_question_contract.json"),
        "semantic_packets_sha256": sha256(INTERNAL / "semantic_packets.ndjson"),
        "evidence_sha256": sha256(INTERNAL / "blind_corpus_evidence.csv"),
    }
    (INTERNAL / "blind_context_manifest.json").write_text(json.dumps(manifest, indent=2))
    print(json.dumps({"prepared": len(rows), **manifest}, indent=2))


def make_batches() -> None:
    packets = [json.loads(line) for line in (INTERNAL / "semantic_packets.ndjson").read_text().splitlines()]
    batch_dir = INTERNAL / "semantic_batches"
    if batch_dir.exists():
        shutil.rmtree(batch_dir)
    batch_dir.mkdir(parents=True)
    batches = []
    current = []
    current_chars = 0
    max_chars = 80000
    for packet in packets:
        text = json.dumps(packet, ensure_ascii=False, separators=(",", ":"))
        size = len(text)
        if current and (len(current) >= 16 or current_chars + size > max_chars):
            batches.append(current)
            current = []
            current_chars = 0
        current.append(packet)
        current_chars += size
    if current:
        batches.append(current)
    manifest = []
    for i, batch in enumerate(batches, 1):
        batch_id = f"echo_batch_{i:03d}"
        path = batch_dir / f"{batch_id}.jsonl"
        with path.open("w") as f:
            for packet in batch:
                f.write(json.dumps(packet, ensure_ascii=False, separators=(",", ":")) + "\n")
        manifest.append({
            "batch_id": batch_id,
            "path": str(path),
            "state": "PENDING",
            "respondent_ids": [p["respondent_id"] for p in batch],
            "source_excel_rows": [p["source_excel_row"] for p in batch],
            "respondent_count": len(batch),
            "attempts": 0,
        })
    (INTERNAL / "batch_manifest.json").write_text(json.dumps(manifest, indent=2))


def validate_batch(batch: dict[str, Any], out: Path) -> tuple[bool, str]:
    try:
        obj = json.loads(out.read_text())
    except Exception as e:
        return False, f"json parse failed: {e}"
    if obj.get("batch_id") != batch["batch_id"]:
        return False, "batch_id mismatch"
    decisions = obj.get("decisions")
    if not isinstance(decisions, list):
        return False, "missing decisions"
    expected = set(batch["respondent_ids"])
    got = [d.get("respondent_id") for d in decisions]
    if set(got) != expected or len(got) != len(set(got)):
        return False, f"coverage mismatch expected={len(expected)} got={len(set(got))}"
    required = AGENT_SCHEMA["properties"]["decisions"]["items"]["required"]
    for d in decisions:
        for key in required:
            if d.get(key) in (None, "", []):
                return False, f"missing {key} for {d.get('respondent_id')}"
        if d["agent_decision"] == "Exclude candidate" and d["binary_prediction"] is not True:
            return False, f"exclude not binary true {d.get('respondent_id')}"
        if d["agent_decision"] != "Exclude candidate" and d["binary_prediction"] is not False:
            return False, f"non-exclude binary true {d.get('respondent_id')}"
    return True, "ok"


def run_one(batch: dict[str, Any], model: str, timeout_s: int) -> dict[str, Any]:
    out_dir = INTERNAL / "semantic_decisions"
    log_dir = INTERNAL / "semantic_logs"
    out_dir.mkdir(exist_ok=True)
    log_dir.mkdir(exist_ok=True)
    out = out_dir / f"{batch['batch_id']}.json"
    if out.exists():
        ok, msg = validate_batch(batch, out)
        if ok:
            return {"batch_id": batch["batch_id"], "state": "COMPLETE", "message": msg}
        out.rename(out.with_suffix(".invalid.json"))
    prompt = (INTERNAL / "agent_prompt_prefix.txt").read_text()
    payload = prompt + "\nBATCH_ID: " + batch["batch_id"] + "\nRAW PACKETS:\n" + Path(batch["path"]).read_text()
    tmp = out.with_suffix(".tmp.json")
    log = log_dir / f"{batch['batch_id']}.log"
    cmd = [
        "codex",
        "exec",
        "--ignore-user-config",
        "--ignore-rules",
        "--skip-git-repo-check",
        "--ephemeral",
        "--sandbox",
        "read-only",
        "--output-schema",
        str(INTERNAL / "agent_output_schema.json"),
        "-o",
        str(tmp),
        "-",
    ]
    if model:
        cmd[2:2] = ["-m", model]
    with log.open("a") as lf:
        lf.write(f"\n--- attempt {batch.get('attempts', 0) + 1} {time.strftime('%Y-%m-%dT%H:%M:%S')} model={model or 'default'} ---\n")
        try:
            proc = subprocess.run(
                cmd,
                input=payload,
                text=True,
                stdout=lf,
                stderr=subprocess.STDOUT,
                timeout=timeout_s,
                cwd=str(Path.home()),
            )
        except subprocess.TimeoutExpired:
            return {"batch_id": batch["batch_id"], "state": "RETRY", "message": "timeout"}
        if proc.returncode != 0:
            return {"batch_id": batch["batch_id"], "state": "RETRY", "message": f"return code {proc.returncode}"}
    if not tmp.exists():
        return {"batch_id": batch["batch_id"], "state": "FAILED_SCHEMA", "message": "no output file"}
    tmp.replace(out)
    ok, msg = validate_batch(batch, out)
    if ok:
        return {"batch_id": batch["batch_id"], "state": "COMPLETE", "message": msg}
    return {"batch_id": batch["batch_id"], "state": "FAILED_SCHEMA", "message": msg}


def run_queue(workers: int, model: str, timeout_s: int, max_batches: int | None = None) -> None:
    manifest_path = INTERNAL / "batch_manifest.json"
    manifest = json.loads(manifest_path.read_text())
    for batch in manifest:
        out = INTERNAL / "semantic_decisions" / f"{batch['batch_id']}.json"
        if out.exists():
            ok, msg = validate_batch(batch, out)
            batch["state"] = "COMPLETE" if ok else "FAILED_SCHEMA"
            batch["message"] = msg
    pending = [b for b in manifest if b["state"] != "COMPLETE"]
    if max_batches:
        pending = pending[:max_batches]
    print(f"pending={len(pending)} complete={sum(1 for b in manifest if b['state']=='COMPLETE')} workers={workers}")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(run_one, batch, model, timeout_s): batch for batch in pending}
        for future in as_completed(futures):
            result = future.result()
            for batch in manifest:
                if batch["batch_id"] == result["batch_id"]:
                    batch["state"] = result["state"]
                    batch["message"] = result["message"]
                    batch["attempts"] = int(batch.get("attempts", 0)) + 1
            manifest_path.write_text(json.dumps(manifest, indent=2))
            complete = sum(1 for b in manifest if b["state"] == "COMPLETE")
            print(json.dumps({"result": result, "complete_batches": complete, "total_batches": len(manifest)}), flush=True)


def make_adjudication() -> None:
    packets = {json.loads(line)["respondent_id"]: json.loads(line) for line in (INTERNAL / "semantic_packets.ndjson").read_text().splitlines()}
    first = [json.loads(line) for line in (INTERNAL / "sealed_blind_prediction_ledger.jsonl").read_text().splitlines()]
    review_ids = set()
    for d in first:
        rid = d["respondent_id"]
        if d["agent_decision"] in {"Exclude candidate", "Review closely"}:
            review_ids.add(rid)
        if float(d["confidence"]) <= 0.65:
            review_ids.add(rid)
    lower = [d for d in first if d["respondent_id"] not in review_ids]
    for d in lower:
        h = int(hashlib.sha1(d["respondent_id"].encode()).hexdigest(), 16)
        if h % 10 == 0:
            review_ids.add(d["respondent_id"])
    batch_dir = INTERNAL / "adjudication_batches"
    decision_dir = INTERNAL / "adjudication_decisions"
    log_dir = INTERNAL / "adjudication_logs"
    for d in (batch_dir, decision_dir, log_dir):
        if d.exists():
            shutil.rmtree(d)
        d.mkdir(parents=True)
    records = []
    for d in first:
        rid = d["respondent_id"]
        if rid in review_ids:
            records.append({"packet": packets[rid], "first_pass_decision": d})
    batches = []
    cur = []
    chars = 0
    for rec in records:
        text = json.dumps(rec, ensure_ascii=False, separators=(",", ":"))
        if cur and (len(cur) >= 10 or chars + len(text) > 70000):
            batches.append(cur)
            cur = []
            chars = 0
        cur.append(rec)
        chars += len(text)
    if cur:
        batches.append(cur)
    manifest = []
    for i, batch in enumerate(batches, 1):
        batch_id = f"echo_adjudication_{i:03d}"
        path = batch_dir / f"{batch_id}.jsonl"
        with path.open("w") as f:
            for rec in batch:
                f.write(json.dumps(rec, ensure_ascii=False, separators=(",", ":")) + "\n")
        manifest.append({
            "batch_id": batch_id,
            "path": str(path),
            "state": "PENDING",
            "respondent_ids": [r["packet"]["respondent_id"] for r in batch],
            "respondent_count": len(batch),
            "attempts": 0,
        })
    prompt = f"""You are Autosurvey's second blind adjudicator. Labels are sealed and unavailable. You receive each respondent packet and the first reviewer's decision. Re-read the respondent independently through the forensic investigator, human advocate, and evidence judge perspectives. You may affirm or revise the decision before sealing.

Skill version: {SKILL_VERSION}

Allowed decisions: Clean keep, Keep with note, Light review, Review closely, Exclude candidate. Only Exclude candidate maps to binary_prediction true.

Audit especially: whether suspicious cues are independent or merely correlated, whether open ends are grounded in the respondent's chain, whether timing is plausible given complexity, whether a one-brand or concise respondent deserves protection, whether duplicate/templated evidence is meaningful, and whether Exclude candidate has convergence across independent evidence families. Keep the final rationale concise and specific.

Return exactly one schema-valid batch object with batch_id and decisions. The decision fields are the final blind adjudicated fields.
"""
    (INTERNAL / "adjudicator_prompt_prefix.txt").write_text(prompt)
    (INTERNAL / "adjudication_manifest.json").write_text(json.dumps(manifest, indent=2))
    print(json.dumps({"adjudication_rows": len(records), "batches": len(batches)}, indent=2))


def run_adjudication_one(batch: dict[str, Any], model: str, timeout_s: int) -> dict[str, Any]:
    out_dir = INTERNAL / "adjudication_decisions"
    log_dir = INTERNAL / "adjudication_logs"
    out = out_dir / f"{batch['batch_id']}.json"
    if out.exists():
        ok, msg = validate_batch(batch, out)
        if ok:
            return {"batch_id": batch["batch_id"], "state": "COMPLETE", "message": msg}
        out.rename(out.with_suffix(".invalid.json"))
    payload = (INTERNAL / "adjudicator_prompt_prefix.txt").read_text()
    payload += "\nBATCH_ID: " + batch["batch_id"] + "\nADJUDICATION PACKETS:\n" + Path(batch["path"]).read_text()
    tmp = out.with_suffix(".tmp.json")
    log = log_dir / f"{batch['batch_id']}.log"
    cmd = [
        "codex",
        "exec",
        "--ignore-user-config",
        "--ignore-rules",
        "--skip-git-repo-check",
        "--ephemeral",
        "--sandbox",
        "read-only",
        "--output-schema",
        str(INTERNAL / "agent_output_schema.json"),
        "-o",
        str(tmp),
        "-",
    ]
    if model:
        cmd[2:2] = ["-m", model]
    with log.open("a") as lf:
        lf.write(f"\n--- adjudication attempt {batch.get('attempts', 0) + 1} {time.strftime('%Y-%m-%dT%H:%M:%S')} model={model or 'default'} ---\n")
        try:
            proc = subprocess.run(cmd, input=payload, text=True, stdout=lf, stderr=subprocess.STDOUT, timeout=timeout_s, cwd=str(Path.home()))
        except subprocess.TimeoutExpired:
            return {"batch_id": batch["batch_id"], "state": "RETRY", "message": "timeout"}
        if proc.returncode != 0:
            return {"batch_id": batch["batch_id"], "state": "RETRY", "message": f"return code {proc.returncode}"}
    if not tmp.exists():
        return {"batch_id": batch["batch_id"], "state": "FAILED_SCHEMA", "message": "no output file"}
    tmp.replace(out)
    ok, msg = validate_batch(batch, out)
    if ok:
        return {"batch_id": batch["batch_id"], "state": "COMPLETE", "message": msg}
    return {"batch_id": batch["batch_id"], "state": "FAILED_SCHEMA", "message": msg}


def run_adjudication_queue(workers: int, model: str, timeout_s: int) -> None:
    manifest_path = INTERNAL / "adjudication_manifest.json"
    manifest = json.loads(manifest_path.read_text())
    for batch in manifest:
        out = INTERNAL / "adjudication_decisions" / f"{batch['batch_id']}.json"
        if out.exists():
            ok, msg = validate_batch(batch, out)
            batch["state"] = "COMPLETE" if ok else "FAILED_SCHEMA"
            batch["message"] = msg
    pending = [b for b in manifest if b["state"] != "COMPLETE"]
    print(f"adjudication pending={len(pending)} complete={sum(1 for b in manifest if b['state']=='COMPLETE')} workers={workers}")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(run_adjudication_one, batch, model, timeout_s): batch for batch in pending}
        for future in as_completed(futures):
            result = future.result()
            for batch in manifest:
                if batch["batch_id"] == result["batch_id"]:
                    batch["state"] = result["state"]
                    batch["message"] = result["message"]
                    batch["attempts"] = int(batch.get("attempts", 0)) + 1
            manifest_path.write_text(json.dumps(manifest, indent=2))
            complete = sum(1 for b in manifest if b["state"] == "COMPLETE")
            print(json.dumps({"result": result, "complete_batches": complete, "total_batches": len(manifest)}), flush=True)


def apply_adjudication() -> None:
    first = [json.loads(line) for line in (INTERNAL / "sealed_blind_prediction_ledger.jsonl").read_text().splitlines()]
    final = {d["respondent_id"]: d for d in first}
    manifest = json.loads((INTERNAL / "adjudication_manifest.json").read_text())
    adjudicated: dict[str, dict[str, Any]] = {}
    for batch in manifest:
        out = INTERNAL / "adjudication_decisions" / f"{batch['batch_id']}.json"
        ok, msg = validate_batch(batch, out) if out.exists() else (False, "missing")
        if not ok:
            raise SystemExit(f"Adjudication batch {batch['batch_id']} invalid: {msg}")
        for d in json.loads(out.read_text())["decisions"]:
            adjudicated[d["respondent_id"]] = d
    for rid, d in adjudicated.items():
        original = final[rid]
        final[rid] = {
            **d,
            "skill_version": SKILL_VERSION,
            "second_reviewed": True,
            "first_pass_agent_decision": original["agent_decision"],
            "first_pass_binary_prediction": original["binary_prediction"],
            "first_pass_authenticity_risk": original["authenticity_risk"],
            "first_pass_client_rejection_probability": original["client_rejection_probability"],
            "first_pass_confidence": original["confidence"],
            "first_pass_semantic_decision_criteria": original["semantic_decision_criteria"],
        }
    for rid, d in list(final.items()):
        if "second_reviewed" not in d:
            d["second_reviewed"] = False
    packets = {json.loads(line)["respondent_id"]: json.loads(line) for line in (INTERNAL / "semantic_packets.ndjson").read_text().splitlines()}
    ledger = INTERNAL / "sealed_blind_prediction_ledger.jsonl"
    with ledger.open("w") as f:
        for rid, d in sorted(final.items(), key=lambda item: packets[item[0]]["source_excel_row"]):
            f.write(json.dumps(d, ensure_ascii=False, separators=(",", ":")) + "\n")
    print(json.dumps({"adjudicated_rows": len(adjudicated), "ledger_sha256": sha256(ledger)}, indent=2))


def assemble(seal: bool = False) -> None:
    manifest = json.loads((INTERNAL / "batch_manifest.json").read_text())
    packets = [json.loads(line) for line in (INTERNAL / "semantic_packets.ndjson").read_text().splitlines()]
    expected = {p["respondent_id"]: p for p in packets}
    decisions = []
    for batch in manifest:
        out = INTERNAL / "semantic_decisions" / f"{batch['batch_id']}.json"
        ok, msg = validate_batch(batch, out) if out.exists() else (False, "missing")
        if not ok:
            raise SystemExit(f"Batch {batch['batch_id']} invalid: {msg}")
        obj = json.loads(out.read_text())
        decisions.extend(obj["decisions"])
    got = [d["respondent_id"] for d in decisions]
    if set(got) != set(expected) or len(got) != len(set(got)):
        raise SystemExit(f"Ledger coverage failed expected={len(expected)} got={len(set(got))}")
    ledger = INTERNAL / "sealed_blind_prediction_ledger.jsonl"
    with ledger.open("w") as f:
        for d in sorted(decisions, key=lambda x: expected[x["respondent_id"]]["source_excel_row"]):
            rec = {"skill_version": SKILL_VERSION, **d}
            f.write(json.dumps(rec, ensure_ascii=False, separators=(",", ":")) + "\n")
    state = "SEALED_PENDING_UNBLIND" if seal else "BLIND_LEDGER_ASSEMBLED"
    seal_manifest = {
        "state": state,
        "sealed_at": time.strftime("%Y-%m-%dT%H:%M:%S") if seal else None,
        "annotated_counterpart_opened": False,
        "rows": len(decisions),
        "source_sha256": sha256(BLIND),
        "question_contract_sha256": sha256(INTERNAL / "frozen_question_contract.json"),
        "semantic_skill_version": SKILL_VERSION,
        "schema_sha256": sha256(INTERNAL / "agent_output_schema.json"),
        "prompt_sha256": sha256(INTERNAL / "agent_prompt_prefix.txt"),
        "batch_manifest_sha256": sha256(INTERNAL / "batch_manifest.json"),
        "ledger_sha256": sha256(ledger),
    }
    (INTERNAL / "seal_manifest.json").write_text(json.dumps(seal_manifest, indent=2))
    print(json.dumps(seal_manifest, indent=2))


def load_annotated_labels() -> dict[str, int]:
    wb = openpyxl.load_workbook(ANNOTATED, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    lower = {h.lower(): h for h in headers}
    uuid_col = lower.get("uuid") or lower.get("participant identifier")
    status_col = lower.get("status")
    if not uuid_col or not status_col:
        raise SystemExit(f"Could not find uuid/status in annotated headers: {headers[:30]}")
    labels = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean_text(row.get(uuid_col))
        if not rid:
            continue
        try:
            labels[rid] = int(row.get(status_col))
        except (TypeError, ValueError):
            pass
    return labels


def metrics() -> None:
    seal_manifest = json.loads((INTERNAL / "seal_manifest.json").read_text())
    if seal_manifest.get("state") != "SEALED_PENDING_UNBLIND":
        raise SystemExit("Refusing to unblind before SEALED_PENDING_UNBLIND")
    labels = load_annotated_labels()
    ledger = [json.loads(line) for line in (INTERNAL / "sealed_blind_prediction_ledger.jsonl").read_text().splitlines()]
    missing = [d["respondent_id"] for d in ledger if d["respondent_id"] not in labels]
    if missing:
        raise SystemExit(f"Label reconciliation failed for {len(missing)} IDs; first={missing[:5]}")
    rows = []
    for d in ledger:
        status = labels[d["respondent_id"]]
        actual = status == 5
        pred = bool(d["binary_prediction"])
        if pred and actual:
            et = "TP"
        elif pred and not actual:
            et = "FP"
        elif not pred and actual:
            et = "FN"
        else:
            et = "TN"
        rows.append({**d, "client_status": status, "actual_rejected": actual, "error_type": et})
    tp = sum(r["error_type"] == "TP" for r in rows)
    fp = sum(r["error_type"] == "FP" for r in rows)
    tn = sum(r["error_type"] == "TN" for r in rows)
    fn = sum(r["error_type"] == "FN" for r in rows)
    n = len(rows)
    def div(a: float, b: float) -> float:
        return a / b if b else 0.0
    precision = div(tp, tp + fp)
    recall = div(tp, tp + fn)
    specificity = div(tn, tn + fp)
    npv = div(tn, tn + fn)
    accuracy = div(tp + tn, n)
    f1 = div(2 * precision * recall, precision + recall)
    neg_f1 = div(2 * npv * specificity, npv + specificity)
    bal = (recall + specificity) / 2
    mcc_den = math.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    mcc = div(tp * tn - fp * fn, mcc_den)
    po = accuracy
    yes_pred = div((tp + fp), n)
    yes_actual = div((tp + fn), n)
    no_pred = div((tn + fn), n)
    no_actual = div((tn + fp), n)
    pe = yes_pred * yes_actual + no_pred * no_actual
    kappa = div(po - pe, 1 - pe)
    brier = statistics.mean((float(r["client_rejection_probability"]) - float(r["actual_rejected"])) ** 2 for r in rows)
    scores = sorted((float(r["client_rejection_probability"]), r["actual_rejected"]) for r in rows)
    pos = sum(y for _, y in scores)
    neg = len(scores) - pos
    rank_sum = sum(i + 1 for i, (_, y) in enumerate(scores) if y)
    auroc = div(rank_sum - pos * (pos + 1) / 2, pos * neg)
    desc = sorted(scores, reverse=True)
    ap = 0.0
    hit = 0
    for i, (_, y) in enumerate(desc, 1):
        if y:
            hit += 1
            ap += hit / i
    auprc = div(ap, pos)
    out = {
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
        "accuracy": accuracy,
        "balanced_accuracy": bal,
        "rejected_precision": precision,
        "rejected_recall": recall,
        "specificity": specificity,
        "NPV": npv,
        "F1_rejected": f1,
        "F1_accepted": neg_f1,
        "macro_F1": (f1 + neg_f1) / 2,
        "MCC": mcc,
        "cohen_kappa": kappa,
        "AUROC": auroc,
        "AUPRC": auprc,
        "brier_score": brier,
        "review_burden": div(sum(r["agent_decision"] in {"Light review", "Review closely", "Exclude candidate"} for r in rows), n),
        "exclusion_burden": div(sum(r["agent_decision"] == "Exclude candidate" for r in rows), n),
        "tier_volumes": dict(Counter(r["agent_decision"] for r in rows)),
        "seal": seal_manifest,
    }
    (INTERNAL / "metrics.json").write_text(json.dumps(out, indent=2))
    with (INTERNAL / "evaluated_rows.jsonl").open("w") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False, separators=(",", ":")) + "\n")
    print(json.dumps(out, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="cmd", required=True)
    sub.add_parser("prepare")
    run = sub.add_parser("run-queue")
    run.add_argument("--workers", type=int, default=6)
    run.add_argument("--model", default="")
    run.add_argument("--timeout", type=int, default=900)
    run.add_argument("--max-batches", type=int)
    sub.add_parser("make-adjudication")
    arun = sub.add_parser("run-adjudication")
    arun.add_argument("--workers", type=int, default=6)
    arun.add_argument("--model", default="")
    arun.add_argument("--timeout", type=int, default=900)
    sub.add_parser("apply-adjudication")
    assemble_p = sub.add_parser("assemble")
    assemble_p.add_argument("--seal", action="store_true")
    sub.add_parser("metrics")
    args = parser.parse_args()
    if args.cmd == "prepare":
        make_prepare()
    elif args.cmd == "run-queue":
        run_queue(args.workers, args.model, args.timeout, args.max_batches)
    elif args.cmd == "make-adjudication":
        make_adjudication()
    elif args.cmd == "run-adjudication":
        run_adjudication_queue(args.workers, args.model, args.timeout)
    elif args.cmd == "apply-adjudication":
        apply_adjudication()
    elif args.cmd == "assemble":
        assemble(args.seal)
    elif args.cmd == "metrics":
        metrics()


if __name__ == "__main__":
    main()
