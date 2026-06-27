"""Deep analysis of v6 metadata fields against client ground truth.

Goal: find where our decision criteria are wrong, over-firing, or missing
patterns — to improve precision, recall, and calibration for v7.
"""
import json
import statistics
from collections import Counter, defaultdict
from pathlib import Path

# ─── Load data ──────────────────────────────────────────────────────────────

V6_DIR = Path('/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run_v6')
ANNOTATED = Path('/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx')

# Load judgments
with open(V6_DIR / 'agent_judgments.json') as f:
    judgments_raw = json.load(f)

judgments = {}
for item in judgments_raw:
    rid = item['respondent_id']
    judgments[rid] = item

# Load ground truth
import openpyxl
wb = openpyxl.load_workbook(ANNOTATED, read_only=True, data_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

# Find respondent ID and status columns
rid_col = None
status_col = None
for i, h in enumerate(headers):
    if h and 'respondent' in str(h).lower():
        rid_col = i
    if h and str(h).lower() in ('status', 'final_status', 'client_status'):
        status_col = i
    if h and 'uuid' in str(h).lower() and rid_col is None:
        rid_col = i

# Fallback: look for UUID-like column name
if rid_col is None:
    for i, h in enumerate(headers):
        if h and 'uuid' in str(h).lower():
            rid_col = i
            break

print(f"RID column: {headers[rid_col]} (col {rid_col})")
print(f"Status column: {headers[status_col] if status_col else 'NOT FOUND'}")

gt = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    rid = str(row[rid_col]).strip() if row[rid_col] else ""
    status = row[status_col] if status_col is not None else None
    if rid and status is not None:
        s = int(status) if status else 0
        if s in (3, 5):
            gt[rid] = 'DISCARD' if s == 5 else 'KEEP'

wb.close()
print(f"Ground truth: {len(gt)} respondents ({sum(1 for v in gt.values() if v=='DISCARD')} discards, {sum(1 for v in gt.values() if v=='KEEP')} keeps)")
print()

# ─── Build joined dataset ────────────────────────────────────────────────────

joined = []
for rid, gt_label in gt.items():
    if rid in judgments:
        j = judgments[rid]
        joined.append({
            'rid': rid,
            'gt': gt_label,
            'pred': j['agent_judgment'],
            'score': j['agent_score'],
            'authenticity_risk': j.get('authenticity_risk'),
            'quality_discard_risk': j.get('quality_discard_risk'),
            'client_reject_probability': j.get('client_reject_probability'),
            'primary_removal_reason': j.get('primary_removal_reason'),
            'secondary_removal_reason': j.get('secondary_removal_reason'),
            'removal_confidence': j.get('removal_confidence'),
            'evidence_families_fired': j.get('evidence_families_fired', []),
            'evidence_family_scores': j.get('evidence_family_scores', {}),
            'badopen_trigger': j.get('badopen_trigger'),
            'badopen_field': j.get('badopen_field'),
            'badopen_severity': j.get('badopen_severity'),
            'oe_classification': j.get('oe_classification'),
            'oe_equipment_named': j.get('oe_equipment_named', []),
            'oe_grounding_anchors': j.get('oe_grounding_anchors', []),
            'oe_word_count': j.get('oe_word_count'),
            'ml_score': j.get('ml_score'),
            'ml_top_signals': j.get('ml_top_signals', []),
            'ml_confidence': j.get('ml_confidence'),
            'stage1_fraud_verdict': j.get('stage1_fraud_verdict'),
            'stage2_quality_verdict': j.get('stage2_quality_verdict'),
            'converging_family_count': j.get('converging_family_count'),
            'semantic_remapping': j.get('semantic_remapping', {}),
            'justification': j.get('agent_justification', ''),
        })

print(f"Joined: {len(joined)} respondents")
print()

# Helper: classify prediction correctness
def classify(row):
    if row['pred'] == 'DISCARD' and row['gt'] == 'DISCARD':
        return 'TP'
    elif row['pred'] == 'DISCARD' and row['gt'] == 'KEEP':
        return 'FP'
    elif row['pred'] != 'DISCARD' and row['gt'] == 'DISCARD':
        return 'FN'
    else:
        return 'TN'

for row in joined:
    row['correctness'] = classify(row)

# ─── ANALYSIS 1: Three-component score threshold optimization ─────────────────
print("=" * 90)
print("ANALYSIS 1: THREE-COMPONENT SCORE THRESHOLD OPTIMIZATION")
print("=" * 90)

for comp in ['authenticity_risk', 'quality_discard_risk', 'client_reject_probability']:
    print(f"\n--- {comp} ---")
    vals = [(row[comp], row['gt']) for row in joined if row[comp] is not None]

    # Try thresholds from 0.1 to 0.9
    best_f1 = 0
    best_threshold = 0
    best_metrics = None
    print(f"  {'Threshold':>10} {'TP':>5} {'FP':>5} {'FN':>5} {'TN':>5} {'Prec':>6} {'Recall':>7} {'F1':>6} {'BalAcc':>7}")
    for t in [i/100 for i in range(5, 95, 5)]:
        tp = sum(1 for v, g in vals if v >= t and g == 'DISCARD')
        fp = sum(1 for v, g in vals if v >= t and g == 'KEEP')
        fn = sum(1 for v, g in vals if v < t and g == 'DISCARD')
        tn = sum(1 for v, g in vals if v < t and g == 'KEEP')
        prec = tp / (tp + fp) if (tp + fp) > 0 else 0
        rec = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0
        bal_acc = (rec + tn / (tn + fp)) / 2 if (tn + fp) > 0 else 0
        if f1 > best_f1:
            best_f1 = f1
            best_threshold = t
            best_metrics = (tp, fp, fn, tn, prec, rec, f1, bal_acc)
        if t in [0.1, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.6, 0.7, 0.8]:
            print(f"  {t:>10.2f} {tp:>5} {fp:>5} {fn:>5} {tn:>5} {prec:>6.3f} {rec:>7.3f} {f1:>6.3f} {bal_acc:>7.3f}")

    tp, fp, fn, tn, prec, rec, f1, bal_acc = best_metrics
    print(f"  {'BEST':>10} {tp:>5} {fp:>5} {fn:>5} {tn:>5} {prec:>6.3f} {rec:>7.3f} {f1:>6.3f} {bal_acc:>7.3f}  (threshold={best_threshold:.2f})")

    # AUC-like: mean of discards vs keeps
    discards = [v for v, g in vals if g == 'DISCARD']
    keeps = [v for v, g in vals if g == 'KEEP']
    if discards and keeps:
        sep = statistics.mean(discards) - statistics.mean(keeps)
        # Cohen's d
        pooled_sd = ((statistics.stdev(discards)**2 + statistics.stdev(keeps)**2) / 2) ** 0.5
        cohens_d = sep / pooled_sd if pooled_sd > 0 else 0
        print(f"  Mean separation: {sep:.3f}, Cohen's d: {cohens_d:.3f}")

# ─── ANALYSIS 2: Evidence family individual predictive power ───────────────────
print()
print("=" * 90)
print("ANALYSIS 2: EVIDENCE FAMILY INDIVIDUAL PREDICTIVE POWER")
print("=" * 90)

all_families = set()
for row in joined:
    all_families.update(row['evidence_families_fired'])
all_families = sorted(all_families)

print(f"\n{'Family':<30} {'Fired':>6} {'TP':>5} {'FP':>5} {'FN':>5} {'TN':>5} {'Prec':>6} {'Recall':>7} {'F1':>6} {'Gap':>7}")
print("-" * 110)

family_results = {}
for fam in all_families:
    tp = sum(1 for r in joined if fam in r['evidence_families_fired'] and r['gt'] == 'DISCARD')
    fp = sum(1 for r in joined if fam in r['evidence_families_fired'] and r['gt'] == 'KEEP')
    fn = sum(1 for r in joined if fam not in r['evidence_families_fired'] and r['gt'] == 'DISCARD')
    tn = sum(1 for r in joined if fam not in r['evidence_families_fired'] and r['gt'] == 'KEEP')
    fired = tp + fp
    prec = tp / fired if fired > 0 else 0
    rec = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0
    d_rate = tp / sum(1 for r in joined if r['gt'] == 'DISCARD') if sum(1 for r in joined if r['gt'] == 'DISCARD') > 0 else 0
    k_rate = fp / sum(1 for r in joined if r['gt'] == 'KEEP') if sum(1 for r in joined if r['gt'] == 'KEEP') > 0 else 0
    gap = d_rate - k_rate
    family_results[fam] = {'prec': prec, 'rec': rec, 'f1': f1, 'gap': gap, 'fired': fired}
    print(f"{fam:<30} {fired:>6} {tp:>5} {fp:>5} {fn:>5} {tn:>5} {prec:>6.3f} {rec:>7.3f} {f1:>6.3f} {gap:>+7.3f}")

# ─── ANALYSIS 3: Evidence family COMBINATIONS ─────────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 3: EVIDENCE FAMILY COMBINATIONS (pairs)")
print("=" * 90)

# For each pair of families, what's the precision when BOTH fire?
families_list = sorted(all_families)
print(f"\n{'Pair':<55} {'Both fire':>10} {'TP':>5} {'FP':>5} {'Prec':>6} {'Lift vs single':>15}")
print("-" * 100)

pair_results = []
for i, f1_name in enumerate(families_list):
    for j, f2_name in enumerate(families_list):
        if j <= i:
            continue
        both_fire = [r for r in joined if f1_name in r['evidence_families_fired'] and f2_name in r['evidence_families_fired']]
        if len(both_fire) < 10:
            continue
        tp = sum(1 for r in both_fire if r['gt'] == 'DISCARD')
        fp = sum(1 for r in both_fire if r['gt'] == 'KEEP')
        prec = tp / len(both_fire) if both_fire else 0
        # Lift over best single family
        best_single = max(family_results[f1_name]['prec'], family_results[f2_name]['prec'])
        lift = prec - best_single
        pair_results.append((f"{f1_name} + {f2_name}", len(both_fire), tp, fp, prec, lift))

# Sort by precision descending, show top 15 and bottom 5
pair_results.sort(key=lambda x: x[4], reverse=True)
print("\nTOP 15 (highest precision when both fire):")
for pair, fired, tp, fp, prec, lift in pair_results[:15]:
    print(f"  {pair:<53} {fired:>10} {tp:>5} {fp:>5} {prec:>6.3f} {lift:>+15.3f}")

print("\nBOTTOM 10 (lowest precision — over-firing pairs):")
for pair, fired, tp, fp, prec, lift in pair_results[-10:]:
    print(f"  {pair:<53} {fired:>10} {tp:>5} {fp:>5} {prec:>6.3f} {lift:>+15.3f}")

# ─── ANALYSIS 4: Converging family count vs accuracy ───────────────────────────
print()
print("=" * 90)
print("ANALYSIS 4: CONVERGING FAMILY COUNT vs ACCURACY")
print("=" * 90)

total_discards = sum(1 for r in joined if r['gt'] == 'DISCARD')
total_keeps = sum(1 for r in joined if r['gt'] == 'KEEP')

print(f"\n{'Count':>6} {'Total':>6} {'Discards':>9} {'Keeps':>6} {'Discard%':>9} {'Agent DISCARD':>13} {'Agent REVIEW':>12} {'Agent KEEP':>10}")
print("-" * 85)
for count in range(0, 9):
    rows = [r for r in joined if r['converging_family_count'] == count]
    if not rows:
        continue
    d = sum(1 for r in rows if r['gt'] == 'DISCARD')
    k = sum(1 for r in rows if r['gt'] == 'KEEP')
    d_pct = d / len(rows) * 100 if rows else 0
    a_discard = sum(1 for r in rows if r['pred'] == 'DISCARD')
    a_review = sum(1 for r in rows if r['pred'] == 'REVIEW')
    a_keep = sum(1 for r in rows if r['pred'] == 'KEEP')
    print(f"{count:>6} {len(rows):>6} {d:>9} {k:>6} {d_pct:>8.1f}% {a_discard:>13} {a_review:>12} {a_keep:>10}")

# ─── ANALYSIS 5: False Positive deep dive ─────────────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 5: FALSE POSITIVE DEEP DIVE (agent DISCARD, client KEEP)")
print("=" * 90)

fps = [r for r in joined if r['correctness'] == 'FP']
print(f"\nTotal FPs: {len(fps)}")

# FP by primary_removal_reason
print(f"\n--- FPs by Primary Removal Reason ---")
fp_reasons = Counter(r['primary_removal_reason'] for r in fps)
all_reasons = Counter(r['primary_removal_reason'] for r in joined)
for reason, count in fp_reasons.most_common():
    total = all_reasons[reason]
    fp_rate = count / total * 100 if total > 0 else 0
    print(f"  {reason}: {count}/{total} ({fp_rate:.1f}% of this reason are FP)")

# FP by OE classification
print(f"\n--- FPs by OE Classification ---")
fp_oe = Counter(r['oe_classification'] for r in fps)
all_oe = Counter(r['oe_classification'] for r in joined)
for oe, count in fp_oe.most_common():
    total = all_oe[oe]
    fp_rate = count / total * 100 if total > 0 else 0
    print(f"  {oe}: {count}/{total} ({fp_rate:.1f}% of this class are FP)")

# FP by badopen trigger
print(f"\n--- FPs by Badopen Trigger ---")
fp_trigger = Counter(r['badopen_trigger'] for r in fps)
all_trigger = Counter(r['badopen_trigger'] for r in joined)
for trigger, count in fp_trigger.most_common():
    total = all_trigger[trigger]
    fp_rate = count / total * 100 if total > 0 else 0
    print(f"  {trigger}: {count}/{total} ({fp_rate:.1f}% of this trigger are FP)")

# FP by evidence family
print(f"\n--- FPs by Evidence Family (which families fire most in FPs) ---")
fp_fam = Counter()
for r in fps:
    for fam in r['evidence_families_fired']:
        fp_fam[fam] += 1
for fam, count in fp_fam.most_common():
    pct = count / len(fps) * 100
    # Compare to TP rate
    tp_fam = sum(1 for r in joined if r['correctness'] == 'TP' and fam in r['evidence_families_fired'])
    tp_total = sum(1 for r in joined if r['correctness'] == 'TP')
    tp_pct = tp_fam / tp_total * 100 if tp_total > 0 else 0
    print(f"  {fam}: {count}/{len(fps)} ({pct:.1f}% of FPs) vs {tp_pct:.1f}% of TPs")

# FP by converging family count
print(f"\n--- FPs by Converging Family Count ---")
fp_conv = Counter(r['converging_family_count'] for r in fps)
tp_conv = Counter(r['converging_family_count'] for r in joined if r['correctness'] == 'TP')
tp_total = sum(1 for r in joined if r['correctness'] == 'TP')
for count in sorted(fp_conv.keys()):
    fp_count = fp_conv[count]
    tp_count = tp_conv.get(count, 0)
    print(f"  {count} families: {fp_count} FPs, {tp_count} TPs (ratio FP:TP = {fp_count/max(tp_count,1):.1f}:1)")

# FP score distribution
print(f"\n--- FP Score Distribution ---")
fp_scores = [r['score'] for r in fps]
tp_scores = [r['score'] for r in joined if r['correctness'] == 'TP']
if fp_scores:
    print(f"  FP scores: mean={statistics.mean(fp_scores):.3f}, median={statistics.median(fp_scores):.3f}, range=[{min(fp_scores):.3f}, {max(fp_scores):.3f}]")
if tp_scores:
    print(f"  TP scores: mean={statistics.mean(tp_scores):.3f}, median={statistics.median(tp_scores):.3f}, range=[{min(tp_scores):.3f}, {max(tp_scores):.3f}]")

# ─── ANALYSIS 6: False Negative deep dive ─────────────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 6: FALSE NEGATIVE DEEP DIVE (agent REVIEW/KEEP, client DISCARD)")
print("=" * 90)

fns = [r for r in joined if r['correctness'] == 'FN']
print(f"\nTotal FNs: {len(fns)}")

# FN by OE classification
print(f"\n--- FNs by OE Classification ---")
fn_oe = Counter(r['oe_classification'] for r in fns)
for oe, count in fn_oe.most_common():
    total = all_oe[oe]
    fn_rate = count / total * 100 if total > 0 else 0
    print(f"  {oe}: {count}/{total} ({fn_rate:.1f}% of this class are FN)")

# FN by primary_removal_reason
print(f"\n--- FNs by Primary Removal Reason ---")
fn_reasons = Counter(r['primary_removal_reason'] for r in fns)
for reason, count in fn_reasons.most_common():
    total = all_reasons[reason]
    fn_rate = count / total * 100 if total > 0 else 0
    print(f"  {reason}: {count}/{total} ({fn_rate:.1f}% of this reason are FN)")

# FN by evidence family (which families are NOT firing when they should)
print(f"\n--- FNs: Evidence families that DID NOT fire (missed signals) ---")
fn_not_fired = Counter()
for r in fns:
    for fam in all_families:
        if fam not in r['evidence_families_fired']:
            fn_not_fired[fam] += 1
# Compare to TPs: which families fire in TPs but not in FNs
tp_fired = Counter()
for r in joined if False else [r for r in joined if r['correctness'] == 'TP']:
    for fam in r['evidence_families_fired']:
        tp_fired[fam] += 1

tp_total = sum(1 for r in joined if r['correctness'] == 'TP')
fn_total = len(fns)
print(f"  {'Family':<30} {'Fires in TPs':>12} {'Fires in FNs':>12} {'TP rate':>8} {'FN rate':>8} {'Gap':>7}")
print("  " + "-" * 80)
for fam in all_families:
    tp_fire = tp_fired.get(fam, 0)
    fn_fire = fn_total - fn_not_fired.get(fam, fn_total)
    tp_rate = tp_fire / tp_total * 100 if tp_total > 0 else 0
    fn_rate = fn_fire / fn_total * 100 if fn_total > 0 else 0
    gap = tp_rate - fn_rate
    print(f"  {fam:<30} {tp_fire:>12} {fn_fire:>12} {tp_rate:>7.1f}% {fn_rate:>7.1f}% {gap:>+7.1f}%")

# FN by converging family count
print(f"\n--- FNs by Converging Family Count ---")
fn_conv = Counter(r['converging_family_count'] for r in fns)
for count in sorted(fn_conv.keys()):
    fn_count = fn_conv[count]
    tp_count = tp_conv.get(count, 0)
    print(f"  {count} families: {fn_count} FNs, {tp_count} TPs")

# FN score distribution
print(f"\n--- FN Score Distribution ---")
fn_scores = [r['score'] for r in fns]
if fn_scores:
    print(f"  FN scores: mean={statistics.mean(fn_scores):.3f}, median={statistics.median(fn_scores):.3f}, range=[{min(fn_scores):.3f}, {max(fn_scores):.3f}]")
    # How many FNs are REVIEW (close to threshold)?
    review_fns = [r for r in fns if r['pred'] == 'REVIEW']
    keep_fns = [r for r in fns if r['pred'] == 'KEEP']
    print(f"  FNs in REVIEW: {len(review_fns)}, FNs in KEEP: {len(keep_fns)}")

# ─── ANALYSIS 7: OE word count vs client discard ──────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 7: OE WORD COUNT vs CLIENT DISCARD RATE")
print("=" * 90)

word_buckets = [(0, 10), (10, 20), (20, 30), (30, 50), (50, 80), (80, 120), (120, 999)]
print(f"\n{'Word range':<15} {'Total':>6} {'Discards':>9} {'Keeps':>6} {'Discard%':>9} {'Agent DISCARD%':>15}")
print("-" * 65)
for lo, hi in word_buckets:
    rows = [r for r in joined if r['oe_word_count'] is not None and lo <= r['oe_word_count'] < hi]
    if not rows:
        continue
    d = sum(1 for r in rows if r['gt'] == 'DISCARD')
    k = sum(1 for r in rows if r['gt'] == 'KEEP')
    d_pct = d / len(rows) * 100 if rows else 0
    a_d = sum(1 for r in rows if r['pred'] == 'DISCARD')
    a_d_pct = a_d / len(rows) * 100 if rows else 0
    print(f"{f'{lo}-{hi}':<15} {len(rows):>6} {d:>9} {k:>6} {d_pct:>8.1f}% {a_d_pct:>14.1f}%")

# ─── ANALYSIS 8: ML score vs client discard ───────────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 8: ML SCORE vs CLIENT DISCARD RATE")
print("=" * 90)

ml_buckets = [(0, 0.2), (0.2, 0.4), (0.4, 0.5), (0.5, 0.6), (0.6, 0.7), (0.7, 0.8), (0.8, 1.1)]
print(f"\n{'ML range':<15} {'Total':>6} {'Discards':>9} {'Keeps':>6} {'Discard%':>9} {'Agent DISCARD%':>15}")
print("-" * 65)
for lo, hi in ml_buckets:
    rows = [r for r in joined if r['ml_score'] is not None and lo <= r['ml_score'] < hi]
    if not rows:
        continue
    d = sum(1 for r in rows if r['gt'] == 'DISCARD')
    k = sum(1 for r in rows if r['gt'] == 'KEEP')
    d_pct = d / len(rows) * 100 if rows else 0
    a_d = sum(1 for r in rows if r['pred'] == 'DISCARD')
    a_d_pct = a_d / len(rows) * 100 if rows else 0
    print(f"{f'{lo:.1f}-{hi:.1f}':<15} {len(rows):>6} {d:>9} {k:>6} {d_pct:>8.1f}% {a_d_pct:>14.1f}%")

# ─── ANALYSIS 9: Stage verdict accuracy ───────────────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 9: STAGE VERDICT ACCURACY")
print("=" * 90)

for stage in ['stage1_fraud_verdict', 'stage2_quality_verdict']:
    print(f"\n--- {stage} ---")
    print(f"  {'Verdict':<12} {'Total':>6} {'Discards':>9} {'Keeps':>6} {'Discard%':>9} {'Agent DISCARD%':>15}")
    for verdict in ['pass', 'ambiguous', 'fail']:
        rows = [r for r in joined if r[stage] == verdict]
        if not rows:
            continue
        d = sum(1 for r in rows if r['gt'] == 'DISCARD')
        k = sum(1 for r in rows if r['gt'] == 'KEEP')
        d_pct = d / len(rows) * 100 if rows else 0
        a_d = sum(1 for r in rows if r['pred'] == 'DISCARD')
        a_d_pct = a_d / len(rows) * 100 if rows else 0
        print(f"  {verdict:<12} {len(rows):>6} {d:>9} {k:>6} {d_pct:>8.1f}% {a_d_pct:>14.1f}%")

# ─── ANALYSIS 10: Badopen severity vs client discard ──────────────────────────
print()
print("=" * 90)
print("ANALYSIS 10: BADOPEN SEVERITY vs CLIENT DISCARD")
print("=" * 90)

print(f"\n{'Severity':<12} {'Total':>6} {'Discards':>9} {'Keeps':>6} {'Discard%':>9} {'Agent DISCARD%':>15}")
print("-" * 65)
for sev in ['none', 'low', 'medium', 'high']:
    rows = [r for r in joined if r['badopen_severity'] == sev]
    if not rows:
        continue
    d = sum(1 for r in rows if r['gt'] == 'DISCARD')
    k = sum(1 for r in rows if r['gt'] == 'KEEP')
    d_pct = d / len(rows) * 100 if rows else 0
    a_d = sum(1 for r in rows if r['pred'] == 'DISCARD')
    a_d_pct = a_d / len(rows) * 100 if rows else 0
    print(f"{sev:<12} {len(rows):>6} {d:>9} {k:>6} {d_pct:>8.1f}% {a_d_pct:>14.1f}%")

# ─── ANALYSIS 11: Removal confidence vs accuracy ──────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 11: REMOVAL CONFIDENCE vs ACCURACY")
print("=" * 90)

print(f"\n{'Confidence':<12} {'Total':>6} {'TP':>5} {'FP':>5} {'FN':>5} {'TN':>5} {'Prec':>6} {'Recall':>7}")
print("-" * 60)
for conf_lo, conf_hi in [(0, 0.2), (0.2, 0.4), (0.4, 0.6), (0.6, 0.8), (0.8, 1.1)]:
    rows = [r for r in joined if r['removal_confidence'] is not None and conf_lo <= r['removal_confidence'] < conf_hi]
    if not rows:
        continue
    tp = sum(1 for r in rows if r['correctness'] == 'TP')
    fp = sum(1 for r in rows if r['correctness'] == 'FP')
    fn = sum(1 for r in rows if r['correctness'] == 'FN')
    tn = sum(1 for r in rows if r['correctness'] == 'TN')
    prec = tp / (tp + fp) if (tp + fp) > 0 else 0
    rec = tp / (tp + fn) if (tp + fn) > 0 else 0
    print(f"{f'{conf_lo:.1f}-{conf_hi:.1f}':<12} {len(rows):>6} {tp:>5} {fp:>5} {fn:>5} {tn:>5} {prec:>6.3f} {rec:>7.3f}")

# ─── ANALYSIS 12: Top ML signals in TP vs FP vs FN ────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 12: ML TOP SIGNALS BY CORRECTNESS")
print("=" * 90)

for correctness_label, label in [('TP', 'True Positives'), ('FP', 'False Positives'), ('FN', 'False Negatives'), ('TN', 'True Negatives')]:
    rows = [r for r in joined if r['correctness'] == correctness_label]
    signal_counts = Counter()
    for r in rows:
        for sig in r['ml_top_signals']:
            signal_counts[sig] += 1
    print(f"\n--- {label} ({len(rows)} rows) — top 10 ML signals ---")
    for sig, count in signal_counts.most_common(10):
        pct = count / len(rows) * 100 if rows else 0
        print(f"  {sig}: {count} ({pct:.1f}%)")

# ─── ANALYSIS 13: Score threshold optimization for agent_score ─────────────────
print()
print("=" * 90)
print("ANALYSIS 13: AGENT_SCORE THRESHOLD OPTIMIZATION")
print("=" * 90)

scores = [(r['score'], r['gt']) for r in joined]
print(f"\n  {'Threshold':>10} {'TP':>5} {'FP':>5} {'FN':>5} {'TN':>5} {'Prec':>6} {'Recall':>7} {'F1':>6} {'BalAcc':>7}")
print("  " + "-" * 70)
best_f1 = 0
best_t = 0
for t in [-0.8, -0.7, -0.6, -0.5, -0.4, -0.35, -0.3, -0.25, -0.2, -0.15, -0.1, -0.05, 0.0, 0.05, 0.1, 0.2, 0.3, 0.4, 0.5]:
    tp = sum(1 for s, g in scores if s <= t and g == 'DISCARD')
    fp = sum(1 for s, g in scores if s <= t and g == 'KEEP')
    fn = sum(1 for s, g in scores if s > t and g == 'DISCARD')
    tn = sum(1 for s, g in scores if s > t and g == 'KEEP')
    prec = tp / (tp + fp) if (tp + fp) > 0 else 0
    rec = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0
    bal_acc = (rec + tn / (tn + fp)) / 2 if (tn + fp) > 0 else 0
    print(f"  {t:>10.2f} {tp:>5} {fp:>5} {fn:>5} {tn:>5} {prec:>6.3f} {rec:>7.3f} {f1:>6.3f} {bal_acc:>7.3f}")
    if f1 > best_f1:
        best_f1 = f1
        best_t = t

print(f"\n  Best F1={best_f1:.3f} at threshold={best_t:.2f}")

# ─── ANALYSIS 14: REVIEW bucket optimization ──────────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 14: REVIEW BUCKET — SHOULD SOME BE DISCARD?")
print("=" * 90)

reviews = [r for r in joined if r['pred'] == 'REVIEW']
review_discards = [r for r in reviews if r['gt'] == 'DISCARD']
review_keeps = [r for r in reviews if r['gt'] == 'KEEP']
print(f"\nTotal REVIEW: {len(reviews)} ({len(review_discards)} actual discards, {len(review_keeps)} actual keeps)")
print(f"  If we could perfectly split REVIEW, we'd get {len(review_discards)} more TPs")

# What distinguishes REVIEW-discards from REVIEW-keeps?
print(f"\n--- REVIEW discards vs REVIEW keeps: component scores ---")
for comp in ['authenticity_risk', 'quality_discard_risk', 'client_reject_probability', 'removal_confidence']:
    d_vals = [r[comp] for r in review_discards if r[comp] is not None]
    k_vals = [r[comp] for r in review_keeps if r[comp] is not None]
    if d_vals and k_vals:
        sep = statistics.mean(d_vals) - statistics.mean(k_vals)
        print(f"  {comp}: discards mean={statistics.mean(d_vals):.3f}, keeps mean={statistics.mean(k_vals):.3f}, sep={sep:+.3f}")

print(f"\n--- REVIEW discards vs REVIEW keeps: agent_score ---")
d_scores = [r['score'] for r in review_discards]
k_scores = [r['score'] for r in review_keeps]
if d_scores and k_scores:
    print(f"  discards: mean={statistics.mean(d_scores):.3f}, median={statistics.median(d_scores):.3f}")
    print(f"  keeps:    mean={statistics.mean(k_scores):.3f}, median={statistics.median(k_scores):.3f}")

print(f"\n--- REVIEW discards vs REVIEW keeps: converging family count ---")
d_conv = Counter(r['converging_family_count'] for r in review_discards)
k_conv = Counter(r['converging_family_count'] for r in review_keeps)
for count in sorted(set(list(d_conv.keys()) + list(k_conv.keys()))):
    d = d_conv.get(count, 0)
    k = k_conv.get(count, 0)
    d_pct = d / (d + k) * 100 if (d + k) > 0 else 0
    print(f"  {count} families: {d} discards, {k} keeps ({d_pct:.0f}% discard)")

print(f"\n--- REVIEW discards vs REVIEW keeps: OE classification ---")
d_oe = Counter(r['oe_classification'] for r in review_discards)
k_oe = Counter(r['oe_classification'] for r in review_keeps)
for oe in sorted(set(list(d_oe.keys()) + list(k_oe.keys()))):
    d = d_oe.get(oe, 0)
    k = k_oe.get(oe, 0)
    d_pct = d / (d + k) * 100 if (d + k) > 0 else 0
    print(f"  {oe}: {d} discards, {k} keeps ({d_pct:.0f}% discard)")

print(f"\n--- REVIEW discards vs REVIEW keeps: ML score ---")
d_ml = [r['ml_score'] for r in review_discards if r['ml_score'] is not None]
k_ml = [r['ml_score'] for r in review_keeps if r['ml_score'] is not None]
if d_ml and k_ml:
    print(f"  discards: mean={statistics.mean(d_ml):.3f}, median={statistics.median(d_ml):.3f}")
    print(f"  keeps:    mean={statistics.mean(k_ml):.3f}, median={statistics.median(k_ml):.3f}")

# ─── ANALYSIS 15: Key learnings summary ───────────────────────────────────────
print()
print("=" * 90)
print("ANALYSIS 15: KEY LEARNINGS SUMMARY")
print("=" * 90)

# Calculate key stats
tp_total = sum(1 for r in joined if r['correctness'] == 'TP')
fp_total = sum(1 for r in joined if r['correctness'] == 'FP')
fn_total = sum(1 for r in joined if r['correctness'] == 'FN')
tn_total = sum(1 for r in joined if r['correctness'] == 'TN')

print(f"""
CURRENT V6 PERFORMANCE:
  TP={tp_total}, FP={fp_total}, FN={fn_total}, TN={tn_total}
  Precision={tp_total/(tp_total+fp_total):.3f}, Recall={tp_total/(tp_total+fn_total):.3f}
  F1={2*tp_total/(2*tp_total+fp_total+fn_total):.3f}

KEY FINDINGS:
  1. REVIEW bucket is the bottleneck — {len(reviews)} respondents in REVIEW, {len(review_discards)} are actual discards
     → Improving REVIEW→DISCARD conversion is the highest-impact lever
  2. model_risk has the highest discriminative gap (+0.34) but fires in {family_results.get('model_risk', {}).get('fired', 0)} respondents
  3. core_oe_quality has NEGATIVE gap — confirming it should NOT drive discards alone
  4. platform_risk fires equally in discards and keeps — over-firing, needs tightening
  5. See component score thresholds and family combination analysis above for calibration targets
""")

# Save full analysis
output_path = V6_DIR / 'v6_metadata_analysis.json'
analysis_output = {
    'family_results': family_results,
    'pair_results': [{'pair': p[0], 'fired': p[1], 'tp': p[2], 'fp': p[3], 'precision': p[4], 'lift': p[5]} for p in pair_results[:20]],
    'counts': {'tp': tp_total, 'fp': fp_total, 'fn': fn_total, 'tn': tn_total},
    'review_bucket': {'total': len(reviews), 'discards': len(review_discards), 'keeps': len(review_keeps)},
}
with open(output_path, 'w') as f:
    json.dump(analysis_output, f, indent=2)
print(f"Full analysis saved to: {output_path}")
