#!/usr/bin/env python3
"""V2 ML pipeline: Dataset-relative features + supplier risk + threshold optimization.

Key improvements over V1:
1. Dataset-relative timing (z-score within dataset)
2. Supplier reject rate computed from TRAINING data only (no leakage)
3. Class balancing via sample weights
4. Threshold optimized on training data (not test)
5. More text features (lexical diversity, repetition, specificity)
6. Cross-respondent features (duplicate cluster sizes)
"""
from __future__ import annotations

import csv
import json
import re
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.metrics import (
    accuracy_score, classification_report, confusion_matrix,
    f1_score, precision_score, recall_score, roc_auc_score,
)

warnings.filterwarnings("ignore")

DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
SIGNAL_MAP = DATA_DIR / "autosurvey-outputs/status-ground-truth-calibration/core_loop_2026-06-24/status_respondent_signal_map.csv"

DATASET_MAP = {
    "106-2502 Delta Water Filtration.xlsx": "260111_Delta Water Filtration.xlsx",
    "109-2601 Echo BH.xlsx": "260300_ECHO.xlsx",
    "153-2602 ODL Switchable Glass.xlsx": "260501_ODL.xlsx",
    "159-2601 Oldcastle Brand Health.xlsx": "260206_OC BH.xlsx",
    "159-2602 Oldcastle Canada.xlsx": "260401_ OC CAN.xlsx",
    "189-2501 SBD Brand Association.xlsx": "260200_SBD.xlsx",
    "287-2501 THD Digital CX.xlsx": "251101_THD CX.xlsx",
    "365-2601 ADDO RaceTrac US GP.xlsx": "260404_ADDO.xlsx",
    "368-2602 Masterlock Conjoint.xlsx": "260403_Masterlock Conjoint.xlsx",
    "999-2601 TFG Contractor Index Q1.xlsx": "251205_TFG Contractor Index Q1.xlsx",
    "999-2602 TFG Contractor Index Q2.xlsx": "260306_TFG Contractor Index Q2.xlsx",
}

ALL_SIGNALS = set()


def clean(v) -> str:
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def norm(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v


def load_signal_map() -> dict[str, dict[str, dict]]:
    by_dataset: dict[str, dict[str, dict]] = defaultdict(dict)
    with open(SIGNAL_MAP) as f:
        for row in csv.DictReader(f):
            ds = row["dataset"]
            rid = row["respondent_key"]
            signals = [s.strip() for s in row["signals"].split(";") if s.strip()]
            for s in signals:
                ALL_SIGNALS.add(s)
            by_dataset[ds][rid] = {
                "signals": signals,
                "decision": row["tfg_decision"],
                "signal_count": int(row["signal_count"]),
            }
    return by_dataset


def lexical_diversity(text: str) -> float:
    """Type-token ratio."""
    words = text.lower().split()
    if not words: return 0.0
    return len(set(words)) / len(words)


def extract_features_from_excel(filepath: Path, signal_map: dict) -> pd.DataFrame:
    """Extract ALL features from a raw Excel file."""
    signal_map_name = DATASET_MAP[filepath.name]
    sm = signal_map.get(signal_map_name, {})

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Identify column groups
    open_text_cols = [(i, h) for i, h in enumerate(headers) if h and (
        str(h).lower().endswith("oe") or str(h).lower() == "outro" or "qcoe" in str(h).lower()
    )]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and "LangAssess" in str(h)]
    rd_search_cols = [(i, h) for i, h in enumerate(headers) if h and str(h).startswith("RD_Search")]
    flag_cols = [(i, h) for i, h in enumerate(headers) if h and any(
        x in str(h) for x in ["TERMFLAGS", "clientflagsr1", "scrutinyflags", "pasted"]
    )]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and re.match(r"q\d+r\d+$", str(h))]

    excluded_prefixes = ("RD_", "LangAssess", "noanswer", "qc5R1_", "conditions",
                         "outroR1_", "qcoe1R1_", "_Pasted", "POSSIBLE", "OWNERSHIP")
    excluded_names = {"qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel",
                      "LangAssessReadEase", "LangAssessNumSen", "LangAssessNumWords",
                      "LangAssessNumSyl", "url", "session", "camp", "bhf", "sfh",
                      "intcode", "record", "uuid", "status", "qtime", "SUPNAME",
                      "ipAddress", "date", "qStateVer", "outro", "outro_Pasted",
                      "CLASSIFY", "CHANNELTRACKING", "RID", "list", "userAgent",
                      "dcua", "start_date", "vlist", "vos", "vbrowser", "vmobiledevice",
                      "vmobileos", "VALIDCLIENT"}
    coded_cols = []
    for i, h in enumerate(headers):
        if not h: continue
        hs = str(h)
        if hs in excluded_names: continue
        if any(hs.startswith(p) for p in excluded_prefixes): continue
        if hs.endswith("oe") or hs.endswith("oth") or hs == "outro": continue
        if "LangAssess" in hs or "RD_" in hs: continue
        coded_cols.append((i, h))

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid_idx = hidx.get("uuid") or hidx.get("record")
        if rid_idx is None or rid_idx >= len(row): continue
        rid = clean(row[rid_idx])
        if not rid or rid not in sm: continue
        rows_data.append((rid, row))

    if not rows_data:
        return pd.DataFrame()

    # Collect all data first for cross-respondent features
    all_oe_texts = []
    all_user_agents = []
    all_ips = []
    all_start_dates = []

    for rid, row in rows_data:
        oe_texts = []
        for i, h in open_text_cols:
            v = row[i] if i < len(row) else None
            text = clean(v)
            if text:
                oe_texts.append(text)
        all_oe_texts.append(" | ".join(oe_texts))

        ua_idx = hidx.get("userAgent")
        if ua_idx is not None and ua_idx < len(row):
            all_user_agents.append(clean(row[ua_idx]))
        else:
            all_user_agents.append("")

        ip_idx = hidx.get("ipAddress")
        if ip_idx is not None and ip_idx < len(row):
            all_ips.append(clean(row[ip_idx]))
        else:
            all_ips.append("")

        sd_idx = hidx.get("start_date")
        if sd_idx is not None and sd_idx < len(row):
            all_start_dates.append(clean(row[sd_idx]))
        else:
            all_start_dates.append("")

    # Compute cross-respondent duplicate counts
    oe_counter = Counter(t.strip().lower() for t in all_oe_texts if t.strip())
    ua_counter = Counter(ua for ua in all_user_agents if ua)
    ip_counter = Counter(ip for ip in all_ips if ip)
    sd_counter = Counter(sd for sd in all_start_dates if sd)

    # Build features
    features = []
    for idx, (rid, row) in enumerate(rows_data):
        feat = {"respondent_id": rid}
        sm_entry = sm[rid]
        signals = set(sm_entry["signals"])

        # === 1. Client signals (one-hot) ===
        for sig in ALL_SIGNALS:
            feat[f"sig_{sig}"] = 1 if sig in signals else 0
        feat["signal_count"] = sm_entry["signal_count"]

        # Signal tier counts
        t1_sigs = {"termflags_nonzero", "long_low_specificity_text", "ai_or_overpolished_text_marker", "generic_placeholder_open_end"}
        t2_sigs = {"rd_searchr3_canada", "rd_searchr1_22", "rd_searchr1_23", "rd_searchr1_20", "qtime_under_dataset_p10",
                   "rd_searchr1_22.0", "rd_searchr1_23.0", "rd_searchr1_20.0"}
        feat["t1_count"] = len(signals & t1_sigs)
        feat["t2_count"] = len(signals & t2_sigs)
        feat["t3_count"] = len(signals - t1_sigs - t2_sigs)

        # === 2. Timing (raw — will be normalized later) ===
        qtime_idx = hidx.get("qtime")
        if qtime_idx is not None and qtime_idx < len(row):
            qtime = row[qtime_idx]
            try: qtime = float(qtime) if qtime else 0
            except: qtime = 0
        else:
            qtime = 0
        feat["qtime_seconds"] = qtime
        feat["qtime_log"] = np.log1p(qtime) if qtime > 0 else 0

        # === 3. LangAssess NLP features ===
        for i, h in lang_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"lang_{h}"] = float(v) if v is not None else 0
            except: feat[f"lang_{h}"] = 0

        # === 4. Open-end text features ===
        oe_texts = all_oe_texts[idx].split(" | ") if all_oe_texts[idx] else []
        oe_texts = [t for t in oe_texts if t]
        oe_lengths = [len(t) for t in oe_texts]
        oe_word_counts = [len(t.split()) for t in oe_texts]

        feat["oe_count"] = len(oe_texts)
        feat["oe_total_chars"] = sum(oe_lengths)
        feat["oe_max_chars"] = max(oe_lengths) if oe_lengths else 0
        feat["oe_min_chars"] = min(oe_lengths) if oe_lengths else 0
        feat["oe_mean_chars"] = np.mean(oe_lengths) if oe_lengths else 0
        feat["oe_total_words"] = sum(oe_word_counts)
        feat["oe_max_words"] = max(oe_word_counts) if oe_word_counts else 0
        feat["oe_mean_words"] = np.mean(oe_word_counts) if oe_word_counts else 0

        all_oe_text = " ".join(oe_texts).lower()
        feat["oe_lexical_diversity"] = lexical_diversity(all_oe_text)
        feat["oe_has_none"] = 1 if any(w in all_oe_text for w in ["none", "n/a", "na", "nothing", "no opinion", "no idea"]) else 0
        feat["oe_has_generic"] = 1 if any(w in all_oe_text for w in ["good", "fine", "ok", "okay", "nice", "great", "yes", "no"]) and oe_word_counts and max(oe_word_counts) <= 3 else 0
        feat["oe_all_caps"] = 1 if oe_texts and any(t.isupper() and len(t) > 5 for t in oe_texts) else 0
        feat["oe_single_word"] = 1 if oe_word_counts and max(oe_word_counts) <= 1 else 0
        feat["oe_very_short"] = 1 if oe_lengths and max(oe_lengths) < 10 else 0
        feat["oe_short"] = 1 if oe_lengths and max(oe_lengths) < 30 else 0

        # === 5. Supplier ===
        sup_idx = hidx.get("SUPNAME")
        if sup_idx is not None and sup_idx < len(row):
            supplier = clean(row[sup_idx])
        else:
            supplier = ""
        feat["supplier_name"] = supplier
        feat["supplier_missing"] = 1 if not supplier or supplier == "MISSING" else 0
        feat["supplier_is_none"] = 1 if supplier == "None" or supplier == "" else 0

        # === 6. Matrix/grid straightlining ===
        matrix_values = []
        for i, h in matrix_cols:
            v = row[i] if i < len(row) else None
            if v is not None and v != "":
                matrix_values.append(norm(v))
        if matrix_values:
            unique_ratio = len(set(matrix_values)) / len(matrix_values)
            feat["matrix_unique_ratio"] = unique_ratio
            feat["matrix_straightline"] = 1 if unique_ratio <= 0.2 and len(matrix_values) >= 5 else 0
            feat["matrix_near_straightline"] = 1 if unique_ratio <= 0.4 and len(matrix_values) >= 5 else 0
            feat["matrix_count"] = len(matrix_values)
            feat["matrix_unique_count"] = len(set(matrix_values))
            val_counts = Counter(matrix_values)
            feat["matrix_most_common_freq"] = val_counts.most_common(1)[0][1] / len(matrix_values)
        else:
            feat["matrix_unique_ratio"] = 1.0
            feat["matrix_straightline"] = 0
            feat["matrix_near_straightline"] = 0
            feat["matrix_count"] = 0
            feat["matrix_unique_count"] = 0
            feat["matrix_most_common_freq"] = 0

        # === 7. Coded answer diversity ===
        coded_values = []
        for i, h in coded_cols:
            v = row[i] if i < len(row) else None
            if v is not None and v != "":
                coded_values.append(str(norm(v)))
        feat["coded_count"] = len(coded_values)
        feat["coded_unique_ratio"] = len(set(coded_values)) / len(coded_values) if coded_values else 1.0
        dk_count = sum(1 for v in coded_values if any(x in v.lower() for x in ["don't know", "dk", "not sure", "no answer"]))
        feat["coded_dk_count"] = dk_count
        feat["coded_dk_ratio"] = dk_count / len(coded_values) if coded_values else 0

        # === 8. RD_Search features ===
        for i, h in rd_search_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"rd_{h}"] = float(v) if v is not None else 0
            except: feat[f"rd_{h}"] = 0

        # === 9. Flag columns ===
        for i, h in flag_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"flag_{h}"] = float(v) if v is not None else 0
            except: feat[f"flag_{h}"] = 0

        # === 10. Cross-respondent duplicates ===
        oe_text = all_oe_texts[idx]
        feat["oe_duplicate_count"] = oe_counter.get(oe_text.strip().lower(), 0) if oe_text.strip() else 0
        feat["oe_is_duplicate"] = 1 if feat["oe_duplicate_count"] > 1 else 0
        feat["ua_duplicate_count"] = ua_counter.get(all_user_agents[idx], 0)
        feat["ua_is_duplicate"] = 1 if feat["ua_duplicate_count"] > 1 else 0
        feat["ip_duplicate_count"] = ip_counter.get(all_ips[idx], 0)
        feat["ip_is_duplicate"] = 1 if feat["ip_duplicate_count"] > 1 else 0
        feat["sd_duplicate_count"] = sd_counter.get(all_start_dates[idx], 0)
        feat["sd_is_duplicate"] = 1 if feat["sd_duplicate_count"] > 1 else 0

        # === 11. Demographic features (categorical) ===
        demo_cols = ["qstate", "REGION", "age", "qager1", "qGender", "q13", "q12",
                     "qHomeType", "q2", "q1", "qIndustry", "qNumEmployees", "q9",
                     "CLASSIFY", "CHANNELTRACKING"]
        for h in demo_cols:
            i = hidx.get(h)
            if i is not None and i < len(row):
                feat[f"demo_{h}"] = clean(row[i])
            else:
                feat[f"demo_{h}"] = ""

        # === 12. Technical features ===
        tech_cols = ["vos", "vbrowser", "vmobiledevice", "vmobileos"]
        for h in tech_cols:
            i = hidx.get(h)
            if i is not None and i < len(row):
                feat[f"tech_{h}"] = clean(row[i])
            else:
                feat[f"tech_{h}"] = ""

        # === 13. Ground truth ===
        feat["label"] = 1 if sm_entry["decision"] == "rejected" else 0
        feat["dataset"] = signal_map_name

        features.append(feat)

    df = pd.DataFrame(features)
    return df


def add_dataset_relative_features(train_df: pd.DataFrame, test_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Add dataset-relative features computed from TRAIN data only."""
    # Compute supplier reject rate from training data
    supplier_rates = train_df.groupby("supplier_name")["label"].agg(["mean", "count"]).reset_index()
    supplier_rates.columns = ["supplier_name", "supplier_reject_rate", "supplier_count"]
    # For suppliers with few samples, shrink toward the global mean
    global_rate = train_df["label"].mean()
    supplier_rates["supplier_reject_rate_smoothed"] = (
        supplier_rates["supplier_count"] * supplier_rates["supplier_reject_rate"] + 20 * global_rate
    ) / (supplier_rates["supplier_count"] + 20)

    # Merge into train and test
    train_df = train_df.merge(supplier_rates[["supplier_name", "supplier_reject_rate_smoothed"]], on="supplier_name", how="left")
    test_df = test_df.merge(supplier_rates[["supplier_name", "supplier_reject_rate_smoothed"]], on="supplier_name", how="left")
    train_df["supplier_reject_rate_smoothed"] = train_df["supplier_reject_rate_smoothed"].fillna(global_rate)
    test_df["supplier_reject_rate_smoothed"] = test_df["supplier_reject_rate_smoothed"].fillna(global_rate)

    # Compute dataset-relative timing (z-score within the test dataset)
    # For train: z-score within each training dataset
    for df in [train_df, test_df]:
        df["qtime_zscore"] = 0.0
        for ds in df["dataset"].unique():
            mask = df["dataset"] == ds
            qtime_vals = df.loc[mask, "qtime_seconds"]
            mean_q = qtime_vals.mean()
            std_q = qtime_vals.std()
            if std_q > 0:
                df.loc[mask, "qtime_zscore"] = (qtime_vals - mean_q) / std_q
            else:
                df.loc[mask, "qtime_zscore"] = 0

        # Dataset-relative signal count
        df["signal_count_zscore"] = 0.0
        for ds in df["dataset"].unique():
            mask = df["dataset"] == ds
            sc_vals = df.loc[mask, "signal_count"]
            mean_s = sc_vals.mean()
            std_s = sc_vals.std()
            if std_s > 0:
                df.loc[mask, "signal_count_zscore"] = (sc_vals - mean_s) / std_s
            else:
                df.loc[mask, "signal_count_zscore"] = 0

    return train_df, test_df


def prepare_features(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.Series]:
    """Prepare features for ML."""
    non_feature_cols = {"respondent_id", "label", "dataset", "supplier_name"}
    feature_cols = [c for c in df.columns if c not in non_feature_cols]

    X = df[feature_cols].copy()
    y = df["label"].copy()

    # Encode categorical columns
    cat_cols = X.select_dtypes(include=["object"]).columns
    for col in cat_cols:
        X[col] = pd.Categorical(X[col]).codes

    X = X.fillna(0)
    return X, y


def find_best_threshold(y_true, y_proba, target_rate=None):
    """Find the threshold that maximizes F1, or matches a target rate."""
    if target_rate is not None:
        # Find threshold that produces the target discard rate
        thresholds = np.linspace(0.01, 0.99, 100)
        best_thresh = 0.5
        best_diff = 1.0
        for t in thresholds:
            pred = (y_proba >= t).astype(int)
            rate = pred.mean()
            diff = abs(rate - target_rate)
            if diff < best_diff:
                best_diff = diff
                best_thresh = t
        return best_thresh

    # Maximize F1
    thresholds = np.linspace(0.01, 0.99, 100)
    best_f1 = 0
    best_thresh = 0.5
    for t in thresholds:
        pred = (y_proba >= t).astype(int)
        f1 = f1_score(y_true, pred, zero_division=0)
        if f1 > best_f1:
            best_f1 = f1
            best_thresh = t
    return best_thresh


def run_lodo_cv():
    print("Loading signal map...")
    signal_map = load_signal_map()
    print(f"  {sum(len(v) for v in signal_map.values())} annotations, {len(ALL_SIGNALS)} unique signals")

    print("\nExtracting features from all datasets...")
    all_dfs = []
    for xlsx_name in DATASET_MAP:
        filepath = DATA_DIR / xlsx_name
        if not filepath.exists(): continue
        print(f"  {xlsx_name}...")
        df = extract_features_from_excel(filepath, signal_map)
        if df is not None and len(df) > 0:
            print(f"    {len(df)} respondents, {len(df.columns)} features")
            all_dfs.append(df)

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\nCombined: {len(combined)} respondents, {len(combined.columns)} features")

    datasets = combined["dataset"].unique()
    results = []

    print(f"\n{'='*120}")
    print(f"LEAVE-ONE-DATASET-OUT CV — V2 (dataset-relative + supplier risk + threshold opt)")
    print(f"{'='*120}")

    for test_ds in datasets:
        train_df = combined[combined["dataset"] != test_ds].copy()
        test_df = combined[combined["dataset"] == test_ds].copy()

        # Add dataset-relative features
        train_df, test_df = add_dataset_relative_features(train_df, test_df)

        X_train, y_train = prepare_features(train_df)
        X_test, y_test = prepare_features(test_df)

        # Align columns
        for col in X_train.columns:
            if col not in X_test.columns:
                X_test[col] = 0
        for col in X_test.columns:
            if col not in X_train.columns:
                X_train[col] = 0
        X_test = X_test[X_train.columns]

        # Class weights
        n_pos = y_train.sum()
        n_neg = len(y_train) - n_pos
        weights = np.where(y_train == 1, len(y_train) / (2 * n_pos), len(y_train) / (2 * n_neg))

        # Train model
        model = GradientBoostingClassifier(
            n_estimators=300,
            max_depth=4,
            learning_rate=0.05,
            subsample=0.8,
            random_state=42,
        )
        model.fit(X_train, y_train, sample_weight=weights)

        # Predict probabilities
        y_proba = model.predict_proba(X_test)[:, 1]

        # Find best threshold on TRAINING data (not test)
        y_train_proba = model.predict_proba(X_train)[:, 1]
        # Use the training data's reject rate as the target
        train_reject_rate = y_train.mean()
        best_thresh = find_best_threshold(y_train, y_train_proba, target_rate=train_reject_rate)

        # Apply threshold
        y_pred = (y_proba >= best_thresh).astype(int)

        # Metrics
        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, zero_division=0)
        rec = recall_score(y_test, y_pred, zero_division=0)
        f1 = f1_score(y_test, y_pred, zero_division=0)
        auc = roc_auc_score(y_test, y_proba) if len(y_test.unique()) > 1 else 0
        discard_rate = y_pred.mean()

        cm = confusion_matrix(y_test, y_pred, labels=[0, 1])
        tn, fp, fn, tp = cm.ravel()

        results.append({
            "dataset": test_ds, "n": len(y_test),
            "client_reject_rate": float(y_test.mean()),
            "agent_discard_rate": float(discard_rate),
            "threshold": best_thresh,
            "accuracy": acc, "precision": prec, "recall": rec, "f1": f1, "auc": auc,
            "tp": int(tp), "fp": int(fp), "tn": int(tn), "fn": int(fn),
        })

        print(f"\n{test_ds}")
        print(f"  N={len(y_test)}, Client reject: {y_test.mean():.1%}, Agent discard: {discard_rate:.1%}, Threshold: {best_thresh:.3f}")
        print(f"  Acc={acc:.1%}  Prec={prec:.1%}  Recall={rec:.1%}  F1={f1:.1%}  AUC={auc:.3f}")
        print(f"  TP={tp}  FP={fp}  TN={tn}  FN={fn}")

        # Top features
        if hasattr(model, "feature_importances_"):
            imp = sorted(zip(X_train.columns, model.feature_importances_), key=lambda x: -x[1])[:10]
            print(f"  Top features: {', '.join(f'{n}={v:.3f}' for n, v in imp)}")

    # Aggregate
    print(f"\n{'='*120}")
    print("AGGREGATE RESULTS")
    print(f"{'='*120}")
    total_tp = sum(r["tp"] for r in results)
    total_fp = sum(r["fp"] for r in results)
    total_tn = sum(r["tn"] for r in results)
    total_fn = sum(r["fn"] for r in results)
    total_n = sum(r["n"] for r in results)

    agg_acc = (total_tp + total_tn) / total_n
    agg_prec = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0
    agg_rec = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0
    agg_f1 = 2 * agg_prec * agg_rec / (agg_prec + agg_rec) if (agg_prec + agg_rec) > 0 else 0

    print(f"  Total: {total_n} respondents")
    print(f"  TP={total_tp}  FP={total_fp}  TN={total_tn}  FN={total_fn}")
    print(f"  Accuracy:  {agg_acc:.1%}")
    print(f"  Precision: {agg_prec:.1%}")
    print(f"  Recall:    {agg_rec:.1%}")
    print(f"  F1:        {agg_f1:.1%}")
    print(f"  Discard rate: {(total_tp + total_fp) / total_n:.1%}")
    print(f"  Client reject rate: {(total_tp + total_fn) / total_n:.1%}")

    # Compare to baselines
    print(f"\n  BASELINE COMPARISON:")
    print(f"    Keep everyone:     Accuracy={1 - (total_tp + total_fn) / total_n:.1%}")
    print(f"    Our V3 pipeline:   Accuracy=75.7%, F1=47.8%")
    print(f"    This ML pipeline:  Accuracy={agg_acc:.1%}, F1={agg_f1:.1%}")

    results_df = pd.DataFrame(results)
    results_df.to_csv(DATA_DIR / "autosurvey-outputs/lodo_cv_v2_results.csv", index=False)
    print(f"\nResults saved to lodo_cv_v2_results.csv")

    return results


if __name__ == "__main__":
    run_lodo_cv()
