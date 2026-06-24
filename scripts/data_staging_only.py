#!/usr/bin/env python3
"""Stage 1: Data staging script for the multi-agent production pipeline.

This script does NO semantic work. It only:
1. Parses the Datamap into question text and value labels
2. Maps each response field to its question text and value label meaning
3. Computes population-level statistics (timing, supplier, duplicates)
4. Assembles structured JSON packets for agents to read in Stage 2

NO scoring. NO classification. NO coherence checks. NO regex on open-ends.
"""
from __future__ import annotations

import hashlib
import json
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
OUTPUT_BASE = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent")

EXCLUDED_FIELDS = {
    "qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel", "LangAssessReadEase",
    "LangAssessNumSen", "LangAssessNumWords", "LangAssessNumSyl", "url", "session",
    "camp", "bhf", "sfh", "intcode", "record", "uuid", "status", "qtime",
    "SUPNAME", "ipAddress", "date", "qStateVer",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions", "outroR1_RD_Review",
                     "qcoe1R1_RD_Review", "LangAssess", "_Pasted")


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb) -> dict[str, dict]:
    dm = wb["Datamap"]
    fields: dict[str, dict] = {}
    current: str | None = None
    for a, b, c in dm.iter_rows(values_only=True):
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            field_name = a_s.split("]", 1)[0][1:]
            qtext = a_s.split("]", 1)[1].lstrip(": ").strip()
            current = field_name
            fields[current] = {"question_text": qtext, "value_labels": {}, "subfield_labels": {}}
            continue
        m = re.match(r"^(q\w+):\s*(.*)", a_s)
        if m:
            current = m.group(1)
            fields[current] = {"question_text": m.group(2), "value_labels": {}, "subfield_labels": {}}
            continue
        if current and b is not None and c is not None:
            b_s, c_s = clean(b), clean(c)
            if b_s.startswith("[") and "]" in b_s:
                subfield = b_s.split("]")[0][1:]
                fields[current]["subfield_labels"][subfield] = c_s
            elif b_s and c_s and not b_s.startswith("Values:"):
                fields[current]["value_labels"][b_s] = c_s
    return fields


def stage_dataset(filepath: Path, output_dir: Path) -> Path:
    """Stage a dataset into agent-ready packets. Returns the output directory."""
    filename = filepath.stem
    print(f"  Staging {filename}...")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fields = load_datamap(wb)
    ws = wb["A1"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Load all rows
    all_rows = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid"))
        if not rid:
            continue
        raw["_rid"] = rid
        raw["_xrow"] = xrow
        all_rows.append(raw)

    # Compute population statistics
    qtimes = [float(r.get("qtime", 0)) for r in all_rows if r.get("qtime")]
    qtime_median = statistics.median(qtimes) if qtimes else 0
    qtime_p10 = statistics.quantiles(qtimes, n=10)[0] if len(qtimes) >= 10 else 0
    qtime_p25 = statistics.quantiles(qtimes, n=4)[0] if len(qtimes) >= 4 else 0
    qtime_p75 = statistics.quantiles(qtimes, n=4)[2] if len(qtimes) >= 4 else 0
    qtime_p90 = statistics.quantiles(qtimes, n=10)[8] if len(qtimes) >= 10 else 0

    suppliers = Counter(clean(str(r.get("SUPNAME"))) or "MISSING" for r in all_rows)

    # Duplicate text detection (exact match on substantive open-end text)
    open_text_hashes = defaultdict(list)
    for r in all_rows:
        for fname, finfo in fields.items():
            if excluded(fname) or fname not in headers:
                continue
            val = r.get(fname)
            if val in (None, ""):
                continue
            # Only check text fields (open-ends)
            text = clean(val)
            if len(text) > 15:  # Only substantive text
                h = hashlib.md5(text.lower().encode()).hexdigest()
                open_text_hashes[h].append((r["_rid"], fname, text))

    duplicate_groups = {}
    for h, entries in open_text_hashes.items():
        if len(entries) > 1:
            duplicate_groups[h] = entries

    # Build per-respondent duplicate membership
    respondent_duplicates = defaultdict(list)
    for h, entries in duplicate_groups.items():
        rids = [e[0] for e in entries]
        for rid, fname, text in entries:
            respondent_duplicates[rid].append({
                "field": fname,
                "text": text[:100],
                "shared_with_count": len(rids),
                "shared_with_ids": [r for r in rids if r != rid][:10],
            })

    # Build staged packets — raw data with value labels looked up
    packets = []
    for r in all_rows:
        rid = r["_rid"]
        qt = float(r["qtime"]) if r.get("qtime") else None

        # Timing percentile
        timing_percentile = None
        if qt:
            if qt < qtime_p10: timing_percentile = "bottom_10"
            elif qt < qtime_p25: timing_percentile = "bottom_25"
            elif qt < qtime_median: timing_percentile = "below_median"
            elif qt < qtime_p75: timing_percentile = "above_median"
            elif qt < qtime_p90: timing_percentile = "top_25"
            else: timing_percentile = "top_10"

        # Build answer chain with question text and value labels
        answer_chain = []
        for fname, finfo in fields.items():
            if excluded(fname) or fname not in headers:
                continue
            qtext = finfo["question_text"]
            labels = finfo["value_labels"]
            sub_labels = finfo["subfield_labels"]

            # Check if this is a multi-select field (has subfield labels)
            if sub_labels:
                checked_items = []
                for h in headers:
                    if not h.startswith(fname + "r"):
                        continue
                    if h.endswith("oe") or h.endswith("oth"):
                        continue
                    val = str(r.get(h, "")).strip()
                    item_label = sub_labels.get(h, h)
                    if val == "1":
                        checked_items.append(item_label)
                if checked_items:
                    answer_chain.append({
                        "field": fname,
                        "question_text": qtext,
                        "answer_type": "multi_select",
                        "checked_items": checked_items,
                    })
                # Check for open-end subfields
                for h in headers:
                    if h.startswith(fname) and (h.endswith("oe") or h.endswith("oth")):
                        val = clean(r.get(h))
                        if val:
                            answer_chain.append({
                                "field": h,
                                "question_text": qtext + " (other specify)",
                                "answer_type": "open_text",
                                "text": val,
                            })
            else:
                val = r.get(fname)
                if val in (None, ""):
                    continue
                sv = str(val).strip()
                label = labels.get(sv, sv)
                # Determine if this is open text or coded
                is_open = fname.endswith("oe") or fname.endswith("othr1") or fname == "outro" or fname.startswith("qcoe")
                if is_open:
                    answer_chain.append({
                        "field": fname,
                        "question_text": qtext,
                        "answer_type": "open_text",
                        "text": clean(val),
                    })
                else:
                    answer_chain.append({
                        "field": fname,
                        "question_text": qtext,
                        "answer_type": "coded",
                        "raw_value": sv,
                        "label": label,
                    })

        packet = {
            "respondent_id": rid,
            "source_excel_row": r["_xrow"],
            "qtime_seconds": qt,
            "qtime_minutes": round(qt / 60, 1) if qt else None,
            "timing_percentile": timing_percentile,
            "supplier": r.get("SUPNAME"),
            "supplier_missing": not r.get("SUPNAME") or clean(str(r.get("SUPNAME"))) == "",
            "duplicate_memberships": respondent_duplicates.get(rid, []),
            "answer_chain": answer_chain,
            "answer_count": len(answer_chain),
        }
        packets.append(packet)

    # Save
    ds_out = output_dir / filename
    ds_out.mkdir(parents=True, exist_ok=True)

    pop_stats = {
        "total_rows": len(packets),
        "timing": {
            "median_seconds": round(qtime_median, 1),
            "median_minutes": round(qtime_median / 60, 1),
            "p10_seconds": round(qtime_p10, 1),
            "p25_seconds": round(qtime_p25, 1),
            "p75_seconds": round(qtime_p75, 1),
            "p90_seconds": round(qtime_p90, 1),
        },
        "supplier_distribution": dict(suppliers.most_common(15)),
        "duplicate_text_groups": len(duplicate_groups),
        "rows_in_duplicate_groups": len(respondent_duplicates),
    }

    (ds_out / "population_stats.json").write_text(json.dumps(pop_stats, indent=2))
    (ds_out / "staged_packets.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in packets) + "\n"
    )

    print(f"    {len(packets)} rows staged, {len(answer_chain)} avg answers/row")
    print(f"    Timing median: {qtime_median:.0f}s ({qtime_median/60:.1f}min)")
    print(f"    Duplicate groups: {len(duplicate_groups)} affecting {len(respondent_duplicates)} rows")

    wb.close()
    return ds_out


def main():
    import sys
    output_dir = OUTPUT_BASE
    output_dir.mkdir(parents=True, exist_ok=True)

    if len(sys.argv) > 1 and sys.argv[1] == "--all":
        datasets = sorted([f for f in DATA_DIR.iterdir() if f.suffix == ".xlsx" and not f.name.startswith(".~")])
        for ds in datasets:
            stage_dataset(ds, output_dir)
    else:
        ds_name = sys.argv[1] if len(sys.argv) > 1 else "106-2502 Delta Water Filtration.xlsx"
        stage_dataset(DATA_DIR / ds_name, output_dir)


if __name__ == "__main__":
    main()
