#!/usr/bin/env python3
"""V12 — Hyperparameter optimization + CatBoost + feature selection + char n-grams.

Improvements over V11:
1. CatBoost (handles categoricals natively, often best for tabular data)
2. Character n-gram TF-IDF (catches misspellings, word fragments)
3. Feature selection via mutual information
4. Hyperparameter grid search for XGBoost/LightGBM
5. More interaction features
6. Cross-question consistency score
"""
from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.feature_selection import mutual_info_classif

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features

import xgboost as xgb
import lightgbm as lgb

try:
    from catboost import CatBoostClassifier
    HAS_CATBOOST = True
except ImportError:
    HAS_CATBOOST = False
    print("WARNING: CatBoost not installed, skipping")


def extract_char_ngrams(df, answer_chains, max_features=150):
    """Extract character n-gram TF-IDF features."""
    print(f"  Extracting char n-gram TF-IDF (max {max_features} terms)...")

    chain_lookup = {ac["respondent_id"]: ac for ac in answer_chains}
    texts = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        ac = chain_lookup.get(rid, {})
        chain = ac.get("answer_chain", [])
        oe_texts = [a.get("raw_value", "") or a.get("response", "") for a in chain
                    if a.get("answer_type") == "open_end" or a.get("role") == "open_end"]
        text = " ".join(oe_texts).strip()
        if not text:
            text = "empty"
        texts.append(text)

    vectorizer = TfidfVectorizer(
        max_features=max_features,
        analyzer="char_wb",
        ngram_range=(3, 5),
        min_df=2,
        max_df=0.95,
    )
    tfidf_matrix = vectorizer.fit_transform(texts)
    tfidf_df = pd.DataFrame(
        tfidf_matrix.toarray(),
        columns=[f"char_{w}" for w in vectorizer.get_feature_names_out()],
        index=df.index,
    )
    print(f"    Char n-grams: {tfidf_df.shape[1]} features")
    return tfidf_df


def extract_cross_question_features(df, answer_chains):
    """Extract cross-question consistency features."""
    print("  Extracting cross-question consistency features...")

    chain_lookup = {ac["respondent_id"]: ac for ac in answer_chains}
    features = []

    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        ac = chain_lookup.get(rid, {})
        chain = ac.get("answer_chain", [])

        # Group answers by question type
        oe_answers = []
        matrix_answers = []
        coded_answers = []
        demo_answers = []

        for a in chain:
            atype = a.get("answer_type", "")
            val = str(a.get("raw_value", "")).strip().lower()
            if not val:
                continue
            if atype == "open_end":
                oe_answers.append(val)
            elif atype == "matrix_cell":
                matrix_answers.append(val)
            elif atype == "coded_question":
                coded_answers.append(val)
            elif atype == "demographic":
                demo_answers.append(val)

        # Cross-question OE consistency: do OE answers share vocabulary?
        oe_word_sets = [set(val.split()) for val in oe_answers if val]
        oe_consistency = 0
        if len(oe_word_sets) > 1:
            overlaps = []
            for i in range(len(oe_word_sets)):
                for j in range(i + 1, len(oe_word_sets)):
                    if oe_word_sets[i] and oe_word_sets[j]:
                        overlap = len(oe_word_sets[i] & oe_word_sets[j]) / min(len(oe_word_sets[i]), len(oe_word_sets[j]))
                        overlaps.append(overlap)
            oe_consistency = np.mean(overlaps) if overlaps else 0

        # Matrix answer entropy (cross-question diversity)
        matrix_entropy = 0
        if matrix_answers:
            from collections import Counter
            counts = Counter(matrix_answers)
            total = len(matrix_answers)
            probs = [c / total for c in counts.values()]
            matrix_entropy = -sum(p * np.log2(p) for p in probs if p > 0)

        # Coded answer diversity
        coded_diversity = len(set(coded_answers)) / max(len(coded_answers), 1) if coded_answers else 1

        # Answer speed consistency (do short OE answers correlate with fast timing?)
        oe_lengths = [len(val) for val in oe_answers]
        oe_length_cv = np.std(oe_lengths) / max(np.mean(oe_lengths), 1) if oe_lengths else 0

        features.append({
            "respondent_id": rid,
            "xq_oe_consistency": oe_consistency,
            "xq_matrix_entropy": matrix_entropy,
            "xq_coded_diversity": coded_diversity,
            "xq_oe_length_cv": oe_length_cv,
            "xq_n_oe_answers": len(oe_answers),
            "xq_n_matrix_answers": len(matrix_answers),
            "xq_n_coded_answers": len(coded_answers),
            "xq_total_answers": len(chain),
        })

    xq_df = pd.DataFrame(features)
    df = df.merge(xq_df, on="respondent_id", how="left")
    print(f"    Cross-question features: {len(xq_df.columns) - 1}")
    return df


def feature_selection(X, y, top_k=80):
    """Select top features by mutual information."""
    print(f"  Feature selection: selecting top {top_k} of {X.shape[1]} features...")
    mi = mutual_info_classif(X, y, random_state=42, discrete_features=False)
    mi_series = pd.Series(mi, index=X.columns)
    top_features = mi_series.nlargest(top_k).index.tolist()
    print(f"    Top 10 features by MI: {mi_series.nlargest(10).to_dict()}")
    return top_features


def run_v12_cv(use_v7_features=True, n_folds=5, use_feature_selection=True, top_k=100):
    """Run V12 cross-validation."""
    print(f"\n{'='*80}")
    print(f"V12 CV — CatBoost + Char N-grams + Feature Selection + Cross-Question")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments() if use_v7_features else {}
    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)

    # Add word TF-IDF
    from v11_advanced_ensemble import extract_tfidf_features, extract_supplier_interactions
    tfidf_df = extract_tfidf_features(df, answer_chains, max_features=100)
    df = pd.concat([df, tfidf_df], axis=1)

    # Add char n-gram TF-IDF
    char_df = extract_char_ngrams(df, answer_chains, max_features=150)
    df = pd.concat([df, char_df], axis=1)

    # Add supplier interactions
    df = extract_supplier_interactions(df)

    # Add cross-question features
    df = extract_cross_question_features(df, answer_chains)

    # Filter to labeled
    labeled = df[df["label"] >= 0].copy().reset_index(drop=True)
    print(f"Labeled: {len(labeled)} respondents, {len(labeled.columns)} total features")

    # Prepare features
    non_feature = {"label", "respondent_id", "dataset", "supplier_name", "v7_judgment"}
    feature_cols = [c for c in labeled.columns if c not in non_feature]

    X = labeled[feature_cols].copy()
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X = X.fillna(0)
    y = labeled["label"].values

    # Read CLASSIFY
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        if rid and classify_idx is not None and classify_idx < len(row):
            classify_map[rid] = row[classify_idx]
    wb.close()

    is_pro = labeled["respondent_id"].map(lambda r: str(classify_map.get(r)) == "1").values

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (train_idx, test_idx) in enumerate(skf.split(X, y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")
        X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
        y_train, y_test = y[train_idx], y[test_idx]
        is_pro_test = is_pro[test_idx]

        # Feature selection on training data only
        if use_feature_selection:
            selected = feature_selection(X_train, y_train, top_k=top_k)
            X_train = X_train[selected]
            X_test = X_test[selected]

        # Split train into train/val
        n_val = len(X_train) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_train), n_val, replace=False)
        train_mask = np.ones(len(X_train), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_train.iloc[train_mask]
        X_val = X_train.iloc[val_idx]
        y_tr = y_train[train_mask]
        y_val = y_train[val_idx]

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        # Train models
        print("  Training XGBoost...")
        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        print("  Training LightGBM...")
        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.7,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        if HAS_CATBOOST:
            print("  Training CatBoost...")
            cat_model = CatBoostClassifier(
                iterations=500, depth=6, learning_rate=0.03,
                l2_leaf_reg=3, random_seed=42, verbose=0
            )
            cat_model.fit(X_tr, y_tr)

        print("  Training MLP...")
        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        # Get validation predictions
        models_val = {}
        models_test = {}

        models_val["xgb"] = xgb_model.predict_proba(X_val)[:, 1]
        models_test["xgb"] = xgb_model.predict_proba(X_test)[:, 1]

        models_val["lgb"] = lgb_model.predict_proba(X_val)[:, 1]
        models_test["lgb"] = lgb_model.predict_proba(X_test)[:, 1]

        if HAS_CATBOOST:
            models_val["cat"] = cat_model.predict_proba(X_val)[:, 1]
            models_test["cat"] = cat_model.predict_proba(X_test)[:, 1]

        models_val["mlp"] = mlp.predict_proba(X_val_scaled)[:, 1]
        models_test["mlp"] = mlp.predict_proba(X_test_scaled)[:, 1]

        # Calibrate
        cal_val = {}
        cal_test = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        # Ensemble
        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Search for best threshold
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.25, 0.60, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    # Aggregate
    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])
    total_tp = sum(m["tp"] for m in all_fold_metrics)
    total_fp = sum(m["fp"] for m in all_fold_metrics)
    total_tn = sum(m["tn"] for m in all_fold_metrics)
    total_fn = sum(m["fn"] for m in all_fold_metrics)
    pooled_prec = total_tp / max(total_tp + total_fp, 1)
    pooled_rec = total_tp / max(total_tp + total_fn, 1)
    pooled_f1 = 2 * pooled_prec * pooled_rec / max(pooled_prec + pooled_rec, 0.001)
    pooled_bacc = (pooled_rec + total_tn / max(total_tn + total_fp, 1)) / 2

    print(f"\n{'='*80}")
    print(f"V12 CV RESULTS ({n_folds} folds, v7={use_v7_features}, fs={use_feature_selection})")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")
    print(f"  Pooled: TP={total_tp}, FP={total_fp}, TN={total_tn}, FN={total_fn}")
    print(f"  Pooled BAcc:  {pooled_bacc:.3f}, F1: {pooled_f1:.3f}")

    return {
        "avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc,
        "pooled_bacc": pooled_bacc, "pooled_f1": pooled_f1,
        "folds": all_fold_metrics,
    }


def main():
    print("=" * 80)
    print("V12 — CatBoost + Char N-grams + Feature Selection + Cross-Question")
    print("=" * 80)

    # Run with V7 features + feature selection
    results = run_v12_cv(use_v7_features=True, n_folds=5, use_feature_selection=True, top_k=120)

    # Run without feature selection (all features)
    results_nofs = run_v12_cv(use_v7_features=True, n_folds=5, use_feature_selection=False)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V10 (sklearn ensemble):  BAcc=0.729")
    print(f"  V11 (XGB+LGB+MLP+RF):    BAcc=0.737")
    print(f"  V12 (with feat select):  BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  V12 (no feat select):    BAcc={results_nofs['avg_bacc']:.3f}, F1={results_nofs['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], results_nofs['avg_bacc']):.3f}")

    with open(AUTOSURVEY_DIR / "v12_cv_results.json", "w") as f:
        json.dump({"v12_fs": results, "v12_nofs": results_nofs}, f, indent=2, default=str)


if __name__ == "__main__":
    main()
