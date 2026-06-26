#!/usr/bin/env python3
"""Compare pipeline output against annotated ground truth and extract misses for agent review.

Loads an annotated dataset (with status column), runs the pipeline, compares
our discards vs client discards (status=5), and writes answer chains for
false negatives (missed discards) and false positives (we discarded, client kept)
to JSON files for subagent review.

Usage:
    python3 compare_to_ground_truth.py <annotated_xlsx> <pipeline_output_dir>
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd

# Add the skills scripts directory
SKILL_SCRIPTS = Path(__file__).parent.parent / "skills" / "cleaning-survey-quality" / "scripts"
sys.path.insert(0, str(SKILL_SCRIPTS))

from survey_pipeline import (
    extract_features_and_chain,
    ml_triage,
    agent_score_respondent,
    reassess_respondent,
    compute_key_signals,
    clean,
    norm,
    parse_datamap,
    classify_field,
)


def load_ground_truth(filepath):
    """Load status labels from an annotated dataset.
    
    status=3 → accepted (KEEP)
    status=5 → rejected (DISCARD)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    rid_col = hidx.get("uuid") or hidx.get("record")
    status_col = hidx.get("status")
    
    if status_col is None:
        print("WARNING: No 'status' column found — file may not be annotated")
        return {}

    ground_truth = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = clean(row[rid_col]) if rid_col and rid_col < len(row) else ""
        status = row[status_col] if status_col and status_col < len(row) else None
        if rid and status is not None:
            try:
                status = int(status)
            except:
                pass
            ground_truth[rid] = status
    
    return ground_truth


def run_comparison(filepath, output_dir):
    """Run pipeline on annotated data and compare against ground truth."""
    filepath = Path(filepath)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*80}")
    print(f"PIPELINE vs GROUND TRUTH COMPARISON")
    print(f"{'='*80}")
    print(f"  Input: {filepath.name}")

    # Load ground truth
    gt = load_ground_truth(filepath)
    n_accepted = sum(1 for s in gt.values() if s == 3)
    n_rejected = sum(1 for s in gt.values() if s == 5)
    n_other = sum(1 for s in gt.values() if s not in (3, 5))
    print(f"  Ground truth: {len(gt)} respondents")
    print(f"    Accepted (status=3): {n_accepted}")
    print(f"    Rejected (status=5): {n_rejected}")
    print(f"    Other: {n_other}")

    # Run pipeline
    print(f"\n[1/4] Extracting features and answer chains...")
    df, datamap, roles, answer_chains = extract_features_and_chain(filepath)
    
    print(f"\n[2/4] ML triage + agent scoring...")
    df = ml_triage(df)

    # Compute matrix prevalence for gating
    matrix_prevalence = (df["matrix_straightline"] == 1).mean() if "matrix_straightline" in df.columns else None
    if matrix_prevalence is not None:
        print(f"  Matrix straightlining prevalence: {matrix_prevalence:.1%} {'(GATED)' if matrix_prevalence > 0.8 else ''}")

    agent_scores = []
    agent_reasons = []
    for idx, row in df.iterrows():
        chain = answer_chains[idx] if idx < len(answer_chains) else {}
        score, reasons = agent_score_respondent(chain, row["ml_triage_score"], matrix_prevalence=matrix_prevalence)
        agent_scores.append(score)
        agent_reasons.append(reasons)
    df["agent_score"] = agent_scores
    df["agent_reasons"] = agent_reasons

    # Reassessment
    final_scores = []
    final_judgments = []
    reassessment_notes_all = []
    for idx, row in df.iterrows():
        if row["agent_score"] < 0:
            chain = answer_chains[idx] if idx < len(answer_chains) else {}
            final_score, judgment, notes = reassess_respondent(
                chain, row["agent_score"], row["agent_reasons"])
            final_scores.append(final_score)
            final_judgments.append(judgment)
            reassessment_notes_all.append(notes)
        else:
            final_scores.append(row["agent_score"])
            final_judgments.append("KEEP")
            reassessment_notes_all.append(["No reassessment needed — positive score"])

    df["final_score"] = final_scores
    df["final_judgment"] = final_judgments
    df["reassessment_notes"] = reassessment_notes_all
    df["key_signals"] = compute_key_signals(df, answer_chains)

    # Add ground truth to df
    df["client_status"] = df["respondent_id"].map(gt)
    df["client_judgment"] = df["client_status"].map({3: "KEEP", 5: "DISCARD"})

    # Compare
    print(f"\n[3/4] Comparing pipeline vs ground truth...")
    
    # Only compare rows that have ground truth
    df_gt = df[df["client_judgment"].notna()].copy()
    
    # Confusion matrix
    tp = ((df_gt["final_judgment"] == "DISCARD") & (df_gt["client_judgment"] == "DISCARD")).sum()
    fp = ((df_gt["final_judgment"] == "DISCARD") & (df_gt["client_judgment"] == "KEEP")).sum()
    fn = ((df_gt["final_judgment"] != "DISCARD") & (df_gt["client_judgment"] == "DISCARD")).sum()
    tn = ((df_gt["final_judgment"] != "DISCARD") & (df_gt["client_judgment"] == "KEEP")).sum()

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    print(f"\n  Confusion Matrix (n={len(df_gt)}):")
    print(f"    True Positives  (we discard, client discards):  {tp}")
    print(f"    False Positives (we discard, client keeps):     {fp}")
    print(f"    False Negatives (we keep, client discards):     {fn}")
    print(f"    True Negatives  (we keep, client keeps):        {tn}")
    print(f"\n  Precision: {precision:.3f}")
    print(f"  Recall:    {recall:.3f}")
    print(f"  F1:        {f1:.3f}")

    # Also check REVIEW band — how many false negatives are in REVIEW?
    fn_in_review = ((df_gt["final_judgment"] == "REVIEW") & (df_gt["client_judgment"] == "DISCARD")).sum()
    fn_in_keep = ((df_gt["final_judgment"] == "KEEP") & (df_gt["client_judgment"] == "DISCARD")).sum()
    print(f"\n  False negative breakdown:")
    print(f"    In REVIEW band: {fn_in_review} (we flagged but didn't discard)")
    print(f"    In KEEP band:   {fn_in_keep} (we completely missed)")

    # Extract misses for agent review
    print(f"\n[4/4] Extracting misses for agent review...")

    # False negatives: client discarded, we didn't
    fn_df = df_gt[(df_gt["client_judgment"] == "DISCARD") & (df_gt["final_judgment"] != "DISCARD")].copy()
    fn_df = fn_df.sort_values("final_score", ascending=False)  # Worst misses first (highest score = most confident keep)

    # False positives: we discarded, client didn't
    fp_df = df_gt[(df_gt["final_judgment"] == "DISCARD") & (df_gt["client_judgment"] == "KEEP")].copy()
    fp_df = fp_df.sort_values("final_score", ascending=True)  # Worst FPs first (lowest score = most confident discard)

    # True positives: both agree on discard
    tp_df = df_gt[(df_gt["final_judgment"] == "DISCARD") & (df_gt["client_judgment"] == "DISCARD")].copy()

    # Build review packets for misses
    def build_review_packet(row, chain, category):
        return {
            "respondent_id": row["respondent_id"],
            "category": category,  # "false_negative" or "false_positive"
            "client_judgment": row["client_judgment"],
            "our_judgment": row["final_judgment"],
            "our_score": round(row["final_score"], 3),
            "ml_triage_score": round(row["ml_triage_score"], 3),
            "agent_score": round(row["agent_score"], 3),
            "key_signals": row["key_signals"] if isinstance(row["key_signals"], list) else [],
            "agent_reasons": row["agent_reasons"] if isinstance(row["agent_reasons"], list) else [],
            "reassessment_notes": row["reassessment_notes"] if isinstance(row.get("reassessment_notes"), list) else [],
            "supplier": chain.get("supplier", ""),
            "supplier_reject_rate": chain.get("supplier_reject_rate", 0),
            "qtime_minutes": chain.get("qtime_minutes", 0),
            "qtime_percentile": chain.get("qtime_percentile", ""),
            "signal_count": chain.get("signal_count", 0),
            "t1_count": chain.get("t1_count", 0),
            "t2_count": chain.get("t2_count", 0),
            "t3_count": chain.get("t3_count", 0),
            "oe_total_chars": chain.get("oe_total_chars", 0),
            "answer_entropy": round(chain.get("answer_entropy", 0), 2),
            "matrix_unique_ratio": round(chain.get("matrix_unique_ratio", 0), 2),
            "matrix_straightline": chain.get("matrix_straightline", 0),
            "oe_dup_count": chain.get("oe_dup_count", 0),
            "ip_dup_count": chain.get("ip_dup_count", 0),
            "lang_readlevel": chain.get("lang_readlevel", 0),
            "answer_chain": chain.get("answer_chain", []),
            "oe_text": chain.get("oe_text", ""),
        }

    # Write false negatives (missed discards)
    fn_packets = []
    for idx, row in fn_df.iterrows():
        chain = answer_chains[df.index.get_loc(idx)] if idx in df.index else {}
        fn_packets.append(build_review_packet(row, chain, "false_negative"))
    
    fn_path = output_dir / "false_negatives_for_review.json"
    with open(fn_path, "w") as f:
        json.dump(fn_packets, f, indent=2)
    print(f"  False negatives (missed discards): {len(fn_packets)} → {fn_path}")

    # Write false positives (wrong discards)
    fp_packets = []
    for idx, row in fp_df.iterrows():
        chain = answer_chains[df.index.get_loc(idx)] if idx in df.index else {}
        fp_packets.append(build_review_packet(row, chain, "false_positive"))
    
    fp_path = output_dir / "false_positives_for_review.json"
    with open(fp_path, "w") as f:
        json.dump(fp_packets, f, indent=2)
    print(f"  False positives (wrong discards): {len(fp_packets)} → {fp_path}")

    # Write true positives (correct discards) for reference
    tp_packets = []
    for idx, row in tp_df.iterrows():
        chain = answer_chains[df.index.get_loc(idx)] if idx in df.index else {}
        tp_packets.append(build_review_packet(row, chain, "true_positive"))
    
    tp_path = output_dir / "true_positives_for_review.json"
    with open(tp_path, "w") as f:
        json.dump(tp_packets, f, indent=2)
    print(f"  True positives (correct discards): {len(tp_packets)} → {tp_path}")

    # Write summary
    summary = {
        "dataset": filepath.name,
        "total_respondents": int(len(df)),
        "annotated_respondents": int(len(df_gt)),
        "client_discards": int(n_rejected),
        "client_keeps": int(n_accepted),
        "our_discards": int((df_gt["final_judgment"] == "DISCARD").sum()),
        "our_reviews": int((df_gt["final_judgment"] == "REVIEW").sum()),
        "our_keeps": int((df_gt["final_judgment"] == "KEEP").sum()),
        "true_positives": int(tp),
        "false_positives": int(fp),
        "false_negatives": int(fn),
        "true_negatives": int(tn),
        "precision": round(precision, 3),
        "recall": round(recall, 3),
        "f1": round(f1, 3),
        "fn_in_review": int(fn_in_review),
        "fn_in_keep": int(fn_in_keep),
    }
    summary_path = output_dir / "comparison_summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"\n  Summary: {summary_path}")

    # Print summary table
    print(f"\n{'='*80}")
    print(f"COMPARISON SUMMARY")
    print(f"{'='*80}")
    print(f"  Client discards: {n_rejected}")
    print(f"  Our discards:    {(df_gt['final_judgment'] == 'DISCARD').sum()}")
    print(f"  Correct catches: {tp} ({tp/n_rejected*100:.1f}% of client discards)" if n_rejected > 0 else "")
    print(f"  Missed:          {fn} ({fn/n_rejected*100:.1f}% of client discards)" if n_rejected > 0 else "")
    print(f"  Wrong discards:  {fp}")
    print(f"  F1: {f1:.3f}")

    return df, fn_packets, fp_packets, tp_packets, summary


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 compare_to_ground_truth.py <annotated_xlsx> [output_dir]")
        return

    filepath = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else filepath.parent / f"{filepath.stem}_comparison"

    run_comparison(filepath, output_dir)


if __name__ == "__main__":
    main()
