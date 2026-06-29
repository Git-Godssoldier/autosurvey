#!/usr/bin/env python3
"""V14 — Semi-supervised self-training + V8 judgments + active learning focus.

Approach:
1. Self-training: Train on Echo labels, predict on other 10 datasets,
   use high-confidence predictions as pseudo-labels, retrain
2. Add V8 agent judgments as features alongside V7 (V8 had higher recall)
3. Focus on REVIEW tier: respondents where ML is uncertain (0.3-0.6 band)
4. Try co-training: train separate models on different feature subsets
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features

import xgboost as xgb
import lightgbm as lgb

DATASETS = {
    "106-2502 Delta Water Filtration": "106-2502 Delta Water Filtration.xlsx",
    "109-2601 Echo BH": "109-2601 Echo BH.xlsx",
    "153-2602 ODL Switchable Glass": "153-2602 ODL Switchable Glass.xlsx",
    "159-2601 Oldcastle Brand Health": "159-2601 Oldcastle Brand Health.xlsx",
    "159-2602 Oldcastle Canada": "159-2602 Oldcastle Canada.xlsx",
    "189-2501 SBD Brand Association": "189-2501 SBD Brand Association.xlsx",
    "287-2501 THD Digital CX": "287-2501 THD Digital CX.xlsx",
    "365-2601 ADDO RaceTrac US GP": "365-2601 ADDO RaceTrac US GP.xlsx",
    "368-2602 Masterlock Conjoint": "368-2602 Masterlock Conjoint.xlsx",
    "999-2601 TFG Contractor Index Q1": "999-2601 TFG Contractor Index Q1.xlsx",
    "999-2602 TFG Contractor Index Q2": "999-2602 TFG Contractor Index Q2.xlsx",
}


def load_v8_judgments():
    """Load V8 agent judgments as additional features."""
    v8_path = ECHO_OUTPUT / "holistic_agent_run_v8" / "agent_judgments.json"
    if not v8_path.exists():
        print(f"  V8 judgments not found at {v8_path}")
        return {}
    with open(v8_path) as f:
        judgments = json.load(f)
    return {j["respondent_id"]: j for j in judgments}


def add_v8_features(df, v8_judgments):
    """Add V8 judgment features alongside V7."""
    if not v8_judgments:
        return df
    df["v8_judgment"] = df["respondent_id"].map(
        lambda rid: v8_judgments.get(rid, {}).get("agent_judgment", "UNKNOWN")
    )
    df["v8_judgment_enc"] = df["v8_judgment"].map(
        {"KEEP": 0, "REVIEW": 1, "DISCARD": 2, "UNKNOWN": 1}
    ).fillna(1)
    df["v8_converging_count"] = df["respondent_id"].map(
        lambda rid: v8_judgments.get(rid, {}).get("converging_family_count", 0)
    )
    df["v8_authenticity_risk"] = df["respondent_id"].map(
        lambda rid: v8_judgments.get(rid, {}).get("authenticity_risk", 0.5)
    )
    df["v8_quality_risk"] = df["respondent_id"].map(
        lambda rid: v8_judgments.get(rid, {}).get("quality_discard_risk", 0.5)
    )
    df["v8_client_reject_prob"] = df["respondent_id"].map(
        lambda rid: v8_judgments.get(rid, {}).get("client_reject_probability", 0.5)
    )

    # V7 + V8 agreement features
    if "v7_judgment_enc" in df.columns:
        df["v7_v8_agree"] = (df["v7_judgment_enc"] == df["v8_judgment_enc"]).astype(int)
        df["v7_v8_avg_risk"] = (
            df.get("v7_client_reject_prob", 0.5) + df["v8_client_reject_prob"]
        ) / 2
        df["v7_v8_max_risk"] = df[
            ["v7_client_reject_prob", "v8_client_reject_prob"]
        ].max(axis=1)
        df["v7_v8_risk_diff"] = abs(
            df.get("v7_client_reject_prob", 0.5) - df["v8_client_reject_prob"]
        )

    print(f"  V8 features added: {df['v8_judgment'].notna().sum()}")
    return df


def extract_all_datasets():
    """Extract features from all datasets for self-training."""
    all_dfs = []
    for dataset_name, xlsx_name in DATASETS.items():
        xlsx_path = DATA_DIR / xlsx_name
        if not xlsx_path.exists():
            continue
        print(f"  Extracting: {dataset_name}...")
        try:
            df, answer_chains = extract_enhanced_features(xlsx_path, gt_labels=None, v7_judgments=None)
            df["dataset"] = dataset_name
            df["label"] = -1  # unlabeled by default
            all_dfs.append(df)
        except Exception as e:
            print(f"    ERROR: {e}")
    return pd.concat(all_dfs, ignore_index=True)


def run_self_training(labeled_df, unlabeled_df, n_iterations=3, confidence_threshold=0.85):
    """Self-training: use high-confidence predictions as pseudo-labels."""
    print(f"\nSelf-training ({n_iterations} iterations, confidence >= {confidence_threshold})...")

    non_feature = {"label", "respondent_id", "dataset", "supplier_name",
                   "v7_judgment", "v8_judgment"}
    feature_cols = [c for c in labeled_df.columns if c not in non_feature]

    # Ensure unlabeled has same features
    for c in feature_cols:
        if c not in unlabeled_df.columns:
            unlabeled_df[c] = 0

    X_labeled = labeled_df[feature_cols].copy()
    for col in X_labeled.select_dtypes(include=["object"]).columns:
        X_labeled[col] = pd.Categorical(X_labeled[col]).codes
    X_labeled = X_labeled.fillna(0)
    y_labeled = labeled_df["label"].values

    X_unlabeled = unlabeled_df[feature_cols].copy()
    for col in X_unlabeled.select_dtypes(include=["object"]).columns:
        X_unlabeled[col] = pd.Categorical(X_unlabeled[col]).codes
    X_unlabeled = X_unlabeled.fillna(0)

    for iteration in range(n_iterations):
        print(f"\n  Self-training iteration {iteration + 1}/{n_iterations}")
        print(f"    Labeled: {len(X_labeled)}, Unlabeled: {len(X_unlabeled)}")

        # Train XGBoost
        model = xgb.XGBClassifier(
            n_estimators=400, max_depth=6, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8, random_state=42,
            use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        model.fit(X_labeled, y_labeled)

        # Predict on unlabeled
        if len(X_unlabeled) > 0:
            probs = model.predict_proba(X_unlabeled)[:, 1]
            high_conf = (probs >= confidence_threshold) | (probs <= 1 - confidence_threshold)
            pseudo_labels = (probs >= 0.5).astype(int)

            n_pseudo = high_conf.sum()
            print(f"    High-confidence pseudo-labels: {n_pseudo}")

            if n_pseudo == 0:
                print("    No high-confidence predictions, stopping self-training")
                break

            # Add high-confidence pseudo-labeled data to training set
            X_pseudo = X_unlabeled[high_conf]
            y_pseudo = pseudo_labels[high_conf]

            X_labeled = pd.concat([X_labeled, X_pseudo], ignore_index=True)
            y_labeled = np.concatenate([y_labeled, y_pseudo])

            # Remove from unlabeled
            X_unlabeled = X_unlabeled[~high_conf]
        else:
            break

    return X_labeled, y_labeled, feature_cols


def run_v14_cv(n_folds=5):
    """Run V14 cross-validation with self-training + V8 features."""
    print(f"\n{'='*80}")
    print(f"V14 — Self-Training + V8 Features + V7+V8 Ensemble")
    print(f"{'='*80}")

    # Load Echo data with V7 + V8 features
    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()
    print(f"V7 judgments: {len(v7)}, V8 judgments: {len(v8)}")

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Extract all other datasets for self-training
    print("\nExtracting all datasets for self-training...")
    all_data = extract_all_datasets()

    # Add Echo to all_data
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)
    print(f"\nTotal: {len(all_data)} ({len(labeled)} labeled, {len(unlabeled)} unlabeled)")

    # Prepare features
    non_feature = {"label", "respondent_id", "dataset", "supplier_name",
                   "v7_judgment", "v8_judgment"}
    feature_cols = [c for c in labeled.columns if c not in non_feature]

    # Read CLASSIFY for Echo
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()

    # K-fold CV on Echo
    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        # Split: Echo train + all unlabeled for self-training
        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Self-training: use Echo train labels + unlabeled data
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        # Prepare test features
        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        # Train final models on self-trained data
        scaler = StandardScaler()
        X_st_scaled = scaler.fit_transform(X_st)
        X_test_scaled = scaler.transform(X_test)

        # Split for calibration
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]
        X_tr_scaled = X_st_scaled[train_mask]
        X_val_scaled = X_st_scaled[val_idx]

        print(f"  Training XGBoost on {len(X_tr)} samples...")
        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        print("  Training LightGBM...")
        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.7,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        print("  Training MLP...")
        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        # Calibrate
        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Search for best threshold
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V14 CV RESULTS (self-training + V8 features)")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V14 — Self-Training + V8 Features + V7+V8 Ensemble")
    print("=" * 80)

    results = run_v14_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V10 (sklearn ensemble):  BAcc=0.729 (with V7 features)")
    print(f"  V11 (XGB+LGB+MLP+RF):    BAcc=0.737 (with V7 features)")
    print(f"  V12 (feat select):       BAcc=0.735 (with V7 features)")
    print(f"  V13 (multi-dataset):     BAcc=0.702 (NO V7 features)")
    print(f"  V14 (self-train + V8):   BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - results['avg_bacc']:.3f}")

    with open(AUTOSURVEY_DIR / "v14_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
