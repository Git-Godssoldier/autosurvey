#!/usr/bin/env python3
"""Analyze raw Excel data to find signals that distinguish FN from TN."""
import json
import openpyxl
from collections import Counter, defaultdict
import statistics

def main():
    # Load V7 judgments
    with open('/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run_v7/agent_judgments.json') as f:
        v7 = {j['respondent_id']: j for j in json.load(f)}

    # Load ground truth
    gt_path = '/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx'
    wb = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    hidx = {h: i for i, h in enumerate(headers) if h}

    gt = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx['uuid']]).strip() if row[hidx['uuid']] else None
        status = row[hidx['status']]
        if rid:
            if status == 5 or status == '5':
                gt[rid] = 'DISCARD'
            elif status == 3 or status == '3':
                gt[rid] = 'KEEP'
    wb.close()

    # Classify
    tp_ids, fp_ids, tn_ids, fn_ids = set(), set(), set(), set()
    for rid, j in v7.items():
        if rid not in gt: continue
        if j['agent_judgment'] == 'DISCARD' and gt[rid] == 'DISCARD': tp_ids.add(rid)
        elif j['agent_judgment'] == 'DISCARD' and gt[rid] == 'KEEP': fp_ids.add(rid)
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'DISCARD': fn_ids.add(rid)
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'KEEP': tn_ids.add(rid)

    print(f"TP={len(tp_ids)}, FP={len(fp_ids)}, TN={len(tn_ids)}, FN={len(fn_ids)}")

    # Load raw data from client annotated file (has all columns)
    wb = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)
    ws = wb.active

    # Columns to analyze
    cols_to_check = [
        'source', 'vlist', 'CLASSIFY', 'REGION', 'qtime',
        'LangAssessReadLevel', 'LangAssessReadEase', 'LangAssessNumWords',
        'TERMFLAGS', 'qc5',
        'conditionsAriens', 'conditionsHD_or_OPE_dealers', 'conditionsOther_channel',
        'age', 'qager1',
        'RD_Searchr0', 'RD_Searchr1', 'RD_Searchr2', 'RD_Searchr3',
        'RD_GetTokenr0', 'RD_GetTokenr1',
        'PROAGE', 'CONAGE',
    ]

    # Also check POSSIBLEBRANDS columns
    brand_cols = [h for h in headers if h and 'POSSIBLEBRANDS' in str(h)]
    cols_to_check.extend(brand_cols[:3])

    # Also check q14 OE columns
    oe_cols = ['q14r6oe', 'q14r9oe', 'q14r14oe', 'q14r15oe', 'q20r15oe']
    cols_to_check.extend(oe_cols)

    # Filter to existing columns
    cols_to_check = [c for c in cols_to_check if c in hidx]

    # Load raw data for all respondents
    raw_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx['uuid']]).strip() if row[hidx['uuid']] else None
        if rid and rid in v7:
            raw_data[rid] = {c: row[hidx[c]] for c in cols_to_check}
    wb.close()

    print(f"\nLoaded raw data for {len(raw_data)} respondents")

    # Analyze each column for FN vs TN discrimination
    print(f"\n{'='*80}")
    print(f"RAW SIGNAL ANALYSIS: FN vs TN")
    print(f"{'='*80}")

    for col in cols_to_check:
        fn_vals = [raw_data[rid][col] for rid in fn_ids if rid in raw_data and raw_data[rid][col] is not None]
        tn_vals = [raw_data[rid][col] for rid in tn_ids if rid in raw_data and raw_data[rid][col] is not None]
        tp_vals = [raw_data[rid][col] for rid in tp_ids if rid in raw_data and raw_data[rid][col] is not None]

        if not fn_vals or not tn_vals:
            continue

        # Check if numeric or categorical
        is_numeric = all(isinstance(v, (int, float)) for v in fn_vals + tn_vals if v is not None)

        if is_numeric:
            fn_mean = statistics.mean(fn_vals) if fn_vals else 0
            tn_mean = statistics.mean(tn_vals) if tn_vals else 0
            tp_mean = statistics.mean(tp_vals) if tp_vals else 0
            gap = fn_mean - tn_mean
            if abs(gap) > 0.01 or col in ['qtime', 'LangAssessReadLevel', 'LangAssessReadEase', 'LangAssessNumWords']:
                print(f"\n  {col} (numeric):")
                print(f"    FN mean={fn_mean:.2f}, TN mean={tn_mean:.2f}, TP mean={tp_mean:.2f}, gap={gap:+.3f}")
        else:
            fn_counter = Counter(str(v)[:40] for v in fn_vals)
            tn_counter = Counter(str(v)[:40] for v in tn_vals)
            # Find values that are much more common in FN than TN
            print(f"\n  {col} (categorical):")
            all_vals = set(list(fn_counter.keys())[:10] + list(tn_counter.keys())[:10])
            for val in sorted(all_vals):
                fn_pct = fn_counter.get(val, 0) / len(fn_vals) * 100
                tn_pct = tn_counter.get(val, 0) / len(tn_vals) * 100
                if abs(fn_pct - tn_pct) > 2.0 or fn_pct > 5:
                    print(f"    '{val}': FN={fn_pct:.1f}%, TN={tn_pct:.1f}%, gap={fn_pct-tn_pct:+.1f}%")

    # Special analysis: source discrimination
    print(f"\n{'='*80}")
    print(f"SOURCE ANALYSIS (supplier-level)")
    print(f"{'='*80}")

    source_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'tn': 0, 'fn': 0})
    for rid in v7:
        if rid not in gt or rid not in raw_data: continue
        src = raw_data[rid].get('source', 'unknown')
        src = str(src)[:30] if src else 'unknown'
        j = v7[rid]
        if j['agent_judgment'] == 'DISCARD' and gt[rid] == 'DISCARD': source_stats[src]['tp'] += 1
        elif j['agent_judgment'] == 'DISCARD' and gt[rid] == 'KEEP': source_stats[src]['fp'] += 1
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'DISCARD': source_stats[src]['fn'] += 1
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'KEEP': source_stats[src]['tn'] += 1

    print(f"\n{'Source':<30} {'Total':>6} {'Client%':>8} {'Agent%':>8} {'FN':>5} {'FP':>5} {'Miss%':>8}")
    for src in sorted(source_stats.keys(), key=lambda x: sum(source_stats[x].values()), reverse=True):
        s = source_stats[src]
        total = s['tp'] + s['fp'] + s['tn'] + s['fn']
        if total < 10: continue
        client_discard = s['tp'] + s['fn']
        agent_discard = s['tp'] + s['fp']
        client_pct = client_discard / total * 100
        agent_pct = agent_discard / total * 100
        miss_pct = s['fn'] / total * 100
        print(f"  {src:<28} {total:>6} {client_pct:>7.1f}% {agent_pct:>7.1f}% {s['fn']:>5} {s['fp']:>5} {miss_pct:>7.1f}%")

    # Special analysis: CLASSIFY discrimination
    print(f"\n{'='*80}")
    print(f"CLASSIFY ANALYSIS")
    print(f"{'='*80}")

    classify_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'tn': 0, 'fn': 0})
    for rid in v7:
        if rid not in gt or rid not in raw_data: continue
        cls = raw_data[rid].get('CLASSIFY', 'unknown')
        cls = str(cls)[:20] if cls else 'unknown'
        j = v7[rid]
        if j['agent_judgment'] == 'DISCARD' and gt[rid] == 'DISCARD': classify_stats[cls]['tp'] += 1
        elif j['agent_judgment'] == 'DISCARD' and gt[rid] == 'KEEP': classify_stats[cls]['fp'] += 1
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'DISCARD': classify_stats[cls]['fn'] += 1
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'KEEP': classify_stats[cls]['tn'] += 1

    print(f"\n{'Class':<20} {'Total':>6} {'Client%':>8} {'Agent%':>8} {'FN':>5} {'FP':>5} {'Prec':>6} {'Rec':>6}")
    for cls in sorted(classify_stats.keys(), key=lambda x: sum(classify_stats[x].values()), reverse=True):
        s = classify_stats[cls]
        total = s['tp'] + s['fp'] + s['tn'] + s['fn']
        if total < 10: continue
        client_discard = s['tp'] + s['fn']
        agent_discard = s['tp'] + s['fp']
        client_pct = client_discard / total * 100
        agent_pct = agent_discard / total * 100
        prec = s['tp'] / (s['tp'] + s['fp']) if (s['tp'] + s['fp']) > 0 else 0
        rec = s['tp'] / (s['tp'] + s['fn']) if (s['tp'] + s['fn']) > 0 else 0
        print(f"  {cls:<18} {total:>6} {client_pct:>7.1f}% {agent_pct:>7.1f}% {s['fn']:>5} {s['fp']:>5} {prec:>5.2f} {rec:>5.2f}")

    # Special analysis: REGION
    print(f"\n{'='*80}")
    print(f"REGION ANALYSIS")
    print(f"{'='*80}")

    region_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'tn': 0, 'fn': 0})
    for rid in v7:
        if rid not in gt or rid not in raw_data: continue
        reg = raw_data[rid].get('REGION', 'unknown')
        reg = str(reg)[:20] if reg else 'unknown'
        j = v7[rid]
        if j['agent_judgment'] == 'DISCARD' and gt[rid] == 'DISCARD': region_stats[reg]['tp'] += 1
        elif j['agent_judgment'] == 'DISCARD' and gt[rid] == 'KEEP': region_stats[reg]['fp'] += 1
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'DISCARD': region_stats[reg]['fn'] += 1
        elif j['agent_judgment'] in ('REVIEW','KEEP') and gt[rid] == 'KEEP': region_stats[reg]['tn'] += 1

    print(f"\n{'Region':<20} {'Total':>6} {'Client%':>8} {'FN':>5} {'FP':>5} {'Miss%':>8}")
    for reg in sorted(region_stats.keys(), key=lambda x: sum(region_stats[x].values()), reverse=True):
        s = region_stats[reg]
        total = s['tp'] + s['fp'] + s['tn'] + s['fn']
        if total < 10: continue
        client_discard = s['tp'] + s['fn']
        client_pct = client_discard / total * 100
        miss_pct = s['fn'] / total * 100
        print(f"  {reg:<18} {total:>6} {client_pct:>7.1f}% {s['fn']:>5} {s['fp']:>5} {miss_pct:>7.1f}%")

    # Special analysis: qtime distribution
    print(f"\n{'='*80}")
    print(f"QTIME (total survey time) ANALYSIS")
    print(f"{'='*80}")

    for label, ids in [('TP', tp_ids), ('FP', fp_ids), ('TN', tn_ids), ('FN', fn_ids)]:
        times = [raw_data[rid].get('qtime') for rid in ids if rid in raw_data and raw_data[rid].get('qtime') is not None]
        if times:
            times = [float(t) for t in times if isinstance(t, (int, float))]
            if times:
                print(f"  {label}: mean={statistics.mean(times):.1f}, median={statistics.median(times):.1f}, n={len(times)}")

    # Special analysis: LangAssess
    print(f"\n{'='*80}")
    print(f"LANGUAGE ASSESSMENT ANALYSIS")
    print(f"{'='*80}")

    for col in ['LangAssessReadLevel', 'LangAssessReadEase', 'LangAssessNumWords']:
        print(f"\n  {col}:")
        for label, ids in [('TP', tp_ids), ('FP', fp_ids), ('TN', tn_ids), ('FN', fn_ids)]:
            vals = [raw_data[rid].get(col) for rid in ids if rid in raw_data and raw_data[rid].get(col) is not None]
            vals = [float(v) for v in vals if isinstance(v, (int, float))]
            if vals:
                print(f"    {label}: mean={statistics.mean(vals):.2f}, median={statistics.median(vals):.2f}, n={len(vals)}")

    # Special analysis: RD_Search values
    print(f"\n{'='*80}")
    print(f"RD_SEARCH ANALYSIS")
    print(f"{'='*80}")

    for col in ['RD_Searchr0', 'RD_Searchr1', 'RD_Searchr2', 'RD_Searchr3']:
        print(f"\n  {col}:")
        for label, ids in [('TP', tp_ids), ('FP', fp_ids), ('TN', tn_ids), ('FN', fn_ids)]:
            vals = [raw_data[rid].get(col) for rid in ids if rid in raw_data and raw_data[rid].get(col) is not None]
            vals = [float(v) for v in vals if isinstance(v, (int, float))]
            if vals:
                print(f"    {label}: mean={statistics.mean(vals):.2f}, median={statistics.median(vals):.2f}")

    # Cross-tabulation: source x CLASSIFY for FN
    print(f"\n{'='*80}")
    print(f"SOURCE x CLASSIFY for FN (where are the misses concentrated?)")
    print(f"{'='*80}")

    fn_cross = Counter()
    fn_total = 0
    for rid in fn_ids:
        if rid not in raw_data: continue
        src = str(raw_data[rid].get('source', '?'))[:15]
        cls = str(raw_data[rid].get('CLASSIFY', '?'))[:10]
        fn_cross[(src, cls)] += 1
        fn_total += 1

    for (src, cls), c in fn_cross.most_common(15):
        print(f"  source={src:<15} CLASSIFY={cls:<10}: {c} ({c/fn_total*100:.1f}%)")

    # Check if FN cluster around specific source values
    print(f"\n{'='*80}")
    print(f"FN WITH ML 0.4-0.5 — raw data deep dive")
    print(f"{'='*80}")

    ml_fn = [v7[rid] for rid in fn_ids if rid in v7 and 0.4 <= v7[rid]['ml_score'] < 0.5]
    print(f"FN with ML 0.4-0.5: {len(ml_fn)}")

    # Check source distribution
    src_dist = Counter()
    for j in ml_fn:
        rid = j['respondent_id']
        if rid in raw_data:
            src = str(raw_data[rid].get('source', '?'))[:20]
            src_dist[src] += 1
    print(f"\nSource distribution (ML 0.4-0.5 FN):")
    for src, c in src_dist.most_common(10):
        print(f"  {src}: {c}")

    # Check CLASSIFY distribution
    cls_dist = Counter()
    for j in ml_fn:
        rid = j['respondent_id']
        if rid in raw_data:
            cls = str(raw_data[rid].get('CLASSIFY', '?'))[:10]
            cls_dist[cls] += 1
    print(f"\nCLASSIFY distribution (ML 0.4-0.5 FN):")
    for cls, c in cls_dist.most_common():
        print(f"  {cls}: {c}")

    # Check qtime
    times = []
    for j in ml_fn:
        rid = j['respondent_id']
        if rid in raw_data and raw_data[rid].get('qtime') is not None:
            t = raw_data[rid]['qtime']
            if isinstance(t, (int, float)):
                times.append(float(t))
    if times:
        print(f"\nQtime (ML 0.4-0.5 FN): mean={statistics.mean(times):.1f}, median={statistics.median(times):.1f}")

    # Compare with TN in same ML band
    ml_tn = [v7[rid] for rid in tn_ids if rid in v7 and 0.4 <= v7[rid]['ml_score'] < 0.5]
    print(f"\nTN with ML 0.4-0.5: {len(ml_tn)}")

    src_dist_tn = Counter()
    for j in ml_tn:
        rid = j['respondent_id']
        if rid in raw_data:
            src = str(raw_data[rid].get('source', '?'))[:20]
            src_dist_tn[src] += 1
    print(f"Source distribution (ML 0.4-0.5 TN):")
    for src, c in src_dist_tn.most_common(10):
        print(f"  {src}: {c}")

    cls_dist_tn = Counter()
    for j in ml_tn:
        rid = j['respondent_id']
        if rid in raw_data:
            cls = str(raw_data[rid].get('CLASSIFY', '?'))[:10]
            cls_dist_tn[cls] += 1
    print(f"\nCLASSIFY distribution (ML 0.4-0.5 TN):")
    for cls, c in cls_dist_tn.most_common():
        print(f"  {cls}: {c}")

    times_tn = []
    for j in ml_tn:
        rid = j['respondent_id']
        if rid in raw_data and raw_data[rid].get('qtime') is not None:
            t = raw_data[rid]['qtime']
            if isinstance(t, (int, float)):
                times_tn.append(float(t))
    if times_tn:
        print(f"Qtime (ML 0.4-0.5 TN): mean={statistics.mean(times_tn):.1f}, median={statistics.median(times_tn):.1f}")

    # LangAssess comparison in ML 0.4-0.5 band
    print(f"\nLangAssess comparison (ML 0.4-0.5):")
    for col in ['LangAssessReadLevel', 'LangAssessReadEase', 'LangAssessNumWords']:
        fn_vals = [raw_data[j['respondent_id']].get(col) for j in ml_fn
                   if j['respondent_id'] in raw_data and raw_data[j['respondent_id']].get(col) is not None]
        tn_vals = [raw_data[j['respondent_id']].get(col) for j in ml_tn
                   if j['respondent_id'] in raw_data and raw_data[j['respondent_id']].get(col) is not None]
        fn_vals = [float(v) for v in fn_vals if isinstance(v, (int, float))]
        tn_vals = [float(v) for v in tn_vals if isinstance(v, (int, float))]
        if fn_vals and tn_vals:
            print(f"  {col}: FN mean={statistics.mean(fn_vals):.2f}, TN mean={statistics.mean(tn_vals):.2f}, gap={statistics.mean(fn_vals)-statistics.mean(tn_vals):+.3f}")

    # Check conditionsAriens / channel mismatch
    print(f"\n{'='*80}")
    print(f"CHANNEL CONDITIONS ANALYSIS")
    print(f"{'='*80}")

    for col in ['conditionsAriens', 'conditionsHD_or_OPE_dealers', 'conditionsOther_channel']:
        print(f"\n  {col}:")
        for label, ids in [('TP', tp_ids), ('FP', fp_ids), ('TN', tn_ids), ('FN', fn_ids)]:
            vals = [raw_data[rid].get(col) for rid in ids if rid in raw_data and raw_data[rid].get(col) is not None]
            vals = [float(v) for v in vals if isinstance(v, (int, float))]
            if vals:
                pct = sum(1 for v in vals if v > 0) / len(vals) * 100
                print(f"    {label}: mean={statistics.mean(vals):.2f}, %active={pct:.1f}%, n={len(vals)}")

if __name__ == '__main__':
    main()
