#!/usr/bin/env python3
"""Delta natural-language proposition mapper.

Translates each question-answer pair into a first-person self-claim proposition.
Scripts stage the question text, value labels, and raw values. The proposition
templates below encode the agent's mapping rules for each field type. The full
set of propositions for a respondent forms their self-claim profile, which the
agent reads as a narrative to check coherence.

This is the evidence staging layer. The agent derives the coherence judgment.
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ANNOTATED = Path("/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260111_Delta Water Filtration.xlsx")
OUT = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/agent-native-delta-transfer-2026-06-24/.autosurvey-internal/t5_semantic")

EXCLUDED_FIELDS = {
    "qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel", "LangAssessReadEase",
    "LangAssessNumSen", "LangAssessNumWords", "LangAssessNumSyl", "url", "session",
    "camp", "bhf", "sfh", "intcode",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions")


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb) -> dict[str, list[str]]:
    qmap: dict[str, list[str]] = {}
    current: str | None = None
    for a, b, c in wb["Datamap"].iter_rows(values_only=True):
        a_s = clean(a)
        # Format 1: [q3]: question text
        if a_s.startswith("[") and "]" in a_s:
            current = a_s.split("]", 1)[0][1:]
            qmap[current] = [a_s]
            continue
        # Format 2: q3: question text
        m = re.match(r"^(q\w+):\s*(.*)", a_s)
        if m:
            current = m.group(1)
            qmap[current] = [a_s]
            continue
        if current and (a_s or b is not None or c is not None):
            if a_s:
                qmap[current].append(a_s)
            if b is not None or c is not None:
                qmap[current].append(f"{b} | {c}")
    return qmap


def field_text(field: str, qmap) -> str:
    lines = qmap.get(field, [])
    if not lines:
        return field
    first = lines[0]
    if "] | " in first:
        return first.split("] | ", 1)[1]
    if "]: " in first:
        return first.split("]: ", 1)[1]
    # Format 2: q3: question text
    m = re.match(r"^q\w+:\s*(.*)", first)
    if m:
        return m.group(1)
    return first


def get_value_labels(field: str, qmap) -> dict[str, str]:
    """Return {coded_value: label} from the datamap."""
    labels = {}
    for line in qmap.get(field, []):
        if " | " not in line:
            continue
        left, right = line.split(" | ", 1)
        lv = clean(left)
        if lv and not lv.startswith("["):
            labels[lv] = clean(right)
    return labels


def get_subfield_labels(field: str, qmap) -> dict[str, str]:
    """Return {subfield_key: label} for multi-select fields (e.g. q3r1 -> 'Water filtration system')."""
    labels = {}
    for line in qmap.get(field, []):
        if " | " not in line:
            continue
        left, right = line.split(" | ", 1)
        lv = clean(left)
        if lv.startswith("[") and "]" in lv:
            key = lv.split("]")[0][1:]
            labels[key] = clean(right)
    return labels


# ---------------------------------------------------------------------------
# Proposition generation rules
# ---------------------------------------------------------------------------

def proposition_for_single_select(field: str, value, qmap) -> str:
    """Coded single-select: 'I [claim] [value label].'"""
    if value in (None, ""):
        return ""
    labels = get_value_labels(field, qmap)
    sv = str(value).strip()
    label = labels.get(sv, sv)
    qtext = field_text(field, qmap)
    # Build first-person proposition using field-specific templates
    return build_single_select_proposition(field, label, qtext)


def build_single_select_proposition(field: str, label: str, qtext: str) -> str:
    """Field-specific proposition templates for natural first-person claims."""
    f = field.lower()
    # Demographics
    if f == "qgender":
        return f"I am {label.lower()}."
    if f == "qager1":
        return f"I am {label} years old."
    if f == "qushhi":
        return f"My annual household income before taxes is {label}."
    if f == "qemploy":
        return f"My current employment status is: {label}."
    if f == "qed":
        return f"My education level is: {label}."
    if f == "qpolitics":
        return f"My political affiliation is: {label}."
    # Home
    if f == "q2":
        return f"I {label.lower()} my home."
    if f == "q31":
        return f"I live in a {label.lower()} area."
    if f == "q32":
        return f"My home is approximately {label}."
    if f == "q33":
        return f"My home value is in the range: {label}."
    if f == "q30":
        return f"My home's water source is: {label}."
    # Industry / role
    if f == "q1":
        return f"I work in: {label}."
    if f == "q13":
        return f"My role in selecting/purchasing items for my home is: {label}."
    if f == "q12":
        return f"Regarding my children's baths: {label}."
    # Product
    if f == "q8a":
        return f"My kitchen sink faucet filtration type is: {label}."
    if f == "q8b":
        return f"The kitchen sink faucet filtration I plan to get is: {label}."
    if f == "q9":
        return f"My bathroom sink faucet filtration type is: {label}."
    # Purchase / concern / knowledge
    if f == "q6":
        return f"In my search for a new water filtration product, I am: {label}."
    if f == "q14":
        return f"My primary reason for buying a water filtration device is: {label}."
    if f == "q15":
        return f"I am {label.lower()} about the quality of water in my home."
    if f == "q16":
        return f"I feel I know {label.lower()} about how a water filter works."
    if f == "q28":
        return f"I would purchase my filter at: {label}."
    # Fallback: generic
    subject = extract_subject(qtext)
    return f"I claim regarding {subject}: {label}."


def proposition_for_multi_select(field_base: str, row: dict, headers: list, qmap) -> list[str]:
    """Multi-select: list checked items as claims using item labels."""
    labels = get_subfield_labels(field_base, qmap)
    props = []
    checked = []
    for h in headers:
        if not h.startswith(field_base + "r"):
            continue
        if h.endswith("oe") or h.endswith("oth"):
            continue
        val = str(row.get(h, "")).strip()
        item_label = labels.get(h, h)
        if val == "1":
            checked.append(item_label)
    if checked:
        qtext = field_text(field_base, qmap)
        verb = infer_multi_select_verb(qtext)
        props.append(f"I {verb}: {', '.join(checked)}.")
    # Check for "None of these" selected
    for h in headers:
        if not h.startswith(field_base + "r"):
            continue
        item_label = labels.get(h, h)
        if "none" in item_label.lower() and str(row.get(h, "")).strip() == "1":
            props.append(f"I claim: {item_label.lower()}.")
    return props


def proposition_for_open_text(field: str, value, qmap) -> str:
    """Open text: 'I said: [text]' about [subject]."""
    if value in (None, ""):
        return ""
    qtext = field_text(field, qmap)
    # Use field-specific subjects for open text
    f = field.lower()
    subjects = {
        "outro": "what this survey was about",
        "q14": "why I decided to buy a water filtration device",
        "q7r11oe": "other research sources I used",
        "q8ar4oe": "other kitchen sink faucet filtration I have",
        "q8br4oe": "other kitchen sink faucet filtration I plan to get",
        "q9r4oe": "other bathroom sink faucet filtration I have",
        "q26r15oe": "other filter brands I am aware of",
        "q28r12oe": "other stores where I would purchase my filter",
        "q30r6oe": "other water sources for my home",
        "qEthnicr8oe": "my race/ethnicity (other)",
    }
    subject = subjects.get(f, extract_subject(qtext))
    return f"I said about {subject}: \"{clean(value)}\""


def proposition_for_matrix(field_base: str, row: dict, qmap) -> list[str]:
    """Matrix cells: translate each product into have/plan/neither claims."""
    labels = get_subfield_labels(field_base, qmap)
    props = []
    # Find all row items (q5r1, q5r2, etc.)
    base_num = re.match(r"(\d+)", field_base.replace("q", ""))
    if not base_num:
        return props
    # For q5: q5r1c1=have, q5r1c2=plan, q5r1c3=neither
    have_items = []
    plan_items = []
    neither_items = []
    for h in sorted(labels):
        if not h.startswith(field_base):
            continue
        item_label = labels[h]
        c1 = str(row.get(f"{h}c1", "")).strip()
        c2 = str(row.get(f"{h}c2", "")).strip()
        c3 = str(row.get(f"{h}c3", "")).strip()
        if c1 == "1":
            have_items.append(item_label)
        if c2 == "1":
            plan_items.append(item_label)
        if c3 == "1":
            neither_items.append(item_label)
    if have_items:
        props.append(f"I currently have water treatment for: {', '.join(have_items)}.")
    if plan_items:
        props.append(f"I plan to get water treatment for: {', '.join(plan_items)}.")
    if neither_items:
        props.append(f"I do not have or plan to get water treatment for: {', '.join(neither_items)}.")
    return props


def proposition_for_numeric(field: str, value, qmap) -> str:
    """Numeric: 'My [subject] is [value].'"""
    if value in (None, ""):
        return ""
    qtext = field_text(field, qmap)
    subject = extract_subject(qtext)
    return f"My {subject}: {value}."


def extract_subject(qtext: str) -> str:
    """Extract a short subject phrase from question text for proposition construction."""
    # Remove leading question words and pipe placeholders
    t = re.sub(r"\[pipe:.*?\]", "", qtext).strip()
    t = re.sub(r"^(Which of the following |What |How |Do you |Are you |Where |When |Why )", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\?*$", "", t).strip()
    # Truncate to a reasonable length
    if len(t) > 80:
        t = t[:77] + "..."
    return t.lower() if t else "this"


def infer_multi_select_verb(qtext: str) -> str:
    """Infer the verb for multi-select propositions."""
    low = qtext.lower()
    if "aware" in low:
        return "am aware of the following"
    if "purchased" in low or "bought" in low:
        return "purchased/installed in my home"
    if "interested" in low or "plan" in low:
        return "am interested in purchasing"
    if "research" in low or "source" in low:
        return "used the following research sources"
    if "concern" in low:
        return "am concerned about the following in my water"
    if "have" in low and "water" in low:
        return "have the following in my home's water"
    if "information" in low or "get info" in low:
        return "would use the following for information"
    if "pay" in low or "willing" in low:
        return "am willing to pay for filtration that"
    return "selected"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    wb = openpyxl.load_workbook(ANNOTATED, read_only=True, data_only=True)
    qmap = load_datamap(wb)
    ws = wb["A1"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Classify fields
    open_fields = {"q7r11oe", "q8ar4oe", "q8br4oe", "q9r4oe", "q26r15oe", "q28r12oe",
                   "q30r6oe", "qEthnicr8oe", "outro", "vosr15oe", "vbrowserr15oe"}
    single_select_fields = {"q1", "q2", "q6", "q8a", "q8b", "q9", "q12", "q13", "q14",
                            "q15", "q16", "q28", "q30", "q31", "q32", "q33",
                            "qUSHHI", "qEmploy", "qEd", "qGender", "qager1", "qPolitics"}
    multi_select_bases = {"q3", "q4", "q7", "q22", "q23", "q26", "q29"}
    matrix_bases = {"q5"}
    # q18, q19 are ranking; q20-q25 are rating grids (treat as multi-select-like)
    rating_grids = {"q20", "q21", "q24", "q25"}

    all_rows = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid"))
        status = raw.get("status")
        if not rid or status is None:
            continue
        raw["_rid"] = rid
        raw["_xrow"] = xrow
        raw["_status"] = int(status)
        all_rows.append(raw)

    t5 = [r for r in all_rows if r["_status"] == 5]
    t3 = [r for r in all_rows if r["_status"] == 3]

    def build_propositions(row: dict) -> dict:
        """Build the full self-claim profile for a respondent."""
        props = []

        # Single-select fields
        for f in single_select_fields:
            if f not in headers:
                continue
            v = row.get(f)
            if v in (None, ""):
                continue
            p = proposition_for_single_select(f, v, qmap)
            if p:
                props.append({"field": f, "type": "single_select", "proposition": p})

        # Multi-select fields
        for base in multi_select_bases:
            sub_props = proposition_for_multi_select(base, row, headers, qmap)
            for p in sub_props:
                props.append({"field": base, "type": "multi_select", "proposition": p})

        # Matrix (q5)
        for base in matrix_bases:
            sub_props = proposition_for_matrix(base, row, qmap)
            for p in sub_props:
                props.append({"field": base, "type": "matrix", "proposition": p})

        # Open text fields
        for f in open_fields:
            if f not in headers:
                continue
            v = row.get(f)
            if v in (None, ""):
                continue
            p = proposition_for_open_text(f, v, qmap)
            if p:
                props.append({"field": f, "type": "open_text", "proposition": p})

        # Rating grids (q20-q25) — summarize using item labels
        for base in rating_grids:
            fields = [h for h in headers if h.startswith(base + "r")]
            labels = get_subfield_labels(base, qmap)
            checked = []
            for h in fields:
                if str(row.get(h, "")).strip() == "1":
                    checked.append(labels.get(h, h))
            if checked:
                qtext = field_text(base, qmap)
                verb = infer_multi_select_verb(qtext)
                props.append({"field": base, "type": "rating_grid", "proposition": f"I {verb}: {', '.join(checked[:8])}{'...' if len(checked)>8 else ''}."})

        return props

    # Build self-claim profiles for all t5 rows + t3 guardrail sample
    t5_profiles = []
    for r in t5:
        props = build_propositions(r)
        t5_profiles.append({
            "respondent_id": r["_rid"],
            "source_excel_row": r["_xrow"],
            "status": r["_status"],
            "qtime_seconds": float(r["qtime"]) if r.get("qtime") else None,
            "supplier": r.get("SUPNAME"),
            "self_claim_propositions": props,
            "proposition_count": len(props),
        })

    # t3 guardrail: 80 rows
    t3_sorted = sorted(t3, key=lambda r: (str(r.get("SUPNAME") or ""), float(r.get("qtime") or 0)))
    step = max(1, len(t3_sorted) // 80)
    t3_guard = t3_sorted[::step][:80]
    t3_profiles = []
    for r in t3_guard:
        props = build_propositions(r)
        t3_profiles.append({
            "respondent_id": r["_rid"],
            "source_excel_row": r["_xrow"],
            "status": r["_status"],
            "qtime_seconds": float(r["qtime"]) if r.get("qtime") else None,
            "supplier": r.get("SUPNAME"),
            "self_claim_propositions": props,
            "proposition_count": len(props),
        })

    (OUT / "t5_self_claim_profiles.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in t5_profiles) + "\n"
    )
    (OUT / "t3_guardrail_self_claim_profiles.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in t3_profiles) + "\n"
    )

    # Print a sample profile for review
    print(f"t5 profiles: {len(t5_profiles)}, t3 guardrail profiles: {len(t3_profiles)}")
    print(f"avg propositions per t5 row: {sum(p['proposition_count'] for p in t5_profiles)/len(t5_profiles):.1f}")
    print()
    print("=== SAMPLE SELF-CLAIM PROFILE (first t5 row) ===")
    sample = t5_profiles[0]
    print(f"Respondent: {sample['respondent_id']} | qtime={sample['qtime_seconds']:.0f}s | supplier={sample['supplier']}")
    for p in sample["self_claim_propositions"]:
        print(f"  [{p['field']:12s}] {p['proposition']}")


if __name__ == "__main__":
    main()
