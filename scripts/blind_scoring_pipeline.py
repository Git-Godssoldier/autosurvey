#!/usr/bin/env python3
"""Blind scoring pipeline for unannotated datasets.

Builds proposition profiles for every row, stages population-level signals,
and prepares review packets for agent semantic judgment. The agent reads
each packet and makes the final discard decision.

No status labels are used. No annotations are consulted. This is the
production-mode pipeline.

Stages:
1. Datamap parsing and field classification
2. Proposition profile construction for every row
3. Population-level signal staging (timing, supplier, duplicates, open-end patterns)
4. Review packet assembly: proposition profile + staged signals per row
5. Agent semantic scoring (the agent reads packets and scores each row)
6. Discard set output

Usage:
    python3 blind_scoring_pipeline.py <dataset.xlsx> --output-dir DIR
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
OUTPUT_BASE = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs")

EXCLUDED_FIELDS = {
    "qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel", "LangAssessReadEase",
    "LangAssessNumSen", "LangAssessNumWords", "LangAssessNumSyl", "url", "session",
    "camp", "bhf", "sfh", "intcode", "record", "uuid", "status", "qtime",
    "SUPNAME", "ipAddress", "date", "qStateVer",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions", "outroR1_RD_Review",
                     "qcoe1R1_RD_Review", "LangAssess", "_Pasted")


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb) -> dict[str, dict]:
    dm = wb["Datamap"]
    fields: dict[str, dict] = {}
    current: str | None = None
    for a, b, c in dm.iter_rows(values_only=True):
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            field_name = a_s.split("]", 1)[0][1:]
            qtext = a_s.split("]", 1)[1].lstrip(": ").strip()
            current = field_name
            fields[current] = {
                "question_text": qtext,
                "value_labels": {},
                "subfield_labels": {},
                "field_type": classify_field(qtext, field_name),
            }
            continue
        m = re.match(r"^(q\w+):\s*(.*)", a_s)
        if m:
            current = m.group(1)
            fields[current] = {
                "question_text": m.group(2),
                "value_labels": {},
                "subfield_labels": {},
                "field_type": classify_field(m.group(2), current),
            }
            continue
        if current and b is not None and c is not None:
            b_s = clean(b)
            c_s = clean(c)
            if b_s.startswith("[") and "]" in b_s:
                subfield = b_s.split("]")[0][1:]
                fields[current]["subfield_labels"][subfield] = c_s
            elif b_s and c_s and not b_s.startswith("Values:"):
                fields[current]["value_labels"][b_s] = c_s
    return fields


def classify_field(qtext: str, field_name: str) -> str:
    low = qtext.lower()
    fname = field_name.lower()
    if fname.endswith("oe") or fname.endswith("othr1") or fname == "outro" or fname.startswith("qcoe"):
        return "open_text"
    if "open text" in low or "open-ended" in low or "please describe" in low or "please specify" in low:
        return "open_text"
    demo_fields = {"qgender", "qager1", "qushhi", "qemploy", "qed", "qpolitics",
                   "qethnic", "qcountry", "qstate", "qzipr1", "qhometype", "qrevenue",
                   "qnumemployees", "qtenure", "qtenurer1", "qtrade", "qindustry"}
    if fname in demo_fields or any(fname.startswith(d) for d in ("qethnic",)):
        return "demographic"
    if re.search(r"which of the following|select all|check all|please select all", low):
        return "multi_select"
    if re.search(r"how (much|well|likely|willing)|rate each|for each|please rank", low):
        return "matrix"
    return "single_select"


def build_propositions(row: dict, fields: dict, headers: list) -> list[dict]:
    props = []
    for fname, finfo in fields.items():
        if excluded(fname) or fname not in headers:
            continue
        ftype = finfo["field_type"]
        qtext = finfo["question_text"]
        labels = finfo["value_labels"]

        if ftype == "open_text":
            value = row.get(fname)
            if value in (None, ""):
                continue
            subject = extract_subject(qtext)
            props.append({"field": fname, "type": "open_text", "proposition": f'I said about {subject}: "{clean(value)}"'})
            continue

        if ftype == "multi_select" or finfo["subfield_labels"]:
            sub_labels = finfo["subfield_labels"]
            checked = []
            for h in headers:
                if not h.startswith(fname + "r"):
                    continue
                if h.endswith("oe") or h.endswith("oth"):
                    continue
                val = str(row.get(h, "")).strip()
                item_label = sub_labels.get(h, h)
                if val == "1":
                    checked.append(item_label)
            if checked:
                verb = infer_multi_select_verb(qtext)
                props.append({"field": fname, "type": "multi_select", "proposition": f"I {verb}: {', '.join(checked[:15])}{'...' if len(checked)>15 else ''}."})
            continue

        if ftype == "demographic":
            value = row.get(fname)
            if value in (None, ""):
                continue
            sv = str(value).strip()
            label = labels.get(sv, sv)
            prop = build_demographic_proposition(fname.lower(), label, sv)
            if prop:
                props.append({"field": fname, "type": "demographic", "proposition": prop})
            continue

        if ftype == "single_select":
            value = row.get(fname)
            if value in (None, ""):
                continue
            sv = str(value).strip()
            label = labels.get(sv, sv)
            subject = extract_subject(qtext)
            props.append({"field": fname, "type": "single_select", "proposition": f"I claim regarding {subject}: {label}."})
            continue

    return props


def build_demographic_proposition(fname: str, label: str, raw: str) -> str | None:
    if fname == "qgender": return f"I am {label.lower()}."
    if fname == "qager1": return f"I am {raw} years old."
    if fname == "qushhi": return f"My annual household income before taxes is {label}."
    if fname == "qemploy": return f"My current employment status is: {label}."
    if fname == "qed": return f"My education level is: {label}."
    if fname == "qrevenue": return f"My company's annual revenue is: {label}."
    if fname == "qnumemployees": return f"My company has {label} employees."
    if fname in ("qtenure", "qtenurer1"): return f"My tenure in my role is: {label}."
    if fname == "qtrade": return f"My trade is: {label}."
    if fname == "qindustry": return f"I work in: {label}."
    if fname == "qstate": return f"I live in {label}."
    if fname == "qcountry": return f"I live in {label}."
    if fname == "qhometype": return f"My home type is: {label}."
    if fname.startswith("qethnic"): return f"My race/ethnicity includes: {label}."
    return f"My demographic: {label}."


def extract_subject(qtext: str) -> str:
    t = re.sub(r"\[pipe:.*?\]", "", qtext).strip()
    t = re.sub(r"^(Which of the following |What |How |Do you |Are you |Where |When |Why |Please |Using the |For each )", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\?*$", "", t).strip()
    if len(t) > 80: t = t[:77] + "..."
    return t.lower() if t else "this"


def infer_multi_select_verb(qtext: str) -> str:
    low = qtext.lower()
    if "aware" in low: return "am aware of the following"
    if "purchased" in low or "bought" in low: return "purchased/installed in my home"
    if "interested" in low or "plan" in low: return "am interested in purchasing"
    if "research" in low or "source" in low: return "used the following research sources"
    if "concern" in low: return "am concerned about the following"
    if "have" in low and "water" in low: return "have the following in my home's water"
    if "information" in low or "get info" in low: return "would use the following for information"
    if "pay" in low or "willing" in low: return "am willing to pay for"
    return "selected"


def stage_population_signals(profiles: list[dict], all_rows: list[dict]) -> dict:
    """Stage population-level signals for the agent to use as context."""
    # Timing distribution
    qtimes = [p["qtime_seconds"] for p in profiles if p["qtime_seconds"]]
    qtime_median = statistics.median(qtimes) if qtimes else 0
    qtime_p10 = statistics.quantiles(qtimes, n=10)[0] if len(qtimes) >= 10 else 0
    qtime_p25 = statistics.quantiles(qtimes, n=4)[0] if len(qtimes) >= 4 else 0

    # Supplier distribution
    suppliers = Counter(p["supplier"] or "MISSING" for p in profiles)
    missing_supplier_rate = suppliers.get("MISSING", 0) / len(profiles)

    # Open-end text duplicate detection
    open_text_hashes = defaultdict(list)
    for p in profiles:
        for prop in p.get("self_claim_propositions", []):
            if prop["type"] == "open_text":
                m = re.search(r'"(.+)"', prop["proposition"])
                if m:
                    text = m.group(1).strip().lower()
                    if len(text) > 15:  # Only substantive text
                        h = hashlib.md5(text.encode()).hexdigest()
                        open_text_hashes[h].append(p["respondent_id"])

    duplicates = {h: rids for h, rids in open_text_hashes.items() if len(rids) > 1}

    # Open-end text pattern staging (for agent reference, NOT for scoring)
    open_text_patterns = Counter()
    for p in profiles:
        for prop in p.get("self_claim_propositions", []):
            if prop["type"] == "open_text":
                m = re.search(r'"(.+)"', prop["proposition"])
                if m:
                    text = m.group(1)
                    tl = text.lower().strip()
                    if re.search(r"thank you|good survey|nice|amazing|love this|very good|great experience|very interesting", tl):
                        open_text_patterns["meta_praise"] += 1
                    if re.search(r"^none$|^n/?a$|^na$|^nothing$|^no$|^idk$|^i don.?t know$", tl):
                        open_text_patterns["placeholder"] += 1
                    if re.search(r"the poll|the study|the survey examined|crucially|in order to determine", tl):
                        open_text_patterns["templated"] += 1
                    if re.search(r"^[a-z]{20,}$|asdf|qwerty|bvnhgjut", tl):
                        open_text_patterns["gibberish"] += 1
                    if re.search(r"go to go to go to|was the one who was the one who|repeat repeat", tl):
                        open_text_patterns["repetition_loop"] += 1
                    if len(text.split()) > 40:
                        open_text_patterns["long_text_40plus"] += 1
                    if len(text.split()) <= 3:
                        open_text_patterns["very_few_words"] += 1

    return {
        "total_rows": len(profiles),
        "timing": {
            "median": round(qtime_median, 1),
            "p10": round(qtime_p10, 1),
            "p25": round(qtime_p25, 1),
        },
        "supplier": {
            "distribution": dict(suppliers.most_common(10)),
            "missing_rate": round(missing_supplier_rate, 3),
        },
        "duplicates": {
            "duplicate_text_groups": len(duplicates),
            "rows_in_duplicate_groups": sum(len(rids) for rids in duplicates.values()),
            "groups": [{"text_hash": h, "respondent_ids": rids} for h, rids in duplicates.items()],
        },
        "open_text_patterns": dict(open_text_patterns),
    }


def build_review_packet(profile: dict, pop_signals: dict) -> dict:
    """Build a review packet for the agent to read and score."""
    # Find if this row is in any duplicate group
    in_duplicate_group = []
    for dg in pop_signals["duplicates"]["groups"]:
        if profile["respondent_id"] in dg["respondent_ids"]:
            in_duplicate_group.append(dg["respondent_ids"])

    # Stage open-end texts for easy reading
    open_end_texts = []
    for prop in profile.get("self_claim_propositions", []):
        if prop["type"] == "open_text":
            m = re.search(r'"(.+)"', prop["proposition"])
            if m:
                open_end_texts.append({
                    "field": prop["field"],
                    "text": m.group(1),
                    "word_count": len(m.group(1).split()),
                })

    # Compute timing percentile
    qt = profile["qtime_seconds"]
    timing_context = None
    if qt:
        if qt < pop_signals["timing"]["p10"]:
            timing_context = "very_fast (bottom 10%)"
        elif qt < pop_signals["timing"]["p25"]:
            timing_context = "fast (bottom 25%)"
        elif qt > pop_signals["timing"]["median"] * 2:
            timing_context = "very_slow (2x median)"
        else:
            timing_context = "normal"

    return {
        "respondent_id": profile["respondent_id"],
        "source_excel_row": profile["source_excel_row"],
        "qtime_seconds": qt,
        "timing_context": timing_context,
        "supplier": profile["supplier"],
        "supplier_missing": not profile["supplier"] or str(profile["supplier"]).strip() == "",
        "in_duplicate_group": in_duplicate_group,
        "open_end_texts": open_end_texts,
        "self_claim_propositions": profile["self_claim_propositions"],
        "proposition_count": profile["proposition_count"],
    }


def process_dataset(filepath: Path, output_dir: Path) -> dict:
    filename = filepath.stem
    print(f"  Processing {filename}...")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fields = load_datamap(wb)
    ws = wb["A1"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    all_rows = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid"))
        if not rid:
            continue
        raw["_rid"] = rid
        raw["_xrow"] = xrow
        all_rows.append(raw)

    # Build proposition profiles for every row
    profiles = []
    for row in all_rows:
        props = build_propositions(row, fields, headers)
        profiles.append({
            "respondent_id": row["_rid"],
            "source_excel_row": row["_xrow"],
            "qtime_seconds": float(row["qtime"]) if row.get("qtime") else None,
            "supplier": row.get("SUPNAME"),
            "self_claim_propositions": props,
            "proposition_count": len(props),
        })

    # Stage population-level signals
    pop_signals = stage_population_signals(profiles, all_rows)

    # Build review packets for every row
    packets = [build_review_packet(p, pop_signals) for p in profiles]

    # Save artifacts
    ds_out = output_dir / filename
    ds_out.mkdir(parents=True, exist_ok=True)

    (ds_out / "population_signals.json").write_text(json.dumps(pop_signals, indent=2))
    (ds_out / "review_packets.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in packets) + "\n"
    )

    # Print summary
    print(f"    {len(profiles)} rows, {sum(p['proposition_count'] for p in profiles)/len(profiles):.1f} avg propositions/row")
    print(f"    Population signals: timing median={pop_signals['timing']['median']:.0f}s, missing_supplier={pop_signals['supplier']['missing_rate']:.1%}")
    print(f"    Open-end patterns: {pop_signals['open_text_patterns']}")
    print(f"    Duplicate text groups: {pop_signals['duplicates']['duplicate_text_groups']}")

    wb.close()
    return {
        "dataset": filename,
        "total_rows": len(profiles),
        "avg_propositions": round(sum(p["proposition_count"] for p in profiles) / len(profiles), 1),
        "population_signals": pop_signals,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--all", action="store_true")
    parser.add_argument("dataset", nargs="?")
    parser.add_argument("--output-dir", default=str(OUTPUT_BASE))
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.all:
        datasets = sorted([f for f in DATA_DIR.iterdir() if f.suffix == ".xlsx" and not f.name.startswith(".~")])
        results = []
        for ds in datasets:
            result = process_dataset(ds, output_dir)
            results.append(result)
        (output_dir / "blind_run_summary.json").write_text(json.dumps(results, indent=2))
        print(f"\nProcessed {len(results)} datasets")
    elif args.dataset:
        ds_path = DATA_DIR / args.dataset
        result = process_dataset(ds_path, output_dir)
        print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
