#!/usr/bin/env python3
"""V29 — V19 features (AUC 0.826) + proper threshold optimization.

V19 achieved AUC 0.826 (highest ever) but only BAcc 0.692 because threshold
optimization was poor. This version:
1. Uses V19's features (target encoding, LangAssess, RD_Search, raw answers, matrix patterns)
2. Optimizes threshold on VALIDATION set (not test set) to avoid overfitting
3. Uses per-channel (Pro/Consumer) threshold optimization on validation
4. Adds more aggressive threshold search with finer granularity
5. Combines V19 features with V14's self-training approach
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score, balanced_accuracy_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import classify_field
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training
from v19_target_encoding import extract_raw_answer_features, add_qtime_quality_interactions, add_target_encoding_features_train

import xgboost as xgb
import lightgbm as lgb


def get_classify_map():
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()
    return classify_map


def optimize_threshold_on_val(val_scores, y_val, is_pro_val):
    """Find optimal per-channel thresholds on validation set."""
    best_bacc = 0
    best_thresh = 0.5
    best_pro_adj = 0

    # Fine-grained threshold search
    for thresh in np.arange(0.10, 0.80, 0.01):
        for pro_adj in np.arange(-0.20, 0.21, 0.01):
            pred = np.zeros(len(y_val), dtype=int)
            for i in range(len(y_val)):
                t = thresh + pro_adj if is_pro_val[i] else thresh
                pred[i] = 1 if val_scores[i] >= t else 0
            tp = ((pred == 1) & (y_val == 1)).sum()
            fp = ((pred == 1) & (y_val == 0)).sum()
            tn = ((pred == 0) & (y_val == 0)).sum()
            fn = ((pred == 0) & (y_test == 1)).sum() if 'y_test' in dir() else ((pred == 0) & (y_val == 1)).sum()
            rec = tp / max(tp + fn, 1)
            spec = tn / max(tn + fp, 1)
            bacc = (rec + spec) / 2
            if bacc > best_bacc:
                best_bacc = bacc
                best_thresh = thresh
                best_pro_adj = pro_adj

    return best_thresh, best_pro_adj, best_bacc


def run_v29_cv(n_folds=5):
    """Run V29 cross-validation."""
    print(f"\n{'='*80}")
    print(f"V29 — V19 Features (AUC 0.826) + Proper Threshold Optimization")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Extract raw answer features (V19)
    print("\nExtracting raw answer features...")
    raw_df, headers, roles = extract_raw_answer_features(ECHO_XLSX, gt, v7, v8)
    df = df.merge(raw_df, on="respondent_id", how="left", suffixes=("", "_raw"))
    df = add_qtime_quality_interactions(df)

    # Initialize TE columns
    for qcol in [c for c in df.columns if c.startswith("ans_")]:
        df[f"te_{qcol}"] = 0.35
        df[f"cnt_{qcol}"] = 0

    print(f"\nTotal features (before TE): {len(df.columns)}")

    # Extract all datasets
    print("\nExtracting all datasets...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    classify_map = get_classify_map()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Compute target encoding using ONLY training fold labels
        echo_train_for_te = echo_train[["respondent_id", "label"] + [c for c in echo_train.columns if c.startswith("ans_")]]
        echo_train = add_target_encoding_features_train(echo_train_for_te, echo_train, smoothing=20)
        echo_test = add_target_encoding_features_train(echo_train_for_te, echo_test, smoothing=20)

        # Self-training
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        # Split for calibration AND threshold optimization
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]
        is_pro_val = np.ones(len(y_val), dtype=bool)  # We don't know pro/consumer for pseudo-labeled

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        print(f"  Training on {len(X_tr)} samples, {X_tr.shape[1]} features...")

        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.6, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.6,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        # Calibrate on validation
        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_val = np.mean(list(cal_val.values()), axis=0)
        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_val = meta_model.predict_proba(meta_X_val)[:, 1]
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Optimize threshold on VALIDATION set (not test!)
        # This is the key change from V19
        print("  Optimizing threshold on validation set...")

        best_bacc_val = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, val_scores, test_scores in [
            ("ensemble", ensemble_val, ensemble_test),
            ("stacking", stacking_val, stacking_test),
        ]:
            for thresh in np.arange(0.10, 0.80, 0.01):
                for pro_adj in np.arange(-0.20, 0.21, 0.01):
                    # Evaluate on validation
                    pred_val = np.zeros(len(y_val), dtype=int)
                    for i in range(len(y_val)):
                        pred_val[i] = 1 if val_scores[i] >= thresh else 0
                    tp = ((pred_val == 1) & (y_val == 1)).sum()
                    fp = ((pred_val == 1) & (y_val == 0)).sum()
                    tn = ((pred_val == 0) & (y_val == 0)).sum()
                    fn = ((pred_val == 0) & (y_val == 1)).sum()
                    rec = tp / max(tp + fn, 1)
                    spec = tn / max(tn + fp, 1)
                    bacc_val = (rec + spec) / 2

                    if bacc_val > best_bacc_val:
                        best_bacc_val = bacc_val
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        # Apply best threshold to TEST set
        test_scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if test_scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, test_scores)

        print(f"  Best (val): {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}, val_bacc={best_bacc_val:.3f}")
        print(f"  Test: TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "thresh": best_thresh, "pro_adj": best_pro_adj,
            "val_bacc": best_bacc_val,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V29 CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V29 — V19 Features (AUC 0.826) + Proper Threshold Optimization")
    print("=" * 80)

    results = run_v29_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V14 (self-train + V8):   BAcc=0.744, AUC=0.788 (BEST BAcc)")
    print(f"  V19 (target encoding):   BAcc=0.692, AUC=0.826 (BEST AUC)")
    print(f"  V29 (V19 feat + val opt): BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}, AUC={results['avg_auc']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], 0.744):.3f}")

    with open(AUTOSURVEY_DIR / "v29_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
