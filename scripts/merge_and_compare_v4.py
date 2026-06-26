#!/usr/bin/env python3
"""Merge agent judgment chunks and compare against ground truth across all 4 datasets.

For each dataset:
1. Merge all agent_judgments_chunk_XX.json files into agent_judgments.json
2. Load annotated ground truth (status column)
3. Compare agent DISCARD/REVIEW/KEEP vs client status=3 (keep) / status=5 (discard)
4. Compute precision, recall, F1

Outputs a cross-dataset summary table.

Usage:
    python3 merge_and_compare_v4.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

# Add the skills scripts directory
SKILL_SCRIPTS = Path(__file__).parent.parent / "skills" / "cleaning-survey-quality" / "scripts"
sys.path.insert(0, str(SKILL_SCRIPTS))

from survey_pipeline import clean


# Dataset configurations: (name, holistic_output_dir, annotated_xlsx_path)
DATASETS = [
    {
        "name": "Delta Water Filtration",
        "short": "Delta",
        "holistic_dir": "/tmp/holistic_106_2502_Delta",
        "annotated_path": "/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260111_Delta Water Filtration.xlsx",
    },
    {
        "name": "SBD Brand Association",
        "short": "SBD",
        "holistic_dir": "/tmp/holistic_260200_SBD",
        "annotated_path": "/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260200_SBD.xlsx",
    },
    {
        "name": "ECHO Brand Health",
        "short": "ECHO",
        "holistic_dir": "/tmp/holistic_260300_ECHO",
        "annotated_path": "/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260300_ECHO.xlsx",
    },
    {
        "name": "ODL Switchable Glass",
        "short": "ODL",
        "holistic_dir": "/tmp/holistic_260501_ODL",
        "annotated_path": "/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260501_ODL.xlsx",
    },
]


def merge_chunk_judgments(holistic_dir):
    """Merge all agent_judgments_chunk_XX.json files into a single list."""
    holistic_dir = Path(holistic_dir)
    all_judgments = []
    chunk_files = sorted(holistic_dir.glob("agent_judgments_chunk_*.json"))

    if not chunk_files:
        print(f"  WARNING: No chunk files found in {holistic_dir}")
        return []

    for chunk_file in chunk_files:
        with open(chunk_file) as f:
            try:
                chunks = json.load(f)
                all_judgments.extend(chunks)
                print(f"  {chunk_file.name}: {len(chunks)} judgments")
            except json.JSONDecodeError as e:
                print(f"  ERROR parsing {chunk_file.name}: {e}")

    # Write merged file
    merged_path = holistic_dir / "agent_judgments.json"
    with open(merged_path, "w") as f:
        json.dump(all_judgments, f, indent=2)
    print(f"  Merged: {len(all_judgments)} total judgments → {merged_path.name}")

    return all_judgments


def load_ground_truth(filepath):
    """Load status labels from an annotated dataset.

    status=3 → accepted (KEEP)
    status=5 → rejected (DISCARD)
    """
    filepath = Path(filepath)
    if not filepath.exists():
        print(f"  WARNING: Annotated file not found: {filepath}")
        return {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    rid_col = hidx.get("uuid") or hidx.get("record")
    status_col = hidx.get("status")

    if status_col is None:
        print("  WARNING: No 'status' column found — file may not be annotated")
        return {}

    ground_truth = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = clean(row[rid_col]) if rid_col and rid_col < len(row) else ""
        status = row[status_col] if status_col and status_col < len(row) else None
        if rid and status is not None:
            try:
                status = int(status)
            except (ValueError, TypeError):
                pass
            ground_truth[rid] = status

    wb.close()
    return ground_truth


def compare_judgments(judgments, ground_truth, dataset_name):
    """Compare agent judgments against ground truth."""
    judgment_lookup = {j["respondent_id"]: j for j in judgments}

    # Only compare respondents that have both agent judgment and ground truth
    compared = []
    for rid, status in ground_truth.items():
        if rid in judgment_lookup and status in (3, 5):
            j = judgment_lookup[rid]
            compared.append({
                "respondent_id": rid,
                "client_status": status,
                "client_judgment": "DISCARD" if status == 5 else "KEEP",
                "agent_score": float(j["agent_score"]),
                "agent_judgment": j["agent_judgment"],
                "agent_justification": j.get("agent_justification", ""),
            })

    if not compared:
        print(f"  WARNING: No comparable respondents for {dataset_name}")
        return None

    # Confusion matrix (DISCARD detection)
    tp = sum(1 for c in compared if c["agent_judgment"] == "DISCARD" and c["client_judgment"] == "DISCARD")
    fp = sum(1 for c in compared if c["agent_judgment"] == "DISCARD" and c["client_judgment"] == "KEEP")
    fn = sum(1 for c in compared if c["agent_judgment"] != "DISCARD" and c["client_judgment"] == "DISCARD")
    tn = sum(1 for c in compared if c["agent_judgment"] != "DISCARD" and c["client_judgment"] == "KEEP")

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    # Also check REVIEW band — how many false negatives are in REVIEW?
    fn_in_review = sum(1 for c in compared if c["agent_judgment"] == "REVIEW" and c["client_judgment"] == "DISCARD")
    fn_in_keep = sum(1 for c in compared if c["agent_judgment"] == "KEEP" and c["client_judgment"] == "DISCARD")

    # False positives breakdown
    fp_from_review = sum(1 for c in compared if c["agent_judgment"] == "DISCARD" and c["client_judgment"] == "KEEP" and c["agent_score"] >= -0.5)

    # Agent judgment distribution
    n_discard = sum(1 for c in compared if c["agent_judgment"] == "DISCARD")
    n_review = sum(1 for c in compared if c["agent_judgment"] == "REVIEW")
    n_keep = sum(1 for c in compared if c["agent_judgment"] == "KEEP")

    # Client distribution
    n_client_discard = sum(1 for c in compared if c["client_judgment"] == "DISCARD")
    n_client_keep = sum(1 for c in compared if c["client_judgment"] == "KEEP")

    result = {
        "dataset": dataset_name,
        "n_compared": len(compared),
        "n_client_discard": n_client_discard,
        "n_client_keep": n_client_keep,
        "n_agent_discard": n_discard,
        "n_agent_review": n_review,
        "n_agent_keep": n_keep,
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "tn": tn,
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(f1, 4),
        "fn_in_review": fn_in_review,
        "fn_in_keep": fn_in_keep,
        "fp_from_review_band": fp_from_review,
    }

    return result, compared


def main():
    print(f"\n{'='*80}")
    print(f"V4 EVIDENCE-FAMILY FRAMEWORK — CROSS-DATASET COMPARISON")
    print(f"{'='*80}")

    all_results = []
    all_compared = {}
    merge_stats = {}

    for ds in DATASETS:
        print(f"\n{'─'*80}")
        print(f"Dataset: {ds['name']} ({ds['short']})")
        print(f"{'─'*80}")

        # Step 1: Merge chunk judgments
        print(f"\n[1/2] Merging agent judgments from {ds['holistic_dir']}...")
        judgments = merge_chunk_judgments(ds["holistic_dir"])
        merge_stats[ds["short"]] = len(judgments)

        # Step 2: Load ground truth and compare
        print(f"\n[2/2] Loading ground truth and comparing...")
        gt = load_ground_truth(ds["annotated_path"])
        n_accepted = sum(1 for s in gt.values() if s == 3)
        n_rejected = sum(1 for s in gt.values() if s == 5)
        print(f"  Ground truth: {len(gt)} total, {n_accepted} accepted, {n_rejected} rejected")

        result = compare_judgments(judgments, gt, ds["name"])
        if result:
            r, compared = result
            all_results.append(r)
            all_compared[ds["short"]] = compared

            print(f"\n  Confusion Matrix (n={r['n_compared']}):")
            print(f"    True Positives  (agent DISCARD, client DISCARD):  {r['tp']}")
            print(f"    False Positives (agent DISCARD, client KEEP):     {r['fp']}")
            print(f"    False Negatives (agent !DISCARD, client DISCARD): {r['fn']}")
            print(f"    True Negatives  (agent !DISCARD, client KEEP):    {r['tn']}")
            print(f"\n  Precision: {r['precision']:.4f}")
            print(f"  Recall:    {r['recall']:.4f}")
            print(f"  F1:        {r['f1']:.4f}")
            print(f"\n  FN breakdown: {r['fn_in_review']} in REVIEW, {r['fn_in_keep']} in KEEP")
            print(f"  Agent distribution: {r['n_agent_discard']} DISCARD, {r['n_agent_review']} REVIEW, {r['n_agent_keep']} KEEP")

    # Cross-dataset summary
    print(f"\n{'='*80}")
    print(f"CROSS-DATASET SUMMARY — V4 EVIDENCE-FAMILY FRAMEWORK")
    print(f"{'='*80}")
    print(f"\n{'Dataset':<12} {'N':>6} {'Client':>8} {'Agent':>8} {'TP':>5} {'FP':>5} {'FN':>5} {'TN':>5} {'Prec':>7} {'Recall':>7} {'F1':>7}")
    print(f"{'─'*12} {'─'*6} {'─'*8} {'─'*8} {'─'*5} {'─'*5} {'─'*5} {'─'*5} {'─'*7} {'─'*7} {'─'*7}")

    total_tp = total_fp = total_fn = total_tn = 0
    for r in all_results:
        client_d = r["n_client_discard"]
        agent_d = r["n_agent_discard"]
        print(f"{r['dataset'][:12]:<12} {r['n_compared']:>6} {client_d:>8} {agent_d:>8} {r['tp']:>5} {r['fp']:>5} {r['fn']:>5} {r['tn']:>5} {r['precision']:>7.3f} {r['recall']:>7.3f} {r['f1']:>7.3f}")
        total_tp += r["tp"]
        total_fp += r["fp"]
        total_fn += r["fn"]
        total_tn += r["tn"]

    # Pooled metrics
    pooled_precision = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0
    pooled_recall = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0
    pooled_f1 = 2 * pooled_precision * pooled_recall / (pooled_precision + pooled_recall) if (pooled_precision + pooled_recall) > 0 else 0

    print(f"{'─'*12} {'─'*6} {'─'*8} {'─'*8} {'─'*5} {'─'*5} {'─'*5} {'─'*5} {'─'*7} {'─'*7} {'─'*7}")
    print(f"{'POOLED':<12} {sum(r['n_compared'] for r in all_results):>6} {sum(r['n_client_discard'] for r in all_results):>8} {sum(r['n_agent_discard'] for r in all_results):>8} {total_tp:>5} {total_fp:>5} {total_fn:>5} {total_tn:>5} {pooled_precision:>7.3f} {pooled_recall:>7.3f} {pooled_f1:>7.3f}")

    # Macro-averaged metrics
    macro_precision = sum(r["precision"] for r in all_results) / len(all_results) if all_results else 0
    macro_recall = sum(r["recall"] for r in all_results) / len(all_results) if all_results else 0
    macro_f1 = sum(r["f1"] for r in all_results) / len(all_results) if all_results else 0
    print(f"{'MACRO-AVG':<12} {'':>6} {'':>8} {'':>8} {'':>5} {'':>5} {'':>5} {'':>5} {macro_precision:>7.3f} {macro_recall:>7.3f} {macro_f1:>7.3f}")

    # Write summary JSON
    output_path = Path("/Users/jeremyalston/Perfect/autosurvey-outputs/cross-dataset-propositions/v4_cross_dataset_comparison.json")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = {
        "version": "v4_evidence_family",
        "datasets": all_results,
        "pooled": {
            "tp": total_tp, "fp": total_fp, "fn": total_fn, "tn": total_tn,
            "precision": round(pooled_precision, 4),
            "recall": round(pooled_recall, 4),
            "f1": round(pooled_f1, 4),
        },
        "macro_averaged": {
            "precision": round(macro_precision, 4),
            "recall": round(macro_recall, 4),
            "f1": round(macro_f1, 4),
        },
        "merge_stats": merge_stats,
    }
    with open(output_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"\nSummary written to: {output_path}")

    # Write per-dataset false negatives and false positives for review
    for ds_short, compared in all_compared.items():
        fns = [c for c in compared if c["client_judgment"] == "DISCARD" and c["agent_judgment"] != "DISCARD"]
        fps = [c for c in compared if c["client_judgment"] == "KEEP" and c["agent_judgment"] == "DISCARD"]
        tps = [c for c in compared if c["client_judgment"] == "DISCARD" and c["agent_judgment"] == "DISCARD"]

        fn_path = Path(f"/Users/jeremyalston/Perfect/autosurvey-outputs/cross-dataset-propositions/v4_{ds_short}_false_negatives.json")
        fp_path = Path(f"/Users/jeremyalston/Perfect/autosurvey-outputs/cross-dataset-propositions/v4_{ds_short}_false_positives.json")
        with open(fn_path, "w") as f:
            json.dump(fns, f, indent=2)
        with open(fp_path, "w") as f:
            json.dump(fps, f, indent=2)
        print(f"  {ds_short}: {len(fns)} FNs, {len(fps)} FPs, {len(tps)} TPs")


if __name__ == "__main__":
    main()
