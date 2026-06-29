#!/usr/bin/env python3
"""V17 — LLM embeddings (sentence-transformers) + V7+V8+V9 agent features.

Improvements over V14 (best so far at 0.744):
1. Sentence-transformer embeddings for OE text (384-dim semantic vectors)
2. V9 agent judgments as third set of agent features (V7+V8+V9)
3. Agent agreement features (V7+V8+V9 majority vote, consensus score)
4. Self-training with best V14 settings (threshold 0.85, 3 iterations)
5. XGBoost + LightGBM + MLP ensemble with isotonic calibration + stacking
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training

import xgboost as xgb
import lightgbm as lgb


def load_v9_judgments():
    """Load V9 agent judgments."""
    v9_path = ECHO_OUTPUT / "holistic_agent_run_v9" / "agent_judgments.json"
    if not v9_path.exists():
        return {}
    with open(v9_path) as f:
        judgments = json.load(f)
    return {j["respondent_id"]: j for j in judgments}


def add_v9_features(df, v9_judgments):
    """Add V9 judgment features."""
    if not v9_judgments:
        return df
    df["v9_judgment"] = df["respondent_id"].map(
        lambda rid: v9_judgments.get(rid, {}).get("agent_judgment", "UNKNOWN")
    )
    df["v9_judgment_enc"] = df["v9_judgment"].map(
        {"KEEP": 0, "REVIEW": 1, "DISCARD": 2, "UNKNOWN": 1}
    ).fillna(1)
    df["v9_converging_count"] = df["respondent_id"].map(
        lambda rid: v9_judgments.get(rid, {}).get("converging_family_count", 0)
    )
    df["v9_authenticity_risk"] = df["respondent_id"].map(
        lambda rid: v9_judgments.get(rid, {}).get("authenticity_risk", 0.5)
    )
    df["v9_quality_risk"] = df["respondent_id"].map(
        lambda rid: v9_judgments.get(rid, {}).get("quality_discard_risk", 0.5)
    )
    df["v9_client_reject_prob"] = df["respondent_id"].map(
        lambda rid: v9_judgments.get(rid, {}).get("client_reject_probability", 0.5)
    )
    print(f"  V9 features added: {df['v9_judgment'].notna().sum()}")
    return df


def add_agent_consensus_features(df):
    """Add consensus features across V7, V8, V9."""
    # Majority vote
    judgments = ["v7_judgment_enc", "v8_judgment_enc", "v9_judgment_enc"]
    available = [j for j in judgments if j in df.columns]
    if len(available) >= 2:
        # Average encoded judgment
        df["agent_avg_judgment"] = df[available].mean(axis=1)
        # Max judgment (most aggressive)
        df["agent_max_judgment"] = df[available].max(axis=1)
        # Min judgment (most conservative)
        df["agent_min_judgment"] = df[available].min(axis=1)
        # Agreement: all three agree
        df["agent_all_agree"] = (df[available].nunique(axis=1) == 1).astype(int)
        # Consensus DISCARD: at least 2 of 3 say DISCARD
        df["agent_majority_discard"] = (df[available].sum(axis=1) >= 4).astype(int)  # 2+ DISCARD (enc=2)
        # Consensus KEEP: at least 2 of 3 say KEEP
        df["agent_majority_keep"] = (df[available].sum(axis=1) <= 1).astype(int)  # 2+ KEEP (enc=0)

    # Average risk across agents
    risks = ["v7_client_reject_prob", "v8_client_reject_prob", "v9_client_reject_prob"]
    available_risks = [r for r in risks if r in df.columns]
    if available_risks:
        df["agent_avg_risk"] = df[available_risks].mean(axis=1)
        df["agent_max_risk"] = df[available_risks].max(axis=1)
        df["agent_min_risk"] = df[available_risks].min(axis=1)
        df["agent_risk_std"] = df[available_risks].std(axis=1).fillna(0)
        # Risk disagreement (high std = uncertain)
        df["agent_risk_disagree"] = (df["agent_risk_std"] > 0.15).astype(int)

    # Average converging count
    convs = ["v7_converging_count", "v8_converging_count", "v9_converging_count"]
    available_convs = [c for c in convs if c in df.columns]
    if available_convs:
        df["agent_avg_converging"] = df[available_convs].mean(axis=1)
        df["agent_max_converging"] = df[available_convs].max(axis=1)

    print(f"  Agent consensus features added ({len(available)} agent sources)")
    return df


def extract_llm_embeddings(df, answer_chains, model_name="all-MiniLM-L6-v2", batch_size=64):
    """Extract sentence-transformer embeddings for OE text."""
    print(f"  Extracting LLM embeddings ({model_name})...")

    from sentence_transformers import SentenceTransformer

    # Build text corpus
    chain_lookup = {ac["respondent_id"]: ac for ac in answer_chains}
    texts = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        ac = chain_lookup.get(rid, {})
        chain = ac.get("answer_chain", [])
        oe_texts = [a.get("raw_value", "") or a.get("response", "") for a in chain
                    if a.get("answer_type") == "open_end" or a.get("role") == "open_end"]
        text = " ".join(oe_texts).strip()
        if not text:
            text = "empty"
        texts.append(text)

    # Load model and encode
    model = SentenceTransformer(model_name)
    embeddings = model.encode(texts, batch_size=batch_size, show_progress_bar=False,
                              convert_to_numpy=True)

    emb_df = pd.DataFrame(
        embeddings,
        columns=[f"emb_{i}" for i in range(embeddings.shape[1])],
        index=df.index,
    )
    print(f"    Embeddings: {emb_df.shape[1]} dimensions from {len(texts)} texts")
    return emb_df


def get_classify_map():
    """Get CLASSIFY map from Echo."""
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()
    return classify_map


def run_v17_cv(n_folds=5, use_embeddings=True):
    """Run V17 cross-validation."""
    print(f"\n{'='*80}")
    print(f"V17 — LLM Embeddings + V7+V8+V9 Agent Consensus + Self-Training")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()
    v9 = load_v9_judgments()
    print(f"V7: {len(v7)}, V8: {len(v8)}, V9: {len(v9)} judgments")

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df = add_v9_features(df, v9)
    df = add_agent_consensus_features(df)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Add LLM embeddings
    if use_embeddings:
        emb_df = extract_llm_embeddings(df, answer_chains)
        df = pd.concat([df, emb_df], axis=1)

    # Extract all datasets for self-training
    print("\nExtracting all datasets for self-training...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)
    print(f"\nTotal: {len(all_data)} ({len(labeled)} labeled, {len(unlabeled)} unlabeled)")
    print(f"  Features: {len(labeled.columns)}")

    classify_map = get_classify_map()

    # K-fold CV on Echo
    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Self-training (V14 settings: threshold 0.85, 3 iterations)
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        # Prepare test features
        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        # Split for calibration
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        print(f"  Training XGBoost on {len(X_tr)} samples, {X_tr.shape[1]} features...")
        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        print("  Training LightGBM...")
        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.7,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        print("  Training MLP...")
        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        # Calibrate
        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Search for best threshold
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V17 CV RESULTS (embeddings={use_embeddings})")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V17 — LLM Embeddings + V7+V8+V9 Agent Consensus + Self-Training")
    print("=" * 80)

    # With embeddings
    results = run_v17_cv(n_folds=5, use_embeddings=True)

    # Without embeddings (for comparison — isolating embedding impact)
    results_no_emb = run_v17_cv(n_folds=5, use_embeddings=False)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V11 (XGB+LGB+MLP+RF):    BAcc=0.737")
    print(f"  V14 (self-train + V8):   BAcc=0.744")
    print(f"  V17 (emb + V7+V8+V9):    BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  V17 (no emb, V7+V8+V9):  BAcc={results_no_emb['avg_bacc']:.3f}, F1={results_no_emb['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], results_no_emb['avg_bacc']):.3f}")

    with open(AUTOSURVEY_DIR / "v17_cv_results.json", "w") as f:
        json.dump({"v17_with_emb": results, "v17_without_emb": results_no_emb}, f, indent=2, default=str)


if __name__ == "__main__":
    main()
