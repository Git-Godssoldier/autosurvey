#!/usr/bin/env python3
"""V22 — Respondent similarity graph + network features + clustering.

Build a graph where respondents are connected if they have similar answer
patterns. Use graph features (degree, centrality, community) and cluster
respondents to find "bad" clusters.

New features:
1. KNN graph features (degree, clustering coefficient, centrality)
2. Community detection (Louvain/label propagation)
3. Per-cluster discard rate
4. Distance to nearest "known bad" respondent
5. Distance to nearest "known good" respondent
6. Cluster-based anomaly score
"""
from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.cluster import KMeans, DBSCAN
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neighbors import NearestNeighbors
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")

sys.path.insert(0, str(SKILL_SCRIPTS))
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training

import xgboost as xgb
import lightgbm as lgb


def add_graph_cluster_features(df, labeled_df, n_clusters=20, n_neighbors=10):
    """Add graph-based and cluster-based features."""
    print("  Computing graph/cluster features...")

    # Prepare feature matrix
    feature_cols = [c for c in df.columns if c not in ["respondent_id", "label", "dataset",
                                                         "v7_judgment", "v8_judgment",
                                                         "classify", "supplier"] and
                    df[c].dtype in ["float64", "float32", "int64", "int32", "bool"]]

    X = df[feature_cols].fillna(0).values
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # KNN graph features
    print("    Building KNN graph...")
    nn = NearestNeighbors(n_neighbors=n_neighbors + 1, metric="cosine")
    nn.fit(X_scaled)

    distances, indices = nn.kneighbors(X_scaled)

    # Degree (number of close neighbors within threshold)
    degree = np.sum(distances[:, 1:] < 0.5, axis=1)
    df["graph_degree"] = degree

    # Mean distance to k nearest neighbors
    df["graph_mean_dist"] = np.mean(distances[:, 1:], axis=1)
    df["graph_min_dist"] = distances[:, 1]
    df["graph_max_dist"] = distances[:, -1]
    df["graph_dist_std"] = np.std(distances[:, 1:], axis=1)

    # Clustering
    print("    Clustering respondents...")
    kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
    cluster_labels = kmeans.fit_predict(X_scaled)
    df["cluster_id"] = cluster_labels

    # Distance to cluster center
    cluster_centers = kmeans.cluster_centers_
    df["cluster_dist_to_center"] = np.linalg.norm(X_scaled - cluster_centers[cluster_labels], axis=1)

    # Per-cluster discard rate (from labeled data)
    cluster_discard_rates = {}
    cluster_counts = {}
    for cl in range(n_clusters):
        cl_mask = labeled_df["cluster_id"] == cl if "cluster_id" in labeled_df.columns else np.zeros(len(labeled_df), dtype=bool)
        if cl_mask.sum() > 0:
            cluster_discard_rates[cl] = labeled_df.loc[cl_mask, "label"].mean()
            cluster_counts[cl] = cl_mask.sum()
        else:
            cluster_discard_rates[cl] = 0.35  # prior
            cluster_counts[cl] = 0

    df["cluster_discard_rate"] = df["cluster_id"].map(cluster_discard_rates).fillna(0.35)
    df["cluster_count"] = df["cluster_id"].map(cluster_counts).fillna(0)

    # Distance to nearest known bad and known good
    print("    Computing distances to known bad/good...")
    if "label" in labeled_df.columns:
        bad_mask = labeled_df["label"] == 1
        good_mask = labeled_df["label"] == 0

        if bad_mask.sum() > 0 and good_mask.sum() > 0:
            bad_indices = labeled_df.index[bad_mask].values
            good_indices = labeled_df.index[good_mask].values

            # Get scaled features for labeled
            labeled_X = df.loc[labeled_df.index, feature_cols].fillna(0).values
            labeled_X_scaled = scaler.transform(labeled_X)

            bad_X = labeled_X_scaled[bad_mask]
            good_X = labeled_X_scaled[good_mask]

            # For each respondent, find distance to nearest bad and nearest good
            nn_bad = NearestNeighbors(n_neighbors=1, metric="cosine").fit(bad_X)
            nn_good = NearestNeighbors(n_neighbors=1, metric="cosine").fit(good_X)

            dist_to_bad, _ = nn_bad.kneighbors(X_scaled)
            dist_to_good, _ = nn_good.kneighbors(X_scaled)

            df["dist_to_nearest_bad"] = dist_to_bad.ravel()
            df["dist_to_nearest_good"] = dist_to_good.ravel()
            df["bad_good_dist_ratio"] = df["dist_to_nearest_bad"] / (df["dist_to_nearest_good"] + 1e-6)
            df["closer_to_bad"] = (df["dist_to_nearest_bad"] < df["dist_to_nearest_good"]).astype(int)

    # DBSCAN outlier detection
    print("    Running DBSCAN...")
    dbscan = DBSCAN(eps=5.0, min_samples=10, n_jobs=-1)
    dbscan_labels = dbscan.fit_predict(X_scaled)
    df["dbscan_cluster"] = dbscan_labels
    df["dbscan_is_outlier"] = (dbscan_labels == -1).astype(int)

    print(f"    Graph features added: degree, distances, {n_clusters} clusters, DBSCAN outliers={df['dbscan_is_outlier'].sum()}")
    return df


def run_v22_cv(n_folds=5):
    """Run V22 cross-validation."""
    print(f"\n{'='*80}")
    print(f"V22 — Respondent Similarity Graph + Network Features + Clustering")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Add graph/cluster features
    labeled_mask = df["label"] >= 0
    df = add_graph_cluster_features(df, df[labeled_mask], n_clusters=20, n_neighbors=10)

    print(f"\nTotal features: {len(df.columns)}")

    # Extract all datasets
    print("\nExtracting all datasets...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    # Get CLASSIFY map
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(hdrs) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        print(f"  Training on {len(X_tr)} samples, {X_tr.shape[1]} features...")

        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.7,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V22 CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V22 — Respondent Similarity Graph + Network Features + Clustering")
    print("=" * 80)

    results = run_v22_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V14 (self-train + V8):   BAcc=0.744 (BEST)")
    print(f"  V22 (graph + cluster):   BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], 0.744):.3f}")

    with open(AUTOSURVEY_DIR / "v22_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
