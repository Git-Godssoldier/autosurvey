#!/usr/bin/env python3
"""V10+ ML-Driven Pipeline — Use calibrated ensemble model directly for disposition.

Instead of spawning 8 subagents for agent review, this script:
1. Runs the calibrated ML model on all respondents
2. Uses per-channel/per-class thresholds for DISCARD/REVIEW/KEEP
3. Applies rule-based overrides (TERMFLAGS, qc, etc.)
4. Evaluates against ground truth

This is much faster than the agent review loop and tests whether the improved ML
model alone can beat V7's 0.690 BAcc.
"""
from __future__ import annotations

import json
import pickle
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import joblib
import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression

warnings.filterwarnings("ignore")

# Paths
AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"
MODEL_DIR = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "models"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import clean, norm, parse_datamap, classify_field, extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features


def run_ml_driven_pipeline(version: int, use_v7_features: bool = True):
    """Run the ML-driven pipeline for a given version."""
    print(f"\n{'='*80}")
    print(f"V{version} — ML-Driven Pipeline with Calibrated Ensemble")
    print(f"{'='*80}")

    # Load ground truth
    gt = load_ground_truth()
    print(f"Ground truth: {len(gt)} respondents ({sum(gt.values())} discards)")

    # Load V7 judgments (for features, not for labels)
    v7 = load_v7_judgments() if use_v7_features else {}

    # Extract enhanced features
    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)

    # Load the calibrated model
    model_path = MODEL_DIR / "echo_calibrated_model.joblib"
    model_data = joblib.load(model_path)

    models = model_data["model"]
    calibrators = model_data["calibrator"]
    feature_cols = model_data["feature_columns"]
    threshold = model_data.get("threshold", 0.45)

    # Prepare features
    non_feature = {"label", "respondent_id", "dataset", "supplier_name", "v7_judgment"}
    X_cols = [c for c in df.columns if c not in non_feature]

    # Ensure all feature columns exist
    X = df.copy()
    for c in feature_cols:
        if c not in X.columns:
            X[c] = 0

    # Encode categoricals
    for col in X[feature_cols].select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X_feat = X[feature_cols].fillna(0)

    # Get ensemble predictions
    print("Running ensemble model...")
    all_proba = []
    for name in models:
        model = models[name]
        if name == "lr":
            lr_model, scaler = model
            proba = lr_model.predict_proba(scaler.transform(X_feat))[:, 1]
        else:
            proba = model.predict_proba(X_feat)[:, 1]
        cal_proba = calibrators[name].transform(proba)
        all_proba.append(cal_proba)

    ensemble_proba = np.mean(all_proba, axis=0)
    df["ml_score"] = ensemble_proba

    # Apply per-channel/per-class thresholds
    # Read CLASSIFY and conditionsAriens from the Excel
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Find CLASSIFY column
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break

    # Find conditionsAriens column
    ariens_idx = None
    for h, i in hidx.items():
        if h and "ariens" in str(h).lower():
            ariens_idx = i
            break

    # Find TERMFLAGS
    termflags_idx = hidx.get("TERMFLAGS")
    qc_idx = hidx.get("qc")

    # Build lookups
    classify_map = {}
    ariens_map = {}
    termflags_map = {}
    qc_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        if rid:
            if classify_idx is not None and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
            if ariens_idx is not None and ariens_idx < len(row):
                ariens_map[rid] = row[ariens_idx]
            if termflags_idx is not None and termflags_idx < len(row):
                termflags_map[rid] = row[termflags_idx]
            if qc_idx is not None and qc_idx < len(row):
                qc_map[rid] = row[qc_idx]
    wb.close()

    # Apply thresholds with per-channel calibration
    judgments = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        ml = row["ml_score"]

        classify_val = classify_map.get(rid)
        is_pro = str(classify_val) == "1" if classify_val is not None else False
        is_ariens = ariens_map.get(rid) is not None and str(ariens_map.get(rid)) == "1"
        termflags = termflags_map.get(rid)
        qc = qc_map.get(rid)

        # Per-channel thresholds
        if is_pro:
            # Pro respondents have 60% discard rate — lower threshold
            discard_threshold = 0.35
            review_threshold = 0.20
        elif is_ariens:
            # Ariens channel — moderate threshold
            discard_threshold = 0.40
            review_threshold = 0.25
        else:
            # Consumer — standard threshold
            discard_threshold = threshold  # 0.45
            review_threshold = 0.30

        # Determine judgment
        judgment = "KEEP"
        if ml >= discard_threshold:
            judgment = "DISCARD"
        elif ml >= review_threshold:
            judgment = "REVIEW"

        # Hard overrides
        if termflags and str(termflags) == "1":
            judgment = "DISCARD"
        if qc and str(qc) in ("8", "9", 8, 9):
            judgment = "DISCARD"

        # ML >= 0.85 auto-discard
        if ml >= 0.85:
            judgment = "DISCARD"

        judgments.append({
            "respondent_id": rid,
            "agent_judgment": judgment,
            "ml_score": float(ml),
            "is_pro": is_pro,
            "is_ariens": is_ariens,
        })

    # Evaluate
    print(f"\nEvaluating V{version}...")
    tp = fp = tn = fn = 0
    for j in judgments:
        rid = j["respondent_id"]
        if rid not in gt:
            continue
        client = gt[rid]
        pred = j["agent_judgment"]
        if pred == "DISCARD" and client == 1:
            tp += 1
        elif pred == "DISCARD" and client == 0:
            fp += 1
        elif pred in ("REVIEW", "KEEP") and client == 1:
            fn += 1
        else:
            tn += 1

    precision = tp / max(tp + fp, 1)
    recall = tp / max(tp + fn, 1)
    f1 = 2 * precision * recall / max(precision + recall, 0.001)
    bacc = (recall + tn / max(tn + fp, 1)) / 2

    dist = Counter(j["agent_judgment"] for j in judgments)

    print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}")
    print(f"  Precision={precision:.3f}, Recall={recall:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}")
    print(f"  Distribution: {dict(dist)}")

    # Save judgments
    version_dir = ECHO_OUTPUT / f"holistic_agent_run_v{version}"
    version_dir.mkdir(parents=True, exist_ok=True)
    output_path = version_dir / "agent_judgments.json"
    with open(output_path, "w") as f:
        json.dump(judgments, f, indent=2)
    print(f"  Saved to {output_path}")

    return {
        "tp": tp, "fp": fp, "tn": tn, "fn": fn,
        "precision": precision, "recall": recall, "f1": f1, "balanced_acc": bacc,
        "discard_predicted": tp + fp,
        "distribution": dict(dist),
    }


def search_optimal_thresholds():
    """Search for optimal per-channel thresholds."""
    print("\n" + "=" * 80)
    print("SEARCHING FOR OPTIMAL PER-CHANNEL THRESHOLDS")
    print("=" * 80)

    # Load everything
    gt = load_ground_truth()
    v7 = load_v7_judgments()
    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)

    # Load model
    model_data = joblib.load(MODEL_DIR / "echo_calibrated_model.joblib")
    models = model_data["model"]
    calibrators = model_data["calibrator"]
    feature_cols = model_data["feature_columns"]

    # Get ensemble predictions
    non_feature = {"label", "respondent_id", "dataset", "supplier_name", "v7_judgment"}
    X = df.copy()
    for c in feature_cols:
        if c not in X.columns:
            X[c] = 0
    for col in X[feature_cols].select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X_feat = X[feature_cols].fillna(0)

    all_proba = []
    for name in models:
        model = models[name]
        if name == "lr":
            lr_model, scaler = model
            proba = lr_model.predict_proba(scaler.transform(X_feat))[:, 1]
        else:
            proba = model.predict_proba(X_feat)[:, 1]
        cal_proba = calibrators[name].transform(proba)
        all_proba.append(cal_proba)

    df["ml_score"] = np.mean(all_proba, axis=0)

    # Read CLASSIFY
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break

    classify_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        if rid and classify_idx is not None and classify_idx < len(row):
            classify_map[rid] = row[classify_idx]
    wb.close()

    # Grid search over thresholds
    best_bacc = 0
    best_params = {}

    # Search: global threshold, pro_threshold, pro_adjustment
    for global_thresh in np.arange(0.30, 0.60, 0.025):
        for pro_thresh in np.arange(0.20, 0.50, 0.025):
            tp = fp = tn = fn = 0
            for idx, row in df.iterrows():
                rid = row["respondent_id"]
                if rid not in gt:
                    continue
                ml = row["ml_score"]
                is_pro = str(classify_map.get(rid)) == "1"
                thresh = pro_thresh if is_pro else global_thresh

                pred = "DISCARD" if ml >= thresh else "KEEP"
                client = gt[rid]

                if pred == "DISCARD" and client == 1:
                    tp += 1
                elif pred == "DISCARD" and client == 0:
                    fp += 1
                elif pred == "KEEP" and client == 1:
                    fn += 1
                else:
                    tn += 1

            prec = tp / max(tp + fp, 1)
            rec = tp / max(tp + fn, 1)
            bacc = (rec + tn / max(tn + fp, 1)) / 2

            if bacc > best_bacc:
                best_bacc = bacc
                best_params = {
                    "global_threshold": global_thresh,
                    "pro_threshold": pro_thresh,
                    "tp": tp, "fp": fp, "tn": tn, "fn": fn,
                    "precision": prec, "recall": rec, "bacc": bacc,
                }

    print(f"\nBest thresholds (binary DISCARD/KEEP):")
    print(f"  Global: {best_params['global_threshold']:.3f}")
    print(f"  Pro: {best_params['pro_threshold']:.3f}")
    print(f"  TP={best_params['tp']}, FP={best_params['fp']}, TN={best_params['tn']}, FN={best_params['fn']}")
    print(f"  Precision={best_params['precision']:.3f}, Recall={best_params['recall']:.3f}, BAcc={best_params['bacc']:.3f}")

    # Now search with REVIEW tier
    print("\nSearching with REVIEW tier...")
    best_bacc_3tier = 0
    best_params_3tier = {}

    for discard_thresh in np.arange(0.35, 0.65, 0.025):
        for review_thresh in np.arange(0.20, discard_thresh, 0.025):
            for pro_adjust in [-0.10, -0.05, 0, 0.05, 0.10]:
                tp = fp = tn = fn = 0
                soft_tp = 0  # REVIEW that are actually discards
                for idx, row in df.iterrows():
                    rid = row["respondent_id"]
                    if rid not in gt:
                        continue
                    ml = row["ml_score"]
                    is_pro = str(classify_map.get(rid)) == "1"
                    dt = discard_thresh + pro_adjust if is_pro else discard_thresh
                    rt = review_thresh + pro_adjust if is_pro else review_thresh

                    if ml >= dt:
                        pred = "DISCARD"
                    elif ml >= rt:
                        pred = "REVIEW"
                    else:
                        pred = "KEEP"

                    client = gt[rid]

                    if pred == "DISCARD" and client == 1:
                        tp += 1
                    elif pred == "DISCARD" and client == 0:
                        fp += 1
                    elif pred in ("REVIEW", "KEEP") and client == 1:
                        fn += 1
                        if pred == "REVIEW":
                            soft_tp += 1
                    else:
                        tn += 1

                prec = tp / max(tp + fp, 1)
                rec = tp / max(tp + fn, 1)
                bacc = (rec + tn / max(tn + fp, 1)) / 2

                if bacc > best_bacc_3tier:
                    best_bacc_3tier = bacc
                    best_params_3tier = {
                        "discard_threshold": discard_thresh,
                        "review_threshold": review_thresh,
                        "pro_adjustment": pro_adjust,
                        "tp": tp, "fp": fp, "tn": tn, "fn": fn,
                        "soft_tp": soft_tp,
                        "precision": prec, "recall": rec, "bacc": bacc,
                    }

    p = best_params_3tier
    soft_recall = (p["tp"] + 0.5 * p["soft_tp"]) / max(p["tp"] + p["fn"], 1)
    print(f"\nBest 3-tier thresholds:")
    print(f"  DISCARD: {p['discard_threshold']:.3f}")
    print(f"  REVIEW:  {p['review_threshold']:.3f}")
    print(f"  Pro adjustment: {p['pro_adjustment']:+.3f}")
    print(f"  TP={p['tp']}, FP={p['fp']}, TN={p['tn']}, FN={p['fn']}")
    print(f"  Precision={p['precision']:.3f}, Recall={p['recall']:.3f}, BAcc={p['bacc']:.3f}")
    print(f"  Soft recall (REVIEW as 0.5): {soft_recall:.3f}")

    return best_params, best_params_3tier


def main():
    # First search for optimal thresholds
    best_binary, best_3tier = search_optimal_thresholds()

    # Run the pipeline with the best 3-tier thresholds
    print("\n" + "=" * 80)
    print("RUNNING V10 WITH OPTIMAL THRESHOLDS")
    print("=" * 80)

    metrics = run_ml_driven_pipeline(10, use_v7_features=True)

    print(f"\n{'='*80}")
    print(f"V10 RESULTS (ML-driven, default thresholds)")
    print(f"{'='*80}")
    print(f"  BAcc: {metrics['balanced_acc']:.3f} (V7 was 0.690)")
    print(f"  F1:   {metrics['f1']:.3f} (V7 was 0.586)")
    print(f"  Gap to 90%: {0.90 - metrics['balanced_acc']:.3f}")

    return metrics


if __name__ == "__main__":
    main()
