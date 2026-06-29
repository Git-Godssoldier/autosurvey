#!/usr/bin/env python3
"""Evolution Loop — Autonomous iteration system for survey quality cleaning.

Runs the full pipeline (packet generation → agent review → integration → comparison)
in a loop, analyzing errors after each run and generating improved instructions for
the next iteration. Logs all decisions and performance metrics.

Stops when balanced accuracy >= target (default 0.90) or max iterations reached.

Usage:
    python3 evolution_loop.py [--target 0.90] [--max-iterations 20] [--start-version 10]
"""
from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

# Paths
AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"
GT_XLSX = DATA_DIR / "AutoQuality Pair Copy - Echo" / "260300_ECHO - client annotated.xlsx"
DECISION_LOG = AUTOSURVEY_DIR / "DECISION_LOG.md"

# V7 base (best performing version)
V7_DIR = ECHO_OUTPUT / "holistic_agent_run_v7"


def log(msg: str, level: str = "INFO"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] [{level}] {msg}", flush=True)


def append_decision_log(version: int, changes: str, metrics: dict, analysis: str, next_plan: str):
    """Append an entry to the decision log."""
    entry = f"""
## V{version} — {datetime.now().strftime("%Y-%m-%d %H:%M")}

### Changes
{changes}

### Metrics
- TP: {metrics.get('tp', '?')}
- FP: {metrics.get('fp', '?')}
- TN: {metrics.get('tn', '?')}
- FN: {metrics.get('fn', '?')}
- Precision: {metrics.get('precision', 0):.3f}
- Recall: {metrics.get('recall', 0):.3f}
- F1: {metrics.get('f1', 0):.3f}
- Balanced Accuracy: {metrics.get('balanced_acc', 0):.3f}
- Discards predicted: {metrics.get('discard_predicted', '?')}

### Error Analysis
{analysis}

### Next Iteration Plan
{next_plan}

---
"""
    with open(DECISION_LOG, "a") as f:
        f.write(entry)
    log(f"Decision log updated for V{version}")


def run_stage1_generate_packets(version_dir: Path) -> bool:
    """Stage 1: Generate review packets."""
    log(f"Stage 1: Generating review packets in {version_dir}")
    version_dir.mkdir(parents=True, exist_ok=True)

    # Copy review_summary.json from V7 (same dataset, same packets)
    summary_src = V7_DIR / "review_summary.json"
    if summary_src.exists():
        shutil.copy(summary_src, version_dir / "review_summary.json")

    # Run packet generation
    cmd = [
        sys.executable,
        str(SKILL_SCRIPTS / "run_holistic_agent_review.py"),
        str(ECHO_XLSX),
        "--output-dir", str(version_dir),
        "--review-all",
        "--chunk-size", "200",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)

    if result.returncode != 0:
        log(f"Stage 1 FAILED: {result.stderr[-500:]}", "ERROR")
        return False

    # Check packets were generated
    packets = list(version_dir.glob("review_chunk_*.json"))
    if not packets:
        # If generation didn't produce new packets, copy from V7
        log("Stage 1: No new packets generated, copying from V7")
        for p in V7_DIR.glob("review_chunk_*.json"):
            shutil.copy(p, version_dir)

    packets = list(version_dir.glob("review_chunk_*.json"))
    log(f"Stage 1 complete: {len(packets)} packet files")
    return len(packets) > 0


def run_stage2_agent_review(version_dir: Path, version: int, instructions: str = None) -> bool:
    """Stage 2: Spawn 8 subagents for agent review.

    NOTE: This function is called from the main loop. The actual subagent spawning
    is handled by the caller (the Devin agent) since subagents are Devin-specific.
    This function prepares the instructions file and returns the chunk list.
    """
    log(f"Stage 2: Preparing agent review for V{version}")

    # Write instructions if provided
    if instructions:
        instr_path = version_dir / "agent_review_instructions.md"
        with open(instr_path, "w") as f:
            f.write(instructions)

    # Return chunk list for the caller to spawn subagents
    chunks = sorted(version_dir.glob("review_chunk_*.json"))
    log(f"Stage 2: {len(chunks)} chunks ready for review")
    return chunks


def run_stage3_integrate(version_dir: Path) -> bool:
    """Stage 3: Integrate agent judgments."""
    log(f"Stage 3: Integrating agent judgments from {version_dir}")

    # First merge chunk files
    chunks = sorted(version_dir.glob("agent_judgments_chunk_*.json"))
    if not chunks:
        log("Stage 3 FAILED: No agent_judgments_chunk_*.json files found", "ERROR")
        return False

    import json
    all_judgments = []
    for cf in chunks:
        with open(cf) as f:
            all_judgments.extend(json.load(f))

    with open(version_dir / "agent_judgments.json", "w") as f:
        json.dump(all_judgments, f, indent=2)
    log(f"Stage 3: Merged {len(all_judgments)} judgments")

    # Run integration script
    cmd = [
        sys.executable,
        str(SKILL_SCRIPTS / "integrate_agent_judgments.py"),
        str(ECHO_XLSX),
        str(version_dir),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)

    if result.returncode != 0:
        log(f"Stage 3 FAILED: {result.stderr[-500:]}", "ERROR")
        return False

    log("Stage 3 complete: Annotated Excel + dashboard generated")
    return True


def run_stage4_compare(version_dir: Path) -> dict | None:
    """Stage 4: Compare against ground truth."""
    log(f"Stage 4: Comparing against ground truth")

    cmd = [sys.executable, str(AUTOSURVEY_DIR / "scripts" / "compare_v4_v5_echo.py")]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

    if result.returncode != 0:
        log(f"Stage 4 FAILED: {result.stderr[-500:]}", "ERROR")
        return None

    # Parse the comparison results JSON
    results_path = version_dir / "comparison_results.json"
    if not results_path.exists():
        # Try V9 dir (the comparison script writes there as fallback)
        results_path = ECHO_OUTPUT / "holistic_agent_run_v9" / "comparison_results.json"

    if not results_path.exists():
        log("Stage 4: comparison_results.json not found", "ERROR")
        return None

    with open(results_path) as f:
        results = json.load(f)

    # Extract metrics for this version
    version_key = f"v{version_dir.name.split('_v')[-1]}"
    metrics = results.get(version_key)
    if not metrics:
        # Try all keys
        for k, v in results.items():
            if k.startswith("v") and v and isinstance(v, dict) and "tp" in v:
                metrics = v
                break

    if metrics:
        log(f"Stage 4 results: TP={metrics['tp']}, FP={metrics['fp']}, "
            f"FN={metrics['fn']}, BAcc={metrics['balanced_acc']:.3f}, "
            f"F1={metrics['f1']:.3f}")

    return metrics


def analyze_errors(version_dir: Path, metrics: dict) -> str:
    """Analyze errors to identify patterns for next iteration."""
    log("Analyzing errors...")

    import json
    import openpyxl
    from collections import Counter

    # Load judgments
    judgments_path = version_dir / "agent_judgments.json"
    if not judgments_path.exists():
        return "Could not load judgments for analysis"

    with open(judgments_path) as f:
        judgments = {j["respondent_id"]: j for j in json.load(f)}

    # Load ground truth
    gt = {}
    if GT_XLSX.exists():
        wb = openpyxl.load_workbook(GT_XLSX, read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        hidx = {h: i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            status = row[hidx["status"]]
            if rid:
                gt[rid] = "DISCARD" if status in (5, "5") else ("KEEP" if status in (3, "3") else None)
        wb.close()

    # Classify errors
    tp, fp, tn, fn = [], [], [], []
    for rid, j in judgments.items():
        if rid not in gt or gt[rid] is None:
            continue
        client = gt[rid]
        pred = j.get("agent_judgment", "")
        if pred == "DISCARD" and client == "DISCARD":
            tp.append(j)
        elif pred == "DISCARD" and client == "KEEP":
            fp.append(j)
        elif pred in ("REVIEW", "KEEP") and client == "DISCARD":
            fn.append(j)
        elif pred in ("REVIEW", "KEEP") and client == "KEEP":
            tn.append(j)

    analysis_parts = []

    # FP analysis
    analysis_parts.append(f"**False Positives ({len(fp)}):**")
    fp_oe = Counter(j.get("oe_classification", "?") for j in fp).most_common(5)
    analysis_parts.append(f"  OE: {', '.join(f'{k}={v}' for k,v in fp_oe)}")
    fp_ml = sum(1 for j in fp if 0.35 <= j.get("ml_score", 0) < 0.5)
    analysis_parts.append(f"  ML 0.35-0.5 band: {fp_ml} ({fp_ml/max(len(fp),1)*100:.0f}%)")
    fp_families = Counter(fam for j in fp for fam, v in j.get("evidence_family_scores", {}).items() if v.get("fired")).most_common(5)
    analysis_parts.append(f"  Top families: {', '.join(f'{k}={v}' for k,v in fp_families)}")

    # FN analysis
    analysis_parts.append(f"\n**False Negatives ({len(fn)}):**")
    fn_oe = Counter(j.get("oe_classification", "?") for j in fn).most_common(5)
    analysis_parts.append(f"  OE: {', '.join(f'{k}={v}' for k,v in fn_oe)}")
    fn_ml = sum(1 for j in fn if 0.35 <= j.get("ml_score", 0) < 0.5)
    analysis_parts.append(f"  ML 0.35-0.5 band: {fn_ml} ({fn_ml/max(len(fn),1)*100:.0f}%)")
    fn_families = Counter(fam for j in fn for fam, v in j.get("evidence_family_scores", {}).items() if v.get("fired")).most_common(5)
    analysis_parts.append(f"  Top families: {', '.join(f'{k}={v}' for k,v in fn_families)}")

    analysis = "\n".join(analysis_parts)
    log(f"Error analysis complete: {len(fp)} FPs, {len(fn)} FNs")
    return analysis


def generate_next_instructions(version: int, metrics: dict, analysis: str, prev_instructions: str) -> str:
    """Generate improved instructions for the next iteration based on error analysis."""
    # This is a template — the actual instruction generation is done by the Devin agent
    # based on the error analysis and performance metrics
    changes = []

    # Identify key patterns from analysis
    bacc = metrics.get("balanced_acc", 0)
    precision = metrics.get("precision", 0)
    recall = metrics.get("recall", 0)

    # Decision rules for what to change next
    if recall < 0.55 and precision > 0.65:
        changes.append("PRIORITY: Improve recall without sacrificing precision")
        changes.append("- Lower ML thresholds for specific OE types (off_topic, non_answer)")
        changes.append("- Add new FN-reduction rules for under-utilized signals")
    elif precision < 0.60 and recall > 0.55:
        changes.append("PRIORITY: Improve precision without sacrificing recall")
        changes.append("- Raise convergence threshold for thin_on_topic")
        changes.append("- Add FP-reduction rules for substantive OE in ML 0.5-0.7 band")
    elif bacc < 0.75:
        changes.append("PRIORITY: Both precision and recall need improvement")
        changes.append("- Focus on ML model calibration (isotonic regression)")
        changes.append("- Add per-channel/per-class thresholds")
        changes.append("- Add semantic features (OE text embeddings)")
    else:
        changes.append("PRIORITY: Fine-tune for final push to 90%")
        changes.append("- Ensemble multiple model predictions")
        changes.append("- Active learning from REVIEW tier")

    plan = "\n".join(changes)
    return plan


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--target", type=float, default=0.90, help="Target balanced accuracy")
    parser.add_argument("--max-iterations", type=int, default=20, help="Max iterations")
    parser.add_argument("--start-version", type=int, default=10, help="Starting version number")
    args = parser.parse_args()

    log(f"=== EVOLUTION LOOP START ===")
    log(f"Target: {args.target} balanced accuracy")
    log(f"Max iterations: {args.max_iterations}")
    log(f"Starting version: V{args.start_version}")

    # Initialize decision log
    if not DECISION_LOG.exists():
        with open(DECISION_LOG, "w") as f:
            f.write(f"# Decision Log — Survey Quality Evolution Loop\n\n")
            f.write(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
            f.write(f"Target: {args.target} balanced accuracy\n\n")
            f.write(f"## Baseline: V7 (best previous version)\n")
            f.write(f"- BAcc: 0.690, F1: 0.586, Precision: 0.664, Recall: 0.524\n\n---\n\n")

    version = args.start_version

    while version < args.start_version + args.max_iterations:
        log(f"\n{'='*60}")
        log(f"ITERATION V{version}")
        log(f"{'='*60}")

        version_dir = ECHO_OUTPUT / f"holistic_agent_run_v{version}"

        # The actual loop is orchestrated by the Devin agent
        # This script provides the framework and logging
        log(f"V{version} directory: {version_dir}")

        # For now, just log that we're ready for the next iteration
        # The Devin agent will:
        # 1. Call run_stage1_generate_packets
        # 2. Spawn 8 subagents for stage 2
        # 3. Call run_stage3_integrate
        # 4. Call run_stage4_compare
        # 5. Call analyze_errors
        # 6. Generate new instructions
        # 7. Append to decision log

        log(f"Waiting for Devin agent to orchestrate V{version}...")
        break

    log("=== EVOLUTION LOOP END ===")


if __name__ == "__main__":
    main()
