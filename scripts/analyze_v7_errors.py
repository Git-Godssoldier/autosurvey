#!/usr/bin/env python3
"""Deep analysis of V7 false negatives and false positives to find improvement opportunities."""
import json
import openpyxl
from pathlib import Path
from collections import Counter, defaultdict

def load_ground_truth(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Find ID and status columns
    id_col = None
    for h in ['uuid', 'UUID', 'RESP_ID', 'resp_id', 'ResponseId', 'RESPID', 'id', 'ID', 'V1']:
        if h in hidx:
            id_col = hidx[h]
            break
    if id_col is None:
        # Try first column
        id_col = 0

    status_col = None
    for h in ['status', 'Status', 'STATUS', 'V2', 'recordstatus']:
        if h in hidx:
            status_col = hidx[h]
            break

    gt = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[id_col]).strip() if row[id_col] else None
        if rid and status_col is not None:
            status = row[status_col]
            if status == 5 or status == '5':
                gt[rid] = 'DISCARD'
            elif status == 3 or status == '3':
                gt[rid] = 'KEEP'
    wb.close()
    return gt

def load_judgments(path):
    with open(path) as f:
        data = json.load(f)
    return {j['respondent_id']: j for j in data}

def main():
    gt_path = '/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx'
    v7_path = '/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run_v7/agent_judgments.json'

    gt = load_ground_truth(gt_path)
    v7 = load_judgments(v7_path)

    # Classify errors
    tp, fp, tn, fn = [], [], [], []
    soft_recall_missed = []  # FN that are also not in REVIEW

    for rid, judgment in v7.items():
        if rid not in gt:
            continue
        client = gt[rid]
        pred = judgment['agent_judgment']

        if pred == 'DISCARD' and client == 'DISCARD':
            tp.append(judgment)
        elif pred == 'DISCARD' and client == 'KEEP':
            fp.append(judgment)
        elif pred in ('REVIEW', 'KEEP') and client == 'DISCARD':
            fn.append(judgment)
            if pred == 'KEEP':
                soft_recall_missed.append(judgment)
        elif pred in ('REVIEW', 'KEEP') and client == 'KEEP':
            tn.append(judgment)

    print(f"TP={len(tp)}, FP={len(fp)}, TN={len(tn)}, FN={len(fn)}")
    print(f"Soft recall missed (FN in KEEP): {len(soft_recall_missed)}")
    print(f"Soft recall (DISCARD+REVIEW catching client discards): {len(tp) + (len(fn) - len(soft_recall_missed))}/{len(tp)+len(fn)} = {(len(tp) + (len(fn) - len(soft_recall_missed)))/(len(tp)+len(fn)):.3f}")

    # ===== FN ANALYSIS =====
    print(f"\n{'='*80}")
    print(f"FALSE NEGATIVES (missed client discards): {len(fn)}")
    print(f"{'='*80}")

    # By judgment
    fn_by_judgment = Counter(j['agent_judgment'] for j in fn)
    print(f"\nBy our judgment: {dict(fn_by_judgment)}")

    # By OE classification
    fn_by_oe = Counter(j['oe_classification'] for j in fn)
    print(f"\nBy OE classification:")
    for oe, c in fn_by_oe.most_common():
        print(f"  {oe}: {c} ({c/len(fn)*100:.1f}%)")

    # By ML score band
    fn_by_ml = Counter()
    for j in fn:
        ml = j['ml_score']
        if ml >= 0.8: fn_by_ml['0.8-1.0'] += 1
        elif ml >= 0.6: fn_by_ml['0.6-0.8'] += 1
        elif ml >= 0.5: fn_by_ml['0.5-0.6'] += 1
        elif ml >= 0.4: fn_by_ml['0.4-0.5'] += 1
        elif ml >= 0.2: fn_by_ml['0.2-0.4'] += 1
        else: fn_by_ml['0.0-0.2'] += 1
    print(f"\nBy ML score band:")
    for band in ['0.0-0.2', '0.2-0.4', '0.4-0.5', '0.5-0.6', '0.6-0.8', '0.8-1.0']:
        print(f"  {band}: {fn_by_ml.get(band, 0)}")

    # By converging family count
    fn_by_conv = Counter(j['converging_family_count'] for j in fn)
    print(f"\nBy converging family count:")
    for c in sorted(fn_by_conv.keys()):
        print(f"  {c}: {fn_by_conv[c]}")

    # By fired families
    fn_families = Counter()
    for j in fn:
        for fam, val in j['evidence_family_scores'].items():
            if val['fired']:
                fn_families[fam] += 1
    print(f"\nFired families in FN:")
    for fam, c in fn_families.most_common():
        print(f"  {fam}: {c} ({c/len(fn)*100:.1f}%)")

    # By primary removal reason
    fn_by_reason = Counter(j['primary_removal_reason'] for j in fn)
    print(f"\nBy primary removal reason: {dict(fn_by_reason)}")

    # By badopen trigger
    fn_by_badopen = Counter(j.get('badopen_trigger', 'none') for j in fn)
    print(f"\nBy badopen trigger:")
    for bt, c in fn_by_badopen.most_common():
        print(f"  {bt}: {c} ({c/len(fn)*100:.1f}%)")

    # FN with 0 converging families and ML < 0.4 — the hardest cases
    hard_fn = [j for j in fn if j['converging_family_count'] == 0 and j['ml_score'] < 0.4]
    print(f"\n--- Hard FN (0 families, ML < 0.4): {len(hard_fn)} ---")
    hard_oe = Counter(j['oe_classification'] for j in hard_fn)
    print(f"  OE: {dict(hard_oe)}")
    hard_badopen = Counter(j.get('badopen_trigger', 'none') for j in hard_fn)
    print(f"  Badopen: {dict(hard_badopen)}")
    hard_reason = Counter(j['primary_removal_reason'] for j in hard_fn)
    print(f"  Reason: {dict(hard_reason)}")

    # FN with ML >= 0.4 but not discarded — should have been caught
    missed_ml = [j for j in fn if j['ml_score'] >= 0.4]
    print(f"\n--- FN with ML >= 0.4 (should have been caught): {len(missed_ml)} ---")
    for j in missed_ml[:15]:
        fired = [k for k,v in j['evidence_family_scores'].items() if v['fired']]
        print(f"  {j['respondent_id']}: ML={j['ml_score']:.3f}, conv={j['converging_family_count']}, "
              f"OE={j['oe_classification']}, judgment={j['agent_judgment']}, "
              f"score={j['agent_score']:.2f}, fired={fired}, "
              f"badopen={j.get('badopen_trigger')}")
        print(f"    justification: {j['agent_justification'][:150]}")

    # FN with converging >= 2 but not discarded
    missed_conv = [j for j in fn if j['converging_family_count'] >= 2 and j['ml_score'] < 0.5]
    print(f"\n--- FN with conv >= 2, ML < 0.5 (convergence not enough): {len(missed_conv)} ---")
    for j in missed_conv[:15]:
        fired = [k for k,v in j['evidence_family_scores'].items() if v['fired']]
        print(f"  {j['respondent_id']}: ML={j['ml_score']:.3f}, conv={j['converging_family_count']}, "
              f"OE={j['oe_classification']}, judgment={j['agent_judgment']}, "
              f"score={j['agent_score']:.2f}, fired={fired}")

    # ===== FP ANALYSIS =====
    print(f"\n{'='*80}")
    print(f"FALSE POSITIVES (wrongly discarded client keeps): {len(fp)}")
    print(f"{'='*80}")

    # By ML band
    fp_by_ml = Counter()
    for j in fp:
        ml = j['ml_score']
        if ml >= 0.8: fp_by_ml['0.8-1.0'] += 1
        elif ml >= 0.6: fp_by_ml['0.6-0.8'] += 1
        elif ml >= 0.5: fp_by_ml['0.5-0.6'] += 1
        elif ml >= 0.4: fp_by_ml['0.4-0.5'] += 1
        elif ml >= 0.2: fp_by_ml['0.2-0.4'] += 1
        else: fp_by_ml['0.0-0.2'] += 1
    print(f"\nBy ML score band:")
    for band in ['0.0-0.2', '0.2-0.4', '0.4-0.5', '0.5-0.6', '0.6-0.8', '0.8-1.0']:
        print(f"  {band}: {fp_by_ml.get(band, 0)}")

    # By OE
    fp_by_oe = Counter(j['oe_classification'] for j in fp)
    print(f"\nBy OE classification: {dict(fp_by_oe)}")

    # By converging family count
    fp_by_conv = Counter(j['converging_family_count'] for j in fp)
    print(f"\nBy converging family count: {dict(sorted(fp_by_conv.items()))}")

    # By fired families
    fp_families = Counter()
    for j in fp:
        for fam, val in j['evidence_family_scores'].items():
            if val['fired']:
                fp_families[fam] += 1
    print(f"\nFired families in FP:")
    for fam, c in fp_families.most_common():
        print(f"  {fam}: {c} ({c/len(fp)*100:.1f}%)")

    # FP with ML >= 0.5 — driven by ML
    fp_ml_driven = [j for j in fp if j['ml_score'] >= 0.5]
    print(f"\n--- FP driven by ML >= 0.5: {len(fp_ml_driven)} ---")
    for j in fp_ml_driven[:15]:
        fired = [k for k,v in j['evidence_family_scores'].items() if v['fired']]
        print(f"  {j['respondent_id']}: ML={j['ml_score']:.3f}, conv={j['converging_family_count']}, "
              f"OE={j['oe_classification']}, score={j['agent_score']:.2f}, fired={fired}")
        print(f"    justification: {j['agent_justification'][:150]}")

    # FP with ML < 0.5 — driven by convergence
    fp_conv_driven = [j for j in fp if j['ml_score'] < 0.5]
    print(f"\n--- FP driven by convergence (ML < 0.5): {len(fp_conv_driven)} ---")
    for j in fp_conv_driven[:15]:
        fired = [k for k,v in j['evidence_family_scores'].items() if v['fired']]
        print(f"  {j['respondent_id']}: ML={j['ml_score']:.3f}, conv={j['converging_family_count']}, "
              f"OE={j['oe_classification']}, score={j['agent_score']:.2f}, fired={fired}")
        print(f"    justification: {j['agent_justification'][:150]}")

    # ===== CROSS-TAB: What features distinguish TP from FP? =====
    print(f"\n{'='*80}")
    print(f"TP vs FP COMPARISON — what distinguishes correct discards from wrong ones?")
    print(f"{'='*80}")

    # ML score comparison
    tp_ml = sum(j['ml_score'] for j in tp) / len(tp)
    fp_ml = sum(j['ml_score'] for j in fp) / len(fp)
    print(f"\nMean ML: TP={tp_ml:.3f}, FP={fp_ml:.3f}, gap={tp_ml-fp_ml:.3f}")

    # Converging families
    tp_conv = sum(j['converging_family_count'] for j in tp) / len(tp)
    fp_conv = sum(j['converging_family_count'] for j in fp) / len(fp)
    print(f"Mean converging families: TP={tp_conv:.2f}, FP={fp_conv:.2f}")

    # OE classification
    tp_oe = Counter(j['oe_classification'] for j in tp)
    fp_oe = Counter(j['oe_classification'] for j in fp)
    print(f"\nOE classification TP vs FP:")
    for oe in set(list(tp_oe.keys()) + list(fp_oe.keys())):
        tp_rate = tp_oe.get(oe, 0) / len(tp) * 100
        fp_rate = fp_oe.get(oe, 0) / len(fp) * 100
        print(f"  {oe}: TP={tp_rate:.1f}%, FP={fp_rate:.1f}%")

    # Family firing rates TP vs FP
    print(f"\nFamily firing rates TP vs FP:")
    for fam in ['model_risk','survey_structure','brand_funnel','source_risk','quota_reconstruction',
                'core_oe_quality','platform_risk','timing_engagement','duplicate_semantics']:
        tp_rate = sum(1 for j in tp if j['evidence_family_scores'][fam]['fired']) / len(tp)
        fp_rate = sum(1 for j in fp if j['evidence_family_scores'][fam]['fired']) / len(fp)
        print(f"  {fam}: TP={tp_rate:.2f}, FP={fp_rate:.2f}, gap={tp_rate-fp_rate:+.2f}")

    # Three-component scores
    tp_auth = sum(j['authenticity_risk'] for j in tp) / len(tp)
    fp_auth = sum(j['authenticity_risk'] for j in fp) / len(fp)
    tp_qual = sum(j['quality_discard_risk'] for j in tp) / len(tp)
    fp_qual = sum(j['quality_discard_risk'] for j in fp) / len(fp)
    tp_client = sum(j['client_reject_probability'] for j in tp) / len(tp)
    fp_client = sum(j['client_reject_probability'] for j in fp) / len(fp)
    print(f"\nThree-component scores:")
    print(f"  authenticity_risk: TP={tp_auth:.3f}, FP={fp_auth:.3f}, gap={tp_auth-fp_auth:+.3f}")
    print(f"  quality_discard_risk: TP={tp_qual:.3f}, FP={fp_qual:.3f}, gap={tp_qual-fp_qual:+.3f}")
    print(f"  client_reject_probability: TP={tp_client:.3f}, FP={fp_client:.3f}, gap={tp_client-fp_client:+.3f}")

    # ===== CROSS-TAB: FN vs TN — what distinguishes missed discards from correctly kept? =====
    print(f"\n{'='*80}")
    print(f"FN vs TN COMPARISON — what are we missing about client discards?")
    print(f"{'='*80}")

    fn_ml = sum(j['ml_score'] for j in fn) / len(fn)
    tn_ml = sum(j['ml_score'] for j in tn) / len(tn)
    print(f"\nMean ML: FN={fn_ml:.3f}, TN={tn_ml:.3f}, gap={fn_ml-tn_ml:+.3f}")

    fn_conv = sum(j['converging_family_count'] for j in fn) / len(fn)
    tn_conv = sum(j['converging_family_count'] for j in tn) / len(tn)
    print(f"Mean converging families: FN={fn_conv:.2f}, TN={tn_conv:.2f}")

    fn_oe = Counter(j['oe_classification'] for j in fn)
    tn_oe = Counter(j['oe_classification'] for j in tn)
    print(f"\nOE classification FN vs TN:")
    for oe in set(list(fn_oe.keys()) + list(tn_oe.keys())):
        fn_rate = fn_oe.get(oe, 0) / len(fn) * 100
        tn_rate = tn_oe.get(oe, 0) / len(tn) * 100
        print(f"  {oe}: FN={fn_rate:.1f}%, TN={tn_rate:.1f}%")

    print(f"\nFamily firing rates FN vs TN:")
    for fam in ['model_risk','survey_structure','brand_funnel','source_risk','quota_reconstruction',
                'core_oe_quality','platform_risk','timing_engagement','duplicate_semantics']:
        fn_rate = sum(1 for j in fn if j['evidence_family_scores'][fam]['fired']) / len(fn)
        tn_rate = sum(1 for j in tn if j['evidence_family_scores'][fam]['fired']) / len(tn)
        print(f"  {fam}: FN={fn_rate:.2f}, TN={tn_rate:.2f}, gap={fn_rate-tn_rate:+.2f}")

    fn_auth = sum(j['authenticity_risk'] for j in fn) / len(fn)
    tn_auth = sum(j['authenticity_risk'] for j in tn) / len(tn)
    fn_qual = sum(j['quality_discard_risk'] for j in fn) / len(fn)
    tn_qual = sum(j['quality_discard_risk'] for j in tn) / len(tn)
    fn_client = sum(j['client_reject_probability'] for j in fn) / len(fn)
    tn_client = sum(j['client_reject_probability'] for j in tn) / len(tn)
    print(f"\nThree-component scores:")
    print(f"  authenticity_risk: FN={fn_auth:.3f}, TN={tn_auth:.3f}, gap={fn_auth-tn_auth:+.3f}")
    print(f"  quality_discard_risk: FN={fn_qual:.3f}, TN={tn_qual:.3f}, gap={fn_qual-tn_qual:+.3f}")
    print(f"  client_reject_probability: FN={fn_client:.3f}, TN={tn_client:.3f}, gap={fn_client-tn_client:+.3f}")

    # Badopen
    fn_badopen = Counter(j.get('badopen_trigger', 'none') for j in fn)
    tn_badopen = Counter(j.get('badopen_trigger', 'none') for j in tn)
    print(f"\nBadopen trigger FN vs TN:")
    for bt in set(list(fn_badopen.keys()) + list(tn_badopen.keys())):
        fn_rate = fn_badopen.get(bt, 0) / len(fn) * 100
        tn_rate = tn_badopen.get(bt, 0) / len(tn) * 100
        print(f"  {bt}: FN={fn_rate:.1f}%, TN={tn_rate:.1f}%")

    # ===== LOAD RAW DATA FOR FN DEEP DIVE =====
    print(f"\n{'='*80}")
    print(f"RAW DATA ANALYSIS OF HARD FN (0 families, ML < 0.4)")
    print(f"{'='*80}")

    # Load raw Excel to see what the client might be seeing
    raw_path = '/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/109-2601 Echo BH.xlsx'
    wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Find ID column
    id_col = None
    for h in ['uuid', 'UUID', 'RESP_ID', 'resp_id', 'ResponseId', 'RESPID', 'id', 'ID', 'V1']:
        if h in hidx:
            id_col = hidx[h]
            break
    if id_col is None:
        id_col = 0

    # Find OE, timing, and other useful columns
    oe_cols = [h for h in headers if h and ('oe' in str(h).lower() or 'open' in str(h).lower() or 'narrative' in str(h).lower() or 'project' in str(h).lower())]
    timing_cols = [h for h in headers if h and ('time' in str(h).lower() or 'duration' in str(h).lower() or 'elapsed' in str(h).lower())]
    print(f"OE columns: {oe_cols[:5]}")
    print(f"Timing columns: {timing_cols[:5]}")

    # Get raw data for hard FN
    hard_fn_ids = {j['respondent_id'] for j in hard_fn}
    raw_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[id_col]).strip() if row[id_col] else None
        if rid in hard_fn_ids:
            raw_data[rid] = row
    wb.close()

    print(f"\nLoaded raw data for {len(raw_data)}/{len(hard_fn_ids)} hard FN")

    # Show OE text and timing for hard FN
    oe_col_idx = None
    for h in oe_cols[:1]:
        if h in hidx:
            oe_col_idx = hidx[h]
            break

    if oe_col_idx is not None:
        print(f"\nHard FN OE text samples (first 20):")
        for j in hard_fn[:20]:
            rid = j['respondent_id']
            raw = raw_data.get(rid)
            oe_text = str(raw[oe_col_idx])[:100] if raw and raw[oe_col_idx] else "N/A"
            print(f"  {rid}: ML={j['ml_score']:.3f}, OE={j['oe_classification']}, badopen={j.get('badopen_trigger')}")
            print(f"    OE: {oe_text}")

    # ===== ANALYZE WHAT THE CLIENT GROUND TRUTH FILE HAS =====
    print(f"\n{'='*80}")
    print(f"CLIENT ANNOTATED FILE — looking for client removal reasons")
    print(f"{'='*80}")

    wb2 = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)
    ws2 = wb2.active
    headers2 = [c.value for c in ws2[1]]
    hidx2 = {h: i for i, h in enumerate(headers2) if h}

    # Find columns that might contain client removal reasons
    reason_cols = [h for h in headers2 if h and any(k in str(h).lower() for k in
                   ['reason', 'remove', 'discard', 'reject', 'flag', 'note', 'comment', 'issue', 'problem'])]
    print(f"Potential reason columns: {reason_cols[:10]}")

    # Find ID column in GT file
    id_col2 = None
    for h in ['uuid', 'UUID', 'RESP_ID', 'resp_id', 'ResponseId', 'RESPID', 'id', 'ID', 'V1']:
        if h in hidx2:
            id_col2 = hidx2[h]
            break
    if id_col2 is None:
        id_col2 = 0

    # Get client reasons for FN
    fn_ids = {j['respondent_id'] for j in fn}
    client_reasons = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        rid = str(row[id_col2]).strip() if row[id_col2] else None
        if rid in fn_ids:
            reasons = {}
            for h in reason_cols[:5]:
                if h in hidx2:
                    val = row[hidx2[h]]
                    if val and str(val).strip():
                        reasons[h] = str(val)[:80]
            if reasons:
                client_reasons[rid] = reasons
    wb2.close()

    print(f"\nClient reasons found for {len(client_reasons)}/{len(fn_ids)} FN")

    # Aggregate client reasons
    reason_values = Counter()
    for rid, reasons in client_reasons.items():
        for h, val in reasons.items():
            reason_values[f"{h}={val}"] += 1

    print(f"\nTop client reason values (FN):")
    for rv, c in reason_values.most_common(30):
        print(f"  {rv}: {c}")

    # Also get client reasons for TP (to see what client says about correct discards)
    tp_ids = {j['respondent_id'] for j in tp}
    wb3 = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)
    ws3 = wb3.active
    client_reasons_tp = {}
    for row in ws3.iter_rows(min_row=2, values_only=True):
        rid = str(row[id_col2]).strip() if row[id_col2] else None
        if rid in tp_ids:
            reasons = {}
            for h in reason_cols[:5]:
                if h in hidx2:
                    val = row[hidx2[h]]
                    if val and str(val).strip():
                        reasons[h] = str(val)[:80]
            if reasons:
                client_reasons_tp[rid] = reasons
    wb3.close()

    reason_values_tp = Counter()
    for rid, reasons in client_reasons_tp.items():
        for h, val in reasons.items():
            reason_values_tp[f"{h}={val}"] += 1

    print(f"\nTop client reason values (TP):")
    for rv, c in reason_values_tp.most_common(30):
        print(f"  {rv}: {c}")

    # ===== ANALYZE ALL COLUMNS IN CLIENT FILE =====
    print(f"\n{'='*80}")
    print(f"ALL COLUMNS IN CLIENT ANNOTATED FILE")
    print(f"{'='*80}")
    wb4 = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)
    ws4 = wb4.active
    headers4 = [c.value for c in ws4[1]]
    for i, h in enumerate(headers4):
        if h:
            print(f"  [{i}] {h}")
    wb4.close()

if __name__ == '__main__':
    main()
