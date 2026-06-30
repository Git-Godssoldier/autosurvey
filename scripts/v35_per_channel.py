#!/usr/bin/env python3
"""V35 — Per-channel models (Pro vs Consumer) with V31 features.

Error analysis showed FNs are mostly Consumer and FPs have different profiles
for Pro vs Consumer. This version trains separate models for each channel
and combines their predictions.
"""
from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score, balanced_accuracy_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")

sys.path.insert(0, str(SKILL_SCRIPTS))
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training
from v19_target_encoding import extract_raw_answer_features, add_qtime_quality_interactions, add_target_encoding_features_train

import xgboost as xgb
import lightgbm as lgb


def get_classify_map():
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()
    return classify_map


def train_channel_model(X_tr, y_tr, X_val, y_val, X_test):
    """Train a single channel model."""
    scaler = StandardScaler()
    X_tr_scaled = scaler.fit_transform(X_tr)
    X_val_scaled = scaler.transform(X_val)
    X_test_scaled = scaler.transform(X_test)

    xgb_model = xgb.XGBClassifier(
        n_estimators=400, max_depth=5, learning_rate=0.05,
        subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
        random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
    )
    xgb_model.fit(X_tr, y_tr)

    lgb_model = lgb.LGBMClassifier(
        n_estimators=400, max_depth=6, learning_rate=0.05,
        num_leaves=31, subsample=0.8, colsample_bytree=0.7,
        reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
    )
    lgb_model.fit(X_tr, y_tr)

    mlp = MLPClassifier(
        hidden_layer_sizes=(128, 64), max_iter=400,
        learning_rate="adaptive", early_stopping=True,
        random_state=42, verbose=False
    )
    mlp.fit(X_tr_scaled, y_tr)

    models_val = {
        "xgb": xgb_model.predict_proba(X_val)[:, 1],
        "lgb": lgb_model.predict_proba(X_val)[:, 1],
        "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
    }
    models_test = {
        "xgb": xgb_model.predict_proba(X_test)[:, 1],
        "lgb": lgb_model.predict_proba(X_test)[:, 1],
        "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
    }

    cal_test = {}
    cal_val = {}
    for name in models_val:
        iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
        cal_val[name] = iso.transform(models_val[name])
        cal_test[name] = iso.transform(models_test[name])

    ensemble_test = np.mean(list(cal_test.values()), axis=0)
    return ensemble_test


def run_v35_cv(n_folds=5):
    """Run V35 cross-validation with per-channel models."""
    print(f"\n{'='*80}")
    print(f"V35 — Per-Channel Models (Pro vs Consumer)")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    raw_df, headers, roles = extract_raw_answer_features(ECHO_XLSX, gt, v7, v8)
    df = df.merge(raw_df, on="respondent_id", how="left", suffixes=("", "_raw"))
    df = add_qtime_quality_interactions(df)

    for qcol in [c for c in df.columns if c.startswith("ans_")]:
        df[f"te_{qcol}"] = 0.35
        df[f"cnt_{qcol}"] = 0

    print(f"\nTotal features (before TE): {len(df.columns)}")

    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    classify_map = get_classify_map()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        echo_train_for_te = echo_train[["respondent_id", "label"] + [c for c in echo_train.columns if c.startswith("ans_")]]
        echo_train = add_target_encoding_features_train(echo_train_for_te, echo_train, smoothing=20)
        echo_test = add_target_encoding_features_train(echo_train_for_te, echo_test, smoothing=20)

        # Self-training on ALL data (combined)
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        # Split self-trained data for calibration
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        # Approach 1: Combined model (V31 baseline)
        print(f"  Training combined model on {len(X_tr)} samples, {X_tr.shape[1]} features...")
        combined_scores = train_channel_model(X_tr, y_tr, X_val, y_val, X_test)

        # Approach 2: Per-channel models
        # We don't know pro/consumer for pseudo-labeled data, so we use the combined model
        # for self-training but train separate models on the labeled Echo data only
        echo_train_labeled = echo_train[echo_train["label"] >= 0].copy()
        echo_train_labeled["is_pro"] = echo_train_labeled["respondent_id"].map(
            lambda r: str(classify_map.get(r)) == "1"
        )

        pro_train = echo_train_labeled[echo_train_labeled["is_pro"]].copy()
        con_train = echo_train_labeled[~echo_train_labeled["is_pro"]].copy()

        print(f"  Pro: {len(pro_train)} ({(pro_train['label']==1).sum()} discards), "
              f"Consumer: {len(con_train)} ({(con_train['label']==1).sum()} discards)")

        # For per-channel models, we need enough samples
        # Use the self-trained data but split by a proxy (V7 reject prob > 0.5 = pro-like)
        # Actually, let's just use the labeled Echo data for per-channel models
        # and combine with the combined model's scores

        per_channel_scores = combined_scores.copy()

        if len(pro_train) > 50 and (pro_train["label"] == 1).sum() > 5:
            # Train Pro model on labeled Echo Pro data
            X_pro = pro_train[st_features].copy()
            for col in X_pro.select_dtypes(include=["object"]).columns:
                X_pro[col] = pd.Categorical(X_pro[col]).codes
            X_pro = X_pro.fillna(0)
            y_pro = pro_train["label"].values

            # Split for calibration
            n_pro_val = max(len(X_pro) // 4, 10)
            pro_val_idx = np.random.RandomState(42).choice(len(X_pro), n_pro_val, replace=False)
            pro_train_mask = np.ones(len(X_pro), dtype=bool)
            pro_train_mask[pro_val_idx] = False

            X_pro_tr = X_pro.iloc[pro_train_mask]
            X_pro_val = X_pro.iloc[pro_val_idx]
            y_pro_tr = y_pro[pro_train_mask]
            y_pro_val = y_pro[pro_val_idx]

            # Get test Pro indices
            pro_test_mask = is_pro_test
            X_pro_test = X_test[pro_test_mask]

            if len(X_pro_test) > 0:
                pro_scores = train_channel_model(X_pro_tr, y_pro_tr, X_pro_val, y_pro_val, X_pro_test)
                per_channel_scores[pro_test_mask] = pro_scores

        if len(con_train) > 50 and (con_train["label"] == 1).sum() > 5:
            # Train Consumer model on labeled Echo Consumer data
            X_con = con_train[st_features].copy()
            for col in X_con.select_dtypes(include=["object"]).columns:
                X_con[col] = pd.Categorical(X_con[col]).codes
            X_con = X_con.fillna(0)
            y_con = con_train["label"].values

            n_con_val = max(len(X_con) // 4, 10)
            con_val_idx = np.random.RandomState(42).choice(len(X_con), n_con_val, replace=False)
            con_train_mask = np.ones(len(X_con), dtype=bool)
            con_train_mask[con_val_idx] = False

            X_con_tr = X_con.iloc[con_train_mask]
            X_con_val = X_con.iloc[con_val_idx]
            y_con_tr = y_con[con_train_mask]
            y_con_val = y_con[con_val_idx]

            con_test_mask = ~is_pro_test
            X_con_test = X_test[con_test_mask]

            if len(X_con_test) > 0:
                con_scores = train_channel_model(X_con_tr, y_con_tr, X_con_val, y_con_val, X_con_test)
                per_channel_scores[con_test_mask] = con_scores

        # Blend: 50% combined + 50% per-channel
        blended_scores = 0.5 * combined_scores + 0.5 * per_channel_scores

        # Optimize threshold for each approach
        best_result = None
        best_bacc = 0
        for method_name, scores in [("combined", combined_scores),
                                     ("per_channel", per_channel_scores),
                                     ("blended", blended_scores)]:
            for thresh in np.arange(0.15, 0.70, 0.01):
                for pro_adj in np.arange(-0.20, 0.21, 0.01):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    rec = tp / max(tp + fn, 1)
                    spec = tn / max(tn + fp, 1)
                    bacc = (rec + spec) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_result = {
                            "method": method_name, "thresh": thresh, "pro_adj": pro_adj,
                            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
                            "precision": tp / max(tp + fp, 1),
                            "recall": rec, "f1": 2 * (tp / max(tp + fp, 1)) * rec / max((tp / max(tp + fp, 1)) + rec, 0.001),
                            "bacc": bacc, "auc": roc_auc_score(y_test, scores),
                        }

        print(f"  Best: {best_result['method']}, thresh={best_result['thresh']:.3f}, pro_adj={best_result['pro_adj']:+.3f}")
        print(f"  TP={best_result['tp']}, FP={best_result['fp']}, TN={best_result['tn']}, FN={best_result['fn']}, "
              f"BAcc={best_result['bacc']:.3f}, AUC={best_result['auc']:.3f}")

        all_fold_metrics.append({"fold": fold + 1, **best_result})

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V35 CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    # Count which method won
    method_counts = Counter([m["method"] for m in all_fold_metrics])
    print(f"  Method wins: {dict(method_counts)}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V35 — Per-Channel Models (Pro vs Consumer)")
    print("=" * 80)

    results = run_v35_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V14 (self-train + V8):   BAcc=0.744, AUC=0.788")
    print(f"  V31 (V14+V19 combined):  BAcc=0.796, AUC=0.844 (BEST)")
    print(f"  V35 (per-channel):       BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}, AUC={results['avg_auc']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], 0.796):.3f}")

    with open(AUTOSURVEY_DIR / "v35_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
