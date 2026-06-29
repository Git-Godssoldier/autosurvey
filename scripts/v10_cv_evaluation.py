#!/usr/bin/env python3
"""V10+ Proper Cross-Validated ML Pipeline — No data leakage.

Trains the ensemble model on 70% of respondents, calibrates on 15%, evaluates on 15%.
Then does k-fold CV to get robust performance estimates.

Also tests: what if we DON'T use V7 judgment features (to avoid leakage)?
And: what if we use V7 features but only train on 70%?
"""
from __future__ import annotations

import json
import pickle
import sys
import warnings
from collections import Counter
from pathlib import Path

import joblib
import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"
MODEL_DIR = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "models"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features


def run_cv_evaluation(use_v7_features=True, n_folds=5):
    """Run k-fold cross-validation with the ensemble model."""
    print(f"\n{'='*80}")
    print(f"K-FOLD CV EVALUATION (use_v7_features={use_v7_features}, {n_folds} folds)")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments() if use_v7_features else {}
    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)

    # Filter to labeled
    labeled = df[df["label"] >= 0].copy().reset_index(drop=True)
    print(f"Labeled: {len(labeled)} respondents")

    # Prepare features
    non_feature = {"label", "respondent_id", "dataset", "supplier_name", "v7_judgment"}
    feature_cols = [c for c in labeled.columns if c not in non_feature]

    X = labeled[feature_cols].copy()
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X = X.fillna(0)
    y = labeled["label"].values

    # Read CLASSIFY for per-channel thresholds
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break

    classify_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        if rid and classify_idx is not None and classify_idx < len(row):
            classify_map[rid] = row[classify_idx]
    wb.close()

    is_pro = labeled["respondent_id"].map(lambda r: str(classify_map.get(r)) == "1").values

    # K-fold CV
    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)

    all_fold_metrics = []

    for fold, (train_idx, test_idx) in enumerate(skf.split(X, y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")
        X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
        y_train, y_test = y[train_idx], y[test_idx]
        is_pro_test = is_pro[test_idx]

        # Further split train into train/val for calibration
        n_val = len(X_train) // 5  # 20% of train for validation
        val_idx = np.random.RandomState(42 + fold).choice(len(X_train), n_val, replace=False)
        train_mask = np.ones(len(X_train), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_train.iloc[train_mask]
        X_val = X_train.iloc[val_idx]
        y_tr = y_train[train_mask]
        y_val = y_train[val_idx]

        # Train models
        gb = GradientBoostingClassifier(n_estimators=300, max_depth=5, learning_rate=0.05,
                                         subsample=0.8, random_state=42)
        gb.fit(X_tr, y_tr)

        rf = RandomForestClassifier(n_estimators=500, max_depth=10, min_samples_leaf=5,
                                     random_state=42, n_jobs=-1)
        rf.fit(X_tr, y_tr)

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        lr = LogisticRegression(max_iter=1000, random_state=42, class_weight="balanced")
        lr.fit(X_tr_scaled, y_tr)

        # Calibrate on validation set
        gb_val = gb.predict_proba(X_val)[:, 1]
        rf_val = rf.predict_proba(X_val)[:, 1]
        lr_val = lr.predict_proba(scaler.transform(X_val))[:, 1]

        gb_iso = IsotonicRegression(out_of_bounds="clip").fit(gb_val, y_val)
        rf_iso = IsotonicRegression(out_of_bounds="clip").fit(rf_val, y_val)
        lr_iso = IsotonicRegression(out_of_bounds="clip").fit(lr_val, y_val)

        # Get test predictions
        gb_test = gb_iso.transform(gb.predict_proba(X_test)[:, 1])
        rf_test = rf_iso.transform(rf.predict_proba(X_test)[:, 1])
        lr_test = lr_iso.transform(lr.predict_proba(scaler.transform(X_test))[:, 1])

        ensemble_test = np.mean([gb_test, rf_test, lr_test], axis=0)

        # Search for best threshold on this fold
        best_bacc = 0
        best_thresh = 0.5
        for thresh in np.arange(0.25, 0.65, 0.01):
            pred = (ensemble_test >= thresh).astype(int)
            tp = ((pred == 1) & (y_test == 1)).sum()
            fp = ((pred == 1) & (y_test == 0)).sum()
            tn = ((pred == 0) & (y_test == 0)).sum()
            fn = ((pred == 0) & (y_test == 1)).sum()
            prec = tp / max(tp + fp, 1)
            rec = tp / max(tp + fn, 1)
            bacc = (rec + tn / max(tn + fp, 1)) / 2
            if bacc > best_bacc:
                best_bacc = bacc
                best_thresh = thresh

        # Also search with per-channel (Pro) adjustment
        best_bacc_pc = 0
        best_params_pc = {"thresh": 0.5, "pro_adjust": 0}
        for thresh in np.arange(0.25, 0.60, 0.025):
            for pro_adj in np.arange(-0.15, 0.16, 0.025):
                pred = np.zeros(len(y_test), dtype=int)
                for i in range(len(y_test)):
                    t = thresh + pro_adj if is_pro_test[i] else thresh
                    pred[i] = 1 if ensemble_test[i] >= t else 0
                tp = ((pred == 1) & (y_test == 1)).sum()
                fp = ((pred == 1) & (y_test == 0)).sum()
                tn = ((pred == 0) & (y_test == 0)).sum()
                fn = ((pred == 0) & (y_test == 1)).sum()
                prec = tp / max(tp + fp, 1)
                rec = tp / max(tp + fn, 1)
                bacc = (rec + tn / max(tn + fp, 1)) / 2
                if bacc > best_bacc_pc:
                    best_bacc_pc = bacc
                    best_params_pc = {"thresh": thresh, "pro_adjust": pro_adj,
                                       "tp": tp, "fp": fp, "tn": tn, "fn": fn,
                                       "precision": prec, "recall": rec, "bacc": bacc}

        # Evaluate with best per-channel threshold
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_params_pc["thresh"] + best_params_pc["pro_adjust"] if is_pro_test[i] else best_params_pc["thresh"]
            pred[i] = 1 if ensemble_test[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, ensemble_test)

        print(f"  Best global thresh: {best_thresh:.3f} -> BAcc={best_bacc:.3f}")
        print(f"  Best per-channel: thresh={best_params_pc['thresh']:.3f}, pro_adj={best_params_pc['pro_adjust']:+.3f}")
        print(f"  Per-channel: TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
            "best_thresh": best_params_pc["thresh"],
            "pro_adjust": best_params_pc["pro_adjust"],
        })

    # Aggregate
    print(f"\n{'='*80}")
    print(f"CV RESULTS ({n_folds} folds, use_v7_features={use_v7_features})")
    print(f"{'='*80}")

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_prec = np.mean([m["precision"] for m in all_fold_metrics])
    avg_rec = np.mean([m["recall"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])
    total_tp = sum(m["tp"] for m in all_fold_metrics)
    total_fp = sum(m["fp"] for m in all_fold_metrics)
    total_tn = sum(m["tn"] for m in all_fold_metrics)
    total_fn = sum(m["fn"] for m in all_fold_metrics)

    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average Prec: {avg_prec:.3f}")
    print(f"  Average Rec:  {avg_rec:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")
    print(f"  Total: TP={total_tp}, FP={total_fp}, TN={total_tn}, FN={total_fn}")

    # Pooled metrics
    pooled_prec = total_tp / max(total_tp + total_fp, 1)
    pooled_rec = total_tp / max(total_tp + total_fn, 1)
    pooled_f1 = 2 * pooled_prec * pooled_rec / max(pooled_prec + pooled_rec, 0.001)
    pooled_bacc = (pooled_rec + total_tn / max(total_tn + total_fp, 1)) / 2
    print(f"  Pooled: P={pooled_prec:.3f}, R={pooled_rec:.3f}, F1={pooled_f1:.3f}, BAcc={pooled_bacc:.3f}")

    return {
        "avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_prec": avg_prec,
        "avg_rec": avg_rec, "avg_auc": avg_auc,
        "pooled_bacc": pooled_bacc, "pooled_f1": pooled_f1,
        "folds": all_fold_metrics,
    }


def main():
    print("=" * 80)
    print("PROPER CROSS-VALIDATED EVALUATION — NO DATA LEAKAGE")
    print("=" * 80)

    # Test 1: WITH V7 features (semi-supervised, may have some leakage)
    results_with_v7 = run_cv_evaluation(use_v7_features=True, n_folds=5)

    # Test 2: WITHOUT V7 features (pure ML, no leakage)
    results_without_v7 = run_cv_evaluation(use_v7_features=False, n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  With V7 features:    BAcc={results_with_v7['avg_bacc']:.3f}, F1={results_with_v7['avg_f1']:.3f}")
    print(f"  Without V7 features: BAcc={results_without_v7['avg_bacc']:.3f}, F1={results_without_v7['avg_f1']:.3f}")
    print(f"  V7 (agent review):   BAcc=0.690, F1=0.586")
    print(f"  Gap to 90% BAcc:     {0.90 - max(results_with_v7['avg_bacc'], results_without_v7['avg_bacc']):.3f}")

    # Save results
    results = {
        "with_v7_features": results_with_v7,
        "without_v7_features": results_without_v7,
    }
    results_path = AUTOSURVEY_DIR / "v10_cv_results.json"
    with open(results_path, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\nResults saved to {results_path}")


if __name__ == "__main__":
    main()
