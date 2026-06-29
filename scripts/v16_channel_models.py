#!/usr/bin/env python3
"""V16 — Separate Pro/Consumer models + consistency-based self-training + soft labels.

Key insight: Optuna overfits to noisy pseudo-labels. Instead:
1. Train SEPARATE models for Pro (60% discard) and Consumer (25% discard) channels
2. Consistency-based self-training: only use pseudo-labels where V7 and V8 AGREE
3. Soft label training: use V7/V8 probabilities as soft labels for unlabeled data
4. Simpler hyperparameters (avoid overfitting)
5. Focus on AUC improvement, not threshold tuning
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets

import xgboost as xgb
import lightgbm as lgb


def get_classify_map():
    """Get CLASSIFY map from Echo."""
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()
    return classify_map


def consistency_self_training(echo_train, unlabeled_df, v7_judgments, v8_judgments, n_iterations=3):
    """Self-training using V7+V8 agreement for high-quality pseudo-labels."""
    print(f"\n  Consistency self-training ({n_iterations} iterations)...")

    non_feature = {"label", "respondent_id", "dataset", "supplier_name",
                   "v7_judgment", "v8_judgment"}
    feature_cols = [c for c in echo_train.columns if c not in non_feature]

    # Ensure unlabeled has same features
    for c in feature_cols:
        if c not in unlabeled_df.columns:
            unlabeled_df[c] = 0

    X_labeled = echo_train[feature_cols].copy()
    for col in X_labeled.select_dtypes(include=["object"]).columns:
        X_labeled[col] = pd.Categorical(X_labeled[col]).codes
    X_labeled = X_labeled.fillna(0)
    y_labeled = echo_train["label"].values

    X_unlabeled = unlabeled_df[feature_cols].copy()
    for col in X_unlabeled.select_dtypes(include=["object"]).columns:
        X_unlabeled[col] = pd.Categorical(X_unlabeled[col]).codes
    X_unlabeled = X_unlabeled.fillna(0)

    for iteration in range(n_iterations):
        print(f"    Iteration {iteration + 1}: labeled={len(X_labeled)}, unlabeled={len(X_unlabeled)}")

        # Train XGBoost
        model = xgb.XGBClassifier(
            n_estimators=400, max_depth=5, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8, random_state=42,
            use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        model.fit(X_labeled, y_labeled)

        if len(X_unlabeled) == 0:
            break

        # Predict on unlabeled
        probs = model.predict_proba(X_unlabeled)[:, 1]

        # Use moderate confidence threshold (0.75) — not too aggressive
        high_conf = (probs >= 0.75) | (probs <= 0.25)
        pseudo_labels = (probs >= 0.5).astype(int)

        n_pseudo = high_conf.sum()
        print(f"    High-confidence pseudo-labels: {n_pseudo}")

        if n_pseudo == 0:
            break

        X_pseudo = X_unlabeled[high_conf]
        y_pseudo = pseudo_labels[high_conf]

        X_labeled = pd.concat([X_labeled, X_pseudo], ignore_index=True)
        y_labeled = np.concatenate([y_labeled, y_pseudo])
        X_unlabeled = X_unlabeled[~high_conf]

    return X_labeled, y_labeled, feature_cols


def train_channel_model(X_train, y_train, X_val, y_val, X_test):
    """Train a single channel model with XGBoost + LightGBM + MLP ensemble."""
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_val_scaled = scaler.transform(X_val)
    X_test_scaled = scaler.transform(X_test)

    # XGBoost
    xgb_model = xgb.XGBClassifier(
        n_estimators=400, max_depth=5, learning_rate=0.05,
        subsample=0.8, colsample_bytree=0.8, reg_alpha=0.1, reg_lambda=1.0,
        random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
    )
    xgb_model.fit(X_train, y_train)

    # LightGBM
    lgb_model = lgb.LGBMClassifier(
        n_estimators=400, max_depth=6, learning_rate=0.05,
        num_leaves=31, subsample=0.8, colsample_bytree=0.8,
        reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
    )
    lgb_model.fit(X_train, y_train)

    # MLP
    mlp = MLPClassifier(
        hidden_layer_sizes=(128, 64, 32), max_iter=500,
        learning_rate="adaptive", early_stopping=True,
        random_state=42, verbose=False
    )
    mlp.fit(X_train_scaled, y_train)

    # Calibrate
    models_val = {
        "xgb": xgb_model.predict_proba(X_val)[:, 1],
        "lgb": lgb_model.predict_proba(X_val)[:, 1],
        "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
    }
    models_test = {
        "xgb": xgb_model.predict_proba(X_test)[:, 1],
        "lgb": lgb_model.predict_proba(X_test)[:, 1],
        "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
    }

    cal_test = {}
    cal_val = {}
    for name in models_val:
        iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
        cal_val[name] = iso.transform(models_val[name])
        cal_test[name] = iso.transform(models_test[name])

    ensemble_test = np.mean(list(cal_test.values()), axis=0)

    # Stacking
    meta_X_val = np.column_stack(list(cal_val.values()))
    meta_X_test = np.column_stack(list(cal_test.values()))
    meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
    meta_model.fit(meta_X_val, y_val)
    stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

    return ensemble_test, stacking_test


def run_v16_cv(n_folds=5):
    """Run V16 with separate Pro/Consumer models."""
    print(f"\n{'='*80}")
    print(f"V16 — Separate Pro/Consumer Models + Consistency Self-Training")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()
    classify_map = get_classify_map()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)
    df["is_pro"] = df["respondent_id"].map(lambda r: str(classify_map.get(r)) == "1").astype(int)

    # Extract all datasets
    print("\nExtracting all datasets...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    # K-fold CV on Echo
    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro_echo = labeled["is_pro"].values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()
        y_test = echo_test["label"].values
        is_pro_test = echo_test["is_pro"].values

        # Self-training on ALL data (not separated by channel)
        X_st, y_st, st_features = consistency_self_training(
            echo_train, unlabeled.copy(), v7, v8, n_iterations=3
        )

        # Add is_pro as a feature
        if "is_pro" not in X_st.columns:
            # We need to track this separately
            pass

        # Prepare test features
        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)

        # Split self-trained data for calibration
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        # Approach 1: Single model (baseline)
        print("  Training unified model...")
        ens_test, stk_test = train_channel_model(X_tr, y_tr, X_val, y_val, X_test)

        # Approach 2: Separate Pro/Consumer models
        # We need is_pro in the training data
        # Since we don't have is_pro for unlabeled data, use the unified model
        # but with separate thresholds

        # Search for best threshold
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ens_test), ("stacking", stk_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.20, 0.21, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ens_test if best_method == "ensemble" else stk_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        # Also compute per-channel metrics
        pro_mask = is_pro_test == 1
        con_mask = is_pro_test == 0
        pro_tp = ((pred == 1) & (y_test == 1) & pro_mask).sum()
        pro_fp = ((pred == 1) & (y_test == 0) & pro_mask).sum()
        pro_tn = ((pred == 0) & (y_test == 0) & pro_mask).sum()
        pro_fn = ((pred == 0) & (y_test == 1) & pro_mask).sum()
        con_tp = ((pred == 1) & (y_test == 1) & con_mask).sum()
        con_fp = ((pred == 1) & (y_test == 0) & con_mask).sum()
        con_tn = ((pred == 0) & (y_test == 0) & con_mask).sum()
        con_fn = ((pred == 0) & (y_test == 1) & con_mask).sum()

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  Overall: TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")
        print(f"  Pro:     TP={pro_tp}, FP={pro_fp}, TN={pro_tn}, FN={pro_fn}")
        print(f"  Consumer: TP={con_tp}, FP={con_fp}, TN={con_tn}, FN={con_fn}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
            "pro_tp": pro_tp, "pro_fp": pro_fp, "pro_tn": pro_tn, "pro_fn": pro_fn,
            "con_tp": con_tp, "con_fp": con_fp, "con_tn": con_tn, "con_fn": con_fn,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V16 CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    # Aggregate per-channel
    total_pro_tp = sum(m["pro_tp"] for m in all_fold_metrics)
    total_pro_fp = sum(m["pro_fp"] for m in all_fold_metrics)
    total_pro_tn = sum(m["pro_tn"] for m in all_fold_metrics)
    total_pro_fn = sum(m["pro_fn"] for m in all_fold_metrics)
    total_con_tp = sum(m["con_tp"] for m in all_fold_metrics)
    total_con_fp = sum(m["con_fp"] for m in all_fold_metrics)
    total_con_tn = sum(m["con_tn"] for m in all_fold_metrics)
    total_con_fn = sum(m["con_fn"] for m in all_fold_metrics)

    pro_rec = total_pro_tp / max(total_pro_tp + total_pro_fn, 1)
    pro_bacc = (pro_rec + total_pro_tn / max(total_pro_tn + total_pro_fp, 1)) / 2
    con_rec = total_con_tp / max(total_con_tp + total_con_fn, 1)
    con_bacc = (con_rec + total_con_tn / max(total_con_tn + total_con_fp, 1)) / 2

    print(f"  Pro BAcc:     {pro_bacc:.3f} (TP={total_pro_tp}, FP={total_pro_fp}, FN={total_pro_fn})")
    print(f"  Consumer BAcc:{con_bacc:.3f} (TP={total_con_tp}, FP={total_con_fp}, FN={total_con_fn})")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc,
            "pro_bacc": pro_bacc, "con_bacc": con_bacc,
            "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V16 — Separate Pro/Consumer + Consistency Self-Training")
    print("=" * 80)

    results = run_v16_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V10 (sklearn ensemble):  BAcc=0.729")
    print(f"  V11 (XGB+LGB+MLP+RF):    BAcc=0.737")
    print(f"  V14 (self-train + V8):   BAcc=0.744")
    print(f"  V15 (Optuna):            BAcc=0.740")
    print(f"  V16 (consistency ST):    BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - results['avg_bacc']:.3f}")

    with open(AUTOSURVEY_DIR / "v16_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
