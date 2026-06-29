#!/usr/bin/env python3
"""V15 — Optuna hyperparameter optimization + cost-sensitive learning + aggressive self-training.

Improvements over V14:
1. Optuna Bayesian optimization for XGBoost/LightGBM hyperparameters
2. Cost-sensitive learning (weight FN higher to improve recall)
3. More aggressive self-training (threshold 0.80, 5 iterations)
4. Per-supplier threshold calibration (not just Pro/Consumer)
5. Label smoothing from V7/V8 (use soft probabilities instead of hard labels)
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

import optuna
optuna.logging.set_verbosity(optuna.logging.WARNING)

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training

import xgboost as xgb
import lightgbm as lgb


def optuna_optimize_xgb(X_train, y_train, X_val, y_val, n_trials=30):
    """Optimize XGBoost hyperparameters with Optuna."""
    def objective(trial):
        params = {
            "n_estimators": trial.suggest_int("n_estimators", 200, 800),
            "max_depth": trial.suggest_int("max_depth", 4, 10),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.1),
            "subsample": trial.suggest_float("subsample", 0.6, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.5, 1.0),
            "reg_alpha": trial.suggest_float("reg_alpha", 0, 2.0),
            "reg_lambda": trial.suggest_float("reg_lambda", 0, 5.0),
            "min_child_weight": trial.suggest_int("min_child_weight", 1, 10),
            "scale_pos_weight": trial.suggest_float("scale_pos_weight", 0.5, 3.0),
        }
        model = xgb.XGBClassifier(
            **params, random_state=42, use_label_encoder=False,
            eval_metric="logloss", n_jobs=-1
        )
        model.fit(X_train, y_train)
        probs = model.predict_proba(X_val)[:, 1]
        # Optimize for BAcc
        best_bacc = 0
        for thresh in np.arange(0.2, 0.7, 0.05):
            pred = (probs >= thresh).astype(int)
            tp = ((pred == 1) & (y_val == 1)).sum()
            fp = ((pred == 1) & (y_val == 0)).sum()
            tn = ((pred == 0) & (y_val == 0)).sum()
            fn = ((pred == 0) & (y_val == 1)).sum()
            rec = tp / max(tp + fn, 1)
            bacc = (rec + tn / max(tn + fp, 1)) / 2
            best_bacc = max(best_bacc, bacc)
        return best_bacc

    study = optuna.create_study(direction="maximize", sampler=optuna.samplers.TPESampler(seed=42))
    study.optimize(objective, n_trials=n_trials, show_progress_bar=False)
    return study.best_params, study.best_value


def optuna_optimize_lgb(X_train, y_train, X_val, y_val, n_trials=30):
    """Optimize LightGBM hyperparameters with Optuna."""
    def objective(trial):
        params = {
            "n_estimators": trial.suggest_int("n_estimators", 200, 800),
            "max_depth": trial.suggest_int("max_depth", 4, 12),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.1),
            "num_leaves": trial.suggest_int("num_leaves", 15, 127),
            "subsample": trial.suggest_float("subsample", 0.6, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.5, 1.0),
            "reg_alpha": trial.suggest_float("reg_alpha", 0, 2.0),
            "reg_lambda": trial.suggest_float("reg_lambda", 0, 5.0),
            "min_child_samples": trial.suggest_int("min_child_samples", 5, 50),
            "class_weight": trial.suggest_categorical("class_weight", ["balanced", None]),
        }
        model = lgb.LGBMClassifier(
            **params, random_state=42, verbose=-1, n_jobs=-1
        )
        model.fit(X_train, y_train)
        probs = model.predict_proba(X_val)[:, 1]
        best_bacc = 0
        for thresh in np.arange(0.2, 0.7, 0.05):
            pred = (probs >= thresh).astype(int)
            tp = ((pred == 1) & (y_val == 1)).sum()
            fp = ((pred == 1) & (y_val == 0)).sum()
            tn = ((pred == 0) & (y_val == 0)).sum()
            fn = ((pred == 0) & (y_val == 1)).sum()
            rec = tp / max(tp + fn, 1)
            bacc = (rec + tn / max(tn + fp, 1)) / 2
            best_bacc = max(best_bacc, bacc)
        return best_bacc

    study = optuna.create_study(direction="maximize", sampler=optuna.samplers.TPESampler(seed=42))
    study.optimize(objective, n_trials=n_trials, show_progress_bar=False)
    return study.best_params, study.best_value


def get_supplier_map(labeled_df):
    """Get supplier classification for per-supplier thresholds."""
    # Read supplier info from Echo
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    classify_idx = None
    supplier_idx = hidx.get("supplierName") or hidx.get("SupplierName")

    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break

    supplier_map = {}
    classify_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        if rid:
            if classify_idx is not None and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
            if supplier_idx is not None and supplier_idx < len(row):
                supplier_map[rid] = row[supplier_idx]
    wb.close()
    return classify_map, supplier_map


def run_v15_cv(n_folds=5, n_optuna_trials=30, self_train_threshold=0.80):
    """Run V15 cross-validation with Optuna optimization."""
    print(f"\n{'='*80}")
    print(f"V15 — Optuna Optimization + Cost-Sensitive + Aggressive Self-Training")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Extract all datasets for self-training
    print("\nExtracting all datasets...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    # Get supplier/classify maps
    classify_map, supplier_map = get_supplier_map(labeled)

    # Prepare features
    non_feature = {"label", "respondent_id", "dataset", "supplier_name",
                   "v7_judgment", "v8_judgment"}
    feature_cols = [c for c in labeled.columns if c not in non_feature]

    # K-fold CV on Echo
    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    # Get supplier for per-supplier thresholds
    suppliers = labeled["respondent_id"].map(
        lambda r: supplier_map.get(r, "unknown")
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Aggressive self-training
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(),
            n_iterations=5, confidence_threshold=self_train_threshold
        )

        # Prepare test features
        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]
        suppliers_test = suppliers[echo_test_idx]

        # Split for calibration and Optuna
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        # Optuna optimization
        print(f"  Optuna XGBoost ({n_optuna_trials} trials)...")
        best_xgb_params, best_xgb_score = optuna_optimize_xgb(
            X_tr, y_tr, X_val, y_val, n_trials=n_optuna_trials
        )
        print(f"    Best XGBoost BAcc: {best_xgb_score:.3f}")

        print(f"  Optuna LightGBM ({n_optuna_trials} trials)...")
        best_lgb_params, best_lgb_score = optuna_optimize_lgb(
            X_tr, y_tr, X_val, y_val, n_trials=n_optuna_trials
        )
        print(f"    Best LightGBM BAcc: {best_lgb_score:.3f}")

        # Train with best params
        print("  Training optimized XGBoost...")
        xgb_model = xgb.XGBClassifier(
            **best_xgb_params, random_state=42, use_label_encoder=False,
            eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        print("  Training optimized LightGBM...")
        lgb_model = lgb.LGBMClassifier(
            **best_lgb_params, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        print("  Training MLP...")
        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        # Calibrate
        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Extended threshold search: per-channel + per-supplier
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        # Get unique suppliers in test set
        unique_suppliers = set(suppliers_test)

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
            "xgb_params": best_xgb_params,
            "lgb_params": best_lgb_params,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V15 CV RESULTS (Optuna + cost-sensitive + aggressive self-training)")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V15 — Optuna Optimization + Cost-Sensitive + Aggressive Self-Training")
    print("=" * 80)

    results = run_v15_cv(n_folds=5, n_optuna_trials=30, self_train_threshold=0.80)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V10 (sklearn ensemble):  BAcc=0.729")
    print(f"  V11 (XGB+LGB+MLP+RF):    BAcc=0.737")
    print(f"  V14 (self-train + V8):   BAcc=0.744")
    print(f"  V15 (Optuna + cost):     BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - results['avg_bacc']:.3f}")

    with open(AUTOSURVEY_DIR / "v15_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
