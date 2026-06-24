#!/usr/bin/env python3
"""Run the Delta transfer benchmark with sealed incumbent and challenger ledgers."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
import statistics
import subprocess
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import openpyxl


BLIND = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/106-2502 Delta Water Filtration.xlsx")
ANNOTATED = Path("/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260111_Delta Water Filtration.xlsx")
RUN_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/agent-native-delta-transfer-2026-06-24")
INTERNAL = RUN_DIR / ".autosurvey-internal"
PUBLIC = RUN_DIR / "public"
ECHO_INTERNAL = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/agent-native-echo-2026-06-24/.autosurvey-internal")
CLIENT_STATUS_THRESHOLD = 0.35
INCUMBENT_VERSION = "incumbent-post-echo-v2026-06-24"
CHALLENGER_VERSION = "challenger-field-first-v2026-06-24"
MODEL = ""

EXCLUDED_FIELDS = {
    "qc",
    "TERMFLAGS",
    "qc5",
    "qc5_Pasted",
    "LangAssessReadLevel",
    "LangAssessReadEase",
    "LangAssessNumSen",
    "LangAssessNumWords",
    "LangAssessNumSyl",
    "url",
    "session",
    "camp",
    "bhf",
    "sfh",
    "intcode",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions")
ACTION_TIERS = ["Clean keep", "Keep with note", "Light review", "Review closely", "Exclude candidate"]


def decision_schema(kind: str) -> dict[str, Any]:
    common = {
        "respondent_id": {"type": "string"},
        "source_excel_row": {"type": "integer"},
        "authenticity_risk": {"type": "string", "enum": ["low", "moderate", "high", "severe"]},
        "client_rejection_probability": {"type": "number", "minimum": 0, "maximum": 1},
        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
        "semantic_decision_criteria": {"type": "string"},
        "signals": {"type": "array", "items": {"type": "string"}, "minItems": 1},
        "signal_evidence": {"type": "array", "items": {"type": "string"}, "minItems": 1},
        "protective_evidence": {"type": "string"},
        "key_relationships": {"type": "string"},
        "adjudication": {"type": "string"},
        "reviewer_id": {"type": "string"},
    }
    if kind == "incumbent":
        required = [
            "respondent_id",
            "source_excel_row",
            "operational_tier",
            "operational_discard",
            "authenticity_risk",
            "predicted_client_status",
            "client_rejection_probability",
            "confidence",
            "semantic_decision_criteria",
            "signals",
            "signal_evidence",
            "protective_evidence",
            "key_relationships",
            "adjudication",
            "reviewer_id",
        ]
        props = {
            **common,
            "operational_tier": {"type": "string", "enum": ACTION_TIERS},
            "operational_discard": {"type": "boolean"},
            "predicted_client_status": {"type": "integer", "enum": [3, 5]},
        }
    else:
        required = [
            "respondent_id",
            "source_excel_row",
            "predicted_client_status",
            "client_rejection_probability",
            "authenticity_assessment",
            "authenticity_risk",
            "operational_tier",
            "operational_discard",
            "field_invalidity_ledger",
            "hard_invalidity_count",
            "soft_concern_count",
            "confidence",
            "semantic_decision_criteria",
            "signals",
            "signal_evidence",
            "protective_evidence",
            "key_relationships",
            "adjudication",
            "reviewer_id",
        ]
        props = {
            **common,
            "predicted_client_status": {"type": "integer", "enum": [3, 5]},
            "authenticity_assessment": {"type": "string"},
            "operational_tier": {"type": "string", "enum": ACTION_TIERS},
            "operational_discard": {"type": "boolean"},
            "field_invalidity_ledger": {"type": "array", "items": {"type": "string"}},
            "hard_invalidity_count": {"type": "integer", "minimum": 0},
            "soft_concern_count": {"type": "integer", "minimum": 0},
        }
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["batch_id", "system_version", "decisions"],
        "properties": {
            "batch_id": {"type": "string"},
            "system_version": {"type": "string"},
            "decisions": {"type": "array", "items": {"type": "object", "additionalProperties": False, "required": required, "properties": props}},
        },
    }


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb: Any) -> dict[str, list[str]]:
    qmap: dict[str, list[str]] = {}
    current: str | None = None
    for a, b, c in wb["Datamap"].iter_rows(values_only=True):
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            current = a_s.split("]", 1)[0][1:]
            qmap[current] = [a_s]
            continue
        if current and (a_s or b is not None or c is not None):
            if a_s:
                qmap[current].append(a_s)
            if b is not None or c is not None:
                qmap[current].append(f"{b} | {c}")
    return qmap


def field_text(field: str, qmap: dict[str, list[str]]) -> str:
    lines = qmap.get(field, [])
    if not lines:
        return field
    first = lines[0]
    if "] | " in first:
        return first.split("] | ", 1)[1]
    if "]: " in first:
        return first.split("]: ", 1)[1]
    return first


def value_text(field: str, value: Any, qmap: dict[str, list[str]]) -> str:
    if value in (None, ""):
        return ""
    sv = str(value)
    for line in qmap.get(field, []):
        if " | " not in line:
            continue
        left, right = line.split(" | ", 1)
        if clean(left) == sv:
            return f"{sv}={right}"
    return sv


def field_role(name: str, qmap: dict[str, list[str]]) -> str:
    low = f"{name} {field_text(name, qmap)}".lower()
    if re.search(r"open|other|specify|remember|why|describe|explain|comment", low):
        return "open_end_or_other_specify"
    if re.search(r"recommend|likely|satisf|importance|attribute|scale|agree", low):
        return "matrix_or_scale"
    if re.search(r"aware|consider|purchase|use|own|brand|product|supplier|store|advert", low):
        return "funnel_or_entity"
    if re.search(r"age|gender|state|income|education|employ|region|home", low):
        return "persona_or_demographic"
    return "closed_or_routed"


def response_type(name: str, qmap: dict[str, list[str]]) -> str:
    lines = qmap.get(name, [])
    if any("Open text" in x for x in lines):
        return "open_text"
    if any("Open numeric" in x for x in lines):
        return "open_numeric"
    if re.search(r"r\d+c\d+", name):
        return "matrix_cell"
    if len([x for x in lines if " | " in x]) > 0:
        return "coded"
    return "raw"


def pct(values: list[float], v: float | None) -> float | None:
    if v is None:
        return None
    return round(sum(1 for x in values if x <= v) / len(values), 3)


def matrix_stats(values: list[Any]) -> dict[str, Any]:
    xs = [clean(v) for v in values if v not in (None, "")]
    if not xs:
        return {}
    counts = Counter(xs)
    run = longest = 1
    prev = None
    for x in xs:
        if x == prev:
            run += 1
        else:
            run = 1
            prev = x
        longest = max(longest, run)
    entropy = 0.0
    for n in counts.values():
        p = n / len(xs)
        entropy -= p * math.log2(p)
    return {
        "answered": len(xs),
        "unique": len(counts),
        "modal_value": counts.most_common(1)[0][0],
        "modal_share": round(counts.most_common(1)[0][1] / len(xs), 3),
        "longest_run": longest,
        "entropy": round(entropy, 3),
    }


def echo_diagnostics() -> dict[str, Any]:
    rows = [json.loads(l) for l in (ECHO_INTERNAL / "evaluated_rows.jsonl").read_text().splitlines()]
    thresholds = {}
    for t in [0.25, 0.35, 0.50, 0.65, 0.75]:
        thresholds[str(t)] = calc_metrics([(float(r["client_rejection_probability"]) >= t, bool(r["actual_rejected"]), float(r["client_rejection_probability"])) for r in rows])
    tiers = {}
    tier_order = ["Clean keep", "Keep with note", "Light review", "Review closely", "Exclude candidate"]
    for cutoff in ["Exclude candidate", "Review closely", "Light review"]:
        idx = tier_order.index(cutoff)
        tiers[cutoff] = calc_metrics([(tier_order.index(r["agent_decision"]) >= idx, bool(r["actual_rejected"]), float(r["client_rejection_probability"])) for r in rows])
    by_tier = {tier: {"rows": 0, "client_rejected": 0} for tier in tier_order}
    for r in rows:
        by_tier[r["agent_decision"]]["rows"] += 1
        by_tier[r["agent_decision"]]["client_rejected"] += int(bool(r["actual_rejected"]))
    fn = [r for r in rows if r["error_type"] == "FN"]
    return {
        "fixed_echo_seal": json.loads((ECHO_INTERNAL / "seal_manifest.json").read_text()),
        "threshold_diagnostics": thresholds,
        "tier_cutoff_diagnostics": tiers,
        "client_label_by_tier": by_tier,
        "false_negative_count": len(fn),
        "false_negative_tier4": sum(r["agent_decision"] == "Review closely" for r in fn),
        "false_negative_probability_above_050": sum(float(r["client_rejection_probability"]) >= 0.50 for r in fn),
        "selected_client_status_threshold": CLIENT_STATUS_THRESHOLD,
        "selected_threshold_rationale": "No fixed threshold met recall>=0.70, precision>=0.50, balanced accuracy>=0.65; 0.35 had the highest MCC among 0.35, 0.50, and 0.65.",
    }


def calc_metrics(items: list[tuple[bool, bool, float]]) -> dict[str, float | int]:
    tp = fp = tn = fn = 0
    for pred, actual, _ in items:
        if pred and actual:
            tp += 1
        elif pred and not actual:
            fp += 1
        elif not pred and actual:
            fn += 1
        else:
            tn += 1
    n = len(items)
    div = lambda a, b: a / b if b else 0.0
    precision = div(tp, tp + fp)
    recall = div(tp, tp + fn)
    specificity = div(tn, tn + fp)
    npv = div(tn, tn + fn)
    f1 = div(2 * precision * recall, precision + recall)
    neg_f1 = div(2 * npv * specificity, npv + specificity)
    den = math.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    mcc = div(tp * tn - fp * fn, den)
    po = div(tp + tn, n)
    pe = div((tp + fp) * (tp + fn) + (tn + fn) * (tn + fp), n * n)
    scores = sorted((s, y) for _, y, s in items)
    pos = sum(y for _, y in scores)
    neg = len(scores) - pos
    rank_sum = sum(i + 1 for i, (_, y) in enumerate(scores) if y)
    auroc = div(rank_sum - pos * (pos + 1) / 2, pos * neg)
    desc = sorted(scores, reverse=True)
    ap = hit = 0.0
    for i, (_, y) in enumerate(desc, 1):
        if y:
            hit += 1
            ap += hit / i
    return {
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
        "accuracy": div(tp + tn, n),
        "balanced_accuracy": (recall + specificity) / 2,
        "rejected_precision": precision,
        "rejected_recall": recall,
        "specificity": specificity,
        "NPV": npv,
        "F1_rejected": f1,
        "F1_accepted": neg_f1,
        "macro_F1": (f1 + neg_f1) / 2,
        "MCC": mcc,
        "cohen_kappa": div(po - pe, 1 - pe),
        "AUROC": auroc,
        "AUPRC": div(ap, pos),
        "Brier": statistics.mean((s - float(y)) ** 2 for _, y, s in items) if items else 0,
        "predicted_rejection_rate": div(tp + fp, n),
        "actual_rejection_rate": div(tp + fn, n),
    }


def prepare() -> None:
    INTERNAL.mkdir(parents=True, exist_ok=True)
    PUBLIC.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(BLIND, read_only=True, data_only=True)
    ws = wb["A1"]
    qmap = load_datamap(wb)
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    qtimes = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid")) or clean(raw.get("record")) or f"row_{xrow}"
        raw["_respondent_id"] = rid
        raw["_source_excel_row"] = xrow
        rows.append(raw)
        try:
            qtimes.append(float(raw.get("qtime")))
        except Exception:
            pass
    raw_path = INTERNAL / "blind_raw_row_packets.ndjson"
    with raw_path.open("w") as f:
        for row in rows:
            vals = {k: v for k, v in row.items() if not k.startswith("_") and not excluded(k)}
            f.write(json.dumps({"respondent_id": row["_respondent_id"], "source_excel_row": row["_source_excel_row"], "raw_values": vals}, ensure_ascii=False, separators=(",", ":")) + "\n")
    open_counts = Counter()
    vector_groups: defaultdict[str, list[str]] = defaultdict(list)
    for row in rows:
        vector = []
        for k, v in row.items():
            if k.startswith("_") or excluded(k) or v in (None, ""):
                continue
            sv = clean(v)
            if field_role(k, qmap) == "open_end_or_other_specify":
                key = re.sub(r"[^a-z0-9 ]+", "", sv.lower())
                if len(key) >= 3:
                    open_counts[key] += 1
            if response_type(k, qmap) in {"coded", "matrix_cell"}:
                vector.append(f"{k}={sv}")
        digest = hashlib.sha1("|".join(vector).encode()).hexdigest()[:16]
        vector_groups[digest].append(row["_respondent_id"])
    contract = {
        "source_workbook": str(BLIND),
        "source_sha256": sha256(BLIND),
        "rows": len(rows),
        "question_contract": {},
        "relation_graph": [
            "Persona and qualification must support downstream product/category experience.",
            "Awareness, consideration, usage, preference, recommendation, ad recall, switching, and loyalty form a chained funnel.",
            "Every open end must answer the specific question and semantic dimension, not merely name something in the broad category.",
            "Uniformity is evaluated by matrix purpose, route, timing, and local protective evidence.",
            "Client-status prediction is broader than authenticity and can include unusable or invalid response behavior.",
        ],
        "echo_diagnostics": echo_diagnostics(),
    }
    for k, lines in qmap.items():
        if excluded(k):
            continue
        contract["question_contract"][k] = {
            "question_id": k,
            "question_text_and_labels": lines[:80],
            "response_type": response_type(k, qmap),
            "field_role": field_role(k, qmap),
            "expected_semantic_dimension": "Answer the exact question text and entity/field role shown here; local validity must be judged before global persona protection.",
            "blank_interpretation": "Blank may be skipped, routed out, missing, or invalid; resolve from local context.",
            "legitimate_answer_patterns": "Concise, rough, misspelled, or uniform answers may be legitimate when locally responsive and coherent with the chain.",
        }
    packets = []
    evidence = []
    for row in rows:
        rid = row["_respondent_id"]
        qtime = None
        try:
            qtime = float(row.get("qtime"))
        except Exception:
            pass
        answers = []
        opens = []
        matrices: defaultdict[str, list[Any]] = defaultdict(list)
        words = 0
        duplicate_opens = []
        vector = []
        for k, v in row.items():
            if k.startswith("_") or excluded(k) or v in (None, ""):
                continue
            sv = clean(v)
            role = field_role(k, qmap)
            rt = response_type(k, qmap)
            entry = {
                "field": k,
                "question": field_text(k, qmap),
                "role": role,
                "response_type": rt,
                "raw_value": v,
                "decoded_value": value_text(k, v, qmap),
            }
            if role == "open_end_or_other_specify":
                opens.append(entry)
                words += len(re.findall(r"\w+", sv))
                key = re.sub(r"[^a-z0-9 ]+", "", sv.lower())
                if len(key) >= 3 and open_counts[key] > 1:
                    duplicate_opens.append(f"{k} repeats {open_counts[key]}x: {sv}")
            else:
                answers.append(entry)
            if rt == "matrix_cell":
                base = re.sub(r"c\d+$", "", k)
                matrices[base].append(v)
            if rt in {"coded", "matrix_cell"}:
                vector.append(f"{k}={sv}")
        digest = hashlib.sha1("|".join(vector).encode()).hexdigest()[:16]
        peers = [x for x in vector_groups[digest] if x != rid][:8]
        packet = {
            "respondent_id": rid,
            "source_excel_row": row["_source_excel_row"],
            "raw_pointer": "blind_raw_row_packets.ndjson",
            "timing": {"qtime_seconds": qtime, "qtime_percentile": pct(qtimes, qtime), "date": row.get("date")},
            "technical": {"supplier": row.get("SUPNAME"), "ip": row.get("ipAddress"), "user_agent": row.get("userAgent")},
            "decoded_non_open_answers": answers[:220],
            "open_ends_verbatim": opens,
            "blind_population_context": {
                "open_end_word_count": words,
                "duplicate_open_end_candidates": duplicate_opens[:12],
                "exact_response_vector_peers": peers,
                "matrix_statistics": {k: matrix_stats(v) for k, v in matrices.items() if matrix_stats(v)},
            },
        }
        packets.append(packet)
        evidence.append({"respondent_id": rid, "source_excel_row": row["_source_excel_row"], "qtime_seconds": qtime, "qtime_percentile": pct(qtimes, qtime), "open_words": words, "duplicate_open_count": len(duplicate_opens), "vector_peer_count": len(peers)})
    (INTERNAL / "question_contract.json").write_text(json.dumps(contract, indent=2, ensure_ascii=False))
    (INTERNAL / "semantic_packets.ndjson").write_text("\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in packets) + "\n")
    with (INTERNAL / "blind_corpus_evidence.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(evidence[0]))
        w.writeheader()
        w.writerows(evidence)
    (INTERNAL / "incumbent_schema.json").write_text(json.dumps(decision_schema("incumbent"), indent=2))
    (INTERNAL / "challenger_schema.json").write_text(json.dumps(decision_schema("challenger"), indent=2))
    write_prompts()
    write_batches("incumbent", 16)
    write_batches("challenger", 12)
    manifest = {"state": "DELTA_BLIND_CONTEXT_FROZEN", "annotated_opened": False, "rows": len(rows), "source_sha256": sha256(BLIND), "question_contract_sha256": sha256(INTERNAL / "question_contract.json"), "packets_sha256": sha256(INTERNAL / "semantic_packets.ndjson"), "client_status_threshold": CLIENT_STATUS_THRESHOLD}
    (INTERNAL / "blind_context_manifest.json").write_text(json.dumps(manifest, indent=2))
    print(json.dumps(manifest, indent=2))


def write_prompts() -> None:
    incumbent = f"""You are the incumbent Autosurvey reviewer. Delta labels are sealed and unavailable. Use the post-Echo methodology, but do not use the new field-first challenger architecture. Review each respondent as a full response chain for authenticity and client cleaning risk.

Return one schema-valid batch object. Use system_version {INCUMBENT_VERSION}. Client-status prediction is separate from operational tier. Set predicted_client_status to 5 when client_rejection_probability >= {CLIENT_STATUS_THRESHOLD}; otherwise 3. Only Exclude candidate sets operational_discard true.

Consider timing, matrices, open-end grounding, funnel consistency, routing, duplicate context, and protective human evidence. Keep rationale specific and cite actual packet evidence."""
    challenger = f"""You are the challenger Autosurvey field-first reviewer. Delta labels are sealed and unavailable. Use system_version {CHALLENGER_VERSION}.

Ordered method:
1. Read the decoded question contract and each answer's field role, response type, question text, answer text, and raw value.
2. Judge local field validity before global story. Classify important answers as responsive, partially responsive, nonresponsive, wrong semantic dimension, off-topic, invalid type, impossible value, route-inconsistent, unsupported other-specify, mechanically repeated, substantively plausible, or locally protected.
3. Separate hard invalidity from soft concern. Hard invalidity includes wrong-question answers, wrong semantic dimension, unsupported other-specifies, off-category entities, impossible allocations, route violations, copied text from another prompt, or invalid matrix structure. Soft concern includes speed, shortness, generic text, broad selection, straightlining, repetition, high positivity, or weak detail.
4. Evaluate relationships: qualification to expertise, awareness to consideration, use to preference, ad exposure to recall, closed answer to open explanation, persona to purchasing behavior, and matrix answers to stated preferences.
5. Use blind population context as evidence, not as an automatic decision.
6. Apply local human protection. A plausible persona or brand chain protects only the specific concern it explains.
7. Predict client status separately from authenticity and operational action.

Provisional principles: two independent hard invalidities strongly support predicted status=5; one hard invalidity plus two independent soft concerns may support status=5; soft concerns alone usually support review rather than exclusion. Predicted status=5 may coexist with Review closely. Exclude candidate requires high-confidence authenticity or usability failure. Use threshold {CLIENT_STATUS_THRESHOLD}: predicted_client_status must be 5 when client_rejection_probability >= {CLIENT_STATUS_THRESHOLD}, otherwise 3.

Return one schema-valid batch object. Cite exact field evidence and include a concise field_invalidity_ledger."""
    (INTERNAL / "incumbent_prompt.txt").write_text(incumbent)
    (INTERNAL / "challenger_prompt.txt").write_text(challenger)


def write_batches(system: str, size: int) -> None:
    packets = [json.loads(l) for l in (INTERNAL / "semantic_packets.ndjson").read_text().splitlines()]
    batch_dir = INTERNAL / f"{system}_batches"
    if batch_dir.exists():
        shutil.rmtree(batch_dir)
    batch_dir.mkdir(parents=True)
    # Deterministic shuffle spreads suppliers/timing clusters across workers while remaining reproducible.
    packets = sorted(packets, key=lambda p: hashlib.sha1(p["respondent_id"].encode()).hexdigest())
    batches = [packets[i : i + size] for i in range(0, len(packets), size)]
    manifest = []
    for i, batch in enumerate(batches, 1):
        batch_id = f"delta_{system}_{i:03d}"
        path = batch_dir / f"{batch_id}.jsonl"
        path.write_text("\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in batch) + "\n")
        manifest.append({"batch_id": batch_id, "path": str(path), "state": "PENDING", "respondent_ids": [p["respondent_id"] for p in batch], "respondent_count": len(batch), "attempts": 0})
    (INTERNAL / f"{system}_manifest.json").write_text(json.dumps(manifest, indent=2))


def validate(system: str, batch: dict[str, Any], out: Path) -> tuple[bool, str]:
    try:
        obj = json.loads(out.read_text())
    except Exception as e:
        return False, str(e)
    if obj.get("batch_id") != batch["batch_id"]:
        return False, "batch_id mismatch"
    dec = obj.get("decisions")
    if not isinstance(dec, list):
        return False, "missing decisions"
    ids = [d.get("respondent_id") for d in dec]
    if set(ids) != set(batch["respondent_ids"]) or len(ids) != len(set(ids)):
        return False, "coverage mismatch"
    schema = decision_schema(system)["properties"]["decisions"]["items"]["required"]
    for d in dec:
        for key in schema:
            if d.get(key) in (None, "", []):
                if key == "field_invalidity_ledger":
                    continue
                return False, f"missing {key} {d.get('respondent_id')}"
        if d["client_rejection_probability"] >= CLIENT_STATUS_THRESHOLD and d["predicted_client_status"] != 5:
            return False, f"threshold mismatch {d.get('respondent_id')}"
        if d["client_rejection_probability"] < CLIENT_STATUS_THRESHOLD and d["predicted_client_status"] != 3:
            return False, f"threshold mismatch {d.get('respondent_id')}"
        if (d["operational_tier"] == "Exclude candidate") != bool(d["operational_discard"]):
            return False, f"discard mismatch {d.get('respondent_id')}"
    return True, "ok"


def run_one(system: str, batch: dict[str, Any], timeout_s: int) -> dict[str, Any]:
    out_dir = INTERNAL / f"{system}_decisions"
    log_dir = INTERNAL / f"{system}_logs"
    out_dir.mkdir(exist_ok=True)
    log_dir.mkdir(exist_ok=True)
    out = out_dir / f"{batch['batch_id']}.json"
    if out.exists():
        ok, msg = validate(system, batch, out)
        if ok:
            return {"batch_id": batch["batch_id"], "state": "COMPLETE", "message": msg}
        out.rename(out.with_suffix(".invalid.json"))
    prompt = (INTERNAL / f"{system}_prompt.txt").read_text()
    payload = prompt + f"\n\nBATCH_ID: {batch['batch_id']}\nRAW PACKETS:\n" + Path(batch["path"]).read_text()
    tmp = out.with_suffix(".tmp.json")
    log = log_dir / f"{batch['batch_id']}.log"
    cmd = ["codex", "exec", "--ignore-user-config", "--ignore-rules", "--skip-git-repo-check", "--ephemeral", "--sandbox", "read-only", "--output-schema", str(INTERNAL / f"{system}_schema.json"), "-o", str(tmp), "-"]
    if MODEL:
        cmd[2:2] = ["-m", MODEL]
    with log.open("a") as lf:
        lf.write(f"\n--- {time.strftime('%Y-%m-%dT%H:%M:%S')} {system} ---\n")
        try:
            proc = subprocess.run(cmd, input=payload, text=True, stdout=lf, stderr=subprocess.STDOUT, timeout=timeout_s, cwd=str(Path.home()))
        except subprocess.TimeoutExpired:
            return {"batch_id": batch["batch_id"], "state": "RETRY", "message": "timeout"}
    if proc.returncode != 0 or not tmp.exists():
        return {"batch_id": batch["batch_id"], "state": "RETRY", "message": f"code {proc.returncode}"}
    tmp.replace(out)
    ok, msg = validate(system, batch, out)
    return {"batch_id": batch["batch_id"], "state": "COMPLETE" if ok else "FAILED_SCHEMA", "message": msg}


def run_queue(system: str, workers: int, timeout_s: int) -> None:
    manifest_path = INTERNAL / f"{system}_manifest.json"
    manifest = json.loads(manifest_path.read_text())
    for b in manifest:
        out = INTERNAL / f"{system}_decisions" / f"{b['batch_id']}.json"
        if out.exists():
            ok, msg = validate(system, b, out)
            b["state"] = "COMPLETE" if ok else "FAILED_SCHEMA"
            b["message"] = msg
    pending = [b for b in manifest if b["state"] != "COMPLETE"]
    print(f"{system} pending={len(pending)} complete={sum(b['state']=='COMPLETE' for b in manifest)} workers={workers}")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futs = {pool.submit(run_one, system, b, timeout_s): b for b in pending}
        for fut in as_completed(futs):
            res = fut.result()
            for b in manifest:
                if b["batch_id"] == res["batch_id"]:
                    b["state"] = res["state"]
                    b["message"] = res["message"]
                    b["attempts"] = int(b.get("attempts", 0)) + 1
            manifest_path.write_text(json.dumps(manifest, indent=2))
            print(json.dumps({"result": res, "complete": sum(b["state"] == "COMPLETE" for b in manifest), "total": len(manifest)}), flush=True)


def assemble(system: str, seal: bool) -> None:
    manifest = json.loads((INTERNAL / f"{system}_manifest.json").read_text())
    packets = {json.loads(l)["respondent_id"]: json.loads(l) for l in (INTERNAL / "semantic_packets.ndjson").read_text().splitlines()}
    decs = []
    for b in manifest:
        out = INTERNAL / f"{system}_decisions" / f"{b['batch_id']}.json"
        ok, msg = validate(system, b, out) if out.exists() else (False, "missing")
        if not ok:
            raise SystemExit(f"{system} batch {b['batch_id']} invalid: {msg}")
        decs.extend(json.loads(out.read_text())["decisions"])
    if set(d["respondent_id"] for d in decs) != set(packets) or len(decs) != len(packets):
        raise SystemExit(f"{system} ledger coverage failed")
    ledger = INTERNAL / f"{system}_sealed_ledger.jsonl"
    version = INCUMBENT_VERSION if system == "incumbent" else CHALLENGER_VERSION
    with ledger.open("w") as f:
        for d in sorted(decs, key=lambda x: packets[x["respondent_id"]]["source_excel_row"]):
            f.write(json.dumps({"system": system, "system_version": version, **d}, ensure_ascii=False, separators=(",", ":")) + "\n")
    seal_doc = {
        "state": "SEALED_PENDING_UNBLIND" if seal else "ASSEMBLED",
        "system": system,
        "rows": len(decs),
        "sealed_at": time.strftime("%Y-%m-%dT%H:%M:%S") if seal else None,
        "annotated_opened": False,
        "source_sha256": sha256(BLIND),
        "question_contract_sha256": sha256(INTERNAL / "question_contract.json"),
        "prompt_sha256": sha256(INTERNAL / f"{system}_prompt.txt"),
        "schema_sha256": sha256(INTERNAL / f"{system}_schema.json"),
        "manifest_sha256": sha256(INTERNAL / f"{system}_manifest.json"),
        "ledger_sha256": sha256(ledger),
        "client_status_threshold": CLIENT_STATUS_THRESHOLD,
    }
    (INTERNAL / f"{system}_seal_manifest.json").write_text(json.dumps(seal_doc, indent=2))
    print(json.dumps(seal_doc, indent=2))


def load_labels() -> dict[str, int]:
    wb = openpyxl.load_workbook(ANNOTATED, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    lower = {h.lower(): h for h in headers}
    uuid_col = lower.get("uuid")
    status_col = lower.get("status")
    if not uuid_col or not status_col:
        raise SystemExit(f"Cannot reconcile annotated labels: {headers[:40]}")
    labels = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(row.get(uuid_col))
        if rid:
            labels[rid] = int(row.get(status_col))
    return labels


def evaluate() -> None:
    for system in ["incumbent", "challenger"]:
        seal = json.loads((INTERNAL / f"{system}_seal_manifest.json").read_text())
        if seal.get("state") != "SEALED_PENDING_UNBLIND":
            raise SystemExit(f"{system} not sealed")
    labels = load_labels()
    results = {}
    evaluated_by_system = {}
    for system in ["incumbent", "challenger"]:
        rows = [json.loads(l) for l in (INTERNAL / f"{system}_sealed_ledger.jsonl").read_text().splitlines()]
        if any(r["respondent_id"] not in labels for r in rows):
            raise SystemExit(f"{system} label reconciliation failed")
        evaluated = []
        for r in rows:
            actual = labels[r["respondent_id"]] == 5
            pred = r["predicted_client_status"] == 5
            if pred and actual:
                et = "TP"
            elif pred and not actual:
                et = "FP"
            elif not pred and actual:
                et = "FN"
            else:
                et = "TN"
            evaluated.append({**r, "client_status": labels[r["respondent_id"]], "actual_rejected": actual, "client_error_type": et})
        items = [(r["predicted_client_status"] == 5, bool(r["actual_rejected"]), float(r["client_rejection_probability"])) for r in evaluated]
        m = calc_metrics(items)
        m.update({
            "always_accept_accuracy": sum(not r["actual_rejected"] for r in evaluated) / len(evaluated),
            "tier_volumes": dict(Counter(r["operational_tier"] for r in evaluated)),
            "operational_discard_volume": sum(r["operational_discard"] for r in evaluated),
            "review_burden": sum(r["operational_tier"] in {"Light review", "Review closely", "Exclude candidate"} for r in evaluated) / len(evaluated),
        })
        results[system] = m
        evaluated_by_system[system] = evaluated
        (INTERNAL / f"{system}_evaluated_rows.jsonl").write_text("\n".join(json.dumps(r, ensure_ascii=False, separators=(",", ":")) for r in evaluated) + "\n")
    inc = {r["respondent_id"]: r for r in evaluated_by_system["incumbent"]}
    chal = {r["respondent_id"]: r for r in evaluated_by_system["challenger"]}
    disagreements = []
    corrected = damaged = 0
    for rid, c in chal.items():
        i = inc[rid]
        if i["predicted_client_status"] != c["predicted_client_status"]:
            disagreements.append({"respondent_id": rid, "incumbent": i["predicted_client_status"], "challenger": c["predicted_client_status"], "actual": c["client_status"]})
            if i["client_error_type"] in {"FP", "FN"} and c["client_error_type"] in {"TP", "TN"}:
                corrected += 1
            if i["client_error_type"] in {"TP", "TN"} and c["client_error_type"] in {"FP", "FN"}:
                damaged += 1
    promoted = (
        results["challenger"]["MCC"] > results["incumbent"]["MCC"]
        and results["challenger"]["balanced_accuracy"] > results["incumbent"]["balanced_accuracy"]
        and results["challenger"]["rejected_recall"] - results["incumbent"]["rejected_recall"] >= 0.15
        and results["challenger"]["rejected_precision"] >= 0.50
        and results["challenger"]["specificity"] >= 0.80
    )
    comparison = {"metrics": results, "disagreements": len(disagreements), "corrected_by_challenger": corrected, "damaged_by_challenger": damaged, "promoted": promoted}
    (INTERNAL / "delta_comparison_metrics.json").write_text(json.dumps(comparison, indent=2))
    print(json.dumps(comparison, indent=2))


def outputs() -> None:
    PUBLIC.mkdir(parents=True, exist_ok=True)
    for p in PUBLIC.iterdir():
        if p.is_file():
            p.unlink()
    metrics = json.loads((INTERNAL / "delta_comparison_metrics.json").read_text())
    inc = [json.loads(l) for l in (INTERNAL / "incumbent_evaluated_rows.jsonl").read_text().splitlines()]
    chal = [json.loads(l) for l in (INTERNAL / "challenger_evaluated_rows.jsonl").read_text().splitlines()]
    raw = {json.loads(l)["respondent_id"]: json.loads(l) for l in (INTERNAL / "blind_raw_row_packets.ndjson").read_text().splitlines()}
    cdict = {r["respondent_id"]: r for r in chal}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Labeled Rows"
    raw_cols = list(next(iter(raw.values()))["raw_values"])
    cols = ["respondent_id", "source_excel_row", "client_status", "actual_rejected", "incumbent_status_prediction", "challenger_status_prediction", "promoted_final_status_prediction", "authenticity_risk", "operational_tier", "client_rejection_probability", "semantic_decision_criteria", "field_invalidity_ledger", "signals", "signal_evidence", "protective_evidence", "incumbent_challenger_agreement", "client_agreement", "error_type", "skill_version"] + raw_cols
    ws.append(cols)
    promoted_system = "challenger" if metrics["promoted"] else "incumbent"
    for ir in inc:
        cr = cdict[ir["respondent_id"]]
        final = cr if promoted_system == "challenger" else ir
        rv = raw[ir["respondent_id"]]["raw_values"]
        vals = [
            ir["respondent_id"], ir["source_excel_row"], ir["client_status"], ir["actual_rejected"],
            ir["predicted_client_status"], cr["predicted_client_status"], final["predicted_client_status"],
            final["authenticity_risk"], final["operational_tier"], final["client_rejection_probability"],
            final["semantic_decision_criteria"], "; ".join(cr.get("field_invalidity_ledger", [])),
            "; ".join(final.get("signals", [])), "; ".join(final.get("signal_evidence", [])),
            final.get("protective_evidence", ""), ir["predicted_client_status"] == cr["predicted_client_status"],
            final["client_error_type"] in {"TP", "TN"}, final["client_error_type"], final["system_version"],
        ] + [rv.get(c, "") for c in raw_cols]
        ws.append(vals)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="263238")
    for col in range(1, min(ws.max_column, 24) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Delta Transfer Benchmark"
    dash["A1"].font = openpyxl.styles.Font(bold=True, size=16)
    r = 3
    for system in ["incumbent", "challenger"]:
        m = metrics["metrics"][system]
        dash.cell(r, 1, system.title())
        dash.cell(r, 1).font = openpyxl.styles.Font(bold=True, size=13)
        for k in ["TP", "FP", "TN", "FN", "accuracy", "always_accept_accuracy", "balanced_accuracy", "rejected_precision", "rejected_recall", "specificity", "NPV", "F1_rejected", "F1_accepted", "macro_F1", "MCC", "cohen_kappa", "AUROC", "AUPRC", "Brier", "predicted_rejection_rate", "actual_rejection_rate", "operational_discard_volume", "review_burden"]:
            r += 1
            dash.cell(r, 1, k)
            dash.cell(r, 2, m[k])
        r += 2
    dash["D3"] = "Promotion decision"
    dash["E3"] = "Promoted challenger" if metrics["promoted"] else "Challenger not promoted"
    dash["D4"] = "Disagreements"
    dash["E4"] = metrics["disagreements"]
    dash["D5"] = "Corrected by challenger"
    dash["E5"] = metrics["corrected_by_challenger"]
    dash["D6"] = "Damaged by challenger"
    dash["E6"] = metrics["damaged_by_challenger"]
    dash["D8"] = "Seal integrity"
    dash["D9"] = "Incumbent ledger hash"
    dash["E9"] = json.loads((INTERNAL / "incumbent_seal_manifest.json").read_text())["ledger_sha256"]
    dash["D10"] = "Challenger ledger hash"
    dash["E10"] = json.loads((INTERNAL / "challenger_seal_manifest.json").read_text())["ledger_sha256"]
    for col in range(1, 7):
        dash.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 26
    xlsx = PUBLIC / "AUTOSURVEY_RESULTS.xlsx"
    wb.save(xlsx)
    md = evolution_report(metrics)
    (PUBLIC / "AUTOSURVEY_EVOLUTION.md").write_text(md)
    print(xlsx)
    print(PUBLIC / "AUTOSURVEY_EVOLUTION.md")


def evolution_report(metrics: dict[str, Any]) -> str:
    inc = metrics["metrics"]["incumbent"]
    chal = metrics["metrics"]["challenger"]
    return f"""# Autosurvey Delta Transfer Evolution Report

## Read This First

We used Echo only as consumed development evidence. The Echo sealed predictions were preserved and never revised. The main Echo diagnosis was that the system ranked some client-rejected rows as risky, but forced binary rejection to the operational discard tier. The fixed Echo diagnostics selected a provisional client-status threshold of `{CLIENT_STATUS_THRESHOLD}` before Delta labels were opened.

For Delta, we froze two systems before unblinding:

- Incumbent: the post-Echo reviewer using the prior full-chain approach.
- Challenger: a field-first reviewer that separates local field validity, client-status prediction, authenticity assessment, and operational tier.

Both ledgers were sealed before the annotated Delta workbook was opened.

## Delta Result

| Metric | Incumbent | Challenger |
|---|---:|---:|
| TP | {inc['TP']} | {chal['TP']} |
| FP | {inc['FP']} | {chal['FP']} |
| TN | {inc['TN']} | {chal['TN']} |
| FN | {inc['FN']} | {chal['FN']} |
| Accuracy | {inc['accuracy']:.3f} | {chal['accuracy']:.3f} |
| Balanced accuracy | {inc['balanced_accuracy']:.3f} | {chal['balanced_accuracy']:.3f} |
| Rejected precision | {inc['rejected_precision']:.3f} | {chal['rejected_precision']:.3f} |
| Rejected recall | {inc['rejected_recall']:.3f} | {chal['rejected_recall']:.3f} |
| Specificity | {inc['specificity']:.3f} | {chal['specificity']:.3f} |
| MCC | {inc['MCC']:.3f} | {chal['MCC']:.3f} |
| AUROC | {inc['AUROC']:.3f} | {chal['AUROC']:.3f} |
| AUPRC | {inc['AUPRC']:.3f} | {chal['AUPRC']:.3f} |

## Promotion Decision

Challenger promoted: **{metrics['promoted']}**.

The challenger corrected {metrics['corrected_by_challenger']} incumbent errors and damaged {metrics['damaged_by_challenger']} rows. Promotion required higher MCC, higher balanced accuracy, at least 0.15 recall gain, precision at or above 0.50, and specificity at or above 0.80.

## What Echo Taught

Echo showed that client status is not identical to provable fraud. TFG's status=5 boundary includes broader unusable response behavior: wrong-field answers, nonresponsive open ends, unsupported other-specifies, mechanical matrices, invalid allocations, and inattentive completion. That distinction is now explicit. Authenticity risk estimates fabrication or gaming; client-status prediction estimates the TFG cleaning boundary; operational tier controls the actual discard set.

## What Delta Tested

Delta tested whether local field validity and separated status prediction transfer to a new survey. The challenger did not merely predict more rejects by tier. It made a separate status prediction using the frozen threshold and kept operational discard as a narrower action.

## Accepted-Respondent Lesson

Accepted respondents can still be brief, rough, repetitive, or uniform. The key guardrail is local responsiveness: a short answer is acceptable when it answers the exact prompt and coheres with the chain. A plausible persona should not protect unrelated invalid fields.

## Next Benchmark

The next benchmark should test whether the field-first client-status ontology remains stable on a non-water-filtration category and whether the threshold of `{CLIENT_STATUS_THRESHOLD}` is still appropriate without tuning on Delta.
"""


def main() -> None:
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd", required=True)
    sub.add_parser("prepare")
    rq = sub.add_parser("run")
    rq.add_argument("system", choices=["incumbent", "challenger"])
    rq.add_argument("--workers", type=int, default=6)
    rq.add_argument("--timeout", type=int, default=900)
    assem = sub.add_parser("assemble")
    assem.add_argument("system", choices=["incumbent", "challenger"])
    assem.add_argument("--seal", action="store_true")
    sub.add_parser("evaluate")
    sub.add_parser("outputs")
    args = p.parse_args()
    if args.cmd == "prepare":
        prepare()
    elif args.cmd == "run":
        run_queue(args.system, args.workers, args.timeout)
    elif args.cmd == "assemble":
        assemble(args.system, args.seal)
    elif args.cmd == "evaluate":
        evaluate()
    elif args.cmd == "outputs":
        outputs()


if __name__ == "__main__":
    main()
