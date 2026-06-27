#!/usr/bin/env python3
"""Compare v4 vs v5 vs v5.1 holistic agent review results against client ground truth for ECHO.

Only status=5 (discarded) and status=3 (accepted) are used as ground truth.
Other status values (0, 1, 2, 4) are excluded entirely.
"""
import json
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl

def load_client_ground_truth(annotated_path):
    """Load status labels from the client-annotated workbook.

    Only status=3 (accepted) and status=5 (discarded) are used.
    Other status values (0, 1, 2, 4, etc.) are excluded entirely —
    they represent incomplete, pending, or other non-final states
    that should NOT be counted as KEEP.
    """
    wb = openpyxl.load_workbook(annotated_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    status_idx = headers.index('status')
    uuid_idx = headers.index('uuid')

    gt = {}
    excluded = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        uuid = row[uuid_idx]
        status = row[status_idx]
        if status == 5:
            gt[uuid] = 'DISCARD'
        elif status == 3:
            gt[uuid] = 'KEEP'
        else:
            excluded += 1
    if excluded:
        print(f"  Excluded {excluded} rows with non-3/5 status (incomplete/pending/other)")
    return gt

def load_judgments(path):
    """Load agent judgments from a JSON file or directory of chunk files.
    Preserves all fields (including v6 27-field schema) for downstream analysis."""
    path = Path(path)
    judgments = {}

    if path.is_dir():
        # Merge all chunk files
        for cf in sorted(path.glob('agent_judgments_chunk_*.json')):
            with open(cf) as f:
                chunk_data = json.load(f)
            for item in chunk_data:
                rid = item.get('respondent_id') or item.get('uuid')
                entry = {
                    'judgment': item.get('agent_judgment', item.get('judgment', '')),
                    'score': item.get('agent_score', item.get('score', 0)),
                    'justification': item.get('agent_justification', item.get('justification', ''))
                }
                # Preserve all v6 fields if present
                for k, v in item.items():
                    if k not in ('respondent_id', 'uuid', 'agent_judgment', 'agent_score', 'agent_justification'):
                        entry[k] = v
                judgments[rid] = entry
    else:
        with open(path) as f:
            data = json.load(f)
        for item in data:
            rid = item.get('respondent_id') or item.get('uuid')
            entry = {
                'judgment': item.get('agent_judgment', item.get('judgment', '')),
                'score': item.get('agent_score', item.get('score', 0)),
                'justification': item.get('agent_justification', item.get('justification', ''))
            }
            for k, v in item.items():
                if k not in ('respondent_id', 'uuid', 'agent_judgment', 'agent_score', 'agent_justification'):
                    entry[k] = v
            judgments[rid] = entry
    return judgments

def compute_metrics(gt, judgments):
    """Compute precision, recall, F1, balanced accuracy for discard detection.

    DISCARD = positive prediction. KEEP/REVIEW = negative prediction.
    Ground truth: status=5 = DISCARD (positive), status=3 = KEEP (negative).
    """
    tp = fp = tn = fn = 0
    matched = 0
    for rid, gt_label in gt.items():
        if rid not in judgments:
            continue
        matched += 1
        pred = judgments[rid]['judgment']

        if gt_label == 'DISCARD':
            if pred == 'DISCARD':
                tp += 1
            else:  # REVIEW or KEEP
                fn += 1
        else:  # gt_label == 'KEEP'
            if pred == 'DISCARD':
                fp += 1
            else:  # REVIEW or KEEP
                tn += 1

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    # Balanced accuracy = (TPR + TNR) / 2
    tpr = recall  # sensitivity
    tnr = tn / (tn + fp) if (tn + fp) > 0 else 0  # specificity
    balanced_acc = (tpr + tnr) / 2

    # Soft recall: DISCARD or REVIEW counts as catching a discard
    soft_tp = sum(1 for rid, gt_label in gt.items()
                  if rid in judgments and gt_label == 'DISCARD'
                  and judgments[rid]['judgment'] in ('DISCARD', 'REVIEW'))
    soft_fn = sum(1 for rid, gt_label in gt.items()
                  if rid in judgments and gt_label == 'DISCARD'
                  and judgments[rid]['judgment'] == 'KEEP')
    soft_recall = soft_tp / (soft_tp + soft_fn) if (soft_tp + soft_fn) > 0 else 0

    return {
        'matched': matched,
        'tp': tp, 'fp': fp, 'tn': tn, 'fn': fn,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'balanced_acc': balanced_acc,
        'tpr': tpr,
        'tnr': tnr,
        'soft_recall': soft_recall,
        'discard_predicted': tp + fp,
        'review_predicted': sum(1 for r in judgments.values() if r['judgment'] == 'REVIEW'),
        'keep_predicted': sum(1 for r in judgments.values() if r['judgment'] == 'KEEP'),
    }

def print_metrics(name, m):
    print(f'\n--- {name} ---')
    print(f'  Matched: {m["matched"]}/{m["matched"]}')
    print(f'  Predicted: DISCARD={m["discard_predicted"]}, REVIEW={m["review_predicted"]}, KEEP={m["keep_predicted"]}')
    print(f'  TP={m["tp"]}, FP={m["fp"]}, TN={m["tn"]}, FN={m["fn"]}')
    print(f'  Precision:       {m["precision"]:.3f}')
    print(f'  Recall (TPR):    {m["recall"]:.3f}')
    print(f'  Specificity(TNR):{m["tnr"]:.3f}')
    print(f'  F1:              {m["f1"]:.3f}')
    print(f'  Balanced Acc:    {m["balanced_acc"]:.3f}')
    print(f'  Soft recall:     {m["soft_recall"]:.3f}')

def main():
    annotated_path = '/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx'
    v4_path = '/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run/all_judgments.json'
    v5_dir = Path('/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run_v5')
    v51_dir = Path('/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run_v51')
    v6_dir = Path('/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run_v6')
    v7_dir = Path('/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent/109-2601 Echo BH/holistic_agent_run_v7')

    print('=' * 80)
    print('ECHO v4 vs v5 vs v5.1 vs v6 vs v7 Comparison Against Client Ground Truth (status=5 only)')
    print('=' * 80)

    # Load ground truth
    gt = load_client_ground_truth(annotated_path)
    gt_discards = sum(1 for v in gt.values() if v == 'DISCARD')
    gt_keeps = sum(1 for v in gt.values() if v == 'KEEP')
    print(f'\nClient ground truth: {len(gt)} respondents (status=3/5 only)')
    print(f'  Discards (status=5): {gt_discards} ({gt_discards/len(gt)*100:.1f}%)')
    print(f'  Keeps (status=3):    {gt_keeps} ({gt_keeps/len(gt)*100:.1f}%)')

    # Captain semantic baseline (from user's reported numbers)
    captain = {
        'matched': 1566, 'tp': 281, 'fp': 203, 'tn': 810, 'fn': 272,
        'precision': 281/(281+203), 'recall': 281/(281+272),
        'tnr': 810/(810+203), 'balanced_acc': (281/(281+272) + 810/(810+203))/2,
        'soft_recall': 0, 'discard_predicted': 281+203, 'review_predicted': 0, 'keep_predicted': 0,
    }
    captain['f1'] = 2*captain['precision']*captain['recall']/(captain['precision']+captain['recall']) if (captain['precision']+captain['recall']) > 0 else 0
    print_metrics('Captain Semantic (prior run)', captain)

    # V4
    v4_judgments = load_judgments(v4_path)
    v4_metrics = compute_metrics(gt, v4_judgments)
    print_metrics('V4 Holistic Agent Review', v4_metrics)

    # V5
    if v5_dir.exists():
        v5_judgments = load_judgments(v5_dir)
        if v5_judgments:
            v5_metrics = compute_metrics(gt, v5_judgments)
            print_metrics('V5 Two-Stage (fraud + quality)', v5_metrics)
        else:
            v5_metrics = None
            print('\n--- V5: No judgments found ---')
    else:
        v5_metrics = None
        print('\n--- V5: Directory not found ---')

    # V5.1
    if v51_dir.exists():
        v51_judgments = load_judgments(v51_dir)
        if v51_judgments:
            v51_metrics = compute_metrics(gt, v51_judgments)
            print_metrics('V5.1 (ML + quota reconstruction)', v51_metrics)
        else:
            v51_metrics = None
            print('\n--- V5.1: No judgments found yet ---')
    else:
        v51_metrics = None
        print('\n--- V5.1: Directory not found ---')

    # V6
    if v6_dir.exists():
        v6_judgments = load_judgments(v6_dir)
        if v6_judgments:
            v6_metrics = compute_metrics(gt, v6_judgments)
            print_metrics('V6 (three-component + ML correlations + 27-field schema)', v6_metrics)
        else:
            v6_metrics = None
            print('\n--- V6: No judgments found yet ---')
    else:
        v6_metrics = None
        print('\n--- V6: Directory not found ---')

    # V7
    if v7_dir.exists():
        v7_judgments = load_judgments(v7_dir)
        if v7_judgments:
            v7_metrics = compute_metrics(gt, v7_judgments)
            print_metrics('V7 (calibrated thresholds + ML-driven + convergence >= 4)', v7_metrics)
        else:
            v7_metrics = None
            print('\n--- V7: No judgments found yet ---')
    else:
        v7_metrics = None
        print('\n--- V7: Directory not found ---')

    # Comparison table
    print(f'\n{"="*90}')
    print(f'{"Metric":<20} {"Captain":>10} {"V4":>10} {"V5":>10} {"V5.1":>10} {"V6":>10} {"V7":>10}')
    print(f'{"-"*90}')
    print(f'{"TP":<20} {captain["tp"]:>10} {v4_metrics["tp"]:>10} {(v5_metrics or {}).get("tp", "-"):>10} {(v51_metrics or {}).get("tp", "-"):>10} {(v6_metrics or {}).get("tp", "-"):>10} {(v7_metrics or {}).get("tp", "-"):>10}')
    print(f'{"FP":<20} {captain["fp"]:>10} {v4_metrics["fp"]:>10} {(v5_metrics or {}).get("fp", "-"):>10} {(v51_metrics or {}).get("fp", "-"):>10} {(v6_metrics or {}).get("fp", "-"):>10} {(v7_metrics or {}).get("fp", "-"):>10}')
    print(f'{"TN":<20} {captain["tn"]:>10} {v4_metrics["tn"]:>10} {(v5_metrics or {}).get("tn", "-"):>10} {(v51_metrics or {}).get("tn", "-"):>10} {(v6_metrics or {}).get("tn", "-"):>10} {(v7_metrics or {}).get("tn", "-"):>10}')
    print(f'{"FN":<20} {captain["fn"]:>10} {v4_metrics["fn"]:>10} {(v5_metrics or {}).get("fn", "-"):>10} {(v51_metrics or {}).get("fn", "-"):>10} {(v6_metrics or {}).get("fn", "-"):>10} {(v7_metrics or {}).get("fn", "-"):>10}')
    print(f'{"Precision":<20} {captain["precision"]:>10.3f} {v4_metrics["precision"]:>10.3f} {(v5_metrics or {}).get("precision", 0):>10.3f} {(v51_metrics or {}).get("precision", 0):>10.3f} {(v6_metrics or {}).get("precision", 0):>10.3f} {(v7_metrics or {}).get("precision", 0):>10.3f}')
    print(f'{"Recall":<20} {captain["recall"]:>10.3f} {v4_metrics["recall"]:>10.3f} {(v5_metrics or {}).get("recall", 0):>10.3f} {(v51_metrics or {}).get("recall", 0):>10.3f} {(v6_metrics or {}).get("recall", 0):>10.3f} {(v7_metrics or {}).get("recall", 0):>10.3f}')
    print(f'{"F1":<20} {captain["f1"]:>10.3f} {v4_metrics["f1"]:>10.3f} {(v5_metrics or {}).get("f1", 0):>10.3f} {(v51_metrics or {}).get("f1", 0):>10.3f} {(v6_metrics or {}).get("f1", 0):>10.3f} {(v7_metrics or {}).get("f1", 0):>10.3f}')
    print(f'{"Balanced Acc":<20} {captain["balanced_acc"]:>10.3f} {v4_metrics["balanced_acc"]:>10.3f} {(v5_metrics or {}).get("balanced_acc", 0):>10.3f} {(v51_metrics or {}).get("balanced_acc", 0):>10.3f} {(v6_metrics or {}).get("balanced_acc", 0):>10.3f} {(v7_metrics or {}).get("balanced_acc", 0):>10.3f}')
    print(f'{"Discards pred":<20} {captain["discard_predicted"]:>10} {v4_metrics["discard_predicted"]:>10} {(v5_metrics or {}).get("discard_predicted", "-"):>10} {(v51_metrics or {}).get("discard_predicted", "-"):>10} {(v6_metrics or {}).get("discard_predicted", "-"):>10} {(v7_metrics or {}).get("discard_predicted", "-"):>10}')
    print(f'{"="*90}')

    # V5.1 FN/FP analysis
    if v51_metrics:
        print(f'\n--- V5.1 False Negatives (missed client discards, status=5) ---')
        fn_samples = []
        for rid, gt_label in gt.items():
            if gt_label == 'DISCARD' and rid in v51_judgments:
                pred = v51_judgments[rid]['judgment']
                if pred != 'DISCARD':
                    fn_samples.append({
                        'rid': rid,
                        'pred': pred,
                        'score': v51_judgments[rid]['score'],
                        'justification': v51_judgments[rid]['justification'][:200]
                    })
        print(f'  Total FN (predicted REVIEW/KEEP for client discard): {len(fn_samples)}')
        for s in fn_samples[:10]:
            print(f'    {s["rid"]} [{s["pred"]} score={s["score"]}]: {s["justification"]}')

        print(f'\n--- V5.1 False Positives (discarded client keeps, status=3) ---')
        fp_samples = []
        for rid, gt_label in gt.items():
            if gt_label == 'KEEP' and rid in v51_judgments:
                pred = v51_judgments[rid]['judgment']
                if pred == 'DISCARD':
                    fp_samples.append({
                        'rid': rid,
                        'score': v51_judgments[rid]['score'],
                        'justification': v51_judgments[rid]['justification'][:200]
                    })
        print(f'  Total FP (predicted DISCARD for client keep): {len(fp_samples)}')
        for s in fp_samples[:10]:
            print(f'    {s["rid"]} (score={s["score"]}): {s["justification"]}')

        # --- V6 metadata analysis (if v6 judgments have the new fields) ---
        print(f'\n{"="*80}')
        print(f'V6 METADATA ANALYSIS (three-component scoring + disposition layer)')
        print(f'{"="*80}')

        # Use v7 judgments if available, otherwise v6, otherwise check v5.1
        v6_or_v51 = v7_judgments if (v7_metrics and v7_judgments) else (v6_judgments if (v6_metrics and v6_judgments) else v51_judgments)
        v6_label = 'V7' if (v7_metrics and v7_judgments) else ('V6' if (v6_metrics and v6_judgments) else 'V5.1')

        if v6_or_v51:
            sample_j = v6_or_v51.get(next(iter(v6_or_v51)), {})
            has_v6 = 'authenticity_risk' in sample_j or 'primary_removal_reason' in sample_j

        if v6_or_v51 and has_v6:
            from collections import Counter

            # Removal reason distribution
            print(f'\n--- Primary Removal Reason Distribution ({v6_label}) ---')
            reason_counts = Counter()
            reason_by_gt = defaultdict(Counter)
            for rid, gt_label in gt.items():
                if rid in v6_or_v51:
                    j = v6_or_v51[rid]
                    reason = j.get('primary_removal_reason', 'unknown')
                    reason_counts[reason] += 1
                    reason_by_gt[gt_label][reason] += 1

            for reason, count in reason_counts.most_common():
                discard_count = reason_by_gt['DISCARD'].get(reason, 0)
                keep_count = reason_by_gt['KEEP'].get(reason, 0)
                print(f'  {reason}: {count} ({discard_count} actual discards, {keep_count} actual keeps)')

            # Badopen trigger distribution
            print(f'\n--- Badopen Trigger Distribution ({v6_label}) ---')
            badopen_counts = Counter()
            badopen_by_gt = defaultdict(Counter)
            for rid, gt_label in gt.items():
                if rid in v6_or_v51:
                    j = v6_or_v51[rid]
                    trigger = j.get('badopen_trigger', 'unknown')
                    badopen_counts[trigger] += 1
                    badopen_by_gt[gt_label][trigger] += 1

            for trigger, count in badopen_counts.most_common():
                discard_count = badopen_by_gt['DISCARD'].get(trigger, 0)
                keep_count = badopen_by_gt['KEEP'].get(trigger, 0)
                precision = discard_count / count if count > 0 else 0
                print(f'  {trigger}: {count} ({discard_count} discards, {keep_count} keeps, precision={precision:.2f})')

            # OE classification distribution
            print(f'\n--- OE Classification Distribution ({v6_label}) ---')
            oe_counts = Counter()
            oe_by_gt = defaultdict(Counter)
            for rid, gt_label in gt.items():
                if rid in v6_or_v51:
                    j = v6_or_v51[rid]
                    oe_class = j.get('oe_classification', 'unknown')
                    oe_counts[oe_class] += 1
                    oe_by_gt[gt_label][oe_class] += 1

            for oe_class, count in oe_counts.most_common():
                discard_count = oe_by_gt['DISCARD'].get(oe_class, 0)
                keep_count = oe_by_gt['KEEP'].get(oe_class, 0)
                discard_rate = discard_count / count if count > 0 else 0
                print(f'  {oe_class}: {count} ({discard_count} discards, {keep_count} keeps, client discard rate={discard_rate:.2f})')

            # Three-component score analysis
            print(f'\n--- Three-Component Score Analysis ({v6_label}) ---')
            component_stats = defaultdict(lambda: {'DISCARD': [], 'KEEP': []})
            for rid, gt_label in gt.items():
                if rid in v6_or_v51:
                    j = v6_or_v51[rid]
                    for comp in ['authenticity_risk', 'quality_discard_risk', 'client_reject_probability']:
                        val = j.get(comp)
                        if val is not None:
                            component_stats[comp][gt_label].append(float(val))

            for comp in ['authenticity_risk', 'quality_discard_risk', 'client_reject_probability']:
                discard_vals = component_stats[comp]['DISCARD']
                keep_vals = component_stats[comp]['KEEP']
                if discard_vals and keep_vals:
                    import statistics
                    print(f'  {comp}:')
                    print(f'    Client discards: mean={statistics.mean(discard_vals):.3f}, median={statistics.median(discard_vals):.3f}')
                    print(f'    Client keeps:    mean={statistics.mean(keep_vals):.3f}, median={statistics.median(keep_vals):.3f}')
                    # Separation = difference of means
                    sep = statistics.mean(discard_vals) - statistics.mean(keep_vals)
                    print(f'    Separation: {sep:.3f} (positive = good discrimination)')

            # Evidence family firing rates: discards vs keeps
            print(f'\n--- Evidence Family Firing Rates (discards vs keeps, {v6_label}) ---')
            family_firing = defaultdict(lambda: {'DISCARD': 0, 'KEEP': 0, 'DISCARD_total': 0, 'KEEP_total': 0})
            for rid, gt_label in gt.items():
                if rid in v6_or_v51:
                    j = v6_or_v51[rid]
                    fired = j.get('evidence_families_fired', [])
                    if isinstance(fired, list):
                        for fam in fired:
                            family_firing[fam][gt_label] += 1
                    family_firing['_total'][gt_label] += 1

            total_discards = family_firing['_total']['DISCARD']
            total_keeps = family_firing['_total']['KEEP']
            for fam in sorted(family_firing.keys()):
                if fam == '_total':
                    continue
                d_count = family_firing[fam]['DISCARD']
                k_count = family_firing[fam]['KEEP']
                d_rate = d_count / total_discards if total_discards > 0 else 0
                k_rate = k_count / total_keeps if total_keeps > 0 else 0
                gap = d_rate - k_rate
                print(f'  {fam}: discards={d_rate:.2f}, keeps={k_rate:.2f}, gap={gap:+.2f}')

            # Stage verdict analysis
            print(f'\n--- Stage Verdict Analysis ({v6_label}) ---')
            stage1_counts = defaultdict(Counter)
            stage2_counts = defaultdict(Counter)
            for rid, gt_label in gt.items():
                if rid in v6_or_v51:
                    j = v6_or_v51[rid]
                    stage1_counts[j.get('stage1_fraud_verdict', 'unknown')][gt_label] += 1
                    stage2_counts[j.get('stage2_quality_verdict', 'unknown')][gt_label] += 1

            print(f'  Stage 1 (Fraud Detection):')
            for verdict in ['pass', 'fail', 'ambiguous']:
                d = stage1_counts[verdict]['DISCARD']
                k = stage1_counts[verdict]['KEEP']
                print(f'    {verdict}: {d} discards, {k} keeps')
            print(f'  Stage 2 (Quality Assessment):')
            for verdict in ['pass', 'fail', 'ambiguous']:
                d = stage2_counts[verdict]['DISCARD']
                k = stage2_counts[verdict]['KEEP']
                print(f'    {verdict}: {d} discards, {k} keeps')

            # ML signal correlation vs actual gap
            print(f'\n--- ML Correlation vs Actual Gap Analysis ---')
            corr_path = Path('/Users/jeremyalston/Perfect/autosurvey/skills/cleaning-survey-quality/evolution/ml-signal-correlation/cross_corpus_signal_correlation.json')
            if corr_path.exists():
                with open(corr_path) as f:
                    corr_data = json.load(f)
                print(f'  Cross-corpus family correlations (13,388 respondents, 11 datasets):')
                for fam, data in sorted(corr_data.get('family_correlations', {}).items(),
                                        key=lambda x: abs(x[1]['mean_correlation']), reverse=True):
                    print(f'    {fam}: mean_corr={data["mean_correlation"]:+.3f} ({data["direction"]}, {data["n_datasets"]} datasets)')
                print(f'  → Families with POSITIVE correlation should be weighted higher in client_reject_probability')
                print(f'  → Families with NEGATIVE correlation (core_oe_quality) should NOT drive discards alone')
        else:
            print(f'  (No v6 metadata fields found in {v6_label} judgments)')

        # Save full comparison
        output = {
            'ground_truth': {'total': len(gt), 'discards': gt_discards, 'keeps': gt_keeps},
            'captain_semantic': captain,
            'v4': v4_metrics,
            'v5': v5_metrics,
            'v5.1': v51_metrics,
            'v6': v6_metrics,
            'v7': v7_metrics,
        }
        output_path = (v7_dir if v7_dir.exists() else (v6_dir if v6_dir.exists() else v51_dir)) / 'comparison_results.json'
        with open(output_path, 'w') as f:
            json.dump(output, f, indent=2)
        print(f'\nFull comparison saved to: {output_path}')

if __name__ == '__main__':
    main()
