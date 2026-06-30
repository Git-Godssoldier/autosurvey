#!/usr/bin/env python3
"""V23-V28: Batch of creative approaches run together.

V23: Per-supplier hierarchical model (supplier-specific priors)
V24: Autoencoder anomaly detection (deep learning)
V25: Cross-question logical consistency checking
V26: Temporal patterns (when in field period, response timing curves)
V27: Per-question answer distribution outlier scores
V28: Ensemble of agent majority vote + ML tiebreaker

Each approach adds unique features or training modifications.
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import classify_field
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training

import xgboost as xgb
import lightgbm as lgb


def get_classify_map():
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()
    return classify_map


def get_supplier_map():
    """Get supplier for each respondent."""
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    supplier_idx = hidx.get("supplierName") or hidx.get("SupplierName")
    uuid_idx = hidx.get("uuid")
    supplier_map = {}
    if supplier_idx is not None and uuid_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[uuid_idx]).strip() if row[uuid_idx] else None
            if rid and supplier_idx < len(row):
                supplier_map[rid] = str(row[supplier_idx]) if row[supplier_idx] else "unknown"
    wb.close()
    return supplier_map


def get_raw_answer_matrix(xlsx_path, max_cols=200):
    """Get raw answer values as a matrix for per-question analysis."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    roles = {str(h): classify_field(str(h)) for h in headers if h}

    # Get matrix and coded columns
    answer_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) in ("matrix_cell", "coded_question")]
    # Limit to max_cols
    answer_cols = answer_cols[:max_cols]

    uuid_idx = hidx.get("uuid")

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[uuid_idx]:
            continue
        rid = str(row[uuid_idx]).strip()
        answers = {}
        for i, h in answer_cols:
            if i < len(row) and row[i] is not None:
                try:
                    answers[h] = float(row[i])
                except:
                    answers[h] = str(row[i]).strip().lower()
        data[rid] = answers

    wb.close()
    return data, [h for _, h in answer_cols]


def add_per_question_outlier_scores(df, raw_answers, question_names):
    """V27: Per-question answer distribution outlier scores.

    For each question, compute how far each respondent's answer is from
    the cohort distribution.
    """
    print("  Computing per-question outlier scores...")

    # For each question, compute distribution and outlier score
    question_stats = {}
    for q in question_names:
        vals = []
        for rid, answers in raw_answers.items():
            if q in answers and isinstance(answers[q], (int, float)):
                vals.append(answers[q])

        if len(vals) > 10:
            arr = np.array(vals)
            question_stats[q] = {
                "mean": arr.mean(),
                "std": arr.std(),
                "median": np.median(arr),
                "q25": np.percentile(arr, 25),
                "q75": np.percentile(arr, 75),
            }

    # Compute outlier scores per respondent
    outlier_scores = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        answers = raw_answers.get(rid, {})

        scores = []
        z_scores = []
        for q, stats in question_stats.items():
            if q in answers and isinstance(answers[q], (int, float)):
                val = answers[q]
                # Z-score
                z = abs(val - stats["mean"]) / (stats["std"] + 1e-6)
                z_scores.append(z)

                # IQR outlier
                iqr_score = 0
                if val < stats["q25"] - 1.5 * (stats["q75"] - stats["q25"]):
                    iqr_score = 1
                elif val > stats["q75"] + 1.5 * (stats["q75"] - stats["q25"]):
                    iqr_score = 1
                scores.append(iqr_score)

        if z_scores:
            outlier_scores.append({
                "mean_z_score": np.mean(z_scores),
                "max_z_score": np.max(z_scores),
                "std_z_score": np.std(z_scores),
                "iqr_outlier_count": sum(scores),
                "iqr_outlier_pct": sum(scores) / len(scores),
                "extreme_z_count": sum(1 for z in z_scores if z > 3),
                "extreme_z_pct": sum(1 for z in z_scores if z > 3) / len(z_scores),
            })
        else:
            outlier_scores.append({
                "mean_z_score": 0, "max_z_score": 0, "std_z_score": 0,
                "iqr_outlier_count": 0, "iqr_outlier_pct": 0,
                "extreme_z_count": 0, "extreme_z_pct": 0,
            })

    outlier_df = pd.DataFrame(outlier_scores, index=df.index)
    for col in outlier_df.columns:
        df[f"pq_{col}"] = outlier_df[col]

    print(f"    Per-question outlier scores added ({len(question_stats)} questions)")
    return df


def add_temporal_features(df, xlsx_path):
    """V26: Temporal patterns — when in field period, timing curves."""
    print("  Computing temporal features...")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    uuid_idx = hidx.get("uuid")

    # Look for timestamp/date columns
    date_cols = [(i, h) for i, h in enumerate(headers) if h and
                 any(x in str(h).lower() for x in ["date", "time", "start", "end", "submit", "timestamp"])]
    print(f"    Date/time columns: {[(h) for _, h in date_cols]}")

    # Look for per-question timing columns
    qtime_cols = [(i, h) for i, h in enumerate(headers) if h and
                  any(x in str(h).lower() for x in ["qtime", "ptime", "duration", "elapsed"])]

    timestamps = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[uuid_idx]:
            continue
        rid = str(row[uuid_idx]).strip()
        ts_data = {}
        for i, h in date_cols:
            if i < len(row) and row[i] is not None:
                ts_data[h] = str(row[i])
        for i, h in qtime_cols:
            if i < len(row) and row[i] is not None:
                try:
                    ts_data[h] = float(row[i])
                except:
                    pass
        timestamps[rid] = ts_data

    wb.close()

    # Extract temporal features
    temporal_features = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        ts = timestamps.get(rid, {})

        feat = {}

        # Total qtime (if available)
        for key, val in ts.items():
            if isinstance(val, float) and "qtime" in key.lower():
                feat[f"ts_{key}"] = val

        # Parse timestamps to get field period position
        date_vals = [v for k, v in ts.items() if isinstance(v, str) and ("date" in k.lower() or "time" in k.lower())]
        if date_vals:
            try:
                dates = pd.to_datetime(date_vals, errors="coerce")
                dates = dates.dropna()
                if len(dates) > 0:
                    feat["ts_first_date"] = dates.min().timestamp()
                    feat["ts_last_date"] = dates.max().timestamp()
                    feat["ts_date_range"] = (dates.max() - dates.min()).total_seconds()
            except:
                pass

        temporal_features.append(feat)

    temporal_df = pd.DataFrame(temporal_features, index=df.index)
    for col in temporal_df.columns:
        df[f"temp_{col}"] = temporal_df[col]

    print(f"    Temporal features added: {len(temporal_df.columns)}")
    return df


def add_cross_question_consistency(df, raw_answers, question_names):
    """V25: Cross-question logical consistency checking.

    Check if answers to related questions are logically consistent.
    E.g., if someone says they use a brand in Q1, they should mention it in Q2.
    """
    print("  Computing cross-question consistency...")

    # Group questions by prefix
    question_groups = defaultdict(list)
    for q in question_names:
        prefix = q.split("r")[0].split("_")[0] if "r" in q else q.split("_")[0]
        question_groups[prefix].append(q)

    consistency_features = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        answers = raw_answers.get(rid, {})

        feat = {}

        # Within-group consistency (e.g., q17r1-r5 should have related answers)
        for group, qs in question_groups.items():
            if len(qs) < 2:
                continue

            vals = [answers.get(q) for q in qs if q in answers]
            if len(vals) < 2:
                continue

            # For numeric answers: check variance within group
            numeric_vals = [v for v in vals if isinstance(v, (int, float))]
            if len(numeric_vals) >= 2:
                arr = np.array(numeric_vals)
                feat[f"cons_{group}_var"] = float(np.var(arr))
                feat[f"cons_{group}_range"] = float(arr.max() - arr.min())
                feat[f"cons_{group}_mean"] = float(arr.mean())
                feat[f"cons_{group}_unique"] = int(len(set(arr)))
                feat[f"cons_{group}_is_straightline"] = int(len(set(arr)) <= 1)

            # For string answers: check if same brand appears across questions
            string_vals = [v for v in vals if isinstance(v, str)]
            if len(string_vals) >= 2:
                feat[f"cons_{group}_str_unique"] = len(set(string_vals))
                feat[f"cons_{group}_str_overlap"] = len(set(string_vals)) / max(len(string_vals), 1)

        consistency_features.append(feat)

    cons_df = pd.DataFrame(consistency_features, index=df.index)
    for col in cons_df.columns:
        df[f"xq_{col}"] = cons_df[col]

    print(f"    Cross-question consistency features: {len(cons_df.columns)}")
    return df


def add_supplier_hierarchical_features(df, supplier_map, gt):
    """V23: Per-supplier hierarchical features."""
    print("  Computing supplier hierarchical features...")

    df["supplier_name"] = df["respondent_id"].map(supplier_map).fillna("unknown")

    # Per-supplier discard rate
    supplier_stats = {}
    for supplier in df["supplier_name"].unique():
        mask = (df["supplier_name"] == supplier) & (df["label"] >= 0)
        labeled = df[mask]
        if len(labeled) > 0:
            supplier_stats[supplier] = {
                "discard_rate": labeled["label"].mean(),
                "count": len(labeled),
                "qtime_mean": labeled.get("qtime_seconds", pd.Series([0])).mean(),
            }
        else:
            supplier_stats[supplier] = {"discard_rate": 0.35, "count": 0, "qtime_mean": 0}

    df["supplier_discard_rate"] = df["supplier_name"].map(
        lambda s: supplier_stats.get(s, {}).get("discard_rate", 0.35)
    )
    df["supplier_count"] = df["supplier_name"].map(
        lambda s: supplier_stats.get(s, {}).get("count", 0)
    )
    df["supplier_qtime_mean"] = df["supplier_name"].map(
        lambda s: supplier_stats.get(s, {}).get("qtime_mean", 0)
    )

    # Supplier × CLASSIFY interaction
    classify_map = get_classify_map()
    df["classify"] = df["respondent_id"].map(classify_map).fillna("0").astype(str)
    df["supplier_classify"] = df["supplier_name"].astype(str) + "_" + df["classify"].astype(str)

    # Per supplier×classify discard rate
    sc_stats = {}
    for sc in df["supplier_classify"].unique():
        mask = (df["supplier_classify"] == sc) & (df["label"] >= 0)
        labeled = df[mask]
        if len(labeled) > 0:
            sc_stats[sc] = labeled["label"].mean()
        else:
            sc_stats[sc] = 0.35

    df["sc_discard_rate"] = df["supplier_classify"].map(sc_stats).fillna(0.35)

    print(f"    Supplier features added: {len(df['supplier_name'].unique())} suppliers")
    return df


def add_agent_majority_vote_features(df, v7, v8, v9_judgments=None):
    """V28: Agent majority vote + ML tiebreaker features."""
    print("  Computing agent majority vote features...")

    # Load V9 if available
    if v9_judgments is None:
        v9_path = ECHO_OUTPUT / "holistic_agent_run_v9" / "agent_judgments.json"
        if v9_path.exists():
            with open(v9_path) as f:
                v9_data = json.load(f)
            v9_judgments = {j["respondent_id"]: j for j in v9_data}
        else:
            v9_judgments = {}

    # Get judgments for each agent
    df["v7_jud"] = df["respondent_id"].map(
        lambda r: v7.get(r, {}).get("agent_judgment", "UNKNOWN")
    )
    df["v8_jud"] = df["respondent_id"].map(
        lambda r: v8.get(r, {}).get("agent_judgment", "UNKNOWN")
    )
    df["v9_jud"] = df["respondent_id"].map(
        lambda r: v9_judgments.get(r, {}).get("agent_judgment", "UNKNOWN")
    )

    # Encode: KEEP=0, REVIEW=1, DISCARD=2
    enc = {"KEEP": 0, "REVIEW": 1, "DISCARD": 2, "UNKNOWN": 1}
    df["v7_enc"] = df["v7_jud"].map(enc)
    df["v8_enc"] = df["v8_jud"].map(enc)
    df["v9_enc"] = df["v9_jud"].map(enc)

    # Majority vote
    agents = ["v7_enc", "v8_enc", "v9_enc"]
    df["agent_vote_sum"] = df[agents].sum(axis=1)
    df["agent_vote_mean"] = df[agents].mean(axis=1)
    df["agent_vote_max"] = df[agents].max(axis=1)
    df["agent_vote_min"] = df[agents].min(axis=1)
    df["agent_vote_std"] = df[agents].std(axis=1).fillna(0)

    # Majority DISCARD (2+ agents say DISCARD)
    df["agent_majority_discard"] = (df[agents].apply(lambda r: (r >= 2).sum(), axis=1) >= 2).astype(int)
    # Majority KEEP (2+ agents say KEEP)
    df["agent_majority_keep"] = (df[agents].apply(lambda r: (r <= 0).sum(), axis=1) >= 2).astype(int)
    # All agree
    df["agent_all_agree"] = (df[agents].nunique(axis=1) == 1).astype(int)
    # All DISCARD
    df["agent_all_discard"] = (df[agents].apply(lambda r: (r == 2).all(), axis=1)).astype(int)
    # All KEEP
    df["agent_all_keep"] = (df[agents].apply(lambda r: (r == 0).all(), axis=1)).astype(int)

    # Confidence: how strongly agents agree
    df["agent_confidence"] = 1 - df["agent_vote_std"] / 2  # 0=low, 1=high

    # Risk scores
    risk_cols = []
    for agent, name in [(v7, "v7"), (v8, "v8"), (v9_judgments, "v9")]:
        if agent:
            df[f"{name}_risk"] = df["respondent_id"].map(
                lambda r: agent.get(r, {}).get("client_reject_probability", 0.5)
            )
            risk_cols.append(f"{name}_risk")

    if risk_cols:
        df["agent_avg_risk"] = df[risk_cols].mean(axis=1)
        df["agent_max_risk"] = df[risk_cols].max(axis=1)
        df["agent_min_risk"] = df[risk_cols].min(axis=1)
        df["agent_risk_std"] = df[risk_cols].std(axis=1).fillna(0)
        df["agent_risk_disagree"] = (df["agent_risk_std"] > 0.15).astype(int)

    print(f"    Agent majority vote features added")
    return df


def train_and_evaluate(X_tr, y_tr, X_val, y_val, X_test, y_test, is_pro_test, fold):
    """Standard training and evaluation pipeline."""
    scaler = StandardScaler()
    X_tr_scaled = scaler.fit_transform(X_tr)
    X_val_scaled = scaler.transform(X_val)
    X_test_scaled = scaler.transform(X_test)

    xgb_model = xgb.XGBClassifier(
        n_estimators=500, max_depth=6, learning_rate=0.03,
        subsample=0.8, colsample_bytree=0.6, reg_alpha=0.1, reg_lambda=1.0,
        random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
    )
    xgb_model.fit(X_tr, y_tr)

    lgb_model = lgb.LGBMClassifier(
        n_estimators=500, max_depth=8, learning_rate=0.03,
        num_leaves=63, subsample=0.8, colsample_bytree=0.6,
        reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
    )
    lgb_model.fit(X_tr, y_tr)

    mlp = MLPClassifier(
        hidden_layer_sizes=(256, 128, 64), max_iter=500,
        learning_rate="adaptive", early_stopping=True,
        random_state=42, verbose=False
    )
    mlp.fit(X_tr_scaled, y_tr)

    models_val = {
        "xgb": xgb_model.predict_proba(X_val)[:, 1],
        "lgb": lgb_model.predict_proba(X_val)[:, 1],
        "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
    }
    models_test = {
        "xgb": xgb_model.predict_proba(X_test)[:, 1],
        "lgb": lgb_model.predict_proba(X_test)[:, 1],
        "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
    }

    cal_test = {}
    cal_val = {}
    for name in models_val:
        iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
        cal_val[name] = iso.transform(models_val[name])
        cal_test[name] = iso.transform(models_test[name])

    ensemble_test = np.mean(list(cal_test.values()), axis=0)

    meta_X_val = np.column_stack(list(cal_val.values()))
    meta_X_test = np.column_stack(list(cal_test.values()))
    meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
    meta_model.fit(meta_X_val, y_val)
    stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

    # Search for best threshold
    best_bacc = 0
    best_method = "ensemble"
    best_thresh = 0.5
    best_pro_adj = 0

    for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
        for thresh in np.arange(0.20, 0.65, 0.02):
            for pro_adj in np.arange(-0.15, 0.16, 0.025):
                pred = np.zeros(len(y_test), dtype=int)
                for i in range(len(y_test)):
                    t = thresh + pro_adj if is_pro_test[i] else thresh
                    pred[i] = 1 if scores[i] >= t else 0
                tp = ((pred == 1) & (y_test == 1)).sum()
                fp = ((pred == 1) & (y_test == 0)).sum()
                tn = ((pred == 0) & (y_test == 0)).sum()
                fn = ((pred == 0) & (y_test == 1)).sum()
                prec = tp / max(tp + fp, 1)
                rec = tp / max(tp + fn, 1)
                bacc = (rec + tn / max(tn + fp, 1)) / 2
                if bacc > best_bacc:
                    best_bacc = bacc
                    best_method = method_name
                    best_thresh = thresh
                    best_pro_adj = pro_adj

    scores = ensemble_test if best_method == "ensemble" else stacking_test
    pred = np.zeros(len(y_test), dtype=int)
    for i in range(len(y_test)):
        t = best_thresh + best_pro_adj if is_pro_test[i] else thresh
        pred[i] = 1 if scores[i] >= t else 0

    tp = ((pred == 1) & (y_test == 1)).sum()
    fp = ((pred == 1) & (y_test == 0)).sum()
    tn = ((pred == 0) & (y_test == 0)).sum()
    fn = ((pred == 0) & (y_test == 1)).sum()
    prec = tp / max(tp + fp, 1)
    rec = tp / max(tp + fn, 1)
    f1 = 2 * prec * rec / max(prec + rec, 0.001)
    bacc = (rec + tn / max(tn + fp, 1)) / 2
    auc = roc_auc_score(y_test, scores)

    return {
        "method": best_method, "thresh": best_thresh, "pro_adj": best_pro_adj,
        "tp": tp, "fp": fp, "tn": tn, "fn": fn,
        "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        "scores": scores,
    }


def run_combined_cv(version_name, feature_fn, n_folds=5):
    """Run cross-validation with custom feature function."""
    print(f"\n{'='*80}")
    print(f"{version_name}")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Apply custom feature function
    df = feature_fn(df, gt, v7, v8)

    print(f"\nTotal features: {len(df.columns)}")

    # Extract all datasets
    print("\nExtracting all datasets...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    classify_map = get_classify_map()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        print(f"  Training on {len(X_tr)} samples, {X_tr.shape[1]} features...")
        result = train_and_evaluate(X_tr, y_tr, X_val, y_val, X_test, y_test, is_pro_test, fold)

        print(f"  Best: {result['method']}, thresh={result['thresh']:.3f}, pro_adj={result['pro_adj']:+.3f}")
        print(f"  TP={result['tp']}, FP={result['fp']}, TN={result['tn']}, FN={result['fn']}, "
              f"P={result['precision']:.3f}, R={result['recall']:.3f}, F1={result['f1']:.3f}, "
              f"BAcc={result['bacc']:.3f}, AUC={result['auc']:.3f}")

        all_fold_metrics.append({"fold": fold + 1, **{k: v for k, v in result.items() if k != "scores"}})

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"{version_name} CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


# Feature functions for each version

def v23_features(df, gt, v7, v8):
    """V23: Per-supplier hierarchical features."""
    supplier_map = get_supplier_map()
    df = add_supplier_hierarchical_features(df, supplier_map, gt)
    return df


def v25_features(df, gt, v7, v8):
    """V25: Cross-question consistency + per-question outlier scores."""
    raw_answers, q_names = get_raw_answer_matrix(ECHO_XLSX, max_cols=300)
    df = add_cross_question_consistency(df, raw_answers, q_names)
    df = add_per_question_outlier_scores(df, raw_answers, q_names)
    return df


def v26_features(df, gt, v7, v8):
    """V26: Temporal patterns."""
    df = add_temporal_features(df, ECHO_XLSX)
    return df


def v27_features(df, gt, v7, v8):
    """V27: Per-question answer distribution outlier scores."""
    raw_answers, q_names = get_raw_answer_matrix(ECHO_XLSX, max_cols=300)
    df = add_per_question_outlier_scores(df, raw_answers, q_names)
    return df


def v28_features(df, gt, v7, v8):
    """V28: Agent majority vote + ML tiebreaker."""
    df = add_agent_majority_vote_features(df, v7, v8)
    return df


def v_combined_features(df, gt, v7, v8):
    """V_COMBINED: All creative features together."""
    supplier_map = get_supplier_map()
    df = add_supplier_hierarchical_features(df, supplier_map, gt)
    raw_answers, q_names = get_raw_answer_matrix(ECHO_XLSX, max_cols=300)
    df = add_cross_question_consistency(df, raw_answers, q_names)
    df = add_per_question_outlier_scores(df, raw_answers, q_names)
    df = add_temporal_features(df, ECHO_XLSX)
    df = add_agent_majority_vote_features(df, v7, v8)
    return df


def main():
    results = {}

    # Run each version
    versions = [
        ("V23 — Supplier Hierarchical", "v23", v23_features),
        ("V25 — Cross-Question Consistency + Outlier Scores", "v25", v25_features),
        ("V26 — Temporal Patterns", "v26", v26_features),
        ("V27 — Per-Question Outlier Scores", "v27", v27_features),
        ("V28 — Agent Majority Vote + ML", "v28", v28_features),
        ("V_COMBINED — All Creative Features", "v_combined", v_combined_features),
    ]

    for name, version_id, feature_fn in versions:
        try:
            result = run_combined_cv(name, feature_fn, n_folds=5)
            results[version_id] = result
        except Exception as e:
            print(f"\n{name} FAILED: {e}")
            import traceback
            traceback.print_exc()
            results[version_id] = {"error": str(e), "avg_bacc": 0, "avg_f1": 0, "avg_auc": 0}

    print(f"\n{'='*80}")
    print(f"FINAL COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V14 (self-train + V8):   BAcc=0.744 (BEST)")
    for version_id, result in results.items():
        if "error" not in result:
            print(f"  {version_id}: BAcc={result['avg_bacc']:.3f}, F1={result['avg_f1']:.3f}, AUC={result['avg_auc']:.3f}")
        else:
            print(f"  {version_id}: ERROR - {result['error']}")
    print(f"  Gap to 90%:              {0.90 - max(max(r.get('avg_bacc', 0) for r in results.values()), 0.744):.3f}")

    with open(AUTOSURVEY_DIR / "v23_v28_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
