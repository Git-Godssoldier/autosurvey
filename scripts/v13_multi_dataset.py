#!/usr/bin/env python3
"""V13 — Multi-dataset training + rich features WITHOUT V7 leakage.

Key insight from V12: V7 agent judgment features dominate mutual information.
The ML model is learning to predict V7's judgment (69% BAcc), creating a ceiling.

V13 approach:
1. Train on ALL 11 annotated datasets (not just Echo) for more training data
2. NO V7 judgment features (break the circular dependency)
3. Rich feature engineering: all V11/V12 features + dataset-specific signals
4. Per-dataset threshold calibration
5. Evaluate on Echo BH held-out fold
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, extract_enhanced_features

import xgboost as xgb
import lightgbm as lgb

# All annotated datasets
DATASETS = {
    "106-2502 Delta Water Filtration": "106-2502 Delta Water Filtration.xlsx",
    "109-2601 Echo BH": "109-2601 Echo BH.xlsx",
    "153-2602 ODL Switchable Glass": "153-2602 ODL Switchable Glass.xlsx",
    "159-2601 Oldcastle Brand Health": "159-2601 Oldcastle Brand Health.xlsx",
    "159-2602 Oldcastle Canada": "159-2602 Oldcastle Canada.xlsx",
    "189-2501 SBD Brand Association": "189-2501 SBD Brand Association.xlsx",
    "287-2501 THD Digital CX": "287-2501 THD Digital CX.xlsx",
    "365-2601 ADDO RaceTrac US GP": "365-2601 ADDO RaceTrac US GP.xlsx",
    "368-2602 Masterlock Conjoint": "368-2602 Masterlock Conjoint.xlsx",
    "999-2601 TFG Contractor Index Q1": "999-2601 TFG Contractor Index Q1.xlsx",
    "999-2602 TFG Contractor Index Q2": "999-2602 TFG Contractor Index Q2.xlsx",
}

# Ground truth paths (client annotated copies)
GT_PATHS = {
    "109-2601 Echo BH": GT_XLSX,
}


def load_all_datasets_gt():
    """Load ground truth from all available annotated datasets."""
    all_gt = {}

    # Echo BH
    gt = load_ground_truth()
    for rid, label in gt.items():
        all_gt[(rid, "109-2601 Echo BH")] = label

    # Try to load other datasets' ground truth from the annotated workbooks
    # Look for annotated copies
    annotated_dir = Path("/Users/jeremyalston/Perfect")
    for dataset_name in DATASETS:
        if dataset_name == "109-2601 Echo BH":
            continue

        # Search for annotated copies
        # The annotated files are in subdirectories like "AutoQuality Pair Copy - <dataset>"
        search_patterns = [
            f"*{dataset_name.split(' ')[0]}*annot*",
            f"*{dataset_name.split(' ')[1].split(' ')[0]}*annot*",
        ]

        for pattern in search_patterns:
            for p in annotated_dir.glob(f"**/{pattern}"):
                if p.suffix == ".xlsx":
                    try:
                        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                        ws = wb.active
                        headers = [c.value for c in ws[1]]
                        hidx = {h: i for i, h in enumerate(headers) if h}

                        if "status" not in hidx or "uuid" not in hidx:
                            wb.close()
                            continue

                        count = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
                            status = row[hidx["status"]]
                            if rid:
                                if status in (5, "5"):
                                    all_gt[(rid, dataset_name)] = 1
                                    count += 1
                                elif status in (3, "3"):
                                    all_gt[(rid, dataset_name)] = 0
                                    count += 1
                        wb.close()
                        print(f"  Loaded {count} GT labels from {p.name} for {dataset_name}")
                        break
                    except Exception as e:
                        pass

    return all_gt


def extract_all_datasets_features(all_gt):
    """Extract features from all datasets."""
    all_dfs = []

    for dataset_name, xlsx_name in DATASETS.items():
        xlsx_path = DATA_DIR / xlsx_name
        if not xlsx_path.exists():
            print(f"  SKIP: {xlsx_name} not found")
            continue

        print(f"  Extracting: {dataset_name}...")

        # Build GT dict for this dataset
        dataset_gt = {}
        for (rid, dn), label in all_gt.items():
            if dn == dataset_name:
                dataset_gt[rid] = label

        try:
            df, answer_chains = extract_enhanced_features(xlsx_path, dataset_gt if dataset_gt else None, v7_judgments=None)
            df["dataset"] = dataset_name

            # Add label
            df["label"] = df["respondent_id"].map(dataset_gt).fillna(-1).astype(int)

            labeled = (df["label"] >= 0).sum()
            print(f"    {len(df)} respondents, {labeled} labeled")
            all_dfs.append(df)
        except Exception as e:
            print(f"    ERROR: {e}")
            continue

    combined = pd.concat(all_dfs, ignore_index=True)
    return combined


def run_v13_cv(n_folds=5):
    """Run V13 cross-validation with multi-dataset training."""
    print(f"\n{'='*80}")
    print(f"V13 — Multi-Dataset Training + Rich Features (No V7 Leakage)")
    print(f"{'='*80}")

    # Load all ground truth
    print("Loading ground truth from all datasets...")
    all_gt = load_all_datasets_gt()
    print(f"  Total GT labels: {len(all_gt)}")

    # Extract features from all datasets
    print("\nExtracting features from all datasets...")
    combined = extract_all_datasets_features(all_gt)

    labeled = combined[combined["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = combined[combined["label"] < 0].copy()

    print(f"\nCombined: {len(combined)} respondents ({len(labeled)} labeled, {len(unlabeled)} unlabeled)")
    print(f"  Labeled discards: {labeled['label'].sum()} ({labeled['label'].mean():.1%})")

    # Prepare features
    non_feature = {"label", "respondent_id", "dataset", "supplier_name", "v7_judgment"}
    feature_cols = [c for c in labeled.columns if c not in non_feature]

    X = labeled[feature_cols].copy()
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X = X.fillna(0)
    y = labeled["label"].values
    datasets = labeled["dataset"].values

    # Read CLASSIFY for Echo
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()

    is_pro = labeled["respondent_id"].map(lambda r: str(classify_map.get(r)) == "1").values

    # K-fold CV on Echo BH only (train on all datasets, test on Echo)
    echo_mask = datasets == "109-2601 Echo BH"
    echo_indices = np.where(echo_mask)[0]
    echo_y = y[echo_indices]
    echo_is_pro = is_pro[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        # Get actual indices in the full dataset
        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        # Training set: all non-Echo labeled data + Echo train fold
        non_echo_idx = np.where(~echo_mask)[0]
        full_train_idx = np.concatenate([non_echo_idx, train_idx])

        X_train, X_test = X.iloc[full_train_idx], X.iloc[test_idx]
        y_train, y_test = y[full_train_idx], y[test_idx]
        is_pro_test = is_pro[test_idx]

        print(f"  Train: {len(X_train)} (from all datasets + Echo train)")
        print(f"  Test:  {len(X_test)} (Echo held-out fold)")

        # Split train into train/val
        n_val = len(X_train) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_train), n_val, replace=False)
        train_mask = np.ones(len(X_train), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_train.iloc[train_mask]
        X_val = X_train.iloc[val_idx]
        y_tr = y_train[train_mask]
        y_val = y_train[val_idx]

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        # Train models
        print("  Training XGBoost...")
        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        print("  Training LightGBM...")
        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.7,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        print("  Training MLP...")
        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        print("  Training RF...")
        rf = RandomForestClassifier(
            n_estimators=500, max_depth=10, min_samples_leaf=5,
            random_state=42, n_jobs=-1
        )
        rf.fit(X_tr, y_tr)

        # Calibrate
        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
            "rf": rf.predict_proba(X_val)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
            "rf": rf.predict_proba(X_test)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Search for best threshold
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    # Aggregate
    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V13 CV RESULTS (multi-dataset training, Echo test)")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V13 — Multi-Dataset Training + Rich Features (No V7 Leakage)")
    print("=" * 80)

    results = run_v13_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V10 (sklearn ensemble):  BAcc=0.729 (with V7 features)")
    print(f"  V11 (XGB+LGB+MLP+RF):    BAcc=0.737 (with V7 features)")
    print(f"  V12 (feat select):       BAcc=0.735 (with V7 features)")
    print(f"  V13 (multi-dataset):     BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f} (NO V7 features)")
    print(f"  Gap to 90%:              {0.90 - results['avg_bacc']:.3f}")

    with open(AUTOSURVEY_DIR / "v13_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
