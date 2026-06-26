#!/usr/bin/env python3
"""Feature extraction + ML pipeline for survey quality scoring.

Extracts ALL features from raw Excel files (356 columns) + client signal map,
then trains a Gradient Boosting model with leave-one-dataset-out cross-validation.

NO target-rate calibration. NO dataset-specific tuning. The model must generalize.
"""
from __future__ import annotations

import csv
import json
import re
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.metrics import (
    accuracy_score,
    classification_report,
    confusion_matrix,
    f1_score,
    precision_score,
    recall_score,
    roc_auc_score,
)

warnings.filterwarnings("ignore")

DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
SIGNAL_MAP = DATA_DIR / "autosurvey-outputs/status-ground-truth-calibration/core_loop_2026-06-24/status_respondent_signal_map.csv"

# Map xlsx filenames to signal map dataset names
DATASET_MAP = {
    "106-2502 Delta Water Filtration.xlsx": "260111_Delta Water Filtration.xlsx",
    "109-2601 Echo BH.xlsx": "260300_ECHO.xlsx",
    "153-2602 ODL Switchable Glass.xlsx": "260501_ODL.xlsx",
    "159-2601 Oldcastle Brand Health.xlsx": "260206_OC BH.xlsx",
    "159-2602 Oldcastle Canada.xlsx": "260401_ OC CAN.xlsx",
    "189-2501 SBD Brand Association.xlsx": "260200_SBD.xlsx",
    "287-2501 THD Digital CX.xlsx": "251101_THD CX.xlsx",
    "365-2601 ADDO RaceTrac US GP.xlsx": "260404_ADDO.xlsx",
    "368-2602 Masterlock Conjoint.xlsx": "260403_Masterlock Conjoint.xlsx",
    "999-2601 TFG Contractor Index Q1.xlsx": "251205_TFG Contractor Index Q1.xlsx",
    "999-2602 TFG Contractor Index Q2.xlsx": "260306_TFG Contractor Index Q2.xlsx",
}

# All client signals seen in the data — we'll one-hot encode these
ALL_SIGNALS = set()
SIGNAL_TIER1 = {
    "termflags_nonzero",
    "long_low_specificity_text",
    "ai_or_overpolished_text_marker",
    "generic_placeholder_open_end",
}
SIGNAL_TIER2 = {
    "rd_searchr3_canada",
    "rd_searchr1_22",
    "rd_searchr1_23",
    "rd_searchr1_20",
    "qtime_under_dataset_p10",
}
# TIER 3 and unclassified signals — we'll still use them as features


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def load_signal_map() -> dict[str, dict[str, dict]]:
    """Load signal map grouped by dataset -> respondent_key -> {signals, decision, signal_count}."""
    by_dataset: dict[str, dict[str, dict]] = defaultdict(dict)
    with open(SIGNAL_MAP) as f:
        for row in csv.DictReader(f):
            ds = row["dataset"]
            rid = row["respondent_key"]
            signals = [s.strip() for s in row["signals"].split(";") if s.strip()]
            for s in signals:
                ALL_SIGNALS.add(s)
            by_dataset[ds][rid] = {
                "signals": signals,
                "decision": row["tfg_decision"],
                "signal_count": int(row["signal_count"]),
                "status": int(row["status"]),
            }
    return by_dataset


def extract_features_from_excel(filepath: Path, signal_map: dict) -> pd.DataFrame:
    """Extract ALL features from a raw Excel file."""
    signal_map_name = DATASET_MAP[filepath.name]
    sm = signal_map.get(signal_map_name, {})

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Build header index
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Identify column groups
    open_text_cols = [(i, h) for i, h in enumerate(headers) if h and (
        str(h).lower().endswith("oe") or str(h).lower() == "outro" or "qcoe" in str(h).lower()
    )]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and "LangAssess" in str(h)]
    rd_search_cols = [(i, h) for i, h in enumerate(headers) if h and str(h).startswith("RD_Search")]
    rd_gettoken_cols = [(i, h) for i, h in enumerate(headers) if h and str(h).startswith("RD_GetToken")]
    flag_cols = [(i, h) for i, h in enumerate(headers) if h and any(
        x in str(h) for x in ["TERMFLAGS", "clientflagsr1", "scrutinyflags", "pasted"]
    )]

    # Identify matrix/grid columns (q3r1-q3r9, q22r1-q22r10, etc.)
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and re.match(r"q\d+r\d+$", str(h))]

    # Identify coded question columns (exclude open_text, lang, rd, flags, technical)
    excluded_prefixes = ("RD_", "LangAssess", "noanswer", "qc5R1_", "conditions",
                         "outroR1_", "qcoe1R1_", "_Pasted", "POSSIBLE", "OWNERSHIP")
    excluded_names = {"qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel",
                      "LangAssessReadEase", "LangAssessNumSen", "LangAssessNumWords",
                      "LangAssessNumSyl", "url", "session", "camp", "bhf", "sfh",
                      "intcode", "record", "uuid", "status", "qtime", "SUPNAME",
                      "ipAddress", "date", "qStateVer", "outro", "outro_Pasted",
                      "CLASSIFY", "CHANNELTRACKING", "RID", "list", "userAgent",
                      "dcua", "start_date", "vlist", "vos", "vbrowser", "vmobiledevice",
                      "vmobileos", "VALIDCLIENT"}
    coded_cols = []
    for i, h in enumerate(headers):
        if not h:
            continue
        hs = str(h)
        if hs in excluded_names:
            continue
        if any(hs.startswith(p) for p in excluded_prefixes):
            continue
        if hs.endswith("oe") or hs.endswith("oth") or hs == "outro":
            continue
        if "LangAssess" in hs or "RD_" in hs:
            continue
        coded_cols.append((i, h))

    # Collect all rows
    rows = []
    respondent_ids = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Get respondent ID from uuid or record
        rid_idx = hidx.get("uuid") or hidx.get("record")
        if rid_idx is None or rid_idx >= len(row):
            continue
        rid = clean(row[rid_idx])
        if not rid:
            continue

        # Check if we have ground truth
        if rid not in sm:
            continue

        respondent_ids.append(rid)
        rows.append(row)

    if not rows:
        return pd.DataFrame()

    # Build feature matrix
    features = []

    for r_idx, (rid, row) in enumerate(zip(respondent_ids, rows)):
        feat = {"respondent_id": rid}

        # === 1. Client signals (one-hot encoded) ===
        sm_entry = sm[rid]
        signals = set(sm_entry["signals"])
        for sig in ALL_SIGNALS:
            feat[f"sig_{sig}"] = 1 if sig in signals else 0
        feat["signal_count"] = sm_entry["signal_count"]
        feat["t1_count"] = len(signals & SIGNAL_TIER1)
        feat["t2_count"] = len(signals & SIGNAL_TIER2)

        # === 2. Timing features ===
        qtime_idx = hidx.get("qtime")
        if qtime_idx is not None and qtime_idx < len(row):
            qtime = row[qtime_idx]
            try:
                qtime = float(qtime) if qtime else 0
            except (ValueError, TypeError):
                qtime = 0
            feat["qtime_seconds"] = qtime
            feat["qtime_minutes"] = qtime / 60.0
            feat["qtime_log"] = np.log1p(qtime) if qtime > 0 else 0
        else:
            feat["qtime_seconds"] = 0
            feat["qtime_minutes"] = 0
            feat["qtime_log"] = 0

        # === 3. LangAssess NLP features ===
        for i, h in lang_cols:
            v = row[i] if i < len(row) else None
            try:
                feat[f"lang_{h}"] = float(v) if v is not None else 0
            except (ValueError, TypeError):
                feat[f"lang_{h}"] = 0

        # === 4. Open-end text features ===
        oe_texts = []
        oe_lengths = []
        oe_word_counts = []
        for i, h in open_text_cols:
            v = row[i] if i < len(row) else None
            text = clean(v)
            if text:
                oe_texts.append(text)
                oe_lengths.append(len(text))
                oe_word_counts.append(len(text.split()))
        feat["oe_count"] = len(oe_texts)
        feat["oe_total_chars"] = sum(oe_lengths)
        feat["oe_max_chars"] = max(oe_lengths) if oe_lengths else 0
        feat["oe_min_chars"] = min(oe_lengths) if oe_lengths else 0
        feat["oe_mean_chars"] = np.mean(oe_lengths) if oe_lengths else 0
        feat["oe_total_words"] = sum(oe_word_counts)
        feat["oe_max_words"] = max(oe_word_counts) if oe_word_counts else 0
        feat["oe_mean_words"] = np.mean(oe_word_counts) if oe_word_counts else 0

        # Open-end quality indicators
        all_oe_text = " ".join(oe_texts).lower()
        feat["oe_has_none"] = 1 if any(w in all_oe_text for w in ["none", "n/a", "na", "nothing", "no opinion"]) else 0
        feat["oe_has_generic"] = 1 if any(
            w in all_oe_text for w in ["good", "fine", "ok", "okay", "nice", "great", "yes", "no"]
        ) and oe_word_counts and max(oe_word_counts) <= 3 else 0
        feat["oe_all_caps"] = 1 if oe_texts and any(t.isupper() and len(t) > 5 for t in oe_texts) else 0
        feat["oe_single_word"] = 1 if oe_word_counts and max(oe_word_counts) <= 1 else 0
        feat["oe_repeated_words"] = 1 if oe_texts and any(
            len(set(t.split())) < len(t.split()) * 0.5 and len(t.split()) > 3
            for t in oe_texts
        ) else 0

        # === 5. Supplier features ===
        sup_idx = hidx.get("SUPNAME")
        if sup_idx is not None and sup_idx < len(row):
            supplier = clean(row[sup_idx])
        else:
            supplier = ""
        feat["supplier_missing"] = 1 if not supplier or supplier == "MISSING" else 0
        feat["supplier_is_none"] = 1 if supplier == "None" or supplier == "" else 0

        # === 6. Matrix/grid straightlining detection ===
        matrix_values = []
        for i, h in matrix_cols:
            v = row[i] if i < len(row) else None
            if v is not None and v != "":
                matrix_values.append(norm(v))
        if matrix_values:
            unique_ratio = len(set(matrix_values)) / len(matrix_values)
            feat["matrix_unique_ratio"] = unique_ratio
            feat["matrix_straightline"] = 1 if unique_ratio <= 0.2 and len(matrix_values) >= 5 else 0
            feat["matrix_near_straightline"] = 1 if unique_ratio <= 0.4 and len(matrix_values) >= 5 else 0
            feat["matrix_count"] = len(matrix_values)
            feat["matrix_unique_count"] = len(set(matrix_values))
            # Most common value frequency
            val_counts = Counter(matrix_values)
            feat["matrix_most_common_freq"] = val_counts.most_common(1)[0][1] / len(matrix_values)
        else:
            feat["matrix_unique_ratio"] = 1.0
            feat["matrix_straightline"] = 0
            feat["matrix_near_straightline"] = 0
            feat["matrix_count"] = 0
            feat["matrix_unique_count"] = 0
            feat["matrix_most_common_freq"] = 0

        # === 7. Coded answer diversity ===
        coded_values = []
        for i, h in coded_cols:
            v = row[i] if i < len(row) else None
            if v is not None and v != "":
                coded_values.append(str(norm(v)))
        feat["coded_count"] = len(coded_values)
        feat["coded_unique_ratio"] = len(set(coded_values)) / len(coded_values) if coded_values else 1.0
        # Check for "don't know" / neutral responses
        dk_count = sum(1 for v in coded_values if any(
            x in v.lower() for x in ["don't know", "dk", "not sure", "no answer", "na"]
        ))
        feat["coded_dk_count"] = dk_count
        feat["coded_dk_ratio"] = dk_count / len(coded_values) if coded_values else 0

        # === 8. RD_Search features ===
        for i, h in rd_search_cols:
            v = row[i] if i < len(row) else None
            try:
                feat[f"rd_{h}"] = float(v) if v is not None else 0
            except (ValueError, TypeError):
                feat[f"rd_{h}"] = 0

        # === 9. Flag columns ===
        for i, h in flag_cols:
            v = row[i] if i < len(row) else None
            try:
                feat[f"flag_{h}"] = float(v) if v is not None else 0
            except (ValueError, TypeError):
                feat[f"flag_{h}"] = 0

        # === 10. Demographic features ===
        demo_cols = ["qstate", "REGION", "age", "qager1", "qGender", "q13", "q12",
                     "qHomeType", "q2", "q1", "qIndustry", "qNumEmployees", "q9",
                     "qEthnicr1", "CLASSIFY", "CHANNELTRACKING"]
        for h in demo_cols:
            i = hidx.get(h)
            if i is not None and i < len(row):
                v = clean(row[i])
                feat[f"demo_{h}"] = v
            else:
                feat[f"demo_{h}"] = ""

        # === 11. Technical features ===
        tech_cols = ["vos", "vbrowser", "vmobiledevice", "vmobileos"]
        for h in tech_cols:
            i = hidx.get(h)
            if i is not None and i < len(row):
                v = clean(row[i])
                feat[f"tech_{h}"] = v
            else:
                feat[f"tech_{h}"] = ""

        # === 12. Cross-respondent duplicate features ===
        # We'll compute these after collecting all rows
        feat["_oe_text"] = " | ".join(oe_texts)

        # === 13. Ground truth ===
        feat["label"] = 1 if sm_entry["decision"] == "rejected" else 0
        feat["dataset"] = signal_map_name

        features.append(feat)

    df = pd.DataFrame(features)

    # === Post-processing: Cross-respondent features ===
    # Duplicate open-end text detection
    oe_counter = Counter()
    for text in df["_oe_text"]:
        if text.strip():
            oe_counter[text.strip().lower()] += 1

    df["oe_duplicate_count"] = df["_oe_text"].apply(
        lambda t: oe_counter.get(t.strip().lower(), 0) if t.strip() else 0
    )
    df["oe_is_duplicate"] = (df["oe_duplicate_count"] > 1).astype(int)

    # Duplicate userAgent detection
    ua_idx = hidx.get("userAgent")
    if ua_idx is not None:
        ua_counter = Counter()
        for row in rows:
            ua = clean(row[ua_idx]) if ua_idx < len(row) else ""
            if ua:
                ua_counter[ua] += 1
        df["ua_duplicate_count"] = [
            ua_counter.get(clean(row[ua_idx]), 0) if ua_idx < len(row) else 0
            for row in rows
        ]
        df["ua_is_duplicate"] = (df["ua_duplicate_count"] > 1).astype(int)
    else:
        df["ua_duplicate_count"] = 0
        df["ua_is_duplicate"] = 0

    # Duplicate IP detection
    ip_idx = hidx.get("ipAddress")
    if ip_idx is not None:
        ip_counter = Counter()
        for row in rows:
            ip = clean(row[ip_idx]) if ip_idx < len(row) else ""
            if ip:
                ip_counter[ip] += 1
        df["ip_duplicate_count"] = [
            ip_counter.get(clean(row[ip_idx]), 0) if ip_idx < len(row) else 0
            for row in rows
        ]
        df["ip_is_duplicate"] = (df["ip_duplicate_count"] > 1).astype(int)
    else:
        df["ip_duplicate_count"] = 0
        df["ip_is_duplicate"] = 0

    # Drop temporary columns
    df = df.drop(columns=["_oe_text"])

    return df


def prepare_features(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.Series]:
    """Prepare features for ML — encode categoricals, drop non-feature columns."""
    non_feature_cols = {"respondent_id", "label", "dataset"}
    feature_cols = [c for c in df.columns if c not in non_feature_cols]

    X = df[feature_cols].copy()
    y = df["label"].copy()

    # Encode categorical columns
    cat_cols = X.select_dtypes(include=["object"]).columns
    for col in cat_cols:
        X[col] = pd.Categorical(X[col]).codes

    # Fill NaN
    X = X.fillna(0)

    return X, y


def leave_one_dataset_out_cv():
    """Run leave-one-dataset-out cross-validation."""
    print("Loading signal map...")
    signal_map = load_signal_map()
    print(f"  Loaded {sum(len(v) for v in signal_map.values())} annotations across {len(signal_map)} datasets")
    print(f"  Total unique signals: {len(ALL_SIGNALS)}")

    print("\nExtracting features from all datasets...")
    all_dfs = []
    for xlsx_name, signal_map_name in DATASET_MAP.items():
        filepath = DATA_DIR / xlsx_name
        if not filepath.exists():
            print(f"  SKIP: {xlsx_name} not found")
            continue
        print(f"  Extracting: {xlsx_name}...")
        df = extract_features_from_excel(filepath, signal_map)
        if df is not None and len(df) > 0:
            print(f"    {len(df)} respondents, {len(df.columns)} features")
            all_dfs.append(df)

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\nCombined: {len(combined)} respondents, {len(combined.columns)} features")
    print(f"  Rejected: {combined['label'].sum()} ({combined['label'].mean():.1%})")
    print(f"  Accepted: {(1-combined['label']).sum()} ({(1-combined['label']).mean():.1%})")

    # Leave-one-dataset-out CV
    datasets = combined["dataset"].unique()
    results = []

    print(f"\n{'='*120}")
    print(f"LEAVE-ONE-DATASET-OUT CROSS-VALIDATION ({len(datasets)} datasets)")
    print(f"{'='*120}")

    for test_ds in datasets:
        train_df = combined[combined["dataset"] != test_ds]
        test_df = combined[combined["dataset"] == test_ds]

        X_train, y_train = prepare_features(train_df)
        X_test, y_test = prepare_features(test_df)

        # Align columns
        for col in X_train.columns:
            if col not in X_test.columns:
                X_test[col] = 0
        for col in X_test.columns:
            if col not in X_train.columns:
                X_train[col] = 0
        X_test = X_test[X_train.columns]

        # Train model
        model = GradientBoostingClassifier(
            n_estimators=200,
            max_depth=5,
            learning_rate=0.1,
            subsample=0.8,
            random_state=42,
        )
        model.fit(X_train, y_train)

        # Predict
        y_pred = model.predict(X_test)
        y_proba = model.predict_proba(X_test)[:, 1]

        # Metrics
        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, zero_division=0)
        rec = recall_score(y_test, y_pred, zero_division=0)
        f1 = f1_score(y_test, y_pred, zero_division=0)
        auc = roc_auc_score(y_test, y_proba) if len(y_test.unique()) > 1 else 0
        discard_rate = y_pred.mean()

        cm = confusion_matrix(y_test, y_pred, labels=[0, 1])
        tn, fp, fn, tp = cm.ravel()

        results.append({
            "dataset": test_ds,
            "n": len(y_test),
            "client_reject_rate": y_test.mean(),
            "agent_discard_rate": discard_rate,
            "accuracy": acc,
            "precision": prec,
            "recall": rec,
            "f1": f1,
            "auc": auc,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
        })

        print(f"\n{test_ds}")
        print(f"  N={len(y_test)}, Client reject: {y_test.mean():.1%}, Agent discard: {discard_rate:.1%}")
        print(f"  Acc={acc:.1%}  Prec={prec:.1%}  Recall={rec:.1%}  F1={f1:.1%}  AUC={auc:.3f}")
        print(f"  TP={tp}  FP={fp}  TN={tn}  FN={fn}")

        # Feature importance for this fold
        if hasattr(model, "feature_importances_"):
            imp = sorted(zip(X_train.columns, model.feature_importances_),
                        key=lambda x: -x[1])[:10]
            print(f"  Top features: {', '.join(f'{n}={v:.3f}' for n, v in imp)}")

    # Aggregate
    print(f"\n{'='*120}")
    print("AGGREGATE RESULTS")
    print(f"{'='*120}")
    total_tp = sum(r["tp"] for r in results)
    total_fp = sum(r["fp"] for r in results)
    total_tn = sum(r["tn"] for r in results)
    total_fn = sum(r["fn"] for r in results)
    total_n = sum(r["n"] for r in results)

    agg_acc = (total_tp + total_tn) / total_n
    agg_prec = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0
    agg_rec = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0
    agg_f1 = 2 * agg_prec * agg_rec / (agg_prec + agg_rec) if (agg_prec + agg_rec) > 0 else 0

    print(f"  Total: {total_n} respondents")
    print(f"  TP={total_tp}  FP={total_fp}  TN={total_tn}  FN={total_fn}")
    print(f"  Accuracy:  {agg_acc:.1%}")
    print(f"  Precision: {agg_prec:.1%}")
    print(f"  Recall:    {agg_rec:.1%}")
    print(f"  F1:        {agg_f1:.1%}")
    print(f"  Discard rate: {(total_tp + total_fp) / total_n:.1%}")
    print(f"  Client reject rate: {(total_tp + total_fn) / total_n:.1%}")

    # Save results
    results_df = pd.DataFrame(results)
    results_df.to_csv(DATA_DIR / "autosurvey-outputs/lodo_cv_results.csv", index=False)
    print(f"\nResults saved to lodo_cv_results.csv")

    # Save combined features for analysis
    combined.to_csv(DATA_DIR / "autosurvey-outputs/all_features.csv", index=False)

    return results


if __name__ == "__main__":
    results = leave_one_dataset_out_cv()
