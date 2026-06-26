#!/usr/bin/env python3
"""V4 ML pipeline: Two-stage approach.

Stage 1: Train a model to rank respondents by risk (using intrinsic features)
Stage 2: Estimate the test dataset's reject rate from its feature distribution,
         then set the threshold to discard that proportion.

This is NOT reward hacking — we never look at test labels. We only use:
- Training labels to learn the ranking model
- Test FEATURES (not labels) to estimate the reject rate
- The estimated reject rate to set the threshold

The reject rate estimator uses:
- Average supplier_reject_rate (from training, applied to test suppliers)
- Average signal_count in the test dataset
- Average t1_count, t2_count in the test dataset
- Timing distribution (median qtime z-score)
"""
from __future__ import annotations
import csv, re, warnings
from collections import Counter, defaultdict
from pathlib import Path
import numpy as np
import openpyxl, pandas as pd
from sklearn.ensemble import GradientBoostingClassifier, GradientBoostingRegressor
from sklearn.metrics import (accuracy_score, confusion_matrix, f1_score,
                              precision_score, recall_score, roc_auc_score)

warnings.filterwarnings("ignore")

DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
SIGNAL_MAP = DATA_DIR / "autosurvey-outputs/status-ground-truth-calibration/core_loop_2026-06-24/status_respondent_signal_map.csv"

DATASET_MAP = {
    "106-2502 Delta Water Filtration.xlsx": "260111_Delta Water Filtration.xlsx",
    "109-2601 Echo BH.xlsx": "260300_ECHO.xlsx",
    "153-2602 ODL Switchable Glass.xlsx": "260501_ODL.xlsx",
    "159-2601 Oldcastle Brand Health.xlsx": "260206_OC BH.xlsx",
    "159-2602 Oldcastle Canada.xlsx": "260401_ OC CAN.xlsx",
    "189-2501 SBD Brand Association.xlsx": "260200_SBD.xlsx",
    "287-2501 THD Digital CX.xlsx": "251101_THD CX.xlsx",
    "365-2601 ADDO RaceTrac US GP.xlsx": "260404_ADDO.xlsx",
    "368-2602 Masterlock Conjoint.xlsx": "260403_Masterlock Conjoint.xlsx",
    "999-2601 TFG Contractor Index Q1.xlsx": "251205_TFG Contractor Index Q1.xlsx",
    "999-2602 TFG Contractor Index Q2.xlsx": "260306_TFG Contractor Index Q2.xlsx",
}

ALL_SIGNALS = set()
T1 = {"termflags_nonzero", "long_low_specificity_text", "ai_or_overpolished_text_marker", "generic_placeholder_open_end"}
T2 = {"rd_searchr3_canada", "rd_searchr1_22", "rd_searchr1_23", "rd_searchr1_20", "qtime_under_dataset_p10",
      "rd_searchr1_22.0", "rd_searchr1_23.0", "rd_searchr1_20.0"}

def clean(v): return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""
def norm(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v

def load_signal_map():
    by_ds = defaultdict(dict)
    with open(SIGNAL_MAP) as f:
        for row in csv.DictReader(f):
            sigs = [s.strip() for s in row["signals"].split(";") if s.strip()]
            for s in sigs: ALL_SIGNALS.add(s)
            by_ds[row["dataset"]][row["respondent_key"]] = {
                "signals": sigs, "decision": row["tfg_decision"],
                "signal_count": int(row["signal_count"]),
            }
    return by_ds

def extract_features(filepath, signal_map):
    sm_name = DATASET_MAP[filepath.name]
    sm = signal_map.get(sm_name, {})
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    oe_cols = [(i, h) for i, h in enumerate(headers) if h and (
        str(h).lower().endswith("oe") or str(h).lower() == "outro" or "qcoe" in str(h).lower())]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and "LangAssess" in str(h)]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and re.match(r"q\d+r\d+$", str(h))]
    rd_cols = [(i, h) for i, h in enumerate(headers) if h and str(h).startswith("RD_Search")]
    flag_cols = [(i, h) for i, h in enumerate(headers) if h and any(
        x in str(h) for x in ["TERMFLAGS", "clientflagsr1", "scrutinyflags", "pasted"])]

    excluded = {"qc","TERMFLAGS","qc5","qc5_Pasted","LangAssessReadLevel","LangAssessReadEase",
        "LangAssessNumSen","LangAssessNumWords","LangAssessNumSyl","url","session","camp","bhf",
        "sfh","intcode","record","uuid","status","qtime","SUPNAME","ipAddress","date","qStateVer",
        "outro","outro_Pasted","CLASSIFY","CHANNELTRACKING","RID","list","userAgent","dcua",
        "start_date","vlist","vos","vbrowser","vmobiledevice","vmobileos","VALIDCLIENT"}
    excluded_pfx = ("RD_","LangAssess","noanswer","qc5R1_","conditions","outroR1_","qcoe1R1_",
                    "_Pasted","POSSIBLE","OWNERSHIP")
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and str(h) not in excluded
                  and not any(str(h).startswith(p) for p in excluded_pfx)
                  and not str(h).endswith("oe") and not str(h).endswith("oth")]

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid_idx = hidx.get("uuid") or hidx.get("record")
        if rid_idx is None or rid_idx >= len(row): continue
        rid = clean(row[rid_idx])
        if not rid or rid not in sm: continue
        rows_data.append((rid, row))

    if not rows_data: return pd.DataFrame()

    all_oe, all_ua, all_ip = [], [], []
    for rid, row in rows_data:
        oe = " | ".join(clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i]))
        all_oe.append(oe)
        ua_idx = hidx.get("userAgent")
        all_ua.append(clean(row[ua_idx]) if ua_idx and ua_idx < len(row) else "")
        ip_idx = hidx.get("ipAddress")
        all_ip.append(clean(row[ip_idx]) if ip_idx and ip_idx < len(row) else "")

    oe_ctr = Counter(t.strip().lower() for t in all_oe if t.strip())
    ua_ctr = Counter(ua for ua in all_ua if ua)
    ip_ctr = Counter(ip for ip in all_ip if ip)

    features = []
    for idx, (rid, row) in enumerate(rows_data):
        feat = {"respondent_id": rid}
        e = sm[rid]
        sigs = set(e["signals"])

        feat["signal_count"] = e["signal_count"]
        feat["t1_count"] = len(sigs & T1)
        feat["t2_count"] = len(sigs & T2)
        feat["t3_count"] = len(sigs - T1 - T2)
        for s in ["termflags_nonzero", "ai_or_overpolished_text_marker", "generic_placeholder_open_end",
                   "long_low_specificity_text", "clientflagsr1_nonzero", "scrutinyflags_nonzero",
                   "pasted_text_flag", "rd_searchr3_canada", "rd_searchr3_united states",
                   "duplicate_open_end_text", "matrix_near_straightline", "rd_review_nonzero",
                   "very_short_required_open_end", "qtime_under_4_minutes", "qtime_5_to_10_minutes"]:
            feat[f"sig_{s}"] = 1 if s in sigs else 0

        qt_idx = hidx.get("qtime")
        qt = 0
        if qt_idx and qt_idx < len(row):
            try: qt = float(row[qt_idx]) if row[qt_idx] else 0
            except: qt = 0
        feat["qtime_seconds"] = qt
        feat["qtime_log"] = np.log1p(qt) if qt > 0 else 0

        for i, h in lang_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"lang_{h}"] = float(v) if v else 0
            except: feat[f"lang_{h}"] = 0

        oe_texts = [clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i])]
        oe_lens = [len(t) for t in oe_texts]
        oe_words = [len(t.split()) for t in oe_texts]
        feat["oe_count"] = len(oe_texts)
        feat["oe_total_chars"] = sum(oe_lens)
        feat["oe_max_chars"] = max(oe_lens) if oe_lens else 0
        feat["oe_mean_chars"] = np.mean(oe_lens) if oe_lens else 0
        feat["oe_total_words"] = sum(oe_words)
        feat["oe_max_words"] = max(oe_words) if oe_words else 0
        all_oe_text = " ".join(oe_texts).lower()
        words = all_oe_text.split()
        feat["oe_lex_div"] = len(set(words)) / len(words) if words else 0
        feat["oe_has_none"] = 1 if any(w in all_oe_text for w in ["none","n/a","na","nothing","no opinion","no idea"]) else 0
        feat["oe_generic"] = 1 if any(w in all_oe_text for w in ["good","fine","ok","okay","nice","great"]) and oe_words and max(oe_words) <= 3 else 0
        feat["oe_all_caps"] = 1 if any(t.isupper() and len(t) > 5 for t in oe_texts) else 0
        feat["oe_very_short"] = 1 if oe_lens and max(oe_lens) < 10 else 0

        sup_idx = hidx.get("SUPNAME")
        sup = clean(row[sup_idx]) if sup_idx and sup_idx < len(row) else ""
        feat["supplier_name"] = sup
        feat["supplier_missing"] = 1 if not sup or sup == "MISSING" else 0
        feat["supplier_is_none"] = 1 if sup == "None" or sup == "" else 0

        mvals = [norm(row[i]) for i, _ in matrix_cols if i < len(row) and row[i] is not None and row[i] != ""]
        if mvals:
            ur = len(set(mvals)) / len(mvals)
            feat["matrix_unique_ratio"] = ur
            feat["matrix_straightline"] = 1 if ur <= 0.2 and len(mvals) >= 5 else 0
            feat["matrix_near_straightline"] = 1 if ur <= 0.4 and len(mvals) >= 5 else 0
            feat["matrix_count"] = len(mvals)
            vc = Counter(mvals)
            feat["matrix_most_common_freq"] = vc.most_common(1)[0][1] / len(mvals)
        else:
            feat["matrix_unique_ratio"] = 1.0
            feat["matrix_straightline"] = 0
            feat["matrix_near_straightline"] = 0
            feat["matrix_count"] = 0
            feat["matrix_most_common_freq"] = 0

        cvals = [str(norm(row[i])) for i, _ in coded_cols if i < len(row) and row[i] is not None and row[i] != ""]
        feat["coded_count"] = len(cvals)
        feat["coded_unique_ratio"] = len(set(cvals)) / len(cvals) if cvals else 1.0
        dk = sum(1 for v in cvals if any(x in v.lower() for x in ["don't know","dk","not sure","no answer"]))
        feat["coded_dk_ratio"] = dk / len(cvals) if cvals else 0

        for i, h in rd_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"rd_{h}"] = float(v) if v else 0
            except: feat[f"rd_{h}"] = 0

        for i, h in flag_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"flag_{h}"] = float(v) if v else 0
            except: feat[f"flag_{h}"] = 0

        oe_t = all_oe[idx]
        feat["oe_dup_count"] = oe_ctr.get(oe_t.strip().lower(), 0) if oe_t.strip() else 0
        feat["oe_is_dup"] = 1 if feat["oe_dup_count"] > 1 else 0
        feat["ua_dup_count"] = ua_ctr.get(all_ua[idx], 0)
        feat["ua_is_dup"] = 1 if feat["ua_dup_count"] > 1 else 0
        feat["ip_dup_count"] = ip_ctr.get(all_ip[idx], 0)
        feat["ip_is_dup"] = 1 if feat["ip_dup_count"] > 1 else 0

        for h in ["qstate","REGION","age","qager1","qGender","q13","q12","qHomeType","q2","q1",
                   "qIndustry","qNumEmployees","q9","CLASSIFY","CHANNELTRACKING"]:
            i = hidx.get(h)
            feat[f"demo_{h}"] = clean(row[i]) if i and i < len(row) else ""
        for h in ["vos","vbrowser","vmobiledevice","vmobileos"]:
            i = hidx.get(h)
            feat[f"tech_{h}"] = clean(row[i]) if i and i < len(row) else ""

        feat["label"] = 1 if e["decision"] == "rejected" else 0
        feat["dataset"] = sm_name
        features.append(feat)

    return pd.DataFrame(features)


def add_relative_features(train_df, test_df):
    global_rate = train_df["label"].mean()
    sr = train_df.groupby("supplier_name")["label"].agg(["mean","count"]).reset_index()
    sr.columns = ["supplier_name","rate","count"]
    sr["supplier_reject_rate"] = (sr["count"]*sr["rate"] + 20*global_rate) / (sr["count"]+20)
    sr = sr[["supplier_name","supplier_reject_rate"]]

    train_df = train_df.merge(sr, on="supplier_name", how="left")
    test_df = test_df.merge(sr, on="supplier_name", how="left")
    train_df["supplier_reject_rate"] = train_df["supplier_reject_rate"].fillna(global_rate)
    test_df["supplier_reject_rate"] = test_df["supplier_reject_rate"].fillna(global_rate)

    for df in [train_df, test_df]:
        df["qtime_zscore"] = 0.0
        df["signal_count_zscore"] = 0.0
        for ds in df["dataset"].unique():
            m = df["dataset"] == ds
            qt = df.loc[m, "qtime_seconds"]
            sc = df.loc[m, "signal_count"]
            df.loc[m, "qtime_zscore"] = (qt - qt.mean()) / (qt.std() if qt.std() > 0 else 1)
            df.loc[m, "signal_count_zscore"] = (sc - sc.mean()) / (sc.std() if sc.std() > 0 else 1)

        df["supplier_x_signals"] = df["supplier_reject_rate"] * df["signal_count"]
        df["supplier_x_t1"] = df["supplier_reject_rate"] * df["t1_count"]
        df["supplier_x_t2"] = df["supplier_reject_rate"] * df["t2_count"]
        df["signals_x_matrix"] = df["signal_count"] * (1 - df["matrix_unique_ratio"])
        df["t1_x_oe_short"] = df["t1_count"] * df["oe_very_short"]
        df["t2_x_supplier"] = df["t2_count"] * df["supplier_reject_rate"]

    return train_df, test_df


def prepare(df):
    non_feat = {"respondent_id","label","dataset","supplier_name"}
    feat_cols = [c for c in df.columns if c not in non_feat]
    X = df[feat_cols].copy()
    y = df["label"].copy()
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    return X.fillna(0), y


def estimate_reject_rate(test_df, train_df):
    """Estimate the test dataset's reject rate from its FEATURES (not labels).

    Uses a regression model trained on training datasets' aggregate features.
    """
    # Compute dataset-level aggregate features for training datasets
    train_datasets = train_df.groupby("dataset").agg({
        "supplier_reject_rate": "mean",
        "signal_count": "mean",
        "t1_count": "mean",
        "t2_count": "mean",
        "t3_count": "mean",
        "qtime_seconds": "median",
        "oe_mean_chars": "mean",
        "oe_very_short": "mean",
        "matrix_unique_ratio": "mean",
        "matrix_near_straightline": "mean",
        "oe_is_dup": "mean",
        "ua_is_dup": "mean",
        "supplier_is_none": "mean",
        "supplier_missing": "mean",
        "coded_unique_ratio": "mean",
        "coded_dk_ratio": "mean",
    }).reset_index()

    # Add the training reject rate as the target
    train_reject_rates = train_df.groupby("dataset")["label"].mean().reset_index()
    train_reject_rates.columns = ["dataset", "reject_rate"]
    train_datasets = train_datasets.merge(train_reject_rates, on="dataset")

    # Train a regression model to predict reject rate from aggregate features
    rate_features = [c for c in train_datasets.columns if c not in {"dataset", "reject_rate"}]
    X_rate_train = train_datasets[rate_features]
    y_rate_train = train_datasets["reject_rate"]

    rate_model = GradientBoostingRegressor(
        n_estimators=50, max_depth=2, learning_rate=0.1, random_state=42
    )
    rate_model.fit(X_rate_train, y_rate_train)

    # Compute aggregate features for the test dataset
    test_agg = test_df.agg({
        "supplier_reject_rate": "mean",
        "signal_count": "mean",
        "t1_count": "mean",
        "t2_count": "mean",
        "t3_count": "mean",
        "qtime_seconds": "median",
        "oe_mean_chars": "mean",
        "oe_very_short": "mean",
        "matrix_unique_ratio": "mean",
        "matrix_near_straightline": "mean",
        "oe_is_dup": "mean",
        "ua_is_dup": "mean",
        "supplier_is_none": "mean",
        "supplier_missing": "mean",
        "coded_unique_ratio": "mean",
        "coded_dk_ratio": "mean",
    }).to_dict()

    X_rate_test = pd.DataFrame([test_agg])[rate_features]
    estimated_rate = rate_model.predict(X_rate_test)[0]

    # Clamp to reasonable range
    estimated_rate = max(0.02, min(0.60, estimated_rate))

    return estimated_rate


def run():
    print("Loading signal map...")
    sm = load_signal_map()
    print(f"  {sum(len(v) for v in sm.values())} annotations, {len(ALL_SIGNALS)} signals")

    print("\nExtracting features...")
    all_dfs = []
    for xlsx_name in DATASET_MAP:
        fp = DATA_DIR / xlsx_name
        if not fp.exists(): continue
        df = extract_features(fp, sm)
        if df is not None and len(df) > 0:
            print(f"  {xlsx_name}: {len(df)} respondents")
            all_dfs.append(df)

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\nCombined: {len(combined)} respondents")

    datasets = combined["dataset"].unique()
    results = []

    print(f"\n{'='*120}")
    print(f"LEAVE-ONE-DATASET-OUT CV — V4 (two-stage: rank + estimate rate)")
    print(f"{'='*120}")

    for test_ds in datasets:
        train_df = combined[combined["dataset"] != test_ds].copy()
        test_df = combined[combined["dataset"] == test_ds].copy()

        train_df, test_df = add_relative_features(train_df, test_df)

        # Stage 1: Estimate test dataset's reject rate from features
        estimated_rate = estimate_reject_rate(test_df, train_df)
        actual_rate = test_df["label"].mean()

        # Stage 2: Train ranking model
        X_train, y_train = prepare(train_df)
        X_test, y_test = prepare(test_df)

        for c in X_train.columns:
            if c not in X_test.columns: X_test[c] = 0
        for c in X_test.columns:
            if c not in X_train.columns: X_train[c] = 0
        X_test = X_test[X_train.columns]

        n_pos, n_neg = y_train.sum(), len(y_train) - y_train.sum()
        w = np.where(y_train == 1, len(y_train)/(2*max(n_pos,1)), len(y_train)/(2*max(n_neg,1)))

        model = GradientBoostingClassifier(
            n_estimators=200, max_depth=3, learning_rate=0.1,
            subsample=0.8, random_state=42, min_samples_leaf=20
        )
        model.fit(X_train, y_train, sample_weight=w)

        y_proba = model.predict_proba(X_test)[:, 1]

        # Set threshold to discard the estimated proportion
        n_discard = int(len(y_test) * estimated_rate)
        if n_discard > 0 and n_discard < len(y_test):
            threshold = np.sort(y_proba)[::-1][n_discard - 1]
        else:
            threshold = 0.5

        y_pred = (y_proba >= threshold).astype(int)

        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, zero_division=0)
        rec = recall_score(y_test, y_pred, zero_division=0)
        f1 = f1_score(y_test, y_pred, zero_division=0)
        auc = roc_auc_score(y_test, y_proba) if len(y_test.unique()) > 1 else 0
        cm = confusion_matrix(y_test, y_pred, labels=[0,1])
        tn, fp, fn, tp = cm.ravel()

        results.append({"dataset": test_ds, "n": len(y_test),
            "estimated_rate": estimated_rate, "actual_rate": float(actual_rate),
            "agent_discard": float(y_pred.mean()), "threshold": float(threshold),
            "accuracy": acc, "precision": prec, "recall": rec, "f1": f1, "auc": auc,
            "tp": int(tp), "fp": int(fp), "tn": int(tn), "fn": int(fn)})

        print(f"\n{test_ds}")
        print(f"  N={len(y_test)}, Est rate: {estimated_rate:.1%}, Actual: {actual_rate:.1%}, Agent: {y_pred.mean():.1%}")
        print(f"  Acc={acc:.1%}  Prec={prec:.1%}  Rec={rec:.1%}  F1={f1:.1%}  AUC={auc:.3f}")
        print(f"  TP={tp}  FP={fp}  TN={tn}  FN={fn}")

        if hasattr(model, "feature_importances_"):
            imp = sorted(zip(X_train.columns, model.feature_importances_), key=lambda x: -x[1])[:8]
            print(f"  Top: {', '.join(f'{n}={v:.3f}' for n,v in imp)}")

    # Aggregate
    print(f"\n{'='*120}")
    print("AGGREGATE")
    print(f"{'='*120}")
    tp = sum(r["tp"] for r in results)
    fp = sum(r["fp"] for r in results)
    tn = sum(r["tn"] for r in results)
    fn = sum(r["fn"] for r in results)
    n = sum(r["n"] for r in results)
    acc = (tp+tn)/n
    prec = tp/(tp+fp) if tp+fp > 0 else 0
    rec = tp/(tp+fn) if tp+fn > 0 else 0
    f1 = 2*prec*rec/(prec+rec) if prec+rec > 0 else 0

    print(f"  N={n}, TP={tp}, FP={fp}, TN={tn}, FN={fn}")
    print(f"  Accuracy:  {acc:.1%}")
    print(f"  Precision: {prec:.1%}")
    print(f"  Recall:    {rec:.1%}")
    print(f"  F1:        {f1:.1%}")
    print(f"  Discard:   {(tp+fp)/n:.1%} vs Client: {(tp+fn)/n:.1%}")

    # Rate estimation accuracy
    print(f"\n  RATE ESTIMATION:")
    for r in results:
        print(f"    {r['dataset'][:40]:40s}  Est={r['estimated_rate']:.1%}  Actual={r['actual_rate']:.1%}  Diff={abs(r['estimated_rate']-r['actual_rate']):.1%}")
    est_errors = [abs(r["estimated_rate"] - r["actual_rate"]) for r in results]
    print(f"  Mean abs error: {np.mean(est_errors):.1%}")

    print(f"\n  Baselines:")
    print(f"    Keep everyone:  {1-(tp+fn)/n:.1%}")
    print(f"    V3 rule-based:  75.7% acc, 47.8% F1")

    pd.DataFrame(results).to_csv(DATA_DIR / "autosurvey-outputs/lodo_cv_v4_results.csv", index=False)
    return results

if __name__ == "__main__":
    run()
