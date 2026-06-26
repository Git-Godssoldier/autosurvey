#!/usr/bin/env python3
"""V6: Accuracy-optimal approach.

Instead of trying to estimate the reject rate (which doesn't work),
use a fixed confidence threshold. Only discard when P(reject) is high enough
that the discard is more likely correct than incorrect.

For accuracy optimization: threshold = accept_rate (≈0.767)
This means: only discard if P(reject) > 76.7%

Also try multiple thresholds and report the accuracy/threshold tradeoff.
"""
from __future__ import annotations
import csv, re, warnings
from collections import Counter, defaultdict
from pathlib import Path
import numpy as np
import openpyxl, pandas as pd
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.metrics import (accuracy_score, confusion_matrix, f1_score,
                              precision_score, recall_score, roc_auc_score)
from sklearn.isotonic import IsotonicRegression

warnings.filterwarnings("ignore")

DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
SIGNAL_MAP = DATA_DIR / "autosurvey-outputs/status-ground-truth-calibration/core_loop_2026-06-24/status_respondent_signal_map.csv"

DATASET_MAP = {
    "106-2502 Delta Water Filtration.xlsx": "260111_Delta Water Filtration.xlsx",
    "109-2601 Echo BH.xlsx": "260300_ECHO.xlsx",
    "153-2602 ODL Switchable Glass.xlsx": "260501_ODL.xlsx",
    "159-2601 Oldcastle Brand Health.xlsx": "260206_OC BH.xlsx",
    "159-2602 Oldcastle Canada.xlsx": "260401_ OC CAN.xlsx",
    "189-2501 SBD Brand Association.xlsx": "260200_SBD.xlsx",
    "287-2501 THD Digital CX.xlsx": "251101_THD CX.xlsx",
    "365-2601 ADDO RaceTrac US GP.xlsx": "260404_ADDO.xlsx",
    "368-2602 Masterlock Conjoint.xlsx": "260403_Masterlock Conjoint.xlsx",
    "999-2601 TFG Contractor Index Q1.xlsx": "251205_TFG Contractor Index Q1.xlsx",
    "999-2602 TFG Contractor Index Q2.xlsx": "260306_TFG Contractor Index Q2.xlsx",
}

ALL_SIGNALS = set()
T1 = {"termflags_nonzero", "long_low_specificity_text", "ai_or_overpolished_text_marker", "generic_placeholder_open_end"}
T2 = {"rd_searchr3_canada", "rd_searchr1_22", "rd_searchr1_23", "rd_searchr1_20", "qtime_under_dataset_p10",
      "rd_searchr1_22.0", "rd_searchr1_23.0", "rd_searchr1_20.0"}

def clean(v): return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""
def norm(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v

def load_signal_map():
    by_ds = defaultdict(dict)
    with open(SIGNAL_MAP) as f:
        for row in csv.DictReader(f):
            sigs = [s.strip() for s in row["signals"].split(";") if s.strip()]
            for s in sigs: ALL_SIGNALS.add(s)
            by_ds[row["dataset"]][row["respondent_key"]] = {
                "signals": sigs, "decision": row["tfg_decision"],
                "signal_count": int(row["signal_count"]),
            }
    return by_ds

def extract_features(filepath, signal_map):
    sm_name = DATASET_MAP[filepath.name]
    sm = signal_map.get(sm_name, {})
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    oe_cols = [(i, h) for i, h in enumerate(headers) if h and (
        str(h).lower().endswith("oe") or str(h).lower() == "outro" or "qcoe" in str(h).lower())]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and "LangAssess" in str(h)]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and re.match(r"q\d+r\d+$", str(h))]
    rd_cols = [(i, h) for i, h in enumerate(headers) if h and str(h).startswith("RD_Search")]
    flag_cols = [(i, h) for i, h in enumerate(headers) if h and any(
        x in str(h) for x in ["TERMFLAGS", "clientflagsr1", "scrutinyflags", "pasted"])]

    excluded = {"qc","TERMFLAGS","qc5","qc5_Pasted","LangAssessReadLevel","LangAssessReadEase",
        "LangAssessNumSen","LangAssessNumWords","LangAssessNumSyl","url","session","camp","bhf",
        "sfh","intcode","record","uuid","status","qtime","SUPNAME","ipAddress","date","qStateVer",
        "outro","outro_Pasted","CLASSIFY","CHANNELTRACKING","RID","list","userAgent","dcua",
        "start_date","vlist","vos","vbrowser","vmobiledevice","vmobileos","VALIDCLIENT"}
    excluded_pfx = ("RD_","LangAssess","noanswer","qc5R1_","conditions","outroR1_","qcoe1R1_",
                    "_Pasted","POSSIBLE","OWNERSHIP")
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and str(h) not in excluded
                  and not any(str(h).startswith(p) for p in excluded_pfx)
                  and not str(h).endswith("oe") and not str(h).endswith("oth")]

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid_idx = hidx.get("uuid") or hidx.get("record")
        if rid_idx is None or rid_idx >= len(row): continue
        rid = clean(row[rid_idx])
        if not rid or rid not in sm: continue
        rows_data.append((rid, row))

    if not rows_data: return pd.DataFrame()

    all_oe, all_ua, all_ip = [], [], []
    for rid, row in rows_data:
        oe = " | ".join(clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i]))
        all_oe.append(oe)
        ua_idx = hidx.get("userAgent")
        all_ua.append(clean(row[ua_idx]) if ua_idx and ua_idx < len(row) else "")
        ip_idx = hidx.get("ipAddress")
        all_ip.append(clean(row[ip_idx]) if ip_idx and ip_idx < len(row) else "")

    oe_ctr = Counter(t.strip().lower() for t in all_oe if t.strip())
    ua_ctr = Counter(ua for ua in all_ua if ua)
    ip_ctr = Counter(ip for ip in all_ip if ip)

    features = []
    for idx, (rid, row) in enumerate(rows_data):
        feat = {"respondent_id": rid}
        e = sm[rid]
        sigs = set(e["signals"])

        feat["signal_count"] = e["signal_count"]
        feat["t1_count"] = len(sigs & T1)
        feat["t2_count"] = len(sigs & T2)
        feat["t3_count"] = len(sigs - T1 - T2)
        for s in ["termflags_nonzero", "ai_or_overpolished_text_marker", "generic_placeholder_open_end",
                   "long_low_specificity_text", "clientflagsr1_nonzero", "scrutinyflags_nonzero",
                   "pasted_text_flag", "rd_searchr3_canada", "rd_searchr3_united states",
                   "duplicate_open_end_text", "matrix_near_straightline", "rd_review_nonzero",
                   "very_short_required_open_end", "qtime_under_4_minutes", "qtime_5_to_10_minutes"]:
            feat[f"sig_{s}"] = 1 if s in sigs else 0

        qt_idx = hidx.get("qtime")
        qt = 0
        if qt_idx and qt_idx < len(row):
            try: qt = float(row[qt_idx]) if row[qt_idx] else 0
            except: qt = 0
        feat["qtime_seconds"] = qt
        feat["qtime_log"] = np.log1p(qt) if qt > 0 else 0

        for i, h in lang_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"lang_{h}"] = float(v) if v else 0
            except: feat[f"lang_{h}"] = 0

        oe_texts = [clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i])]
        oe_lens = [len(t) for t in oe_texts]
        oe_words = [len(t.split()) for t in oe_texts]
        feat["oe_count"] = len(oe_texts)
        feat["oe_total_chars"] = sum(oe_lens)
        feat["oe_max_chars"] = max(oe_lens) if oe_lens else 0
        feat["oe_mean_chars"] = np.mean(oe_lens) if oe_lens else 0
        feat["oe_total_words"] = sum(oe_words)
        feat["oe_max_words"] = max(oe_words) if oe_words else 0
        all_oe_text = " ".join(oe_texts).lower()
        words = all_oe_text.split()
        feat["oe_lex_div"] = len(set(words)) / len(words) if words else 0
        feat["oe_has_none"] = 1 if any(w in all_oe_text for w in ["none","n/a","na","nothing","no opinion","no idea"]) else 0
        feat["oe_generic"] = 1 if any(w in all_oe_text for w in ["good","fine","ok","okay","nice","great"]) and oe_words and max(oe_words) <= 3 else 0
        feat["oe_all_caps"] = 1 if any(t.isupper() and len(t) > 5 for t in oe_texts) else 0
        feat["oe_very_short"] = 1 if oe_lens and max(oe_lens) < 10 else 0

        sup_idx = hidx.get("SUPNAME")
        sup = clean(row[sup_idx]) if sup_idx and sup_idx < len(row) else ""
        feat["supplier_name"] = sup
        feat["supplier_missing"] = 1 if not sup or sup == "MISSING" else 0
        feat["supplier_is_none"] = 1 if sup == "None" or sup == "" else 0

        mvals = [norm(row[i]) for i, _ in matrix_cols if i < len(row) and row[i] is not None and row[i] != ""]
        if mvals:
            ur = len(set(mvals)) / len(mvals)
            feat["matrix_unique_ratio"] = ur
            feat["matrix_straightline"] = 1 if ur <= 0.2 and len(mvals) >= 5 else 0
            feat["matrix_near_straightline"] = 1 if ur <= 0.4 and len(mvals) >= 5 else 0
            feat["matrix_count"] = len(mvals)
            vc = Counter(mvals)
            feat["matrix_most_common_freq"] = vc.most_common(1)[0][1] / len(mvals)
        else:
            feat["matrix_unique_ratio"] = 1.0
            feat["matrix_straightline"] = 0
            feat["matrix_near_straightline"] = 0
            feat["matrix_count"] = 0
            feat["matrix_most_common_freq"] = 0

        cvals = [str(norm(row[i])) for i, _ in coded_cols if i < len(row) and row[i] is not None and row[i] != ""]
        feat["coded_count"] = len(cvals)
        feat["coded_unique_ratio"] = len(set(cvals)) / len(cvals) if cvals else 1.0
        dk = sum(1 for v in cvals if any(x in v.lower() for x in ["don't know","dk","not sure","no answer"]))
        feat["coded_dk_ratio"] = dk / len(cvals) if cvals else 0

        for i, h in rd_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"rd_{h}"] = float(v) if v else 0
            except: feat[f"rd_{h}"] = 0

        for i, h in flag_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"flag_{h}"] = float(v) if v else 0
            except: feat[f"flag_{h}"] = 0

        oe_t = all_oe[idx]
        feat["oe_dup_count"] = oe_ctr.get(oe_t.strip().lower(), 0) if oe_t.strip() else 0
        feat["oe_is_dup"] = 1 if feat["oe_dup_count"] > 1 else 0
        feat["ua_dup_count"] = ua_ctr.get(all_ua[idx], 0)
        feat["ua_is_dup"] = 1 if feat["ua_dup_count"] > 1 else 0
        feat["ip_dup_count"] = ip_ctr.get(all_ip[idx], 0)
        feat["ip_is_dup"] = 1 if feat["ip_dup_count"] > 1 else 0

        for h in ["qstate","REGION","age","qager1","qGender","q13","q12","qHomeType","q2","q1",
                   "qIndustry","qNumEmployees","q9","CLASSIFY","CHANNELTRACKING"]:
            i = hidx.get(h)
            feat[f"demo_{h}"] = clean(row[i]) if i and i < len(row) else ""
        for h in ["vos","vbrowser","vmobiledevice","vmobileos"]:
            i = hidx.get(h)
            feat[f"tech_{h}"] = clean(row[i]) if i and i < len(row) else ""

        feat["label"] = 1 if e["decision"] == "rejected" else 0
        feat["dataset"] = sm_name
        features.append(feat)

    return pd.DataFrame(features)


def add_relative_features(train_df, test_df):
    global_rate = train_df["label"].mean()
    sr = train_df.groupby("supplier_name")["label"].agg(["mean","count"]).reset_index()
    sr.columns = ["supplier_name","rate","count"]
    sr["supplier_reject_rate"] = (sr["count"]*sr["rate"] + 20*global_rate) / (sr["count"]+20)
    sr = sr[["supplier_name","supplier_reject_rate"]]
    train_df = train_df.merge(sr, on="supplier_name", how="left")
    test_df = test_df.merge(sr, on="supplier_name", how="left")
    train_df["supplier_reject_rate"] = train_df["supplier_reject_rate"].fillna(global_rate)
    test_df["supplier_reject_rate"] = test_df["supplier_reject_rate"].fillna(global_rate)

    for df in [train_df, test_df]:
        df["qtime_zscore"] = 0.0
        df["signal_count_zscore"] = 0.0
        for ds in df["dataset"].unique():
            m = df["dataset"] == ds
            qt = df.loc[m, "qtime_seconds"]
            sc = df.loc[m, "signal_count"]
            df.loc[m, "qtime_zscore"] = (qt - qt.mean()) / (qt.std() if qt.std() > 0 else 1)
            df.loc[m, "signal_count_zscore"] = (sc - sc.mean()) / (sc.std() if sc.std() > 0 else 1)
        df["supplier_x_signals"] = df["supplier_reject_rate"] * df["signal_count"]
        df["supplier_x_t1"] = df["supplier_reject_rate"] * df["t1_count"]
        df["supplier_x_t2"] = df["supplier_reject_rate"] * df["t2_count"]
        df["signals_x_matrix"] = df["signal_count"] * (1 - df["matrix_unique_ratio"])
        df["t1_x_oe_short"] = df["t1_count"] * df["oe_very_short"]
        df["t2_x_supplier"] = df["t2_count"] * df["supplier_reject_rate"]
    return train_df, test_df


def prepare(df):
    non_feat = {"respondent_id","label","dataset","supplier_name"}
    feat_cols = [c for c in df.columns if c not in non_feat]
    X = df[feat_cols].copy()
    y = df["label"].copy()
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    return X.fillna(0), y


def run():
    print("Loading signal map...")
    sm = load_signal_map()

    print("Extracting features...")
    all_dfs = []
    for xlsx_name in DATASET_MAP:
        fp = DATA_DIR / xlsx_name
        if not fp.exists(): continue
        df = extract_features(fp, sm)
        if df is not None and len(df) > 0:
            all_dfs.append(df)
    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"Combined: {len(combined)} respondents")

    datasets = combined["dataset"].unique()

    # Try multiple fixed thresholds
    thresholds = [0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90]

    print(f"\n{'='*120}")
    print(f"THRESHOLD SWEEP — V6 (fixed confidence threshold, calibrated probs)")
    print(f"{'='*120}")

    all_results = {}

    for threshold in thresholds:
        results = []
        for test_ds in datasets:
            train_df = combined[combined["dataset"] != test_ds].copy()
            test_df = combined[combined["dataset"] == test_ds].copy()
            train_df, test_df = add_relative_features(train_df, test_df)

            X_train, y_train = prepare(train_df)
            X_test, y_test = prepare(test_df)

            for c in X_train.columns:
                if c not in X_test.columns: X_test[c] = 0
            for c in X_test.columns:
                if c not in X_train.columns: X_train[c] = 0
            X_test = X_test[X_train.columns]

            n_pos, n_neg = y_train.sum(), len(y_train) - y_train.sum()
            w = np.where(y_train == 1, len(y_train)/(2*max(n_pos,1)), len(y_train)/(2*max(n_neg,1)))

            model = GradientBoostingClassifier(
                n_estimators=200, max_depth=3, learning_rate=0.1,
                subsample=0.8, random_state=42, min_samples_leaf=20
            )
            model.fit(X_train, y_train, sample_weight=w)

            y_train_proba = model.predict_proba(X_train)[:, 1]
            y_test_proba = model.predict_proba(X_test)[:, 1]

            # Calibrate
            iso = IsotonicRegression(out_of_bounds='clip')
            iso.fit(y_train_proba, y_train)
            y_test_cal = iso.transform(y_test_proba)

            y_pred = (y_test_cal >= threshold).astype(int)

            acc = accuracy_score(y_test, y_pred)
            prec = precision_score(y_test, y_pred, zero_division=0)
            rec = recall_score(y_test, y_pred, zero_division=0)
            f1 = f1_score(y_test, y_pred, zero_division=0)
            cm = confusion_matrix(y_test, y_pred, labels=[0,1])
            tn, fp, fn, tp = cm.ravel()

            results.append({"tp": tp, "fp": fp, "tn": tn, "fn": fn, "n": len(y_test),
                            "acc": acc, "prec": prec, "rec": rec, "f1": f1})

        tp = sum(r["tp"] for r in results)
        fp = sum(r["fp"] for r in results)
        tn = sum(r["tn"] for r in results)
        fn = sum(r["fn"] for r in results)
        n = sum(r["n"] for r in results)
        acc = (tp+tn)/n
        prec = tp/(tp+fp) if tp+fp > 0 else 0
        rec = tp/(tp+fn) if tp+fn > 0 else 0
        f1 = 2*prec*rec/(prec+rec) if prec+rec > 0 else 0
        discard = (tp+fp)/n

        all_results[threshold] = {"acc": acc, "prec": prec, "rec": rec, "f1": f1,
                                   "discard": discard, "tp": tp, "fp": fp, "tn": tn, "fn": fn}
        print(f"  Threshold={threshold:.2f}: Acc={acc:.1%}  Prec={prec:.1%}  Rec={rec:.1%}  F1={f1:.1%}  Discard={discard:.1%}  TP={tp}  FP={fp}  TN={tn}  FN={fn}")

    # Find best threshold for accuracy
    best_t = max(all_results, key=lambda t: all_results[t]["acc"])
    best = all_results[best_t]
    print(f"\n  Best threshold for accuracy: {best_t:.2f}")
    print(f"    Accuracy:  {best['acc']:.1%}")
    print(f"    Precision: {best['prec']:.1%}")
    print(f"    Recall:    {best['rec']:.1%}")
    print(f"    F1:        {best['f1']:.1%}")
    print(f"    Discard:   {best['discard']:.1%}")
    print(f"\n  Baselines: Keep everyone=76.7%, V3 rule-based=75.7%/47.8%")

    # Also try: use training data to find the accuracy-optimal threshold
    # (instead of sweeping on test data, which is cheating)
    print(f"\n{'='*120}")
    print(f"TRAINING-OPTIMAL THRESHOLD (not cheating)")
    print(f"{'='*120}")

    # For each LODO fold, find the accuracy-optimal threshold on TRAINING data
    # then apply it to test data
    results = []
    for test_ds in datasets:
        train_df = combined[combined["dataset"] != test_ds].copy()
        test_df = combined[combined["dataset"] == test_ds].copy()
        train_df, test_df = add_relative_features(train_df, test_df)

        X_train, y_train = prepare(train_df)
        X_test, y_test = prepare(test_df)

        for c in X_train.columns:
            if c not in X_test.columns: X_test[c] = 0
        for c in X_test.columns:
            if c not in X_train.columns: X_train[c] = 0
        X_test = X_test[X_train.columns]

        n_pos, n_neg = y_train.sum(), len(y_train) - y_train.sum()
        w = np.where(y_train == 1, len(y_train)/(2*max(n_pos,1)), len(y_train)/(2*max(n_neg,1)))

        model = GradientBoostingClassifier(
            n_estimators=200, max_depth=3, learning_rate=0.1,
            subsample=0.8, random_state=42, min_samples_leaf=20
        )
        model.fit(X_train, y_train, sample_weight=w)

        y_train_proba = model.predict_proba(X_train)[:, 1]
        y_test_proba = model.predict_proba(X_test)[:, 1]

        # Calibrate on training data
        iso = IsotonicRegression(out_of_bounds='clip')
        iso.fit(y_train_proba, y_train)
        y_train_cal = iso.transform(y_train_proba)
        y_test_cal = iso.transform(y_test_proba)

        # Find accuracy-optimal threshold on TRAINING data
        best_train_acc = 0
        best_t = 0.5
        for t in np.linspace(0.01, 0.99, 200):
            pred = (y_train_cal >= t).astype(int)
            acc = accuracy_score(y_train, pred)
            if acc > best_train_acc:
                best_train_acc = acc
                best_t = t

        y_pred = (y_test_cal >= best_t).astype(int)

        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, zero_division=0)
        rec = recall_score(y_test, y_pred, zero_division=0)
        f1 = f1_score(y_test, y_pred, zero_division=0)
        auc = roc_auc_score(y_test, y_test_proba) if len(y_test.unique()) > 1 else 0
        cm = confusion_matrix(y_test, y_pred, labels=[0,1])
        tn, fp, fn, tp = cm.ravel()

        results.append({"dataset": test_ds, "n": len(y_test), "threshold": best_t,
            "accuracy": acc, "precision": prec, "recall": rec, "f1": f1, "auc": auc,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn})

        print(f"\n{test_ds}")
        print(f"  N={len(y_test)}, Thresh: {best_t:.3f}, Acc={acc:.1%}  Prec={prec:.1%}  Rec={rec:.1%}  F1={f1:.1%}  AUC={auc:.3f}")
        print(f"  TP={tp}  FP={fp}  TN={tn}  FN={fn}")

    tp = sum(r["tp"] for r in results)
    fp = sum(r["fp"] for r in results)
    tn = sum(r["tn"] for r in results)
    fn = sum(r["fn"] for r in results)
    n = sum(r["n"] for r in results)
    acc = (tp+tn)/n
    prec = tp/(tp+fp) if tp+fp > 0 else 0
    rec = tp/(tp+fn) if tp+fn > 0 else 0
    f1 = 2*prec*rec/(prec+rec) if prec+rec > 0 else 0

    print(f"\n{'='*120}")
    print(f"AGGREGATE (training-optimal threshold)")
    print(f"{'='*120}")
    print(f"  N={n}, TP={tp}, FP={fp}, TN={tn}, FN={fn}")
    print(f"  Accuracy:  {acc:.1%}")
    print(f"  Precision: {prec:.1%}")
    print(f"  Recall:    {rec:.1%}")
    print(f"  F1:        {f1:.1%}")
    print(f"  Discard:   {(tp+fp)/n:.1%}")
    print(f"\n  Baselines: Keep everyone=76.7%, V3 rule-based=75.7%/47.8%")

    pd.DataFrame(results).to_csv(DATA_DIR / "autosurvey-outputs/lodo_cv_v6_results.csv", index=False)
    return results

if __name__ == "__main__":
    run()
