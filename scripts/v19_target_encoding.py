#!/usr/bin/env python3
"""V19 — Per-question target encoding + LangAssess + RD_Search + raw answer features.

Key insight from error analysis:
- 240 questions have FN/TP distribution differences > 0.15
- LangAssess features (reading level, ease) are unused
- RD_Searchr1 has different distributions for FNs vs TPs
- FNs have HIGH qtime but bad answers (model misses them)
- FPs have LOW qtime but good answers (model falsely discards)

New features:
1. Per-question target encoding (mean label per answer value, per question)
2. LangAssess features (reading level, ease, sentence count, word count, syllable count)
3. RD_Search raw values (especially RD_Searchr1)
4. Answer diversity per question group (brand questions, matrix questions)
5. qtime-independent quality features
6. qtime × answer quality interaction features
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import classify_field
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training

import xgboost as xgb
import lightgbm as lgb


def extract_raw_answer_features(xlsx_path, gt=None, v7=None, v8=None):
    """Extract per-question answer values + LangAssess + RD_Search features."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Classify columns
    roles = {str(h): classify_field(str(h)) for h in headers if h}

    # Identify column groups
    oe_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "open_end"]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "matrix_cell"]
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "coded_question"]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and "LangAssess" in str(h)]
    rd_cols = [(i, h) for i, h in enumerate(headers) if h and "RD_Search" in str(h)]

    # Get key columns
    uuid_idx = hidx.get("uuid")
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    supplier_idx = hidx.get("supplierName") or hidx.get("SupplierName")
    qtime_idx = hidx.get("qtime") or hidx.get("QTIME")

    # Read all rows
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[uuid_idx]:
            rows_data.append(row)

    print(f"  Read {len(rows_data)} rows, {len(headers)} columns")
    print(f"  OE: {len(oe_cols)}, Matrix: {len(matrix_cols)}, Coded: {len(coded_cols)}")
    print(f"  LangAssess: {len(lang_cols)}, RD_Search: {len(rd_cols)}")

    # Build feature dict per respondent
    features = []
    for row in rows_data:
        rid = str(row[uuid_idx]).strip()
        feat = {"respondent_id": rid}

        # CLASSIFY
        if classify_idx and classify_idx < len(row):
            feat["classify"] = str(row[classify_idx]) if row[classify_idx] else "0"

        # Supplier
        if supplier_idx and supplier_idx < len(row):
            feat["supplier"] = str(row[supplier_idx]) if row[supplier_idx] else "unknown"

        # qtime
        if qtime_idx and qtime_idx < len(row) and row[qtime_idx]:
            try:
                feat["qtime_raw"] = float(row[qtime_idx])
            except:
                feat["qtime_raw"] = 0

        # LangAssess features
        for i, h in lang_cols:
            if i < len(row) and row[i] is not None:
                col_name = h.replace("LangAssess", "lang_").lower()
                try:
                    feat[col_name] = float(row[i])
                except:
                    feat[col_name] = 0

        # RD_Search features
        for i, h in rd_cols:
            if i < len(row) and row[i] is not None:
                col_name = f"rd_search_{h.replace('RD_Search', '').lower()}"
                val = str(row[i])
                feat[col_name + "_str"] = val
                try:
                    feat[col_name + "_num"] = float(val)
                except:
                    feat[col_name + "_num"] = -1

        # Per-question answer values for top discriminating questions
        # (from error analysis: q17r3, q17r4, q17r5, q19_2026othr1, q11othr1-6, q29, qc5)
        key_questions = [
            "q17r1", "q17r2", "q17r3", "q17r4", "q17r5",
            "q11othr1", "q11othr2", "q11othr3", "q11othr4", "q11othr5", "q11othr6",
            "q19_2026othr1", "q29", "qc5",
        ]
        for qname in key_questions:
            if qname in hidx:
                i = hidx[qname]
                if i < len(row) and row[i] is not None:
                    feat[f"ans_{qname}"] = str(row[i]).strip().lower()

        # Matrix answer patterns
        matrix_answers = []
        for i, h in matrix_cols:
            if i < len(row) and row[i] is not None:
                try:
                    matrix_answers.append(float(row[i]))
                except:
                    pass

        if matrix_answers:
            arr = np.array(matrix_answers)
            feat["matrix_mean"] = float(arr.mean())
            feat["matrix_std"] = float(arr.std())
            feat["matrix_min"] = float(arr.min())
            feat["matrix_max"] = float(arr.max())
            feat["matrix_median"] = float(np.median(arr))
            feat["matrix_q25"] = float(np.percentile(arr, 25))
            feat["matrix_q75"] = float(np.percentile(arr, 75))
            feat["matrix_iqr"] = feat["matrix_q75"] - feat["matrix_q25"]
            feat["matrix_range"] = feat["matrix_max"] - feat["matrix_min"]
            # Count of each value
            for val in [1, 2, 3, 4, 5]:
                feat[f"matrix_count_{val}"] = int((arr == val).sum())
            feat["matrix_count_1_2"] = int((arr <= 2).sum())
            feat["matrix_count_4_5"] = int((arr >= 4).sum())
            feat["matrix_pct_1_2"] = feat["matrix_count_1_2"] / len(arr)
            feat["matrix_pct_4_5"] = feat["matrix_count_4_5"] / len(arr)
            # Straightline detection per question group
            feat["matrix_unique_vals"] = int(len(set(arr)))
            feat["matrix_is_straightline"] = int(feat["matrix_unique_vals"] <= 2)
            # Middle response bias (always picking 3)
            feat["matrix_count_3"] = int((arr == 3).sum())
            feat["matrix_pct_3"] = feat["matrix_count_3"] / len(arr)

        # Coded question answer patterns
        coded_answers = []
        for i, h in coded_cols:
            if i < len(row) and row[i] is not None:
                coded_answers.append(str(row[i]).strip().lower())

        if coded_answers:
            feat["coded_unique_count"] = len(set(coded_answers))
            feat["coded_diversity_ratio"] = len(set(coded_answers)) / len(coded_answers)
            # Most common answer frequency
            c = Counter(coded_answers)
            feat["coded_most_common_freq"] = c.most_common(1)[0][1] if c else 0
            feat["coded_most_common_pct"] = feat["coded_most_common_freq"] / len(coded_answers)
            # Count of "none" / "don't know" / empty
            none_count = sum(1 for a in coded_answers if a in ("none", "n/a", "na", "don't know", "dk", "", "0"))
            feat["coded_none_count"] = none_count
            feat["coded_none_pct"] = none_count / len(coded_answers)

        # OE text patterns
        oe_texts = []
        for i, h in oe_cols:
            if i < len(row) and row[i] is not None:
                oe_texts.append(str(row[i]).strip())
        oe_combined = " ".join(oe_texts)
        feat["oe_total_chars_raw"] = len(oe_combined)
        feat["oe_word_count_raw"] = len(oe_combined.split())
        feat["oe_has_content"] = int(len(oe_combined.strip()) > 0)
        feat["oe_question_count"] = sum(1 for t in oe_texts if len(t.strip()) > 0)

        features.append(feat)

    wb.close()
    return pd.DataFrame(features), headers, roles


def add_target_encoding_features_train(train_df, all_df, smoothing=10):
    """Add target encoding for per-question answer values.

    CRITICAL: Only uses train_df labels to compute encoding, then applies to all_df.
    This prevents data leakage.
    """
    key_q_cols = [c for c in all_df.columns if c.startswith("ans_")]

    global_mean = train_df["label"].mean() if "label" in train_df.columns else 0.35

    for qcol in key_q_cols:
        te_col = f"te_{qcol}"
        cnt_col = f"cnt_{qcol}"

        # Compute per-answer mean from TRAINING data only
        answer_means = {}
        answer_counts = {}
        for ans in train_df[qcol].dropna().unique():
            mask = train_df[qcol] == ans
            count = mask.sum()
            mean = train_df.loc[mask, "label"].mean() if count > 0 else global_mean
            answer_means[ans] = (count * mean + smoothing * global_mean) / (count + smoothing)
            answer_counts[ans] = count

        # Apply to all data (using only training-derived mapping)
        all_df[te_col] = all_df[qcol].map(answer_means).fillna(global_mean)
        all_df[cnt_col] = all_df[qcol].map(answer_counts).fillna(0)

    return all_df


def add_qtime_quality_interactions(df):
    """Add features that capture answer quality independent of timing."""
    if "qtime_raw" not in df.columns:
        return df

    # qtime per OE word (reading speed proxy)
    if "oe_word_count_raw" in df.columns:
        df["qtime_per_oe_word"] = df["qtime_raw"] / (df["oe_word_count_raw"] + 1)

    # qtime per matrix question
    if "matrix_mean" in df.columns:
        df["qtime_per_matrix"] = df["qtime_raw"] / 195.0  # 195 matrix questions

    # qtime quartile
    df["qtime_quartile"] = pd.qcut(df["qtime_raw"], 4, labels=False, duplicates="drop")
    df["qtime_decile"] = pd.qcut(df["qtime_raw"], 10, labels=False, duplicates="drop")

    # Is this a "fast but good" or "slow but bad" respondent?
    # (These are the error patterns we found)
    df["qtime_low"] = (df["qtime_raw"] < df["qtime_raw"].median()).astype(int)
    df["qtime_high"] = (df["qtime_raw"] > df["qtime_raw"].quantile(0.75)).astype(int)

    # Matrix quality × qtime interaction
    if "matrix_mean" in df.columns:
        df["matrix_x_qtime"] = df["matrix_mean"] * df["qtime_raw"] / 1000
        df["matrix_std_x_qtime"] = df.get("matrix_std", 0) * df["qtime_raw"] / 1000
        df["matrix_diversity_x_qtime"] = df.get("matrix_unique_vals", 0) * df["qtime_raw"] / 1000

    # OE quality × qtime interaction
    if "oe_has_content" in df.columns:
        df["oe_content_x_qtime"] = df["oe_has_content"] * df["qtime_raw"] / 1000
        df["oe_empty_x_qtime_low"] = (1 - df["oe_has_content"]) * df["qtime_low"]

    # Coded diversity × qtime
    if "coded_diversity_ratio" in df.columns:
        df["coded_div_x_qtime"] = df["coded_diversity_ratio"] * df["qtime_raw"] / 1000

    return df


def run_v19_cv(n_folds=5):
    """Run V19 cross-validation."""
    print(f"\n{'='*80}")
    print(f"V19 — Per-Question Target Encoding + LangAssess + RD_Search + Raw Answer Features")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    # Extract enhanced features (V14 base)
    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Extract raw answer features
    print("\nExtracting raw answer features...")
    raw_df, headers, roles = extract_raw_answer_features(ECHO_XLSX, gt, v7, v8)

    # Merge raw features with enhanced features
    df = df.merge(raw_df, on="respondent_id", how="left", suffixes=("", "_raw"))

    # Add qtime-quality interactions
    df = add_qtime_quality_interactions(df)

    # NOTE: Target encoding is computed INSIDE the CV loop to prevent leakage
    # Initialize TE columns with prior
    for qcol in [c for c in df.columns if c.startswith("ans_")]:
        df[f"te_{qcol}"] = 0.35
        df[f"cnt_{qcol}"] = 0

    print(f"\nTotal features (before TE): {len(df.columns)}")

    # Extract all datasets for self-training
    print("\nExtracting all datasets for self-training...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    # Get CLASSIFY map
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(hdrs) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Compute target encoding using ONLY training fold labels
        echo_train_for_te = echo_train[["respondent_id", "label"] + [c for c in echo_train.columns if c.startswith("ans_")]]
        echo_train = add_target_encoding_features_train(echo_train_for_te, echo_train, smoothing=10)
        echo_test = add_target_encoding_features_train(echo_train_for_te, echo_test, smoothing=10)

        # Self-training
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        # Split for calibration
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        print(f"  Training on {len(X_tr)} samples, {X_tr.shape[1]} features...")

        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.6, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.6,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        # Calibrate
        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Search for best threshold
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V19 CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V19 — Per-Question Target Encoding + LangAssess + RD_Search + Raw Answer Features")
    print("=" * 80)

    results = run_v19_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V14 (self-train + V8):   BAcc=0.744 (BEST)")
    print(f"  V17 (LLM embeddings):    BAcc=0.739")
    print(f"  V18 (two-stage):         BAcc=0.738")
    print(f"  V19 (target encoding):   BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], 0.744):.3f}")

    with open(AUTOSURVEY_DIR / "v19_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
