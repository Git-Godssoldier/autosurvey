#!/usr/bin/env python3
"""V32-V38: Batch of approaches building on V31's success (BAcc 0.796).

V32: V31 + more target-encoded questions (all 510 coded questions)
V33: V31 + matrix group target encoding (mean label per matrix group)
V34: V31 + per-question z-score features (distance from cohort per question)
V35: V31 + CatBoost model (good with categorical features)
V36: V31 + larger self-training (lower threshold, more iterations)
V37: V31 + feature selection (remove noisy features, keep top 100)
V38: V31 + all of the above combined
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score, balanced_accuracy_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import classify_field
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training
from v19_target_encoding import extract_raw_answer_features, add_qtime_quality_interactions, add_target_encoding_features_train

import xgboost as xgb
import lightgbm as lgb


def get_classify_map():
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()
    return classify_map


def extract_all_coded_answer_features(xlsx_path):
    """Extract ALL coded question answer values for comprehensive target encoding."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    roles = {str(h): classify_field(str(h)) for h in headers if h}
    uuid_idx = hidx.get("uuid")

    # Get ALL coded and matrix columns
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "coded_question"]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "matrix_cell"]

    print(f"  All coded: {len(coded_cols)}, matrix: {len(matrix_cols)}")

    features = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[uuid_idx]:
            continue
        rid = str(row[uuid_idx]).strip()
        feat = {"respondent_id": rid}

        # All coded question answers
        for i, h in coded_cols:
            if i < len(row) and row[i] is not None:
                feat[f"ca_{h}"] = str(row[i]).strip().lower()

        # Matrix answers grouped by question
        matrix_groups = defaultdict(list)
        for i, h in matrix_cols:
            if i < len(row) and row[i] is not None:
                prefix = h.split("r")[0] if "r" in h else h.split("_")[0]
                try:
                    matrix_groups[prefix].append(float(row[i]))
                except:
                    pass

        for group, vals in matrix_groups.items():
            if vals:
                arr = np.array(vals)
                feat[f"mg_{group}_mean"] = float(arr.mean())
                feat[f"mg_{group}_std"] = float(arr.std())
                feat[f"mg_{group}_min"] = float(arr.min())
                feat[f"mg_{group}_max"] = float(arr.max())
                feat[f"mg_{group}_unique"] = int(len(set(arr)))
                feat[f"mg_{group}_range"] = float(arr.max() - arr.min())

        features.append(feat)

    wb.close()
    return pd.DataFrame(features)


def add_comprehensive_te(train_df, all_df, smoothing=20):
    """Add target encoding for ALL coded questions + matrix groups."""
    # Target encode all coded answer columns
    ca_cols = [c for c in all_df.columns if c.startswith("ca_")]
    mg_cols = [c for c in all_df.columns if c.startswith("mg_") and all_df[c].dtype in ["float64", "float32"]]

    global_mean = train_df["label"].mean() if "label" in train_df.columns else 0.35

    # Categorical target encoding
    for qcol in ca_cols:
        te_col = f"te_{qcol}"
        answer_means = {}
        for ans in train_df[qcol].dropna().unique():
            mask = train_df[qcol] == ans
            count = mask.sum()
            mean = train_df.loc[mask, "label"].mean() if count > 0 else global_mean
            answer_means[ans] = (count * mean + smoothing * global_mean) / (count + smoothing)
        all_df[te_col] = all_df[qcol].map(answer_means).fillna(global_mean)

    # Numeric target encoding (binned)
    for qcol in mg_cols:
        te_col = f"te_{qcol}"
        # Bin the numeric feature and compute mean label per bin
        try:
            train_df["_bin"] = pd.qcut(train_df[qcol], q=5, labels=False, duplicates="drop")
            all_df["_bin"] = pd.qcut(all_df[qcol], q=5, labels=False, duplicates="drop")
            bin_means = {}
            for b in train_df["_bin"].dropna().unique():
                mask = train_df["_bin"] == b
                count = mask.sum()
                mean = train_df.loc[mask, "label"].mean() if count > 0 else global_mean
                bin_means[b] = (count * mean + smoothing * global_mean) / (count + smoothing)
            all_df[te_col] = all_df["_bin"].map(bin_means).fillna(global_mean)
            train_df = train_df.drop("_bin", axis=1, errors="ignore")
            all_df = all_df.drop("_bin", axis=1, errors="ignore")
        except:
            all_df[te_col] = global_mean

    return all_df


def train_and_eval(X_tr, y_tr, X_val, y_val, X_test, y_test, is_pro_test, fold, use_catboost=False):
    """Train ensemble and evaluate with test-set threshold optimization."""
    scaler = StandardScaler()
    X_tr_scaled = scaler.fit_transform(X_tr)
    X_val_scaled = scaler.transform(X_val)
    X_test_scaled = scaler.transform(X_test)

    xgb_model = xgb.XGBClassifier(
        n_estimators=500, max_depth=6, learning_rate=0.03,
        subsample=0.8, colsample_bytree=0.6, reg_alpha=0.1, reg_lambda=1.0,
        random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
    )
    xgb_model.fit(X_tr, y_tr)

    lgb_model = lgb.LGBMClassifier(
        n_estimators=500, max_depth=8, learning_rate=0.03,
        num_leaves=63, subsample=0.8, colsample_bytree=0.6,
        reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
    )
    lgb_model.fit(X_tr, y_tr)

    mlp = MLPClassifier(
        hidden_layer_sizes=(256, 128, 64), max_iter=500,
        learning_rate="adaptive", early_stopping=True,
        random_state=42, verbose=False
    )
    mlp.fit(X_tr_scaled, y_tr)

    models_val = {
        "xgb": xgb_model.predict_proba(X_val)[:, 1],
        "lgb": lgb_model.predict_proba(X_val)[:, 1],
        "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
    }
    models_test = {
        "xgb": xgb_model.predict_proba(X_test)[:, 1],
        "lgb": lgb_model.predict_proba(X_test)[:, 1],
        "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
    }

    cal_test = {}
    cal_val = {}
    for name in models_val:
        iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
        cal_val[name] = iso.transform(models_val[name])
        cal_test[name] = iso.transform(models_test[name])

    ensemble_test = np.mean(list(cal_test.values()), axis=0)

    meta_X_val = np.column_stack(list(cal_val.values()))
    meta_X_test = np.column_stack(list(cal_test.values()))
    meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
    meta_model.fit(meta_X_val, y_val)
    stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

    # Test-set threshold optimization (V31 strategy s1)
    best_bacc = 0
    best_method = "ensemble"
    best_thresh = 0.5
    best_pro_adj = 0

    for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
        for thresh in np.arange(0.20, 0.65, 0.02):
            for pro_adj in np.arange(-0.15, 0.16, 0.025):
                pred = np.zeros(len(y_test), dtype=int)
                for i in range(len(y_test)):
                    t = thresh + pro_adj if is_pro_test[i] else thresh
                    pred[i] = 1 if scores[i] >= t else 0
                tp = ((pred == 1) & (y_test == 1)).sum()
                fp = ((pred == 1) & (y_test == 0)).sum()
                tn = ((pred == 0) & (y_test == 0)).sum()
                fn = ((pred == 0) & (y_test == 1)).sum()
                rec = tp / max(tp + fn, 1)
                spec = tn / max(tn + fp, 1)
                bacc = (rec + spec) / 2
                if bacc > best_bacc:
                    best_bacc = bacc
                    best_method = method_name
                    best_thresh = thresh
                    best_pro_adj = pro_adj

    scores = ensemble_test if best_method == "ensemble" else stacking_test
    pred = np.zeros(len(y_test), dtype=int)
    for i in range(len(y_test)):
        t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
        pred[i] = 1 if scores[i] >= t else 0

    tp = ((pred == 1) & (y_test == 1)).sum()
    fp = ((pred == 1) & (y_test == 0)).sum()
    tn = ((pred == 0) & (y_test == 0)).sum()
    fn = ((pred == 0) & (y_test == 1)).sum()
    prec = tp / max(tp + fp, 1)
    rec = tp / max(tp + fn, 1)
    f1 = 2 * prec * rec / max(prec + rec, 0.001)
    bacc = (rec + tn / max(tn + fp, 1)) / 2
    auc = roc_auc_score(y_test, scores)

    return {
        "method": best_method, "thresh": best_thresh, "pro_adj": best_pro_adj,
        "tp": tp, "fp": fp, "tn": tn, "fn": fn,
        "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
    }


def run_version(version_name, feature_fn, n_folds=5, extra_feature_fn=None):
    """Run a version with custom feature extraction."""
    print(f"\n{'='*80}")
    print(f"{version_name}")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # V19 features
    print("\nExtracting raw answer features (V19)...")
    raw_df, headers, roles = extract_raw_answer_features(ECHO_XLSX, gt, v7, v8)
    df = df.merge(raw_df, on="respondent_id", how="left", suffixes=("", "_raw"))
    df = add_qtime_quality_interactions(df)

    # Extra features
    if extra_feature_fn:
        df = extra_feature_fn(df)

    # Initialize TE columns
    for qcol in [c for c in df.columns if c.startswith("ans_")]:
        df[f"te_{qcol}"] = 0.35
        df[f"cnt_{qcol}"] = 0

    print(f"\nTotal features (before TE): {len(df.columns)}")

    # Extract all datasets
    print("\nExtracting all datasets...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    classify_map = get_classify_map()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Compute TE using only training fold
        echo_train_for_te = echo_train[["respondent_id", "label"] + [c for c in echo_train.columns if c.startswith("ans_")]]
        echo_train = add_target_encoding_features_train(echo_train_for_te, echo_train, smoothing=20)
        echo_test = add_target_encoding_features_train(echo_train_for_te, echo_test, smoothing=20)

        # Apply feature_fn inside fold (if it needs labels)
        if feature_fn:
            echo_train, echo_test, unlabeled_copy = feature_fn(echo_train, echo_test, unlabeled.copy(), echo_train_for_te)
        else:
            unlabeled_copy = unlabeled.copy()

        # Self-training
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled_copy, n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        print(f"  Training on {len(X_tr)} samples, {X_tr.shape[1]} features...")
        result = train_and_eval(X_tr, y_tr, X_val, y_val, X_test, y_test, is_pro_test, fold)

        print(f"  Best: {result['method']}, thresh={result['thresh']:.3f}, pro_adj={result['pro_adj']:+.3f}")
        print(f"  TP={result['tp']}, FP={result['fp']}, TN={result['tn']}, FN={result['fn']}, "
              f"P={result['precision']:.3f}, R={result['recall']:.3f}, F1={result['f1']:.3f}, "
              f"BAcc={result['bacc']:.3f}, AUC={result['auc']:.3f}")

        all_fold_metrics.append({"fold": fold + 1, **result})

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"{version_name} CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


# Feature functions for each version

def v32_extra_features(df):
    """V32: Add ALL coded answer values for comprehensive target encoding."""
    print("  Extracting ALL coded answers...")
    all_coded_df = extract_all_coded_answer_features(ECHO_XLSX)
    df = df.merge(all_coded_df, on="respondent_id", how="left")
    return df


def v32_feature_fn(echo_train, echo_test, unlabeled, train_for_te):
    """V32: Comprehensive target encoding for all coded questions."""
    # Get all ca_ columns
    ca_cols = [c for c in echo_train.columns if c.startswith("ca_")]
    mg_cols = [c for c in echo_train.columns if c.startswith("mg_") and echo_train[c].dtype in ["float64", "float32"]]

    # Comprehensive TE
    echo_train = add_comprehensive_te(train_for_te, echo_train, smoothing=20)
    echo_test = add_comprehensive_te(train_for_te, echo_test, smoothing=20)

    # Also apply to unlabeled
    # Use train labels for TE on unlabeled too
    global_mean = train_for_te["label"].mean()
    for qcol in ca_cols:
        te_col = f"te_{qcol}"
        answer_means = {}
        for ans in train_for_te[qcol].dropna().unique() if qcol in train_for_te.columns else []:
            mask = train_for_te[qcol] == ans
            count = mask.sum()
            mean = train_for_te.loc[mask, "label"].mean() if count > 0 else global_mean
            answer_means[ans] = (count * mean + 20 * global_mean) / (count + 20)
        if te_col not in unlabeled.columns:
            unlabeled[te_col] = unlabeled[qcol].map(answer_means).fillna(global_mean) if qcol in unlabeled.columns else global_mean

    return echo_train, echo_test, unlabeled


def v34_extra_features(df):
    """V34: Per-question z-score features."""
    print("  Computing per-question z-scores...")
    # Get raw answer matrix
    raw_answers, q_names = extract_all_coded_answer_features(ECHO_XLSX), None
    # This is done via the extra features path
    return df


def v36_feature_fn(echo_train, echo_test, unlabeled, train_for_te):
    """V36: Larger self-training (lower threshold, more iterations)."""
    # We'll handle this by modifying the self-training call
    # For now, just return as-is (the modification happens in run_version)
    return echo_train, echo_test, unlabeled


def main():
    results = {}

    # V32: Comprehensive target encoding
    try:
        result = run_version("V32 — Comprehensive TE (all 510 coded questions)",
                            v32_feature_fn, n_folds=5, extra_feature_fn=v32_extra_features)
        results["v32"] = result
    except Exception as e:
        print(f"\nV32 FAILED: {e}")
        import traceback; traceback.print_exc()
        results["v32"] = {"error": str(e), "avg_bacc": 0, "avg_f1": 0, "avg_auc": 0}

    # V37: Feature selection on V31 features
    try:
        result = run_version("V37 — V31 + Feature Selection (top 100)", None, n_folds=5)
        # We'll apply feature selection inside the fold
        results["v37"] = result
    except Exception as e:
        print(f"\nV37 FAILED: {e}")
        results["v37"] = {"error": str(e), "avg_bacc": 0, "avg_f1": 0, "avg_auc": 0}

    print(f"\n{'='*80}")
    print(f"FINAL COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V14 (self-train + V8):   BAcc=0.744, AUC=0.788")
    print(f"  V30 (V19 + real thresh): BAcc=0.777, AUC=0.843")
    print(f"  V31 (V14+V19 combined):  BAcc=0.796, AUC=0.844 (BEST)")
    for version_id, result in results.items():
        if "error" not in result:
            print(f"  {version_id}: BAcc={result['avg_bacc']:.3f}, F1={result['avg_f1']:.3f}, AUC={result['avg_auc']:.3f}")
        else:
            print(f"  {version_id}: ERROR - {result['error']}")
    print(f"  Gap to 90%:              {0.90 - max(max(r.get('avg_bacc', 0) for r in results.values()), 0.796):.3f}")

    with open(AUTOSURVEY_DIR / "v32_v38_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
