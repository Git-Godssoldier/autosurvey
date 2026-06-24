#!/usr/bin/env python3
"""Generalized cross-dataset proposition mapper and coherence analyzer.

Works across all TFG annotated datasets by reading the Datamap dynamically
to build first-person self-claim propositions. No dataset-specific field
templates — all propositions are generated from the Datamap question text
and value labels.

The coherence analysis uses generalized break detection that doesn't depend
on specific field names (q14, q22, etc.) but instead on field roles identified
from the Datamap.

Usage:
    python3 generalized_proposition_mapper.py <dataset.xlsx> [--output-dir DIR]
    python3 generalized_proposition_mapper.py --all  # process all datasets
"""
from __future__ import annotations

import argparse
import json
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

DATA_DIR = Path("/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer")
OUTPUT_BASE = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/cross-dataset-propositions")

EXCLUDED_FIELDS = {
    "qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel", "LangAssessReadEase",
    "LangAssessNumSen", "LangAssessNumWords", "LangAssessNumSyl", "url", "session",
    "camp", "bhf", "sfh", "intcode", "record", "uuid", "status", "qtime",
    "SUPNAME", "ipAddress", "date", "qStateVer",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions", "outroR1_RD_Review",
                     "qcoe1R1_RD_Review", "LangAssess", "_Pasted")


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb) -> dict[str, dict]:
    """Parse the Datamap into a structured field dictionary.

    Returns: {field_name: {question_text, value_labels: {code: label}, subfield_labels: {subfield: label}, field_type}}
    """
    dm = wb["Datamap"]
    fields: dict[str, dict] = {}
    current: str | None = None

    for a, b, c in dm.iter_rows(values_only=True):
        a_s = clean(a)

        # Format 1: [field]: question text
        if a_s.startswith("[") and "]" in a_s:
            field_name = a_s.split("]", 1)[0][1:]
            qtext = a_s.split("]", 1)[1].lstrip(": ").strip()
            current = field_name
            fields[current] = {
                "question_text": qtext,
                "value_labels": {},
                "subfield_labels": {},
                "field_type": classify_field(qtext, field_name),
            }
            continue

        # Format 2: field: question text (Delta format)
        m = re.match(r"^(q\w+):\s*(.*)", a_s)
        if m:
            current = m.group(1)
            fields[current] = {
                "question_text": m.group(2),
                "value_labels": {},
                "subfield_labels": {},
                "field_type": classify_field(m.group(2), current),
            }
            continue

        # Value labels: col B = code, col C = label
        if current and b is not None and c is not None:
            b_s = clean(b)
            c_s = clean(c)
            # Subfield reference: [q3r1] | "Water filtration system"
            if b_s.startswith("[") and "]" in b_s:
                subfield = b_s.split("]")[0][1:]
                fields[current]["subfield_labels"][subfield] = c_s
            # Coded value: 1 | "Yes"
            elif b_s and c_s and not b_s.startswith("Values:"):
                fields[current]["value_labels"][b_s] = c_s

    return fields


def classify_field(qtext: str, field_name: str) -> str:
    """Classify a field's role from its question text and name."""
    low = qtext.lower()
    fname = field_name.lower()

    # Open text fields
    if fname.endswith("oe") or fname.endswith("othr1") or fname == "outro" or fname.startswith("qcoe"):
        return "open_text"
    if "open text" in low or "open-ended" in low or "please describe" in low or "please specify" in low:
        return "open_text"

    # Demographics
    demo_fields = {"qgender", "qager1", "qushhi", "qemploy", "qed", "qpolitics",
                   "qethnic", "qcountry", "qstate", "qzipr1", "qhometype", "qrevenue",
                   "qnumemployees", "qtenure", "qtenurer1", "qtrade", "qindustry"}
    if fname in demo_fields or any(fname.startswith(d) for d in ("qethnic",)):
        return "demographic"

    # Multi-select (has subfield labels)
    if re.search(r"which of the following|select all|check all|please select all", low):
        return "multi_select"

    # Matrix
    if re.search(r"how (much|well|likely|willing)|rate each|for each|please rank", low):
        return "matrix"

    # Single select (default for coded fields)
    return "single_select"


def build_proposition(field_name: str, value, field_info: dict, row: dict, headers: list, all_fields: dict) -> str | None:
    """Build a first-person proposition for a single field value."""
    if value in (None, ""):
        return None

    ftype = field_info["field_type"]
    qtext = field_info["question_text"]
    labels = field_info["value_labels"]

    if ftype == "open_text":
        # Open text: "I said about [subject]: [text]"
        subject = extract_subject(qtext)
        return f'I said about {subject}: "{clean(value)}"'

    if ftype == "demographic":
        sv = str(value).strip()
        label = labels.get(sv, sv)
        return build_demographic_proposition(field_name.lower(), label, sv)

    if ftype == "single_select":
        sv = str(value).strip()
        label = labels.get(sv, sv)
        subject = extract_subject(qtext)
        return f"I claim regarding {subject}: {label}."

    return None


def build_demographic_proposition(fname: str, label: str, raw: str) -> str:
    """Build demographic-specific propositions."""
    if fname == "qgender":
        return f"I am {label.lower()}."
    if fname == "qager1":
        return f"I am {raw} years old."
    if fname == "qushhi":
        return f"My annual household income before taxes is {label}."
    if fname == "qemploy":
        return f"My current employment status is: {label}."
    if fname == "qed":
        return f"My education level is: {label}."
    if fname == "qrevenue":
        return f"My company's annual revenue is: {label}."
    if fname == "qnumemployees":
        return f"My company has {label} employees."
    if fname == "qtenure" or fname == "qtenurer1":
        return f"My tenure in my role is: {label}."
    if fname == "qtrade":
        return f"My trade is: {label}."
    if fname == "qindustry":
        return f"I work in: {label}."
    if fname == "qstate":
        return f"I live in {label}."
    if fname == "qcountry":
        return f"I live in {label}."
    if fname == "qhometype":
        return f"My home type is: {label}."
    if fname.startswith("qethnic"):
        return f"My race/ethnicity includes: {label}."
    # Generic fallback
    return f"My demographic: {label}."


def build_multi_select_propositions(field_base: str, row: dict, headers: list, field_info: dict) -> list[str]:
    """Build propositions for multi-select fields."""
    sub_labels = field_info["subfield_labels"]
    checked = []
    for h in headers:
        if not h.startswith(field_base + "r"):
            continue
        if h.endswith("oe") or h.endswith("oth"):
            continue
        val = str(row.get(h, "")).strip()
        item_label = sub_labels.get(h, h)
        if val == "1":
            checked.append(item_label)

    props = []
    if checked:
        qtext = field_info["question_text"]
        verb = infer_multi_select_verb(qtext)
        props.append(f"I {verb}: {', '.join(checked[:15])}{'...' if len(checked)>15 else ''}.")

    # Check for "None of these"
    for h in headers:
        if not h.startswith(field_base + "r"):
            continue
        item_label = sub_labels.get(h, h)
        if "none" in item_label.lower() and str(row.get(h, "")).strip() == "1":
            props.append(f"I claim: {item_label.lower()}.")
    return props


def extract_subject(qtext: str) -> str:
    """Extract a short subject phrase from question text."""
    t = re.sub(r"\[pipe:.*?\]", "", qtext).strip()
    t = re.sub(r"^(Which of the following |What |How |Do you |Are you |Where |When |Why |Please |Using the |For each )", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\?*$", "", t).strip()
    if len(t) > 80:
        t = t[:77] + "..."
    return t.lower() if t else "this"


def infer_multi_select_verb(qtext: str) -> str:
    """Infer the verb for multi-select propositions."""
    low = qtext.lower()
    if "aware" in low:
        return "am aware of the following"
    if "purchased" in low or "bought" in low:
        return "purchased/installed in my home"
    if "interested" in low or "plan" in low:
        return "am interested in purchasing"
    if "research" in low or "source" in low:
        return "used the following research sources"
    if "concern" in low:
        return "am concerned about the following"
    if "have" in low and "water" in low:
        return "have the following in my home's water"
    if "information" in low or "get info" in low:
        return "would use the following for information"
    if "pay" in low or "willing" in low:
        return "am willing to pay for"
    if "select" in low or "which" in low:
        return "selected"
    return "selected"


def identify_field_roles(fields: dict, headers: list) -> dict:
    """Identify field roles for coherence analysis across datasets.

    Returns: {role: [field_names]} mapping.
    """
    roles = defaultdict(list)

    for fname, info in fields.items():
        if excluded(fname):
            continue
        ftype = info["field_type"]
        qtext = info["question_text"].lower()

        # Open-end / feedback
        if ftype == "open_text":
            if fname == "outro" or fname.startswith("qcoe"):
                roles["feedback_open_end"].append(fname)
            else:
                roles["other_specify"].append(fname)
            continue

        # Industry / screener
        if "industry" in qtext or fname.lower() == "qindustry":
            roles["industry"].append(fname)
            continue

        # Role / trade
        if "role" in qtext or "trade" in qtext or fname.lower().startswith("qrole") or fname.lower() == "qtrade":
            roles["role_trade"].append(fname)
            continue

        # Demographics
        if ftype == "demographic":
            roles["demographic"].append(fname)
            continue

        # Multi-select
        if ftype == "multi_select" or info["subfield_labels"]:
            if "aware" in qtext or "brand" in qtext:
                roles["brand_awareness"].append(fname)
            elif "concern" in qtext:
                roles["concerns"].append(fname)
            elif "have" in qtext and ("water" in qtext or "contamin" in qtext):
                roles["contaminants"].append(fname)
            elif "research" in qtext or "source" in qtext:
                roles["research_sources"].append(fname)
            elif "purchas" in qtext or "bought" in qtext:
                roles["purchased"].append(fname)
            elif "interest" in qtext or "plan" in qtext:
                roles["interested"].append(fname)
            else:
                roles["multi_select_other"].append(fname)
            continue

        # Single select
        if ftype == "single_select":
            if "concern" in qtext:
                roles["concern_level"].append(fname)
            elif "know" in qtext or "familiar" in qtext:
                roles["knowledge_level"].append(fname)
            elif "reason" in qtext or "prompted" in qtext or "why" in qtext:
                roles["purchase_reason"].append(fname)
            elif "search" in qtext or "shopping" in qtext:
                roles["search_stage"].append(fname)
            elif "where" in qtext or "purchase" in qtext:
                roles["purchase_location"].append(fname)
            elif "source" in qtext and "water" in qtext:
                roles["water_source"].append(fname)
            else:
                roles["single_select_other"].append(fname)
            continue

    return dict(roles)


def analyze_coherence_generalized(profile: dict, field_roles: dict, fields: dict) -> dict:
    """Generalized coherence analysis that works across datasets.

    Instead of hardcoding field names (q14, q22, etc.), this uses the field
    role mapping to find coherence breaks in whatever fields exist.
    """
    concerns = []
    hard_invalidities = []

    props_by_field = {p["field"]: p["proposition"] for p in profile.get("self_claim_propositions", [])}
    props_by_type = defaultdict(list)
    for p in profile.get("self_claim_propositions", []):
        props_by_type[p["type"]].append(p["proposition"])

    # 1. Open-end / feedback analysis
    feedback_fields = field_roles.get("feedback_open_end", [])
    for ff in feedback_fields:
        if ff in props_by_field:
            text = props_by_field[ff].lower()
            # Nonresponsive outro
            if re.search(r"thank you|good survey|nice|amazing|great experience|love this|very good", text):
                if not hard_invalidities:
                    concerns.append("feedback_nonresponsive: survey feedback is meta-praise, not substantive")
            # Non-English
            if re.search(r"zła woda|badanie|filtrów|dotyczące|encuesta|examen", text):
                hard_invalidities.append("feedback_nonenglish: survey feedback is non-English")
            # Gibberish
            if re.search(r"^[a-z]{20,}$|asdf|qwerty|bvnhgjut", text):
                hard_invalidities.append("feedback_gibberish: survey feedback is gibberish")
            # Templated
            if re.search(r"the poll examined|the study aimed|crucially, the poll|in order to determine|this (research|study|survey) (aimed|examined|sought)", text):
                concerns.append("feedback_templated: survey feedback reads as templated/LLM-generated")

    # 2. Other-specify field validity
    other_specify_fields = field_roles.get("other_specify", [])
    for of in other_specify_fields:
        if of in props_by_field:
            text = props_by_field[of].lower()
            # Check if other-specify is nonresponsive
            if re.search(r"thank you|good survey|nice|amazing|love", text):
                hard_invalidities.append(f"other_specify_nonresponsive: '{of}' contains survey-meta instead of entity name")
            # Check if other-specify is a brand name when it should be something else
            # (This is dataset-specific and hard to generalize without knowing the field context)

    # 3. Purchase reason validity (if purchase_reason role exists)
    reason_fields = field_roles.get("purchase_reason", [])
    for rf in reason_fields:
        if rf in props_by_field:
            text = props_by_field[rf]
            # Extract the answer part
            m = re.search(r'is:\s*(.+)', text)
            reason_text = m.group(1).strip().rstrip(".") if m else text
            reason_low = reason_text.lower()

            # Brand name only
            reason_words = re.search(r"taste|smell|health|contamin|chlorine|lead|hard water|safety|quality|family|kid|water|filter|clean|drink|install|concern|pestic|chemical|mineral|rust|iron|bacteria|virus|skin|hair|odor|discolor|because|want|need|decided|improve|reduce|remove|safer|better|project|job|client|home|house|business", reason_low)
            brand_only = re.match(r"^\s*(samsung|delta|moen|kohler|brita|pur|aquasana|culligan|brizo|canopy|apec|waterdrop|ispring)\s*$", reason_low)
            if brand_only:
                hard_invalidities.append(f"purchase_reason_wrong_dimension: reason is just a brand name ('{reason_text[:40]}')")

            # Templated
            templated = re.search(r"primary driver is|the benefits should be|deciding to buy.*is primarily prompted|it's a great technology and a beautiful|regarding the communication i received", reason_low)
            if templated and not re.search(r"my (family|home|house|kid|child|business|client)", reason_low):
                concerns.append(f"purchase_reason_templated: reason reads as templated/generic ('{reason_text[:50]}')")

            # Nonresponsive
            if re.search(r"thank you|good survey|nice|amazing|great experience|love this|it's essential|easy to use and unique", reason_low) and not reason_words:
                hard_invalidities.append(f"purchase_reason_nonresponsive: reason is nonresponsive ('{reason_text[:40]}')")

    # 4. Over-claiming patterns (generalized)
    # Count multi-select items checked
    multi_props = props_by_type.get("multi_select", [])
    for prop in multi_props:
        # Count items in the proposition
        m = re.search(r":\s*(.+?)\.?\s*$", prop)
        if m:
            items = [x.strip() for x in m.group(1).split(",") if x.strip() and x.strip() != "..."]
            if len(items) >= 8:
                # Check if this is brand awareness or contaminants
                if "aware" in prop.lower():
                    if len(items) >= 12:
                        concerns.append(f"brand_overclaim: aware of {len(items)} items")
                elif "concern" in prop.lower() or "water" in prop.lower():
                    concerns.append(f"multi_select_overclaim: selected {len(items)} items")
                elif "research" in prop.lower() or "source" in prop.lower():
                    if len(items) >= 6:
                        concerns.append(f"source_overclaim: used {len(items)} research sources")

    # 5. Concern vs knowledge consistency
    concern_level_fields = field_roles.get("concern_level", [])
    knowledge_level_fields = field_roles.get("knowledge_level", [])
    if concern_level_fields and knowledge_level_fields:
        concern_prop = props_by_field.get(concern_level_fields[0], "")
        knowledge_prop = props_by_field.get(knowledge_level_fields[0], "")
        if re.search(r"extremely|very concerned", concern_prop, re.I) and re.search(r"nothing|very little|not much|don't know", knowledge_prop, re.I):
            concerns.append("concern_knowledge_break: claims extreme concern but no knowledge")
        if re.search(r"not at all|slightly|not concerned", concern_prop, re.I) and re.search(r"great deal|a lot|expert", knowledge_prop, re.I):
            concerns.append("concern_knowledge_break: claims high knowledge but no concern")

    # 6. Industry exclusion (market research)
    industry_fields = field_roles.get("industry", [])
    for ind_f in industry_fields:
        if ind_f in props_by_field:
            ind_text = props_by_field[ind_f].lower()
            if "market research" in ind_text:
                hard_invalidities.append("industry_exclusion: works in market research")

    # Classify
    if hard_invalidities:
        coherence_status = "incoherent_hard"
    elif len(concerns) >= 3:
        coherence_status = "incoherent_soft"
    elif len(concerns) >= 1:
        coherence_status = "minor_concern"
    else:
        coherence_status = "coherent"

    return {
        "respondent_id": profile["respondent_id"],
        "status": profile["status"],
        "qtime_seconds": profile["qtime_seconds"],
        "supplier": profile["supplier"],
        "coherence_status": coherence_status,
        "hard_invalidities": hard_invalidities,
        "soft_concerns": concerns,
        "hard_invalidity_count": len(hard_invalidities),
        "soft_concern_count": len(concerns),
    }


def process_dataset(filepath: Path, output_dir: Path) -> dict:
    """Process a single dataset: build propositions, run coherence analysis, compute lift."""
    filename = filepath.stem
    print(f"  Processing {filename}...")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fields = load_datamap(wb)
    ws = wb["A1"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Identify field roles
    field_roles = identify_field_roles(fields, headers)

    # Load all rows
    all_rows = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid"))
        status = raw.get("status")
        if not rid or status is None:
            continue
        raw["_rid"] = rid
        raw["_xrow"] = xrow
        raw["_status"] = int(status)
        all_rows.append(raw)

    t5 = [r for r in all_rows if r["_status"] == 5]
    t3 = [r for r in all_rows if r["_status"] == 3]

    if not t5 or not t3:
        print(f"    WARNING: missing t5 ({len(t5)}) or t3 ({len(t3)})")
        wb.close()
        return {"dataset": filename, "error": "missing populations"}

    # Build propositions for each row
    def build_profile(row: dict) -> dict:
        props = []
        for fname, finfo in fields.items():
            if excluded(fname) or fname not in headers:
                continue
            ftype = finfo["field_type"]

            if ftype == "multi_select" or finfo["subfield_labels"]:
                sub_props = build_multi_select_propositions(fname.split("r")[0] if "r" in fname else fname, row, headers, finfo)
                for p in sub_props:
                    props.append({"field": fname, "type": "multi_select", "proposition": p})
            else:
                value = row.get(fname)
                p = build_proposition(fname, value, finfo, row, headers, fields)
                if p:
                    props.append({"field": fname, "type": ftype, "proposition": p})

        return {
            "respondent_id": row["_rid"],
            "source_excel_row": row["_xrow"],
            "status": row["_status"],
            "qtime_seconds": float(row["qtime"]) if row.get("qtime") else None,
            "supplier": row.get("SUPNAME"),
            "self_claim_propositions": props,
            "proposition_count": len(props),
        }

    t5_profiles = [build_profile(r) for r in t5]

    # t3 guardrail: 80 rows (or all if <80)
    t3_sorted = sorted(t3, key=lambda r: (str(r.get("SUPNAME") or ""), float(r.get("qtime") or 0)))
    step = max(1, len(t3_sorted) // min(80, len(t3_sorted)))
    t3_guard = t3_sorted[::step][:80]
    t3_profiles = [build_profile(r) for r in t3_guard]

    # Coherence analysis
    t5_coherence = [analyze_coherence_generalized(p, field_roles, fields) for p in t5_profiles]
    t3_coherence = [analyze_coherence_generalized(p, field_roles, fields) for p in t3_profiles]

    # Compute lift
    t5_status = Counter(r["coherence_status"] for r in t5_coherence)
    t3_status = Counter(r["coherence_status"] for r in t3_coherence)

    t5_hard = Counter()
    for r in t5_coherence:
        for h in r["hard_invalidities"]:
            t5_hard[h.split(":")[0]] += 1
    t3_hard = Counter()
    for r in t3_coherence:
        for h in r["hard_invalidities"]:
            t3_hard[h.split(":")[0]] += 1

    t5_soft = Counter()
    for r in t5_coherence:
        for s in r["soft_concerns"]:
            t5_soft[s.split(":")[0]] += 1
    t3_soft = Counter()
    for r in t3_coherence:
        for s in r["soft_concerns"]:
            t3_soft[s.split(":")[0]] += 1

    all_signal_types = set(list(t5_soft.keys()) + list(t3_soft.keys()) + list(t5_hard.keys()) + list(t3_hard.keys()))
    lift = {}
    for s in all_signal_types:
        t5_rate = (t5_soft.get(s, 0) + t5_hard.get(s, 0)) / len(t5_coherence)
        t3_rate = (t3_soft.get(s, 0) + t3_hard.get(s, 0)) / len(t3_coherence)
        lift[s] = {
            "t5_rate": round(t5_rate, 4),
            "t3_rate": round(t3_rate, 4),
            "lift": round(t5_rate / max(t3_rate, 0.001), 2),
        }

    # Timing comparison
    t5_qtimes = [r["qtime_seconds"] for r in t5_profiles if r["qtime_seconds"]]
    t3_qtimes = [float(r.get("qtime")) for r in t3 if r.get("qtime")]
    t5_qtime_avg = statistics.mean(t5_qtimes) if t5_qtimes else 0
    t3_qtime_avg = statistics.mean(t3_qtimes) if t3_qtimes else 0

    # Supplier
    t5_missing_sup = sum(1 for r in t5_profiles if not r["supplier"] or str(r["supplier"]).strip() == "") / len(t5_profiles)
    t3_missing_sup = sum(1 for r in t3 if not r.get("SUPNAME") or str(r["SUPNAME"]).strip() == "") / len(t3)

    result = {
        "dataset": filename,
        "t5_total": len(t5),
        "t3_total": len(t3),
        "t3_guardrail": len(t3_profiles),
        "field_roles": {k: v[:10] for k, v in field_roles.items()},
        "avg_propositions_per_row": round(sum(p["proposition_count"] for p in t5_profiles) / len(t5_profiles), 1) if t5_profiles else 0,
        "t5_coherence_status": dict(t5_status),
        "t3_coherence_status": dict(t3_status),
        "t5_hard_invalidity_types": dict(t5_hard),
        "t3_hard_invalidity_types": dict(t3_hard),
        "t5_soft_concern_types": dict(t5_soft),
        "t3_soft_concern_types": dict(t3_soft),
        "lift_signals": lift,
        "timing": {
            "t5_avg_qtime": round(t5_qtime_avg, 1),
            "t3_avg_qtime": round(t3_qtime_avg, 1),
            "t5_slower": t5_qtime_avg > t3_qtime_avg,
        },
        "supplier": {
            "t5_missing_rate": round(t5_missing_sup, 3),
            "t3_missing_rate": round(t3_missing_sup, 3),
            "lift": round(t5_missing_sup / max(t3_missing_sup, 0.001), 2),
        },
    }

    # Save artifacts
    ds_out = output_dir / filename
    ds_out.mkdir(parents=True, exist_ok=True)
    (ds_out / "t5_self_claim_profiles.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in t5_profiles) + "\n"
    )
    (ds_out / "t3_guardrail_profiles.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in t3_profiles) + "\n"
    )
    (ds_out / "t5_coherence_analysis.jsonl").write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, separators=(",", ":")) for r in t5_coherence) + "\n"
    )
    (ds_out / "t3_coherence_analysis.jsonl").write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, separators=(",", ":")) for r in t3_coherence) + "\n"
    )
    (ds_out / "dataset_summary.json").write_text(json.dumps(result, indent=2))

    wb.close()
    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--all", action="store_true", help="Process all datasets")
    parser.add_argument("dataset", nargs="?", help="Single dataset filename")
    parser.add_argument("--output-dir", default=str(OUTPUT_BASE))
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.all:
        datasets = sorted([f for f in DATA_DIR.iterdir() if f.suffix == ".xlsx"])
        results = []
        for ds in datasets:
            try:
                result = process_dataset(ds, output_dir)
                results.append(result)
            except Exception as e:
                print(f"  ERROR processing {ds.name}: {e}")
                results.append({"dataset": ds.stem, "error": str(e)})

        # Save aggregated results
        (output_dir / "cross_dataset_summary.json").write_text(json.dumps(results, indent=2))
        print(f"\nProcessed {len(results)} datasets. Summary saved to {output_dir / 'cross_dataset_summary.json'}")

        # Print summary table
        print(f"\n{'Dataset':<45} {'t5':>5} {'t3':>5} {'incoh_hard':>11} {'incoh_soft':>11} {'coherent':>9} {'avg_props':>9}")
        for r in results:
            if "error" in r:
                print(f"{r['dataset']:<45} ERROR: {r['error'][:40]}")
                continue
            t5s = r["t5_coherence_status"]
            print(f"{r['dataset']:<45} {r['t5_total']:>5} {r['t3_total']:>5} {t5s.get('incoherent_hard',0):>11} {t5s.get('incoherent_soft',0):>11} {t5s.get('coherent',0):>9} {r['avg_propositions_per_row']:>9}")

    elif args.dataset:
        ds_path = DATA_DIR / args.dataset
        result = process_dataset(ds_path, output_dir)
        print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
