#!/usr/bin/env python3
"""Cross-corpus ML signal correlation analysis.

Trains ML models across all annotated datasets to find which signals
statistically separate accepted (status=3) from rejected (status=5)
respondents. Outputs:
  1. Per-dataset signal importance rankings
  2. Cross-dataset universal signals (fire across all/most datasets)
  3. Dataset-specific signals (strong in one but not others)
  4. Signal family correlation matrix
  5. JSON output for integration into review packets

Usage:
    python3 cross_corpus_signal_correlation.py
"""
import json
import sys
import warnings
from pathlib import Path
from collections import defaultdict, Counter
import numpy as np

warnings.filterwarnings('ignore')

# Add skill scripts to path
SKILL_SCRIPTS = Path(__file__).parent.parent / "skills" / "cleaning-survey-quality" / "scripts"
sys.path.insert(0, str(SKILL_SCRIPTS))

from sklearn.ensemble import GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import roc_auc_score, classification_report
from sklearn.model_selection import cross_val_score

# The continuous-evolution file with all 11 datasets
CORPUS_PATH = '/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/continuous-evolution/public/AUTOSURVEY_RESULTS.xlsx'

# Signal family mapping — group raw features into evidence families
# Based on actual signal names from the AutoQuality pipeline
SIGNAL_FAMILY_MAP = {
    'core_oe_quality': [
        'thin_open_end', 'polished_ungrounded_open_end', 'weak_persona_context',
        'sig_very_short_required_open_end', 'sig_oe_word_count_low',
        'sig_oe_gibberish', 'sig_oe_off_topic', 'sig_oe_thin',
        'sig_oe_product_review', 'sig_oe_benefit_stack',
    ],
    'platform_risk': [
        'sig_termflags', 'sig_qc_flag', 'sig_rd_search_high',
        'sig_rd_search_moderate', 'sig_readlevel_high', 'sig_readlevel_low',
        'sig_non_english', 'rd_searchr1', 'termflags',
    ],
    'model_risk': ['ml_triage_score'],
    'source_risk': [
        'sig_supplier_reject_high', 'sig_supplier_reject_medium',
        'sig_source_risk',
    ],
    'duplicate_semantics': [
        'duplicate_open_chain', 'sig_ai_text_suspicion', 'sig_oe_duplicate_high',
        'sig_paraphrase_cluster',
    ],
    'survey_structure': [
        'survey_meta_substitution', 'sig_classify_mismatch',
        'sig_channel_brand_mismatch', 'sig_pro_no_pro_evidence',
        'sig_structure_incoherence', 'classify',
    ],
    'brand_funnel': [
        'sig_wrong_brand_universe', 'sig_brand_funnel_incoherence',
        'sig_no_brand_awareness', 'sig_garbled_brand', 'sig_nps_generic',
    ],
    'timing_engagement': [
        'low_total_duration', 'text_time_mismatch', 'high_matrix_uniformity',
        'sig_timing_bottom10', 'sig_timing_bottom25',
        'sig_straightlining_high', 'sig_matrix_prevalence_high', 'qtime',
    ],
    'quota_reconstruction': [
        'sig_quota_overfilled', 'sig_quota_pro_cell',
        'sig_quota_channel_mismatch',
    ],
}


def load_corpus():
    """Load the continuous-evolution corpus with all datasets."""
    import openpyxl
    wb = openpyxl.load_workbook(CORPUS_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    ds_idx = headers.index('autosurvey_dataset')
    status_idx = headers.index('autosurvey_client_status')
    uuid_idx = headers.index('uuid')
    signals_idx = headers.index('autosurvey_signals') if 'autosurvey_signals' in headers else None
    json_idx = headers.index('autosurvey_original_response_json')

    # Also get agent signals and risk scores
    auth_risk_idx = headers.index('autosurvey_authenticity_risk') if 'autosurvey_authenticity_risk' in headers else None
    client_rej_idx = headers.index('autosurvey_client_rejection_probability') if 'autosurvey_client_rejection_probability' in headers else None
    conf_idx = headers.index('autosurvey_confidence') if 'autosurvey_confidence' in headers else None

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ds = row[ds_idx]
        status = row[status_idx]
        uuid = row[uuid_idx]
        signals_raw = row[signals_idx] if signals_idx else ''
        json_raw = row[json_idx]
        auth_risk = row[auth_risk_idx] if auth_risk_idx else None
        client_rej = row[client_rej_idx] if client_rej_idx else None
        conf = row[conf_idx] if conf_idx else None

        if status not in (3, 5, '3', '5'):
            continue

        # Parse signals
        signals = []
        if signals_raw:
            if isinstance(signals_raw, str):
                try:
                    signals = json.loads(signals_raw)
                except json.JSONDecodeError:
                    signals = [s.strip() for s in signals_raw.split(',') if s.strip()]
            elif isinstance(signals_raw, list):
                signals = signals_raw

        # Parse raw response JSON for feature extraction
        raw_response = {}
        if json_raw and isinstance(json_raw, str):
            try:
                raw_response = json.loads(json_raw)
            except json.JSONDecodeError:
                pass

        rows.append({
            'dataset': ds,
            'status': int(status),
            'uuid': uuid,
            'signals': signals,
            'raw_response': raw_response,
            'authenticity_risk': float(auth_risk) if auth_risk else None,
            'client_rejection_prob': float(client_rej) if client_rej else None,
            'confidence': float(conf) if conf else None,
        })

    return rows


def extract_signal_features(rows):
    """Extract binary signal features from the signal lists."""
    # Collect all unique signal names
    all_signals = set()
    for r in rows:
        for s in r['signals']:
            if isinstance(s, str):
                all_signals.add(s)
            elif isinstance(s, dict):
                all_signals.add(s.get('signal', s.get('name', str(s))))

    all_signals = sorted(all_signals)
    signal_idx = {s: i for i, s in enumerate(all_signals)}

    # Build feature matrix
    X = np.zeros((len(rows), len(all_signals)))
    y = np.array([1 if r['status'] == 5 else 0 for r in rows])
    datasets = [r['dataset'] for r in rows]

    for i, r in enumerate(rows):
        for s in r['signals']:
            if isinstance(s, str):
                if s in signal_idx:
                    X[i, signal_idx[s]] = 1
            elif isinstance(s, dict):
                key = s.get('signal', s.get('name'))
                if key and key in signal_idx:
                    X[i, signal_idx[key]] = 1

    # Add raw response features
    raw_features = extract_raw_response_features(rows)
    if raw_features.shape[1] > 0:
        X = np.hstack([X, raw_features])
        raw_feature_names = [
            'TERMFLAGS', 'RD_Searchr1', 'qtime', 'q1', 'q15', 'q16',
            'PRODUCT2RATE', 'OWNERSHIP', 'CLASSIFY', 'REGION', 'age',
        ][:raw_features.shape[1]]
    else:
        raw_feature_names = []

    feature_names = list(all_signals) + raw_feature_names

    return X, y, feature_names, datasets


def extract_raw_response_features(rows):
    """Extract numeric features from raw response JSON."""
    # Key fields to extract across datasets
    numeric_fields = [
        'TERMFLAGS', 'RD_Searchr1', 'qtime', 'q1', 'q15', 'q16',
        'PRODUCT2RATE', 'OWNERSHIP', 'CLASSIFY', 'REGION', 'age',
    ]

    features = []
    for r in rows:
        raw = r['raw_response']
        feat = []
        for field in numeric_fields:
            val = raw.get(field)
            if val is None or val == '':
                feat.append(0)
            else:
                try:
                    feat.append(float(val))
                except (ValueError, TypeError):
                    feat.append(0)
        features.append(feat)

    return np.array(features)


def train_per_dataset_models(X, y, feature_names, datasets):
    """Train a model per dataset and extract signal importance."""
    results = {}
    unique_datasets = sorted(set(datasets))

    for ds in unique_datasets:
        mask = np.array([d == ds for d in datasets])
        X_ds = X[mask]
        y_ds = y[mask]

        n_discard = int(y_ds.sum())
        n_keep = int(len(y_ds) - n_discard)

        if n_discard < 10 or n_keep < 10:
            print(f"  {ds}: skipping (too few samples: {n_discard} discards, {n_keep} keeps)")
            continue

        # Train gradient boosting
        try:
            clf = GradientBoostingClassifier(
                n_estimators=100, max_depth=3, random_state=42
            )
            clf.fit(X_ds, y_ds)

            # Get feature importance
            importances = clf.feature_importances_
            top_indices = np.argsort(importances)[::-1][:20]

            top_signals = []
            for idx in top_indices:
                if importances[idx] > 0:
                    top_signals.append({
                        'feature': feature_names[idx],
                        'importance': float(importances[idx]),
                        'family': classify_signal_to_family(feature_names[idx]),
                    })

            # Cross-validated AUC
            try:
                auc_scores = cross_val_score(
                    clf, X_ds, y_ds, cv=3, scoring='roc_auc'
                )
                auc = float(auc_scores.mean())
            except Exception:
                auc = 0.0

            results[ds] = {
                'n_total': int(len(y_ds)),
                'n_discard': n_discard,
                'n_keep': n_keep,
                'discard_rate': n_discard / len(y_ds),
                'auc': auc,
                'top_signals': top_signals,
            }

            print(f"  {ds}: {len(y_ds)} respondents, {n_discard} discards ({n_discard/len(y_ds)*100:.1f}%), AUC={auc:.3f}")

        except Exception as e:
            print(f"  {ds}: ERROR - {e}")

    return results


def classify_signal_to_family(feature_name):
    """Classify a feature name into an evidence family."""
    name_lower = feature_name.lower()

    # Check compound signals — split on ';' and check each part
    parts = [p.strip().lower() for p in name_lower.split(';')]

    for part in parts:
        for family, patterns in SIGNAL_FAMILY_MAP.items():
            for pattern in patterns:
                if pattern.lower() == part or pattern.lower() in part:
                    return family

    # Heuristic fallback on the full name
    if any(k in name_lower for k in ['thin_open', 'polished_ungrounded', 'weak_persona', 'oe_', 'open_end', 'word_count', 'gibberish', 'off_topic']):
        return 'core_oe_quality'
    if any(k in name_lower for k in ['termflag', 'rd_search', 'readlevel', 'qc_', 'non_english']):
        return 'platform_risk'
    if any(k in name_lower for k in ['ml_', 'model_', 'triage']):
        return 'model_risk'
    if any(k in name_lower for k in ['supplier', 'source', 'reject_rate']):
        return 'source_risk'
    if any(k in name_lower for k in ['duplicate', 'ai_text', 'paraphrase']):
        return 'duplicate_semantics'
    if any(k in name_lower for k in ['survey_meta', 'classify', 'channel', 'pro_', 'structure', 'meta_substitution']):
        return 'survey_structure'
    if any(k in name_lower for k in ['brand', 'funnel', 'nps', 'awareness']):
        return 'brand_funnel'
    if any(k in name_lower for k in ['low_total_duration', 'text_time_mismatch', 'matrix_uniformity', 'timing', 'straightline', 'matrix', 'speed', 'qtime']):
        return 'timing_engagement'
    if any(k in name_lower for k in ['quota', 'cell']):
        return 'quota_reconstruction'
    if 'no_strong_staged' in name_lower:
        return 'overall_signal_strength'
    return 'other'


def find_universal_signals(per_dataset_results):
    """Find signals that are important across multiple datasets."""
    signal_datasets = defaultdict(list)
    signal_importances = defaultdict(list)

    for ds, result in per_dataset_results.items():
        for sig in result['top_signals']:
            name = sig['feature']
            signal_datasets[name].append(ds)
            signal_importances[name].append(sig['importance'])

    universal = []
    for name, datasets in signal_datasets.items():
        if len(datasets) >= 3:  # appears in 3+ datasets
            universal.append({
                'feature': name,
                'family': classify_signal_to_family(name),
                'n_datasets': len(datasets),
                'datasets': datasets,
                'mean_importance': float(np.mean(signal_importances[name])),
                'max_importance': float(np.max(signal_importances[name])),
            })

    universal.sort(key=lambda x: x['n_datasets'], reverse=True)
    return universal


def find_dataset_specific_signals(per_dataset_results):
    """Find signals strong in one dataset but absent in others."""
    all_signals = set()
    for result in per_dataset_results.values():
        for sig in result['top_signals']:
            all_signals.add(sig['feature'])

    specific = {}
    for ds, result in per_dataset_results.items():
        ds_specific = []
        other_datasets = [d for d in per_dataset_results if d != ds]
        for sig in result['top_signals']:
            # Check if this signal appears in other datasets
            appears_elsewhere = False
            for other_ds in other_datasets:
                for other_sig in per_dataset_results[other_ds]['top_signals']:
                    if other_sig['feature'] == sig['feature']:
                        appears_elsewhere = True
                        break
                if appears_elsewhere:
                    break
            if not appears_elsewhere and sig['importance'] > 0.01:
                ds_specific.append(sig)
        if ds_specific:
            specific[ds] = ds_specific

    return specific


def compute_family_correlation(X, y, feature_names, datasets):
    """Compute correlation between each evidence family and discard status."""
    family_scores = defaultdict(lambda: {'correlations': [], 'datasets': []})

    for ds in sorted(set(datasets)):
        mask = np.array([d == ds for d in datasets])
        X_ds = X[mask]
        y_ds = y[mask]

        if y_ds.sum() < 10 or (len(y_ds) - y_ds.sum()) < 10:
            continue

        # Group features by family
        family_feature_indices = defaultdict(list)
        for i, name in enumerate(feature_names):
            family = classify_signal_to_family(name)
            family_feature_indices[family].append(i)

        for family, indices in family_feature_indices.items():
            if not indices:
                continue
            # Compute family-level score = mean of features in family
            family_score = X_ds[:, indices].mean(axis=1)
            # Correlation with discard
            if family_score.std() > 0:
                corr = float(np.corrcoef(family_score, y_ds)[0, 1])
            else:
                corr = 0.0
            family_scores[family]['correlations'].append(corr)
            family_scores[family]['datasets'].append(ds)

    # Aggregate
    family_summary = {}
    for family, data in family_scores.items():
        if data['correlations']:
            family_summary[family] = {
                'mean_correlation': float(np.mean(data['correlations'])),
                'median_correlation': float(np.median(data['correlations'])),
                'std_correlation': float(np.std(data['correlations'])),
                'n_datasets': len(data['correlations']),
                'datasets': data['datasets'],
                'correlations': dict(zip(data['datasets'], data['correlations'])),
                'direction': 'positive' if np.mean(data['correlations']) > 0 else 'negative',
            }

    return family_summary


def train_global_model(X, y, feature_names, datasets):
    """Train a global model across all datasets."""
    print("\nTraining global model across all datasets...")

    clf = GradientBoostingClassifier(n_estimators=200, max_depth=4, random_state=42)
    clf.fit(X, y)

    importances = clf.feature_importances_
    top_indices = np.argsort(importances)[::-1][:30]

    top_signals = []
    for idx in top_indices:
        if importances[idx] > 0:
            top_signals.append({
                'feature': feature_names[idx],
                'importance': float(importances[idx]),
                'family': classify_signal_to_family(feature_names[idx]),
            })

    # AUC
    try:
        auc_scores = cross_val_score(clf, X, y, cv=5, scoring='roc_auc')
        auc = float(auc_scores.mean())
    except Exception:
        auc = 0.0

    print(f"  Global AUC: {auc:.3f}")
    print(f"  Top 10 signals:")
    for s in top_signals[:10]:
        print(f"    {s['feature']}: {s['importance']:.4f} ({s['family']})")

    return {
        'auc': auc,
        'top_signals': top_signals,
        'n_total': int(len(y)),
        'n_discard': int(y.sum()),
        'n_keep': int(len(y) - y.sum()),
    }


def main():
    print("=" * 80)
    print("CROSS-CORPUS ML SIGNAL CORRELATION ANALYSIS")
    print("=" * 80)

    output_dir = Path('/Users/jeremyalston/Perfect/autosurvey/skills/cleaning-survey-quality/evolution/ml-signal-correlation')
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load corpus
    print("\n[1/5] Loading annotated corpus...")
    rows = load_corpus()
    print(f"  Loaded {len(rows)} respondents across {len(set(r['dataset'] for r in rows))} datasets")

    # Dataset distribution
    ds_dist = Counter(r['dataset'] for r in rows)
    for ds, count in sorted(ds_dist.items()):
        n_discard = sum(1 for r in rows if r['dataset'] == ds and r['status'] == 5)
        print(f"    {ds}: {count} respondents, {n_discard} discards ({n_discard/count*100:.1f}%)")

    # Extract features
    print("\n[2/5] Extracting signal features...")
    X, y, feature_names, datasets = extract_signal_features(rows)
    print(f"  Feature matrix: {X.shape[0]} respondents x {X.shape[1]} features")

    # Train per-dataset models
    print("\n[3/5] Training per-dataset ML models...")
    per_dataset = train_per_dataset_models(X, y, feature_names, datasets)

    # Find universal and dataset-specific signals
    print("\n[4/5] Analyzing signal correlations...")
    universal = find_universal_signals(per_dataset)
    specific = find_dataset_specific_signals(per_dataset)
    family_corr = compute_family_correlation(X, y, feature_names, datasets)

    print(f"\n  Universal signals (appear in 3+ datasets): {len(universal)}")
    for s in universal[:15]:
        print(f"    {s['feature']}: {s['n_datasets']} datasets, mean_importance={s['mean_importance']:.4f} ({s['family']})")

    print(f"\n  Evidence family correlations with client discard:")
    for family, data in sorted(family_corr.items(), key=lambda x: abs(x[1]['mean_correlation']), reverse=True):
        print(f"    {family}: mean_corr={data['mean_correlation']:.3f}, n_datasets={data['n_datasets']}, direction={data['direction']}")

    # Train global model
    print("\n[5/5] Training global model...")
    global_model = train_global_model(X, y, feature_names, datasets)

    # Save results
    results = {
        'corpus_stats': {
            'total_respondents': int(len(rows)),
            'total_discards': int(y.sum()),
            'total_keeps': int(len(y) - y.sum()),
            'n_datasets': len(set(datasets)),
            'dataset_distribution': dict(ds_dist),
        },
        'per_dataset_models': per_dataset,
        'universal_signals': universal,
        'dataset_specific_signals': {k: v for k, v in specific.items()},
        'family_correlations': family_corr,
        'global_model': global_model,
    }

    output_path = output_dir / 'cross_corpus_signal_correlation.json'
    with open(output_path, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\nResults saved to: {output_path}")

    # Also save a readable summary
    summary_path = output_dir / 'signal_correlation_summary.md'
    with open(summary_path, 'w') as f:
        f.write("# Cross-Corpus ML Signal Correlation Analysis\n\n")
        f.write(f"**Corpus:** {len(rows)} respondents across {len(set(datasets))} datasets\n\n")

        f.write("## Evidence Family Correlations with Client Discard\n\n")
        f.write("| Family | Mean Correlation | N Datasets | Direction |\n")
        f.write("|--------|-----------------|------------|----------|\n")
        for family, data in sorted(family_corr.items(), key=lambda x: abs(x[1]['mean_correlation']), reverse=True):
            f.write(f"| {family} | {data['mean_correlation']:.3f} | {data['n_datasets']} | {data['direction']} |\n")

        f.write("\n## Universal Signals (3+ datasets)\n\n")
        f.write("| Signal | Family | N Datasets | Mean Importance |\n")
        f.write("|--------|--------|------------|----------------|\n")
        for s in universal[:20]:
            f.write(f"| {s['feature']} | {s['family']} | {s['n_datasets']} | {s['mean_importance']:.4f} |\n")

        f.write("\n## Per-Dataset Top Signals\n\n")
        for ds, result in sorted(per_dataset.items()):
            f.write(f"### {ds} (AUC={result['auc']:.3f}, {result['n_discard']} discards)\n\n")
            f.write("| Signal | Importance | Family |\n")
            f.write("|--------|------------|--------|\n")
            for sig in result['top_signals'][:10]:
                f.write(f"| {sig['feature']} | {sig['importance']:.4f} | {sig['family']} |\n")
            f.write("\n")

        f.write("## Global Model\n\n")
        f.write(f"- AUC: {global_model['auc']:.3f}\n")
        f.write(f"- Total: {global_model['n_total']} respondents ({global_model['n_discard']} discards)\n\n")
        f.write("| Signal | Importance | Family |\n")
        f.write("|--------|------------|--------|\n")
        for sig in global_model['top_signals'][:15]:
            f.write(f"| {sig['feature']} | {sig['importance']:.4f} | {sig['family']} |\n")

    print(f"Summary saved to: {summary_path}")
    print(f"\n{'='*80}")
    print("COMPLETE")
    print(f"{'='*80}")


if __name__ == '__main__':
    main()
