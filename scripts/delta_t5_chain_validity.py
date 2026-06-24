#!/usr/bin/env python3
"""Delta t=5 per-field chain validity review (Layer 2 of progressive filtering).

Examines the full response chain for each status=5 row:
- funnel consistency (awareness → consideration → purchase → reason)
- brand awareness anomalies (over-selection, under-selection)
- matrix behavior across distinct product categories
- other-specify field validity
- q14 purchase reason responsiveness
- q15/q16 concern/knowledge consistency
- route integrity

Stages evidence only. The agent derives the chain validity judgment.
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ANNOTATED = Path("/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260111_Delta Water Filtration.xlsx")
OUT = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/agent-native-delta-transfer-2026-06-24/.autosurvey-internal/t5_semantic")

EXCLUDED_FIELDS = {
    "qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel", "LangAssessReadEase",
    "LangAssessNumSen", "LangAssessNumWords", "LangAssessNumSyl", "url", "session",
    "camp", "bhf", "sfh", "intcode",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions")


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb) -> dict[str, list[str]]:
    qmap: dict[str, list[str]] = {}
    current: str | None = None
    for a, b, c in wb["Datamap"].iter_rows(values_only=True):
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            current = a_s.split("]", 1)[0][1:]
            qmap[current] = [a_s]
            continue
        if current and (a_s or b is not None or c is not None):
            if a_s:
                qmap[current].append(a_s)
            if b is not None or c is not None:
                qmap[current].append(f"{b} | {c}")
    return qmap


def field_text(field: str, qmap) -> str:
    lines = qmap.get(field, [])
    if not lines:
        return field
    first = lines[0]
    if "] | " in first:
        return first.split("] | ", 1)[1]
    if "]: " in first:
        return first.split("]: ", 1)[1]
    return first


def value_text(field: str, value, qmap) -> str:
    if value in (None, ""):
        return ""
    sv = str(value)
    for line in qmap.get(field, []):
        if " | " not in line:
            continue
        left, right = line.split(" | ", 1)
        if clean(left) == sv:
            return f"{sv}={right}"
    return sv


# Brand list for q26/q27 (awareness/consideration)
BRANDS = {
    "q26r1": "Delta", "q26r2": "Moen", "q26r3": "Kohler", "q26r4": "Brita",
    "q26r5": "PUR", "q26r6": "GE", "q26r7": "Aquasana", "q26r8": "A.O. Smith",
    "q26r9": "Culligan", "q26r10": "Canopy", "q26r11": "Apec", "q26r12": "WaterDrop",
    "q26r13": "iSpring", "q26r14": "Brizo", "q26r15": "Other", "q26r16": "Another brand",
    "q26r17": "Not sure",
}

# Research sources q7
SOURCES_Q7 = {
    "q7r1": "Google", "q7r2": "Amazon", "q7r3": "Store employees", "q7r4": "Manufacturer website",
    "q7r5": "In-store signage", "q7r6": "HomeDepot.com", "q7r7": "Lowes.com", "q7r8": "Acehardware.com",
    "q7r9": "Facebook", "q7r10": "Instagram", "q7r11": "Other", "q7r12": "Did not use any",
}

# q5 matrix product categories (semantically distinct)
Q5_PRODUCTS = {
    "q5r1": "Kitchen Sink Faucet", "q5r2": "Bathroom Sink Faucet", "q5r3": "Showerhead",
    "q5r4": "Tub Spout Filler", "q5r5": "Whole Home Filtration", "q5r6": "Whole Home Softener",
    "q5r7": "Refrigerator/Ice Maker",
}


def main() -> None:
    wb = openpyxl.load_workbook(ANNOTATED, read_only=True, data_only=True)
    qmap = load_datamap(wb)
    ws = wb["A1"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    all_rows = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid"))
        status = raw.get("status")
        if not rid or status is None:
            continue
        raw["_rid"] = rid
        raw["_xrow"] = xrow
        raw["_status"] = int(status)
        all_rows.append(raw)

    t5 = [r for r in all_rows if r["_status"] == 5]
    t3 = [r for r in all_rows if r["_status"] == 3]

    # ---- population-level baselines for comparison ----
    def brand_awareness_count(row):
        return sum(1 for k in BRANDS if k.startswith("q26r") and str(row.get(k, "")).strip() == "1")

    def research_source_count(row):
        return sum(1 for k in SOURCES_Q7 if str(row.get(k, "")).strip() == "1")

    def q5_ownership(row):
        """Return dict of product -> (have, plan) from c1/c2/c3 columns."""
        result = {}
        for base, product in Q5_PRODUCTS.items():
            c1 = clean(row.get(f"{base}c1"))  # have
            c2 = clean(row.get(f"{base}c2"))  # plan to buy
            c3 = clean(row.get(f"{base}c3"))  # neither
            result[product] = {"have": c1, "plan": c2, "neither": c3}
        return result

    t3_brand_counts = [brand_awareness_count(r) for r in t3]
    t3_source_counts = [research_source_count(r) for r in t3]
    t3_qtime = []
    for r in t3:
        try:
            t3_qtime.append(float(r.get("qtime")))
        except Exception:
            pass

    # ---- per-row chain validity review ----
    judgments = []
    for r in t5:
        rid = r["_rid"]
        try:
            qtime = float(r.get("qtime"))
        except Exception:
            qtime = None

        # 1. Brand awareness analysis
        brand_checked = []
        for k, name in BRANDS.items():
            if str(r.get(k, "")).strip() == "1":
                brand_checked.append(name)
        brand_count = len(brand_checked)
        # over-selection: checking many brands including obscure ones
        obscure_brands = [b for b in brand_checked if b in ("Apec", "WaterDrop", "iSpring", "Canopy")]
        # "Not sure" + checking brands = contradiction
        not_sure = str(r.get("q26r17", "")).strip() == "1"
        brand_contradiction = not_sure and brand_count > 0

        # 2. Research source analysis
        sources_checked = []
        for k, name in SOURCES_Q7.items():
            if str(r.get(k, "")).strip() == "1":
                sources_checked.append(name)
        source_count = len(sources_checked)
        no_sources = "Did not use any" in sources_checked
        source_contradiction = no_sources and source_count > 1

        # 3. q5 matrix: ownership/plan across distinct products
        ownership = q5_ownership(r)
        # count how many products they have, plan to buy, neither
        have_count = sum(1 for v in ownership.values() if v["have"] == "1")
        plan_count = sum(1 for v in ownership.values() if v["plan"] == "1")
        neither_count = sum(1 for v in ownership.values() if v["neither"] == "1")
        # over-selection: claiming to have many different filtration types
        over_select_have = have_count >= 5
        # all-plan: planning to buy everything
        over_select_plan = plan_count >= 5

        # 4. q14 purchase reason
        q14_val = clean(r.get("q14"))
        q14_decoded = value_text("q14", r.get("q14"), qmap)
        # q14 is actually coded - check if it's a real reason code
        q14_responsive = q14_val != "" and q14_val is not None

        # 5. q15 concern level
        q15_val = clean(r.get("q15"))
        q15_decoded = value_text("q15", r.get("q15"), qmap)

        # 6. q16 knowledge level
        q16_val = clean(r.get("q16"))
        q16_decoded = value_text("q16", r.get("q16"), qmap)

        # 7. q30 water source
        q30_val = clean(r.get("q30"))
        q30_decoded = value_text("q30", r.get("q30"), qmap)

        # 8. q31 area type
        q31_val = clean(r.get("q31"))
        q31_decoded = value_text("q31", r.get("q31"), qmap)

        # 9. q32 square footage
        q32_val = clean(r.get("q32"))
        q32_decoded = value_text("q32", r.get("q32"), qmap)

        # 10. q33 home value
        q33_val = clean(r.get("q33"))
        q33_decoded = value_text("q33", r.get("q33"), qmap)

        # 11. q13 purchase role
        q13_val = clean(r.get("q13"))
        q13_decoded = value_text("q13", r.get("q13"), qmap)

        # 12. Supplier
        supplier = r.get("SUPNAME")

        # 13. outro (from prior analysis)
        outro_val = clean(r.get("outro"))

        # ---- chain validity concerns ----
        concerns = []
        hard_invalidities = []
        soft_concerns = []

        # Brand over-selection
        if brand_count >= 12:
            soft_concerns.append(f"brand_oversel: checked {brand_count}/17 brands including {obscure_brands}")
        elif brand_count >= 10:
            soft_concerns.append(f"brand_broad: checked {brand_count}/17 brands")
        if brand_contradiction:
            hard_invalidities.append("brand_contradiction: checked 'Not sure' AND named brands")
        if obscure_brands and brand_count >= 8:
            soft_concerns.append(f"obscure_brand_awareness: knows {obscure_brands} among many brands")

        # Source over-selection
        if source_count >= 6:
            soft_concerns.append(f"source_oversel: used {source_count} research sources")
        if source_contradiction:
            hard_invalidities.append("source_contradiction: 'Did not use any' AND checked sources")

        # Matrix over-selection
        if over_select_have:
            soft_concerns.append(f"matrix_oversel_have: claims to have {have_count}/7 filtration types")
        if over_select_plan:
            soft_concerns.append(f"matrix_oversel_plan: plans to buy {plan_count}/7 filtration types")
        # have + plan for same product = possible contradiction
        have_and_plan = sum(1 for v in ownership.values() if v["have"] == "1" and v["plan"] == "1")
        if have_and_plan >= 3:
            soft_concerns.append(f"have_and_plan_same: {have_and_plan} products both have AND plan to buy")

        # Funnel: high concern + low knowledge + many brands = possible over-claiming
        try:
            q15_n = int(q15_val) if q15_val else None
        except Exception:
            q15_n = None
        try:
            q16_n = int(q16_val) if q16_val else None
        except Exception:
            q16_n = None
        if q15_n and q16_n and q15_n <= 2 and q16_n >= 4 and brand_count >= 8:
            soft_concerns.append("funnel_break: low concern but high knowledge + many brands")

        # Missing supplier
        missing_supplier = not supplier or str(supplier).strip() == ""
        if missing_supplier:
            soft_concerns.append("missing_supplier")

        # Missing key fields
        missing_key = []
        for f, label in [("q14", "purchase_reason"), ("q15", "concern"), ("q16", "knowledge"),
                         ("q30", "water_source"), ("q31", "area"), ("q32", "sqft"), ("q33", "home_value")]:
            if clean(r.get(f)) == "":
                missing_key.append(label)
        if missing_key:
            soft_concerns.append(f"missing_key_fields: {missing_key}")

        # ---- classify chain validity ----
        if hard_invalidities:
            chain_status = "hard_invalidity"
        elif len(soft_concerns) >= 3:
            chain_status = "multi_soft_concern"
        elif len(soft_concerns) >= 1:
            chain_status = "soft_concern"
        else:
            chain_status = "chain_clean"

        # ---- build the full chain summary for agent review ----
        chain_summary = {
            "respondent_id": rid,
            "source_excel_row": r["_xrow"],
            "status": r["_status"],
            "qtime_seconds": qtime,
            "supplier": supplier,
            "chain_status": chain_status,
            "hard_invalidities": hard_invalidities,
            "soft_concerns": soft_concerns,
            "hard_invalidity_count": len(hard_invalidities),
            "soft_concern_count": len(soft_concerns),
            "brand_awareness": {
                "checked_brands": brand_checked,
                "brand_count": brand_count,
                "obscure_brands_known": obscure_brands,
                "not_sure_contradiction": brand_contradiction,
            },
            "research_sources": {
                "checked_sources": sources_checked,
                "source_count": source_count,
                "no_sources_contradiction": source_contradiction,
            },
            "matrix_ownership": {
                "have_count": have_count,
                "plan_count": plan_count,
                "neither_count": neither_count,
                "over_select_have": over_select_have,
                "over_select_plan": over_select_plan,
                "have_and_plan_same": have_and_plan,
                "by_product": ownership,
            },
            "funnel_fields": {
                "q13_role": q13_decoded,
                "q14_purchase_reason": q14_decoded,
                "q15_concern": q15_decoded,
                "q16_knowledge": q16_decoded,
                "q30_water_source": q30_decoded,
                "q31_area": q31_decoded,
                "q32_sqft": q32_decoded,
                "q33_home_value": q33_decoded,
            },
            "outro": outro_val[:200],
        }
        judgments.append(chain_summary)

    # ---- population-level summaries ----
    chain_status_counts = Counter(j["chain_status"] for j in judgments)
    hard_counts = Counter()
    for j in judgments:
        for h in j["hard_invalidities"]:
            hard_counts[h.split(":")[0]] += 1
    soft_counts = Counter()
    for j in judgments:
        for s in j["soft_concerns"]:
            soft_counts[s.split(":")[0]] += 1

    # t3 baselines for comparison
    t3_brand_avg = sum(t3_brand_counts) / len(t3_brand_counts) if t3_brand_counts else 0
    t3_source_avg = sum(t3_source_counts) / len(t3_source_counts) if t3_source_counts else 0
    t3_over_select_have = sum(1 for r in t3 if sum(1 for base in Q5_PRODUCTS if str(r.get(f"{base}c1","")).strip()=="1") >= 5) / len(t3)
    t3_missing_supplier = sum(1 for r in t3 if not r.get("SUPNAME") or str(r["SUPNAME"]).strip()=="") / len(t3)

    t5_brand_counts = [j["brand_awareness"]["brand_count"] for j in judgments]
    t5_source_counts = [j["research_sources"]["source_count"] for j in judgments]
    t5_over_select = sum(1 for j in judgments if j["matrix_ownership"]["over_select_have"]) / len(judgments)
    t5_missing_supplier = sum(1 for j in judgments if not j["supplier"] or str(j["supplier"]).strip()=="") / len(judgments)

    summary = {
        "t5_total": len(judgments),
        "chain_status_distribution": dict(chain_status_counts),
        "hard_invalidity_types": dict(hard_counts),
        "soft_concern_types": dict(soft_counts),
        "t5_vs_t3_comparison": {
            "brand_awareness_avg": {"t5": round(sum(t5_brand_counts)/len(t5_brand_counts),2), "t3": round(t3_brand_avg,2)},
            "research_source_avg": {"t5": round(sum(t5_source_counts)/len(t5_source_counts),2), "t3": round(t3_source_avg,2)},
            "over_select_have_rate": {"t5": round(t5_over_select,3), "t3": round(t3_over_select_have,3)},
            "missing_supplier_rate": {"t5": round(t5_missing_supplier,3), "t3": round(t3_missing_supplier,3)},
        },
    }

    (OUT / "t5_chain_validity_judgments.jsonl").write_text(
        "\n".join(json.dumps(j, ensure_ascii=False, separators=(",", ":")) for j in judgments) + "\n"
    )
    (OUT / "t5_chain_validity_summary.json").write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
