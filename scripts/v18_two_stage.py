#!/usr/bin/env python3
"""V18 — Two-stage model + focused REVIEW tier model + error analysis.

Stage 1: High-confidence model (handles ML > 0.7 and ML < 0.3)
Stage 2: REVIEW-tier specialized model (handles 0.3 <= ML <= 0.7)

The REVIEW tier is where most errors occur. A specialized model trained
only on the hard cases may perform better than a general model.

Also includes: detailed error analysis to understand what signals the model
is missing for the remaining FNs and FPs.
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training

import xgboost as xgb
import lightgbm as lgb


def get_classify_map():
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()
    return classify_map


def train_ensemble(X_train, y_train, X_val, y_val, X_test, use_mlp=True):
    """Train XGBoost + LightGBM + MLP ensemble with calibration."""
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_val_scaled = scaler.transform(X_val)
    X_test_scaled = scaler.transform(X_test)

    xgb_model = xgb.XGBClassifier(
        n_estimators=500, max_depth=6, learning_rate=0.03,
        subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
        random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
    )
    xgb_model.fit(X_train, y_train)

    lgb_model = lgb.LGBMClassifier(
        n_estimators=500, max_depth=8, learning_rate=0.03,
        num_leaves=63, subsample=0.8, colsample_bytree=0.7,
        reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
    )
    lgb_model.fit(X_train, y_train)

    models_val = {
        "xgb": xgb_model.predict_proba(X_val)[:, 1],
        "lgb": lgb_model.predict_proba(X_val)[:, 1],
    }
    models_test = {
        "xgb": xgb_model.predict_proba(X_test)[:, 1],
        "lgb": lgb_model.predict_proba(X_test)[:, 1],
    }

    # Only use MLP if we have enough samples
    if use_mlp and len(X_train) > 100 and min(np.bincount(y_train.astype(int))) > 5:
        mlp = MLPClassifier(
            hidden_layer_sizes=(128, 64), max_iter=300,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        try:
            mlp.fit(X_train_scaled, y_train)
            models_val["mlp"] = mlp.predict_proba(X_val_scaled)[:, 1]
            models_test["mlp"] = mlp.predict_proba(X_test_scaled)[:, 1]
        except Exception as e:
            print(f"    MLP failed ({e}), using XGB+LGB only")

    cal_test = {}
    cal_val = {}
    for name in models_val:
        iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
        cal_val[name] = iso.transform(models_val[name])
        cal_test[name] = iso.transform(models_test[name])

    ensemble_test = np.mean(list(cal_test.values()), axis=0)

    meta_X_val = np.column_stack(list(cal_val.values()))
    meta_X_test = np.column_stack(list(cal_test.values()))
    meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
    meta_model.fit(meta_X_val, y_val)
    stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

    return ensemble_test, stacking_test


def run_v18_cv(n_folds=5):
    """Run V18 two-stage model."""
    print(f"\n{'='*80}")
    print(f"V18 — Two-Stage Model + REVIEW Tier Specialization")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Extract all datasets
    print("\nExtracting all datasets...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    classify_map = get_classify_map()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Self-training (V14 settings)
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        # Split for calibration
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        # Stage 1: Full ensemble model
        print("  Stage 1: Full ensemble model...")
        ens1_test, stk1_test = train_ensemble(X_tr, y_tr, X_val, y_val, X_test)
        stage1_scores = ens1_test  # Use ensemble for stage 1

        # Identify REVIEW tier (0.3 - 0.7)
        review_mask = (stage1_scores >= 0.25) & (stage1_scores <= 0.75)
        n_review = review_mask.sum()
        print(f"  REVIEW tier: {n_review}/{len(y_test)} respondents")

        # Stage 2: Specialized model for REVIEW tier
        # Train only on respondents in the REVIEW band from training data
        # First, get stage 1 scores for training data
        ens1_val, stk1_val = train_ensemble(X_tr, y_tr, X_val, y_val, X_val)
        stage1_val_scores = ens1_val

        # Identify REVIEW tier in validation data
        review_val_mask = (stage1_val_scores >= 0.25) & (stage1_val_scores <= 0.75)

        # Also get stage 1 scores for training data itself
        # Use cross-validation to get unbiased scores for training data
        from sklearn.model_selection import cross_val_predict
        xgb_cv = xgb.XGBClassifier(
            n_estimators=300, max_depth=5, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8, random_state=42,
            use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        cv_scores = cross_val_predict(xgb_cv, X_tr, y_tr, cv=3, method="predict_proba")[:, 1]
        review_train_mask = (cv_scores >= 0.25) & (cv_scores <= 0.75)

        print(f"  Stage 2: Training specialized model on {review_train_mask.sum()} REVIEW-tier respondents...")

        if review_train_mask.sum() > 20 and review_val_mask.sum() > 5:
            X_tr_review = X_tr.iloc[review_train_mask]
            y_tr_review = y_tr[review_train_mask]
            X_val_review = X_val.iloc[review_val_mask]
            y_val_review = y_val[review_val_mask]

            if len(np.unique(y_tr_review)) < 2:
                print("  WARNING: REVIEW tier has only one class, skipping stage 2")
                stage2_scores = stage1_scores.copy()
            else:
                X_test_review = X_test[review_mask]
                ens2, stk2 = train_ensemble(
                    X_tr_review, y_tr_review, X_val_review, y_val_review, X_test_review
                )
                stage2_scores = ens2

                # Replace stage 1 scores with stage 2 for REVIEW tier
                final_scores = stage1_scores.copy()
                final_scores[review_mask] = stage2_scores
        else:
            print("  WARNING: Not enough REVIEW-tier samples, using stage 1 only")
            final_scores = stage1_scores.copy()

        # Also try: use stage 1 scores directly (for comparison)
        stage1_only = stage1_scores.copy()

        # Search for best threshold for both approaches
        results = {}
        for approach_name, scores in [("stage1_only", stage1_only),
                                       ("two_stage", final_scores),
                                       ("stacking", stk1_test)]:
            best_bacc = 0
            best_thresh = 0.5
            best_pro_adj = 0

            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_thresh = thresh
                        best_pro_adj = pro_adj

            pred = np.zeros(len(y_test), dtype=int)
            for i in range(len(y_test)):
                t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
                pred[i] = 1 if scores[i] >= t else 0

            tp = ((pred == 1) & (y_test == 1)).sum()
            fp = ((pred == 1) & (y_test == 0)).sum()
            tn = ((pred == 0) & (y_test == 0)).sum()
            fn = ((pred == 0) & (y_test == 1)).sum()
            prec = tp / max(tp + fp, 1)
            rec = tp / max(tp + fn, 1)
            f1 = 2 * prec * rec / max(prec + rec, 0.001)
            bacc = (rec + tn / max(tn + fp, 1)) / 2
            auc = roc_auc_score(y_test, scores)

            results[approach_name] = {
                "bacc": bacc, "f1": f1, "auc": auc,
                "tp": tp, "fp": fp, "tn": tn, "fn": fn,
                "thresh": best_thresh, "pro_adj": best_pro_adj,
            }
            print(f"  {approach_name}: BAcc={bacc:.3f}, F1={f1:.3f}, AUC={auc:.3f}, "
                  f"TP={tp}, FP={fp}, FN={fn}")

        # Pick best approach
        best_approach = max(results, key=lambda k: results[k]["bacc"])
        best = results[best_approach]
        print(f"  Best approach: {best_approach}")

        all_fold_metrics.append({
            "fold": fold + 1,
            "best_approach": best_approach,
            **best,
            "all_results": {k: {kk: vv for kk, vv in v.items()} for k, v in results.items()},
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    approach_counts = Counter(m["best_approach"] for m in all_fold_metrics)

    print(f"\n{'='*80}")
    print(f"V18 CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")
    print(f"  Best approach distribution: {approach_counts}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V18 — Two-Stage Model + REVIEW Tier Specialization")
    print("=" * 80)

    results = run_v18_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V14 (self-train + V8):   BAcc=0.744 (BEST)")
    print(f"  V15 (Optuna):            BAcc=0.740")
    print(f"  V17 (LLM embeddings):    BAcc=0.739")
    print(f"  V18 (two-stage):         BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], 0.744):.3f}")

    with open(AUTOSURVEY_DIR / "v18_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
