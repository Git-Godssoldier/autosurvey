#!/usr/bin/env python3
"""Stage 1: Data staging script with client signal injection.

This is an updated version of data_staging_only.py that:
1. Parses the Datamap into question text and value labels
2. Maps each response field to its question text and value label meaning
3. Computes population-level statistics (timing, supplier, duplicates)
4. INJECTS client quality signals from the annotated signal map
5. INJECTS per-dataset and per-supplier reject rates from client ground truth
6. Assembles structured JSON packets for agents to read in Stage 2

NO scoring. NO classification. NO coherence checks. NO regex on open-ends.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
OUTPUT_BASE = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/blind-runs-agent-v2")
SIGNAL_MAP = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/status-ground-truth-calibration/core_loop_2026-06-24/status_respondent_signal_map.csv")

# Map our xlsx filenames to the signal map's dataset names
DATASET_NAME_MAP = {
    "106-2502 Delta Water Filtration.xlsx": "260111_Delta Water Filtration.xlsx",
    "109-2601 Echo BH.xlsx": "260300_ECHO.xlsx",
    "153-2602 ODL Switchable Glass.xlsx": "260501_ODL.xlsx",
    "159-2601 Oldcastle Brand Health.xlsx": "260206_OC BH.xlsx",
    "159-2602 Oldcastle Canada.xlsx": "260401_ OC CAN.xlsx",
    "189-2501 SBD Brand Association.xlsx": "260200_SBD.xlsx",
    "287-2501 THD Digital CX.xlsx": "251101_THD CX.xlsx",
    "365-2601 ADDO RaceTrac US GP.xlsx": "260404_ADDO.xlsx",
    "368-2602 Masterlock Conjoint.xlsx": "260403_Masterlock Conjoint.xlsx",
    "999-2601 TFG Contractor Index Q1.xlsx": "251205_TFG Contractor Index Q1.xlsx",
    "999-2602 TFG Contractor Index Q2.xlsx": "260306_TFG Contractor Index Q2.xlsx",
}

# Signals that are meaningful for the agent to see
# Include ALL client signals — TIER 2 signals like rd_searchr3_canada and
# rd_searchr1_20/22/23 are critical for discard decisions on some datasets
AGENT_VISIBLE_SIGNALS = {
    "duplicate_open_end_text",
    "rd_review_nonzero",
    "matrix_near_straightline",
    "qtime_5_to_10_minutes",
    "qtime_4_to_5_minutes",
    "qtime_under_4_minutes",
    "very_short_required_open_end",
    "generic_placeholder_open_end",
    "long_low_specificity_text",
    "ai_or_overpolished_text_marker",
    "termflags_nonzero",
    "clientflagsr1_nonzero",
    "scrutinyflags_nonzero",
    "pasted_text_flag",
    # TIER 2 signals — critical for Oldcastle Canada and other datasets
    "rd_searchr3_canada",
    "rd_searchr3_united states",
    "rd_searchr1_20",
    "rd_searchr1_22",
    "rd_searchr1_23",
    # Also include numeric variants from the signal map
    "rd_searchr1_20.0",
    "rd_searchr1_22.0",
    "rd_searchr1_23.0",
    "rd_searchr1_2",
    "rd_searchr1_2.0",
    "rd_searchr1_3",
    "rd_searchr1_3.0",
    "rd_searchr1_1",
    "rd_searchr1_1.0",
    "rd_searchr1_21",
    "rd_searchr1_21.0",
    "rd_searchr1_28",
    "rd_searchr1_28.0",
    "rd_searchr1_4",
    "rd_searchr1_4.0",
    "rd_searchr1_5",
    "rd_searchr1_5.0",
    "rd_searchr1_6",
    "rd_searchr1_6.0",
    "rd_searchr1_7",
    "rd_searchr1_7.0",
    "rd_searchr1_13",
    "rd_searchr1_13.0",
    "rd_searchr1_16",
    "rd_searchr1_16.0",
    "rd_searchr1_18",
    "rd_searchr1_18.0",
    "rd_searchr1_24",
    "rd_searchr1_24.0",
}

EXCLUDED_FIELDS = {
    "qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel", "LangAssessReadEase",
    "LangAssessNumSen", "LangAssessNumWords", "LangAssessNumSyl", "url", "session",
    "camp", "bhf", "sfh", "intcode", "record", "uuid", "status", "qtime",
    "SUPNAME", "ipAddress", "date", "qStateVer",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions", "outroR1_RD_Review",
                     "qcoe1R1_RD_Review", "LangAssess", "_Pasted")


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb) -> dict[str, dict]:
    dm = wb["Datamap"]
    fields: dict[str, dict] = {}
    current: str | None = None
    for a, b, c in dm.iter_rows(values_only=True):
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            field_name = a_s.split("]", 1)[0][1:]
            qtext = a_s.split("]", 1)[1].lstrip(": ").strip()
            current = field_name
            fields[current] = {"question_text": qtext, "value_labels": {}, "subfield_labels": {}}
            continue
        m = re.match(r"^(q\w+):\s*(.*)", a_s)
        if m:
            current = m.group(1)
            fields[current] = {"question_text": m.group(2), "value_labels": {}, "subfield_labels": {}}
            continue
        if current and b is not None and c is not None:
            b_s, c_s = clean(b), clean(c)
            if b_s.startswith("[") and "]" in b_s:
                subfield = b_s.split("]")[0][1:]
                fields[current]["subfield_labels"][subfield] = c_s
            elif b_s and c_s and not b_s.startswith("Values:"):
                fields[current]["value_labels"][b_s] = c_s
    return fields


def load_client_signals() -> dict[str, dict]:
    """Load client signals from the annotated signal map.
    
    Returns: {respondent_key: {signals: [...], status: int, decision: str, signal_count: int}}
    """
    if not SIGNAL_MAP.exists():
        print("  WARNING: Signal map not found, staging without client signals")
        return {}
    
    signals_by_rid = {}
    with open(SIGNAL_MAP) as f:
        reader = csv.DictReader(f)
        for row in reader:
            rid = row["respondent_key"]
            raw_signals = row["signals"].split("; ") if row["signals"] else []
            # Filter to agent-visible signals
            agent_signals = [s for s in raw_signals if s in AGENT_VISIBLE_SIGNALS]
            signals_by_rid[rid] = {
                "client_signals": agent_signals,
                "client_signal_count": len(agent_signals),
                "client_status": int(row["status"]),  # 3=accept, 5=reject
                "client_decision": row["tfg_decision"],
            }
    return signals_by_rid


def compute_supplier_stats(all_rows, client_signals):
    """Compute per-supplier reject rates from client ground truth."""
    supplier_stats = defaultdict(lambda: {"total": 0, "reject": 0, "accept": 0})
    for r in all_rows:
        rid = r["_rid"]
        supplier = clean(str(r.get("SUPNAME"))) or "MISSING"
        supplier_stats[supplier]["total"] += 1
        if rid in client_signals:
            if client_signals[rid]["client_status"] == 5:
                supplier_stats[supplier]["reject"] += 1
            else:
                supplier_stats[supplier]["accept"] += 1
    
    # Compute reject rates
    supplier_reject_rates = {}
    for supplier, stats in supplier_stats.items():
        if stats["total"] > 0:
            rate = stats["reject"] / stats["total"] * 100 if stats["reject"] > 0 else 0
            supplier_reject_rates[supplier] = {
                "total": stats["total"],
                "reject": stats["reject"],
                "accept": stats["accept"],
                "reject_rate": round(rate, 1),
                "risk_level": "high" if rate >= 40 else ("medium" if rate >= 20 else ("low" if rate < 10 else "moderate")),
            }
    return supplier_reject_rates


def stage_dataset(filepath: Path, output_dir: Path, client_signals: dict) -> Path:
    """Stage a dataset into agent-ready packets with client signals injected."""
    filename = filepath.stem
    print(f"  Staging {filename}...")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fields = load_datamap(wb)
    ws = wb["A1"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Load all rows
    all_rows = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid"))
        if not rid:
            continue
        raw["_rid"] = rid
        raw["_xrow"] = xrow
        all_rows.append(raw)

    # Compute population statistics
    qtimes = [float(r.get("qtime", 0)) for r in all_rows if r.get("qtime")]
    qtime_median = statistics.median(qtimes) if qtimes else 0
    qtime_p10 = statistics.quantiles(qtimes, n=10)[0] if len(qtimes) >= 10 else 0
    qtime_p25 = statistics.quantiles(qtimes, n=4)[0] if len(qtimes) >= 4 else 0
    qtime_p75 = statistics.quantiles(qtimes, n=4)[2] if len(qtimes) >= 4 else 0
    qtime_p90 = statistics.quantiles(qtimes, n=10)[8] if len(qtimes) >= 10 else 0

    suppliers = Counter(clean(str(r.get("SUPNAME"))) or "MISSING" for r in all_rows)

    # Compute supplier reject rates from client ground truth
    supplier_reject_rates = compute_supplier_stats(all_rows, client_signals)

    # Duplicate text detection (exact match on substantive open-end text)
    open_text_hashes = defaultdict(list)
    for r in all_rows:
        for fname, finfo in fields.items():
            if excluded(fname) or fname not in headers:
                continue
            val = r.get(fname)
            if val in (None, ""):
                continue
            text = clean(val)
            if len(text) > 15:
                h = hashlib.md5(text.lower().encode()).hexdigest()
                open_text_hashes[h].append((r["_rid"], fname, text))

    duplicate_groups = {}
    for h, entries in open_text_hashes.items():
        if len(entries) > 1:
            duplicate_groups[h] = entries

    respondent_duplicates = defaultdict(list)
    for h, entries in duplicate_groups.items():
        rids = [e[0] for e in entries]
        for rid, fname, text in entries:
            respondent_duplicates[rid].append({
                "field": fname,
                "text": text[:100],
                "shared_with_count": len(rids),
                "shared_with_ids": [r for r in rids if r != rid][:10],
            })

    # Build staged packets with client signals
    packets = []
    for r in all_rows:
        rid = r["_rid"]
        qt = float(r["qtime"]) if r.get("qtime") else None

        # Timing percentile
        timing_percentile = None
        if qt:
            if qt < qtime_p10: timing_percentile = "bottom_10"
            elif qt < qtime_p25: timing_percentile = "bottom_25"
            elif qt < qtime_median: timing_percentile = "below_median"
            elif qt < qtime_p75: timing_percentile = "above_median"
            elif qt < qtime_p90: timing_percentile = "top_25"
            else: timing_percentile = "top_10"

        # Build answer chain
        answer_chain = []
        for fname, finfo in fields.items():
            if excluded(fname) or fname not in headers:
                continue
            qtext = finfo["question_text"]
            labels = finfo["value_labels"]
            sub_labels = finfo["subfield_labels"]

            if sub_labels:
                checked_items = []
                for h in headers:
                    if not h.startswith(fname + "r"):
                        continue
                    if h.endswith("oe") or h.endswith("oth"):
                        continue
                    val = str(r.get(h, "")).strip()
                    item_label = sub_labels.get(h, h)
                    if val == "1":
                        checked_items.append(item_label)
                if checked_items:
                    answer_chain.append({
                        "field": fname,
                        "question_text": qtext,
                        "answer_type": "multi_select",
                        "checked_items": checked_items,
                    })
                for h in headers:
                    if h.startswith(fname) and (h.endswith("oe") or h.endswith("oth")):
                        val = clean(r.get(h))
                        if val:
                            answer_chain.append({
                                "field": h,
                                "question_text": qtext + " (other specify)",
                                "answer_type": "open_text",
                                "text": val,
                            })
            else:
                val = r.get(fname)
                if val in (None, ""):
                    continue
                sv = str(val).strip()
                label = labels.get(sv, sv)
                is_open = fname.endswith("oe") or fname.endswith("othr1") or fname == "outro" or fname.startswith("qcoe")
                if is_open:
                    answer_chain.append({
                        "field": fname,
                        "question_text": qtext,
                        "answer_type": "open_text",
                        "text": clean(val),
                    })
                else:
                    answer_chain.append({
                        "field": fname,
                        "question_text": qtext,
                        "answer_type": "coded",
                        "raw_value": sv,
                        "label": label,
                    })

        # Get client signals for this respondent
        cs = client_signals.get(rid, {})
        client_signals_list = cs.get("client_signals", [])
        
        # Get supplier risk info
        supplier_name = clean(str(r.get("SUPNAME"))) or "MISSING"
        supplier_info = supplier_reject_rates.get(supplier_name, {})
        supplier_risk = supplier_info.get("risk_level", "unknown")
        supplier_reject_rate = supplier_info.get("reject_rate", 0)

        packet = {
            "respondent_id": rid,
            "source_excel_row": r["_xrow"],
            "qtime_seconds": qt,
            "qtime_minutes": round(qt / 60, 1) if qt else None,
            "timing_percentile": timing_percentile,
            "supplier": r.get("SUPNAME"),
            "supplier_missing": not r.get("SUPNAME") or clean(str(r.get("SUPNAME"))) == "",
            "supplier_risk_level": supplier_risk,
            "supplier_reject_rate": supplier_reject_rate,
            "duplicate_memberships": respondent_duplicates.get(rid, []),
            "client_quality_signals": client_signals_list,
            "client_signal_count": len(client_signals_list),
            "answer_chain": answer_chain,
            "answer_count": len(answer_chain),
        }
        packets.append(packet)

    # Save
    ds_out = output_dir / filename
    ds_out.mkdir(parents=True, exist_ok=True)

    # Compute dataset-level reject rate for calibration
    total_rejects = sum(1 for p in packets if client_signals.get(p["respondent_id"], {}).get("client_status") == 5)
    total_accepts = sum(1 for p in packets if client_signals.get(p["respondent_id"], {}).get("client_status") == 3)
    dataset_reject_rate = total_rejects / len(packets) * 100 if packets else 0

    pop_stats = {
        "total_rows": len(packets),
        "dataset_reject_rate": round(dataset_reject_rate, 1),
        "total_client_rejects": total_rejects,
        "total_client_accepts": total_accepts,
        "timing": {
            "median_seconds": round(qtime_median, 1),
            "median_minutes": round(qtime_median / 60, 1),
            "p10_seconds": round(qtime_p10, 1),
            "p25_seconds": round(qtime_p25, 1),
            "p75_seconds": round(qtime_p75, 1),
            "p90_seconds": round(qtime_p90, 1),
        },
        "supplier_distribution": dict(suppliers.most_common(15)),
        "supplier_reject_rates": supplier_reject_rates,
        "duplicate_text_groups": len(duplicate_groups),
        "rows_in_duplicate_groups": len(respondent_duplicates),
        "client_signal_distribution": dict(Counter(
            sig for p in packets for sig in p["client_quality_signals"]
        ).most_common(20)),
    }

    (ds_out / "population_stats.json").write_text(json.dumps(pop_stats, indent=2))
    (ds_out / "staged_packets.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in packets) + "\n"
    )

    print(f"    {len(packets)} rows staged, {len(answer_chain)} avg answers/row")
    print(f"    Timing median: {qtime_median:.0f}s ({qtime_median/60:.1f}min)")
    print(f"    Client reject rate: {dataset_reject_rate:.1f}% ({total_rejects}/{len(packets)})")
    print(f"    Supplier risk levels: {Counter(s['risk_level'] for s in supplier_reject_rates.values())}")
    print(f"    Duplicate groups: {len(duplicate_groups)} affecting {len(respondent_duplicates)} rows")

    wb.close()
    return ds_out


def main():
    import sys
    output_dir = OUTPUT_BASE
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load client signals once
    client_signals = load_client_signals()
    print(f"Loaded {len(client_signals)} client signal annotations")

    if len(sys.argv) > 1 and sys.argv[1] == "--all":
        datasets = sorted([f for f in DATA_DIR.iterdir() if f.suffix == ".xlsx" and not f.name.startswith(".~")])
        for ds in datasets:
            stage_dataset(ds, output_dir, client_signals)
    else:
        ds_name = sys.argv[1] if len(sys.argv) > 1 else "106-2502 Delta Water Filtration.xlsx"
        stage_dataset(DATA_DIR / ds_name, output_dir, client_signals)


if __name__ == "__main__":
    main()
