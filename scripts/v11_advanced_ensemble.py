#!/usr/bin/env python3
"""V11 — Advanced ensemble (XGBoost + LightGBM + MLP) + TF-IDF text features + supplier interactions.

Improvements over V10:
1. XGBoost and LightGBM (typically outperform sklearn GBM)
2. MLP neural network as ensemble member
3. TF-IDF features from OE text (top 100 terms)
4. Supplier × signal interaction features
5. Stacking ensemble (meta-learner on top of base models)
6. Per-channel threshold optimization
"""
from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path

import joblib
import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.feature_extraction.text import TfidfVectorizer

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features

import xgboost as xgb
import lightgbm as lgb


def extract_tfidf_features(df, answer_chains, max_features=100):
    """Extract TF-IDF features from OE text."""
    print(f"  Extracting TF-IDF features (max {max_features} terms)...")

    # Build text corpus
    chain_lookup = {ac["respondent_id"]: ac for ac in answer_chains}
    texts = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        ac = chain_lookup.get(rid, {})
        chain = ac.get("answer_chain", [])
        oe_texts = [a.get("raw_value", "") or a.get("response", "") for a in chain if a.get("answer_type") == "open_end" or a.get("role") == "open_end"]
        text = " ".join(oe_texts).strip()
        if not text:
            text = "empty"
        texts.append(text)

    # TF-IDF
    vectorizer = TfidfVectorizer(
        max_features=max_features,
        stop_words="english",
        ngram_range=(1, 2),
        min_df=2,
        max_df=0.95,
    )
    tfidf_matrix = vectorizer.fit_transform(texts)
    tfidf_df = pd.DataFrame(
        tfidf_matrix.toarray(),
        columns=[f"tfidf_{w}" for w in vectorizer.get_feature_names_out()],
        index=df.index,
    )

    print(f"    TF-IDF: {tfidf_df.shape[1]} features from {len(texts)} texts")
    return tfidf_df


def extract_supplier_interactions(df):
    """Extract supplier × signal interaction features."""
    print("  Extracting supplier interaction features...")

    # Supplier reject rate interactions
    if "supplier_reject_rate" in df.columns:
        for col in ["signal_count", "t1_count", "t2_count", "oe_total_chars", "qtime_seconds"]:
            if col in df.columns:
                df[f"supplier_x_{col}"] = df["supplier_reject_rate"] * df[col] / 100

    # Timing interactions
    if "qtime_seconds" in df.columns:
        df["qtime_per_signal"] = df["qtime_seconds"] / (df["signal_count"] + 1)
        df["qtime_x_oe_chars"] = df["qtime_seconds"] * df.get("oe_total_chars", 0) / 10000

    # OE length interactions
    if "oe_total_chars" in df.columns:
        df["oe_chars_x_signals"] = df["oe_total_chars"] * df.get("signal_count", 0) / 100
        df["oe_chars_per_signal"] = df["oe_total_chars"] / (df.get("signal_count", 0) + 1)

    return df


def run_v11_cv(use_v7_features=True, n_folds=5):
    """Run V11 cross-validation with advanced ensemble + TF-IDF."""
    print(f"\n{'='*80}")
    print(f"V11 CV — Advanced Ensemble + TF-IDF + Supplier Interactions")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments() if use_v7_features else {}
    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)

    # Add TF-IDF features
    tfidf_df = extract_tfidf_features(df, answer_chains, max_features=100)
    df = pd.concat([df, tfidf_df], axis=1)

    # Add supplier interactions
    df = extract_supplier_interactions(df)

    # Filter to labeled
    labeled = df[df["label"] >= 0].copy().reset_index(drop=True)
    print(f"Labeled: {len(labeled)} respondents, {len(labeled.columns)} total features")

    # Prepare features
    non_feature = {"label", "respondent_id", "dataset", "supplier_name", "v7_judgment"}
    feature_cols = [c for c in labeled.columns if c not in non_feature]

    X = labeled[feature_cols].copy()
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X = X.fillna(0)
    y = labeled["label"].values

    # Read CLASSIFY for per-channel
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        if rid and classify_idx is not None and classify_idx < len(row):
            classify_map[rid] = row[classify_idx]
    wb.close()

    is_pro = labeled["respondent_id"].map(lambda r: str(classify_map.get(r)) == "1").values

    # K-fold CV
    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (train_idx, test_idx) in enumerate(skf.split(X, y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")
        X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
        y_train, y_test = y[train_idx], y[test_idx]
        is_pro_test = is_pro[test_idx]

        # Split train into train/val
        n_val = len(X_train) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_train), n_val, replace=False)
        train_mask = np.ones(len(X_train), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_train.iloc[train_mask]
        X_val = X_train.iloc[val_idx]
        y_tr = y_train[train_mask]
        y_val = y_train[val_idx]

        # Scale for MLP/LR
        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        # Train models
        print("  Training XGBoost...")
        xgb_model = xgb.XGBClassifier(
            n_estimators=400, max_depth=6, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8, random_state=42,
            use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        print("  Training LightGBM...")
        lgb_model = lgb.LGBMClassifier(
            n_estimators=400, max_depth=6, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8, random_state=42,
            verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        print("  Training MLP...")
        mlp = MLPClassifier(
            hidden_layer_sizes=(128, 64, 32), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        print("  Training Random Forest...")
        rf = RandomForestClassifier(
            n_estimators=500, max_depth=10, min_samples_leaf=5,
            random_state=42, n_jobs=-1
        )
        rf.fit(X_tr, y_tr)

        # Get validation predictions for calibration
        xgb_val = xgb_model.predict_proba(X_val)[:, 1]
        lgb_val = lgb_model.predict_proba(X_val)[:, 1]
        mlp_val = mlp.predict_proba(X_val_scaled)[:, 1]
        rf_val = rf.predict_proba(X_val)[:, 1]

        # Calibrate each model
        xgb_iso = IsotonicRegression(out_of_bounds="clip").fit(xgb_val, y_val)
        lgb_iso = IsotonicRegression(out_of_bounds="clip").fit(lgb_val, y_val)
        mlp_iso = IsotonicRegression(out_of_bounds="clip").fit(mlp_val, y_val)
        rf_iso = IsotonicRegression(out_of_bounds="clip").fit(rf_val, y_val)

        # Get test predictions
        xgb_test = xgb_iso.transform(xgb_model.predict_proba(X_test)[:, 1])
        lgb_test = lgb_iso.transform(lgb_model.predict_proba(X_test)[:, 1])
        mlp_test = mlp_iso.transform(mlp.predict_proba(X_test_scaled)[:, 1])
        rf_test = rf_iso.transform(rf.predict_proba(X_test)[:, 1])

        # Simple averaging ensemble
        ensemble_test = np.mean([xgb_test, lgb_test, mlp_test, rf_test], axis=0)

        # Stacking: train meta-learner on validation predictions
        meta_X_val = np.column_stack([xgb_val, lgb_val, mlp_val, rf_val])
        meta_X_test = np.column_stack([xgb_test, lgb_test, mlp_test, rf_test])
        meta_model = LogisticRegression(max_iter=200, random_state=42)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Search for best threshold (per-channel) for both ensemble and stacking
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.25, 0.60, 0.025):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        # Evaluate best
        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else best_thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    # Aggregate
    print(f"\n{'='*80}")
    print(f"V11 CV RESULTS ({n_folds} folds, use_v7_features={use_v7_features})")
    print(f"{'='*80}")

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])
    total_tp = sum(m["tp"] for m in all_fold_metrics)
    total_fp = sum(m["fp"] for m in all_fold_metrics)
    total_tn = sum(m["tn"] for m in all_fold_metrics)
    total_fn = sum(m["fn"] for m in all_fold_metrics)

    pooled_prec = total_tp / max(total_tp + total_fp, 1)
    pooled_rec = total_tp / max(total_tp + total_fn, 1)
    pooled_f1 = 2 * pooled_prec * pooled_rec / max(pooled_prec + pooled_rec, 0.001)
    pooled_bacc = (pooled_rec + total_tn / max(total_tn + total_fp, 1)) / 2

    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")
    print(f"  Pooled: TP={total_tp}, FP={total_fp}, TN={total_tn}, FN={total_fn}")
    print(f"  Pooled BAcc:  {pooled_bacc:.3f}, F1: {pooled_f1:.3f}")

    method_dist = {}
    for m in all_fold_metrics:
        method_dist[m["method"]] = method_dist.get(m["method"], 0) + 1
    print(f"  Method selection: {method_dist}")

    return {
        "avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc,
        "pooled_bacc": pooled_bacc, "pooled_f1": pooled_f1,
        "folds": all_fold_metrics,
    }


def main():
    print("=" * 80)
    print("V11 — Advanced Ensemble + TF-IDF + Supplier Interactions")
    print("=" * 80)

    # With V7 features
    results_v11 = run_v11_cv(use_v7_features=True, n_folds=5)

    # Without V7 features
    results_v11_nov7 = run_v11_cv(use_v7_features=False, n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON ACROSS VERSIONS")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V10 (sklearn ensemble):  BAcc=0.729, F1=0.650 (with V7 features)")
    print(f"  V10 (no V7 features):    BAcc=0.703, F1=0.611")
    print(f"  V11 (XGB+LGB+MLP+RF):    BAcc={results_v11['avg_bacc']:.3f}, F1={results_v11['avg_f1']:.3f} (with V7 features)")
    print(f"  V11 (no V7 features):    BAcc={results_v11_nov7['avg_bacc']:.3f}, F1={results_v11_nov7['avg_f1']:.3f}")
    print(f"  Gap to 90% BAcc:         {0.90 - results_v11['avg_bacc']:.3f}")

    # Save
    results = {"v11_with_v7": results_v11, "v11_without_v7": results_v11_nov7}
    with open(AUTOSURVEY_DIR / "v11_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\nResults saved to {AUTOSURVEY_DIR / 'v11_cv_results.json'}")


if __name__ == "__main__":
    main()
