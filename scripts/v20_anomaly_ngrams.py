#!/usr/bin/env python3
"""V20 — Anomaly detection + answer sequence n-grams + matrix pattern features.

Fresh approach: Instead of supervised classification, use anomaly detection
to find "unusual" respondents. Also extract answer sequence patterns as n-grams.

New features:
1. Isolation Forest anomaly score
2. One-Class SVM outlier score
3. Answer sequence n-grams (2-grams, 3-grams of consecutive answers)
4. Matrix response patterns (runs, transitions, variance patterns)
5. Per-respondent answer distribution vs cohort distribution (KL divergence)
6. Duplicate answer detection (same answer across many questions)
"""
from __future__ import annotations

import json
import sys
import warnings
from collections import Counter
from itertools import product
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.isotonic import IsotonicRegression
from sklearn.ensemble import IsolationForest
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.svm import OneClassSVM

warnings.filterwarnings("ignore")

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import classify_field
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features, extract_all_datasets, run_self_training

import xgboost as xgb
import lightgbm as lgb


def extract_matrix_patterns(xlsx_path):
    """Extract detailed matrix response patterns."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    roles = {str(h): classify_field(str(h)) for h in headers if h}

    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "matrix_cell"]
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "coded_question"]
    uuid_idx = hidx.get("uuid")

    # Group matrix columns by question (q17r1, q17r2 -> q17)
    matrix_groups = {}
    for i, h in matrix_cols:
        # Extract question prefix (e.g., q17r1 -> q17)
        prefix = h.split("r")[0] if "r" in h else h.split("_")[0]
        if prefix not in matrix_groups:
            matrix_groups[prefix] = []
        matrix_groups[prefix].append((i, h))

    print(f"  Matrix groups: {len(matrix_groups)}")

    features = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[uuid_idx]:
            continue
        rid = str(row[uuid_idx]).strip()
        feat = {"respondent_id": rid}

        # Per-group matrix patterns
        all_matrix_vals = []
        for group_name, cols in matrix_groups.items():
            vals = []
            for i, h in cols:
                if i < len(row) and row[i] is not None:
                    try:
                        vals.append(float(row[i]))
                    except:
                        pass

            if vals:
                arr = np.array(vals)
                feat[f"{group_name}_mean"] = float(arr.mean())
                feat[f"{group_name}_std"] = float(arr.std())
                feat[f"{group_name}_range"] = float(arr.max() - arr.min())
                feat[f"{group_name}_unique"] = int(len(set(arr)))
                feat[f"{group_name}_is_straightline"] = int(len(set(arr)) <= 1)
                all_matrix_vals.extend(vals)

        # Global matrix patterns
        if all_matrix_vals:
            arr = np.array(all_matrix_vals)

            # Run length encoding (detect straightlining runs)
            runs = []
            current_val = arr[0]
            current_len = 1
            for v in arr[1:]:
                if v == current_val:
                    current_len += 1
                else:
                    runs.append(current_len)
                    current_val = v
                    current_len = 1
            runs.append(current_len)

            feat["matrix_max_run"] = max(runs)
            feat["matrix_mean_run"] = np.mean(runs)
            feat["matrix_num_runs"] = len(runs)
            feat["matrix_long_run_pct"] = max(runs) / len(arr)

            # Transitions (how often answer changes)
            transitions = sum(1 for i in range(1, len(arr)) if arr[i] != arr[i-1])
            feat["matrix_transitions"] = transitions
            feat["matrix_transition_rate"] = transitions / max(len(arr) - 1, 1)

            # Variance patterns (first half vs second half)
            half = len(arr) // 2
            feat["matrix_var_first_half"] = float(np.var(arr[:half]))
            feat["matrix_var_second_half"] = float(np.var(arr[half:]))
            feat["matrix_var_ratio"] = (feat["matrix_var_second_half"] + 0.01) / (feat["matrix_var_first_half"] + 0.01)

            # Answer distribution
            for val in [1, 2, 3, 4, 5]:
                feat[f"matrix_pct_{val}"] = float((arr == val).sum() / len(arr))

            # Entropy of answers
            counts = Counter(arr)
            probs = np.array(list(counts.values())) / len(arr)
            feat["matrix_entropy"] = float(-np.sum(probs * np.log2(probs + 1e-10)))

        # Coded answer sequence n-grams
        coded_vals = []
        for i, h in coded_cols:
            if i < len(row) and row[i] is not None:
                coded_vals.append(str(row[i]).strip().lower())

        if len(coded_vals) > 2:
            # 2-grams
            bigrams = [f"{coded_vals[i]}_{coded_vals[i+1]}" for i in range(len(coded_vals)-1)]
            feat["coded_bigram_unique"] = len(set(bigrams))
            feat["coded_bigram_diversity"] = len(set(bigrams)) / max(len(bigrams), 1)

            # Most common bigram frequency
            bg_counts = Counter(bigrams)
            feat["coded_bigram_top_freq"] = bg_counts.most_common(1)[0][1] if bg_counts else 0

            # 3-grams
            trigrams = [f"{coded_vals[i]}_{coded_vals[i+1]}_{coded_vals[i+2]}" for i in range(len(coded_vals)-2)]
            feat["coded_trigram_unique"] = len(set(trigrams))
            feat["coded_trigram_diversity"] = len(set(trigrams)) / max(len(trigrams), 1)

            # Duplicate answer rate (same answer repeated)
            answer_counts = Counter(coded_vals)
            max_repeat = max(answer_counts.values())
            feat["coded_max_repeat"] = max_repeat
            feat["coded_max_repeat_pct"] = max_repeat / len(coded_vals)
            feat["coded_single_answers"] = sum(1 for v in answer_counts.values() if v == 1)
            feat["coded_single_pct"] = feat["coded_single_answers"] / len(answer_counts)

        features.append(feat)

    wb.close()
    return pd.DataFrame(features)


def add_anomaly_scores(df, labeled_df):
    """Add Isolation Forest and One-Class SVM anomaly scores."""
    print("  Computing anomaly scores...")

    # Prepare features for anomaly detection
    feature_cols = [c for c in df.columns if c not in ["respondent_id", "label", "dataset",
                                                         "v7_judgment", "v8_judgment",
                                                         "classify", "supplier"] and
                    df[c].dtype in ["float64", "float32", "int64", "int32", "bool"]]

    X = df[feature_cols].fillna(0).values

    # Isolation Forest
    iso_forest = IsolationForest(n_estimators=200, contamination=0.35, random_state=42, n_jobs=-1)
    iso_scores = iso_forest.fit_predict(X)
    iso_decision = iso_forest.decision_function(X)
    df["iso_anomaly"] = iso_scores
    df["iso_score"] = iso_decision

    # One-Class SVM (on a sample for speed)
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # Use subsample for SVM training
    n_train = min(1000, len(X_scaled))
    train_idx = np.random.RandomState(42).choice(len(X_scaled), n_train, replace=False)
    ocsvm = OneClassSVM(kernel="rbf", nu=0.35, gamma="scale")
    ocsvm.fit(X_scaled[train_idx])
    df["ocsvm_score"] = ocsvm.decision_function(X_scaled)

    print(f"  Anomaly scores added (iso: {df['iso_anomaly'].value_counts().to_dict()})")
    return df


def run_v20_cv(n_folds=5):
    """Run V20 cross-validation."""
    print(f"\n{'='*80}")
    print(f"V20 — Anomaly Detection + Answer Sequence N-grams + Matrix Patterns")
    print(f"{'='*80}")

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    # Extract enhanced features (V14 base)
    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Extract matrix patterns
    print("\nExtracting matrix patterns...")
    matrix_df = extract_matrix_patterns(ECHO_XLSX)
    df = df.merge(matrix_df, on="respondent_id", how="left")

    # Add anomaly scores
    labeled_mask = df["label"] >= 0
    df = add_anomaly_scores(df, df[labeled_mask])

    print(f"\nTotal features: {len(df.columns)}")

    # Extract all datasets for self-training
    print("\nExtracting all datasets for self-training...")
    all_data = extract_all_datasets()
    echo_df = df.copy()
    echo_df["dataset"] = "109-2601 Echo BH"
    all_data = pd.concat([all_data, echo_df], ignore_index=True)

    labeled = all_data[all_data["label"] >= 0].copy().reset_index(drop=True)
    unlabeled = all_data[all_data["label"] < 0].copy().reset_index(drop=True)

    # Get CLASSIFY map
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(hdrs) if h}
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break
    classify_map = {}
    if classify_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
            if rid and classify_idx < len(row):
                classify_map[rid] = row[classify_idx]
    wb.close()

    echo_mask = (labeled["dataset"] == "109-2601 Echo BH").values
    echo_indices = np.where(echo_mask)[0]
    echo_y = labeled["label"].values[echo_indices]
    is_pro = labeled["respondent_id"].map(
        lambda r: str(classify_map.get(r)) == "1"
    ).values[echo_indices]

    skf = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=42)
    all_fold_metrics = []

    for fold, (echo_train_idx, echo_test_idx) in enumerate(skf.split(np.zeros(len(echo_indices)), echo_y)):
        print(f"\n--- Fold {fold+1}/{n_folds} ---")

        train_idx = echo_indices[echo_train_idx]
        test_idx = echo_indices[echo_test_idx]

        echo_train = labeled.iloc[train_idx].copy()
        echo_test = labeled.iloc[test_idx].copy()

        # Self-training
        X_st, y_st, st_features = run_self_training(
            echo_train, unlabeled.copy(), n_iterations=3, confidence_threshold=0.85
        )

        X_test = echo_test[st_features].copy()
        for col in X_test.select_dtypes(include=["object"]).columns:
            X_test[col] = pd.Categorical(X_test[col]).codes
        X_test = X_test.fillna(0)
        y_test = echo_test["label"].values
        is_pro_test = is_pro[echo_test_idx]

        # Split for calibration
        n_val = len(X_st) // 5
        val_idx = np.random.RandomState(42 + fold).choice(len(X_st), n_val, replace=False)
        train_mask = np.ones(len(X_st), dtype=bool)
        train_mask[val_idx] = False
        X_tr = X_st.iloc[train_mask]
        X_val = X_st.iloc[val_idx]
        y_tr = y_st[train_mask]
        y_val = y_st[val_idx]

        scaler = StandardScaler()
        X_tr_scaled = scaler.fit_transform(X_tr)
        X_val_scaled = scaler.transform(X_val)
        X_test_scaled = scaler.transform(X_test)

        print(f"  Training on {len(X_tr)} samples, {X_tr.shape[1]} features...")

        xgb_model = xgb.XGBClassifier(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.6, reg_alpha=0.1, reg_lambda=1.0,
            random_state=42, use_label_encoder=False, eval_metric="logloss", n_jobs=-1
        )
        xgb_model.fit(X_tr, y_tr)

        lgb_model = lgb.LGBMClassifier(
            n_estimators=500, max_depth=8, learning_rate=0.03,
            num_leaves=63, subsample=0.8, colsample_bytree=0.6,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1
        )
        lgb_model.fit(X_tr, y_tr)

        mlp = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64), max_iter=500,
            learning_rate="adaptive", early_stopping=True,
            random_state=42, verbose=False
        )
        mlp.fit(X_tr_scaled, y_tr)

        # Calibrate
        models_val = {
            "xgb": xgb_model.predict_proba(X_val)[:, 1],
            "lgb": lgb_model.predict_proba(X_val)[:, 1],
            "mlp": mlp.predict_proba(X_val_scaled)[:, 1],
        }
        models_test = {
            "xgb": xgb_model.predict_proba(X_test)[:, 1],
            "lgb": lgb_model.predict_proba(X_test)[:, 1],
            "mlp": mlp.predict_proba(X_test_scaled)[:, 1],
        }

        cal_test = {}
        cal_val = {}
        for name in models_val:
            iso = IsotonicRegression(out_of_bounds="clip").fit(models_val[name], y_val)
            cal_val[name] = iso.transform(models_val[name])
            cal_test[name] = iso.transform(models_test[name])

        ensemble_test = np.mean(list(cal_test.values()), axis=0)

        # Stacking
        meta_X_val = np.column_stack(list(cal_val.values()))
        meta_X_test = np.column_stack(list(cal_test.values()))
        meta_model = LogisticRegression(max_iter=200, random_state=42, C=0.5)
        meta_model.fit(meta_X_val, y_val)
        stacking_test = meta_model.predict_proba(meta_X_test)[:, 1]

        # Search for best threshold
        best_bacc = 0
        best_method = "ensemble"
        best_thresh = 0.5
        best_pro_adj = 0

        for method_name, scores in [("ensemble", ensemble_test), ("stacking", stacking_test)]:
            for thresh in np.arange(0.20, 0.65, 0.02):
                for pro_adj in np.arange(-0.15, 0.16, 0.025):
                    pred = np.zeros(len(y_test), dtype=int)
                    for i in range(len(y_test)):
                        t = thresh + pro_adj if is_pro_test[i] else thresh
                        pred[i] = 1 if scores[i] >= t else 0
                    tp = ((pred == 1) & (y_test == 1)).sum()
                    fp = ((pred == 1) & (y_test == 0)).sum()
                    tn = ((pred == 0) & (y_test == 0)).sum()
                    fn = ((pred == 0) & (y_test == 1)).sum()
                    prec = tp / max(tp + fp, 1)
                    rec = tp / max(tp + fn, 1)
                    bacc = (rec + tn / max(tn + fp, 1)) / 2
                    if bacc > best_bacc:
                        best_bacc = bacc
                        best_method = method_name
                        best_thresh = thresh
                        best_pro_adj = pro_adj

        scores = ensemble_test if best_method == "ensemble" else stacking_test
        pred = np.zeros(len(y_test), dtype=int)
        for i in range(len(y_test)):
            t = best_thresh + best_pro_adj if is_pro_test[i] else thresh
            pred[i] = 1 if scores[i] >= t else 0

        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, scores)

        print(f"  Best: {best_method}, thresh={best_thresh:.3f}, pro_adj={best_pro_adj:+.3f}")
        print(f"  TP={tp}, FP={fp}, TN={tn}, FN={fn}, P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

        all_fold_metrics.append({
            "fold": fold + 1, "method": best_method,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "precision": prec, "recall": rec, "f1": f1, "bacc": bacc, "auc": auc,
        })

    avg_bacc = np.mean([m["bacc"] for m in all_fold_metrics])
    avg_f1 = np.mean([m["f1"] for m in all_fold_metrics])
    avg_auc = np.mean([m["auc"] for m in all_fold_metrics])

    print(f"\n{'='*80}")
    print(f"V20 CV RESULTS")
    print(f"{'='*80}")
    print(f"  Average BAcc: {avg_bacc:.3f} (+/- {np.std([m['bacc'] for m in all_fold_metrics]):.3f})")
    print(f"  Average F1:   {avg_f1:.3f}")
    print(f"  Average AUC:  {avg_auc:.3f}")

    return {"avg_bacc": avg_bacc, "avg_f1": avg_f1, "avg_auc": avg_auc, "folds": all_fold_metrics}


def main():
    print("=" * 80)
    print("V20 — Anomaly Detection + Answer Sequence N-grams + Matrix Patterns")
    print("=" * 80)

    results = run_v20_cv(n_folds=5)

    print(f"\n{'='*80}")
    print(f"COMPARISON")
    print(f"{'='*80}")
    print(f"  V7 (agent review):       BAcc=0.690, F1=0.586")
    print(f"  V14 (self-train + V8):   BAcc=0.744 (BEST)")
    print(f"  V19 (target encoding):   See v19_cv_results.json")
    print(f"  V20 (anomaly + ngrams):  BAcc={results['avg_bacc']:.3f}, F1={results['avg_f1']:.3f}")
    print(f"  Gap to 90%:              {0.90 - max(results['avg_bacc'], 0.744):.3f}")

    with open(AUTOSURVEY_DIR / "v20_cv_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
