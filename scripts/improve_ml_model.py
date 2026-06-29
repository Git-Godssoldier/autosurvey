#!/usr/bin/env python3
"""ML Model Improvement — Retrain with better calibration + new features.

Key improvements:
1. Train on Echo BH specifically (per-dataset model, not cross-dataset)
2. Add V7 agent judgments as semi-supervised labels
3. Isotonic regression calibration
4. Add semantic features (OE text length, word count, specificity score)
5. Add cross-question consistency features
6. Add per-channel/per-class features (CLASSIFY, conditionsAriens)
7. Train/val/test split for proper evaluation

Output: Improved model saved to models/echo_calibrated_model.joblib
"""
from __future__ import annotations

import json
import pickle
import re
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score, classification_report, confusion_matrix,
    f1_score, precision_score, recall_score, roc_auc_score,
    brier_score_loss
)
from sklearn.model_selection import StratifiedShuffleSplit
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC

warnings.filterwarnings("ignore")

# Paths
AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"
V7_JUDGMENTS = ECHO_OUTPUT / "holistic_agent_run_v7" / "agent_judgments.json"
MODEL_DIR = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "models"

# Add skill scripts to path
sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import clean, norm, parse_datamap, classify_field, extract_features_and_chain


def load_ground_truth():
    """Load client ground truth labels."""
    wb = openpyxl.load_workbook(GT_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    hidx = {h: i for i, h in enumerate(headers) if h}

    gt = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        status = row[hidx["status"]]
        if rid:
            if status in (5, "5"):
                gt[rid] = 1  # discard
            elif status in (3, "3"):
                gt[rid] = 0  # keep
    wb.close()
    return gt


def load_v7_judgments():
    """Load V7 agent judgments as additional features."""
    if not V7_JUDGMENTS.exists():
        return {}
    with open(V7_JUDGMENTS) as f:
        judgments = json.load(f)
    return {j["respondent_id"]: j for j in judgments}


def extract_enhanced_features(xlsx_path, gt_labels=None, v7_judgments=None):
    """Extract features from Excel + add enhanced semantic/consistency features."""
    print(f"Extracting enhanced features from {xlsx_path.name}...")

    # Use existing pipeline for base features
    df, datamap, roles, answer_chains = extract_features_and_chain(xlsx_path)

    # Add ground truth labels if available
    if gt_labels:
        df["label"] = df["respondent_id"].map(gt_labels).fillna(-1).astype(int)
        print(f"  Labeled: {(df['label'] >= 0).sum()} / {len(df)}")

    # Add V7 judgment features (semi-supervised)
    if v7_judgments:
        df["v7_judgment"] = df["respondent_id"].map(
            lambda rid: v7_judgments.get(rid, {}).get("agent_judgment", "UNKNOWN")
        )
        df["v7_score"] = df["respondent_id"].map(
            lambda rid: v7_judgments.get(rid, {}).get("agent_score", 0)
        )
        df["v7_judgment_enc"] = df["v7_judgment"].map({"KEEP": 0, "REVIEW": 1, "DISCARD": 2, "UNKNOWN": 1}).fillna(1)
        df["v7_converging_count"] = df["respondent_id"].map(
            lambda rid: v7_judgments.get(rid, {}).get("converging_family_count", 0)
        )
        df["v7_authenticity_risk"] = df["respondent_id"].map(
            lambda rid: v7_judgments.get(rid, {}).get("authenticity_risk", 0.5)
        )
        df["v7_quality_risk"] = df["respondent_id"].map(
            lambda rid: v7_judgments.get(rid, {}).get("quality_discard_risk", 0.5)
        )
        df["v7_client_reject_prob"] = df["respondent_id"].map(
            lambda rid: v7_judgments.get(rid, {}).get("client_reject_probability", 0.5)
        )
        print(f"  V7 features added: {df['v7_judgment'].notna().sum()}")

    # Enhanced semantic features from answer chains
    print("  Adding enhanced semantic features...")
    # Build lookup: respondent_id -> answer chain dict
    chain_lookup = {ac["respondent_id"]: ac for ac in answer_chains}

    semantic_features = []
    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        ac = chain_lookup.get(rid, {})
        chain = ac.get("answer_chain", [])

        # OE text analysis
        oe_texts = [a.get("raw_value", "") or a.get("response", "") for a in chain if a.get("answer_type") == "open_end" or a.get("role") == "open_end"]
        all_oe = " ".join(oe_texts)

        # Specificity score: unique words / total words
        words = re.findall(r"\b\w+\b", all_oe.lower())
        unique_words = set(words)
        specificity = len(unique_words) / max(len(words), 1)

        # Equipment mentions (OPE-specific)
        ope_words = {"mower", "blower", "trimmer", "chainsaw", "snowblower", "edger",
                     "hedge", "tractor", "zero", "turn", "rider", "walk", "behind",
                     "battery", "gas", "electric", "cordless", "backpack", "handheld"}
        ope_mentions = sum(1 for w in words if w in ope_words)

        # Brand mentions
        brands = {"aeg", "ariens", "black", "deck", "bosch", "briggs", "stratton",
                  "cub", "cadet", "deere", "dewalt", "echo", "efco", "ego",
                  "ferris", "grillo", "honda", "husqvarna", "jonsered", "kawasaki",
                  "kress", "kruger", "makita", "milwaukee", "molkop", "poulan",
                  "ryobi", "scag", "snapper", "stihl", "toro", "troy", "bilt",
                  "weedeater", "worx", "yamaha", "zama"}
        brand_mentions = sum(1 for w in words if w in brands)

        # Grounding anchors (temporal, locational, sensory)
        temporal = {"yesterday", "today", "last", "week", "month", "year", "recently",
                    "before", "after", "during", "when", "while", "ago", "since"}
        locational = {"backyard", "front", "yard", "garden", "driveway", "sidewalk",
                      "property", "house", "home", "acre", "lot", "field", "pasture"}
        sensory = {"saw", "felt", "heard", "smelled", "touched", "noticed", "realized",
                   "experienced", "struggled", "enjoyed", "hated", "loved"}
        temporal_count = sum(1 for w in words if w in temporal)
        locational_count = sum(1 for w in words if w in locational)
        sensory_count = sum(1 for w in words if w in sensory)
        grounding = temporal_count + locational_count + sensory_count

        # First person pronouns
        first_person = sum(1 for w in words if w in {"i", "me", "my", "mine", "we", "our", "us"})

        # Cross-question consistency: do OE answers reference each other?
        n_oe = len(oe_texts)
        oe_overlap = 0
        if n_oe > 1:
            for i in range(n_oe):
                for j in range(i + 1, n_oe):
                    w1 = set(re.findall(r"\b\w+\b", oe_texts[i].lower()))
                    w2 = set(re.findall(r"\b\w+\b", oe_texts[j].lower()))
                    if w1 and w2:
                        overlap = len(w1 & w2) / min(len(w1), len(w2))
                        oe_overlap += overlap
            oe_overlap = oe_overlap / max(n_oe * (n_oe - 1) / 2, 1)

        # Answer chain length consistency
        chain_lengths = [len(a.get("response", "")) for a in chain if a.get("response")]
        length_var = np.var(chain_lengths) if len(chain_lengths) > 1 else 0
        length_mean = np.mean(chain_lengths) if chain_lengths else 0

        # Matrix answer diversity (cross-question)
        matrix_answers = []
        for a in chain:
            if a.get("role") == "matrix_cell" and a.get("response"):
                matrix_answers.append(str(a["response"]).strip())
        matrix_unique = len(set(matrix_answers)) / max(len(matrix_answers), 1) if matrix_answers else 1

        semantic_features.append({
            "respondent_id": rid,
            "oe_word_count": len(words),
            "oe_unique_words": len(unique_words),
            "oe_specificity": specificity,
            "ope_mentions": ope_mentions,
            "brand_mentions": brand_mentions,
            "temporal_anchors": temporal_count,
            "locational_anchors": locational_count,
            "sensory_anchors": sensory_count,
            "grounding_total": grounding,
            "first_person_count": first_person,
            "oe_field_count": n_oe,
            "oe_cross_overlap": oe_overlap,
            "chain_length_var": length_var,
            "chain_length_mean": length_mean,
            "matrix_diversity": matrix_unique,
            "oe_total_chars": len(all_oe),
            "oe_avg_chars": len(all_oe) / max(n_oe, 1),
        })

    sem_df = pd.DataFrame(semantic_features)
    df = df.merge(sem_df, on="respondent_id", how="left")
    print(f"  Enhanced features: {len(sem_df.columns) - 1} semantic features added")

    return df, answer_chains


def train_calibrated_model(df, target_col="label"):
    """Train a calibrated model with train/val/test split."""
    # Filter to labeled rows only
    labeled = df[df[target_col] >= 0].copy()
    print(f"\nTraining on {len(labeled)} labeled respondents")
    print(f"  Discards: {labeled[target_col].sum()} ({labeled[target_col].mean():.1%})")
    print(f"  Keeps: {(1 - labeled[target_col]).sum()} ({(1 - labeled[target_col]).mean():.1%})")

    # Prepare features
    non_feature = {target_col, "respondent_id", "dataset", "supplier_name", "v7_judgment"}
    feature_cols = [c for c in labeled.columns if c not in non_feature]

    X = labeled[feature_cols].copy()
    y = labeled[target_col].copy()

    # Encode categoricals
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X = X.fillna(0)

    # Split: 70% train, 15% val, 15% test
    sss1 = StratifiedShuffleSplit(n_splits=1, test_size=0.30, random_state=42)
    train_idx, temp_idx = next(sss1.split(X, y))
    X_train, X_temp = X.iloc[train_idx], X.iloc[temp_idx]
    y_train, y_temp = y.iloc[train_idx], y.iloc[temp_idx]

    sss2 = StratifiedShuffleSplit(n_splits=1, test_size=0.50, random_state=42)
    val_idx, test_idx = next(sss2.split(X_temp, y_temp))
    X_val, X_test = X_temp.iloc[val_idx], X_temp.iloc[test_idx]
    y_val, y_test = y_temp.iloc[val_idx], y_temp.iloc[test_idx]

    print(f"  Train: {len(X_train)} (discard rate: {y_train.mean():.1%})")
    print(f"  Val:   {len(X_val)} (discard rate: {y_val.mean():.1%})")
    print(f"  Test:  {len(X_test)} (discard rate: {y_test.mean():.1%})")

    # Train multiple models
    models = {}

    # 1. Gradient Boosting (like current model)
    print("\nTraining Gradient Boosting...")
    gb = GradientBoostingClassifier(
        n_estimators=300, max_depth=5, learning_rate=0.05,
        subsample=0.8, random_state=42
    )
    gb.fit(X_train, y_train)
    gb_proba = gb.predict_proba(X_test)[:, 1]
    models["gb"] = gb

    # 2. Random Forest
    print("Training Random Forest...")
    rf = RandomForestClassifier(
        n_estimators=500, max_depth=10, min_samples_leaf=5,
        random_state=42, n_jobs=-1
    )
    rf.fit(X_train, y_train)
    rf_proba = rf.predict_proba(X_test)[:, 1]
    models["rf"] = rf

    # 3. Logistic Regression (for calibration baseline)
    print("Training Logistic Regression...")
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    lr = LogisticRegression(max_iter=1000, random_state=42, class_weight="balanced")
    lr.fit(X_train_scaled, y_train)
    lr_proba = lr.predict_proba(X_test_scaled)[:, 1]
    models["lr"] = (lr, scaler)

    # Calibrate each model with isotonic regression on validation set
    print("\nCalibrating models with isotonic regression...")
    calibrated = {}

    for name, model in models.items():
        if name == "lr":
            lr_model, lr_scaler = model
            val_proba = lr_model.predict_proba(lr_scaler.transform(X_val))[:, 1]
            test_proba = lr_proba
        else:
            val_proba = model.predict_proba(X_val)[:, 1]
            test_proba = model.predict_proba(X_test)[:, 1]

        iso = IsotonicRegression(out_of_bounds="clip")
        iso.fit(val_proba, y_val)
        cal_proba = iso.transform(test_proba)

        calibrated[name] = {
            "model": model,
            "calibrator": iso,
            "test_proba": cal_proba,
            "raw_test_proba": test_proba,
        }

        # Evaluate
        pred = (cal_proba >= 0.5).astype(int)
        tp = ((pred == 1) & (y_test == 1)).sum()
        fp = ((pred == 1) & (y_test == 0)).sum()
        tn = ((pred == 0) & (y_test == 0)).sum()
        fn = ((pred == 0) & (y_test == 1)).sum()
        prec = tp / max(tp + fp, 1)
        rec = tp / max(tp + fn, 1)
        f1 = 2 * prec * rec / max(prec + rec, 0.001)
        bacc = (rec + tn / max(tn + fp, 1)) / 2
        auc = roc_auc_score(y_test, cal_proba)
        brier = brier_score_loss(y_test, cal_proba)

        print(f"  {name}: P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}, Brier={brier:.4f}")

    # Ensemble: average of all calibrated models
    print("\nBuilding ensemble...")
    ensemble_proba = np.mean([c["test_proba"] for c in calibrated.values()], axis=0)
    pred = (ensemble_proba >= 0.5).astype(int)
    tp = ((pred == 1) & (y_test == 1)).sum()
    fp = ((pred == 1) & (y_test == 0)).sum()
    tn = ((pred == 0) & (y_test == 0)).sum()
    fn = ((pred == 0) & (y_test == 1)).sum()
    prec = tp / max(tp + fp, 1)
    rec = tp / max(tp + fn, 1)
    f1 = 2 * prec * rec / max(prec + rec, 0.001)
    bacc = (rec + tn / max(tn + fp, 1)) / 2
    auc = roc_auc_score(y_test, ensemble_proba)
    print(f"  Ensemble: P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}, AUC={auc:.3f}")

    # Find best threshold on validation set
    print("\nFinding optimal threshold on validation set...")
    ensemble_val_proba = np.mean([
        c["calibrator"].transform(
            c["model"].predict_proba(X_val)[:, 1] if name != "lr"
            else c["model"][0].predict_proba(c["model"][1].transform(X_val))[:, 1]
        )
        for name, c in calibrated.items()
    ], axis=0)

    best_threshold = 0.5
    best_bacc = 0
    for thresh in np.arange(0.3, 0.7, 0.01):
        pred_val = (ensemble_val_proba >= thresh).astype(int)
        tp_v = ((pred_val == 1) & (y_val == 1)).sum()
        fp_v = ((pred_val == 1) & (y_val == 0)).sum()
        tn_v = ((pred_val == 0) & (y_val == 0)).sum()
        fn_v = ((pred_val == 0) & (y_val == 1)).sum()
        prec_v = tp_v / max(tp_v + fp_v, 1)
        rec_v = tp_v / max(tp_v + fn_v, 1)
        bacc_v = (rec_v + tn_v / max(tn_v + fp_v, 1)) / 2
        if bacc_v > best_bacc:
            best_bacc = bacc_v
            best_threshold = thresh

    print(f"  Best threshold: {best_threshold:.2f} (val BAcc={best_bacc:.3f})")

    # Evaluate with best threshold
    pred = (ensemble_proba >= best_threshold).astype(int)
    tp = ((pred == 1) & (y_test == 1)).sum()
    fp = ((pred == 1) & (y_test == 0)).sum()
    tn = ((pred == 0) & (y_test == 0)).sum()
    fn = ((pred == 0) & (y_test == 1)).sum()
    prec = tp / max(tp + fp, 1)
    rec = tp / max(tp + fn, 1)
    f1 = 2 * prec * rec / max(prec + rec, 0.001)
    bacc = (rec + tn / max(tn + fp, 1)) / 2
    print(f"  Test with best threshold: P={prec:.3f}, R={rec:.3f}, F1={f1:.3f}, BAcc={bacc:.3f}")

    # Save the best model
    # Pick the best individual model or ensemble
    best_model_name = "ensemble"
    best_model_data = {
        "model": {name: c["model"] for name, c in calibrated.items()},
        "calibrator": {name: c["calibrator"] for name, c in calibrated.items()},
        "feature_columns": feature_cols,
        "threshold": best_threshold,
        "supplier_rates": {},
        "global_reject_rate": float(y_train.mean()),
        "type": "ensemble_calibrated",
        "metadata": {
            "sklearn_version": __import__("sklearn").__version__,
            "train_size": len(X_train),
            "val_size": len(X_val),
            "test_size": len(X_test),
            "test_metrics": {
                "precision": float(prec),
                "recall": float(rec),
                "f1": float(f1),
                "balanced_acc": float(bacc),
                "auc": float(auc),
            },
        },
    }

    model_path = MODEL_DIR / "echo_calibrated_model.joblib"
    import joblib
    joblib.dump(best_model_data, model_path)
    print(f"\nModel saved to {model_path}")

    # Also save as pkl for compatibility
    pkl_path = MODEL_DIR / "echo_calibrated_model.pkl"
    with open(pkl_path, "wb") as f:
        pickle.dump(best_model_data, f)

    return {
        "test_metrics": {"precision": prec, "recall": rec, "f1": f1, "balanced_acc": bacc, "auc": auc},
        "best_threshold": best_threshold,
        "model_path": str(model_path),
    }


def main():
    print("=" * 80)
    print("ML MODEL IMPROVEMENT — Calibrated Ensemble with Enhanced Features")
    print("=" * 80)

    # Load ground truth
    print("\n1. Loading ground truth...")
    gt = load_ground_truth()
    print(f"   {len(gt)} labeled respondents ({sum(gt.values())} discards)")

    # Load V7 judgments
    print("\n2. Loading V7 agent judgments...")
    v7 = load_v7_judgments()
    print(f"   {len(v7)} V7 judgments loaded")

    # Extract enhanced features
    print("\n3. Extracting enhanced features...")
    df, chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    print(f"   {len(df)} respondents, {len(df.columns)} total features")

    # Train calibrated model
    print("\n4. Training calibrated ensemble model...")
    results = train_calibrated_model(df)

    print(f"\n{'='*80}")
    print(f"RESULTS")
    print(f"{'='*80}")
    print(f"  Test BAcc: {results['test_metrics']['balanced_acc']:.3f}")
    print(f"  Test F1:   {results['test_metrics']['f1']:.3f}")
    print(f"  Test AUC:  {results['test_metrics']['auc']:.3f}")
    print(f"  Best threshold: {results['best_threshold']:.2f}")
    print(f"  Model saved: {results['model_path']}")

    return results


if __name__ == "__main__":
    main()
