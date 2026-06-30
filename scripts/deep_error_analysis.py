#!/usr/bin/env python3
"""Deep error analysis on V14 — understand what FNs and FPs look like.

Examines the specific respondents that V14 gets wrong and looks for
patterns that could be turned into new features.
"""
from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

AUTOSURVEY_DIR = Path("/Users/jeremyalston/Perfect/autosurvey")
SKILL_SCRIPTS = AUTOSURVEY_DIR / "skills" / "cleaning-survey-quality" / "scripts"
DATA_DIR = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
ECHO_XLSX = DATA_DIR / "109-2601 Echo BH.xlsx"
GT_XLSX = Path("/Users/jeremyalston/Perfect/AutoQuality Pair Copy - Echo/260300_ECHO - client annotated.xlsx")
ECHO_OUTPUT = DATA_DIR / "autosurvey-outputs" / "blind-runs-agent" / "109-2601 Echo BH"

sys.path.insert(0, str(SKILL_SCRIPTS))
from survey_pipeline import extract_features_and_chain, clean, norm
from improve_ml_model import load_ground_truth, load_v7_judgments, extract_enhanced_features
from v14_self_training import load_v8_judgments, add_v8_features


def deep_error_analysis():
    """Analyze V14 errors in detail."""
    print("=" * 80)
    print("DEEP ERROR ANALYSIS ON V14")
    print("=" * 80)

    gt = load_ground_truth()
    v7 = load_v7_judgments()
    v8 = load_v8_judgments()

    df, answer_chains = extract_enhanced_features(ECHO_XLSX, gt, v7)
    df = add_v8_features(df, v8)
    df["label"] = df["respondent_id"].map(gt).fillna(-1).astype(int)

    # Load the raw Excel for detailed analysis
    wb = openpyxl.load_workbook(ECHO_XLSX, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Read all rows
    raw_rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = str(row[hidx["uuid"]]).strip() if row[hidx["uuid"]] else None
        if rid:
            raw_rows[rid] = row

    # Get CLASSIFY
    classify_idx = None
    for h, i in hidx.items():
        if h and "CLASSIFY" in str(h).upper():
            classify_idx = i
            break

    # Get supplier
    supplier_idx = hidx.get("supplierName") or hidx.get("SupplierName")

    # Get markers
    markers_idx = hidx.get("markers")

    # Get all OE columns
    from survey_pipeline import classify_field
    roles = {str(h): classify_field(str(h)) for h in headers if h}
    oe_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "open_end"]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "matrix_cell"]
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "coded_question"]

    print(f"\nOE columns: {len(oe_cols)}")
    print(f"Matrix columns: {len(matrix_cols)}")
    print(f"Coded columns: {len(coded_cols)}")

    # Analyze each labeled respondent
    labeled = df[df["label"] >= 0].copy()
    chain_lookup = {ac["respondent_id"]: ac for ac in answer_chains}

    # Get V7 and V8 judgments
    v7_judgments = {j["respondent_id"]: j for j in v7.values()} if isinstance(v7, list) else v7
    v8_judgments = {j["respondent_id"]: j for j in v8.values()} if isinstance(v8, list) else v8

    # Categorize respondents
    categories = {
        "tp": [], "fp": [], "fn": [], "tn": [],
        "v7_correct": [], "v7_wrong": [],
        "v8_correct": [], "v8_wrong": [],
        "both_wrong_discard": [], "both_wrong_keep": [],
    }

    for idx, row in labeled.iterrows():
        rid = row["respondent_id"]
        client_label = row["label"]
        v7_j = v7.get(rid, {}).get("agent_judgment", "UNKNOWN")
        v8_j = v8.get(rid, {}).get("agent_judgment", "UNKNOWN")

        v7_pred = 1 if v7_j == "DISCARD" else 0
        v8_pred = 1 if v8_j == "DISCARD" else 0
        v7_correct = (v7_pred == client_label)
        v8_correct = (v8_pred == client_label)

        if v7_correct:
            categories["v7_correct"].append(rid)
        else:
            categories["v7_wrong"].append(rid)

        if v8_correct:
            categories["v8_correct"].append(rid)
        else:
            categories["v8_wrong"].append(rid)

        if not v7_correct and not v8_correct:
            if client_label == 1:
                categories["both_wrong_discard"].append(rid)  # Both missed a discard
            else:
                categories["both_wrong_keep"].append(rid)  # Both falsely discarded

    print(f"\n=== AGENT ACCURACY ===")
    print(f"V7 correct: {len(categories['v7_correct'])}/{len(labeled)} ({len(categories['v7_correct'])/len(labeled)*100:.1f}%)")
    print(f"V8 correct: {len(categories['v8_correct'])}/{len(labeled)} ({len(categories['v8_correct'])/len(labeled)*100:.1f}%)")
    print(f"Both wrong (missed discard): {len(categories['both_wrong_discard'])}")
    print(f"Both wrong (false discard): {len(categories['both_wrong_keep'])}")

    # Analyze the "both wrong" cases — these are the hardest
    print(f"\n=== HARDEST CASES (BOTH V7 AND V8 WRONG) ===")

    for category, rids in [("Missed Discards", categories["both_wrong_discard"]),
                            ("False Discards", categories["both_wrong_keep"])]:
        print(f"\n--- {category} ({len(rids)} respondents) ---")

        # Analyze OE text for these respondents
        oe_lengths = []
        oe_texts = []
        supplier_counts = Counter()
        classify_counts = Counter()
        qtime_vals = []
        signal_counts = []
        ml_scores = []

        for rid in rids:
            ac = chain_lookup.get(rid, {})
            chain = ac.get("answer_chain", [])
            oe_items = [a for a in chain if a.get("answer_type") == "open_end"]
            oe_text = " ".join(a.get("raw_value", "") for a in oe_items)
            oe_lengths.append(len(oe_text))
            oe_texts.append(oe_text[:200])

            raw = raw_rows.get(rid, ())
            if classify_idx and classify_idx < len(raw):
                classify_counts[str(raw[classify_idx])] += 1
            if supplier_idx and supplier_idx < len(raw):
                supplier_counts[str(raw[supplier_idx])] += 1

            qtime_vals.append(ac.get("qtime_seconds", 0))
            signal_counts.append(ac.get("signal_count", 0))

            row_data = labeled[labeled["respondent_id"] == rid].iloc[0]
            ml_scores.append(row_data.get("v7_client_reject_prob", 0.5))

        print(f"  OE length: mean={np.mean(oe_lengths):.0f}, median={np.median(oe_lengths):.0f}")
        print(f"  QTime: mean={np.mean(qtime_vals):.0f}s, median={np.median(qtime_vals):.0f}s")
        print(f"  Signal count: mean={np.mean(signal_counts):.1f}")
        print(f"  V7 reject prob: mean={np.mean(ml_scores):.3f}")
        print(f"  CLASSIFY: {dict(classify_counts.most_common(5))}")
        print(f"  Suppliers: {dict(supplier_counts.most_common(5))}")

        # Show some example OE texts
        print(f"\n  Example OE texts (first 5):")
        for i, text in enumerate(oe_texts[:5]):
            print(f"    [{i}] {text[:150]}")

    # Analyze what features differ between TPs, FPs, FNs, TNs
    print(f"\n=== FEATURE COMPARISON: TPs vs FPs vs FNs vs TNs ===")

    # Use V7 as proxy for V14 (since V14 uses V7 features)
    for idx, row in labeled.iterrows():
        rid = row["respondent_id"]
        client_label = row["label"]
        v7_j = v7.get(rid, {}).get("agent_judgment", "UNKNOWN")
        v7_pred = 1 if v7_j == "DISCARD" else 0

        if v7_pred == 1 and client_label == 1:
            categories["tp"].append(rid)
        elif v7_pred == 1 and client_label == 0:
            categories["fp"].append(rid)
        elif v7_pred == 0 and client_label == 1:
            categories["fn"].append(rid)
        else:
            categories["tn"].append(rid)

    features_to_compare = [
        "oe_total_chars", "qtime_seconds", "signal_count", "t1_count",
        "matrix_straightline", "answer_entropy", "supplier_reject_rate",
        "oe_specificity", "grounding_total", "first_person_count",
        "ope_mentions", "brand_mentions", "oe_word_count",
        "v7_client_reject_prob", "v8_client_reject_prob",
    ]

    print(f"\n{'Feature':<30} {'TP':>10} {'FP':>10} {'FN':>10} {'TN':>10}")
    print("-" * 70)

    for feat in features_to_compare:
        if feat not in labeled.columns:
            continue
        vals = {}
        for cat, rids in [("tp", categories["tp"]), ("fp", categories["fp"]),
                          ("fn", categories["fn"]), ("tn", categories["tn"])]:
            cat_vals = labeled[labeled["respondent_id"].isin(rids)][feat]
            vals[cat] = cat_vals.mean() if len(cat_vals) > 0 else 0

        print(f"{feat:<30} {vals['tp']:>10.2f} {vals['fp']:>10.2f} {vals['fn']:>10.2f} {vals['tn']:>10.2f}")

    # Look at specific question patterns that might distinguish FNs from TPs
    print(f"\n=== PER-QUESTION ANALYSIS: FNs vs TPs ===")
    print("(What questions do FNs answer differently from TPs?)")

    fn_rids = categories["fn"]
    tp_rids = categories["tp"]

    # Compare answers to specific questions
    question_diffs = []
    for i, h in enumerate(headers):
        if not h or i >= len(next(iter(raw_rows.values()), [])):
            continue
        role = roles.get(str(h), "")
        if role not in ("coded_question", "matrix_cell", "demographic"):
            continue

        fn_vals = []
        tp_vals = []
        for rid in fn_rids:
            raw = raw_rows.get(rid)
            if raw and i < len(raw) and raw[i] is not None:
                fn_vals.append(str(raw[i]))
        for rid in tp_rids:
            raw = raw_rows.get(rid)
            if raw and i < len(raw) and raw[i] is not None:
                tp_vals.append(str(raw[i]))

        if len(fn_vals) > 5 and len(tp_vals) > 5:
            fn_dist = Counter(fn_vals)
            tp_dist = Counter(tp_vals)
            # Compute distribution difference
            all_vals = set(fn_dist.keys()) | set(tp_dist.keys())
            diff = sum(abs(fn_dist.get(v, 0) / len(fn_vals) - tp_dist.get(v, 0) / len(tp_vals)) for v in all_vals) / 2
            if diff > 0.15:
                question_diffs.append((h, diff, fn_dist.most_common(3), tp_dist.most_common(3)))

    question_diffs.sort(key=lambda x: -x[1])
    print(f"\nQuestions with biggest FN vs TP distribution differences:")
    for h, diff, fn_top, tp_top in question_diffs[:15]:
        print(f"  {h} (diff={diff:.3f})")
        print(f"    FN: {fn_top}")
        print(f"    TP: {tp_top}")

    # Look at the markers field
    print(f"\n=== MARKERS FIELD ANALYSIS ===")
    if markers_idx is not None:
        fn_markers = Counter()
        tp_markers = Counter()
        for rid in fn_rids:
            raw = raw_rows.get(rid)
            if raw and markers_idx < len(raw) and raw[markers_idx]:
                marker = str(raw[markers_idx])
                if "bad:" in marker:
                    fn_markers["has_bad"] += 1
                elif "qualified" in marker:
                    fn_markers["qualified"] += 1
                else:
                    fn_markers["other"] += 1
        for rid in tp_rids:
            raw = raw_rows.get(rid)
            if raw and markers_idx < len(raw) and raw[markers_idx]:
                marker = str(raw[markers_idx])
                if "bad:" in marker:
                    tp_markers["has_bad"] += 1
                elif "qualified" in marker:
                    tp_markers["qualified"] += 1
                else:
                    tp_markers["other"] += 1

        print(f"  FN markers: {dict(fn_markers)}")
        print(f"  TP markers: {dict(tp_markers)}")

    # Look at RD_Search values
    print(f"\n=== RD_SEARCH ANALYSIS ===")
    rd_cols = [(i, h) for i, h in enumerate(headers) if h and "RD_Search" in str(h)]
    print(f"  RD_Search columns: {len(rd_cols)}")

    for i, h in rd_cols[:5]:
        fn_vals = []
        tp_vals = []
        for rid in fn_rids:
            raw = raw_rows.get(rid)
            if raw and i < len(raw) and raw[i] is not None:
                fn_vals.append(str(raw[i]))
        for rid in tp_rids:
            raw = raw_rows.get(rid)
            if raw and i < len(raw) and raw[i] is not None:
                tp_vals.append(str(raw[i]))

        fn_dist = Counter(fn_vals)
        tp_dist = Counter(tp_vals)
        print(f"  {h}:")
        print(f"    FN: {fn_dist.most_common(5)}")
        print(f"    TP: {tp_dist.most_common(5)}")

    # Look at LangAssess columns
    print(f"\n=== LANGASSESS ANALYSIS ===")
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and "LangAssess" in str(h)]
    print(f"  LangAssess columns: {len(lang_cols)}")

    for i, h in lang_cols[:10]:
        fn_vals = []
        tp_vals = []
        tn_vals = []
        for rid in fn_rids:
            raw = raw_rows.get(rid)
            if raw and i < len(raw) and raw[i] is not None:
                try:
                    fn_vals.append(float(raw[i]))
                except:
                    pass
        for rid in tp_rids:
            raw = raw_rows.get(rid)
            if raw and i < len(raw) and raw[i] is not None:
                try:
                    tp_vals.append(float(raw[i]))
                except:
                    pass
        for rid in categories["tn"]:
            raw = raw_rows.get(rid)
            if raw and i < len(raw) and raw[i] is not None:
                try:
                    tn_vals.append(float(raw[i]))
                except:
                    pass

        if fn_vals and tp_vals and tn_vals:
            print(f"  {h}: FN={np.mean(fn_vals):.2f}, TP={np.mean(tp_vals):.2f}, TN={np.mean(tn_vals):.2f}")

    wb.close()

    # Summary of potential new features
    print(f"\n{'='*80}")
    print(f"POTENTIAL NEW FEATURES IDENTIFIED")
    print(f"{'='*80}")
    print(f"1. Per-question answer values (not just aggregates) — {len(question_diffs)} questions with FN/TP differences > 0.15")
    print(f"2. Markers field — quota/bad flags")
    print(f"3. RD_Search raw values (not just binary flags)")
    print(f"4. LangAssess detailed columns")
    print(f"5. Supplier × question interaction patterns")
    print(f"6. Both-wrong cases: {len(categories['both_wrong_discard'])} missed discards, {len(categories['both_wrong_keep'])} false discards")


if __name__ == "__main__":
    deep_error_analysis()
