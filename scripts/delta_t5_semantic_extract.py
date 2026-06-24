#!/usr/bin/env python3
"""Extract status=5 (rejected) Delta population with full response chains for semantic review.

Stages evidence only. The agent derives meaning and weight per semantic-signal-expansion.md.
Read-only: does not modify source workbooks.
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ANNOTATED = Path("/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer/260111_Delta Water Filtration.xlsx")
OUT = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/agent-native-delta-transfer-2026-06-24/.autosurvey-internal/t5_semantic")
OUT.mkdir(parents=True, exist_ok=True)

EXCLUDED_FIELDS = {
    "qc", "TERMFLAGS", "qc5", "qc5_Pasted", "LangAssessReadLevel", "LangAssessReadEase",
    "LangAssessNumSen", "LangAssessNumWords", "LangAssessNumSyl", "url", "session",
    "camp", "bhf", "sfh", "intcode",
}
EXCLUDED_PREFIXES = ("RD_", "noanswer", "qc5R1_", "conditions")


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def excluded(name: str) -> bool:
    return name in EXCLUDED_FIELDS or any(name.startswith(p) for p in EXCLUDED_PREFIXES)


def load_datamap(wb) -> dict[str, list[str]]:
    qmap: dict[str, list[str]] = {}
    current: str | None = None
    for a, b, c in wb["Datamap"].iter_rows(values_only=True):
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            current = a_s.split("]", 1)[0][1:]
            qmap[current] = [a_s]
            continue
        if current and (a_s or b is not None or c is not None):
            if a_s:
                qmap[current].append(a_s)
            if b is not None or c is not None:
                qmap[current].append(f"{b} | {c}")
    return qmap


def field_text(field: str, qmap) -> str:
    lines = qmap.get(field, [])
    if not lines:
        return field
    first = lines[0]
    if "] | " in first:
        return first.split("] | ", 1)[1]
    if "]: " in first:
        return first.split("]: ", 1)[1]
    return first


def value_text(field: str, value, qmap) -> str:
    if value in (None, ""):
        return ""
    sv = str(value)
    for line in qmap.get(field, []):
        if " | " not in line:
            continue
        left, right = line.split(" | ", 1)
        if clean(left) == sv:
            return f"{sv}={right}"
    return sv


def field_role(name: str, qmap) -> str:
    low = f"{name} {field_text(name, qmap)}".lower()
    if re.search(r"open|other|specify|remember|why|describe|explain|comment", low):
        return "open_end_or_other_specify"
    if re.search(r"recommend|likely|satisf|importance|attribute|scale|agree", low):
        return "matrix_or_scale"
    if re.search(r"aware|consider|purchase|use|own|brand|product|supplier|store|advert", low):
        return "funnel_or_entity"
    if re.search(r"age|gender|state|income|education|employ|region|home", low):
        return "persona_or_demographic"
    return "closed_or_routed"


def response_type(name: str, qmap) -> str:
    lines = qmap.get(name, [])
    if any("Open text" in x for x in lines):
        return "open_text"
    if any("Open numeric" in x for x in lines):
        return "open_numeric"
    if re.search(r"r\d+c\d+", name):
        return "matrix_cell"
    if len([x for x in lines if " | " in x]) > 0:
        return "coded"
    return "raw"


def main() -> None:
    wb = openpyxl.load_workbook(ANNOTATED, read_only=True, data_only=True)
    qmap = load_datamap(wb)
    ws = wb["A1"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    si = headers.index("status")
    ui = headers.index("uuid")

    # Open-end fields and key context fields
    open_fields = [h for h in headers if not excluded(h) and field_role(h, qmap) == "open_end_or_other_specify"]
    matrix_bases: set[str] = set()
    for h in headers:
        if not excluded(h) and response_type(h, qmap) == "matrix_cell":
            matrix_bases.add(re.sub(r"c\d+$", "", h))

    # population open-text duplicate detection (across full pop)
    open_counts: Counter = Counter()
    all_rows = []
    for xrow, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: norm(v) for h, v in zip(headers, values) if h}
        rid = clean(raw.get("uuid"))
        status = raw.get("status")
        if not rid or status is None:
            continue
        raw["_rid"] = rid
        raw["_xrow"] = xrow
        raw["_status"] = int(status)
        all_rows.append(raw)
        for f in open_fields:
            sv = clean(raw.get(f))
            key = re.sub(r"[^a-z0-9 ]+", "", sv.lower())
            if len(key) >= 3:
                open_counts[key] += 1

    # status-3 vs status-5 duplicate-open cohort (rejected-only duplicates)
    by_status: dict[int, set[str]] = {3: set(), 5: set()}
    for r in all_rows:
        for f in open_fields:
            sv = clean(r.get(f))
            key = re.sub(r"[^a-z0-9 ]+", "", sv.lower())
            if len(key) >= 3:
                by_status[r["_status"]].add(key)
    rejected_only = {k for k in open_counts if k in by_status[5] and k not in by_status[3]}

    def build_packet(r: dict) -> dict:
        rid = r["_rid"]
        status = r["_status"]
        try:
            qtime = float(r.get("qtime"))
        except Exception:
            qtime = None
        opens = []
        matrices: dict[str, list] = defaultdict(list)
        decoded_closed = []
        words = 0
        dup_opens = []
        rejected_dup_opens = []
        for f in headers:
            if excluded(f) or f in ("uuid", "status"):
                continue
            v = r.get(f)
            if v in (None, ""):
                continue
            sv = clean(v)
            role = field_role(f, qmap)
            rt = response_type(f, qmap)
            entry = {
                "field": f,
                "question": field_text(f, qmap),
                "role": role,
                "response_type": rt,
                "raw_value": v,
                "decoded_value": value_text(f, v, qmap),
            }
            if role == "open_end_or_other_specify":
                opens.append(entry)
                words += len(re.findall(r"\w+", sv))
                key = re.sub(r"[^a-z0-9 ]+", "", sv.lower())
                if len(key) >= 3 and open_counts[key] > 1:
                    dup_opens.append({"field": f, "count": open_counts[key], "text": sv})
                # rejected-only DUPLICATE: appears >1 time overall AND only in status=5 rows
                if key in rejected_only and open_counts[key] > 1:
                    rejected_dup_opens.append({"field": f, "count": open_counts[key], "text": sv})
            else:
                decoded_closed.append(entry)
            if rt == "matrix_cell":
                base = re.sub(r"c\d+$", "", f)
                matrices[base].append(v)
        # matrix straightline summary
        matrix_summary = {}
        for base, vals in matrices.items():
            xs = [clean(v) for v in vals if v not in (None, "")]
            if not xs:
                continue
            counts = Counter(xs)
            matrix_summary[base] = {
                "question": field_text(base, qmap),
                "answered": len(xs),
                "unique": len(counts),
                "modal_share": round(counts.most_common(1)[0][1] / len(xs), 3),
                "longest_run": max((sum(1 for _ in g) for k, g in __import__("itertools").groupby(xs)), default=0),
            }
        return {
            "respondent_id": rid,
            "source_excel_row": r["_xrow"],
            "status": status,
            "timing": {"qtime_seconds": qtime, "date": r.get("date")},
            "technical": {"supplier": r.get("SUPNAME"), "ip": r.get("ipAddress")},
            "open_ends_verbatim": opens,
            "decoded_non_open_answers": decoded_closed[:200],
            "matrix_statistics": matrix_summary,
            "population_context": {
                "open_end_word_count": words,
                "duplicate_open_count": len(dup_opens),
                "duplicate_opens": dup_opens[:15],
                "rejected_only_duplicate_opens": rejected_dup_opens[:15],
            },
        }

    t5 = [build_packet(r) for r in all_rows if r["_status"] == 5]
    t3 = [build_packet(r) for r in all_rows if r["_status"] == 3]

    (OUT / "t5_population_semantic_packets.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in t5) + "\n"
    )
    # guardrail sample: 80 status-3 rows spread across suppliers + qtime
    t3_sorted = sorted(t3, key=lambda p: (p["technical"].get("supplier") or "", p["timing"].get("qtime_seconds") or 0))
    step = max(1, len(t3_sorted) // 80)
    guard = t3_sorted[::step][:80]
    (OUT / "t3_guardrail_sample.ndjson").write_text(
        "\n".join(json.dumps(p, ensure_ascii=False, separators=(",", ":")) for p in guard) + "\n"
    )

    # question-set authenticity map source: open-end fields with prompt text
    open_map = []
    for f in open_fields:
        open_map.append({
            "field": f,
            "question": field_text(f, qmap),
            "response_type": response_type(f, qmap),
            "value_labels": [l for l in qmap.get(f, []) if " | " in l][:30],
        })
    (OUT / "open_end_question_map.json").write_text(json.dumps(open_map, indent=2, ensure_ascii=False))

    summary = {
        "status_5_count": len(t5),
        "status_3_count": len(t3),
        "guardrail_sample_count": len(guard),
        "open_end_fields": len(open_fields),
        "matrix_bases": sorted(matrix_bases),
        "rejected_only_duplicate_keys": len(rejected_only),
    }
    (OUT / "extract_summary.json").write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
