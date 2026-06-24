import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

SCRIPT_PATH = (
    Path(__file__).resolve().parents[1]
    / "skills"
    / "reporting-survey-quality"
    / "scripts"
    / "build_continuous_evolution_loop.py"
)
spec = importlib.util.spec_from_file_location("continuous_evolution", SCRIPT_PATH)
continuous = importlib.util.module_from_spec(spec)
assert spec.loader is not None
sys.modules[spec.name] = continuous
spec.loader.exec_module(continuous)


class ContinuousEvolutionTests(unittest.TestCase):
    def test_leakage_denylist_blocks_client_outcome_fields(self):
        cols = ["uuid", "status", "markers", "client_decision", "qtime", "outro", "TERMFLAGS"]
        blocked = continuous.leakage_columns(cols)
        self.assertIn("status", blocked)
        self.assertIn("markers", blocked)
        self.assertIn("client_decision", blocked)
        self.assertIn("TERMFLAGS", blocked)
        self.assertNotIn("outro", blocked)

    def test_confusion_orientation_positive_class_is_rejected(self):
        y = np.array([1, 1, 0, 0])
        pred = np.array([1, 0, 1, 0])
        counts = continuous.confusion_counts(y, pred)
        self.assertEqual(counts, {"tp": 1, "fp": 1, "tn": 1, "fn": 1})

    def test_inner_threshold_returns_valid_threshold_and_orientation(self):
        rows = []
        for dataset in ["a", "b"]:
            for i in range(20):
                label = 1 if i < 10 else 0
                row = {"dataset_id": dataset, "client_label": label, "open_chain_hash": f"{dataset}-{i}"}
                for feature in continuous.FEATURE_COLUMNS:
                    row[feature] = 1.0 - label
                rows.append(row)
        threshold, invert = continuous.inner_threshold(pd.DataFrame(rows))
        self.assertIsInstance(invert, bool)
        self.assertGreaterEqual(threshold, 0.0)
        self.assertLessEqual(threshold, 1.0)

    def test_public_workbook_has_exactly_two_visible_sheets(self):
        data = pd.DataFrame(
            [
                {
                    "record": "1",
                    "uuid": "abc",
                    "autosurvey_original_response_json": "{}",
                    **{col: "" for col in continuous.PUBLIC_APPEND_COLUMNS},
                }
            ]
        )
        data["autosurvey_dataset"] = "demo"
        data["autosurvey_respondent_id"] = "abc"
        data["autosurvey_agent_decision"] = "Clean keep"
        data["autosurvey_client_agreement"] = "agree"
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "AUTOSURVEY_RESULTS.xlsx"
            continuous.write_public_workbook(path, data, [["Release status", ""], ["Terminal state", "IMPROVING_NOT_MET"]])
            wb = load_workbook(path, read_only=True)
            self.assertEqual(wb.sheetnames, ["Labeled Rows", "Dashboard"])
            self.assertEqual([ws.sheet_state for ws in wb.worksheets], ["visible", "visible"])

    def test_terminal_state_refuses_partial_99_claim(self):
        metrics = {
            "accuracy": 0.995,
            "balanced_accuracy": 0.99,
            "macro_f1": 0.99,
            "rejected_recall": 0.80,
            "rejected_precision": 0.99,
            "mcc": 0.96,
        }
        datasets = pd.DataFrame({"dataset_id": ["a"], "accuracy": [0.99], "rejected_recall": [0.80]})
        self.assertEqual(continuous.terminal_state(metrics, datasets), "IMPROVING_NOT_MET")


if __name__ == "__main__":
    unittest.main()
