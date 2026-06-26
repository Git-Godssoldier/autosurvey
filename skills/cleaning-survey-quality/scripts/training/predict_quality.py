#!/usr/bin/env python3
"""Survey Quality Prediction — Combined ML + Agent Rules + Semantic Parsing.

This is the production prediction pipeline that combines:
1. ML model (trained on 11 annotated datasets, 13,388 respondents)
2. Agent analysis rules (TIER 1/2/3 signal classification, discard rules)
3. Semantic structure parsing (Datamap, field roles, answer chains)

The ML model provides a RISK RANKING. The agent rules provide SPECIFIC DISCARD REASONS.
Together, they produce a final determination with justification.

Usage:
    python3 predict_quality.py <xlsx_path> [--output <path>] [--threshold <float>]
"""
from __future__ import annotations

import json
import pickle
import re
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.isotonic import IsotonicRegression

warnings.filterwarnings("ignore")

MODEL_DIR = Path(__file__).parent.parent.parent / "models"

# Signal tiers (from agent analysis training)
T1_SIGNALS = {"termflags_nonzero", "long_low_specificity_text", "ai_or_overpolished_text_marker", "generic_placeholder_open_end"}
T2_SIGNALS = {"rd_searchr3_canada", "rd_searchr1_22", "rd_searchr1_23", "rd_searchr1_20", "qtime_under_dataset_p10",
              "rd_searchr1_22.0", "rd_searchr1_23.0", "rd_searchr1_20.0"}
T3_SIGNALS = {"duplicate_open_end_text", "rd_review_nonzero", "matrix_near_straightline",
              "rd_searchr3_united states", "qtime_5_to_10_minutes", "qtime_under_4_minutes",
              "very_short_required_open_end", "qtime_4_to_5_minutes"}


def clean(v): return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""
def norm(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v


def parse_datamap(wb):
    """Parse the Datamap sheet to get question text and value labels."""
    dm = {}
    if "Datamap" not in wb.sheetnames:
        return dm
    ws = wb["Datamap"]
    for row in ws.iter_rows(values_only=True):
        a, b, c = (row + (None, None, None))[:3]
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            field_name = a_s.split("]", 1)[0][1:]
            qtext = clean(b)
            labels = {}
            dm[field_name] = {"question_text": qtext, "labels": labels}
        elif dm:
            # Find current field
            for k, v in dm.items():
                if v["question_text"] and not v.get("_done"):
                    if a_s and b is not None:
                        v["labels"][str(norm(a))] = clean(b)
                    break
    return dm


def classify_field(field_name):
    """Classify field role for semantic structure parsing."""
    fn = field_name.lower()
    if fn in ("uuid", "record", "rid"): return "id"
    if fn == "status": return "label"
    if fn == "markers": return "quota"
    if fn == "qtime": return "timing"
    if fn in ("supname", "supplier"): return "supplier"
    if fn == "ipaddress": return "ip"
    if fn == "useragent": return "user_agent"
    if fn in ("date", "start_date"): return "timestamp"
    if fn in ("vos", "vbrowser", "vmobiledevice", "vmobileos"): return "technical"
    if fn.startswith("langassess"): return "nlp_metadata"
    if fn.startswith("rd_"): return "review_metadata"
    if fn.startswith("termflags") or "flag" in fn: return "quality_flag"
    if fn.endswith("oe") or fn == "outro" or "qcoe" in fn: return "open_end"
    if re.match(r"q\d+r\d+$", fn): return "matrix_cell"
    if fn in ("age", "qager1", "qgender", "qstate", "region", "q13", "q12",
              "qhomtype", "q2", "q1", "qindustry", "qnumemployees", "q9",
              "qethnicr1", "classify", "channeltracking"): return "demographic"
    if "pasted" in fn: return "paste_flag"
    if fn.startswith("noanswer"): return "no_answer"
    if fn.startswith("conditions"): return "routing"
    if fn.startswith("own") or fn.startswith("possible"): return "product_attr"
    return "coded_question"


def extract_features(filepath, signal_map=None):
    """Extract features from any survey Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Parse datamap
    datamap = parse_datamap(wb)

    # Classify fields
    roles = {str(h): classify_field(str(h)) for h in headers if h}

    # Group columns
    oe_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "open_end"]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "nlp_metadata"]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "matrix_cell"]
    rd_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "review_metadata"]
    flag_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "quality_flag"]
    paste_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "paste_flag"]
    demo_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "demographic"]
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "coded_question"]

    # Collect rows
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid_idx = hidx.get("uuid") or hidx.get("record")
        if rid_idx is None or rid_idx >= len(row): continue
        rid = clean(row[rid_idx])
        if not rid: continue
        rows_data.append((rid, row))

    if not rows_data:
        return pd.DataFrame(), datamap, roles

    # Cross-respondent duplicates
    all_oe, all_ua, all_ip = [], [], []
    for rid, row in rows_data:
        oe = " | ".join(clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i]))
        all_oe.append(oe)
        ua_idx = hidx.get("userAgent")
        all_ua.append(clean(row[ua_idx]) if ua_idx and ua_idx < len(row) else "")
        ip_idx = hidx.get("ipAddress")
        all_ip.append(clean(row[ip_idx]) if ip_idx and ip_idx < len(row) else "")

    oe_ctr = Counter(t.strip().lower() for t in all_oe if t.strip())
    ua_ctr = Counter(ua for ua in all_ua if ua)
    ip_ctr = Counter(ip for ip in all_ip if ip)

    # Build features
    features = []
    for idx, (rid, row) in enumerate(rows_data):
        feat = {"respondent_id": rid}

        # Client signals (if available)
        sm = signal_map or {}
        e = sm.get(rid, {"signals": [], "signal_count": 0})
        sigs = set(e.get("signals", []))
        feat["signal_count"] = e.get("signal_count", 0)
        feat["t1_count"] = len(sigs & T1_SIGNALS)
        feat["t2_count"] = len(sigs & T2_SIGNALS)
        feat["t3_count"] = len(sigs - T1_SIGNALS - T2_SIGNALS)

        # Timing
        qt_idx = hidx.get("qtime")
        qt = 0
        if qt_idx and qt_idx < len(row):
            try: qt = float(row[qt_idx]) if row[qt_idx] else 0
            except: qt = 0
        feat["qtime_seconds"] = qt
        feat["qtime_log"] = np.log1p(qt) if qt > 0 else 0

        # LangAssess
        for i, h in lang_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"lang_{h}"] = float(v) if v else 0
            except: feat[f"lang_{h}"] = 0

        # Open-end text
        oe_texts = [clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i])]
        oe_lens = [len(t) for t in oe_texts]
        oe_words = [len(t.split()) for t in oe_texts]
        feat["oe_count"] = len(oe_texts)
        feat["oe_total_chars"] = sum(oe_lens)
        feat["oe_max_chars"] = max(oe_lens) if oe_lens else 0
        feat["oe_mean_chars"] = np.mean(oe_lens) if oe_lens else 0
        feat["oe_total_words"] = sum(oe_words)
        feat["oe_max_words"] = max(oe_words) if oe_words else 0
        all_oe_text = " ".join(oe_texts).lower()
        words = all_oe_text.split()
        feat["oe_lex_div"] = len(set(words)) / len(words) if words else 0
        feat["oe_has_none"] = 1 if any(w in all_oe_text for w in ["none","n/a","na","nothing","no opinion","no idea"]) else 0
        feat["oe_generic"] = 1 if any(w in all_oe_text for w in ["good","fine","ok","okay","nice","great"]) and oe_words and max(oe_words) <= 3 else 0
        feat["oe_all_caps"] = 1 if any(t.isupper() and len(t) > 5 for t in oe_texts) else 0
        feat["oe_very_short"] = 1 if oe_lens and max(oe_lens) < 10 else 0
        feat["oe_short"] = 1 if oe_lens and max(oe_lens) < 30 else 0

        # Supplier
        sup_idx = hidx.get("SUPNAME")
        sup = clean(row[sup_idx]) if sup_idx and sup_idx < len(row) else ""
        feat["supplier_name"] = sup
        feat["supplier_missing"] = 1 if not sup or sup == "MISSING" else 0
        feat["supplier_is_none"] = 1 if sup == "None" or sup == "" else 0

        # Matrix
        mvals = [norm(row[i]) for i, _ in matrix_cols if i < len(row) and row[i] is not None and row[i] != ""]
        if mvals:
            ur = len(set(mvals)) / len(mvals)
            feat["matrix_unique_ratio"] = ur
            feat["matrix_straightline"] = 1 if ur <= 0.2 and len(mvals) >= 5 else 0
            feat["matrix_near_straightline"] = 1 if ur <= 0.4 and len(mvals) >= 5 else 0
            feat["matrix_count"] = len(mvals)
            feat["matrix_unique_count"] = len(set(mvals))
            vc = Counter(mvals)
            feat["matrix_most_common_freq"] = vc.most_common(1)[0][1] / len(mvals)
        else:
            feat.update({"matrix_unique_ratio": 1.0, "matrix_straightline": 0,
                         "matrix_near_straightline": 0, "matrix_count": 0,
                         "matrix_unique_count": 0, "matrix_most_common_freq": 0})

        # Coded diversity
        cvals = [str(norm(row[i])) for i, _ in coded_cols if i < len(row) and row[i] is not None and row[i] != ""]
        feat["coded_count"] = len(cvals)
        feat["coded_unique_ratio"] = len(set(cvals)) / len(cvals) if cvals else 1.0
        dk = sum(1 for v in cvals if any(x in v.lower() for x in ["don't know","dk","not sure","no answer"]))
        feat["coded_dk_count"] = dk
        feat["coded_dk_ratio"] = dk / len(cvals) if cvals else 0

        # RD_Search
        for i, h in rd_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"rd_{h}"] = float(v) if v else 0
            except: feat[f"rd_{h}"] = 0

        # Flags
        for i, h in flag_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"flag_{h}"] = float(v) if v else 0
            except: feat[f"flag_{h}"] = 0

        # Paste flags
        for i, h in paste_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"paste_{h}"] = float(v) if v else 0
            except: feat[f"paste_{h}"] = 0

        # Duplicates
        oe_t = all_oe[idx]
        feat["oe_dup_count"] = oe_ctr.get(oe_t.strip().lower(), 0) if oe_t.strip() else 0
        feat["oe_is_dup"] = 1 if feat["oe_dup_count"] > 1 else 0
        feat["ua_dup_count"] = ua_ctr.get(all_ua[idx], 0)
        feat["ua_is_dup"] = 1 if feat["ua_dup_count"] > 1 else 0
        feat["ip_dup_count"] = ip_ctr.get(all_ip[idx], 0)
        feat["ip_is_dup"] = 1 if feat["ip_dup_count"] > 1 else 0

        # Demographics
        for i, h in demo_cols:
            feat[f"demo_{h}"] = clean(row[i]) if i < len(row) else ""
        for h in ["vos", "vbrowser", "vmobiledevice", "vmobileos"]:
            i = hidx.get(h)
            feat[f"tech_{h}"] = clean(row[i]) if i and i < len(row) else ""

        feat["dataset"] = filepath.name
        features.append(feat)

    df = pd.DataFrame(features)

    # Dataset-relative z-scores
    for col in ["qtime_seconds", "signal_count"]:
        df[f"{col}_zscore"] = 0.0
        vals = df[col]
        std = vals.std()
        df[f"{col}_zscore"] = (vals - vals.mean()) / (std if std > 0 else 1)

    return df, datamap, roles


def apply_agent_rules(df, datamap, roles):
    """Apply agent analysis rules to generate rule-based determinations.

    These rules come from the exemplar library and agent training.
    They provide SPECIFIC DISCARD REASONS that complement the ML model.
    """
    determinations = []

    for _, row in df.iterrows():
        reasons = []
        det = "KEEP"
        risk_score = 0.0

        # Rule 1: TIER 1 signals = AUTO-DISCARD
        if row.get("t1_count", 0) > 0:
            det = "DISCARD"
            reasons.append(f"TIER 1 signal present ({row['t1_count']} signals)")

        # Rule 2: Matrix straightline + very short open-end
        if row.get("matrix_straightline", 0) == 1 and row.get("oe_very_short", 0) == 1:
            if det != "DISCARD":
                det = "REVIEW"
            reasons.append("Matrix straightline + very short open-end")
            risk_score += 0.3

        # Rule 3: Duplicate open-end text + duplicate IP
        if row.get("oe_is_dup", 0) == 1 and row.get("ip_is_dup", 0) == 1:
            if det != "DISCARD":
                det = "REVIEW"
            reasons.append("Duplicate open-end text + duplicate IP")
            risk_score += 0.2

        # Rule 4: All caps open-end
        if row.get("oe_all_caps", 0) == 1:
            reasons.append("All-caps open-end text")
            risk_score += 0.1

        # Rule 5: Generic/none open-end
        if row.get("oe_has_none", 0) == 1:
            reasons.append("None/N/A in open-end")
            risk_score += 0.1
        if row.get("oe_generic", 0) == 1:
            reasons.append("Generic open-end response")
            risk_score += 0.1

        # Rule 6: Very fast completion (bottom 10% by z-score)
        if row.get("qtime_seconds_zscore", 0) < -1.28:
            reasons.append(f"Very fast completion (z-score: {row['qtime_seconds_zscore']:.2f})")
            risk_score += 0.15

        # Rule 7: High signal count
        if row.get("signal_count", 0) >= 5:
            reasons.append(f"High signal count ({row['signal_count']})")
            risk_score += 0.2
            if det == "KEEP":
                det = "REVIEW"

        # Rule 8: TIER 2 signals present
        if row.get("t2_count", 0) > 0:
            reasons.append(f"TIER 2 signal present ({row['t2_count']} signals)")
            risk_score += 0.15

        determinations.append({
            "respondent_id": row["respondent_id"],
            "rule_determination": det,
            "rule_reasons": reasons,
            "rule_risk_score": risk_score,
        })

    return pd.DataFrame(determinations)


def predict(filepath, output_path=None, threshold=None):
    """Full prediction pipeline: ML model + agent rules + semantic parsing."""
    # Load model
    model_path = MODEL_DIR / "survey_quality_model.pkl"
    if not model_path.exists():
        print(f"ERROR: Model not found at {model_path}")
        print("Run: python3 survey_quality_ml.py train")
        return None

    with open(model_path, "rb") as f:
        model_data = pickle.load(f)

    model = model_data["model"]
    iso = model_data["calibrator"]
    default_threshold = model_data["threshold"]
    train_features = model_data["feature_columns"]
    supplier_rates = model_data["supplier_rates"]
    global_rate = model_data["global_reject_rate"]

    if threshold is None:
        threshold = default_threshold

    print(f"Predicting: {filepath.name}")
    print(f"  Model threshold: {threshold:.3f}")
    print(f"  Features: {len(train_features)}")

    # Extract features
    df, datamap, roles = extract_features(filepath)
    if df.empty:
        print("ERROR: No data extracted")
        return None
    print(f"  Respondents: {len(df)}")
    print(f"  Datamap fields: {len(datamap)}")
    print(f"  Field roles: {Counter(roles.values())}")

    # Add supplier risk
    df["supplier_reject_rate"] = df["supplier_name"].map(supplier_rates).fillna(global_rate)
    df["supplier_x_signals"] = df["supplier_reject_rate"] * df["signal_count"]
    df["supplier_x_t1"] = df["supplier_reject_rate"] * df["t1_count"]
    df["supplier_x_t2"] = df["supplier_reject_rate"] * df["t2_count"]
    df["signals_x_matrix"] = df["signal_count"] * (1 - df["matrix_unique_ratio"])

    # Prepare ML features
    non_feat = {"respondent_id", "label", "dataset", "supplier_name"}
    feat_cols = [c for c in df.columns if c not in non_feat]
    X = df[feat_cols].copy()
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    X = X.fillna(0)

    # Align with training features
    for c in train_features:
        if c not in X.columns: X[c] = 0
    X = X[train_features]

    # ML prediction
    y_proba = model.predict_proba(X)[:, 1]
    y_cal = iso.transform(y_proba)

    # Agent rules
    rule_dets = apply_agent_rules(df, datamap, roles)

    # Combine ML + rules
    results = []
    for i in range(len(df)):
        ml_score = float(y_cal[i])
        ml_pred = "DISCARD" if ml_score >= threshold else "KEEP"

        rule_row = rule_dets.iloc[i]
        rule_det = rule_row["rule_determination"]
        rule_reasons = rule_row["rule_reasons"]
        rule_risk = rule_row["rule_risk_score"]

        # Combined determination:
        # - DISCARD if both ML and rules say DISCARD
        # - DISCARD if ML says DISCARD with high confidence OR rules say DISCARD (TIER 1)
        # - REVIEW if either says REVIEW
        # - KEEP otherwise
        if ml_pred == "DISCARD" and rule_det == "DISCARD":
            final_det = "DISCARD"
            confidence = "HIGH"
        elif rule_det == "DISCARD":  # TIER 1 signal
            final_det = "DISCARD"
            confidence = "HIGH"
        elif ml_pred == "DISCARD" and ml_score > threshold + 0.1:
            final_det = "DISCARD"
            confidence = "MEDIUM"
        elif ml_pred == "DISCARD" or rule_det == "REVIEW":
            final_det = "REVIEW"
            confidence = "MEDIUM"
        else:
            final_det = "KEEP"
            confidence = "LOW"

        all_reasons = []
        if ml_pred == "DISCARD":
            all_reasons.append(f"ML risk score: {ml_score:.3f} (threshold: {threshold:.3f})")
        all_reasons.extend(rule_reasons)

        results.append({
            "respondent_id": df.iloc[i]["respondent_id"],
            "determination": final_det,
            "confidence": confidence,
            "ml_score": round(ml_score, 4),
            "rule_risk_score": round(rule_risk, 4),
            "combined_risk": round(ml_score * 0.6 + rule_risk * 0.4, 4),
            "reasons": all_reasons,
            "supplier": df.iloc[i]["supplier_name"],
            "qtime_seconds": df.iloc[i]["qtime_seconds"],
            "signal_count": df.iloc[i]["signal_count"],
        })

    results_df = pd.DataFrame(results)

    # Summary
    n = len(results_df)
    n_discard = (results_df["determination"] == "DISCARD").sum()
    n_review = (results_df["determination"] == "REVIEW").sum()
    n_keep = (results_df["determination"] == "KEEP").sum()

    print(f"\n{'='*80}")
    print(f"PREDICTION RESULTS: {filepath.name}")
    print(f"{'='*80}")
    print(f"  Total: {n}")
    print(f"  DISCARD: {n_discard} ({n_discard/n:.1%})")
    print(f"  REVIEW:  {n_review} ({n_review/n:.1%})")
    print(f"  KEEP:    {n_keep} ({n_keep/n:.1%})")
    print(f"  Mean ML score: {results_df['ml_score'].mean():.3f}")
    print(f"  Mean combined risk: {results_df['combined_risk'].mean():.3f}")

    # Top features
    if hasattr(model, "feature_importances_"):
        imp = sorted(zip(train_features, model.feature_importances_), key=lambda x: -x[1])[:10]
        print(f"\n  Top ML features:")
        for name, score in imp:
            print(f"    {name}: {score:.4f}")

    # Save
    if output_path is None:
        output_path = filepath.parent / f"{filepath.stem}_quality_predictions.csv"
    results_df.to_csv(output_path, index=False)
    print(f"\n  Saved to: {output_path}")

    # Also save NDJSON
    ndjson_path = output_path.with_suffix(".ndjson")
    with open(ndjson_path, "w") as f:
        for _, row in results_df.iterrows():
            f.write(json.dumps(row.to_dict()) + "\n")
    print(f"  NDJSON: {ndjson_path}")

    return results_df


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 predict_quality.py <xlsx_path> [--output <path>] [--threshold <float>]")
        return

    filepath = Path(sys.argv[1])
    output_path = None
    threshold = None

    for i, arg in enumerate(sys.argv[2:], 2):
        if arg == "--output" and i + 1 < len(sys.argv):
            output_path = Path(sys.argv[i + 1])
        elif arg == "--threshold" and i + 1 < len(sys.argv):
            threshold = float(sys.argv[i + 1])

    predict(filepath, output_path, threshold)


if __name__ == "__main__":
    main()
