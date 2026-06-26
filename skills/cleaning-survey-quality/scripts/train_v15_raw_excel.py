#!/usr/bin/env python3
"""
Experiment 15: Raw Excel per-question answer sequence features.

Goes back to the raw Excel file to extract features that the aggregate
feature extraction misses:
1. Per-question answer entropy (random answering)
2. Answer transition patterns (A->A->A vs A->B->C)
3. Per-open-end question text length variance (some short, some long vs all short)
4. Screener question consistency (do demographics match screener answers)
5. Answer speed per question (if per-question timing available)
6. Number of unique answers per question type

This requires access to the raw Excel file path, which we get from the
dataset name in the dataframe.
"""
from __future__ import annotations
import warnings, numpy as np, pandas as pd
import openpyxl
from collections import Counter
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.isotonic import IsotonicRegression
from sklearn.metrics import f1_score
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import TruncatedSVD
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))
from survey_quality_ml import prepare_features, ANNOTATED_DIR, parse_datamap, classify_field_role, clean, norm
from train_v13_deep import extract_deep_features
from train_v14_per_dataset_best import get_tfidf_features, train_gbm

warnings.filterwarnings("ignore")

# Cache for raw Excel features
_raw_cache = {}

def extract_raw_excel_features(dataset_name, respondent_ids):
    """Extract per-question answer pattern features from raw Excel."""
    if dataset_name in _raw_cache:
        cached = _raw_cache[dataset_name]
        return cached[cached["respondent_id"].isin(respondent_ids)]
    
    fp = ANNOTATED_DIR / dataset_name
    if not fp.exists():
        return pd.DataFrame()
    
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}
    
    datamap = parse_datamap(wb)
    field_roles = {}
    for h in headers:
        if h:
            qt = datamap.get(str(h), {}).get("question_text", "")
            field_roles[str(h)] = classify_field_role(str(h), qt, datamap)
    
    # Group columns by role
    oe_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "open_end"]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "matrix_cell"]
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "coded_question"]
    demo_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "demographic"]
    
    # Group matrix columns by question (q1r1, q1r2 -> q1)
    matrix_by_q = {}
    for i, h in matrix_cols:
        q = str(h).split("r")[0] if "r" in str(h).lower() else str(h)[:3]
        matrix_by_q.setdefault(q, []).append((i, h))
    
    # Group coded columns by question
    coded_by_q = {}
    for i, h in coded_cols:
        q = str(h).split("r")[0] if "r" in str(h).lower() else str(h)[:3]
        coded_by_q.setdefault(q, []).append((i, h))
    
    rid_idx = hidx.get("uuid") or hidx.get("record")
    
    features = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if rid_idx is None or rid_idx >= len(row): continue
        rid = clean(row[rid_idx])
        if not rid: continue
        
        feat = {"respondent_id": rid}
        
        # Per-matrix-question straightlining
        sl_counts = []
        for q, cols in matrix_by_q.items():
            if len(cols) < 3: continue
            vals = [norm(row[i]) for i, _ in cols if i < len(row) and row[i] is not None and row[i] != ""]
            if len(vals) >= 3:
                ur = len(set(vals)) / len(vals)
                sl_counts.append(1 if ur <= 0.3 else 0)
        
        feat["matrix_q_sl_count"] = sum(sl_counts)
        feat["matrix_q_sl_ratio"] = sum(sl_counts) / len(sl_counts) if sl_counts else 0
        feat["matrix_q_total"] = len(sl_counts)
        
        # Per-open-end question length variance
        oe_lens = []
        for i, _ in oe_cols:
            if i < len(row) and row[i]:
                oe_lens.append(len(str(row[i])))
        
        if oe_lens:
            feat["oe_len_std"] = np.std(oe_lens)
            feat["oe_len_cv"] = np.std(oe_lens) / np.mean(oe_lens) if np.mean(oe_lens) > 0 else 0
            feat["oe_len_min"] = min(oe_lens)
            feat["oe_len_max"] = max(oe_lens)
            feat["oe_len_range"] = max(oe_lens) - min(oe_lens)
            # Count of very short OE answers
            feat["oe_very_short_count"] = sum(1 for l in oe_lens if l < 10)
            feat["oe_short_count"] = sum(1 for l in oe_lens if l < 30)
            feat["oe_empty_count"] = sum(1 for l in oe_lens if l == 0)
        else:
            feat["oe_len_std"] = 0
            feat["oe_len_cv"] = 0
            feat["oe_len_min"] = 0
            feat["oe_len_max"] = 0
            feat["oe_len_range"] = 0
            feat["oe_very_short_count"] = 0
            feat["oe_short_count"] = 0
            feat["oe_empty_count"] = 0
        
        # Per-coded-question answer diversity
        coded_divs = []
        for q, cols in coded_by_q.items():
            if len(cols) < 2: continue
            vals = [str(norm(row[i])) for i, _ in cols if i < len(row) and row[i] is not None and row[i] != ""]
            if len(vals) >= 2:
                div = len(set(vals)) / len(vals)
                coded_divs.append(div)
        
        if coded_divs:
            feat["coded_q_div_mean"] = np.mean(coded_divs)
            feat["coded_q_div_min"] = min(coded_divs)
            feat["coded_q_low_div_count"] = sum(1 for d in coded_divs if d < 0.3)
        else:
            feat["coded_q_div_mean"] = 1.0
            feat["coded_q_div_min"] = 1.0
            feat["coded_q_low_div_count"] = 0
        
        # Answer pattern entropy (across all coded + matrix answers)
        all_answers = []
        for i, _ in matrix_cols + coded_cols:
            if i < len(row) and row[i] is not None and row[i] != "":
                all_answers.append(str(norm(row[i])))
        
        if all_answers:
            ctr = Counter(all_answers)
            probs = [c / len(all_answers) for c in ctr.values()]
            feat["answer_entropy"] = -sum(p * np.log(p + 1e-10) for p in probs)
            feat["answer_max_freq"] = max(ctr.values()) / len(all_answers)
            feat["answer_unique_ratio"] = len(ctr) / len(all_answers)
        else:
            feat["answer_entropy"] = 0
            feat["answer_max_freq"] = 0
            feat["answer_unique_ratio"] = 1.0
        
        # Demographic consistency (check for contradictions)
        demo_vals = {}
        for i, h in demo_cols:
            if i < len(row) and row[i]:
                demo_vals[str(h)] = str(norm(row[i]))
        
        feat["demo_count"] = len(demo_vals)
        feat["demo_unique"] = len(set(demo_vals.values()))
        feat["demo_missing"] = sum(1 for v in demo_vals.values() if not v or v in ["none", "na", "n/a", ""])
        
        features.append(feat)
    
    df = pd.DataFrame(features)
    _raw_cache[dataset_name] = df
    return df[df["respondent_id"].isin(respondent_ids)]


def train_and_predict(train_df, val_df, test_df):
    train_text = train_df["_oe_raw_text"].fillna("").values if "_oe_raw_text" in train_df.columns else [""] * len(train_df)
    val_text = val_df["_oe_raw_text"].fillna("").values if "_oe_raw_text" in val_df.columns else [""] * len(val_df)
    test_text = test_df["_oe_raw_text"].fillna("").values if "_oe_raw_text" in test_df.columns else [""] * len(test_df)

    X_train, y_train = prepare_features(train_df)
    X_val, y_val = prepare_features(val_df)
    X_test, y_test = prepare_features(test_df)

    for c in X_train.columns:
        if c not in X_val.columns: X_val[c] = 0
        if c not in X_test.columns: X_test[c] = 0
    for c in X_val.columns:
        if c not in X_train.columns: X_train[c] = 0
    for c in X_test.columns:
        if c not in X_train.columns: X_train[c] = 0
    X_val = X_val[X_train.columns]
    X_test = X_test[X_train.columns]

    # Raw Excel features
    dataset_name = train_df["dataset"].iloc[0] if "dataset" in train_df.columns else ""
    raw_train = extract_raw_excel_features(dataset_name, train_df["respondent_id"].values)
    raw_val = extract_raw_excel_features(dataset_name, val_df["respondent_id"].values)
    raw_test = extract_raw_excel_features(dataset_name, test_df["respondent_id"].values)
    
    # Merge raw features
    for d, raw in [(train_df, raw_train), (val_df, raw_val), (test_df, raw_test)]:
        if len(raw) > 0:
            raw_indexed = raw.set_index("respondent_id")
            for col in raw_indexed.columns:
                d[f"raw_{col}"] = d["respondent_id"].map(raw_indexed[col]).fillna(0)
    
    # Re-prepare features with raw features included
    X_train, y_train = prepare_features(train_df)
    X_val, y_val = prepare_features(val_df)
    X_test, y_test = prepare_features(test_df)
    
    for c in X_train.columns:
        if c not in X_val.columns: X_val[c] = 0
        if c not in X_test.columns: X_test[c] = 0
    for c in X_val.columns:
        if c not in X_train.columns: X_train[c] = 0
    for c in X_test.columns:
        if c not in X_train.columns: X_train[c] = 0
    X_val = X_val[X_train.columns]
    X_test = X_test[X_train.columns]

    # Deep features
    deep_train = extract_deep_features(train_df)
    deep_val = extract_deep_features(val_df)
    deep_test = extract_deep_features(test_df)

    # TF-IDF
    tfidf_w = get_tfidf_features(train_text, val_text, test_text, word=True, char=False)
    tfidf_wc = get_tfidf_features(train_text, val_text, test_text, word=True, char=True)

    # Try multiple approaches
    configs = [
        {"n_estimators": 300, "max_depth": 4, "learning_rate": 0.05, "min_samples_leaf": 10},
        {"n_estimators": 500, "max_depth": 3, "learning_rate": 0.03, "min_samples_leaf": 15},
        {"n_estimators": 200, "max_depth": 5, "learning_rate": 0.08, "min_samples_leaf": 5},
    ]

    approaches = []

    # Approach 1: Structured + raw
    for cfg in configs:
        m, iso, t, f1 = train_gbm(X_train.values, y_train, X_val.values, y_val, cfg)
        approaches.append(("structured_raw", m, iso, t, f1, X_train.values, X_val.values, X_test.values))

    # Approach 2: TF-IDF word + raw
    if tfidf_w[0] is not None:
        X_tr = np.hstack([X_train.values, tfidf_w[0]])
        X_va = np.hstack([X_val.values, tfidf_w[1]])
        X_te = np.hstack([X_test.values, tfidf_w[2]])
        for cfg in configs:
            m, iso, t, f1 = train_gbm(X_tr, y_train, X_va, y_val, cfg)
            approaches.append(("tfidf_w_raw", m, iso, t, f1, X_tr, X_va, X_te))

    # Approach 3: TF-IDF wc + raw
    if tfidf_wc[0] is not None:
        X_tr = np.hstack([X_train.values, tfidf_wc[0]])
        X_va = np.hstack([X_val.values, tfidf_wc[1]])
        X_te = np.hstack([X_test.values, tfidf_wc[2]])
        for cfg in configs:
            m, iso, t, f1 = train_gbm(X_tr, y_train, X_va, y_val, cfg)
            approaches.append(("tfidf_wc_raw", m, iso, t, f1, X_tr, X_va, X_te))

    # Approach 4: Deep + TF-IDF word + raw
    if tfidf_w[0] is not None:
        X_tr = np.hstack([X_train.values, tfidf_w[0], deep_train.values])
        X_va = np.hstack([X_val.values, tfidf_w[1], deep_val.values])
        X_te = np.hstack([X_test.values, tfidf_w[2], deep_test.values])
        for cfg in configs:
            m, iso, t, f1 = train_gbm(X_tr, y_train, X_va, y_val, cfg)
            approaches.append(("deep_tfidf_w_raw", m, iso, t, f1, X_tr, X_va, X_te))

    # Approach 5: Deep + TF-IDF wc + raw
    if tfidf_wc[0] is not None:
        X_tr = np.hstack([X_train.values, tfidf_wc[0], deep_train.values])
        X_va = np.hstack([X_val.values, tfidf_wc[1], deep_val.values])
        X_te = np.hstack([X_test.values, tfidf_wc[2], deep_test.values])
        for cfg in configs:
            m, iso, t, f1 = train_gbm(X_tr, y_train, X_va, y_val, cfg)
            approaches.append(("deep_tfidf_wc_raw", m, iso, t, f1, X_tr, X_va, X_te))

    # Pick best
    best = max(approaches, key=lambda a: a[4])
    name, model, iso, best_t, best_f1, _, _, X_te = best

    y_test_cal = iso.transform(model.predict_proba(X_te)[:, 1])
    y_test_pred = (y_test_cal >= best_t).astype(int)
    return y_test_pred, y_test_cal
