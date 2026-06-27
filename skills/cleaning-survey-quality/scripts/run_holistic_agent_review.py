#!/usr/bin/env python3
"""Holistic Agent Review Pipeline — Phase 1: Generate comprehensive review packets.

This script builds review packets that include the FULL answer chain with all
signals: per-grid straightlining analysis, all open-end fields with question
text, survey defender flags (qc, TERMFLAGS, RD_Search), timing, supplier risk,
LangAssess readability, duplicate counts, and ML triage score.

Key insight from miss analysis:
- The `outro` field is a QC question asking "describe what this survey was about"
  → Generic topic restatements are EXPECTED, not fraud signals
- q14 ("What prompted you to decide to buy?") is the key open-end for personal experience
- Missing q14 is NOT a discard signal for this client
- ML triage > 0.7 correlates with true discards
- Per-grid straightlining matters (some grids are naturally uniform)

Usage:
    python3 run_holistic_agent_review.py <xlsx_path> [--output-dir DIR] [--review-all] [--chunk-size N]
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl

# Add the skills scripts directory
SKILL_SCRIPTS = Path(__file__).parent
sys.path.insert(0, str(SKILL_SCRIPTS))

from survey_pipeline import (
    extract_features_and_chain,
    ml_triage,
    clean,
    norm,
    parse_datamap,
    classify_field,
)


def identify_grid_groups(headers, roles):
    """Group matrix columns by question prefix (e.g., q3r1-q3r12 = grid q3)."""
    grids = defaultdict(list)
    for i, h in enumerate(headers):
        if h and roles.get(str(h)) == "matrix_cell":
            # Extract grid prefix: q3r1 -> q3, q5r1c1 -> q5
            m = re.match(r"(q\d+)[a-z]?\d+", str(h).lower())
            if m:
                grids[m.group(1)].append((i, str(h)))
    return dict(grids)


def analyze_per_grid_straightlining(row, grids, hidx):
    """Analyze straightlining for each grid separately.
    
    Returns dict: {grid_name: {unique_ratio, straightline, n_items, values}}
    """
    results = {}
    for grid_name, cols in grids.items():
        vals = []
        for i, h in cols:
            if i < len(row) and row[i] is not None and row[i] != "":
                vals.append(norm(row[i]))
        if not vals:
            continue
        unique = len(set(vals))
        ur = unique / len(vals)
        results[grid_name] = {
            "n_items": len(vals),
            "n_unique": unique,
            "unique_ratio": round(ur, 3),
            "straightline": 1 if ur <= 0.2 and len(vals) >= 3 else 0,
            "values": [str(v) for v in vals],
        }
    return results


def get_all_oe_fields(row, oe_cols, datamap):
    """Get all open-end fields with their question text and response."""
    results = []
    for i, h in oe_cols:
        if i >= len(row):
            continue
        val = clean(row[i])
        if not val:
            continue
        dm = datamap.get(str(h), {})
        qtext = dm.get("question_text", str(h))
        results.append({
            "field": str(h),
            "question": qtext[:200],
            "response": val[:500],
            "char_count": len(val),
        })
    return results


def get_defender_signals(row, hidx):
    """Extract ALL survey defender / quality control / platform signals.
    
    Returns dict with raw signals + a human-readable summary string.
    """
    signals = {}
    summary_parts = []

    # === 1. QC FLAG (quality control — platform's own QC check) ===
    qc_idx = hidx.get("qc")
    if qc_idx and qc_idx < len(row):
        qc_val = norm(row[qc_idx])
        if qc_val and qc_val != 0:
            qc_labels = {
                1: "Not select 3", 2: "State mismatch", 3: "Red Herring",
                4: "REGION mismatch", 5: "AGE mismatch", 6: "RD /REVIEW rejection",
                7: "OE screening", 8: "RD /SEARCH threat rejection",
                9: "RD /SEARCH duplicate rejection", 10: "RD /SEARCH country rejection",
                11: "Speeder", 12: "Exceeded number of terms",
            }
            signals["qc_flag"] = qc_val
            signals["qc_label"] = qc_labels.get(qc_val, f"Unknown QC flag {qc_val}")
            summary_parts.append(f"QC FLAG: {qc_labels.get(qc_val, qc_val)}")

    # === 2. TERMFLAGS (platform termination flag — FRAUD DETECTED) ===
    tf_idx = hidx.get("TERMFLAGS")
    if tf_idx and tf_idx < len(row):
        tf = norm(row[tf_idx])
        if tf is not None:
            signals["termflags"] = tf
            if tf != 0:
                summary_parts.append(f"TERMFLAGS={tf} (PLATFORM FRAUD FLAG — auto-discard)")

    # === 3. RD_Search (research defender — bot/fraud/threat detection) ===
    rd_search_labels = {
        "RD_Searchr0": "search_threat_potential",
        "RD_Searchr1": "search_threat_score",
        "RD_Searchr2": "search_respondent_risk",
        "RD_Searchr3": "search_country",
        "RD_Searchr4": "search_flag",
        "RD_Searchr5": "search_duplicate_potential",
        "RD_Searchr6": "search_duplicate_score",
        "RD_Searchr7": "search_duplicate_flag",
    }
    rd_values = {}
    for field, label in rd_search_labels.items():
        idx = hidx.get(field)
        if idx and idx < len(row):
            val = row[idx]
            if val is not None and val != "":
                rd_values[label] = norm(val) if not isinstance(val, str) else val
                signals[field] = rd_values[label]
    # Summarize RD_Search
    threat_score = rd_values.get("search_threat_score")
    if threat_score is not None:
        try:
            ts = int(threat_score)
            if ts >= 25:
                summary_parts.append(f"RD_Search threat score={ts} (ELEVATED)")
            elif ts >= 20:
                summary_parts.append(f"RD_Search threat score={ts} (moderate)")
        except (ValueError, TypeError):
            pass
    country = rd_values.get("search_country")
    if country and country != "United States":
        summary_parts.append(f"RD_Search country={country} (NON-US)")
    dup_pot = rd_values.get("search_duplicate_potential")
    if dup_pot and str(dup_pot).lower() not in ("low", "0", "none"):
        summary_parts.append(f"RD_Search duplicate={dup_pot}")

    # === 4. RD_GetToken (token verification / bot detection) ===
    for field in ["RD_GetTokenr0", "RD_GetTokenr1"]:
        idx = hidx.get(field)
        if idx and idx < len(row):
            val = row[idx]
            if val is not None and val != "":
                signals[field] = str(val)[:50]

    # === 5. outroR1_RD_Review (outro text review — language/paste/profanity/similarity) ===
    outro_review_labels = {
        "outroR1_RD_Reviewr0": "outro_language_detected",
        "outroR1_RD_Reviewr1": "outro_pasted_response",
        "outroR1_RD_Reviewr2": "outro_profanity_check",
        "outroR1_RD_Reviewr3": "outro_composite_score",
        "outroR1_RD_Reviewr4": "outro_similarity_flag",
    }
    outro_review = {}
    for field, label in outro_review_labels.items():
        idx = hidx.get(field)
        if idx and idx < len(row):
            val = row[idx]
            if val is not None and val != "":
                outro_review[label] = norm(val) if not isinstance(val, str) else val
                signals[field] = outro_review[label]
    # Summarize outro review
    lang = outro_review.get("outro_language_detected")
    if lang and lang != "English":
        summary_parts.append(f"Outro language={lang} (NON-ENGLISH)")
    pasted = outro_review.get("outro_pasted_response")
    if pasted and str(pasted) != "0":
        summary_parts.append(f"Outro PASTED response detected")
    profanity = outro_review.get("outro_profanity_check")
    if profanity and str(profanity) != "0":
        summary_parts.append(f"Outro profanity detected")
    composite = outro_review.get("outro_composite_score")
    if composite:
        try:
            cs = int(composite)
            if cs >= 10:
                summary_parts.append(f"Outro composite score={cs} (HIGH — suspicious)")
            elif cs >= 5:
                summary_parts.append(f"Outro composite score={cs} (moderate)")
        except (ValueError, TypeError):
            pass
    sim_flag = outro_review.get("outro_similarity_flag")
    if sim_flag and str(sim_flag) == "0":
        summary_parts.append("Outro similarity flag=0 (flagged by platform)")

    # === 6. LangAssess (readability — AI text detection proxy) ===
    lang_fields = {
        "LangAssessReadLevel": "read_level",
        "LangAssessReadEase": "read_ease",
        "LangAssessNumSen": "num_sentences",
        "LangAssessNumWords": "num_words",
        "LangAssessNumSyl": "num_syllables",
    }
    lang_vals = {}
    for field, label in lang_fields.items():
        idx = hidx.get(field)
        if idx and idx < len(row):
            val = row[idx]
            if val is not None:
                try:
                    lang_vals[label] = float(val)
                except (ValueError, TypeError):
                    lang_vals[label] = val
    signals["lang_assess"] = lang_vals
    # Summarize LangAssess — high read level with short text = possible AI
    rl = lang_vals.get("read_level")
    nw = lang_vals.get("num_words")
    if rl is not None and nw is not None:
        try:
            rl_f = float(rl)
            nw_f = float(nw)
            if rl_f >= 15 and nw_f <= 15:
                summary_parts.append(f"ReadLevel={rl_f:.1f} with only {int(nw_f)} words (VERY HIGH readability for short text — possible AI-generated)")
            elif rl_f >= 17:
                summary_parts.append(f"ReadLevel={rl_f:.1f} (VERY HIGH — possible AI-generated)")
            elif rl_f < 2 and nw_f > 5:
                summary_parts.append(f"ReadLevel={rl_f:.1f} (VERY LOW — possible incoherence)")
        except (ValueError, TypeError):
            pass

    # === 7. vlist (participant source — some sources are higher risk) ===
    vlist_idx = hidx.get("vlist")
    if vlist_idx and vlist_idx < len(row):
        vlist_val = norm(row[vlist_idx])
        if vlist_val is not None:
            signals["vlist"] = vlist_val

    # === 8. decLang (declared language) ===
    decLang_idx = hidx.get("decLang")
    if decLang_idx and decLang_idx < len(row):
        dl = row[decLang_idx]
        if dl is not None and str(dl).strip():
            signals["decLang"] = str(dl).strip()
            if str(dl).strip().lower() not in ("english", "en", "us"):
                summary_parts.append(f"Declared language={dl} (NON-ENGLISH)")

    # === 9. vdropout (last seen question — detects dropouts/partials) ===
    vdropout_idx = hidx.get("vdropout")
    if vdropout_idx and vdropout_idx < len(row):
        vd = row[vdropout_idx]
        if vd is not None and str(vd).strip():
            signals["vdropout"] = str(vd).strip()

    # === 10. IP / UserAgent duplicates (detected elsewhere, but noted here) ===
    # These are computed in the cross-respondent section

    # Build the summary string
    if not summary_parts:
        signals["defender_summary"] = "No platform defender signals triggered."
    else:
        signals["defender_summary"] = " | ".join(summary_parts)

    return signals


def get_key_answers(row, hidx, datamap):
    """Extract key single-choice and demographic answers for context.

    Dynamically discovers fields from the workbook rather than hardcoding
    Delta-specific field names. Falls back to known field names only when
    dynamic discovery fails.
    """
    answers = {}

    # Dynamically extract ALL single-choice fields with datamap labels
    # This works for any dataset, not just Delta
    for h, idx in hidx.items():
        if idx >= len(row):
            continue
        val = norm(row[idx])
        if val is None or val == "":
            continue
        dm = datamap.get(str(h), {})
        labels = dm.get("labels", {})
        qtext = dm.get("question_text", "")
        # If the field has value labels, it's a coded single/multi-select
        if labels and str(val) in labels:
            answers[str(h)] = {
                "question": qtext[:150] if qtext else str(h),
                "value": str(val),
                "label": labels[str(val)],
            }

    return answers


def get_survey_structure(row, hidx, datamap):
    """Extract survey-structure fields that carry discard signal lift.

    These fields are NOT semantic content — they are classification, quota,
    channel, and supplier-structure fields that the client's cleaning process
    uses but that semantic review alone cannot see.

    Evidence from ECHO annotated data:
    - CLASSIFY=1 (pro) rejects at 60.4% vs CLASSIFY=2 (consumer) at 30.5%
    - conditionsAriens=1 rejects at 59.2%
    - list 25 rejects at 45.3% vs list 139 at 30.0%
    - PROAGE=1 rejects at 65.5%
    - TERMFLAGS=1 rejects at 66.7%
    """
    structure = {}

    # Classification fields
    classify_idx = hidx.get("CLASSIFY")
    if classify_idx and classify_idx < len(row):
        val = norm(row[classify_idx])
        if val is not None:
            dm = datamap.get("CLASSIFY", {})
            labels = dm.get("labels", {})
            structure["classify"] = {
                "value": val,
                "label": labels.get(str(val), f"CLASSIFY={val}"),
            }

    proage_idx = hidx.get("PROAGE")
    if proage_idx and proage_idx < len(row):
        val = norm(row[proage_idx])
        if val is not None:
            structure["proage"] = val

    conage_idx = hidx.get("CONAGE")
    if conage_idx and conage_idx < len(row):
        val = norm(row[conage_idx])
        if val is not None:
            structure["conage"] = val

    # Channel/condition fields
    condition_fields = [
        "conditionsAriens", "conditionsHD_or_OPE_dealers", "conditionsOther_channel",
        "conditionsWest", "conditionsMidwest", "conditionsSouth", "conditionsNortheast",
    ]
    conditions = {}
    for field in condition_fields:
        idx = hidx.get(field)
        if idx and idx < len(row):
            val = norm(row[idx])
            if val is not None:
                conditions[field] = val
    if conditions:
        structure["conditions"] = conditions

    # Supplier/list fields
    list_idx = hidx.get("list")
    if list_idx and list_idx < len(row):
        val = norm(row[list_idx])
        if val is not None:
            structure["list"] = val

    source_idx = hidx.get("source")
    if source_idx and source_idx < len(row):
        val = row[source_idx]
        if val is not None and str(val).strip():
            structure["source"] = str(val).strip()

    vlist_idx = hidx.get("vlist")
    if vlist_idx and vlist_idx < len(row):
        val = norm(row[vlist_idx])
        if val is not None:
            structure["vlist"] = val

    # Device/user agent category (dcua) — correlates with reject rate
    dcua_idx = hidx.get("dcua")
    if dcua_idx and dcua_idx < len(row):
        val = row[dcua_idx]
        if val is not None and str(val).strip():
            structure["dcua"] = str(val).strip()

    # FIRMREV (firm revenue — pro classification)
    firmrev_idx = hidx.get("FIRMREV")
    if firmrev_idx and firmrev_idx < len(row):
        val = norm(row[firmrev_idx])
        if val is not None:
            dm = datamap.get("FIRMREV", {})
            labels = dm.get("labels", {})
            structure["firmrev"] = {
                "value": val,
                "label": labels.get(str(val), str(val)),
            }

    return structure


def get_brand_funnel(row, hidx, datamap, headers):
    """Extract brand-funnel fields for consistency checking.

    Brand funnel fields carry strong discard signal (BRANDS2RATEQuota appeared
    1040 times across 553 ECHO discards). We extract:
    - Brand awareness fields (q1 multi-select)
    - Brands rated (q11a* matrix)
    - Brand consideration (q17*)
    - Brand recommendation (q19*, q20*)
    - NPS / satisfaction (q29, q30*)
    - Brand purchase / ownership (q14*, q16*)
    """
    funnel = {}

    # Find brand-related fields by scanning headers for known patterns
    brand_field_patterns = [
        # Awareness / multi-select brand fields
        (r"^q1r\d+$", "awareness"),
        (r"^q3r\d+$", "awareness_detail"),
        # Brand rating matrices
        (r"^q11ar\d+c\d+$", "brand_rating"),
        (r"^q11ar\d+oe$", "brand_other"),
        # Brand consideration / recommendation
        (r"^q17r\d+$", "consideration"),
        (r"^q17r\d+oe$", "consideration_oe"),
        (r"^q19_2026r\d+$", "recommendation"),
        (r"^q19_2026othr\d+$", "recommendation_oe"),
        (r"^q20r\d+oe$", "purchase_location_oe"),
        # NPS / satisfaction
        (r"^q29$", "nps_verbatim"),
        (r"^q30_2026r\d+$", "satisfaction"),
        (r"^q30_2026r\d+oe$", "satisfaction_oe"),
        # Brand purchase
        (r"^q14r\d+oe$", "purchase_oe"),
        (r"^q16r\d+$", "brand_knowledge"),
        # Possible brands
        (r"^POSSIBLEBRANDSr\d+$", "possible_brands"),
        (r"^q23_2026_Lr\d+$", "brand_share"),
        (r"^q18r\d+$", "brand_familiarity"),
    ]

    import re as _re
    for i, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h)
        for pattern, funnel_stage in brand_field_patterns:
            if _re.match(pattern, h_str, _re.IGNORECASE):
                if i < len(row):
                    val = row[i]
                    if val is not None and str(val).strip():
                        dm = datamap.get(h_str, {})
                        labels = dm.get("labels", {})
                        qtext = dm.get("question_text", "")
                        if isinstance(val, (int, float)):
                            label = labels.get(str(int(val)), str(val))
                        else:
                            label = labels.get(str(val), str(val))
                        if funnel_stage not in funnel:
                            funnel[funnel_stage] = []
                        funnel[funnel_stage].append({
                            "field": h_str,
                            "value": str(val),
                            "label": label,
                            "question": qtext[:120] if qtext else "",
                        })
                break

    return funnel


def get_quota_reconstruction(row, hidx, datamap, headers, df=None, idx=None):
    """Reconstruct quota bucket membership from raw fields.

    The client's discard process uses quota markers (CLASSIFYQuota, GenderQuota,
    RegionQuota, BRANDS2RATEQuota, ChannelQuota, TotalQuota, etc.) that are NOT
    in the raw workbook. However, the underlying quota dimensions ARE present as
    raw fields. We reconstruct the quota bucket membership to give agents visibility
    into which quota cells this respondent occupies.

    Key quota dimensions (ECHO-specific, generalizable patterns):
    - CLASSIFY → CLASSIFYQuota (pro vs consumer)
    - REGION → RegionQuota
    - Gender (from q3 or gender field) → GenderQuota
    - Age (from age or qager1) → AgeQuota / CONAgeQuota
    - Channel conditions → ChannelQuota (Ariens, HD or OPE dealers, Other channel)
    - Brands rated (q11a* matrix) → BRANDS2RATEQuota
    - Total → TotalQuota (all respondents)
    """
    quota = {
        "classify_bucket": None,
        "region_bucket": None,
        "gender_bucket": None,
        "age_bucket": None,
        "channel_bucket": None,
        "brands_rated_count": 0,
        "brands_rated": [],
        "quota_cells": [],
        "population_in_cells": {},
    }

    # CLASSIFY bucket
    classify_idx = hidx.get("CLASSIFY")
    if classify_idx and classify_idx < len(row):
        val = norm(row[classify_idx])
        if val is not None:
            dm = datamap.get("CLASSIFY", {})
            labels = dm.get("labels", {})
            label = labels.get(str(int(val)) if isinstance(val, (int, float)) else str(val), str(val))
            quota["classify_bucket"] = {"value": val, "label": label}
            quota["quota_cells"].append(f"CLASSIFYQuota/{label}")

    # REGION bucket
    region_idx = hidx.get("REGION")
    if region_idx and region_idx < len(row):
        val = norm(row[region_idx])
        if val is not None:
            dm = datamap.get("REGION", {})
            labels = dm.get("labels", {})
            label = labels.get(str(int(val)) if isinstance(val, (int, float)) else str(val), str(val))
            quota["region_bucket"] = {"value": val, "label": label}
            quota["quota_cells"].append(f"RegionQuota/{label}")

    # Gender bucket — look for gender field
    gender_idx = hidx.get("q3") or hidx.get("gender") or hidx.get("Gender")
    if gender_idx and gender_idx < len(row):
        val = norm(row[gender_idx])
        if val is not None:
            dm = datamap.get("q3", datamap.get("gender", {}))
            labels = dm.get("labels", {})
            label = labels.get(str(int(val)) if isinstance(val, (int, float)) else str(val), str(val))
            quota["gender_bucket"] = {"value": val, "label": label}
            quota["quota_cells"].append(f"GenderQuota/{label}")

    # Age bucket
    age_idx = hidx.get("age") or hidx.get("qager1") or hidx.get("CONAGE") or hidx.get("PROAGE")
    if age_idx and age_idx < len(row):
        val = norm(row[age_idx])
        if val is not None:
            field_name = "age" if hidx.get("age") else ("qager1" if hidx.get("qager1") else ("CONAGE" if hidx.get("CONAGE") else "PROAGE"))
            dm = datamap.get(field_name, {})
            labels = dm.get("labels", {})
            label = labels.get(str(int(val)) if isinstance(val, (int, float)) else str(val), str(val))
            quota["age_bucket"] = {"value": val, "label": label, "field": field_name}
            if field_name in ("CONAGE", "PROAGE"):
                quota["quota_cells"].append(f"CONAgeQuota/{label}" if field_name == "CONAGE" else f"PROAgeQuota/{label}")
            else:
                quota["quota_cells"].append(f"AgeQuota/{label}")

    # Channel bucket — from conditions fields
    channel_parts = []
    condition_fields = {
        "conditionsAriens": "Ariens",
        "conditionsHD_or_OPE_dealers": "HD or OPE dealers",
        "conditionsOther_channel": "Other channel",
    }
    for field, label in condition_fields.items():
        cidx = hidx.get(field)
        if cidx and cidx < len(row):
            val = norm(row[cidx])
            if val == 1:
                channel_parts.append(label)
                quota["quota_cells"].append(f"ChannelQuota/{label}")

    if channel_parts:
        quota["channel_bucket"] = channel_parts

    # Brands rated — count q11a* matrix fields with nonzero values
    import re as _re
    brands_rated = []
    for i, h in enumerate(headers):
        if h and _re.match(r"^q11ar\d+c4$", str(h), _re.IGNORECASE):
            if i < len(row):
                val = row[i]
                if val is not None and str(val).strip() and str(val) != "0":
                    brands_rated.append(str(h))
    quota["brands_rated_count"] = len(brands_rated)
    quota["brands_rated"] = brands_rated
    if brands_rated:
        quota["quota_cells"].append(f"BRANDS2RATEQuota/{len(brands_rated)}_brands")

    # Total quota (all respondents)
    quota["quota_cells"].append("TotalQuota/Total")

    # Population counts per cell (if df is available)
    if df is not None and idx is not None:
        try:
            classify_val = quota.get("classify_bucket", {}).get("value")
            region_val = quota.get("region_bucket", {}).get("value")
            if classify_val is not None:
                same_classify = (df["CLASSIFY"] == classify_val).sum() if "CLASSIFY" in df.columns else 0
                quota["population_in_cells"]["CLASSIFYQuota"] = int(same_classify)
            if region_val is not None:
                same_region = (df["REGION"] == region_val).sum() if "REGION" in df.columns else 0
                quota["population_in_cells"]["RegionQuota"] = int(same_region)
        except Exception:
            pass

    return quota


def load_ml_signal_correlations():
    """Load cross-corpus ML signal correlation data for injection into review packets.

    Returns a dict with:
    - family_correlations: per-family mean correlation with client discard
    - universal_signals: signals that fire across 3+ datasets
    - global_top_signals: top features from the global model
    - per_dataset_signals: dataset-specific top signals (if dataset name matches)
    """
    corr_path = Path(__file__).parent.parent / "evolution" / "ml-signal-correlation" / "cross_corpus_signal_correlation.json"
    if not corr_path.exists():
        return None
    try:
        with open(corr_path) as f:
            data = json.load(f)
        return {
            "family_correlations": data.get("family_correlations", {}),
            "universal_signals": data.get("universal_signals", [])[:10],
            "global_top_signals": data.get("global_model", {}).get("top_signals", [])[:10],
            "global_auc": data.get("global_model", {}).get("auc", 0),
            "corpus_stats": data.get("corpus_stats", {}),
        }
    except Exception:
        return None


def build_holistic_review_packets(filepath, output_dir, review_all=False, chunk_size=200):
    """Build comprehensive review packets for all respondents."""
    filepath = Path(filepath)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*80}")
    print(f"HOLISTIC AGENT REVIEW — Building Review Packets")
    print(f"{'='*80}")
    print(f"  Input: {filepath.name}")
    print(f"  Output: {output_dir}")

    # Load cross-corpus ML signal correlations
    ml_correlations = load_ml_signal_correlations()
    if ml_correlations:
        print(f"  ML correlations loaded: {ml_correlations['corpus_stats'].get('total_respondents', 0)} respondents, "
              f"{ml_correlations['corpus_stats'].get('n_datasets', 0)} datasets, "
              f"global AUC={ml_correlations['global_auc']:.3f}")
    else:
        print(f"  ML correlations: not found (run cross_corpus_signal_correlation.py first)")

    # Load workbook
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    datamap = parse_datamap(wb)
    roles = {str(h): classify_field(str(h)) for h in headers if h}

    # Identify column groups
    oe_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "open_end"]

    # Also identify open-end fields from datamap (fields with no value labels = text fields)
    # This catches fields like q14 that don't end in "oe"
    oe_field_names = {h for _, h in oe_cols}
    technical_qtext = {"captured variable", "open text response", "open numeric response",
                       "open-ended response", ""}
    for i, h in enumerate(headers):
        if h and str(h) not in oe_field_names:
            dm = datamap.get(str(h), {})
            qtext = dm.get("question_text", "").lower().strip()
            # If the field has question text but no labels, it's likely an open-end text field
            # Exclude technical "Captured variable" and generic "Open text response" fields
            if dm.get("question_text") and not dm.get("labels") and qtext not in technical_qtext:
                # Verify it's not a known non-OE field
                if roles.get(str(h)) not in ("id", "label", "timing", "supplier", "ip",
                    "user_agent", "timestamp", "technical", "nlp_metadata", "review_metadata",
                    "quality_flag", "paste_flag", "no_answer", "routing", "quota"):
                    oe_cols.append((i, h))
                    oe_field_names.add(str(h))

    lang_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "nlp_metadata"]
    grids = identify_grid_groups(headers, roles)

    print(f"  Open-end fields: {len(oe_cols)}")
    print(f"  Grid groups: {len(grids)} ({list(grids.keys())[:10]}...)")
    print(f"  Datamap fields: {len(datamap)}")

    # Run the standard pipeline for ML triage + features
    print(f"\n  Running ML triage + feature extraction...")
    df, _, _, answer_chains = extract_features_and_chain(filepath)
    df = ml_triage(df)

    # Compute population stats
    matrix_prevalence = (df["matrix_straightline"] == 1).mean() if "matrix_straightline" in df.columns else 0
    qtime_median = float(df["qtime_minutes"].median()) if "qtime_minutes" in df.columns else 0
    qtime_p10 = float(np.percentile(df[df["qtime_minutes"] > 0]["qtime_minutes"], 10)) if "qtime_minutes" in df.columns else 0
    qtime_p25 = float(np.percentile(df[df["qtime_minutes"] > 0]["qtime_minutes"], 25)) if "qtime_minutes" in df.columns else 0

    # Supplier reject rates
    sup_rates = df.groupby("supplier_name")["signal_count"].mean().to_dict() if "signal_count" in df.columns else {}
    global_rate = float(df["signal_count"].mean()) if "signal_count" in df.columns else 0

    # OE duplicate counts (across all respondents)
    all_oe_combined = []
    for idx, row in df.iterrows():
        chain = answer_chains[idx] if idx < len(answer_chains) else {}
        all_oe_combined.append(chain.get("oe_text", ""))
    oe_ctr = Counter(t.strip().lower() for t in all_oe_combined if t.strip())

    # Per-field OE duplicate counts
    oe_field_texts = defaultdict(list)
    rows_raw = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows_raw:
        for i, h in oe_cols:
            if i < len(row):
                val = clean(row[i])
                if val:
                    oe_field_texts[str(h)].append(val.strip().lower())
    oe_field_ctr = {h: Counter(texts) for h, texts in oe_field_texts.items()}

    # === Cross-respondent text similarity detection ("too clean to be real" pattern) ===
    # Detect AI-generated text by finding clusters of respondents whose OE answers
    # share suspiciously similar phrasing, vocabulary, and sentence structure.
    # Real humans write differently from each other; AI-generated respondents use
    # templates with minor variations.
    print(f"\n  Computing cross-respondent text similarity (AI template detection)...")
    from difflib import SequenceMatcher

    # For each OE field, compute pairwise similarity for all respondents
    # Then flag respondents whose text is >60% similar to 5+ others
    oe_similarity = {}  # rid -> {field: {"max_similarity": float, "n_similar": int, "similar_to": [rids]}}
    for i, h in oe_cols:
        field = str(h)
        texts = []  # (rid, text)
        for idx_row, row in enumerate(rows_raw):
            if i < len(row):
                val = clean(row[i])
                if val and len(val.strip()) > 20:  # Only compare substantive text
                    rid = df.iloc[idx_row]["respondent_id"] if idx_row < len(df) else str(idx_row)
                    texts.append((rid, val.strip().lower()))

        if len(texts) < 10:
            continue

        # Compute pairwise similarity (sample if too many to avoid O(n^2) blowup)
        # For >500 texts, use a sampling approach: compare each text to a random sample
        import random
        random.seed(42)
        sample_size = min(len(texts), 300)

        for rid_a, text_a in texts:
            if rid_a not in oe_similarity:
                oe_similarity[rid_a] = {}
            if field not in oe_similarity[rid_a]:
                oe_similarity[rid_a][field] = {"max_similarity": 0, "n_similar": 0, "similar_to": []}

            # Compare to a sample of other texts
            others = [t for t in texts if t[0] != rid_a]
            if len(others) > sample_size:
                others = random.sample(others, sample_size)

            n_similar = 0
            max_sim = 0
            similar_to = []
            for rid_b, text_b in others:
                sim = SequenceMatcher(None, text_a, text_b, autojunk=False).ratio()
                if sim > max_sim:
                    max_sim = sim
                if sim > 0.6:
                    n_similar += 1
                    if len(similar_to) < 5:
                        similar_to.append(rid_b)

            oe_similarity[rid_a][field]["max_similarity"] = round(max_sim, 3)
            oe_similarity[rid_a][field]["n_similar"] = n_similar
            oe_similarity[rid_a][field]["similar_to"] = similar_to

    # Compute AI-text suspicion score per respondent
    # High similarity to many others = likely from a template = AI-generated
    ai_suspicion = {}  # rid -> {"score": 0-1, "fields_flagged": [...], "details": str}
    for rid, fields in oe_similarity.items():
        flagged = []
        max_n_similar = 0
        max_sim = 0
        for field, info in fields.items():
            if info["n_similar"] >= 5 or info["max_similarity"] > 0.8:
                flagged.append(field)
                max_n_similar = max(max_n_similar, info["n_similar"])
                max_sim = max(max_sim, info["max_similarity"])

        if flagged:
            # Score: more fields flagged + higher similarity = higher suspicion
            score = min(1.0, (len(flagged) * 0.3) + (max_n_similar * 0.05) + (max_sim * 0.2))
            detail_parts = []
            for field in flagged:
                info = fields[field]
                detail_parts.append(f"{field}: {info['n_similar']} similar (max={info['max_similarity']})")
            ai_suspicion[rid] = {
                "score": round(score, 3),
                "fields_flagged": flagged,
                "details": " | ".join(detail_parts),
            }

    n_ai_flagged = len(ai_suspicion)
    print(f"    AI text suspicion: {n_ai_flagged} respondents flagged ({n_ai_flagged/len(df)*100:.1f}%)")

    # Build review packets
    print(f"\n  Building review packets for {len(df)} respondents...")

    review_packets = []
    for idx, df_row in df.iterrows():
        rid = df_row["respondent_id"]
        raw_row = rows_raw[idx] if idx < len(rows_raw) else None
        if raw_row is None:
            continue

        chain = answer_chains[idx] if idx < len(answer_chains) else {}

        # Per-grid straightlining analysis
        grid_analysis = analyze_per_grid_straightlining(raw_row, grids, hidx)
        n_grids_sl = sum(1 for g in grid_analysis.values() if g["straightline"])
        n_grids_total = len(grid_analysis)
        grids_sl_list = [g for g, v in grid_analysis.items() if v["straightline"]]

        # All OE fields with question text
        oe_fields = get_all_oe_fields(raw_row, oe_cols, datamap)

        # Defender signals (includes defender_summary string)
        defender = get_defender_signals(raw_row, hidx)

        # AI text suspicion (cross-respondent similarity)
        ai_info = ai_suspicion.get(rid, {"score": 0, "fields_flagged": [], "details": ""})

        # Key demographic/choice answers (dynamically discovered)
        key_answers = get_key_answers(raw_row, hidx, datamap)

        # Survey-structure fields (CLASSIFY, PROAGE, conditions, list, etc.)
        survey_structure = get_survey_structure(raw_row, hidx, datamap)

        # Brand funnel fields (awareness → rating → consideration → NPS)
        brand_funnel = get_brand_funnel(raw_row, hidx, datamap, headers)

        # Quota reconstruction (CLASSIFYQuota, RegionQuota, GenderQuota, ChannelQuota, etc.)
        quota_reconstruction = get_quota_reconstruction(raw_row, hidx, datamap, headers, df=df, idx=idx)

        # LangAssess
        lang = {}
        for i, h in lang_cols:
            if i < len(raw_row) and raw_row[i] is not None:
                lang[str(h)] = float(raw_row[i]) if raw_row[i] else 0

        # Timing
        qtime_min = float(df_row.get("qtime_minutes", 0))
        qtime_pct = df_row.get("qtime_percentile", "")

        # Supplier
        supplier = str(df_row.get("supplier_name", ""))
        sup_rate = float(df_row.get("supplier_reject_rate", 0))

        # ML triage
        ml_score = float(df_row.get("ml_triage_score", 0.5))

        # OE duplicate counts per field
        oe_dups = {}
        for oe in oe_fields:
            field = oe["field"]
            text = oe["response"].strip().lower()
            if text:
                ctr = oe_field_ctr.get(field, Counter())
                oe_dups[field] = ctr.get(text, 0)

        # Build the comprehensive packet
        packet = {
            "respondent_id": rid,
            "ml_triage_score": round(ml_score, 3),
            "timing": {
                "minutes": round(qtime_min, 1),
                "percentile": qtime_pct,
                "median_minutes": round(qtime_median, 1),
            },
            "supplier": {
                "name": supplier,
                "reject_rate": round(sup_rate, 1),
            },
            "defender_signals": defender,
            "defender_summary": defender.get("defender_summary", "No platform defender signals triggered."),
            "ai_text_suspicion": ai_info,
            "lang_assess": lang,
            "matrix_analysis": {
                "total_grids": n_grids_total,
                "grids_straightlined": n_grids_sl,
                "straightlined_grids": grids_sl_list,
                "per_grid": {g: {"unique_ratio": v["unique_ratio"], "n_items": v["n_items"], "n_unique": v["n_unique"]}
                            for g, v in grid_analysis.items()},
            },
            "open_end_responses": oe_fields,
            "oe_duplicate_counts": oe_dups,
            "key_answers": key_answers,
            "survey_structure": survey_structure,
            "brand_funnel": brand_funnel,
            "quota_reconstruction": quota_reconstruction,
            "signal_count": int(df_row.get("signal_count", 0)),
            "t1_count": int(df_row.get("t1_count", 0)),
            "t2_count": int(df_row.get("t2_count", 0)),
            "answer_entropy": round(float(df_row.get("answer_entropy", 0)), 2),
            "matrix_unique_ratio": round(float(df_row.get("matrix_unique_ratio", 0)), 3),
            "oe_total_chars": int(df_row.get("oe_total_chars", 0)),
        }

        # Inject cross-corpus ML signal correlations (same for all packets, but agents need it in context)
        if ml_correlations:
            packet["ml_signal_correlations"] = {
                "family_correlations": {
                    fam: {
                        "mean_correlation": data["mean_correlation"],
                        "direction": data["direction"],
                        "n_datasets": data["n_datasets"],
                    }
                    for fam, data in ml_correlations["family_correlations"].items()
                },
                "universal_signals": [
                    {"feature": s["feature"], "family": s["family"], "n_datasets": s["n_datasets"]}
                    for s in ml_correlations["universal_signals"]
                ],
                "global_top_signals": [
                    {"feature": s["feature"], "family": s["family"], "importance": s["importance"]}
                    for s in ml_correlations["global_top_signals"]
                ],
                "corpus_stats": {
                    "total_respondents": ml_correlations["corpus_stats"].get("total_respondents", 0),
                    "n_datasets": ml_correlations["corpus_stats"].get("n_datasets", 0),
                },
            }
        review_packets.append(packet)

    print(f"  Total packets: {len(review_packets)}")

    # Write packets in chunks
    n_chunks = (len(review_packets) + chunk_size - 1) // chunk_size
    print(f"  Chunking into {n_chunks} files of ~{chunk_size} each...")

    for chunk_idx in range(n_chunks):
        start = chunk_idx * chunk_size
        end = min(start + chunk_size, len(review_packets))
        chunk = review_packets[start:end]
        chunk_path = output_dir / f"review_chunk_{chunk_idx:02d}.json"
        with open(chunk_path, "w") as f:
            json.dump(chunk, f, indent=2)
        print(f"    Chunk {chunk_idx:02d}: {len(chunk)} packets → {chunk_path.name}")

    # Write agent instructions
    instructions = build_agent_instructions(len(review_packets), n_chunks, matrix_prevalence, filepath.name)
    instructions_path = output_dir / "agent_review_instructions.md"
    with open(instructions_path, "w") as f:
        f.write(instructions)
    print(f"  Instructions: {instructions_path.name}")

    # Write summary
    summary = {
        "dataset": filepath.name,
        "total_respondents": len(review_packets),
        "n_chunks": n_chunks,
        "chunk_size": chunk_size,
        "matrix_prevalence": round(matrix_prevalence, 3),
        "median_qtime_minutes": round(qtime_median, 1),
        "oe_fields": [h for _, h in oe_cols],
        "grid_groups": list(grids.keys()),
    }
    summary_path = output_dir / "review_summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)

    print(f"\n{'='*80}")
    print(f"COMPLETE — {len(review_packets)} review packets in {n_chunks} chunks")
    print(f"{'='*80}")
    print(f"\nNext: Spawn {n_chunks} subagents to review the chunks.")
    print(f"Each subagent reads review_chunk_XX.json and writes agent_judgments_chunk_XX.json")

    return review_packets


def build_agent_instructions(n_respondents, n_chunks, matrix_prevalence, dataset_name):
    """Build detailed instructions for the reviewing agent."""

    return f"""# Holistic Agent Review Instructions — v6 (Three-Component + Disposition Layer)

## Task

Read the file `review_chunk_XX.json` (assigned to you). It contains ~200 respondent review packets.

For each respondent, you must read ALL signals holistically and produce a STRUCTURED judgment
with three separate risk scores, a disposition layer, per-evidence-family firing, a badopen
audit trail, OE semantic classification, and ML analysis context.

## Three-Component Scoring

We no longer treat every removal as a "bad respondent" signal. We separate:

1. **authenticity_risk** (0.0–1.0): Is this respondent real?
   - 0.0 = clearly genuine human
   - 0.3 = minor platform concerns
   - 0.5 = ambiguous (could be AI or human)
   - 0.7 = likely fraud (platform flags, AI text, gibberish)
   - 0.9 = certain fraud (TERMFLAGS=1 + AI text + gibberish)

2. **quality_discard_risk** (0.0–1.0): Does the content fail the PM quality bar?
   - 0.0 = substantive, grounded, on-topic
   - 0.3 = thin but on-topic
   - 0.5 = off-topic or fails answer-role
   - 0.7 = short non-answer or product-copy register
   - 0.9 = gibberish or completely nonresponsive

3. **client_reject_probability** (0.0–1.0): Would the client discard this?
   - This accounts for quota, badopen, classification, and source logic
   - 0.0 = client would keep (substantive OE, coherent structure)
   - 0.3 = client might review (thin OE but on-topic)
   - 0.5 = client likely discards (badopen trigger + thin OE)
   - 0.7 = client very likely discards (off-topic + ML high + quota over-filled)
   - 0.9 = client certainly discards (TERMFLAGS + gibberish + wrong brand universe)

## Disposition Layer

Assign a primary and secondary removal reason from this enum:
- `quality_auth_failure` — OE fails role test, off-topic, gibberish, AI text, platform fraud
- `quota_balancing` — respondent in over-filled quota cell, marginal quality
- `eligibility_screenout` — fails screener, wrong classification branch
- `partial_incomplete` — missing required fields, abandoned survey
- `vendor_source` — supplier reject rate, RD_Search threat
- `manual_admin` — human reviewer judgment, not detectable from data
- `unknown_mixed` — converging signals but no single clear cause
- `none` — no removal reason (KEEP)

## Badopen Audit Trail

For every respondent, classify the badopen trigger (even if the OE is good — use `none`):
- `duplicate_text` — OE text matches or closely paraphrases other respondents
- `too_short` — OE is under ~20 characters with no substantive content
- `pasted_text` — OE shows formatting artifacts, line breaks, or copy-paste markers
- `wrong_topic` — OE describes a project in the wrong domain
- `profanity` — OE contains profanity or abusive content
- `ai_like_similarity` — OE has AI-template language, benefit-stack, feature-list register
- `nonresponsive` — OE doesn't answer the question (e.g., "N/A", "good", "yes")
- `human_reviewer` — OE is borderline; a human would need to decide
- `none` — OE passes all badopen checks

## Output Format

Write to `agent_judgments_chunk_XX.json` as a JSON array. EVERY field is required.
Use null for fields that don't apply, but include all keys.

```json
[
  {{
    "respondent_id": "abc123",
    "agent_score": -0.7,
    "agent_judgment": "DISCARD",
    "agent_justification": "Core OE ('mowing and blowing') is on-topic but thin — no equipment named, no project detail. CLASSIFY=2 (consumer) but brand funnel shows no brand awareness. ML=0.72. Thin answer with converging quality concerns.",

    "authenticity_risk": 0.3,
    "quality_discard_risk": 0.7,
    "client_reject_probability": 0.6,

    "primary_removal_reason": "quality_auth_failure",
    "secondary_removal_reason": "vendor_source",
    "removal_confidence": 0.65,

    "evidence_families_fired": ["core_oe_quality", "model_risk", "brand_funnel"],
    "evidence_family_scores": {{
      "core_oe_quality": {{"fired": true, "score": 0.7, "trigger": "thin_on_topic"}},
      "platform_risk": {{"fired": false, "score": 0.1, "trigger": null}},
      "model_risk": {{"fired": true, "score": 0.72, "trigger": "ml_0.72_high"}},
      "source_risk": {{"fired": false, "score": 0.2, "trigger": null}},
      "duplicate_semantics": {{"fired": false, "score": 0.1, "trigger": null}},
      "survey_structure": {{"fired": false, "score": 0.2, "trigger": null}},
      "brand_funnel": {{"fired": true, "score": 0.5, "trigger": "no_brand_awareness_consistency"}},
      "timing_engagement": {{"fired": false, "score": 0.2, "trigger": null}},
      "quota_reconstruction": {{"fired": false, "score": 0.1, "trigger": null}}
    }},

    "badopen_trigger": "nonresponsive",
    "badopen_field": "qc5",
    "badopen_evidence": "Names generic tasks (mowing, blowing) without specific equipment or project narrative",
    "badopen_severity": "medium",

    "oe_classification": "thin_on_topic",
    "oe_equipment_named": [],
    "oe_grounding_anchors": [],
    "oe_word_count": 18,

    "ml_score": 0.72,
    "ml_top_signals": ["sig_rd_searchr1_2", "sig_very_short_required_open_end"],
    "ml_confidence": "high",

    "semantic_remapping": {{
      "core_oe_field": "qc5",
      "core_oe_role": "project_description",
      "classify_branch": "consumer",
      "channel_condition": "HD_or_OPE_dealers",
      "quota_cells": ["CLASSIFYQuota/Con", "RegionQuota/Northeast"]
    }},

    "stage1_fraud_verdict": "pass",
    "stage2_quality_verdict": "fail",
    "converging_family_count": 3
  }}
]
```

## Field Reference

### Required Fields (all 21 fields for every respondent)

| Field | Type | Values |
|-------|------|--------|
| respondent_id | string | from packet |
| agent_score | float | -1.0 to +1.0 |
| agent_judgment | string | DISCARD / REVIEW / KEEP |
| agent_justification | string | 2-4 sentences citing specific evidence |
| authenticity_risk | float | 0.0–1.0 |
| quality_discard_risk | float | 0.0–1.0 |
| client_reject_probability | float | 0.0–1.0 |
| primary_removal_reason | string | enum (see Disposition Layer) |
| secondary_removal_reason | string | enum or null |
| removal_confidence | float | 0.0–1.0 |
| evidence_families_fired | array | list of family names that fired |
| evidence_family_scores | object | per-family {{fired, score, trigger}} |
| badopen_trigger | string | enum (see Badopen Audit Trail) |
| badopen_field | string | OE field name or null |
| badopen_evidence | string | specific text evidence or null |
| badopen_severity | string | low / medium / high / none |
| oe_classification | string | substantive/thin_on_topic/off_topic/non_answer/gibberish/product_review/benefit_stack |
| oe_equipment_named | array | list of equipment/brand names mentioned |
| oe_grounding_anchors | array | list of grounding details (temporal, locational, sensory) |
| oe_word_count | int | word count of core OE |
| ml_score | float | from packet |
| ml_top_signals | array | top ML feature names contributing |
| ml_confidence | string | low / medium / high |
| semantic_remapping | object | {{core_oe_field, core_oe_role, classify_branch, channel_condition, quota_cells}} |
| stage1_fraud_verdict | string | pass / fail / ambiguous |
| stage2_quality_verdict | string | pass / fail / ambiguous |
| converging_family_count | int | number of families that fired |

### Decision Mapping

- **agent_score** = -(authenticity_risk * 0.4 + quality_discard_risk * 0.4 + client_reject_probability * 0.2)
  - Adjusted up/down by converging family count and grounding evidence
- **agent_judgment**:
  - DISCARD: agent_score < -0.3 AND removal_confidence > 0.5
  - REVIEW: agent_score -0.3 to 0 OR removal_confidence 0.3–0.5
  - KEEP: agent_score > 0 AND no families fired

## CORE FRAMEWORK: Three-Layer Review (v6)

AutoQuality operates through three linked layers:

### Layer 0 — Semantic Remapping (done before scoring)

Each workbook is semantically remapped before the agent reviews it. We do NOT treat fields as
generic columns. We reconstruct the survey contract:
- Which fields are screeners, quotas, brand funnels, matrices, timing fields, vendor fields,
  open ends, review markers, and final status fields
- Coded values are translated into their real survey meaning (e.g., CLASSIFY=1 → "professional",
  CLASSIFY=2 → "consumer")
- The review is based on what the respondent actually said or selected, not just raw codes

The `semantic_remapping` field in your output captures this: identify the core OE field, its
question role, the classification branch, the channel condition, and the quota cells.

### Layer 1 — ML Analysis (statistical guide, not final answer)

ML models are trained across annotated files to identify which signals separate accepted and
rejected respondents at scale. The ML score guides the review but is NOT the final answer.

ML signals include: source risk, timing distributions, duplicate text, marker fields, bad-open
flags, matrix behavior, and cross-field interactions.

Use the ML score as:
- **ml >= 0.8**: Very high risk — strong statistical evidence of discard. Combine with OE quality.
- **ml >= 0.7**: High risk — statistical evidence supports discard if OE is thin.
- **ml 0.5–0.7**: Ambiguous — use as tiebreaker with other families.
- **ml < 0.4**: Low risk — protective, but does NOT override a failed core OE.

Record `ml_score`, `ml_top_signals` (which features drove the score), and `ml_confidence`.

### Layer 1b — Cross-Corpus Signal Correlations (from 13,388 respondents across 11 datasets)

Each packet includes a `ml_signal_correlations` field with cross-corpus ML findings.
USE THIS TO CALIBRATE YOUR EVIDENCE FAMILY WEIGHTING:

**Family correlations with client discard (positive = client discards when this family fires):**
- `platform_risk`: mean_corr=+0.086 (STRONGEST — RD_Searchr1 is a top universal signal)
- `timing_engagement`: mean_corr=+0.073 (STRONG — qtime/low_total_duration fire in 10/11 datasets)
- `duplicate_semantics`: mean_corr=+0.042 (moderate — duplicate_open_chain fires in 5 datasets)
- `survey_structure`: mean_corr=+0.023 (weak positive — CLASSIFY matters in some datasets)
- `core_oe_quality`: mean_corr=-0.030 (NEGATIVE — client does NOT primarily discard on OE quality!)

**Key insight**: The client's discard process is driven MORE by platform risk and timing
than by OE quality. A thin-but-on-topic OE is NOT a strong predictor of client discard.
Do NOT discard solely for thin OE — require convergence with platform_risk or timing_engagement.

**Universal signals (fire across 3+ datasets, ranked by n_datasets):**
1. `low_total_duration` — 10 datasets (timing_engagement)
2. `text_time_mismatch` — 9 datasets (timing_engagement)
3. `no_strong_staged_signal` — 9 datasets (overall signal weakness)
4. `survey_meta_substitution` — 7 datasets (survey_structure)
5. `thin_open_end` — 6 datasets (core_oe_quality)
6. `RD_Searchr1` — 6 datasets (platform_risk)
7. `CLASSIFY` — 6 datasets (survey_structure)
8. `weak_persona_context` — 5 datasets (core_oe_quality)
9. `duplicate_open_chain` — 5 datasets (duplicate_semantics)

**Global model top features:**
1. CLASSIFY (importance=0.355) — survey_structure
2. qtime (importance=0.149) — timing_engagement
3. RD_Searchr1 (importance=0.137) — platform_risk
4. thin_open_end (importance=0.081) — core_oe_quality
5. no_strong_staged_signal (importance=0.055) — overall signal weakness

Use these correlations to weight your `client_reject_probability`:
- Platform risk + timing firing → HIGH client_reject_probability (even if OE is thin, not failed)
- OE quality alone firing → LOW client_reject_probability (client keeps thin OEs)
- Survey structure (CLASSIFY mismatch) + timing → MEDIUM client_reject_probability

### Layer 2 — Two-Stage Agent Review

**Stage 1 — Fraud Detection**: Is this respondent authentic? (bots, AI, platform flags, gibberish, non-English, duplicate chains)
→ Record as `authenticity_risk` and `stage1_fraud_verdict`

**Stage 2 — PM Quality Assessment**: Does this respondent meet the quality bar for this survey? (substantive engagement, brand funnel consistency, classification coherence, on-topic depth)
→ Record as `quality_discard_risk` and `stage2_quality_verdict`

### Layer 3 — Client Rejection Probability

Separate from authenticity and quality: would the client's own process discard this respondent?
The client may discard for quota, badopen, eligibility, or admin reasons that are NOT quality
failures. Record as `client_reject_probability`.

**Key principle**: A respondent can be authentic (low authenticity_risk) AND substantive (low
quality_discard_risk) but still discarded by the client (high client_reject_probability) due to
quota balancing or badopen triggers. Conversely, a respondent can be inauthentic (high
authenticity_risk) but kept by the client if they passed the client's review process.

The `primary_removal_reason` and `secondary_removal_reason` fields capture WHICH type of
removal this is — separating quality/auth failures from quota, eligibility, partial, admin,
and source exclusions.

### The Master Rule

**A row is discard-like when the core open end fails its question role, lacks grounded chain evidence, AND converges with at least one independent risk family — OR when the respondent fails the PM quality bar (thin engagement, brand funnel incoherence, classification mismatch, off-topic content).**

The nine independent evidence families:
1. **Core OE Quality** — answer-role test, grounded detail, substantiveness
2. **Platform Risk** — TERMFLAGS, qc, RD_Search, non-English
3. **Model Risk** — ML triage score
4. **Source Risk** — supplier reject rate, RD_Search threat
5. **Duplicate Semantics** — text similarity, paraphrase clusters
6. **Survey Structure** — CLASSIFY, PROAGE, conditions, list, channel coherence
7. **Brand Funnel Consistency** — awareness → rating → consideration → NPS chain
8. **Timing & Engagement** — speed, straightlining, matrix patterns
9. **Quota Reconstruction** — quota cell membership, over-filled cells, quota-aware quality bar

---

## Evidence Family 1: Core Open-End Quality (THE ANCHOR)

The core open-end field is the most important signal. Identify which OE field asks for personal motivation, experience, or project description. Read the question text to determine this.

### 1A. Answer-Role Test

Does the answer actually answer the question that was asked?

For a project-description question ("Describe a recent project involving outdoor power equipment"):
- **FAILS the role**: "Mowing and blowing/raking leaves" (names generic tasks, no project narrative), "Basic yard maintenance" (category language, no personal project)
- **PASSES the role**: "Cleaned up fallen branches after a storm using my battery-powered chainsaw and blower" (specific project, equipment named, temporal anchor)

For a purchase-motivation question ("What prompted you to decide to buy?"):
- **FAILS the role**: "Water filtration systems" (names the topic, not a motivation), "Health concerns" (category language)
- **PASSES the role**: "My water started smelling like chlorine after the city changed treatment plants" (specific personal trigger)

### 1B. Substantive Engagement Test (PM Quality Bar)

This is the key test that separates fraud detection from quality assessment. A respondent can be authentic but still fail the PM quality bar.

**The test**: Does the core OE demonstrate substantive engagement with the survey's specific topic?

- **Substantive**: Names specific equipment/products, describes a real project with temporal/locational anchors, shows understanding of the product category
- **Thin but on-topic**: Names the right category but with no detail ("Mowing and blowing", "Basic yard maintenance") — this PASSES fraud detection but FAILS the PM quality bar for strict clients
- **Off-topic**: Describes a project in the wrong domain (gardening/fertilizer when the survey is about outdoor power EQUIPMENT) — this is a quality failure
- **Generic first-person**: "I love cutting grasses" — authentic but not substantive

**IMPORTANT**: The substantive engagement threshold varies by dataset. Some clients tolerate generic first-person (Delta), others require specific detail (ECHO). When in doubt without calibration data, treat thin-but-on-topic as REVIEW, not KEEP. If ANY other family fires (ML >= 0.6, brand funnel inconsistency, survey structure mismatch), upgrade to DISCARD.

### 1C. Grounded First-Person Test

First-person pronouns are NOT protective by themselves. The test is whether the content is GROUNDED in lived experience.

**Grounded evidence** connects to:
- A concrete event ("after the storm last week", "when my neighbor showed me their new mower")
- A household condition ("my yard has a steep hill", "our property is 2 acres")
- A product use detail ("I change the oil every season", "the pull cord broke")
- A sensory issue ("the engine kept stalling", "leaves everywhere")
- A specific place ("here in Colorado where the snow is heavy")
- A buying context ("was at Home Depot getting mulch", "needed a new trimmer after mine died")

**NOT grounded** (even if first-person):
- "I was concerned about the quality" (no concrete anchor)
- "I wanted a better yard" (benefit-stack language, no lived context)

### 1D. Off-Topic Detection

Check whether the core OE is actually about the survey's subject matter. Read the question text to determine what the survey is about, then check:

- Is the described project in the right domain? (OPE survey → chainsaws, mowers, trimmers, blowers — NOT gardening, pest control, indoor projects)
- Does the equipment mentioned match the survey topic? (OPE survey → power tools — NOT shovels, hand clippers, fertilizer)
- Is the project plausible for the survey audience? (Consumer OPE → residential yard work — NOT commercial landscaping unless CLASSIFY=pro)

**Off-topic core OE is a quality failure**, even if the text is authentic and well-written. Discard or review depending on severity.

### 1E. Product-Copy Register

Marketing-like language is suspicious:
- Benefit-stack: "enhancing curb appeal", "ensuring consistent depth", "multi-stage filtration"
- Feature-list: "contaminant removal technologies", "advanced filtration"
- **Stronger when the answer has no lived context**

### 1F. Synthetic Detail Detection

Some details sound specific but are common synthetic clusters:
- "my skin" / "my family" / "my health" — common fake specifics
- "chlorine" / "contaminants" / "hard water" — product-category vocabulary

Genuinely specific details are idiosyncratic:
- Named brands: "Stihl", "Husqvarna", "Echo" (not just "a chainsaw")
- Named conditions: "eczema", "psoriasis" (not just "skin")
- Unusual sensory details: "smelled like rotten eggs", "left orange stains"

---

## Evidence Family 2: Platform Risk

Read the `defender_summary` field FIRST.

- **TERMFLAGS=1**: Platform fraud flag. STRONG (-0.7). Override to REVIEW only with unusually strong human evidence (specific named event + chain-consistent + no AI markers).
- **qc=8 or qc=9**: AUTOMATIC DISCARD (-1.0)
- **qc=6**: RD /REVIEW rejection — strong discard
- **qc=7**: OE screening failure — strong concern
- **qc=11**: Speeder — moderate concern
- **RD_Search threat >= 25**: Elevated. >= 20: Moderate.
- **Non-English**: Automatic discard for US surveys
- **LangAssess read_level >= 17**: Very high (AI suspicion). >= 15 with short text: AI suspicion.

---

## Evidence Family 3: Model Risk

- **ML >= 0.8**: Very high. If core OE also fails, discard. If core OE is grounded, cap at REVIEW.
- **ML >= 0.7**: High. If core OE is thin/generic, discard. If grounded, cap at REVIEW.
- **ML 0.5-0.7**: Ambiguous. Tiebreaker — if any other family fires, lean discard.
- **ML < 0.4**: Low risk. Protective, but does NOT override a failed core OE.

---

## Evidence Family 4: Source Risk

- **Supplier reject_rate > 30%**: High risk. Multiplier.
- **Supplier reject_rate > 20%**: Medium risk.
- **RD_Search threat >= 20**: Moderate.
- Source risk alone is NOT sufficient for discard. It is a multiplier.

---

## Evidence Family 5: Duplicate Semantics

- **`ai_text_suspicion` on core OE field**: STRONG. Text is similar to others — likely templated.
- **`ai_text_suspicion` on outro ONLY**: DOWNWEIGHT. Topical inevitability. But + weak core OE + any risk family = meaningful.
- **`oe_duplicate_counts`**: High count (>10) on personal text = suspicious. On generic text = topical inevitability.
- **Paraphrase clusters**: Does the core OE tell the same story frame as many others?

---

## Evidence Family 6: Survey Structure (NEW — v5)

Read the `survey_structure` field. These fields carry 2x+ discrimination power but are NOT semantic content — they are classification, quota, and channel fields.

### 6A. Classification (CLASSIFY)

- **CLASSIFY=1 (professional)**: Pro respondents are held to a HIGHER standard. They should show professional purchasing patterns (dealer channels, commercial equipment, volume). A pro who answers like a consumer is a quality failure.
- **CLASSIFY=2 (consumer)**: Consumer respondents are held to the standard of a homeowner who uses OPE. They should show residential yard work with retail-purchased equipment.
- **Mismatch**: A CLASSIFY=1 (pro) respondent who describes a tiny residential yard with one mower is classification-incoherent. A CLASSIFY=2 (consumer) who describes commercial landscaping is also suspicious.

### 6B. Pro Age (PROAGE) / Consumer Age (CONAGE)

- PROAGE present means the respondent is in the pro branch. Their answers should reflect professional experience.
- CONAGE present means consumer branch. Their answers should reflect consumer experience.
- A respondent with PROAGE but no professional evidence in their OE is a quality concern.

### 6C. Channel Conditions

- `conditionsAriens=1`: Respondent is in the Ariens dealer channel. Their brand answers should include Ariens products.
- `conditionsHD_or_OPE_dealers=1`: Home Depot / OPE dealer channel. Brand answers should match.
- `conditionsOther_channel=1`: Other channel.
- **Channel-brand mismatch**: A respondent in the Ariens channel who never mentions Ariens in any brand field is a quality concern.

### 6D. Supplier/List

- `list` / `vlist` / `source`: Different suppliers have different reject rates. This is context, not proof.
- But if a supplier has a known high reject rate AND the respondent shows other concerns, the combination is stronger.

### How survey structure combines:

- CLASSIFY mismatch + thin core OE → strong quality concern (lean DISCARD)
- Channel-brand mismatch + brand funnel incoherence → strong quality concern
- PROAGE present + no professional evidence → quality concern (REVIEW)
- Survey structure alone is NOT sufficient for discard. It is a multiplier on core OE quality.

---

## Evidence Family 7: Brand Funnel Consistency (NEW — v5)

Read the `brand_funnel` field. Check whether the brand funnel is internally consistent.

### 7A. Awareness → Consideration → Recommendation Chain

- Does the respondent claim awareness of brands they later cannot rate or consider?
- Does the respondent recommend brands they did not claim awareness of?
- Does the respondent rate brands they did not claim awareness of?

### 7B. Brand Name Quality in OE Fields

- Are brand names in open-end fields real brands? (Stihl, Husqvarna, Echo, Honda, Ryobi, Toro, Craftsman = real OPE brands)
- Are brand names garbled or fake? ("Harmmer", "china", "Mercedes" for OPE = wrong brand universe)
- Are brand names consistent across fields? (q17r1 says "Honda" but q17r2 says "Costoc" = inconsistency)

### 7C. Share Allocation

- If the respondent allocated share across brands, is the allocation plausible?
- Equal share to all brands = potential straightlining in the brand battery
- All share to one brand = possible extreme opinion (not necessarily fraud)

### 7D. NPS / Satisfaction Verbatim

- The NPS verbatim (q29) should be about a specific brand, not generic praise
- "Effective work and power" = generic, not brand-specific
- "They're pretty proud of their name and put it in VERY LARGE letters on their products" = brand-specific, grounded

### How brand funnel combines:

- Brand funnel incoherence + thin core OE → strong quality concern (lean DISCARD)
- Wrong brand universe (non-OPE brands in OPE survey) → quality failure
- Brand funnel alone is NOT sufficient for discard. It combines with core OE quality.

---

## Evidence Family 8: Timing & Engagement

- **bottom_10% timing**: Very fast. Concern, not sufficient alone.
- **bottom_25% timing**: Fast. Mild concern.
- **above_median timing**: Protective.
- **Very fast + failed core OE + any risk family** = converging → discard
- **Straightlining**: If prevalence >80% (see dataset context), straightlining alone is NOT discriminative.

---

## Evidence Family 9: Quota Reconstruction (NEW — v5.1)

Read the `quota_reconstruction` field. The client's discard process uses quota markers that are NOT in the raw workbook. We reconstruct quota bucket membership from raw fields to give visibility into which quota cells this respondent occupies.

### 9A. Quota Cell Membership

The `quota_cells` array shows which quota buckets this respondent fills:
- `CLASSIFYQuota/{{label}}` — pro or consumer classification bucket
- `RegionQuota/{{label}}` — geographic region bucket
- `GenderQuota/{{label}}` — gender bucket
- `AgeQuota/{{label}}` or `CONAgeQuota/{{label}}` — age bracket bucket
- `ChannelQuota/{{label}}` — channel condition bucket (Ariens, HD or OPE dealers, Other)
- `BRANDS2RATEQuota/{{N}}_brands` — number of brands rated
- `TotalQuota/Total` — all respondents

### 9B. How to use quota reconstruction

**Quota reconstruction is NOT a direct discard signal.** The client uses quota markers to track which buckets are filled, but the discard decision is based on `badopen` (open-end quality) — NOT on quota membership itself. Both kept and discarded respondents have the same quota markers; the difference is the `bad:` prefix on discards.

However, quota reconstruction IS useful for:
1. **Understanding which quota cells are over-represented** — if a respondent is in a cell with many other discards, the cell may be over-filled and the client is more likely to discard
2. **Identifying classification-channel-brand coherence** — a respondent in the Ariens channel quota who never mentions Ariens has a channel-brand mismatch
3. **Pro vs consumer quota context** — CLASSIFY=1 (pro) respondents are in a smaller, more scrutinized quota bucket with higher discard rates (60.4% on ECHO)

### 9C. Quota-aware quality bar

The client's `badopen` standard is applied WITHIN quota cells. This means:
- A respondent in a over-filled quota cell (many respondents in the same bucket) faces a stricter quality bar
- A respondent in an under-filled cell may be kept despite marginal quality
- The `population_in_cells` field shows how many other respondents share this respondent's CLASSIFY and REGION buckets

### How quota reconstruction combines:

- Over-filled quota cell + thin core OE → stronger quality concern (the client is more likely to apply badopen)
- CLASSIFY=1 (pro) quota + consumer-like answers → classification mismatch (already covered in Family 6)
- Channel quota mismatch (Ariens channel, no Ariens brand) → brand funnel concern (already covered in Family 7)
- Quota reconstruction is a MULTIPLIER on other families, not a standalone signal

---

## The Decision Algorithm

For each respondent, work through these steps and fill in ALL output fields:

### Step 1: Semantic Remapping
Read `defender_summary`, `survey_structure`, `brand_funnel`, `quota_reconstruction`.
Identify: What is the core OE field? What role does it play? What is the classification branch?
What channel condition? What quota cells? → Fill `semantic_remapping` object.

### Step 2: Stage 1 — Fraud Detection (authenticity_risk)
Read platform signals: TERMFLAGS, qc, RD_Search, ReadLevel, ai_text_suspicion.
Is this respondent real? → Fill `authenticity_risk` (0–1) and `stage1_fraud_verdict`.

### Step 3: Stage 2 — OE Quality Assessment (quality_discard_risk)
Identify the core OE field. Apply answer-role test, substantive engagement test,
grounded first-person test, off-topic detection. → Fill `oe_classification`,
`oe_equipment_named`, `oe_grounding_anchors`, `oe_word_count`, `quality_discard_risk`,
`stage2_quality_verdict`.

### Step 4: Badopen Audit Trail
Classify the badopen trigger for the core OE. Is it too_short, wrong_topic, nonresponsive,
ai_like_similarity, duplicate_text, pasted_text, profanity, human_reviewer, or none?
→ Fill `badopen_trigger`, `badopen_field`, `badopen_evidence`, `badopen_severity`.

### Step 5: ML Analysis
Read `ml_triage_score`. What are the top ML signals? How confident is the model?
→ Fill `ml_score`, `ml_top_signals`, `ml_confidence`.

### Step 6: Evidence Family Scoring
For EACH of the 9 evidence families, assess: did it fire? What score (0–1)? What trigger?
→ Fill `evidence_family_scores` object and `evidence_families_fired` array.
→ Fill `converging_family_count`.

### Step 7: Client Rejection Probability
Considering quota cells, badopen trigger, classification coherence, source risk:
Would the client's own process discard this? → Fill `client_reject_probability` (0–1).

### Step 8: Disposition Layer
What is the PRIMARY removal reason? What is the SECONDARY?
→ Fill `primary_removal_reason`, `secondary_removal_reason`, `removal_confidence`.

### Step 9: Final Decision
Compute `agent_score` from the three components:
  agent_score = -(authenticity_risk * 0.4 + quality_discard_risk * 0.4 + client_reject_probability * 0.2)
  Adjust: +0.1 if grounded equipment named, +0.05 if above_median timing,
  -0.05 per additional converging family beyond 2.
→ Fill `agent_score`, `agent_judgment`, `agent_justification`.

### Decision thresholds:

**DISCARD** (score < -0.3 AND removal_confidence > 0.5) when:
- Core OE fails its role AND >= 1 risk family fires
- Core OE is thin/generic AND >= 2 families fire
- Core OE is off-topic AND any risk family fires
- TERMFLAGS=1 AND core OE is not unusually specifically grounded
- ML >= 0.8 AND core OE fails role or is generic
- Brand funnel shows wrong brand universe AND core OE is thin
- CLASSIFY mismatch AND core OE is thin
- Core OE is a short non-answer AND >= 1 risk family fires
- primary_removal_reason = quality_auth_failure with high confidence

**REVIEW** (score -0.3 to 0 OR removal_confidence 0.3–0.5) when:
- Core OE is thin but on-topic with no other concerns
- Core OE is grounded but one risk family fires
- ML >= 0.7 AND core OE is grounded and specific
- primary_removal_reason = quota_balancing or unknown_mixed
- Mixed signals where neither KEEP nor DISCARD is clearly warranted

**KEEP** (score > 0 AND no families fired) when:
- Core OE passes answer-role test AND shows substantive engagement AND no risk family fires
- Core OE is specific, grounded, and chain-consistent AND ML < 0.5 AND no platform flags
- primary_removal_reason = none

### Thin-but-on-topic rule (IMPORTANT):
An answer like "Mowing and blowing/raking leaves" is authentic and on-topic for an OPE survey,
but it does NOT demonstrate substantive engagement. It names generic tasks without equipment,
project narrative, or personal detail.

For the three-component model:
- authenticity_risk = LOW (this is a real human answer)
- quality_discard_risk = MEDIUM (thin but on-topic)
- client_reject_probability = VARIABLE (depends on client's badopen standard)

Without calibration data, score this as REVIEW (-0.1 to -0.2), not KEEP.
If ANY other family fires (ML >= 0.6, brand funnel inconsistency, survey structure mismatch),
upgrade to DISCARD.

### Short non-answer rule:
Short answers (<25 chars) to a motivation/project question are almost always role failures.
"It's essential", "Snow Blower", "Home Depot" are NOT project descriptions.
badopen_trigger = too_short or nonresponsive.
When paired with ANY risk family signal, this becomes a discard.

### Accepted-row similarity guardrail:
Before discarding, ask: what makes this row different from accepted rows with the same surface
flaw? The discard should come from CONVERGENCE of multiple families, not from any single signal
common among accepted rows. If the client kept similar rows, lower client_reject_probability.

---

## Packet Field Reference

- `respondent_id`: Unique identifier
- `ml_triage_score`: 0-1 risk probability from pre-trained model
- `timing`: minutes + percentile (bottom_10, bottom_25, above_median, above_p75)
- `supplier`: name + reject_rate
- `defender_summary`: Consolidated platform signals (read this first)
- `defender_signals`: Raw platform signal values
- `ai_text_suspicion`: Cross-respondent text similarity (score, fields_flagged)
- `lang_assess`: Readability metrics (read_level, num_words, etc.)
- `matrix_analysis`: Per-grid straightlining analysis
- `open_end_responses`: All OE fields with question text + response + char_count
- `oe_duplicate_counts`: How many respondents share the same OE text per field
- `key_answers`: All coded single-choice fields with labels (dynamically discovered)
- `survey_structure`: CLASSIFY, PROAGE, CONAGE, conditions, list, source, dcua, FIRMREV
- `brand_funnel`: Awareness, rating, consideration, recommendation, NPS, satisfaction fields
- `quota_reconstruction`: Quota cell membership (CLASSIFYQuota, RegionQuota, GenderQuota, ChannelQuota, BRANDS2RATEQuota, TotalQuota) + population counts per cell
- `answer_entropy`: Variety in the answer chain
- `oe_total_chars`: Total characters across all OE fields

## Dataset Context
- Dataset: {dataset_name}
- Total respondents: {n_respondents}
- Matrix straightlining prevalence: {matrix_prevalence:.1%} {'(VERY HIGH — straightlining alone is NOT discriminative)' if matrix_prevalence > 0.8 else ''}
- You are reviewing chunk XX of {n_chunks}
- NOTE: Read the question text in each OE field to identify the survey topic, core OE field, and expected evidence type. Adapt all tests to the actual survey subject.
"""


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 run_holistic_agent_review.py <xlsx_path> [--output-dir DIR] [--review-all] [--chunk-size N]")
        return

    filepath = Path(sys.argv[1])
    output_dir = None
    review_all = False
    chunk_size = 200

    for i, arg in enumerate(sys.argv[2:], 2):
        if arg == "--output-dir" and i + 1 < len(sys.argv):
            output_dir = Path(sys.argv[i + 1])
        elif arg == "--review-all":
            review_all = True
        elif arg == "--chunk-size" and i + 1 < len(sys.argv):
            chunk_size = int(sys.argv[i + 1])

    build_holistic_review_packets(filepath, output_dir or filepath.parent / "holistic_review_output", review_all, chunk_size)


if __name__ == "__main__":
    main()
