#!/usr/bin/env python3
"""Phase 3: Integrate agent judgments into the annotated Excel + dashboard.

Reads agent_judgments.json (produced by subagent review) and re-generates
the annotated Excel and dashboard with the agent's natural-language
justifications and scores.

Usage:
    python3 integrate_agent_judgments.py <xlsx_path> <output_dir>
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

# Add the skills scripts directory
SKILL_SCRIPTS = Path(__file__).parent.parent / "skills" / "cleaning-survey-quality" / "scripts"
sys.path.insert(0, str(SKILL_SCRIPTS))

from survey_pipeline import (
    extract_features_and_chain,
    ml_triage,
    agent_score_respondent,
    classify_open_end_semantic,
    compute_key_signals,
    write_annotated_excel,
    write_dashboard,
)


def integrate_agent_judgments(filepath, output_dir):
    """Integrate agent judgments into the output."""
    filepath = Path(filepath)
    output_dir = Path(output_dir)

    # Load agent judgments
    judgments_path = output_dir / "agent_judgments.json"
    if not judgments_path.exists():
        print(f"ERROR: Agent judgments not found at {judgments_path}")
        print("Run the subagent review first to produce agent_judgments.json")
        return None

    with open(judgments_path) as f:
        agent_judgments = json.load(f)

    print(f"\n{'='*80}")
    print(f"INTEGRATING AGENT JUDGMENTS")
    print(f"{'='*80}")
    print(f"  Input: {filepath.name}")
    print(f"  Agent judgments: {len(agent_judgments)}")

    # Build lookup
    judgment_lookup = {j["respondent_id"]: j for j in agent_judgments}

    # Load review packets for defender_summary + ai_text_suspicion
    defender_lookup = {}
    ai_suspicion_lookup = {}
    for chunk_file in sorted(output_dir.glob("review_chunk_*.json")):
        with open(chunk_file) as f:
            packets = json.load(f)
        for p in packets:
            rid = p["respondent_id"]
            defender_lookup[rid] = p.get("defender_summary", "")
            ai_info = p.get("ai_text_suspicion", {})
            if ai_info and ai_info.get("score", 0) > 0:
                ai_suspicion_lookup[rid] = f"score={ai_info['score']:.2f} fields={ai_info.get('fields_flagged', [])} | {ai_info.get('details', '')}"
            else:
                ai_suspicion_lookup[rid] = "none"

    # Extract v6 metadata fields from judgments (with fallbacks for v5 judgments)
    v6_metadata_lookup = {}
    for j in agent_judgments:
        rid = j["respondent_id"]
        v6_metadata_lookup[rid] = {
            "authenticity_risk": j.get("authenticity_risk"),
            "quality_discard_risk": j.get("quality_discard_risk"),
            "client_reject_probability": j.get("client_reject_probability"),
            "primary_removal_reason": j.get("primary_removal_reason"),
            "secondary_removal_reason": j.get("secondary_removal_reason"),
            "removal_confidence": j.get("removal_confidence"),
            "evidence_families_fired": j.get("evidence_families_fired", []),
            "badopen_trigger": j.get("badopen_trigger"),
            "badopen_field": j.get("badopen_field"),
            "badopen_evidence": j.get("badopen_evidence"),
            "badopen_severity": j.get("badopen_severity"),
            "oe_classification": j.get("oe_classification"),
            "oe_equipment_named": j.get("oe_equipment_named", []),
            "oe_grounding_anchors": j.get("oe_grounding_anchors", []),
            "oe_word_count": j.get("oe_word_count"),
            "ml_top_signals": j.get("ml_top_signals", []),
            "ml_confidence": j.get("ml_confidence"),
            "stage1_fraud_verdict": j.get("stage1_fraud_verdict"),
            "stage2_quality_verdict": j.get("stage2_quality_verdict"),
            "converging_family_count": j.get("converging_family_count"),
        }

    # Re-run feature extraction
    print(f"\n[1/4] Extracting features...")
    df, datamap, roles, answer_chains = extract_features_and_chain(filepath)
    df = ml_triage(df)

    # Compute matrix prevalence
    matrix_prevalence = (df["matrix_straightline"] == 1).mean() if "matrix_straightline" in df.columns else None

    # Get rule-based scores for all respondents
    print(f"\n[2/4] Computing rule-based scores...")
    rule_scores = []
    rule_reasons = []
    for idx, row in df.iterrows():
        chain = answer_chains[idx] if idx < len(answer_chains) else {}
        score, reasons = agent_score_respondent(chain, row["ml_triage_score"], matrix_prevalence=matrix_prevalence)
        rule_scores.append(score)
        rule_reasons.append(reasons)
    df["rule_based_score"] = rule_scores
    df["rule_reasons"] = rule_reasons

    # Override with agent judgments where available
    print(f"\n[3/4] Integrating agent judgments...")
    final_scores = []
    final_judgments = []
    agent_justifications = []
    agent_scores_list = []
    reassessment_notes = []

    n_agent = 0
    n_rule = 0

    for idx, row in df.iterrows():
        rid = row["respondent_id"]
        if rid in judgment_lookup:
            j = judgment_lookup[rid]
            final_scores.append(float(j["agent_score"]))
            final_judgments.append(j["agent_judgment"])
            agent_justifications.append(j.get("agent_justification", ""))
            agent_scores_list.append(float(j["agent_score"]))
            reassessment_notes.append(["Agent reviewed this respondent"])
            n_agent += 1
        else:
            # Use rule-based score
            score = row["rule_based_score"]
            final_scores.append(score)
            agent_scores_list.append(score)
            agent_justifications.append("")
            if score >= 0:
                final_judgments.append("KEEP")
            else:
                final_judgments.append("REVIEW")
            reassessment_notes.append(["Rule-based — not reviewed by agent"])
            n_rule += 1

    df["agent_score"] = agent_scores_list
    df["final_score"] = final_scores
    df["final_judgment"] = final_judgments
    df["agent_justification"] = agent_justifications
    df["reassessment_notes"] = reassessment_notes
    df["key_signals"] = compute_key_signals(df, answer_chains)
    df["agent_reasons"] = rule_reasons  # Keep rule reasons for reference

    print(f"  Agent-reviewed: {n_agent}")
    print(f"  Rule-based only: {n_rule}")

    n_discard = int((df["final_judgment"] == "DISCARD").sum())
    n_review = int((df["final_judgment"] == "REVIEW").sum())
    n_keep = int((df["final_judgment"] == "KEEP").sum())
    print(f"\n  Final judgments:")
    print(f"    DISCARD: {n_discard} ({n_discard/len(df):.1%})")
    print(f"    REVIEW:  {n_review} ({n_review/len(df):.1%})")
    print(f"    KEEP:    {n_keep} ({n_keep/len(df):.1%})")

    # Generate outputs with agent justifications
    print(f"\n[4/4] Generating outputs with agent justifications...")

    excel_path = output_dir / f"{filepath.stem}_annotated.xlsx"
    write_annotated_excel_with_agent(filepath, df, excel_path, defender_lookup, ai_suspicion_lookup, v6_metadata_lookup)
    print(f"  Annotated Excel: {excel_path}")

    dashboard_path = output_dir / f"{filepath.stem}_dashboard.html"
    write_dashboard_with_agent(df, answer_chains, filepath.name, dashboard_path)
    print(f"  Dashboard: {dashboard_path}")

    # Summary
    summary = {
        "dataset": filepath.name,
        "total_respondents": int(len(df)),
        "agent_reviewed": n_agent,
        "rule_based_only": n_rule,
        "discard": n_discard,
        "review": n_review,
        "keep": n_keep,
        "mean_agent_score": float(df["agent_score"].mean()),
        "mean_final_score": float(df["final_score"].mean()),
        "mean_ml_triage": float(df["ml_triage_score"].mean()),
    }
    summary_path = output_dir / "summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"  Summary: {summary_path}")

    print(f"\n{'='*80}")
    print(f"COMPLETE — Agent judgments integrated")
    print(f"{'='*80}")

    return df


def write_annotated_excel_with_agent(original_path, df, excel_path, defender_lookup=None, ai_suspicion_lookup=None):
    """Write annotated Excel with agent justifications."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb = openpyxl.load_workbook(original_path)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    rid_col = hidx.get("uuid") or hidx.get("record")
    if rid_col is None:
        print("WARNING: No respondent ID column found")
        return

    # Build lookups
    score_lookup = dict(zip(df["respondent_id"], df["final_score"]))
    judgment_lookup = dict(zip(df["respondent_id"], df["final_judgment"]))
    ml_lookup = dict(zip(df["respondent_id"], df["ml_triage_score"]))
    agent_score_lookup = dict(zip(df["respondent_id"], df["agent_score"]))
    signals_lookup = dict(zip(df["respondent_id"], df["key_signals"]))
    justification_lookup = dict(zip(df["respondent_id"], df["agent_justification"]))
    notes_lookup = {}
    for _, row in df.iterrows():
        notes = row.get("reassessment_notes", [])
        notes_lookup[row["respondent_id"]] = "; ".join(notes) if isinstance(notes, list) else str(notes)

    # Add annotation columns
    n_cols = len(headers)
    annotation_headers = [
        "ML_Triage_Score",
        "Agent_Score",
        "Final_Score",
        "Final_Judgment",
        "Agent_Justification",
        "Key_Signals",
        "Reassessment_Notes",
        "Defender_Summary",
        "AI_Text_Suspicion",
    ]
    for i, h in enumerate(annotation_headers):
        col = n_cols + 1 + i
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    row_num = 2
    for row in ws.iter_rows(min_row=2, values_only=False):
        rid_val = row[rid_col].value if rid_col < len(row) else None
        rid = str(rid_val).strip() if rid_val else ""
        if not rid or rid not in score_lookup:
            row_num += 1
            continue

        ws.cell(row=row_num, column=n_cols + 1, value=round(float(ml_lookup[rid]), 4))
        ws.cell(row=row_num, column=n_cols + 2, value=round(float(agent_score_lookup[rid]), 4))
        ws.cell(row=row_num, column=n_cols + 3, value=round(float(score_lookup[rid]), 4))

        judgment = judgment_lookup[rid]
        cell = ws.cell(row=row_num, column=n_cols + 4, value=judgment)
        if judgment == "DISCARD":
            cell.fill = red_fill
        elif judgment == "REVIEW":
            cell.fill = yellow_fill
        else:
            cell.fill = green_fill

        # Agent justification (the key new column)
        just = justification_lookup.get(rid, "")
        ws.cell(row=row_num, column=n_cols + 5, value=just)

        signals = signals_lookup[rid]
        ws.cell(row=row_num, column=n_cols + 6, value="; ".join(signals) if isinstance(signals, list) else str(signals))
        ws.cell(row=row_num, column=n_cols + 7, value=notes_lookup.get(rid, ""))
        ws.cell(row=row_num, column=n_cols + 8, value=(defender_lookup or {}).get(rid, ""))
        ws.cell(row=row_num, column=n_cols + 9, value=(ai_suspicion_lookup or {}).get(rid, "none"))

        row_num += 1

    # Auto-fit column widths
    for i in range(len(annotation_headers)):
        col = n_cols + 1 + i
        if i == 4:  # Justification
            width = 50
        elif i == 7:  # Defender_Summary
            width = 50
        elif i == 8:  # AI_Text_Suspicion
            width = 40
        else:
            width = 25
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(excel_path)


def write_dashboard_with_agent(df, answer_chains, dataset_name, output_path):
    """Write dashboard with agent justifications in the discard table."""
    # Reuse the standard dashboard but add agent justifications to the discard table
    from survey_pipeline import write_dashboard
    # The standard dashboard doesn't show justifications, so we'll write a custom one
    # that includes them in the discard table

    n = len(df)
    n_discard = int((df["final_judgment"] == "DISCARD").sum())
    n_review = int((df["final_judgment"] == "REVIEW").sum())
    n_keep = int((df["final_judgment"] == "KEEP").sum())

    # Score distribution
    buckets = {"-1.0 to -0.5": 0, "-0.5 to 0.0": 0, "0.0 to 0.5": 0, "0.5 to 1.0": 0}
    for s in df["final_score"]:
        if s < -0.5: buckets["-1.0 to -0.5"] += 1
        elif s < 0: buckets["-0.5 to 0.0"] += 1
        elif s < 0.5: buckets["0.0 to 0.5"] += 1
        else: buckets["0.5 to 1.0"] += 1

    # Supplier analysis
    sup_stats = df.groupby("supplier_name").agg(
        count=("respondent_id", "count"),
        mean_score=("final_score", "mean"),
        discards=("final_judgment", lambda x: (x == "DISCARD").sum()),
    ).sort_values("count", ascending=False).head(15)

    # Top signals
    from collections import Counter
    signal_counts = Counter()
    for signals in df["key_signals"]:
        if isinstance(signals, list):
            for s in signals:
                key = s.split(":")[0].split("(")[0].strip()
                signal_counts[key] += 1

    # Discard table with agent justifications
    discard_df = df[df["final_judgment"] == "DISCARD"].sort_values("final_score")
    discard_rows = []
    for _, row in discard_df.head(100).iterrows():
        signals = "; ".join(row["key_signals"][:3]) if isinstance(row["key_signals"], list) else ""
        just = row.get("agent_justification", "") or ""
        discard_rows.append({
            "id": row["respondent_id"],
            "score": round(float(row["final_score"]), 2),
            "ml": round(float(row["ml_triage_score"]), 2),
            "supplier": str(row.get("supplier_name", "")),
            "justification": just,
            "signals": signals,
        })

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Survey Quality Dashboard — {dataset_name}</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; color: #333; }}
  .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
  h1 {{ font-size: 24px; margin-bottom: 8px; }}
  h2 {{ font-size: 18px; margin: 24px 0 12px; padding-bottom: 8px; border-bottom: 2px solid #ddd; }}
  .subtitle {{ color: #666; margin-bottom: 20px; font-size: 14px; }}
  .cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 24px; }}
  .card {{ background: white; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
  .card .label {{ font-size: 12px; color: #888; text-transform: uppercase; letter-spacing: 0.5px; }}
  .card .value {{ font-size: 28px; font-weight: 700; margin-top: 4px; }}
  .card.discard .value {{ color: #d32f2f; }}
  .card.review .value {{ color: #f9a825; }}
  .card.keep .value {{ color: #388e3c; }}
  .card.total .value {{ color: #1976d2; }}
  .card .pct {{ font-size: 12px; color: #888; margin-top: 2px; }}
  .panel {{ background: white; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 24px; }}
  .panel h3 {{ font-size: 14px; color: #555; margin-bottom: 12px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .bar-row {{ display: flex; align-items: center; margin-bottom: 6px; }}
  .bar-label {{ width: 120px; font-size: 13px; color: #555; }}
  .bar-track {{ flex: 1; height: 20px; background: #eee; border-radius: 4px; overflow: hidden; }}
  .bar-fill {{ height: 100%; border-radius: 4px; }}
  .bar-fill.red {{ background: #e53935; }}
  .bar-fill.yellow {{ background: #fdd835; }}
  .bar-fill.green {{ background: #43a047; }}
  .bar-value {{ width: 60px; text-align: right; font-size: 13px; color: #555; margin-left: 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ text-align: left; padding: 8px; background: #f5f5f5; border-bottom: 2px solid #ddd; font-weight: 600; }}
  td {{ padding: 8px; border-bottom: 1px solid #eee; vertical-align: top; }}
  tr:hover {{ background: #f9f9f9; }}
  .score-negative {{ color: #d32f2f; font-weight: 700; }}
  .score-positive {{ color: #388e3c; font-weight: 700; }}
  .justification {{ font-size: 12px; color: #555; line-height: 1.4; max-width: 400px; }}
  .signal-list {{ list-style: none; }}
  .signal-list li {{ padding: 4px 0; border-bottom: 1px solid #eee; font-size: 13px; }}
  .signal-list li:last-child {{ border-bottom: none; }}
  .signal-count {{ float: right; color: #888; font-weight: 600; }}
  .agent-badge {{ display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 11px; font-weight: 600; background: #e3f2fd; color: #1565c0; margin-left: 8px; }}
  .footer {{ margin-top: 24px; padding: 12px; text-align: center; color: #888; font-size: 12px; }}
</style>
</head>
<body>
<div class="container">
  <h1>Survey Quality Dashboard <span class="agent-badge">Agent-Reviewed</span></h1>
  <div class="subtitle">{dataset_name} — {n} respondents — Agent judgments integrated</div>

  <div class="cards">
    <div class="card total"><div class="label">Total</div><div class="value">{n}</div></div>
    <div class="card discard"><div class="label">Discard</div><div class="value">{n_discard}</div><div class="pct">{n_discard/n:.1%}</div></div>
    <div class="card review"><div class="label">Review</div><div class="value">{n_review}</div><div class="pct">{n_review/n:.1%}</div></div>
    <div class="card keep"><div class="label">Keep</div><div class="value">{n_keep}</div><div class="pct">{n_keep/n:.1%}</div></div>
  </div>

  <div class="panel">
    <h3>Agent Score Distribution (-1 to +1)</h3>
"""
    for label, count in buckets.items():
        pct = count / n * 100 if n > 0 else 0
        color = "red" if "-0.5" in label or "-1.0" in label else ("yellow" if "0.0" in label else "green")
        html += f'    <div class="bar-row"><div class="bar-label">{label}</div><div class="bar-track"><div class="bar-fill {color}" style="width: {pct}%"></div></div><div class="bar-value">{count}</div></div>\n'

    html += f"""  </div>

  <div class="panel">
    <h3>Top Population Signals</h3>
    <ul class="signal-list">
"""
    for signal, count in signal_counts.most_common(15):
        pct = count / n * 100
        html += f'      <li>{signal}<span class="signal-count">{count} ({pct:.1f}%)</span></li>\n'

    html += f"""    </ul>
  </div>

  <div class="panel">
    <h3>Supplier Analysis (Top 15)</h3>
    <table>
      <thead><tr><th>Supplier</th><th>N</th><th>Mean Score</th><th>Discards</th></tr></thead>
      <tbody>
"""
    for sup, row in sup_stats.iterrows():
        if not sup: sup = "(missing)"
        score_class = "score-negative" if row["mean_score"] < 0 else "score-positive"
        html += f'      <tr><td>{str(sup)[:25]}</td><td>{int(row["count"])}</td><td class="{score_class}">{row["mean_score"]:.2f}</td><td>{int(row["discards"])}</td></tr>\n'

    html += """    </tbody>
    </table>
  </div>

  <div class="panel">
    <h3>Discard Set with Agent Justifications """ + f"({n_discard} respondents)</h3>"
    if n_discard == 0:
        html += "    <p style='padding:12px;color:#888;'>No discards in this run.</p>\n"
    else:
        html += f"""    <table>
      <thead><tr><th>Respondent ID</th><th>Score</th><th>ML</th><th>Supplier</th><th>Agent Justification</th><th>Key Signals</th></tr></thead>
      <tbody>
"""
        for r in discard_rows:
            score_class = "score-negative" if r["score"] < 0 else "score-positive"
            html += f'      <tr><td>{r["id"]}</td><td class="{score_class}">{r["score"]}</td><td>{r["ml"]}</td><td>{r["supplier"][:20]}</td><td class="justification">{r["justification"]}</td><td style="font-size:11px;">{r["signals"][:60]}</td></tr>\n'

        html += "      </tbody>\n    </table>\n"

    html += f"""  </div>

  <div class="footer">
    Generated by Autosurvey Survey Quality Pipeline with Agent Review — {dataset_name}
  </div>
</div>
</body>
</html>"""

    with open(output_path, "w") as f:
        f.write(html)


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 integrate_agent_judgments.py <xlsx_path> <output_dir>")
        return

    integrate_agent_judgments(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
