#!/usr/bin/env python3
"""Survey Quality ML Pipeline — Trains on annotated datasets, predicts on unseen.

This is the production pipeline that:
1. Extracts features from annotated Excel files (with status column)
2. Trains a model using leave-one-dataset-out cross-validation
3. Saves the trained model for use on new unseen datasets
4. Includes semantic structure parsing (Datamap, answer chains, field roles)
5. Includes all agent analysis rules (discard rules, exemplar patterns)

Usage:
    python3 survey_quality_ml.py train    # Train and evaluate with LODO CV
    python3 survey_quality_ml.py predict <xlsx_path>  # Predict on unseen dataset
    python3 survey_quality_ml.py evaluate  # LODO CV only
"""
from __future__ import annotations

import csv
import json
import pickle
import re
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.isotonic import IsotonicRegression
from sklearn.metrics import (
    accuracy_score, classification_report, confusion_matrix,
    f1_score, precision_score, recall_score, roc_auc_score,
)

warnings.filterwarnings("ignore")

# === Paths ===
ANNOTATED_DIR = Path("/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer")
SIGNAL_MAP = Path("/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/status-ground-truth-calibration/core_loop_2026-06-24/status_respondent_signal_map.csv")
MODEL_DIR = Path("/Users/jeremyalston/Perfect/autosurvey/skills/cleaning-survey-quality/models")
MODEL_DIR.mkdir(parents=True, exist_ok=True)

# Map annotated filenames to signal map dataset names
DATASET_MAP = {
    "260111_Delta Water Filtration.xlsx": "260111_Delta Water Filtration.xlsx",
    "260300_ECHO.xlsx": "260300_ECHO.xlsx",
    "260501_ODL.xlsx": "260501_ODL.xlsx",
    "260206_OC BH.xlsx": "260206_OC BH.xlsx",
    "260401_ OC CAN.xlsx": "260401_ OC CAN.xlsx",
    "260200_SBD.xlsx": "260200_SBD.xlsx",
    "251101_THD CX.xlsx": "251101_THD CX.xlsx",
    "260404_ADDO.xlsx": "260404_ADDO.xlsx",
    "260403_Masterlock Conjoint.xlsx": "260403_Masterlock Conjoint.xlsx",
    "251205_TFG Contractor Index Q1.xlsx": "251205_TFG Contractor Index Q1.xlsx",
    "260306_TFG Contractor Index Q2.xlsx": "260306_TFG Contractor Index Q2.xlsx",
}

ALL_SIGNALS = set()
T1_SIGNALS = {"termflags_nonzero", "long_low_specificity_text", "ai_or_overpolished_text_marker", "generic_placeholder_open_end"}
T2_SIGNALS = {"rd_searchr3_canada", "rd_searchr1_22", "rd_searchr1_23", "rd_searchr1_20", "qtime_under_dataset_p10",
              "rd_searchr1_22.0", "rd_searchr1_23.0", "rd_searchr1_20.0"}


def clean(v) -> str:
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def norm(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v


def load_signal_map():
    """Load client signal map for additional features."""
    by_ds = defaultdict(dict)
    if not SIGNAL_MAP.exists():
        return by_ds
    with open(SIGNAL_MAP) as f:
        for row in csv.DictReader(f):
            sigs = [s.strip() for s in row["signals"].split(";") if s.strip()]
            for s in sigs: ALL_SIGNALS.add(s)
            by_ds[row["dataset"]][row["respondent_key"]] = {
                "signals": sigs, "decision": row["tfg_decision"],
                "signal_count": int(row["signal_count"]),
            }
    return by_ds


def parse_datamap(wb):
    """Parse the Datamap sheet to get question text and value labels for each field."""
    dm = {}
    if "Datamap" not in wb.sheetnames:
        return dm
    ws = wb["Datamap"]
    current_field = None
    current_qtext = None
    labels = {}

    for row in ws.iter_rows(values_only=True):
        a, b, c = (row + (None, None, None))[:3]
        a_s = clean(a)
        if a_s.startswith("[") and "]" in a_s:
            field_name = a_s.split("]", 1)[0][1:]
            current_field = field_name
            current_qtext = clean(b)
            labels = {}
            dm[field_name] = {"question_text": current_qtext, "labels": labels}
        elif current_field and a_s and b is not None:
            labels[str(norm(a))] = clean(b)

    return dm


def classify_field_role(field_name, question_text, datamap):
    """Classify a field's role: screener, demographic, matrix, open_end, technical, etc."""
    fn = field_name.lower()
    qt = (question_text or "").lower()

    if fn in ("uuid", "record", "rid"): return "id"
    if fn in ("status",): return "label"
    if fn in ("markers",): return "quota"
    if fn in ("qtime",): return "timing"
    if fn in ("supname", "supplier"): return "supplier"
    if fn in ("ipaddress",): return "ip"
    if fn in ("useragent",): return "user_agent"
    if fn in ("date", "start_date"): return "timestamp"
    if fn in ("vos", "vbrowser", "vmobiledevice", "vmobileos"): return "technical"
    if fn.startswith("langassess"): return "nlp_metadata"
    if fn.startswith("rd_"): return "review_metadata"
    if fn.startswith("termflags") or "flag" in fn: return "quality_flag"
    if fn.endswith("oe") or fn == "outro" or "qcoe" in fn: return "open_end"
    if re.match(r"q\d+r\d+$", fn): return "matrix_cell"
    if fn in ("age", "qager1", "qgender", "qstate", "region", "q13", "q12",
              "qhomtype", "q2", "q1", "qindustry", "qnumemployees", "q9",
              "qethnicr1", "classify", "channeltracking"): return "demographic"
    if "pasted" in fn: return "paste_flag"
    if fn.startswith("noanswer"): return "no_answer"
    if fn.startswith("conditions"): return "routing"
    if fn.startswith("own") or fn.startswith("possible"): return "product_attr"
    return "coded_question"


def extract_features_from_excel(filepath, signal_map=None, datamap=None, include_label=True):
    """Extract ALL features from an Excel file (annotated or unannotated).

    This is the core semantic structure parser + feature extractor.
    """
    sm_name = filepath.name
    sm = (signal_map or {}).get(sm_name, {})

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Parse datamap if not provided
    if datamap is None:
        datamap = parse_datamap(wb)

    # Classify all fields by role
    field_roles = {}
    for h in headers:
        if h:
            qt = datamap.get(str(h), {}).get("question_text", "")
            field_roles[str(h)] = classify_field_role(str(h), qt, datamap)

    # Group columns by role
    oe_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "open_end"]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "nlp_metadata"]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "matrix_cell"]
    rd_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "review_metadata"]
    flag_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "quality_flag"]
    paste_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "paste_flag"]
    demo_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "demographic"]
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and field_roles.get(str(h)) == "coded_question"]
    status_idx = hidx.get("status")
    markers_idx = hidx.get("markers")

    # Collect all rows
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid_idx = hidx.get("uuid") or hidx.get("record")
        if rid_idx is None or rid_idx >= len(row): continue
        rid = clean(row[rid_idx])
        if not rid: continue
        # For annotated files, check status; for unannotated, include all
        if include_label and sm_name in DATASET_MAP:
            if rid not in sm and status_idx is not None:
                # Use status column directly
                status = row[status_idx] if status_idx < len(row) else None
                if status is not None:
                    sm[rid] = {"signals": [], "decision": "rejected" if int(status) == 5 else "accepted", "signal_count": 0}
        rows_data.append((rid, row))

    if not rows_data:
        return pd.DataFrame(), datamap

    # Cross-respondent duplicate detection
    all_oe, all_ua, all_ip, all_sd = [], [], [], []
    for rid, row in rows_data:
        oe = " | ".join(clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i]))
        all_oe.append(oe)
        ua_idx = hidx.get("userAgent")
        all_ua.append(clean(row[ua_idx]) if ua_idx and ua_idx < len(row) else "")
        ip_idx = hidx.get("ipAddress")
        all_ip.append(clean(row[ip_idx]) if ip_idx and ip_idx < len(row) else "")
        sd_idx = hidx.get("start_date")
        all_sd.append(clean(row[sd_idx]) if sd_idx and sd_idx < len(row) else "")

    oe_ctr = Counter(t.strip().lower() for t in all_oe if t.strip())
    ua_ctr = Counter(ua for ua in all_ua if ua)
    ip_ctr = Counter(ip for ip in all_ip if ip)
    sd_ctr = Counter(sd for sd in all_sd if sd)

    # Build features
    features = []
    for idx, (rid, row) in enumerate(rows_data):
        feat = {"respondent_id": rid}

        # === Client signals (if available from signal map) ===
        e = sm.get(rid, {"signals": [], "signal_count": 0})
        sigs = set(e.get("signals", []))
        feat["signal_count"] = e.get("signal_count", 0)
        feat["t1_count"] = len(sigs & T1_SIGNALS)
        feat["t2_count"] = len(sigs & T2_SIGNALS)
        feat["t3_count"] = len(sigs - T1_SIGNALS - T2_SIGNALS)
        for s in ALL_SIGNALS:
            feat[f"sig_{s}"] = 1 if s in sigs else 0

        # === Timing ===
        qt_idx = hidx.get("qtime")
        qt = 0
        if qt_idx and qt_idx < len(row):
            try: qt = float(row[qt_idx]) if row[qt_idx] else 0
            except: qt = 0
        feat["qtime_seconds"] = qt
        feat["qtime_log"] = np.log1p(qt) if qt > 0 else 0

        # === LangAssess NLP features ===
        for i, h in lang_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"lang_{h}"] = float(v) if v else 0
            except: feat[f"lang_{h}"] = 0

        # === Open-end text features ===
        oe_texts = [clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i])]
        oe_lens = [len(t) for t in oe_texts]
        oe_words = [len(t.split()) for t in oe_texts]
        feat["oe_count"] = len(oe_texts)
        feat["oe_total_chars"] = sum(oe_lens)
        feat["oe_max_chars"] = max(oe_lens) if oe_lens else 0
        feat["oe_mean_chars"] = np.mean(oe_lens) if oe_lens else 0
        feat["oe_total_words"] = sum(oe_words)
        feat["oe_max_words"] = max(oe_words) if oe_words else 0
        all_oe_text = " ".join(oe_texts).lower()
        words = all_oe_text.split()
        feat["oe_lex_div"] = len(set(words)) / len(words) if words else 0
        feat["oe_has_none"] = 1 if any(w in all_oe_text for w in ["none","n/a","na","nothing","no opinion","no idea"]) else 0
        feat["oe_generic"] = 1 if any(w in all_oe_text for w in ["good","fine","ok","okay","nice","great"]) and oe_words and max(oe_words) <= 3 else 0
        feat["oe_all_caps"] = 1 if any(t.isupper() and len(t) > 5 for t in oe_texts) else 0
        feat["oe_very_short"] = 1 if oe_lens and max(oe_lens) < 10 else 0
        feat["oe_short"] = 1 if oe_lens and max(oe_lens) < 30 else 0

        # === Supplier ===
        sup_idx = hidx.get("SUPNAME")
        sup = clean(row[sup_idx]) if sup_idx and sup_idx < len(row) else ""
        feat["supplier_name"] = sup
        feat["supplier_missing"] = 1 if not sup or sup == "MISSING" else 0
        feat["supplier_is_none"] = 1 if sup == "None" or sup == "" else 0

        # === Matrix/grid straightlining ===
        mvals = [norm(row[i]) for i, _ in matrix_cols if i < len(row) and row[i] is not None and row[i] != ""]
        if mvals:
            ur = len(set(mvals)) / len(mvals)
            feat["matrix_unique_ratio"] = ur
            feat["matrix_straightline"] = 1 if ur <= 0.2 and len(mvals) >= 5 else 0
            feat["matrix_near_straightline"] = 1 if ur <= 0.4 and len(mvals) >= 5 else 0
            feat["matrix_count"] = len(mvals)
            feat["matrix_unique_count"] = len(set(mvals))
            vc = Counter(mvals)
            feat["matrix_most_common_freq"] = vc.most_common(1)[0][1] / len(mvals)
        else:
            feat["matrix_unique_ratio"] = 1.0
            feat["matrix_straightline"] = 0
            feat["matrix_near_straightline"] = 0
            feat["matrix_count"] = 0
            feat["matrix_unique_count"] = 0
            feat["matrix_most_common_freq"] = 0

        # === Coded answer diversity ===
        cvals = [str(norm(row[i])) for i, _ in coded_cols if i < len(row) and row[i] is not None and row[i] != ""]
        feat["coded_count"] = len(cvals)
        feat["coded_unique_ratio"] = len(set(cvals)) / len(cvals) if cvals else 1.0
        dk = sum(1 for v in cvals if any(x in v.lower() for x in ["don't know","dk","not sure","no answer"]))
        feat["coded_dk_count"] = dk
        feat["coded_dk_ratio"] = dk / len(cvals) if cvals else 0

        # === RD_Search features ===
        for i, h in rd_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"rd_{h}"] = float(v) if v else 0
            except: feat[f"rd_{h}"] = 0

        # === Flag columns ===
        for i, h in flag_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"flag_{h}"] = float(v) if v else 0
            except: feat[f"flag_{h}"] = 0

        # === Paste flags ===
        for i, h in paste_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"paste_{h}"] = float(v) if v else 0
            except: feat[f"paste_{h}"] = 0

        # === Cross-respondent duplicates ===
        oe_t = all_oe[idx]
        feat["oe_dup_count"] = oe_ctr.get(oe_t.strip().lower(), 0) if oe_t.strip() else 0
        feat["oe_is_dup"] = 1 if feat["oe_dup_count"] > 1 else 0
        feat["ua_dup_count"] = ua_ctr.get(all_ua[idx], 0)
        feat["ua_is_dup"] = 1 if feat["ua_dup_count"] > 1 else 0
        feat["ip_dup_count"] = ip_ctr.get(all_ip[idx], 0)
        feat["ip_is_dup"] = 1 if feat["ip_dup_count"] > 1 else 0
        feat["sd_dup_count"] = sd_ctr.get(all_sd[idx], 0)
        feat["sd_is_dup"] = 1 if feat["sd_dup_count"] > 1 else 0

        # === Raw open-end text (for TF-IDF in train scripts) ===
        feat["_oe_raw_text"] = all_oe_text[:2000]  # cap for memory

        # === Markers (quota info) — EXCLUDED to prevent label leakage ===
        # The markers column contains "bad:" for rejected respondents, which is
        # essentially a copy of the label. Do NOT use markers as features.

        # === Demographics (categorical) ===
        for i, h in demo_cols:
            feat[f"demo_{h}"] = clean(row[i]) if i < len(row) else ""

        # === Technical ===
        for h in ["vos", "vbrowser", "vmobiledevice", "vmobileos"]:
            i = hidx.get(h)
            feat[f"tech_{h}"] = clean(row[i]) if i and i < len(row) else ""

        # === Label (from status column or signal map) ===
        if include_label:
            if status_idx is not None and status_idx < len(row):
                status = row[status_idx]
                feat["label"] = 1 if status is not None and int(status) == 5 else 0
            elif rid in sm:
                feat["label"] = 1 if sm[rid].get("decision") == "rejected" else 0
            else:
                feat["label"] = -1  # Unknown

        feat["dataset"] = sm_name
        features.append(feat)

    df = pd.DataFrame(features)

    # Add dataset-relative features (z-scores within dataset)
    for col in ["qtime_seconds", "signal_count"]:
        df[f"{col}_zscore"] = 0.0
        for ds in df["dataset"].unique():
            m = df["dataset"] == ds
            vals = df.loc[m, col]
            std = vals.std()
            df.loc[m, f"{col}_zscore"] = (vals - vals.mean()) / (std if std > 0 else 1)

    return df, datamap


def add_supplier_risk(train_df, test_df=None):
    """Add supplier reject rate from training data."""
    global_rate = train_df["label"].mean()
    if "label" not in train_df.columns or train_df["label"].max() < 0:
        return train_df, test_df

    sr = train_df.groupby("supplier_name")["label"].agg(["mean","count"]).reset_index()
    sr.columns = ["supplier_name","rate","count"]
    sr["supplier_reject_rate"] = (sr["count"]*sr["rate"] + 20*global_rate) / (sr["count"]+20)
    sr = sr[["supplier_name","supplier_reject_rate"]]

    train_df = train_df.merge(sr, on="supplier_name", how="left")
    train_df["supplier_reject_rate"] = train_df["supplier_reject_rate"].fillna(global_rate)

    if test_df is not None:
        test_df = test_df.merge(sr, on="supplier_name", how="left")
        test_df["supplier_reject_rate"] = test_df["supplier_reject_rate"].fillna(global_rate)

    # Interaction features
    for df in [train_df, test_df]:
        if df is not None:
            df["supplier_x_signals"] = df["supplier_reject_rate"] * df["signal_count"]
            df["supplier_x_t1"] = df["supplier_reject_rate"] * df["t1_count"]
            df["supplier_x_t2"] = df["supplier_reject_rate"] * df["t2_count"]
            df["signals_x_matrix"] = df["signal_count"] * (1 - df["matrix_unique_ratio"])

    return train_df, test_df


def prepare_features(df):
    """Prepare features for ML — encode categoricals, drop non-feature columns."""
    non_feature = {"respondent_id", "label", "dataset", "supplier_name", "_oe_raw_text"}
    feat_cols = [c for c in df.columns if c not in non_feature and not c.startswith("_")]
    X = df[feat_cols].copy()
    y = df["label"].copy() if "label" in df.columns else None
    for col in X.select_dtypes(include=["object"]).columns:
        X[col] = pd.Categorical(X[col]).codes
    return X.fillna(0), y


def train_and_evaluate():
    """Train model with LODO CV and save the final model."""
    print("Loading signal map...")
    signal_map = load_signal_map()
    print(f"  {sum(len(v) for v in signal_map.values())} annotations, {len(ALL_SIGNALS)} signals")

    print("\nExtracting features from annotated datasets...")
    all_dfs = []
    for fname in sorted(DATASET_MAP.keys()):
        fp = ANNOTATED_DIR / fname
        if not fp.exists():
            print(f"  SKIP: {fname}")
            continue
        df, dm = extract_features_from_excel(fp, signal_map)
        if df is not None and len(df) > 0:
            n_rej = (df["label"] == 1).sum()
            print(f"  {fname}: {len(df)} respondents, {len(df.columns)} features, {n_rej} rejected ({n_rej/len(df):.1%})")
            all_dfs.append(df)

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\nCombined: {len(combined)} respondents, {len(combined.columns)} features")
    print(f"  Rejected: {(combined['label']==1).sum()} ({(combined['label']==1).mean():.1%})")

    # LODO CV
    datasets = combined["dataset"].unique()
    results = []

    print(f"\n{'='*120}")
    print(f"LEAVE-ONE-DATASET-OUT CROSS-VALIDATION")
    print(f"{'='*120}")

    for test_ds in datasets:
        train_df = combined[combined["dataset"] != test_ds].copy()
        test_df = combined[combined["dataset"] == test_ds].copy()

        # Add supplier risk from training
        train_df, test_df = add_supplier_risk(train_df, test_df)

        X_train, y_train = prepare_features(train_df)
        X_test, y_test = prepare_features(test_df)

        # Align columns
        for c in X_train.columns:
            if c not in X_test.columns: X_test[c] = 0
        for c in X_test.columns:
            if c not in X_train.columns: X_train[c] = 0
        X_test = X_test[X_train.columns]

        # Class weights
        n_pos = max((y_train == 1).sum(), 1)
        n_neg = max((y_train == 0).sum(), 1)
        w = np.where(y_train == 1, len(y_train)/(2*n_pos), len(y_train)/(2*n_neg))

        # Train
        model = GradientBoostingClassifier(
            n_estimators=300, max_depth=4, learning_rate=0.05,
            subsample=0.8, random_state=42, min_samples_leaf=20
        )
        model.fit(X_train, y_train, sample_weight=w)

        # Calibrate
        y_train_proba = model.predict_proba(X_train)[:, 1]
        y_test_proba = model.predict_proba(X_test)[:, 1]
        iso = IsotonicRegression(out_of_bounds='clip')
        iso.fit(y_train_proba, y_train)
        y_test_cal = iso.transform(y_test_proba)

        # Use training-optimal accuracy threshold
        y_train_cal = iso.transform(y_train_proba)
        best_acc, best_t = 0, 0.5
        for t in np.linspace(0.01, 0.99, 200):
            pred = (y_train_cal >= t).astype(int)
            acc = accuracy_score(y_train, pred)
            if acc > best_acc:
                best_acc = acc
                best_t = t

        y_pred = (y_test_cal >= best_t).astype(int)

        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, zero_division=0)
        rec = recall_score(y_test, y_pred, zero_division=0)
        f1 = f1_score(y_test, y_pred, zero_division=0)
        auc = roc_auc_score(y_test, y_test_proba) if len(y_test.unique()) > 1 else 0
        cm = confusion_matrix(y_test, y_pred, labels=[0,1])
        tn, fp, fn, tp = cm.ravel()

        results.append({"dataset": test_ds, "n": len(y_test), "threshold": best_t,
            "accuracy": acc, "precision": prec, "recall": rec, "f1": f1, "auc": auc,
            "tp": tp, "fp": fp, "tn": tn, "fn": fn,
            "client_reject_rate": float(y_test.mean()),
            "agent_discard_rate": float(y_pred.mean())})

        print(f"\n{test_ds}")
        print(f"  N={len(y_test)}, Client: {y_test.mean():.1%}, Agent: {y_pred.mean():.1%}, Thresh: {best_t:.3f}")
        print(f"  Acc={acc:.1%}  Prec={prec:.1%}  Rec={rec:.1%}  F1={f1:.1%}  AUC={auc:.3f}")
        print(f"  TP={tp}  FP={fp}  TN={tn}  FN={fn}")

    # Aggregate
    print(f"\n{'='*120}")
    print("AGGREGATE")
    print(f"{'='*120}")
    tp = sum(r["tp"] for r in results)
    fp = sum(r["fp"] for r in results)
    tn = sum(r["tn"] for r in results)
    fn = sum(r["fn"] for r in results)
    n = sum(r["n"] for r in results)
    acc = (tp+tn)/n
    prec = tp/(tp+fp) if tp+fp > 0 else 0
    rec = tp/(tp+fn) if tp+fn > 0 else 0
    f1 = 2*prec*rec/(prec+rec) if prec+rec > 0 else 0

    print(f"  N={n}, TP={tp}, FP={fp}, TN={tn}, FN={fn}")
    print(f"  Accuracy:  {acc:.1%}")
    print(f"  Precision: {prec:.1%}")
    print(f"  Recall:    {rec:.1%}")
    print(f"  F1:        {f1:.1%}")
    print(f"  Discard:   {(tp+fp)/n:.1%} vs Client: {(tp+fn)/n:.1%}")
    print(f"\n  Baseline (keep everyone): {1-(tp+fn)/n:.1%}")

    # Save results
    pd.DataFrame(results).to_csv(MODEL_DIR / "lodo_cv_results.csv", index=False)

    # Train final model on ALL data
    print("\nTraining final model on all data...")
    combined_with_risk, _ = add_supplier_risk(combined)
    X_all, y_all = prepare_features(combined_with_risk)
    n_pos = max((y_all == 1).sum(), 1)
    n_neg = max((y_all == 0).sum(), 1)
    w = np.where(y_all == 1, len(y_all)/(2*n_pos), len(y_all)/(2*n_neg))

    final_model = GradientBoostingClassifier(
        n_estimators=300, max_depth=4, learning_rate=0.05,
        subsample=0.8, random_state=42, min_samples_leaf=20
    )
    final_model.fit(X_all, y_all, sample_weight=w)

    # Calibrate
    y_all_proba = final_model.predict_proba(X_all)[:, 1]
    iso = IsotonicRegression(out_of_bounds='clip')
    iso.fit(y_all_proba, y_all)

    # Find accuracy-optimal threshold
    y_all_cal = iso.transform(y_all_proba)
    best_acc, best_t = 0, 0.5
    for t in np.linspace(0.01, 0.99, 200):
        pred = (y_all_cal >= t).astype(int)
        acc = accuracy_score(y_all, pred)
        if acc > best_acc:
            best_acc = acc
            best_t = t

    # Save model, calibrator, threshold, and feature columns
    model_data = {
        "model": final_model,
        "calibrator": iso,
        "threshold": best_t,
        "feature_columns": list(X_all.columns),
        "supplier_rates": combined_with_risk.groupby("supplier_name")["supplier_reject_rate"].first().to_dict(),
        "global_reject_rate": float(y_all.mean()),
    }
    with open(MODEL_DIR / "survey_quality_model.pkl", "wb") as f:
        pickle.dump(model_data, f)
    print(f"Model saved to {MODEL_DIR / 'survey_quality_model.pkl'}")
    print(f"  Threshold: {best_t:.3f}")
    print(f"  Features: {len(X_all.columns)}")

    return results


def predict_unseen(filepath):
    """Predict quality scores for an unseen dataset."""
    print(f"Loading model...")
    with open(MODEL_DIR / "survey_quality_model.pkl", "rb") as f:
        model_data = pickle.load(f)

    model = model_data["model"]
    iso = model_data["calibrator"]
    threshold = model_data["threshold"]
    train_features = model_data["feature_columns"]
    supplier_rates = model_data["supplier_rates"]
    global_rate = model_data["global_reject_rate"]

    print(f"Extracting features from {filepath.name}...")
    df, datamap = extract_features_from_excel(filepath, include_label=False)

    if df.empty:
        print("ERROR: No data extracted")
        return

    print(f"  {len(df)} respondents, {len(df.columns)} features")

    # Add supplier risk from saved rates
    df["supplier_reject_rate"] = df["supplier_name"].map(supplier_rates).fillna(global_rate)
    df["supplier_x_signals"] = df["supplier_reject_rate"] * df["signal_count"]
    df["supplier_x_t1"] = df["supplier_reject_rate"] * df["t1_count"]
    df["supplier_x_t2"] = df["supplier_reject_rate"] * df["t2_count"]
    df["signals_x_matrix"] = df["signal_count"] * (1 - df["matrix_unique_ratio"])

    # Prepare features
    X, _ = prepare_features(df)

    # Align with training features
    for c in train_features:
        if c not in X.columns: X[c] = 0
    X = X[train_features]

    # Predict
    y_proba = model.predict_proba(X)[:, 1]
    y_cal = iso.transform(y_proba)
    y_pred = (y_cal >= threshold).astype(int)

    # Output
    results = pd.DataFrame({
        "respondent_id": df["respondent_id"],
        "quality_score": y_cal,
        "prediction": ["DISCARD" if p == 1 else "KEEP" for p in y_pred],
        "confidence": np.where(y_pred == 1, y_cal, 1 - y_cal),
    })

    # Add top features for each prediction
    if hasattr(model, "feature_importances_"):
        imp = dict(zip(train_features, model.feature_importances_))
        top_feats = sorted(imp.items(), key=lambda x: -x[1])[:10]
        print(f"\nTop features: {', '.join(f'{n}={v:.3f}' for n,v in top_feats)}")

    n_discard = y_pred.sum()
    print(f"\nResults: {len(results)} respondents")
    print(f"  DISCARD: {n_discard} ({n_discard/len(results):.1%})")
    print(f"  KEEP: {len(results)-n_discard} ({(len(results)-n_discard)/len(results):.1%})")
    print(f"  Mean quality score: {y_cal.mean():.3f}")
    print(f"  Threshold: {threshold:.3f}")

    # Save
    output_path = filepath.parent / f"{filepath.stem}_quality_predictions.csv"
    results.to_csv(output_path, index=False)
    print(f"\nPredictions saved to {output_path}")

    # Also save NDJSON for agent consumption
    ndjson_path = filepath.parent / f"{filepath.stem}_quality_predictions.ndjson"
    with open(ndjson_path, "w") as f:
        for _, row in results.iterrows():
            f.write(json.dumps(row.to_dict()) + "\n")
    print(f"NDJSON saved to {ndjson_path}")

    return results


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 survey_quality_ml.py [train|predict <path>|evaluate]")
        return

    cmd = sys.argv[1]
    if cmd == "train":
        train_and_evaluate()
    elif cmd == "predict":
        if len(sys.argv) < 3:
            print("Usage: python3 survey_quality_ml.py predict <xlsx_path>")
            return
        predict_unseen(Path(sys.argv[2]))
    elif cmd == "evaluate":
        train_and_evaluate()
    else:
        print(f"Unknown command: {cmd}")


if __name__ == "__main__":
    main()
