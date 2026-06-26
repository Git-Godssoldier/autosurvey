#!/usr/bin/env python3
"""Survey Quality Pipeline — ML triage + agent scoring + final output.

New architecture:
1. ML model provides triage risk probability (0-1) for each respondent
2. Agent reads each respondent's answer chain and assigns a -1 to +1 score:
   -1 = clear discard, -0.5 = discard lean, 0 = uncertain, +0.5 = keep lean, +1 = clear keep
3. All respondents with score < 0 get a second agent reassessment
4. Final discards selected from the reassessed pool
5. Output: annotated Excel + single-file HTML dashboard

Usage:
    python3 survey_pipeline.py <xlsx_path> [--output-dir DIR]
"""
from __future__ import annotations

import json
import pickle
import re
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.isotonic import IsotonicRegression

warnings.filterwarnings("ignore")

MODEL_DIR = Path(__file__).parent.parent / "models"

# Signal tiers from agent training
T1_SIGNALS = {"termflags_nonzero", "long_low_specificity_text", "ai_or_overpolished_text_marker", "generic_placeholder_open_end"}
T2_SIGNALS = {"rd_searchr3_canada", "rd_searchr1_22", "rd_searchr1_23", "rd_searchr1_20", "qtime_under_dataset_p10",
              "rd_searchr1_22.0", "rd_searchr1_23.0", "rd_searchr1_20.0"}

# Powerful signals from feature importance analysis (SKILL.md)
POWERFUL_SIGNALS = {
    "lang_readlevel": {"name": "LangAssess ReadLevel", "tier": 1, "description": "Readability of open-end text. Low readability correlates with low quality."},
    "supplier_x_signals": {"name": "Supplier × Signals interaction", "tier": 1, "description": "High-reject supplier + quality signals = very likely bad."},
    "supplier_reject_rate": {"name": "Supplier reject rate", "tier": 1, "description": "Historical reject rate for this supplier."},
    "answer_entropy": {"name": "Answer entropy", "tier": 2, "description": "Low entropy = repetitive answers = potential straightlining."},
    "oe_total_chars": {"name": "Open-end total chars", "tier": 2, "description": "Very short open-ends are a strong discard signal."},
    "matrix_straightline": {"name": "Matrix straightlining", "tier": 2, "description": "Same answer across all matrix rows."},
    "matrix_most_common_freq": {"name": "Matrix most common freq", "tier": 2, "description": "High frequency of one answer = straightlining."},
    "dup_shared": {"name": "Duplicate fingerprint sharing", "tier": 3, "description": "Sharing IP/UA/oe with many others = potential fraud."},
    "time_per_q": {"name": "Time per question", "tier": 3, "description": "Very low time per question = speeding."},
}


def clean(v): return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""
def norm(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v


def parse_datamap(wb):
    """Parse the Datamap sheet to get question text and value labels.

    Datamap format:
        [fieldName]: Question text   | (col B empty) | (col C empty)
        Values: 1-12                  |               |
        (empty)                       | 1             | Label for value 1
        (empty)                       | 2             | Label for value 2
    """
    dm = {}
    if "Datamap" not in wb.sheetnames:
        return dm
    ws = wb["Datamap"]
    current_field = None
    for row in ws.iter_rows(values_only=True):
        a, b, c = (row + (None, None, None))[:3]
        a_s = clean(a)
        b_s = clean(b)
        c_s = clean(c)

        if a_s.startswith("[") and "]" in a_s:
            # New field definition: [fieldName]: Question text
            bracket_end = a_s.index("]")
            field_name = a_s[1:bracket_end]
            # Question text is everything after "]: " in column A
            qtext = a_s[bracket_end + 1:].lstrip(": ").strip()
            # If qtext is empty, try column B
            if not qtext:
                qtext = b_s
            dm[field_name] = {"question_text": qtext, "labels": {}}
            current_field = field_name
        elif current_field:
            # Value label rows — can have empty column A
            # Format 1: (empty) | value | label
            # Format 2: value | label | (empty)
            # Format 3: value | label | label2
            if not a_s.startswith("Values:") and not a_s.startswith("Open") and not a_s.startswith("RD_"):
                if b_s and c_s:
                    dm[current_field]["labels"][str(norm(b))] = c_s
                elif a_s and b_s and not c_s:
                    dm[current_field]["labels"][str(norm(a_s))] = b_s
    return dm


def classify_field(field_name):
    fn = field_name.lower()
    if fn in ("uuid", "record", "rid"): return "id"
    if fn == "status": return "label"
    if fn == "markers": return "quota"
    if fn == "qtime": return "timing"
    if fn in ("supname", "supplier"): return "supplier"
    if fn == "ipaddress": return "ip"
    if fn == "useragent": return "user_agent"
    if fn in ("date", "start_date"): return "timestamp"
    if fn in ("vos", "vbrowser", "vmobiledevice", "vmobileos"): return "technical"
    if fn.startswith("langassess"): return "nlp_metadata"
    if fn.startswith("rd_"): return "review_metadata"
    if fn.startswith("termflags") or "flag" in fn: return "quality_flag"
    if fn.endswith("oe") or fn == "outro" or "qcoe" in fn: return "open_end"
    if re.match(r"q\d+r\d+$", fn): return "matrix_cell"
    if fn in ("age", "qager1", "qgender", "qstate", "region", "q13", "q12",
              "qhomtype", "q2", "q1", "qindustry", "qnumemployees", "q9",
              "qethnicr1", "classify", "channeltracking"): return "demographic"
    if "pasted" in fn: return "paste_flag"
    if fn.startswith("noanswer"): return "no_answer"
    if fn.startswith("conditions"): return "routing"
    if fn.startswith("own") or fn.startswith("possible"): return "product_attr"
    return "coded_question"


def extract_features_and_chain(filepath, signal_map=None):
    """Extract ML features AND build the answer chain for agent review.
    
    Returns:
        df: DataFrame with ML features
        datamap: parsed datamap
        roles: field role classification
        answer_chains: list of answer chain dicts per respondent
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    datamap = parse_datamap(wb)
    roles = {str(h): classify_field(str(h)) for h in headers if h}

    oe_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "open_end"]
    lang_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "nlp_metadata"]
    matrix_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "matrix_cell"]
    rd_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "review_metadata"]
    flag_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "quality_flag"]
    paste_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "paste_flag"]
    demo_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "demographic"]
    coded_cols = [(i, h) for i, h in enumerate(headers) if h and roles.get(str(h)) == "coded_question"]

    # Collect rows
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid_idx = hidx.get("uuid") or hidx.get("record")
        if rid_idx is None or rid_idx >= len(row): continue
        rid = clean(row[rid_idx])
        if not rid: continue
        rows_data.append((rid, row))

    if not rows_data:
        return pd.DataFrame(), datamap, roles, []

    # Cross-respondent duplicates
    all_oe, all_ua, all_ip = [], [], []
    for rid, row in rows_data:
        oe = " | ".join(clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i]))
        all_oe.append(oe)
        ua_idx = hidx.get("userAgent")
        all_ua.append(clean(row[ua_idx]) if ua_idx and ua_idx < len(row) else "")
        ip_idx = hidx.get("ipAddress")
        all_ip.append(clean(row[ip_idx]) if ip_idx and ip_idx < len(row) else "")

    oe_ctr = Counter(t.strip().lower() for t in all_oe if t.strip())
    ua_ctr = Counter(ua for ua in all_ua if ua)
    ip_ctr = Counter(ip for ip in all_ip if ip)

    # Compute population stats for timing percentiles
    all_qtimes = []
    for rid, row in rows_data:
        qt_idx = hidx.get("qtime")
        qt = 0
        if qt_idx and qt_idx < len(row):
            try: qt = float(row[qt_idx]) if row[qt_idx] else 0
            except: qt = 0
        all_qtimes.append(qt)
    qtime_arr = np.array(all_qtimes)
    qtime_p10 = np.percentile(qtime_arr[qtime_arr > 0], 10) if (qtime_arr > 0).any() else 0
    qtime_p25 = np.percentile(qtime_arr[qtime_arr > 0], 25) if (qtime_arr > 0).any() else 0
    qtime_median = np.median(qtime_arr[qtime_arr > 0]) if (qtime_arr > 0).any() else 0

    # Supplier reject rates from population
    sup_rates = {}
    sup_counts = defaultdict(lambda: [0, 0])  # [total, signals]
    for idx, (rid, row) in enumerate(rows_data):
        sup_idx = hidx.get("SUPNAME")
        sup = clean(row[sup_idx]) if sup_idx and sup_idx < len(row) else ""
        sm = signal_map or {}
        e = sm.get(rid, {"signals": [], "signal_count": 0})
        sig_count = e.get("signal_count", 0)
        sup_counts[sup][0] += 1
        sup_counts[sup][1] += sig_count
    for sup, (total, sigs) in sup_counts.items():
        sup_rates[sup] = sigs / total if total > 0 else 0
    global_rate = sum(s for _, s in sup_counts.values()) / max(sum(t for t, _ in sup_counts.values()), 1)

    features = []
    answer_chains = []
    for idx, (rid, row) in enumerate(rows_data):
        feat = {"respondent_id": rid}

        # Client signals
        sm = signal_map or {}
        e = sm.get(rid, {"signals": [], "signal_count": 0})
        sigs = set(e.get("signals", []))
        feat["signal_count"] = e.get("signal_count", 0)
        feat["t1_count"] = len(sigs & T1_SIGNALS)
        feat["t2_count"] = len(sigs & T2_SIGNALS)
        feat["t3_count"] = len(sigs - T1_SIGNALS - T2_SIGNALS)

        # Timing
        qt_idx = hidx.get("qtime")
        qt = 0
        if qt_idx and qt_idx < len(row):
            try: qt = float(row[qt_idx]) if row[qt_idx] else 0
            except: qt = 0
        feat["qtime_seconds"] = qt
        feat["qtime_log"] = np.log1p(qt) if qt > 0 else 0
        feat["qtime_minutes"] = qt / 60.0

        # Timing percentile
        if qt > 0:
            if qt <= qtime_p10:
                feat["qtime_percentile"] = "bottom_10"
            elif qt <= qtime_p25:
                feat["qtime_percentile"] = "bottom_25"
            elif qt <= qtime_median:
                feat["qtime_percentile"] = "below_median"
            else:
                feat["qtime_percentile"] = "above_median"
        else:
            feat["qtime_percentile"] = "unknown"

        # LangAssess
        for i, h in lang_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"lang_{h}"] = float(v) if v else 0
            except: feat[f"lang_{h}"] = 0

        # Open-end text
        oe_texts = [clean(row[i]) for i, _ in oe_cols if i < len(row) and clean(row[i])]
        oe_lens = [len(t) for t in oe_texts]
        oe_words = [len(t.split()) for t in oe_texts]
        feat["oe_count"] = len(oe_texts)
        feat["oe_total_chars"] = sum(oe_lens)
        feat["oe_max_chars"] = max(oe_lens) if oe_lens else 0
        feat["oe_mean_chars"] = np.mean(oe_lens) if oe_lens else 0
        feat["oe_total_words"] = sum(oe_words)
        feat["oe_max_words"] = max(oe_words) if oe_words else 0
        all_oe_text = " ".join(oe_texts).lower()
        words = all_oe_text.split()
        feat["oe_lex_div"] = len(set(words)) / len(words) if words else 0
        feat["oe_has_none"] = 1 if any(w in all_oe_text for w in ["none","n/a","na","nothing","no opinion","no idea"]) else 0
        feat["oe_generic"] = 1 if any(w in all_oe_text for w in ["good","fine","ok","okay","nice","great"]) and oe_words and max(oe_words) <= 3 else 0
        feat["oe_all_caps"] = 1 if any(t.isupper() and len(t) > 5 for t in oe_texts) else 0
        feat["oe_very_short"] = 1 if oe_lens and max(oe_lens) < 10 else 0
        feat["oe_short"] = 1 if oe_lens and max(oe_lens) < 30 else 0

        # Supplier
        sup_idx = hidx.get("SUPNAME")
        sup = clean(row[sup_idx]) if sup_idx and sup_idx < len(row) else ""
        feat["supplier_name"] = sup
        feat["supplier_missing"] = 1 if not sup or sup == "MISSING" else 0
        feat["supplier_is_none"] = 1 if sup == "None" or sup == "" else 0
        feat["supplier_reject_rate"] = sup_rates.get(sup, global_rate) * 100

        # Matrix
        mvals = [norm(row[i]) for i, _ in matrix_cols if i < len(row) and row[i] is not None and row[i] != ""]
        if mvals:
            ur = len(set(mvals)) / len(mvals)
            feat["matrix_unique_ratio"] = ur
            feat["matrix_straightline"] = 1 if ur <= 0.2 and len(mvals) >= 5 else 0
            feat["matrix_near_straightline"] = 1 if ur <= 0.4 and len(mvals) >= 5 else 0
            feat["matrix_count"] = len(mvals)
            feat["matrix_unique_count"] = len(set(mvals))
            vc = Counter(mvals)
            feat["matrix_most_common_freq"] = vc.most_common(1)[0][1] / len(mvals)
        else:
            feat.update({"matrix_unique_ratio": 1.0, "matrix_straightline": 0,
                         "matrix_near_straightline": 0, "matrix_count": 0,
                         "matrix_unique_count": 0, "matrix_most_common_freq": 0})

        # Coded diversity
        cvals = [str(norm(row[i])) for i, _ in coded_cols if i < len(row) and row[i] is not None and row[i] != ""]
        feat["coded_count"] = len(cvals)
        feat["coded_unique_ratio"] = len(set(cvals)) / len(cvals) if cvals else 1.0
        dk = sum(1 for v in cvals if any(x in v.lower() for x in ["don't know","dk","not sure","no answer"]))
        feat["coded_dk_count"] = dk
        feat["coded_dk_ratio"] = dk / len(cvals) if cvals else 0

        # Answer entropy (raw)
        all_answers = [str(v) for v in (mvals + cvals) if v is not None]
        if all_answers:
            vc = Counter(all_answers)
            probs = np.array([c / len(all_answers) for c in vc.values()])
            feat["answer_entropy"] = float(-np.sum(probs * np.log2(probs + 1e-10)))
            feat["answer_max_freq"] = max(vc.values()) / len(all_answers)
            feat["answer_unique_ratio"] = len(vc) / len(all_answers)
        else:
            feat["answer_entropy"] = 0
            feat["answer_max_freq"] = 0
            feat["answer_unique_ratio"] = 1.0

        # RD_Search
        for i, h in rd_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"rd_{h}"] = float(v) if v else 0
            except: feat[f"rd_{h}"] = 0

        # Flags
        for i, h in flag_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"flag_{h}"] = float(v) if v else 0
            except: feat[f"flag_{h}"] = 0

        # Paste flags
        for i, h in paste_cols:
            v = row[i] if i < len(row) else None
            try: feat[f"paste_{h}"] = float(v) if v else 0
            except: feat[f"paste_{h}"] = 0

        # Duplicates
        oe_t = all_oe[idx]
        feat["oe_dup_count"] = oe_ctr.get(oe_t.strip().lower(), 0) if oe_t.strip() else 0
        feat["oe_is_dup"] = 1 if feat["oe_dup_count"] > 1 else 0
        feat["ua_dup_count"] = ua_ctr.get(all_ua[idx], 0)
        feat["ua_is_dup"] = 1 if feat["ua_dup_count"] > 1 else 0
        feat["ip_dup_count"] = ip_ctr.get(all_ip[idx], 0)
        feat["ip_is_dup"] = 1 if feat["ip_dup_count"] > 1 else 0

        # Demographics
        for i, h in demo_cols:
            feat[f"demo_{h}"] = clean(row[i]) if i < len(row) else ""
        for h in ["vos", "vbrowser", "vmobiledevice", "vmobileos"]:
            i = hidx.get(h)
            feat[f"tech_{h}"] = clean(row[i]) if i and i < len(row) else ""

        # Engineered features
        feat["supplier_x_signals"] = feat["supplier_reject_rate"] * feat["signal_count"] / 100
        feat["supplier_x_t1"] = feat["supplier_reject_rate"] * feat["t1_count"] / 100
        feat["supplier_x_t2"] = feat["supplier_reject_rate"] * feat["t2_count"] / 100
        feat["signals_x_matrix"] = feat["signal_count"] * (1 - feat["matrix_unique_ratio"])

        feat["dataset"] = filepath.name
        features.append(feat)

        # Build answer chain for agent review
        chain = []
        for i, h in enumerate(headers):
            if not h: continue
            role = roles.get(str(h), "")
            if role in ("id", "label", "quota", "timing", "nlp_metadata", "review_metadata",
                        "quality_flag", "paste_flag", "no_answer", "routing", "user_agent", "ip"):
                continue
            v = row[i] if i < len(row) else None
            if v is None or v == "": continue
            dm = datamap.get(str(h), {})
            qtext = dm.get("question_text", str(h))
            labels = dm.get("labels", {})
            label = labels.get(str(norm(v)), str(v))
            chain.append({
                "field": str(h),
                "question_text": qtext,
                "answer_type": role,
                "raw_value": str(norm(v)),
                "label": str(label),
            })
        answer_chains.append({
            "respondent_id": rid,
            "answer_chain": chain,
            "oe_text": all_oe_text,
            "qtime_seconds": qt,
            "qtime_minutes": qt / 60.0,
            "qtime_percentile": feat["qtime_percentile"],
            "supplier": sup,
            "supplier_reject_rate": feat["supplier_reject_rate"],
            "signals": list(sigs),
            "signal_count": feat["signal_count"],
            "t1_count": feat["t1_count"],
            "t2_count": feat["t2_count"],
            "t3_count": feat["t3_count"],
            "oe_dup_count": feat["oe_dup_count"],
            "ua_dup_count": feat["ua_dup_count"],
            "ip_dup_count": feat["ip_dup_count"],
            "matrix_unique_ratio": feat["matrix_unique_ratio"],
            "matrix_straightline": feat["matrix_straightline"],
            "answer_entropy": feat["answer_entropy"],
            "oe_total_chars": feat["oe_total_chars"],
            "lang_readlevel": feat.get("lang_LangAssessReadLevel", 0),
        })

    df = pd.DataFrame(features)
    for col in ["qtime_seconds", "signal_count"]:
        df[f"{col}_zscore"] = 0.0
        vals = df[col]
        std = vals.std()
        df[f"{col}_zscore"] = (vals - vals.mean()) / (std if std > 0 else 1)

    return df, datamap, roles, answer_chains


def ml_triage(df, model_path=MODEL_DIR / "survey_quality_model.pkl"):
    """Run ML model to get triage risk probabilities (0-1)."""
    # Try joblib first (more portable), then pickle
    joblib_path = model_path.with_suffix(".joblib")

    model_data = None
    load_error = None

    for path in [joblib_path, model_path]:
        if path.exists():
            try:
                if path.suffix == ".joblib":
                    import joblib
                    model_data = joblib.load(path)
                else:
                    with open(path, "rb") as f:
                        model_data = pickle.load(f)
                break
            except Exception as e:
                load_error = e
                continue

    if model_data is None:
        if load_error:
            print(f"WARNING: ML model failed to load: {load_error}")
            print("This is likely a scikit-learn version mismatch.")
            print(f"  Current sklearn: {getattr(__import__('sklearn'), '__version__', 'unknown')}")
            meta = getattr(model_data, '_metadata', None) if model_data else None
            if meta:
                print(f"  Model trained with sklearn: {meta.get('sklearn_version', 'unknown')}")
        else:
            print(f"WARNING: Model not found at {model_path} or {joblib_path}")
        print("Using rule-based triage only (ml_triage_score=0.5 for all).")
        print("This significantly weakens the pipeline.")
        df["ml_triage_score"] = 0.5
        return df

    try:
        model = model_data["model"]
        iso = model_data["calibrator"]
        train_features = model_data["feature_columns"]
        supplier_rates = model_data.get("supplier_rates", {})
        global_rate = model_data.get("global_reject_rate", 0.25)

        # Use model's supplier rates if available (more accurate)
        if supplier_rates:
            df["supplier_reject_rate"] = df["supplier_name"].map(supplier_rates).fillna(global_rate) * 100
            df["supplier_x_signals"] = df["supplier_reject_rate"] * df["signal_count"] / 100
            df["supplier_x_t1"] = df["supplier_reject_rate"] * df["t1_count"] / 100
            df["supplier_x_t2"] = df["supplier_reject_rate"] * df["t2_count"] / 100

        non_feat = {"respondent_id", "label", "dataset", "supplier_name"}
        feat_cols = [c for c in df.columns if c not in non_feat]
        X = df[feat_cols].copy()
        for col in X.select_dtypes(include=["object"]).columns:
            X[col] = pd.Categorical(X[col]).codes
        X = X.fillna(0)

        for c in train_features:
            if c not in X.columns: X[c] = 0
        X = X[train_features]

        y_proba = model.predict_proba(X)[:, 1]
        y_cal = iso.transform(y_proba)

        df["ml_triage_score"] = y_cal
        print(f"  ML triage: {len(df)} respondents scored (model loaded successfully)")
        return df
    except Exception as e:
        print(f"WARNING: ML model loaded but prediction failed: {e}")
        print("Using rule-based triage only (ml_triage_score=0.5 for all).")
        print("This significantly weakens the pipeline. Check sklearn version compatibility.")
        df["ml_triage_score"] = 0.5
        return df


def compute_key_signals(df, answer_chains):
    """Compute the key signals that drove the score for each respondent."""
    signals_list = []
    for idx, row in df.iterrows():
        chain = answer_chains[idx] if idx < len(answer_chains) else {}
        signals = []

        # TIER 1 signals
        if row.get("t1_count", 0) > 0:
            signals.append(f"TIER1: {row['t1_count']} critical signal(s)")

        # Supplier risk
        sr = row.get("supplier_reject_rate", 0)
        if sr > 30:
            signals.append(f"High-risk supplier ({sr:.0f}% reject rate)")
        elif sr > 20:
            signals.append(f"Medium-risk supplier ({sr:.0f}% reject rate)")

        # Timing
        if row.get("qtime_percentile") == "bottom_10":
            signals.append(f"Very fast completion ({row.get('qtime_minutes', 0):.1f} min, bottom 10%)")
        elif row.get("qtime_percentile") == "bottom_25":
            signals.append(f"Fast completion ({row.get('qtime_minutes', 0):.1f} min, bottom 25%)")

        # Open-end quality
        oe_chars = row.get("oe_total_chars", 0)
        if oe_chars < 10 and oe_chars >= 0:
            signals.append("Very short open-end response")
        elif oe_chars < 30:
            signals.append("Short open-end response")

        # Matrix straightlining
        if row.get("matrix_straightline", 0) == 1:
            signals.append("Matrix straightlining detected")
        elif row.get("matrix_near_straightline", 0) == 1:
            signals.append("Near-straightline matrix pattern")

        # Answer entropy
        ent = row.get("answer_entropy", 0)
        if ent < 1.0:
            signals.append(f"Low answer entropy ({ent:.2f})")

        # Duplicates
        if row.get("oe_is_dup", 0) == 1:
            signals.append(f"Duplicate open-end text (shared with {row['oe_dup_count']} others)")
        if row.get("ip_is_dup", 0) == 1:
            signals.append(f"Duplicate IP (shared with {row['ip_dup_count']} others)")

        # LangAssess
        rl = row.get("lang_LangAssessReadLevel", 0)
        if rl > 0 and rl < 5:
            signals.append(f"Low readability score ({rl:.1f})")

        # Signal count
        sc = row.get("signal_count", 0)
        if sc >= 5:
            signals.append(f"High signal count ({sc})")

        if not signals:
            signals.append("No significant quality signals detected")

        signals_list.append(signals)

    return signals_list


def classify_open_end_semantic(oe_text):
    """Classify open-end text semantically based on agent review findings.
    
    Returns (classification, is_personal, is_discard_signal).
    
    Classifications from agent review of 303 missed discards:
    - 'first_person_personal': Uses first-person language with personal experience (KEEP signal)
    - 'third_person_meta': Describes what the survey was about in third person (DISCARD signal)
    - 'generic_topic_restatement': Restates survey topic without personal insight (DISCARD signal)
    - 'off_topic_incoherent': Off-topic, incoherent, or non-answer (DISCARD signal)
    - 'templated_truncated': Truncated mid-sentence or templated phrase (DISCARD signal)
    - 'generic_praise': Generic praise like "thank you" or "good survey" (DISCARD signal)
    - 'gibberish': Keyboard mashing or nonsense (DISCARD signal)
    - 'substantive': On-topic with some substance (neutral)
    """
    if not oe_text or len(oe_text.strip()) < 3:
        return "empty", False, True

    oe = oe_text.lower().strip()

    # Gibberish detection: high consonant ratio or keyboard mashing
    if len(oe) > 10:
        consonants = sum(1 for c in oe if c in "bcdfghjklmnpqrstvwxyz")
        vowels = sum(1 for c in oe if c in "aeiou")
        if vowels > 0 and consonants / vowels > 5:
            return "gibberish", False, True
        # Check for repeated keyboard patterns
        if any(p in oe for p in ["bvnhgj", "vcgtr", "asdf", "qwer"]):
            return "gibberish", False, True

    # Generic praise
    if any(p in oe for p in ["thank you", "good survey", "nice survey", "great survey", "everything is good", "all good", "this is so good"]):
        if len(oe) < 40:
            return "generic_praise", False, True

    # Third-person meta-description (biggest missed pattern)
    third_person_markers = [
        "this survey was about", "this survey is about", "the survey was about",
        "the survey is about", "it was about", "it is about",
        "questions were asked", "customers were asked", "respondents were asked",
        "the purpose of this study", "researching the habits",
        "asked about their", "focused on", "mainly about",
        "it looked into", "assessing the importance",
    ]
    if any(m in oe for m in third_person_markers):
        return "third_person_meta", False, True

    # Off-topic or incoherent
    off_topic_markers = [
        "good night", "hope you", "the government is", "conspiracy",
        "this isn't just", "fraud masquerading",
    ]
    if any(m in oe for m in off_topic_markers):
        return "off_topic_incoherent", False, True

    # Very short non-answers
    if len(oe) < 15 and any(w in oe for w in ["i need it", "just need", "none", "n/a", "na", "no"]):
        return "off_topic_incoherent", False, True

    # Templated/truncated (ends mid-word or mid-sentence)
    if len(oe) > 30 and not oe.endswith(('.', '!', '?')) and oe[-1] not in '.!?':
        # Check if it looks truncated (ends with partial word)
        last_word = oe.split()[-1] if oe.split() else ""
        if len(last_word) <= 2 and not last_word.endswith(('ed', 'er', 'ly', 'ng')):
            return "templated_truncated", False, True
        # Check for common templated starts
        templated_starts = [
            "i decided to buy", "i wanted cleaner", "i wanted better",
            "i purchased", "i bought",
        ]
        if any(oe.startswith(t) for t in templated_starts) and len(oe) < 100:
            return "templated_truncated", False, True

    # First-person personal experience (KEEP signal)
    first_person_markers = ["i ", "my ", "me ", "we ", "our ", "i'm", "i've", "i'd"]
    personal_experience_markers = [
        "wife", "husband", "family", "kids", "children", "home", "house",
        "taste", "smell", "skin", "hair", "shower", "bath", "kitchen",
        "calcium", "hard water", "chlorine", "lead", "contaminants",
        "cost", "expensive", "cheap", "afford", "budget",
        "health", "safe", "safer", "concerned", "worried",
    ]
    has_first_person = any(m in oe for m in first_person_markers)
    has_personal = any(m in oe for m in personal_experience_markers)
    if has_first_person and has_personal:
        return "first_person_personal", True, False

    # Generic topic restatement (mentions water/filtration but no personal angle)
    if any(w in oe for w in ["water quality", "water filtration", "water filter", "hard water", "clean water", "water system"]):
        if not has_first_person:
            return "generic_topic_restatement", False, True

    # Substantive (on-topic, some content)
    if len(oe) > 30:
        return "substantive", False, False

    # Default: short but not clearly bad
    return "short_neutral", False, False


def agent_score_respondent(chain, ml_triage_score, matrix_prevalence=None):
    """Agent scoring function: assigns -1 to +1 score based on answer chain signals.
    
    The ML triage score is ONE input — it flags who to look at carefully.
    The agent score is based on the CONVERGENCE of signals observed in the
    answer chain, not a mechanical inversion of the ML score.
    
    Score mapping:
    -1.0 = clear discard (TIER 1 signal, or multiple strong converging signals)
    -0.5 = discard lean (high-risk supplier + signals, or incoherent profile)
     0.0 = uncertain (some concerns but not converging)
    +0.5 = likely keep (minor concerns but coherent chain)
    +1.0 = clear keep (no signals, coherent chain, reasonable timing)
    
    Args:
        chain: answer chain dict with signals and text
        ml_triage_score: ML model risk probability (0-1)
        matrix_prevalence: fraction of respondents with matrix_straightline=1.
            If >0.8, matrix straightlining is not discriminative and is downweighted.
    """
    score = 0.0
    reasons = []
    concerns = 0  # Track converging concerns

    # Matrix prevalence gating: if >80% of respondents have straightlining,
    # it's not discriminative (from agent review of false positives)
    matrix_gated = matrix_prevalence is not None and matrix_prevalence > 0.8

    # === STRONG SIGNALS (each can drive score negative on its own) ===

    # TIER 1 signals → strong discard (-0.5 each)
    if chain.get("t1_count", 0) > 0:
        score -= 0.5
        concerns += 3
        reasons.append(f"TIER 1 signal present ({chain['t1_count']} signals)")

    # TIER 2 signals → discard lean (-0.2 each)
    if chain.get("t2_count", 0) > 0:
        score -= 0.2
        concerns += 2
        reasons.append(f"TIER 2 signal present ({chain['t2_count']} signals)")

    # === SEMANTIC OPEN-END CLASSIFICATION (from agent review) ===
    # This is the biggest improvement: detecting third-person meta-descriptions,
    # generic topic restatements, and off-topic answers that rules miss.
    oe_text = chain.get("oe_text", "")
    oe_class, is_personal, is_discard_signal = classify_open_end_semantic(oe_text)

    if is_discard_signal:
        if oe_class in ("gibberish", "off_topic_incoherent"):
            score -= 0.3
            concerns += 2
            reasons.append(f"Open-end: {oe_class} ({oe_text[:50]}...)")
        elif oe_class in ("third_person_meta", "generic_topic_restatement"):
            score -= 0.25  # Increased — this is the biggest missed pattern per agent review
            concerns += 2  # Also counts as 2 concerns for convergence
            reasons.append(f"Open-end: {oe_class} — describes survey topic, not personal experience")
        elif oe_class == "templated_truncated":
            score -= 0.15
            concerns += 1
            reasons.append(f"Open-end: templated/truncated phrase")
        elif oe_class == "generic_praise":
            score -= 0.2
            concerns += 1
            reasons.append(f"Open-end: generic praise with no content")
        elif oe_class == "empty":
            score -= 0.15
            concerns += 1
            reasons.append("Open-end: empty or missing")

    # === MODERATE SIGNALS (need convergence to drive negative) ===

    # Supplier risk (but don't apply to individuals with no personal signals)
    sr = chain.get("supplier_reject_rate", 0)
    if sr > 30:
        score -= 0.1  # Reduced from 0.15 — agent review showed over-weighting
        concerns += 1
        reasons.append(f"High-risk supplier ({sr:.0f}% reject rate)")
    elif sr > 20:
        score -= 0.05  # Reduced from 0.08
        concerns += 1

    # Timing
    pct = chain.get("qtime_percentile", "")
    if pct == "bottom_10":
        score -= 0.12
        concerns += 1
        reasons.append("Very fast completion (bottom 10%)")
    elif pct == "bottom_25":
        score -= 0.06
        concerns += 1

    # Open-end length (but only if not already flagged by semantic classification)
    oe_chars = chain.get("oe_total_chars", 0)
    if oe_class not in ("empty",) and not is_discard_signal:
        if oe_chars < 10:
            score -= 0.1
            concerns += 1
            reasons.append("Very short open-end")
        elif oe_chars < 30:
            score -= 0.05

    # Matrix straightlining (gated by prevalence)
    if chain.get("matrix_straightline", 0) == 1:
        if matrix_gated:
            # Not discriminative alone, but still counts as a concern for convergence
            concerns += 1
        else:
            score -= 0.12
            concerns += 1
            reasons.append("Matrix straightlining")
    elif chain.get("matrix_near_straightline", 0) == 1:
        if not matrix_gated:
            score -= 0.04

    # Answer entropy
    ent = chain.get("answer_entropy", 0)
    if ent < 0.5:
        score -= 0.1
        concerns += 1
        reasons.append(f"Very low answer entropy ({ent:.2f})")
    elif ent < 1.0:
        score -= 0.05

    # Duplicates (but check if generic/topical)
    oe_dup = chain.get("oe_dup_count", 0)
    if oe_dup > 10:
        # From agent review: "water filtration systems" is topical inevitability
        # Only weight as concern if the text is unusual, not generic
        oe_lower = oe_text.lower().strip()
        generic_phrases = ["water filtration", "water filter", "water quality", "water system",
                          "filtration system", "home appliances", "kitchen faucet"]
        is_generic = any(p in oe_lower for p in generic_phrases)
        if is_generic and oe_dup < 50:
            # Generic text shared by many — not a strong fraud signal
            score -= 0.02
        else:
            score -= 0.08
            concerns += 1
            reasons.append(f"Open-end shared with {oe_dup} others")

    if chain.get("ip_dup_count", 0) > 10:
        score -= 0.08
        concerns += 1
        reasons.append(f"IP shared with {chain['ip_dup_count']} others")

    # LangAssess readability (but don't compound with short open-end)
    rl = chain.get("lang_readlevel", 0)
    if rl > 0 and rl < 3 and oe_chars > 25:
        score -= 0.06
        concerns += 1
        reasons.append(f"Very low readability ({rl:.1f})")

    # High signal count
    sc = chain.get("signal_count", 0)
    if sc >= 5:
        score -= 0.08
        concerns += 1
        reasons.append(f"High signal count ({sc})")

    # ML triage as a moderate signal
    if ml_triage_score > 0.7:
        score -= 0.1
        concerns += 1
        reasons.append(f"ML triage flags high risk ({ml_triage_score:.2f})")
    elif ml_triage_score > 0.5:
        score -= 0.05

    # === PROTECTIVE SIGNALS (from agent review of false positives) ===

    if chain.get("t1_count", 0) == 0 and chain.get("t2_count", 0) == 0:
        # First-person personal open-end is strong protective factor
        if is_personal:
            score += 0.15
            reasons.append("First-person personal open-end experience (protective)")

        # Substantive open-end
        if oe_chars > 100 and not is_discard_signal:
            score += 0.08
            reasons.append("Substantive open-end response")

        # High answer diversity
        if ent > 2.5:
            score += 0.05

        # Low-risk supplier
        if sr < 15 and sr > 0:
            score += 0.05

        # Generous timing
        if pct == "above_median":
            score += 0.05

        # Natural human errors (misspellings indicate human, not bot)
        if oe_text and any(m in oe_text.lower() for m in ["filtation", "wager", "abouy", "choldrildren", "becuase", "recieve"]):
            score += 0.05
            reasons.append("Natural misspellings indicate human respondent (protective)")

    # Convergence bonus
    if concerns >= 4:
        score -= 0.1
        reasons.append(f"Multiple converging concerns ({concerns})")
    elif concerns >= 3:
        score -= 0.05

    # Clamp to [-1, 1]
    score = max(-1.0, min(1.0, score))

    if not reasons:
        reasons.append("No significant signals — clean profile")

    return score, reasons


def reassess_respondent(chain, initial_score, initial_reasons):
    """Second-pass agent reassessment for respondents with score < 0.
    
    The agent score already encodes signal convergence (multiple signals = more negative).
    The reassessment checks for protective factors that might override the negative score,
    then assigns final judgment based on score thresholds:
    
    Score < -0.3: Strong negative → DISCARD (unless 2+ strong protective factors)
    Score < -0.1: Moderate negative → REVIEW
    Score < 0:    Slight negative → REVIEW
    Score >= 0:   KEEP (not reassessed)
    """
    reasons = list(initial_reasons)
    notes = []

    # Check protective factors
    protective = 0
    oe_chars = chain.get("oe_total_chars", 0)
    ent = chain.get("answer_entropy", 0)
    sr = chain.get("supplier_reject_rate", 0)
    pct = chain.get("qtime_percentile", "")

    if oe_chars > 150:
        protective += 1
        notes.append("Substantive detailed open-end (protective)")
    if ent > 3.0:
        protective += 1
        notes.append("Very high answer diversity (protective)")
    if sr > 0 and sr < 12:
        protective += 1
        notes.append("Very low-risk supplier (protective)")
    if pct == "above_median" and chain.get("qtime_minutes", 0) > 15:
        protective += 1
        notes.append("Generous completion time (protective)")

    # TIER 1 always discards regardless of protection
    if chain.get("t1_count", 0) > 0:
        return min(initial_score, -0.8), "DISCARD", notes + ["TIER 1 signal — confirmed discard"]

    # Score-based judgment with protective factor override
    if initial_score < -0.3:
        if protective >= 2:
            # Strong protective factors pull back from discard
            return max(initial_score, -0.15), "REVIEW", notes + ["Strong protective factors — pulled back from discard to review"]
        else:
            return initial_score, "DISCARD", notes + [f"Score {initial_score:.2f} — confirmed discard"]
    elif initial_score < -0.1:
        return initial_score, "REVIEW", notes + [f"Score {initial_score:.2f} — moderate concern, routing to review"]
    else:
        return initial_score, "REVIEW", notes + [f"Score {initial_score:.2f} — slight concern, routing to review"]


def run_pipeline(filepath, output_dir=None):
    """Full pipeline: ML triage → agent scoring → reassessment → output."""
    filepath = Path(filepath)
    if output_dir is None:
        output_dir = filepath.parent / f"{filepath.stem}_quality_output"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*80}")
    print(f"SURVEY QUALITY PIPELINE")
    print(f"{'='*80}")
    print(f"  Input: {filepath.name}")
    print(f"  Output: {output_dir}")

    # Step 1: Extract features + answer chains
    print(f"\n[1/5] Extracting features and answer chains...")
    df, datamap, roles, answer_chains = extract_features_and_chain(filepath)
    print(f"  Respondents: {len(df)}")
    print(f"  Datamap fields: {len(datamap)}")
    print(f"  Field roles: {Counter(roles.values())}")

    # Step 2: ML triage scoring
    print(f"\n[2/5] ML triage scoring...")
    df = ml_triage(df)
    print(f"  Mean triage score: {df['ml_triage_score'].mean():.3f}")
    print(f"  High-risk (>0.6): {(df['ml_triage_score'] > 0.6).sum()}")
    print(f"  Medium-risk (0.3-0.6): {((df['ml_triage_score'] >= 0.3) & (df['ml_triage_score'] <= 0.6)).sum()}")
    print(f"  Low-risk (<0.3): {(df['ml_triage_score'] < 0.3).sum()}")

    # Step 3: Agent scoring (-1 to +1)
    print(f"\n[3/5] Agent scoring (-1 to +1)...")

    # Compute matrix prevalence for gating
    matrix_prevalence = (df["matrix_straightline"] == 1).mean() if "matrix_straightline" in df.columns else None
    if matrix_prevalence is not None:
        print(f"  Matrix straightlining prevalence: {matrix_prevalence:.1%} {'(GATED - not discriminative)' if matrix_prevalence > 0.8 else ''}")

    agent_scores = []
    agent_reasons = []
    for idx, row in df.iterrows():
        chain = answer_chains[idx] if idx < len(answer_chains) else {}
        score, reasons = agent_score_respondent(chain, row["ml_triage_score"], matrix_prevalence=matrix_prevalence)
        agent_scores.append(score)
        agent_reasons.append(reasons)
    df["agent_score"] = agent_scores
    df["agent_reasons"] = agent_reasons

    n_negative = (df["agent_score"] < 0).sum()
    n_positive = (df["agent_score"] >= 0).sum()
    print(f"  Score < 0 (needs reassessment): {n_negative}")
    print(f"  Score >= 0: {n_positive}")
    print(f"  Score distribution: min={df['agent_score'].min():.2f}, median={df['agent_score'].median():.2f}, max={df['agent_score'].max():.2f}")

    # Step 4: Reassess respondents with score < 0
    print(f"\n[4/5] Reassessing {n_negative} respondents with negative scores...")
    final_scores = []
    final_judgments = []
    reassessment_notes_all = []
    for idx, row in df.iterrows():
        if row["agent_score"] < 0:
            chain = answer_chains[idx] if idx < len(answer_chains) else {}
            final_score, judgment, notes = reassess_respondent(
                chain, row["agent_score"], row["agent_reasons"])
            final_scores.append(final_score)
            final_judgments.append(judgment)
            reassessment_notes_all.append(notes)
        else:
            final_scores.append(row["agent_score"])
            final_judgments.append("KEEP")
            reassessment_notes_all.append(["No reassessment needed — positive score"])

    df["final_score"] = final_scores
    df["final_judgment"] = final_judgments
    df["reassessment_notes"] = reassessment_notes_all

    # Compute key signals
    df["key_signals"] = compute_key_signals(df, answer_chains)

    n_discard = (df["final_judgment"] == "DISCARD").sum()
    n_review = (df["final_judgment"] == "REVIEW").sum()
    n_keep = (df["final_judgment"] == "KEEP").sum()
    print(f"  Final judgments:")
    print(f"    DISCARD: {n_discard} ({n_discard/len(df):.1%})")
    print(f"    REVIEW:  {n_review} ({n_review/len(df):.1%})")
    print(f"    KEEP:    {n_keep} ({n_keep/len(df):.1%})")

    # Step 5: Generate outputs
    print(f"\n[5/5] Generating outputs...")

    # Annotated Excel
    excel_path = output_dir / f"{filepath.stem}_annotated.xlsx"
    write_annotated_excel(filepath, df, output_dir, excel_path)
    print(f"  Annotated Excel: {excel_path}")

    # Dashboard HTML
    dashboard_path = output_dir / f"{filepath.stem}_dashboard.html"
    write_dashboard(df, answer_chains, filepath.name, dashboard_path)
    print(f"  Dashboard: {dashboard_path}")

    # Summary JSON
    summary = {
        "dataset": filepath.name,
        "total_respondents": int(len(df)),
        "discard": int(n_discard),
        "review": int(n_review),
        "keep": int(n_keep),
        "mean_agent_score": float(df["agent_score"].mean()),
        "mean_final_score": float(df["final_score"].mean()),
        "mean_ml_triage": float(df["ml_triage_score"].mean()),
    }
    summary_path = output_dir / "summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"  Summary: {summary_path}")

    print(f"\n{'='*80}")
    print(f"COMPLETE")
    print(f"{'='*80}")

    return df


def write_annotated_excel(original_path, df, output_dir, excel_path):
    """Write the original Excel with added annotation columns."""
    # Load original workbook
    wb = openpyxl.load_workbook(original_path)
    ws = wb["A1"] if "A1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hidx = {h: i for i, h in enumerate(headers) if h}

    # Find respondent ID column
    rid_col = hidx.get("uuid") or hidx.get("record")
    if rid_col is None:
        print("WARNING: No respondent ID column found, cannot annotate Excel")
        return

    # Build lookup from df
    score_lookup = dict(zip(df["respondent_id"], df["final_score"]))
    judgment_lookup = dict(zip(df["respondent_id"], df["final_judgment"]))
    ml_lookup = dict(zip(df["respondent_id"], df["ml_triage_score"]))
    signals_lookup = dict(zip(df["respondent_id"], df["key_signals"]))
    reasons_lookup = {}
    notes_lookup = {}
    for _, row in df.iterrows():
        reasons_lookup[row["respondent_id"]] = "; ".join(row["agent_reasons"]) if isinstance(row["agent_reasons"], list) else str(row["agent_reasons"])
        notes_lookup[row["respondent_id"]] = "; ".join(row["reassessment_notes"]) if isinstance(row.get("reassessment_notes"), list) else ""

    # Add annotation columns
    from openpyxl.styles import PatternFill, Font, Alignment
    n_cols = len(headers)
    annotation_headers = [
        "ML_Triage_Score",
        "Agent_Score",
        "Final_Score",
        "Final_Judgment",
        "Key_Signals",
        "Agent_Reasons",
        "Reassessment_Notes",
    ]
    for i, h in enumerate(annotation_headers):
        col = n_cols + 1 + i
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Fill annotation rows
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    row_num = 2
    for row in ws.iter_rows(min_row=2, values_only=False):
        rid_val = row[rid_col].value if rid_col < len(row) else None
        rid = clean(rid_val) if rid_val else ""
        if not rid or rid not in score_lookup:
            row_num += 1
            continue

        ws.cell(row=row_num, column=n_cols + 1, value=round(ml_lookup[rid], 4))
        ws.cell(row=row_num, column=n_cols + 2, value=round(dict(zip(df["respondent_id"], df["agent_score"]))[rid], 4))
        ws.cell(row=row_num, column=n_cols + 3, value=round(score_lookup[rid], 4))
        judgment = judgment_lookup[rid]
        cell = ws.cell(row=row_num, column=n_cols + 4, value=judgment)
        if judgment == "DISCARD":
            cell.fill = red_fill
        elif judgment == "REVIEW":
            cell.fill = yellow_fill
        else:
            cell.fill = green_fill

        signals = signals_lookup[rid]
        ws.cell(row=row_num, column=n_cols + 5, value="; ".join(signals) if isinstance(signals, list) else str(signals))
        ws.cell(row=row_num, column=n_cols + 6, value=reasons_lookup.get(rid, ""))
        ws.cell(row=row_num, column=n_cols + 7, value=notes_lookup.get(rid, ""))

        row_num += 1

    # Auto-fit column widths for annotation columns
    for i in range(len(annotation_headers)):
        col = n_cols + 1 + i
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

    wb.save(excel_path)


def write_dashboard(df, answer_chains, dataset_name, output_path):
    """Write a self-contained HTML dashboard."""
    import base64

    # Compute statistics
    n = len(df)
    n_discard = int((df["final_judgment"] == "DISCARD").sum())
    n_review = int((df["final_judgment"] == "REVIEW").sum())
    n_keep = int((df["final_judgment"] == "KEEP").sum())

    # Score distribution buckets
    buckets = {"-1.0 to -0.5": 0, "-0.5 to 0.0": 0, "0.0 to 0.5": 0, "0.5 to 1.0": 0}
    for s in df["final_score"]:
        if s < -0.5: buckets["-1.0 to -0.5"] += 1
        elif s < 0: buckets["-0.5 to 0.0"] += 1
        elif s < 0.5: buckets["0.0 to 0.5"] += 1
        else: buckets["0.5 to 1.0"] += 1

    # Supplier analysis
    sup_stats = df.groupby("supplier_name").agg(
        count=("respondent_id", "count"),
        mean_score=("final_score", "mean"),
        discards=("final_judgment", lambda x: (x == "DISCARD").sum()),
        reviews=("final_judgment", lambda x: (x == "REVIEW").sum()),
    ).sort_values("count", ascending=False).head(15)

    # Top signals
    signal_counts = Counter()
    for signals in df["key_signals"]:
        if isinstance(signals, list):
            for s in signals:
                # Extract signal type (first part before colon or parenthesis)
                key = s.split(":")[0].split("(")[0].strip()
                signal_counts[key] += 1

    # Timing distribution
    timing_buckets = Counter(df["qtime_percentile"].fillna("unknown"))

    # LangAssess distribution
    lang_scores = df.get("lang_LangAssessReadLevel", pd.Series(dtype=float))
    lang_buckets = {"0-3": 0, "3-6": 0, "6-9": 0, "9-12": 0, "12+": 0}
    for s in lang_scores:
        if s < 3: lang_buckets["0-3"] += 1
        elif s < 6: lang_buckets["3-6"] += 1
        elif s < 9: lang_buckets["6-9"] += 1
        elif s < 12: lang_buckets["9-12"] += 1
        else: lang_buckets["12+"] += 1

    # Discard respondents table
    discard_df = df[df["final_judgment"] == "DISCARD"].sort_values("final_score")
    discard_rows = []
    for _, row in discard_df.head(50).iterrows():
        signals = "; ".join(row["key_signals"][:3]) if isinstance(row["key_signals"], list) else ""
        reasons = "; ".join(row["agent_reasons"][:2]) if isinstance(row["agent_reasons"], list) else ""
        discard_rows.append({
            "id": row["respondent_id"],
            "score": round(row["final_score"], 2),
            "ml": round(row["ml_triage_score"], 2),
            "supplier": row.get("supplier_name", ""),
            "qtime": f"{row.get('qtime_minutes', 0):.1f}m",
            "signals": signals,
            "reasons": reasons,
        })

    # Build HTML
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Survey Quality Dashboard — {dataset_name}</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; color: #333; }}
  .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
  h1 {{ font-size: 24px; margin-bottom: 8px; }}
  h2 {{ font-size: 18px; margin: 24px 0 12px; padding-bottom: 8px; border-bottom: 2px solid #ddd; }}
  .subtitle {{ color: #666; margin-bottom: 20px; font-size: 14px; }}
  .cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 24px; }}
  .card {{ background: white; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
  .card .label {{ font-size: 12px; color: #888; text-transform: uppercase; letter-spacing: 0.5px; }}
  .card .value {{ font-size: 28px; font-weight: 700; margin-top: 4px; }}
  .card.discard .value {{ color: #d32f2f; }}
  .card.review .value {{ color: #f9a825; }}
  .card.keep .value {{ color: #388e3c; }}
  .card.total .value {{ color: #1976d2; }}
  .card .pct {{ font-size: 12px; color: #888; margin-top: 2px; }}
  .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 24px; }}
  .panel {{ background: white; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
  .panel h3 {{ font-size: 14px; color: #555; margin-bottom: 12px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .bar-chart {{ margin-top: 8px; }}
  .bar-row {{ display: flex; align-items: center; margin-bottom: 6px; }}
  .bar-label {{ width: 120px; font-size: 13px; color: #555; }}
  .bar-track {{ flex: 1; height: 20px; background: #eee; border-radius: 4px; overflow: hidden; }}
  .bar-fill {{ height: 100%; border-radius: 4px; transition: width 0.3s; }}
  .bar-fill.red {{ background: #e53935; }}
  .bar-fill.orange {{ background: #fb8c00; }}
  .bar-fill.yellow {{ background: #fdd835; }}
  .bar-fill.green {{ background: #43a047; }}
  .bar-fill.blue {{ background: #1e88e5; }}
  .bar-value {{ width: 60px; text-align: right; font-size: 13px; color: #555; margin-left: 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ text-align: left; padding: 8px; background: #f5f5f5; border-bottom: 2px solid #ddd; font-weight: 600; }}
  td {{ padding: 8px; border-bottom: 1px solid #eee; }}
  tr:hover {{ background: #f9f9f9; }}
  .score-cell {{ font-weight: 700; }}
  .score-negative {{ color: #d32f2f; }}
  .score-positive {{ color: #388e3c; }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 11px; font-weight: 600; }}
  .badge.discard {{ background: #ffcdd2; color: #b71c1c; }}
  .badge.review {{ background: #fff9c4; color: #e65100; }}
  .badge.keep {{ background: #c8e6c9; color: #1b5e20; }}
  .signal-list {{ list-style: none; }}
  .signal-list li {{ padding: 4px 0; border-bottom: 1px solid #eee; font-size: 13px; }}
  .signal-list li:last-child {{ border-bottom: none; }}
  .signal-count {{ float: right; color: #888; font-weight: 600; }}
  .footer {{ margin-top: 24px; padding: 12px; text-align: center; color: #888; font-size: 12px; }}
</style>
</head>
<body>
<div class="container">
  <h1>Survey Quality Dashboard</h1>
  <div class="subtitle">{dataset_name} — {n} respondents — Generated by Autosurvey Pipeline</div>

  <!-- Summary Cards -->
  <div class="cards">
    <div class="card total">
      <div class="label">Total Respondents</div>
      <div class="value">{n}</div>
    </div>
    <div class="card discard">
      <div class="label">Discard</div>
      <div class="value">{n_discard}</div>
      <div class="pct">{n_discard/n:.1%} of total</div>
    </div>
    <div class="card review">
      <div class="label">Review</div>
      <div class="value">{n_review}</div>
      <div class="pct">{n_review/n:.1%} of total</div>
    </div>
    <div class="card keep">
      <div class="label">Keep</div>
      <div class="value">{n_keep}</div>
      <div class="pct">{n_keep/n:.1%} of total</div>
    </div>
  </div>

  <!-- Score Distribution + Timing -->
  <div class="grid">
    <div class="panel">
      <h3>Agent Score Distribution (-1 to +1)</h3>
      <div class="bar-chart">
"""
    for label, count in buckets.items():
        pct = count / n * 100 if n > 0 else 0
        color = "red" if "-0.5" in label or "-1.0" in label else ("yellow" if "0.0" in label else "green")
        html += f'        <div class="bar-row"><div class="bar-label">{label}</div><div class="bar-track"><div class="bar-fill {color}" style="width: {pct}%"></div></div><div class="bar-value">{count}</div></div>\n'

    html += """      </div>
      <p style="margin-top:12px;font-size:12px;color:#888;">Scores < 0 were reassessed by the agent. Final discards selected from reassessed pool.</p>
    </div>

    <div class="panel">
      <h3>Timing Distribution</h3>
      <div class="bar-chart">
"""
    timing_order = ["bottom_10", "bottom_25", "below_median", "above_median", "unknown"]
    timing_labels = {"bottom_10": "Bottom 10%", "bottom_25": "Bottom 25%", "below_median": "Below Median", "above_median": "Above Median", "unknown": "Unknown"}
    for key in timing_order:
        count = timing_buckets.get(key, 0)
        pct = count / n * 100 if n > 0 else 0
        color = "red" if "bottom_10" == key else ("orange" if "bottom_25" == key else ("yellow" if "below" in key else "green"))
        html += f'        <div class="bar-row"><div class="bar-label">{timing_labels.get(key, key)}</div><div class="bar-track"><div class="bar-fill {color}" style="width: {pct}%"></div></div><div class="bar-value">{count}</div></div>\n'

    html += """      </div>
    </div>
  </div>

  <!-- Top Signals + Supplier Analysis -->
  <div class="grid">
    <div class="panel">
      <h3>Top Population Signals</h3>
      <ul class="signal-list">
"""
    for signal, count in signal_counts.most_common(15):
        pct = count / n * 100
        html += f'        <li>{signal}<span class="signal-count">{count} ({pct:.1f}%)</span></li>\n'

    html += """      </ul>
    </div>

    <div class="panel">
      <h3>Supplier Analysis (Top 15)</h3>
      <table>
        <thead><tr><th>Supplier</th><th>N</th><th>Mean Score</th><th>Discards</th><th>Reviews</th></tr></thead>
        <tbody>
"""
    for sup, row in sup_stats.iterrows():
        if not sup or sup == "": sup = "(missing)"
        score_class = "score-negative" if row["mean_score"] < 0 else "score-positive"
        html += f'        <tr><td>{sup[:25]}</td><td>{int(row["count"])}</td><td class="score-cell {score_class}">{row["mean_score"]:.2f}</td><td>{int(row["discards"])}</td><td>{int(row["reviews"])}</td></tr>\n'

    html += """      </tbody>
      </table>
    </div>
  </div>

  <!-- LangAssess + ML Triage -->
  <div class="grid">
    <div class="panel">
      <h3>LangAssess Readability Distribution</h3>
      <div class="bar-chart">
"""
    lang_colors = {"0-3": "red", "3-6": "orange", "6-9": "yellow", "9-12": "green", "12+": "green"}
    for label, count in lang_buckets.items():
        pct = count / n * 100 if n > 0 else 0
        html += f'        <div class="bar-row"><div class="bar-label">{label}</div><div class="bar-track"><div class="bar-fill {lang_colors[label]}" style="width: {pct}%"></div></div><div class="bar-value">{count}</div></div>\n'

    html += """      </div>
      <p style="margin-top:8px;font-size:12px;color:#888;">Low readability (<3) is the most universal quality signal across all datasets.</p>
    </div>

    <div class="panel">
      <h3>ML Triage Score Distribution</h3>
      <div class="bar-chart">
"""
    ml_buckets = {"0.0-0.2 (low risk)": 0, "0.2-0.4": 0, "0.4-0.6 (uncertain)": 0, "0.6-0.8": 0, "0.8-1.0 (high risk)": 0}
    for s in df["ml_triage_score"]:
        if s < 0.2: ml_buckets["0.0-0.2 (low risk)"] += 1
        elif s < 0.4: ml_buckets["0.2-0.4"] += 1
        elif s < 0.6: ml_buckets["0.4-0.6 (uncertain)"] += 1
        elif s < 0.8: ml_buckets["0.6-0.8"] += 1
        else: ml_buckets["0.8-1.0 (high risk)"] += 1

    ml_colors = {"0.0-0.2 (low risk)": "green", "0.2-0.4": "green", "0.4-0.6 (uncertain)": "yellow", "0.6-0.8": "orange", "0.8-1.0 (high risk)": "red"}
    for label, count in ml_buckets.items():
        pct = count / n * 100 if n > 0 else 0
        html += f'        <div class="bar-row"><div class="bar-label">{label}</div><div class="bar-track"><div class="bar-fill {ml_colors[label]}" style="width: {pct}%"></div></div><div class="bar-value">{count}</div></div>\n'

    html += """      </div>
      <p style="margin-top:8px;font-size:12px;color:#888;">ML triage flags high-risk respondents for agent review. Not a standalone classifier.</p>
    </div>
  </div>

  <!-- Discard Table -->
  <div class="panel" style="margin-bottom:24px;">
    <h3>Discard Set ({n_discard} respondents)</h3>
"""
    if n_discard == 0:
        html += "    <p style='padding:12px;color:#888;'>No discards in this run.</p>\n"
    else:
        html += """    <table>
      <thead><tr><th>Respondent ID</th><th>Final Score</th><th>ML Triage</th><th>Supplier</th><th>QTime</th><th>Key Signals</th><th>Agent Reasons</th></tr></thead>
      <tbody>
"""
        for r in discard_rows:
            score_class = "score-negative" if r["score"] < 0 else "score-positive"
            html += f'        <tr><td>{r["id"]}</td><td class="score-cell {score_class}">{r["score"]}</td><td>{r["ml"]}</td><td>{r["supplier"][:20]}</td><td>{r["qtime"]}</td><td style="font-size:11px;">{r["signals"][:80]}</td><td style="font-size:11px;">{r["reasons"][:80]}</td></tr>\n'

        html += "      </tbody>\n    </table>\n"

    html += f"""  </div>

  <div class="footer">
    Generated by Autosurvey Survey Quality Pipeline — {dataset_name}
  </div>
</div>
</body>
</html>"""

    with open(output_path, "w") as f:
        f.write(html)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 survey_pipeline.py <xlsx_path> [--output-dir DIR]")
        return

    filepath = Path(sys.argv[1])
    output_dir = None

    for i, arg in enumerate(sys.argv[2:], 2):
        if arg == "--output-dir" and i + 1 < len(sys.argv):
            output_dir = Path(sys.argv[i + 1])

    run_pipeline(filepath, output_dir)


if __name__ == "__main__":
    main()
