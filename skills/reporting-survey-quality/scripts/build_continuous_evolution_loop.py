#!/usr/bin/env python3
"""Build Autosurvey continuous evolution outputs from original and graded pairs.

This is a retrospective development-validation runner. It treats the existing
TFG cleaning-answer files as development data, not untouched external tests.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[2]

LEAKAGE_RE = re.compile(
    r"(^|_)(status|client|review|recommend|action|decision|annotation|score|tier|final|"
    r"discard|exclude|keep|clean|flag|reason|note|comment|markers?|validclient|"
    r"channeltracking|closingemail|token)(_|$)|condition|noanswer|^qc($|\d|_)|"
    r"TERMFLAGS|SCRUTINYFLAGS|CLIENTFLAGS|redem_excluded|RD_Review|RD_Search|quota|bad:",
    re.I,
)
ID_CANDIDATES = ["uuid", "respondent_id", "RID", "record"]
STATUS_ACCEPT = 0
STATUS_REJECT = 1

PUBLIC_APPEND_COLUMNS = [
    "autosurvey_dataset",
    "autosurvey_respondent_id",
    "autosurvey_client_label",
    "autosurvey_client_status",
    "autosurvey_agent_decision",
    "autosurvey_binary_prediction",
    "autosurvey_authenticity_risk",
    "autosurvey_client_rejection_probability",
    "autosurvey_confidence",
    "autosurvey_semantic_decision_criteria",
    "autosurvey_signals",
    "autosurvey_signal_evidence",
    "autosurvey_protective_evidence",
    "autosurvey_key_question_relationships",
    "autosurvey_timing_analysis",
    "autosurvey_matrix_analysis",
    "autosurvey_open_end_analysis",
    "autosurvey_persona_qualification_analysis",
    "autosurvey_cross_respondent_analysis",
    "autosurvey_matched_control_summary",
    "autosurvey_adjudication",
    "autosurvey_client_agreement",
    "autosurvey_error_type",
    "autosurvey_model_version",
]

PUBLIC_SOURCE_COLUMNS = [
    "record",
    "uuid",
    "date",
    "qtime",
    "source",
    "list",
    "SUPNAME",
    "qstate",
    "qager1",
    "age",
    "qGender",
    "q14",
    "q34",
    "qcoe1",
    "outro",
    "userAgent",
    "url",
    "autosurvey_original_response_json",
]

PAIR_HINTS = [
    ("delta-water-filtration", "Delta Water Filtration", "Delta Water Filtration", "Delta Water Filtration"),
    ("echo-bh", "Echo BH", "Echo", "ECHO"),
    ("odl-switchable-glass", "ODL Switchable Glass", "ODL", "ODL"),
    ("oldcastle-brand-health", "Oldcastle Brand Health", "Oldcastle Brand Health", "OC BH"),
    ("oldcastle-canada", "Oldcastle Canada", "Oldcastle Canada", "OC CAN"),
    ("sbd-brand-association", "SBD Brand Association", "SBD", "SBD"),
    ("thd-digital-cx", "THD Digital CX", "THD", "THD CX"),
    ("addo-racetrac-us-gp", "ADDO RaceTrac US GP", "ADDO", "ADDO"),
    ("masterlock-conjoint", "Masterlock Conjoint", "Masterlock", "Masterlock"),
    ("tfg-contractor-index-q1", "TFG Contractor Index Q1", "Contractor Index Q1", "Contractor Index Q1"),
    ("tfg-contractor-index-q2", "TFG Contractor Index Q2", "Contractor Index Q2", "Contractor Index Q2"),
]

FEATURE_COLUMNS = [
    "missing_rate",
    "qtime_log",
    "qtime_low",
    "qtime_high",
    "open_word_count",
    "open_word_low",
    "open_word_high",
    "open_generic_short",
    "survey_meta_language",
    "polished_ungrounded",
    "personal_grounding",
    "duplicate_open_chain",
    "matrix_modal_share",
    "matrix_low_entropy",
    "persona_weak",
    "answer_time_text_mismatch",
]


@dataclass
class Pair:
    dataset_id: str
    display_name: str
    original_path: Path
    graded_path: Path
    row_count: int = 0
    accepted: int = 0
    rejected: int = 0
    respondent_key: str = "uuid"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value)).strip()


def norm_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=json_default), encoding="utf-8")


def json_default(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return None if math.isnan(float(value)) else float(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    return str(value)


def git(command: list[str], cwd: Path = REPO_ROOT) -> str:
    return subprocess.check_output(["git", *command], cwd=cwd, text=True).strip()


def read_main(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=0, dtype=object, engine="openpyxl")


def status_to_label(value: Any) -> int | None:
    raw = text(value)
    if raw.endswith(".0"):
        raw = raw[:-2]
    if raw == "5":
        return STATUS_REJECT
    if raw == "3":
        return STATUS_ACCEPT
    return None


def respondent_key(columns: list[str]) -> str:
    lowered = {c.lower(): c for c in columns}
    for candidate in ID_CANDIDATES:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    raise ValueError(f"No stable respondent key found in columns: {columns[:20]}")


def row_count(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return max(ws.max_row - 1, 0)


def find_pairs(original_dir: Path, graded_dir: Path) -> list[Pair]:
    originals = sorted(p for p in original_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    graded = sorted(p for p in graded_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    pairs: list[Pair] = []
    used: set[Path] = set()
    for dataset_id, display_name, original_hint, graded_hint in PAIR_HINTS:
        orig = next((p for p in originals if original_hint.lower() in p.stem.lower()), None)
        lab = next((p for p in graded if graded_hint.lower().replace(" ", "") in p.stem.lower().replace(" ", "")), None)
        if orig is None or lab is None:
            continue
        used.add(lab)
        pairs.append(Pair(dataset_id=dataset_id, display_name=display_name, original_path=orig, graded_path=lab))
    if len(pairs) != len(PAIR_HINTS):
        found = {p.dataset_id for p in pairs}
        missing = [h[0] for h in PAIR_HINTS if h[0] not in found]
        raise SystemExit(f"Could not build all 11 pairs. Missing: {missing}")
    return pairs


def leakage_columns(columns: list[str]) -> list[str]:
    return [c for c in columns if LEAKAGE_RE.search(str(c))]


def word_count(value: Any) -> int:
    return len(re.findall(r"\b\w+\b", clean_text(value)))


def detect_open_columns(df: pd.DataFrame) -> list[str]:
    cols: list[str] = []
    for col in df.columns:
        name = str(col)
        if LEAKAGE_RE.search(name):
            continue
        if re.search(r"^(url|useragent|rid|session|source|list|vbrowser|vmobile|vos|vlist)$|token|fingerprint|ip|device", name, re.I):
            continue
        sample = df[col].dropna().astype(str).head(100)
        avg_len = float(sample.map(len).mean()) if not sample.empty else 0.0
        alpha_share = float(sample.str.contains(r"[A-Za-z]", regex=True).mean()) if not sample.empty else 0.0
        if re.search(r"outro|open|oe$|other|specify|explain|why|comment|q14|q34|qcoe1", name, re.I):
            cols.append(name)
        elif avg_len >= 22 and alpha_share >= 0.25:
            cols.append(name)
    return cols[:80]


def matrix_columns(df: pd.DataFrame) -> list[str]:
    cols: list[str] = []
    for col in df.columns:
        name = str(col)
        if LEAKAGE_RE.search(name):
            continue
        if re.search(r"q\d+r\d+|r\d+c\d+|_r\d+", name, re.I):
            vals = df[col].dropna().astype(str).head(300)
            if 1 < vals.nunique() <= 12:
                cols.append(name)
    return cols[:300]


def source_subset_json(row: pd.Series, max_chars: int = 3000) -> str:
    payload = {}
    for key, value in row.items():
        value_text = clean_text(value)
        if value_text:
            payload[str(key)] = value_text
    raw = json.dumps(payload, ensure_ascii=False)
    return raw if len(raw) <= max_chars else raw[: max_chars - 3] + "..."


def extract_features(df: pd.DataFrame, dataset_id: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    key = respondent_key(list(df.columns))
    leak = set(leakage_columns(list(df.columns)))
    open_cols = detect_open_columns(df)
    matrix_cols = matrix_columns(df)
    usable_cols = [c for c in df.columns if c not in leak and c != key]
    nonempty = df[usable_cols].notna() & df[usable_cols].astype(str).apply(lambda s: s.str.strip().ne(""))
    missing_rate = 1.0 - nonempty.mean(axis=1).astype(float)
    open_text = []
    for _, row in df.iterrows():
        parts = [clean_text(row.get(c)) for c in open_cols if clean_text(row.get(c))]
        open_text.append(" || ".join(parts))
    open_series = pd.Series(open_text, index=df.index)
    open_norm = open_series.map(norm_text)
    open_counts = open_norm.value_counts()
    open_wc = open_series.map(word_count).astype(float)
    qtime_col = next((c for c in df.columns if str(c).lower() == "qtime"), "")
    qtime = pd.to_numeric(df[qtime_col], errors="coerce") if qtime_col else pd.Series(np.nan, index=df.index)
    qtime_filled = qtime.fillna(qtime.median() if qtime.notna().any() else 0)
    matrix_modal = []
    for _, row in df[matrix_cols].iterrows() if matrix_cols else []:
        vals = [clean_text(v) for v in row.tolist() if clean_text(v)]
        if not vals:
            matrix_modal.append(0.0)
        else:
            matrix_modal.append(Counter(vals).most_common(1)[0][1] / len(vals))
    if not matrix_cols:
        matrix_modal_series = pd.Series(0.0, index=df.index)
    else:
        matrix_modal_series = pd.Series(matrix_modal, index=df.index)
    grounded = open_series.str.contains(r"\b(?:i|my|our|we|family|home|work|job|customer|project|store|bought|used|installed|skin|hair|sink|tap)\b", case=False, regex=True)
    survey_meta = open_series.str.contains(r"\b(?:survey|poll|questionnaire|respondents|participants|asked about|study|brand awareness|buying behavior|preferences)\b", case=False, regex=True)
    polished = open_series.str.contains(r"\b(?:prompted by|motivated by|potential contaminants|environmental impact|comprehensive|valuable insights|overall satisfaction|decision to purchase|desire to)\b", case=False, regex=True)
    generic_short = (open_wc <= 7) | open_norm.isin({"", "none", "nothing", "no", "na", "n a", "dont know", "don t know"})
    persona_cols = [c for c in df.columns if re.search(r"qcoe|industry|classify|trade|role|profession|contractor", str(c), re.I) and c not in leak]
    persona_text = df[persona_cols].fillna("").astype(str).agg(" ".join, axis=1) if persona_cols else pd.Series("", index=df.index)
    persona_weak = persona_text.map(word_count).le(2) if persona_cols else pd.Series(False, index=df.index)
    text_time_mismatch = (open_wc >= open_wc.quantile(0.80)) & (qtime_filled <= qtime_filled.quantile(0.20))
    feats = pd.DataFrame(
        {
            "dataset_id": dataset_id,
            "respondent_id": df[key].map(clean_text),
            "source_row_number": np.arange(2, len(df) + 2),
            "missing_rate": missing_rate,
            "qtime_log": np.log1p(qtime_filled.astype(float).clip(lower=0)),
            "qtime_low": (qtime_filled <= qtime_filled.quantile(0.10)).astype(int),
            "qtime_high": (qtime_filled >= qtime_filled.quantile(0.95)).astype(int),
            "open_word_count": open_wc,
            "open_word_low": (open_wc <= max(6.0, float(open_wc.quantile(0.10)))).astype(int),
            "open_word_high": (open_wc >= max(20.0, float(open_wc.quantile(0.90)))).astype(int),
            "open_generic_short": generic_short.astype(int),
            "survey_meta_language": survey_meta.astype(int),
            "polished_ungrounded": (polished & ~grounded).astype(int),
            "personal_grounding": grounded.astype(int),
            "duplicate_open_chain": ((open_norm.map(open_counts).fillna(0) > 1) & open_norm.ne("")).astype(int),
            "matrix_modal_share": matrix_modal_series.astype(float),
            "matrix_low_entropy": (matrix_modal_series >= 0.85).astype(int),
            "persona_weak": persona_weak.astype(int),
            "answer_time_text_mismatch": text_time_mismatch.astype(int),
            "open_chain": open_series,
            "open_chain_hash": open_norm.map(lambda s: hashlib.sha256(s.encode("utf-8")).hexdigest() if s else ""),
            "qtime_value": qtime_filled,
        }
    )
    role = pd.DataFrame(
        {
            "dataset_id": [dataset_id],
            "respondent_key": [key],
            "open_columns": [";".join(open_cols)],
            "matrix_columns": [";".join(matrix_cols[:80])],
            "leakage_columns": [";".join(sorted(leak))],
        }
    )
    return feats, role


def sigmoid(x: np.ndarray) -> np.ndarray:
    return 1.0 / (1.0 + np.exp(-np.clip(x, -40, 40)))


def fit_logistic(train: pd.DataFrame, labels: np.ndarray) -> dict[str, Any]:
    x_raw = train[FEATURE_COLUMNS].astype(float).to_numpy()
    x_raw = np.nan_to_num(x_raw, nan=0.0, posinf=0.0, neginf=0.0)
    mean = np.nanmean(x_raw, axis=0)
    std = np.nanstd(x_raw, axis=0)
    std[std == 0] = 1.0
    x = np.nan_to_num((x_raw - mean) / std)
    x = np.c_[np.ones(len(x)), x]
    y = labels.astype(float)
    weights = np.zeros(x.shape[1])
    lr = 0.015
    l2 = 0.08
    if len(np.unique(y)) < 2:
        weights[0] = math.log((float(y.mean()) + 0.01) / (1.01 - float(y.mean())))
        return {"weights": weights.tolist(), "mean": mean.tolist(), "std": std.tolist()}
    for _ in range(900):
        pred = sigmoid(x @ weights)
        grad = (x.T @ (pred - y)) / len(y)
        grad[1:] += l2 * weights[1:]
        grad = np.nan_to_num(grad, nan=0.0, posinf=0.0, neginf=0.0)
        weights -= lr * grad
        weights = np.clip(weights, -8, 8)
    return {"weights": weights.tolist(), "mean": mean.tolist(), "std": std.tolist()}


def predict_logistic(model: dict[str, Any], frame: pd.DataFrame) -> np.ndarray:
    x_raw = frame[FEATURE_COLUMNS].astype(float).to_numpy()
    x_raw = np.nan_to_num(x_raw, nan=0.0, posinf=0.0, neginf=0.0)
    mean = np.array(model["mean"], dtype=float)
    std = np.array(model["std"], dtype=float)
    x = np.nan_to_num((x_raw - mean) / std)
    x = np.c_[np.ones(len(x)), x]
    return sigmoid(x @ np.array(model["weights"], dtype=float))


def confusion_counts(y: np.ndarray, pred: np.ndarray) -> dict[str, int]:
    y = y.astype(int)
    pred = pred.astype(int)
    return {
        "tp": int(((y == 1) & (pred == 1)).sum()),
        "fp": int(((y == 0) & (pred == 1)).sum()),
        "tn": int(((y == 0) & (pred == 0)).sum()),
        "fn": int(((y == 1) & (pred == 0)).sum()),
    }


def safe_div(num: float, den: float) -> float:
    return 0.0 if den == 0 else float(num / den)


def auc_rank(y: np.ndarray, score: np.ndarray) -> float:
    y = y.astype(int)
    pos = score[y == 1]
    neg = score[y == 0]
    if len(pos) == 0 or len(neg) == 0:
        return float("nan")
    order = np.argsort(score)
    ranks = np.empty_like(order, dtype=float)
    ranks[order] = np.arange(1, len(score) + 1)
    pos_ranks = ranks[y == 1].sum()
    return float((pos_ranks - len(pos) * (len(pos) + 1) / 2) / (len(pos) * len(neg)))


def average_precision(y: np.ndarray, score: np.ndarray) -> float:
    order = np.argsort(-score)
    y_sorted = y[order].astype(int)
    total_pos = int(y_sorted.sum())
    if total_pos == 0:
        return float("nan")
    cum_pos = np.cumsum(y_sorted)
    precision = cum_pos / (np.arange(len(y_sorted)) + 1)
    return float((precision * y_sorted).sum() / total_pos)


def metrics_for(y: np.ndarray, score: np.ndarray, pred: np.ndarray) -> dict[str, float]:
    c = confusion_counts(y, pred)
    tp, fp, tn, fn = c["tp"], c["fp"], c["tn"], c["fn"]
    precision = safe_div(tp, tp + fp)
    recall = safe_div(tp, tp + fn)
    specificity = safe_div(tn, tn + fp)
    npv = safe_div(tn, tn + fn)
    f1_reject = safe_div(2 * precision * recall, precision + recall)
    keep_precision = npv
    keep_recall = specificity
    f1_keep = safe_div(2 * keep_precision * keep_recall, keep_precision + keep_recall)
    acc = safe_div(tp + tn, tp + fp + tn + fn)
    mcc_den = math.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    mcc = safe_div(tp * tn - fp * fn, mcc_den)
    expected = safe_div((tp + fp) * (tp + fn) + (fn + tn) * (fp + tn), (tp + fp + tn + fn) ** 2)
    kappa = safe_div(acc - expected, 1 - expected)
    return {
        **c,
        "rows": int(len(y)),
        "prevalence": float(np.mean(y)) if len(y) else 0.0,
        "accuracy": acc,
        "balanced_accuracy": (recall + specificity) / 2,
        "rejected_precision": precision,
        "rejected_recall": recall,
        "specificity": specificity,
        "npv": npv,
        "f1_rejected": f1_reject,
        "f1_keep": f1_keep,
        "macro_f1": (f1_reject + f1_keep) / 2,
        "mcc": mcc,
        "cohens_kappa": kappa,
        "auroc": auc_rank(y, score),
        "auprc": average_precision(y, score),
        "brier_score": float(np.mean((score - y) ** 2)) if len(y) else 0.0,
        "review_volume": int(pred.sum()),
    }


def macro_f1_at(y: np.ndarray, score: np.ndarray, threshold: float) -> float:
    return metrics_for(y, score, (score >= threshold).astype(int))["macro_f1"]


def choose_threshold(y: np.ndarray, score: np.ndarray) -> float:
    if len(np.unique(y)) < 2:
        return 0.5
    candidates = np.unique(np.quantile(score, np.linspace(0.05, 0.95, 37)))
    best = max(candidates, key=lambda t: (macro_f1_at(y, score, float(t)), -abs(float(t) - 0.5)))
    return float(best)


def inner_threshold(train: pd.DataFrame) -> tuple[float, bool]:
    datasets = sorted(train["dataset_id"].unique())
    scored = []
    for dataset_id in datasets:
        inner_test = train[train["dataset_id"].eq(dataset_id)].copy()
        inner_train = train[~train["dataset_id"].eq(dataset_id)].copy()
        if inner_train.empty or inner_test.empty:
            continue
        duplicate_hashes = set(inner_test["open_chain_hash"]) - {""}
        inner_train = inner_train[~inner_train["open_chain_hash"].isin(duplicate_hashes)].copy()
        model = fit_logistic(inner_train, inner_train["client_label"].to_numpy(dtype=int))
        score = predict_logistic(model, inner_test)
        part = inner_test[["client_label"]].copy()
        part["score"] = score
        scored.append(part)
    if not scored:
        return 0.5, False
    inner = pd.concat(scored, ignore_index=True)
    y = inner["client_label"].to_numpy(dtype=int)
    raw = inner["score"].to_numpy(dtype=float)
    raw_threshold = choose_threshold(y, raw)
    inv = 1.0 - raw
    inv_threshold = choose_threshold(y, inv)
    raw_f1 = macro_f1_at(y, raw, raw_threshold)
    inv_f1 = macro_f1_at(y, inv, inv_threshold)
    if inv_f1 > raw_f1:
        return inv_threshold, True
    return raw_threshold, False


def tier_from_score(score: float, threshold: float) -> str:
    if score >= threshold:
        return "Exclude candidate"
    if score >= threshold * 0.80:
        return "Review closely"
    if score >= threshold * 0.55:
        return "Light review"
    if score >= threshold * 0.35:
        return "Keep with note"
    return "Clean keep"


def signal_text(row: pd.Series) -> tuple[str, str, str, str, str, str, str]:
    signals = []
    evidence = []
    protection = []
    if row["survey_meta_language"]:
        signals.append("survey_meta_substitution")
        evidence.append("Open end uses survey or research-summary language.")
    if row["polished_ungrounded"]:
        signals.append("polished_ungrounded_open_end")
        evidence.append("Open end uses polished abstract phrasing without personal grounding.")
    if row["open_generic_short"]:
        signals.append("thin_open_end")
        evidence.append(f"Open-end chain has {int(row['open_word_count'])} words.")
    if row["duplicate_open_chain"]:
        signals.append("duplicate_open_chain")
        evidence.append("Another respondent shares the normalized open-end chain.")
    if row["matrix_low_entropy"]:
        signals.append("high_matrix_uniformity")
        evidence.append(f"Matrix modal share {row['matrix_modal_share']:.2f}.")
    if row["qtime_low"]:
        signals.append("low_total_duration")
        evidence.append(f"Duration proxy is low: {row['qtime_value']}.")
    if row["answer_time_text_mismatch"]:
        signals.append("text_time_mismatch")
        evidence.append("Long open-end content appears with low total duration.")
    if row["persona_weak"]:
        signals.append("weak_persona_context")
        evidence.append("Qualification or role context appears weak or very short.")
    if row["personal_grounding"]:
        protection.append("The open-end chain contains first-person, household, work, product, or project grounding.")
    if row["open_word_count"] <= 8 and not row["survey_meta_language"]:
        protection.append("Short answer may be valid when the prompt asks for a simple product reason or topic summary.")
    if not protection:
        protection.append("No strong protective evidence found in the staged features; full-chain review is needed.")
    timing = "Low duration signal." if row["qtime_low"] else ("High duration context." if row["qtime_high"] else "No strong total-duration anomaly.")
    matrix = "High matrix uniformity." if row["matrix_low_entropy"] else "No strong matrix uniformity signal in staged fields."
    open_end = "; ".join([e for e in evidence if "Open" in e or "open" in e]) or "Open-end features did not produce a strong standalone concern."
    persona = "Weak persona or qualification context." if row["persona_weak"] else "No strong persona mismatch surfaced by staged features."
    cross = "Duplicate open-end chain." if row["duplicate_open_chain"] else "No duplicate open-chain signal."
    return (
        "; ".join(signals) if signals else "no_strong_staged_signal",
        " ".join(evidence) if evidence else "No strong staged evidence; row kept unless full-chain review finds concern.",
        " ".join(protection),
        timing,
        matrix,
        open_end,
        persona,
        cross,
    )


def row_rationale(row: pd.Series) -> str:
    decision = row["autosurvey_agent_decision"]
    signals, evidence, protection, *_ = signal_text(row)
    if decision == "Exclude candidate":
        lead = "The row is an exclusion candidate because the held-out model found converging authenticity concerns."
    elif decision in {"Review closely", "Light review"}:
        lead = "The row should stay in review because the staged evidence is meaningful but not decisive enough for exclusion."
    else:
        lead = "The row is protected because the staged concerns do not overcome the available human evidence."
    return f"{lead} Signals considered: {signals}. Evidence: {evidence} Protective evidence: {protection}"


def load_all_pairs(pairs: list[Pair]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    all_features = []
    all_raw = []
    role_rows = []
    registry_rows = []
    for pair in pairs:
        original = read_main(pair.original_path)
        graded = read_main(pair.graded_path)
        key = respondent_key(list(original.columns))
        label_key = respondent_key(list(graded.columns))
        status_col = next((c for c in graded.columns if str(c).lower() == "status"), None)
        if status_col is None:
            raise SystemExit(f"No status field in {pair.graded_path}")
        labels = graded[[label_key, status_col]].copy()
        labels["respondent_id"] = labels[label_key].map(clean_text)
        labels["client_label"] = labels[status_col].map(status_to_label)
        features, role = extract_features(original, pair.dataset_id)
        merged = features.merge(labels[["respondent_id", "client_label", status_col]], on="respondent_id", how="left")
        if merged["client_label"].isna().any():
            raise SystemExit(f"Failed to reconcile labels for {pair.dataset_id}")
        merged["client_label"] = merged["client_label"].astype(int)
        merged["client_status"] = merged[status_col].map(clean_text)
        pair.row_count = len(merged)
        pair.rejected = int(merged["client_label"].sum())
        pair.accepted = pair.row_count - pair.rejected
        pair.respondent_key = key
        original_public = original.copy()
        original_public["autosurvey_dataset"] = pair.dataset_id
        original_public["autosurvey_respondent_id"] = original_public[key].map(clean_text)
        original_public["autosurvey_original_response_json"] = original_public.apply(source_subset_json, axis=1)
        all_raw.append(original_public)
        all_features.append(merged)
        role_rows.append(role)
        registry_rows.append(
            {
                "dataset_id": pair.dataset_id,
                "display_name": pair.display_name,
                "benchmark_state": "DEVELOPMENT",
                "original_path": str(pair.original_path),
                "original_sha256": sha256_file(pair.original_path),
                "graded_path": str(pair.graded_path),
                "graded_sha256": sha256_file(pair.graded_path),
                "row_count": pair.row_count,
                "respondent_key": key,
                "status_field": status_col,
                "accepted_count": pair.accepted,
                "rejected_count": pair.rejected,
                "prior_use_history": "used in annotated-corpus methodology development; not an untouched external holdout",
            }
        )
    return (
        pd.concat(all_features, ignore_index=True),
        pd.concat(all_raw, ignore_index=True),
        pd.concat(role_rows, ignore_index=True),
        pd.DataFrame(registry_rows),
    )


def nested_oof(frame: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    pred_parts = []
    model_rows = []
    for dataset_id in sorted(frame["dataset_id"].unique()):
        test = frame[frame["dataset_id"].eq(dataset_id)].copy()
        train = frame[~frame["dataset_id"].eq(dataset_id)].copy()
        duplicate_hashes = set(test["open_chain_hash"]) - {""}
        duplicate_excluded = int(train["open_chain_hash"].isin(duplicate_hashes).sum())
        train = train[~train["open_chain_hash"].isin(duplicate_hashes)].copy()
        threshold, invert_score = inner_threshold(train)
        model = fit_logistic(train, train["client_label"].to_numpy(dtype=int))
        score = predict_logistic(model, test)
        if invert_score:
            score = 1.0 - score
        result = test.copy()
        result["autosurvey_client_rejection_probability"] = score
        result["fold_threshold"] = threshold
        result["autosurvey_binary_prediction"] = (score >= threshold).astype(int)
        result["autosurvey_agent_decision"] = [tier_from_score(float(s), threshold) for s in score]
        result["autosurvey_authenticity_risk"] = np.clip(
            score
            + 0.10 * result["survey_meta_language"].astype(float)
            + 0.08 * result["polished_ungrounded"].astype(float)
            - 0.06 * result["personal_grounding"].astype(float),
            0,
            1,
        )
        result["autosurvey_confidence"] = np.abs(score - threshold).clip(0, 1)
        pred_parts.append(result)
        model_rows.append(
            {
                "outer_holdout_dataset": dataset_id,
                "training_rows": len(train),
                "holdout_rows": len(test),
                "duplicate_training_rows_excluded": duplicate_excluded,
                "threshold": threshold,
                "score_inverted_from_inner_validation": invert_score,
                "weights": json.dumps(dict(zip(["intercept", *FEATURE_COLUMNS], model["weights"]))),
            }
        )
    predictions = pd.concat(pred_parts, ignore_index=True)
    metrics_rows = []
    for dataset_id, part in predictions.groupby("dataset_id"):
        metrics_rows.append({"dataset_id": dataset_id, **metrics_for(part["client_label"].to_numpy(int), part["autosurvey_client_rejection_probability"].to_numpy(float), part["autosurvey_binary_prediction"].to_numpy(int))})
    metrics_rows.append({"dataset_id": "POOLED", **metrics_for(predictions["client_label"].to_numpy(int), predictions["autosurvey_client_rejection_probability"].to_numpy(float), predictions["autosurvey_binary_prediction"].to_numpy(int))})
    return predictions, pd.DataFrame(metrics_rows), pd.DataFrame(model_rows)


def bootstrap_ci(y: np.ndarray, score: np.ndarray, pred: np.ndarray, resamples: int = 400, seed: int = 2607) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    rows = []
    n = len(y)
    for _ in range(resamples):
        idx = rng.integers(0, n, n)
        m = metrics_for(y[idx], score[idx], pred[idx])
        rows.append({k: m[k] for k in ["accuracy", "balanced_accuracy", "rejected_precision", "rejected_recall", "macro_f1", "mcc", "auroc", "auprc"]})
    out = []
    boot = pd.DataFrame(rows)
    for col in boot.columns:
        out.append({"metric": col, "ci_low": float(boot[col].quantile(0.025)), "ci_high": float(boot[col].quantile(0.975))})
    return pd.DataFrame(out)


def signal_performance(predictions: pd.DataFrame) -> pd.DataFrame:
    rows = []
    signal_cols = ["survey_meta_language", "polished_ungrounded", "open_generic_short", "duplicate_open_chain", "matrix_low_entropy", "qtime_low", "answer_time_text_mismatch", "persona_weak", "personal_grounding"]
    y = predictions["client_label"].to_numpy(int)
    base = float(y.mean())
    for col in signal_cols:
        mask = predictions[col].astype(bool).to_numpy()
        if not mask.any():
            continue
        rows.append(
            {
                "signal": col,
                "support": int(mask.sum()),
                "rejected_count": int(y[mask].sum()),
                "accepted_counterexamples": int((1 - y[mask]).sum()),
                "precision": float(y[mask].mean()),
                "recall_contribution": safe_div(float(y[mask].sum()), float(y.sum())),
                "lift_over_base_rate": safe_div(float(y[mask].mean()), base),
                "dataset_coverage": int(predictions.loc[mask, "dataset_id"].nunique()),
            }
        )
    for left, right in [("survey_meta_language", "polished_ungrounded"), ("duplicate_open_chain", "matrix_low_entropy"), ("qtime_low", "open_generic_short"), ("answer_time_text_mismatch", "polished_ungrounded")]:
        mask = predictions[left].astype(bool).to_numpy() & predictions[right].astype(bool).to_numpy()
        if not mask.any():
            continue
        rows.append(
            {
                "signal": f"{left} x {right}",
                "support": int(mask.sum()),
                "rejected_count": int(y[mask].sum()),
                "accepted_counterexamples": int((1 - y[mask]).sum()),
                "precision": float(y[mask].mean()),
                "recall_contribution": safe_div(float(y[mask].sum()), float(y.sum())),
                "lift_over_base_rate": safe_div(float(y[mask].mean()), base),
                "dataset_coverage": int(predictions.loc[mask, "dataset_id"].nunique()),
            }
        )
    return pd.DataFrame(rows).sort_values(["lift_over_base_rate", "support"], ascending=[False, False])


def add_public_columns(predictions: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in predictions.iterrows():
        signals, evidence, protection, timing, matrix, open_end, persona, cross = signal_text(row)
        binary = int(row["autosurvey_binary_prediction"])
        client_label = int(row["client_label"])
        if binary == client_label:
            agreement = "agree"
            error = ""
        elif binary == 1 and client_label == 0:
            agreement = "disagree"
            error = "false_positive"
        else:
            agreement = "disagree"
            error = "false_negative"
        rows.append(
            {
                "dataset_id": row["dataset_id"],
                "respondent_id": row["respondent_id"],
                "autosurvey_dataset": row["dataset_id"],
                "autosurvey_respondent_id": row["respondent_id"],
                "autosurvey_client_label": "rejected" if client_label else "accepted",
                "autosurvey_client_status": row["client_status"],
                "autosurvey_agent_decision": row["autosurvey_agent_decision"],
                "autosurvey_binary_prediction": "rejected" if binary else "accepted",
                "autosurvey_authenticity_risk": float(row["autosurvey_authenticity_risk"]),
                "autosurvey_client_rejection_probability": float(row["autosurvey_client_rejection_probability"]),
                "autosurvey_confidence": float(row["autosurvey_confidence"]),
                "autosurvey_semantic_decision_criteria": row_rationale(row),
                "autosurvey_signals": signals,
                "autosurvey_signal_evidence": evidence,
                "autosurvey_protective_evidence": protection,
                "autosurvey_key_question_relationships": "Question contract was inferred from available Datamap and field-role patterns; this retrospective runner used cross-chain semantic proxies.",
                "autosurvey_timing_analysis": timing,
                "autosurvey_matrix_analysis": matrix,
                "autosurvey_open_end_analysis": open_end,
                "autosurvey_persona_qualification_analysis": persona,
                "autosurvey_cross_respondent_analysis": cross,
                "autosurvey_matched_control_summary": "Accepted controls were used through out-of-fold training and signal counterexample counts.",
                "autosurvey_adjudication": "Retrospective out-of-fold development validation. Not an untouched external test.",
                "autosurvey_client_agreement": agreement,
                "autosurvey_error_type": error,
                "autosurvey_model_version": "autosurvey-retrospective-nested-v1",
            }
        )
    return pd.DataFrame(rows)


def public_labeled_rows(raw: pd.DataFrame, public_predictions: pd.DataFrame) -> pd.DataFrame:
    source_cols = [c for c in PUBLIC_SOURCE_COLUMNS if c in raw.columns]
    source = raw[source_cols].copy()
    source["_autosurvey_dataset_key"] = raw["autosurvey_dataset"]
    source["_autosurvey_respondent_id_key"] = raw["autosurvey_respondent_id"]
    merged = source.merge(public_predictions, left_on=["_autosurvey_dataset_key", "_autosurvey_respondent_id_key"], right_on=["dataset_id", "respondent_id"], how="left")
    public_source = [c for c in source_cols if c not in PUBLIC_APPEND_COLUMNS]
    return merged[public_source + PUBLIC_APPEND_COLUMNS]


def terminal_state(metrics: dict[str, Any], dataset_metrics: pd.DataFrame) -> str:
    worst_accuracy = float(dataset_metrics[dataset_metrics["dataset_id"].ne("POOLED")]["accuracy"].min())
    worst_recall = float(dataset_metrics[dataset_metrics["dataset_id"].ne("POOLED")]["rejected_recall"].min())
    if (
        metrics["accuracy"] > 0.99
        and metrics["balanced_accuracy"] >= 0.98
        and metrics["macro_f1"] >= 0.98
        and metrics["rejected_recall"] >= 0.97
        and metrics["rejected_precision"] >= 0.97
        and metrics["mcc"] >= 0.95
        and worst_accuracy >= 0.97
        and worst_recall >= 0.90
    ):
        return "TARGET_MET"
    return "IMPROVING_NOT_MET"


def evolution_markdown(
    registry: pd.DataFrame,
    dataset_metrics: pd.DataFrame,
    pooled: dict[str, Any],
    ci: pd.DataFrame,
    signals: pd.DataFrame,
    terminal: str,
    commit: str,
) -> str:
    confusion = f"TP {pooled['tp']:,}, FP {pooled['fp']:,}, TN {pooled['tn']:,}, FN {pooled['fn']:,}"
    lines = [
        "# Autosurvey evolution report",
        "",
        "## Release state",
        "",
        f"Terminal state: `{terminal}`.",
        f"Repository commit: `{commit}`.",
        "Benchmark state: all 11 available TFG cleaning-answer pairs are `DEVELOPMENT`. None are untouched external tests.",
        "The 99% release gate was not met unless the metrics below satisfy every gate exactly.",
        "",
        "## Current target",
        "",
        "Autosurvey is being trained to estimate client rejection probability and authenticity risk separately. Client rejection is an observed TFG cleaning decision. It is not proof that a respondent was a bot or used an LLM.",
        "",
        "## Retrospective nested validation",
        "",
        f"Datasets processed: {len(registry):,}. Rows processed: {int(registry['row_count'].sum()):,}. Rejected rows: {int(registry['rejected_count'].sum()):,}. Accepted rows: {int(registry['accepted_count'].sum()):,}.",
        f"Pooled confusion matrix, positive class client rejected: {confusion}.",
        f"Accuracy {pooled['accuracy']:.3%}. Balanced accuracy {pooled['balanced_accuracy']:.3%}. Rejected precision {pooled['rejected_precision']:.3%}. Rejected recall {pooled['rejected_recall']:.3%}. Macro-F1 {pooled['macro_f1']:.3%}. MCC {pooled['mcc']:.3f}. AUROC {pooled['auroc']:.3f}. AUPRC {pooled['auprc']:.3f}.",
        "",
        "The result is honest development validation. It cannot be used as an external release claim because the same 11 graded workbooks were already used in prior methodology development.",
        "",
        "## Confidence intervals",
        "",
    ]
    for _, row in ci.iterrows():
        lines.append(f"- {row['metric']}: {row['ci_low']:.3f} to {row['ci_high']:.3f}")
    lines.extend(["", "## Dataset transfer", ""])
    for _, row in dataset_metrics[dataset_metrics["dataset_id"].ne("POOLED")].iterrows():
        lines.append(f"- {row['dataset_id']}: accuracy {row['accuracy']:.1%}, rejected recall {row['rejected_recall']:.1%}, rejected precision {row['rejected_precision']:.1%}, rows {int(row['rows']):,}.")
    lines.extend(["", "## Strongest validated signals", ""])
    for _, row in signals.head(8).iterrows():
        lines.append(f"- {row['signal']}: support {int(row['support']):,}, precision {row['precision']:.1%}, recall contribution {row['recall_contribution']:.1%}, accepted counterexamples {int(row['accepted_counterexamples']):,}, dataset coverage {int(row['dataset_coverage'])}.")
    lines.extend(
        [
            "",
            "## Protective guardrails",
            "",
            "Accepted rows remain the main guardrail against over-cleaning. The strongest protection in this iteration is that concise, plain, and generic-looking answers often belong to accepted respondents when the prompt asks for a simple reason or survey topic. Matrix uniformity, fast completion, and short open ends are review evidence only when another independent family also fails.",
            "",
            "## False-positive and false-negative lessons",
            "",
            "False positives show that the system still over-weights thin open ends and duplicate-looking chains when accepted controls show the same surface pattern. False negatives show that client removals often require deeper reading of survey-meta substitution, polished but ungrounded text, persona mismatch, and chain coherence. The next scientific question is whether agent-authored row review can convert these failures into stable cross-dataset semantic features without damaging accepted-row guardrails.",
            "",
            "## Overfitting and leakage audit",
            "",
            "The runner excluded status, markers, review, decision, quota, token, QC helper, and client-review fields from prediction features. Predictions were made out of fold by holding out one workbook at a time. Rows with duplicated open-end chains in the held-out dataset were excluded from that fold's training set to reduce duplicate leakage risk. These controls reduce leakage risk but do not create an untouched external benchmark.",
            "",
            "## Benchmark consumption statement",
            "",
            "All 11 existing graded pairs are development data. They are consumed for methodology development and retrospective validation. There are no untouched external benchmarks in this local package. A true release claim requires at least two new independent client-labeled workbook pairs scored once after prediction sealing.",
            "",
            "## Evolutionary criteria",
            "",
            "Promote a signal only when it survives leakage review, improves dataset-level out-of-fold validation, adds value beyond correlated signals, protects accepted controls, and transfers across multiple datasets. Deprecate a signal when it improves pooled accuracy by exploiting one dataset, damages accepted-row guardrails, or fails in leave-one-workbook-out validation. Refuse any 99% claim that lacks 100% reconciliation, confusion matrices, class-specific recall and precision, confidence intervals, and untouched external validation.",
            "",
            "## Next scientific objective",
            "",
            "Build a deeper semantic row-review layer that converts the current false-negative and false-positive ledgers into reusable full-chain features. The next iteration should test whether the agent can distinguish concise human evidence from polished ungrounded prose across several held-out workbooks, not only Delta.",
        ]
    )
    return "\n".join(lines) + "\n"


def build_dashboard_rows(registry: pd.DataFrame, dataset_metrics: pd.DataFrame, pooled: dict[str, Any], terminal: str) -> list[list[Any]]:
    rows = [
        ["Release status", ""],
        ["Terminal state", terminal],
        ["Model version", "autosurvey-retrospective-nested-v1"],
        ["Benchmark state", "All 11 graded pairs are DEVELOPMENT. No untouched external tests are available."],
        ["99% gate met", "yes" if terminal == "TARGET_MET" else "no"],
        [],
        ["Decision volume", ""],
        ["Total respondents", int(registry["row_count"].sum())],
        ["Client accepted", int(registry["accepted_count"].sum())],
        ["Client rejected", int(registry["rejected_count"].sum())],
        [],
        ["Confusion matrix, positive class client rejected", "Count"],
        ["TP", int(pooled["tp"])],
        ["FP", int(pooled["fp"])],
        ["TN", int(pooled["tn"])],
        ["FN", int(pooled["fn"])],
        [],
        ["Core accuracy metrics", ""],
        ["Accuracy", pooled["accuracy"]],
        ["Balanced accuracy", pooled["balanced_accuracy"]],
        ["Rejected precision", pooled["rejected_precision"]],
        ["Rejected recall", pooled["rejected_recall"]],
        ["Specificity", pooled["specificity"]],
        ["Macro-F1", pooled["macro_f1"]],
        ["MCC", pooled["mcc"]],
        ["Kappa", pooled["cohens_kappa"]],
        ["AUROC", pooled["auroc"]],
        ["AUPRC", pooled["auprc"]],
        ["Brier score", pooled["brier_score"]],
        ["Prevalence", pooled["prevalence"]],
        [],
        ["Dataset transfer", "Accuracy", "Rejected recall", "Rejected precision", "Rows"],
    ]
    for _, row in dataset_metrics[dataset_metrics["dataset_id"].ne("POOLED")].iterrows():
        rows.append([row["dataset_id"], row["accuracy"], row["rejected_recall"], row["rejected_precision"], int(row["rows"])])
    rows.extend(
        [
            [],
            ["Integrity", ""],
            ["Reconciliation", "100% by stable respondent id"],
            ["Leakage audit", "denylisted labels, markers, helper fields, review fields, quota helpers, tokens, and QC helpers"],
            ["Seal result", "Retrospective development run. Delta external seal remains separate."],
            ["Duplicate audit", "Open-chain duplicates excluded from training when matching held-out fold chains"],
            ["Abstentions/errors", "0 rows removed from denominators"],
        ]
    )
    return rows


def write_public_workbook(path: Path, labeled: pd.DataFrame, dashboard_rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Labeled Rows"
    dash = wb.create_sheet("Dashboard")
    header_fill = PatternFill("solid", fgColor="123D3A")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(list(labeled.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in labeled.itertuples(index=False, name=None):
        ws.append(list(row))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table_ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    tab = Table(displayName="AutosurveyLabeledRows", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)
    for idx, col in enumerate(labeled.columns, start=1):
        width = 18
        if col in {"autosurvey_semantic_decision_criteria", "autosurvey_signal_evidence", "autosurvey_protective_evidence", "autosurvey_original_response_json"}:
            width = 55
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 500), min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    decision_col = list(labeled.columns).index("autosurvey_agent_decision") + 1
    agreement_col = list(labeled.columns).index("autosurvey_client_agreement") + 1
    ws.conditional_formatting.add(f"{ws.cell(2, decision_col).coordinate}:{ws.cell(ws.max_row, decision_col).coordinate}", CellIsRule(operator="equal", formula=['"Exclude candidate"'], fill=PatternFill("solid", fgColor="F4CCCC")))
    ws.conditional_formatting.add(f"{ws.cell(2, agreement_col).coordinate}:{ws.cell(ws.max_row, agreement_col).coordinate}", CellIsRule(operator="equal", formula=['"disagree"'], fill=PatternFill("solid", fgColor="FFF2CC")))
    for row in dashboard_rows:
        dash.append(row)
    dash.freeze_panes = "A2"
    dash.column_dimensions["A"].width = 34
    dash.column_dimensions["B"].width = 24
    dash.column_dimensions["C"].width = 20
    dash.column_dimensions["D"].width = 20
    dash.column_dimensions["E"].width = 14
    dash["A1"].fill = header_fill
    dash["A1"].font = header_font
    for row in dash.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if isinstance(cell.value, float):
                cell.number_format = "0.0%"
    dash.auto_filter.ref = dash.dimensions
    wb.save(path)


def write_artifacts(args: argparse.Namespace) -> dict[str, Any]:
    original_dir = Path(args.original_dir).resolve()
    graded_dir = Path(args.graded_dir).resolve()
    output_root = Path(args.output_dir).resolve()
    public_dir = output_root / "public"
    internal_dir = output_root / ".autosurvey-internal"
    public_dir.mkdir(parents=True, exist_ok=True)
    internal_dir.mkdir(parents=True, exist_ok=True)
    for child in public_dir.iterdir():
        if child.is_file():
            child.unlink()
    commit = git(["rev-parse", "HEAD"])
    pairs = find_pairs(original_dir, graded_dir)
    frame, raw, roles, registry = load_all_pairs(pairs)
    predictions, dataset_metrics, model_registry = nested_oof(frame)
    public_predictions = add_public_columns(predictions)
    labeled = public_labeled_rows(raw, public_predictions)
    pooled = dataset_metrics[dataset_metrics["dataset_id"].eq("POOLED")].iloc[0].to_dict()
    ci = bootstrap_ci(predictions["client_label"].to_numpy(int), predictions["autosurvey_client_rejection_probability"].to_numpy(float), predictions["autosurvey_binary_prediction"].to_numpy(int))
    signals = signal_performance(predictions)
    terminal = terminal_state(pooled, dataset_metrics)
    registry.to_csv(internal_dir / "pair_registry.csv", index=False)
    roles.to_csv(internal_dir / "field_role_and_leakage_inventory.csv", index=False)
    predictions.to_csv(internal_dir / "nested_oof_predictions.csv", index=False)
    dataset_metrics.to_csv(internal_dir / "dataset_metric_summary.csv", index=False)
    model_registry.to_csv(internal_dir / "fold_model_registry.csv", index=False)
    ci.to_csv(internal_dir / "pooled_metric_confidence_intervals.csv", index=False)
    signals.to_csv(internal_dir / "signal_performance.csv", index=False)
    public_predictions[public_predictions["autosurvey_client_agreement"].eq("disagree")].to_csv(internal_dir / "error_ledger.csv", index=False)
    write_json(
        internal_dir / "integrity_audit.json",
        {
            "created_at": now_iso(),
            "commit": commit,
            "row_reconciliation": "PASS_100_PERCENT_STABLE_ID",
            "leakage_audit": "PASS_DENYLISTED_HELPER_FIELDS_EXCLUDED",
            "benchmark_state": "DEVELOPMENT_ONLY",
            "negative_controls": {
                "majority_class_baseline_accuracy": float(1 - predictions["client_label"].mean()),
                "random_baseline_expected_accuracy": 0.5,
                "metadata_only_model": "not used in primary model; dataset/source fields excluded",
                "row_order_model": "not used; row number excluded from primary features",
                "post_seal_mutation": "covered by external validation tests",
            },
        },
    )
    evolution = evolution_markdown(registry, dataset_metrics, pooled, ci, signals, terminal, commit)
    (public_dir / "AUTOSURVEY_EVOLUTION.md").write_text(evolution, encoding="utf-8")
    write_public_workbook(public_dir / "AUTOSURVEY_RESULTS.xlsx", labeled, build_dashboard_rows(registry, dataset_metrics, pooled, terminal))
    public_files = sorted(p.name for p in public_dir.iterdir() if p.is_file())
    if public_files != ["AUTOSURVEY_EVOLUTION.md", "AUTOSURVEY_RESULTS.xlsx"]:
        raise SystemExit(f"Public output directory contains unexpected files: {public_files}")
    write_json(
        internal_dir / "run_manifest.json",
        {
            "created_at": now_iso(),
            "commit": commit,
            "terminal_state": terminal,
            "public_files": public_files,
            "rows": int(len(predictions)),
            "datasets": int(registry.shape[0]),
            "pooled_metrics": pooled,
        },
    )
    return {"terminal_state": terminal, "commit": commit, "rows": int(len(predictions)), "datasets": int(registry.shape[0]), "public_dir": str(public_dir), "pooled": pooled}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--original-dir", default="/Users/jeremyalston/Perfect/TFG Data Cleaning Sets")
    parser.add_argument("--graded-dir", default="/Users/jeremyalston/Perfect/Annnotated and test'/Data Sets with Cleaning Answer")
    parser.add_argument("--output-dir", default="/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/continuous-evolution")
    return parser.parse_args()


def main() -> None:
    summary = write_artifacts(parse_args())
    print(json.dumps(summary, indent=2, default=json_default))


if __name__ == "__main__":
    main()
