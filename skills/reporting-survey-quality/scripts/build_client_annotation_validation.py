#!/usr/bin/env python3
"""Compare autosurvey artifacts against a client annotated workbook."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook


REQUIRED_CLIENT_COLUMNS = [
    "Respondent Flags",
    "Respondent Score",
    "Recommended_Action",
]


FLAG_FAMILIES: list[tuple[str, Callable[[dict[str, object]], bool]]] = [
    ("client_any_flag", lambda row: text(row.get("Respondent Flags")) not in {"", "No concerns"}),
    (
        "client_light_or_close_review",
        lambda row: text(row.get("Recommended_Action")) in {"Light review", "Review closely"},
    ),
    ("client_review_closely", lambda row: text(row.get("Recommended_Action")) == "Review closely"),
    ("client_qtime_under_4_minutes", lambda row: text(row.get("qtime_Under_4_Minutes")) == "Yes"),
    (
        "client_preferred_brand_inconsistency",
        lambda row: text(row.get("Preferred_Brand_Inconsistent_With_Consideration_Recommendation")) == "Yes",
    ),
    ("client_q32_straightline", lambda row: text(row.get("Q32_Straightline")) == "Yes"),
    (
        "client_outro_off_topic",
        lambda row: "off-topic" in text(row.get("outro_Topic_Relevance")).lower()
        or "outro off-topic" in text(row.get("Respondent Flags")).lower(),
    ),
    (
        "client_moderate_ai_open_end",
        lambda row: "moderate ai/open-end concern" in text(row.get("Respondent Flags")).lower(),
    ),
    (
        "client_high_ai_open_end",
        lambda row: "high ai/open-end concern" in text(row.get("Respondent Flags")).lower(),
    ),
]


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def pct(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return "0.0%"
    return f"{(numerator / denominator) * 100:.1f}%"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--client-workbook", type=Path, required=True)
    parser.add_argument("--sheet", default="A1")
    parser.add_argument("--id-column", default="uuid")
    return parser.parse_args()


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path) if path.exists() else pd.DataFrame()


def client_rows(path: Path, sheet: str, id_column: str) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    if sheet not in workbook.sheetnames:
        raise SystemExit(f"Sheet {sheet!r} was not found in {path}")
    ws = workbook[sheet]
    headers = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]
    index = {text(header): column + 1 for column, header in enumerate(headers) if text(header)}
    missing = [column for column in REQUIRED_CLIENT_COLUMNS + [id_column] if column not in index]
    if missing:
        raise SystemExit(f"Client workbook is missing required columns: {', '.join(missing)}")

    rows: list[dict[str, object]] = []
    for row_index in range(2, ws.max_row + 1):
        key = text(ws.cell(row_index, index[id_column]).value)
        if not key:
            continue
        row = {"client_row_number": row_index, "respondent_key": key}
        for header, column in index.items():
            row[header] = ws.cell(row_index, column).value
        rows.append(row)

    summary: dict[str, object] = {}
    if "Cleaning Summary" in workbook.sheetnames:
        summary_ws = workbook["Cleaning Summary"]
        for row_index in range(1, summary_ws.max_row + 1):
            key = text(summary_ws.cell(row_index, 1).value)
            value = summary_ws.cell(row_index, 2).value
            if key:
                summary[key] = value
    return rows, summary


def indexed(df: pd.DataFrame) -> dict[str, dict[str, object]]:
    if df.empty or "respondent_key" not in df.columns:
        return {}
    rows: dict[str, dict[str, object]] = {}
    for row in df.to_dict(orient="records"):
        key = text(row.get("respondent_key"))
        if key and key not in rows:
            rows[key] = row
    return rows


def detect_discard_claims(run_dir: Path, discard_count: int) -> list[dict[str, object]]:
    claims: list[dict[str, object]] = []
    patterns = [
        (re.compile(r"\bzero final discards\b|\bno rows should be escalated\b|\bdiscard set is empty\b", re.I), 0),
        (re.compile(r"\brecommend(?:ed)?\s+0\s+rows\b|\b0 rows? (?:are )?recommended\b", re.I), 0),
        (re.compile(r"\brecommend(?:ed)?\s+(\d+)\s+rows?\b", re.I), None),
        (re.compile(r"\b(\d+)\s+rows? (?:are )?recommended for exclusion review\b", re.I), None),
    ]
    for name in [
        "agent_findings_essay.md",
        "agent_escalation_packet.md",
        "agent_positive_insights_report.md",
        "deep_findings_analysis.md",
        "agent_final_visual_findings_report.md",
    ]:
        path = run_dir / name
        if not path.exists():
            continue
        body = path.read_text(encoding="utf-8", errors="replace")
        file_claims: list[int] = []
        for pattern, fixed in patterns:
            for match in pattern.finditer(body):
                if fixed is not None:
                    file_claims.append(fixed)
                else:
                    file_claims.append(int(match.group(1)))
        unique_claims = sorted(set(file_claims))
        claims.append(
            {
                "artifact": name,
                "discard_count_claims": unique_claims,
                "matches_discard_set": (not unique_claims) or unique_claims == [discard_count],
            }
        )
    return claims


def main() -> None:
    args = parse_args()
    run_dir = args.run_dir.expanduser().resolve()
    client_workbook = args.client_workbook.expanduser().resolve()
    clients, workbook_summary = client_rows(client_workbook, args.sheet, args.id_column)

    row_scores = indexed(read_csv(run_dir / "row_scores.csv"))
    judgments = indexed(read_csv(run_dir / "agent_review_judgment_table.csv"))
    discards = indexed(read_csv(run_dir / "agent_discard_set.csv"))

    validation_rows: list[dict[str, object]] = []
    for row in clients:
        key = text(row.get("respondent_key"))
        scored = row_scores.get(key, {})
        judgment = judgments.get(key, {})
        discard = discards.get(key, {})
        family_hits = [name for name, predicate in FLAG_FAMILIES if predicate(row)]
        validation_rows.append(
            {
                "respondent_key": key,
                "client_row_number": row.get("client_row_number"),
                "client_recommended_action": text(row.get("Recommended_Action")),
                "client_respondent_flags": text(row.get("Respondent Flags")),
                "client_respondent_score": row.get("Respondent Score"),
                "client_flag_families": "; ".join(family_hits),
                "autosurvey_computed_action": text(scored.get("computed_action")),
                "autosurvey_second_pass_decision": text(scored.get("second_pass_decision")),
                "autosurvey_reviewed_by_agent": key in judgments,
                "autosurvey_agent_final_decision": text(judgment.get("agent_final_decision")),
                "autosurvey_review_theme": text(judgment.get("review_theme")),
                "autosurvey_in_discard_set": key in discards,
                "validation_note": validation_note(row, key, judgments, discards),
            }
        )

    validation = pd.DataFrame(validation_rows)
    validation.to_csv(run_dir / "client_annotation_validation.csv", index=False)

    family_summary = []
    for name, predicate in FLAG_FAMILIES:
        keys = [text(row.get("respondent_key")) for row in clients if predicate(row)]
        reviewed = [key for key in keys if key in judgments]
        discarded = [key for key in keys if key in discards]
        missed = [key for key in keys if key not in judgments]
        family_summary.append(
            {
                "family": name,
                "client_rows": len(keys),
                "autosurvey_reviewed": len(reviewed),
                "autosurvey_review_coverage": pct(len(reviewed), len(keys)),
                "autosurvey_discarded": len(discarded),
                "missed_sample": missed[:12],
            }
        )

    action_counts = Counter(text(row.get("Recommended_Action")) for row in clients)
    flag_counts = Counter(text(row.get("Respondent Flags")) for row in clients)
    over_reviewed = [
        key
        for key in judgments
        if text(next((row for row in clients if text(row.get("respondent_key")) == key), {}).get("Respondent Flags"))
        == "No concerns"
    ]
    client_keep_discards = [
        key
        for key in discards
        if text(next((row for row in clients if text(row.get("respondent_key")) == key), {}).get("Recommended_Action"))
        == "Keep"
    ]
    artifact_claims = detect_discard_claims(run_dir, len(discards))
    summary = {
        "client_workbook": str(client_workbook),
        "client_sheet": args.sheet,
        "client_rows": len(clients),
        "client_summary": workbook_summary,
        "client_action_counts": dict(action_counts),
        "top_client_flag_counts": dict(flag_counts.most_common(20)),
        "autosurvey_scored_rows": len(row_scores),
        "autosurvey_agent_reviewed_rows": len(judgments),
        "autosurvey_discard_rows": len(discards),
        "family_summary": family_summary,
        "autosurvey_reviewed_client_no_concerns": len(over_reviewed),
        "autosurvey_reviewed_client_no_concerns_sample": over_reviewed[:20],
        "autosurvey_discard_rows_client_kept": client_keep_discards,
        "artifact_discard_claims": artifact_claims,
    }
    (run_dir / "client_annotation_validation_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    (run_dir / "client_annotation_validation.md").write_text(markdown(summary, validation), encoding="utf-8")
    print(run_dir / "client_annotation_validation.md")


def validation_note(
    client_row: dict[str, object],
    key: str,
    judgments: dict[str, dict[str, object]],
    discards: dict[str, dict[str, object]],
) -> str:
    client_action = text(client_row.get("Recommended_Action"))
    client_flags = text(client_row.get("Respondent Flags"))
    if key in discards and client_action == "Keep":
        return "Autosurvey discard conflicts with client keep baseline. Requires analyst explanation and PM review."
    if client_action in {"Light review", "Review closely"} and key not in judgments:
        return "Client routed this row to review, but autosurvey did not include it in final semantic review."
    if client_flags != "No concerns" and key not in judgments:
        return "Client flagged this row, but autosurvey did not include it in final semantic review."
    if client_flags == "No concerns" and key in judgments:
        return "Autosurvey reviewed this row despite no client concern. Explain the added signal or demote it."
    if key in judgments:
        return "Autosurvey included this row in final semantic review."
    return "No baseline discrepancy."


def markdown(summary: dict[str, object], validation: pd.DataFrame) -> str:
    lines = [
        "# Client annotation validation",
        "",
        f"Client workbook: `{summary['client_workbook']}`",
        f"Client rows: {summary['client_rows']}",
        f"Autosurvey scored rows: {summary['autosurvey_scored_rows']}",
        f"Autosurvey agent-reviewed rows: {summary['autosurvey_agent_reviewed_rows']}",
        f"Autosurvey discard rows: {summary['autosurvey_discard_rows']}",
        "",
        "## Client baseline counts",
        "",
    ]
    for action, count in summary["client_action_counts"].items():
        lines.append(f"- {action}: {count}")
    lines.extend(["", "Top client flags:"])
    for flag, count in summary["top_client_flag_counts"].items():
        lines.append(f"- {flag}: {count}")

    lines.extend(["", "## Autosurvey coverage by client signal", ""])
    for row in summary["family_summary"]:
        lines.append(
            f"- {row['family']}: client rows {row['client_rows']}; "
            f"autosurvey reviewed {row['autosurvey_reviewed']} ({row['autosurvey_review_coverage']}); "
            f"autosurvey discarded {row['autosurvey_discarded']}."
        )
        if row["missed_sample"]:
            lines.append(f"  Missed sample: {', '.join(row['missed_sample'])}")

    lines.extend(
        [
            "",
            "## Validation risks",
            "",
            f"- Autosurvey reviewed {summary['autosurvey_reviewed_client_no_concerns']} rows that the client marked `No concerns`.",
            f"- Autosurvey discard rows with client action `Keep`: {', '.join(summary['autosurvey_discard_rows_client_kept']) or 'none'}.",
            "",
            "## Artifact consistency",
            "",
        ]
    )
    for claim in summary["artifact_discard_claims"]:
        status = "matches" if claim["matches_discard_set"] else "does not match"
        claims = claim["discard_count_claims"] or ["no explicit count found"]
        lines.append(f"- `{claim['artifact']}` {status} discard set count. Claims: {claims}.")

    mismatches = validation[validation["validation_note"].astype(str).ne("No baseline discrepancy.")]
    lines.extend(["", "## Rows needing analyst review", ""])
    if mismatches.empty:
        lines.append("No row-level baseline discrepancies were found.")
    else:
        for _, row in mismatches.head(40).iterrows():
            lines.append(
                f"- {row['respondent_key']} | client {row['client_recommended_action']} | "
                f"autosurvey {row['autosurvey_agent_final_decision'] or row['autosurvey_second_pass_decision']} | "
                f"{row['validation_note']}"
            )
        if len(mismatches) > 40:
            lines.append(f"- +{len(mismatches) - 40} more rows in `client_annotation_validation.csv`.")

    lines.extend(
        [
            "",
            "## Standard",
            "",
            "Autosurvey should match or explain every client review surface before claiming a benchmark pass. It should then go further by adding full-chain reading, field-role mapping, false-positive guardrails, demographic and aggregate insights, and next-pass learning.",
            "",
        ]
    )
    return "\n".join(lines)


if __name__ == "__main__":
    main()
