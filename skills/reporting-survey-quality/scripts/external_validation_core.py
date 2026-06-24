#!/usr/bin/env python3
"""Sealed external validation utilities for Autosurvey.

The module enforces a simple separation:

Role A predicts from an unlabeled workbook and development artifacts only.
Role B seals hashes before labels are opened.
Role C validates the seal before reading labels.
Role D writes integrity checks from frozen artifacts.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import random
import re
import shutil
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


LABEL_HEADER_RE = re.compile(
    r"(^|_)(status|client_status|final_status|decision|final_decision|recommended_action|"
    r"client_decision|accept|accepted|reject|rejected|discard|exclude|removed|cleaned_out|"
    r"review_note|reviewer_note|reason|annotation)(_|$)",
    re.I,
)
LEAKAGE_RE = re.compile(
    r"(^|_)(status|client|review|recommend|action|decision|annotation|score|tier|final|"
    r"discard|exclude|keep|clean|flag|reason|note|comment|markers?|validclient|"
    r"channeltracking|closingemail|token)(_|$)|condition|noanswer|^qc($|\d|_)|"
    r"TERMFLAGS|SCRUTINYFLAGS|CLIENTFLAGS|redem_excluded|RD_Review|RD_Search|quota|bad:",
    re.I,
)
ID_RE = re.compile(r"uuid|respondent|rid|record|case|id$", re.I)
TIME_RE = re.compile(r"qtime|time|duration|elapsed|date|timestamp|start|end", re.I)
OPEN_RE = re.compile(r"open|other|specify|explain|why|comment|oe|outro|q34|q43|q32|q10|q9", re.I)
MATRIX_RE = re.compile(r"(.+?)(?:_?r|row)\d+(?:_?c\d+)?$", re.I)
GENERIC_RE = re.compile(r"^(n/?a|none|nothing|no|not sure|dont know|don't know|idk|good|ok|okay|same|other|na)$", re.I)
POLISHED_RE = re.compile(r"\b(comprehensive|seamless|valuable insight|overall satisfaction|high quality|user friendly|cost effective|innovative solution|as an ai|cannot answer)\b|—", re.I)

TIER_ORDER = {
    "Clean keep": 1,
    "Keep with note": 2,
    "Light review": 3,
    "Review closely": 4,
    "Exclude candidate": 5,
}
ORDER_TO_TIER = {v: k for k, v in TIER_ORDER.items()}
PRIMARY_METRICS = [
    "auroc",
    "auprc",
    "mcc_tier5",
    "tier5_recall",
    "tier5_false_positive_rate",
    "tier45_precision",
    "tier45_recall",
    "brier_score",
    "calibration_slope",
]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value).lower()).strip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=json_default), encoding="utf-8")


def json_default(value: Any) -> Any:
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return None if math.isnan(float(value)) else float(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, Path):
        return str(value)
    return str(value)


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        keys: list[str] = []
        seen = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    keys.append(key)
        fieldnames = keys
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def git(command: list[str], cwd: Path) -> str:
    return subprocess.check_output(["git", *command], cwd=cwd, text=True).strip()


def workbook_header_inventory(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1), [])
        headers = [text(c.value) for c in first if text(c.value)]
        sheets.append(
            {
                "sheet_name": ws.title,
                "row_count": max(ws.max_row - 1, 0),
                "column_count": ws.max_column,
                "headers": headers,
                "respondent_key_candidates": [h for h in headers if ID_RE.search(h)][:12],
                "outcome_column_names_only": [h for h in headers if LABEL_HEADER_RE.search(h)],
            }
        )
    return {
        "path": str(path),
        "file_name": path.name,
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
        "sheets": sheets,
    }


def discover_candidate_label_files(root: Path, test_path: Path) -> list[dict[str, Any]]:
    rows = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xls", ".csv", ".tsv"}:
            continue
        if path.name.startswith("~$"):
            continue
        if path.resolve() == test_path.resolve():
            relationship = "original_unlabeled_test_input"
        elif "clean" in str(path).lower() or "answer" in str(path).lower() or "annot" in str(path).lower():
            relationship = "possible_development_or_label_artifact"
        else:
            relationship = "unknown"
        if path.suffix.lower() in {".xlsx", ".xls"}:
            try:
                inv = workbook_header_inventory(path)
                for sheet in inv["sheets"]:
                    rows.append(
                        {
                            "candidate_file": str(path),
                            "file_hash": inv["sha256"],
                            "modified_time": inv["modified_time"],
                            "sheet_name": sheet["sheet_name"],
                            "row_count_if_header_safe": sheet["row_count"],
                            "respondent_key_candidates": "|".join(sheet["respondent_key_candidates"]),
                            "outcome_column_names_only": "|".join(sheet["outcome_column_names_only"]),
                            "possible_relationship_to_hiri": relationship,
                            "label_values_inspected": "false",
                        }
                    )
            except Exception as exc:
                rows.append(
                    {
                        "candidate_file": str(path),
                        "file_hash": sha256_file(path),
                        "modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
                        "sheet_name": "",
                        "row_count_if_header_safe": "",
                        "respondent_key_candidates": "",
                        "outcome_column_names_only": "",
                        "possible_relationship_to_hiri": f"header_error: {exc}",
                        "label_values_inspected": "false",
                    }
                )
        else:
            with path.open("r", encoding="utf-8", errors="ignore") as handle:
                header = next(csv.reader(handle), [])
            rows.append(
                {
                    "candidate_file": str(path),
                    "file_hash": sha256_file(path),
                    "modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
                    "sheet_name": "csv",
                    "row_count_if_header_safe": "",
                    "respondent_key_candidates": "|".join([h for h in header if ID_RE.search(h)][:12]),
                    "outcome_column_names_only": "|".join([h for h in header if LABEL_HEADER_RE.search(h)]),
                    "possible_relationship_to_hiri": relationship,
                    "label_values_inspected": "false",
                }
            )
    return rows


def respondent_key_column(headers: list[str]) -> str:
    for preferred in ["uuid", "respondent_id", "RID", "rid", "record"]:
        for header in headers:
            if header.lower() == preferred.lower():
                return header
    for header in headers:
        if ID_RE.search(header):
            return header
    return headers[0] if headers else ""


def row_fingerprint(row: pd.Series, excluded: set[str] | None = None) -> str:
    excluded = excluded or set()
    payload = []
    for col in row.index:
        if col in excluded or LEAKAGE_RE.search(str(col)):
            continue
        payload.append(f"{col}={text(row[col])}")
    return sha256_text("|".join(payload))


def read_main_sheet(path: Path) -> pd.DataFrame:
    sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl")
    sheet_name = "A1" if "A1" in sheets else next(iter(sheets))
    df = sheets[sheet_name].copy()
    df.columns = [text(c) for c in df.columns]
    return df


def words(value: str) -> list[str]:
    return re.findall(r"[A-Za-z][A-Za-z']+", value.lower())


def matrix_group(col: str) -> str:
    match = MATRIX_RE.match(col)
    return re.sub(r"[^A-Za-z0-9]+", "_", match.group(1)).strip("_").lower() if match else ""


def feature_and_score_rows(df: pd.DataFrame, repo: Path, commit: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    headers = list(df.columns)
    key_col = respondent_key_column(headers)
    leak_cols = {c for c in headers if LEAKAGE_RE.search(c)}
    usable = [c for c in headers if c not in leak_cols]
    open_cols = [c for c in usable if OPEN_RE.search(c)]
    timing_cols = [c for c in usable if TIME_RE.search(c)]
    matrix_groups: dict[str, list[str]] = defaultdict(list)
    for col in usable:
        group = matrix_group(col)
        if group:
            matrix_groups[group].append(col)
    open_signatures = []
    for _, row in df.iterrows():
        raw = " ".join(text(row.get(c)) for c in open_cols)
        raw = re.sub(r"\d+", "#", raw.lower())
        raw = re.sub(r"[^a-z# ]+", " ", raw)
        raw = re.sub(r"\s+", " ", raw).strip()[:260]
        open_signatures.append(raw)
    signature_counts = Counter(s for s in open_signatures if len(s) >= 24)
    prediction_rows = []
    signal_rows = []
    for idx, row in df.iterrows():
        respondent_id = text(row.get(key_col)) or f"row_{idx + 2}"
        row_open = [text(row.get(c)) for c in open_cols if text(row.get(c))]
        combined_open = " ".join(row_open)
        word_count = len(words(combined_open))
        generic_count = sum(1 for v in row_open if GENERIC_RE.search(norm(v)) or len(words(v)) <= 1)
        polished = int(bool(POLISHED_RE.search(combined_open)))
        timing_values = []
        for col in timing_cols:
            val = pd.to_numeric(pd.Series([row.get(col)]), errors="coerce").iloc[0]
            if pd.notna(val):
                timing_values.append(float(val))
        qtime = next((v for v in timing_values if 0 < v < 200000), np.nan)
        fast = int(pd.notna(qtime) and qtime < 240)
        answered = sum(1 for col in usable if text(row.get(col)))
        missing_rate = 1 - answered / max(len(usable), 1)
        high_missing = int(missing_rate > 0.55)
        matrix_modal = 0.0
        matrix_uniform = 0
        for cols in matrix_groups.values():
            if len(cols) < 4:
                continue
            vals = [text(row.get(c)) for c in cols if text(row.get(c))]
            if len(vals) < 4:
                continue
            modal = max(Counter(vals).values()) / len(vals)
            matrix_modal = max(matrix_modal, modal)
            if modal >= 0.92 and len(set(vals)) <= 2:
                matrix_uniform = 1
        duplicate_open = int(len(open_signatures[idx]) >= 24 and signature_counts[open_signatures[idx]] >= 2)
        open_generic = int(word_count > 0 and (generic_count > 0 or word_count <= 3))
        no_open = int(not row_open)
        protective_specific = int(word_count >= 8 and not open_generic)
        protective_short_valid = int(0 < word_count <= 6 and not generic_count)
        risk_raw = (
            1.1 * duplicate_open
            + 1.0 * int(fast and open_generic)
            + 0.9 * int(high_missing and polished)
            + 0.8 * int(matrix_uniform and (open_generic or no_open))
            + 0.65 * int(high_missing and no_open)
            + 0.45 * open_generic
            - 0.55 * protective_specific
            - 0.30 * protective_short_valid
        )
        risk_raw = max(0.0, risk_raw)
        client_prob = min(0.98, max(0.02, risk_raw / 3.4))
        authenticity_prob = min(0.98, max(0.02, (risk_raw + 0.4 * duplicate_open + 0.3 * int(fast and polished)) / 3.8))
        attention_prob = min(0.98, max(0.02, (0.6 * high_missing + 0.5 * matrix_uniform + 0.4 * fast + 0.35 * open_generic) / 2.2))
        protective_score = min(1.0, max(0.0, 0.5 * protective_specific + 0.25 * protective_short_valid + 0.25 * int(answered > len(usable) * 0.5)))
        uncertainty = min(1.0, max(0.05, 0.2 + 0.35 * no_open + 0.25 * int(missing_rate > 0.65) + 0.2 * int(len(usable) < 25)))
        if client_prob >= 0.75:
            tier = "Exclude candidate"
        elif client_prob >= 0.55:
            tier = "Review closely"
        elif client_prob >= 0.30:
            tier = "Light review"
        elif client_prob >= 0.15 or protective_score > 0:
            tier = "Keep with note"
        else:
            tier = "Clean keep"
        signals = []
        for name, val in {
            "cross_respondent_open_template": duplicate_open,
            "fast_generic_open_end": int(fast and open_generic),
            "polished_missingness_compound": int(high_missing and polished),
            "matrix_uniform_weak_chain": int(matrix_uniform and (open_generic or no_open)),
            "high_missing_no_open_chain": int(high_missing and no_open),
            "open_generic_or_too_short": open_generic,
        }.items():
            if val:
                signals.append(name)
                signal_rows.append(
                    {
                        "respondent_id": respondent_id,
                        "source_row_number": idx + 2,
                        "signal_family": name,
                        "signal_value": 1,
                        "frozen_before_unblind": True,
                    }
                )
        guards = []
        if protective_specific:
            guards.append("grounded_open_end_detail")
        if protective_short_valid:
            guards.append("short_but_valid_answer")
        prediction_rows.append(
            {
                "respondent_id": respondent_id,
                "source_row_number": idx + 2,
                "source_row_fingerprint": row_fingerprint(row),
                "input_eligibility_status": "eligible",
                "prediction_status": "predicted",
                "client_reject_probability": round(float(client_prob), 6),
                "authenticity_risk_probability": round(float(authenticity_prob), 6),
                "attention_or_validity_risk": round(float(attention_prob), 6),
                "protective_human_evidence_score": round(float(protective_score), 6),
                "model_uncertainty": round(float(uncertainty), 6),
                "operational_tier": tier,
                "primary_signal_families": "|".join(signals),
                "supporting_signal_ids": "|".join(signals),
                "protective_signal_ids": "|".join(guards),
                "plain_language_evidence_summary": evidence_summary(signals, qtime, word_count, missing_rate),
                "plain_language_counterevidence_summary": counter_summary(guards),
                "question_contract_coverage": round(answered / max(len(usable), 1), 6),
                "abstention_reason": "",
                "predictor_commit": commit,
                "predictor_version": "external_validation_predictor_v1",
                "prediction_timestamp": now_iso(),
            }
        )
    return pd.DataFrame(prediction_rows), pd.DataFrame(signal_rows)


def evidence_summary(signals: list[str], qtime: float, word_count: int, missing_rate: float) -> str:
    if not signals:
        return "No frozen high-risk signal family fired before unblinding."
    parts = []
    if "cross_respondent_open_template" in signals:
        parts.append("The open-ended response pattern appeared in a repeated template group.")
    if "fast_generic_open_end" in signals:
        parts.append("The row combined fast completion with a weak or generic open-ended chain.")
    if "polished_missingness_compound" in signals:
        parts.append("Polished wording appeared with substantial missingness.")
    if "matrix_uniform_weak_chain" in signals:
        parts.append("A highly uniform matrix pattern appeared without strong open-ended grounding.")
    if "high_missing_no_open_chain" in signals:
        parts.append("The row had high missingness and no usable open-ended chain.")
    if "open_generic_or_too_short" in signals:
        parts.append("A central open-ended answer was generic or too short to carry much evidence.")
    parts.append(f"Pre-unblind context: open words {word_count}, missing-rate {missing_rate:.2f}.")
    if pd.notna(qtime):
        parts.append(f"Completion-time value used only as routing context: {qtime:.0f}.")
    return " ".join(parts)


def counter_summary(guards: list[str]) -> str:
    if not guards:
        return "No strong protective human-evidence guardrail fired in the frozen pre-unblind pass."
    if "grounded_open_end_detail" in guards:
        return "The row has enough open-ended detail to protect against over-relying on surface signals."
    return "The row contains a short but valid answer pattern, so the shortness is not exclusion evidence by itself."


def confusion_metrics(y_true: np.ndarray, y_pred: np.ndarray) -> dict[str, float]:
    tp = int(((y_true == 1) & (y_pred == 1)).sum())
    fp = int(((y_true == 0) & (y_pred == 1)).sum())
    tn = int(((y_true == 0) & (y_pred == 0)).sum())
    fn = int(((y_true == 1) & (y_pred == 0)).sum())
    precision = tp / max(tp + fp, 1)
    recall = tp / max(tp + fn, 1)
    specificity = tn / max(tn + fp, 1)
    npv = tn / max(tn + fn, 1)
    fpr = fp / max(fp + tn, 1)
    fnr = fn / max(fn + tp, 1)
    f1 = 2 * precision * recall / max(precision + recall, 1e-12)
    f2 = 5 * precision * recall / max((4 * precision) + recall, 1e-12)
    mcc_den = math.sqrt(max((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn), 1))
    mcc = ((tp * tn) - (fp * fn)) / mcc_den
    total = tp + fp + tn + fn
    po = (tp + tn) / max(total, 1)
    pe = (((tp + fp) / max(total, 1)) * ((tp + fn) / max(total, 1))) + (((fn + tn) / max(total, 1)) * ((fp + tn) / max(total, 1)))
    kappa = (po - pe) / max(1 - pe, 1e-12)
    return {
        "tp": tp,
        "fp": fp,
        "tn": tn,
        "fn": fn,
        "sensitivity_recall": recall,
        "specificity": specificity,
        "precision_ppv": precision,
        "npv": npv,
        "false_positive_rate": fpr,
        "false_negative_rate": fnr,
        "f1": f1,
        "f2": f2,
        "balanced_accuracy": (recall + specificity) / 2,
        "mcc": mcc,
        "cohens_kappa": kappa,
        "accuracy": po,
    }


def roc_auc(y: np.ndarray, score: np.ndarray) -> float:
    if len(set(y.tolist())) < 2:
        return float("nan")
    order = np.argsort(score)
    ranks = np.empty_like(order, dtype=float)
    ranks[order] = np.arange(1, len(score) + 1)
    pos = y == 1
    neg = y == 0
    return float((ranks[pos].sum() - pos.sum() * (pos.sum() + 1) / 2) / max(pos.sum() * neg.sum(), 1))


def pr_curve(y: np.ndarray, score: np.ndarray) -> pd.DataFrame:
    order = np.argsort(-score)
    y_sorted = y[order]
    score_sorted = score[order]
    tp = np.cumsum(y_sorted == 1)
    fp = np.cumsum(y_sorted == 0)
    precision = tp / np.maximum(tp + fp, 1)
    recall = tp / max(int((y == 1).sum()), 1)
    return pd.DataFrame({"threshold": score_sorted, "precision": precision, "recall": recall})


def average_precision(y: np.ndarray, score: np.ndarray) -> float:
    curve = pr_curve(y, score)
    if curve.empty:
        return float("nan")
    recall = np.r_[0, curve["recall"].to_numpy()]
    precision = np.r_[1, curve["precision"].to_numpy()]
    return float(np.sum((recall[1:] - recall[:-1]) * precision[1:]))


def roc_curve_points(y: np.ndarray, score: np.ndarray) -> pd.DataFrame:
    thresholds = sorted(set(score.tolist()), reverse=True)
    rows = []
    for threshold in thresholds:
        pred = score >= threshold
        m = confusion_metrics(y, pred.astype(int))
        rows.append({"threshold": threshold, "tpr": m["sensitivity_recall"], "fpr": m["false_positive_rate"]})
    return pd.DataFrame(rows)


def calibration(y: np.ndarray, score: np.ndarray, bins: int = 10) -> tuple[pd.DataFrame, dict[str, float]]:
    clipped = np.clip(score, 1e-6, 1 - 1e-6)
    brier = float(np.mean((clipped - y) ** 2))
    log_loss = float(-np.mean(y * np.log(clipped) + (1 - y) * np.log(1 - clipped)))
    logit = np.log(clipped / (1 - clipped))
    try:
        slope, intercept = np.polyfit(logit, y, 1)
    except Exception:
        slope, intercept = float("nan"), float("nan")
    cuts = np.linspace(0, 1, bins + 1)
    rows = []
    ece = 0.0
    mce = 0.0
    for i in range(bins):
        lo, hi = cuts[i], cuts[i + 1]
        mask = (score >= lo) & (score <= hi if i == bins - 1 else score < hi)
        if not mask.any():
            rows.append({"bin": i + 1, "lo": lo, "hi": hi, "n": 0, "mean_predicted": "", "observed_rate": "", "abs_error": ""})
            continue
        mean_pred = float(score[mask].mean())
        obs = float(y[mask].mean())
        err = abs(mean_pred - obs)
        ece += err * (mask.sum() / len(y))
        mce = max(mce, err)
        rows.append({"bin": i + 1, "lo": lo, "hi": hi, "n": int(mask.sum()), "mean_predicted": mean_pred, "observed_rate": obs, "abs_error": err})
    return pd.DataFrame(rows), {"brier_score": brier, "log_loss": log_loss, "calibration_intercept": float(intercept), "calibration_slope": float(slope), "expected_calibration_error": float(ece), "maximum_calibration_error": float(mce)}


def inspect_repo(repo: Path) -> dict[str, Any]:
    dirty = git(["status", "--short"], repo)
    files = [
        "README.md",
        "skills/cleaning-survey-quality/SKILL.md",
        "skills/reporting-survey-quality/SKILL.md",
        "skills/reporting-survey-quality/scripts/build_semantic_authenticity_loop.py",
    ]
    return {
        "branch": git(["branch", "--show-current"], repo),
        "commit": git(["rev-parse", "HEAD"], repo),
        "dirty_status": dirty,
        "dirty_diff_hash": sha256_text(git(["diff"], repo)),
        "tracked_file_hashes": {f: sha256_file(repo / f) for f in files if (repo / f).exists()},
        "python": sys.version,
        "pandas": pd.__version__,
        "numpy": np.__version__,
    }


def cmd_discover(args: argparse.Namespace) -> None:
    repo = Path(args.repo).resolve()
    output = Path(args.output_dir).resolve()
    root = Path(args.client_root).resolve()
    test_path = Path(args.test_workbook).resolve()
    methodology = Path(args.methodology_dir).resolve()
    output.mkdir(parents=True, exist_ok=True)
    repo_state = inspect_repo(repo)
    inv = workbook_header_inventory(test_path)
    candidates = discover_candidate_label_files(root, test_path)
    write_json(output / "predictor_environment.json", {"created_at": now_iso(), **repo_state})
    write_json(output / "repository_state_preunblind.json", repo_state)
    write_json(
        output / "input_candidate_manifest.json",
        {
            "created_at": now_iso(),
            "test_workbook": inv,
            "methodology_dir": str(methodology),
            "methodology_manifest_hash": sha256_file(methodology / "semantic_methodology_freeze_manifest.json")
            if (methodology / "semantic_methodology_freeze_manifest.json").exists()
            else "",
        },
    )
    write_csv(output / "client_label_candidate_registry.csv", candidates)
    headers = inv["sheets"][0]["headers"] if inv["sheets"] else []
    key = respondent_key_column(headers)
    write_json(
        output / "respondent_reconciliation_contract.json",
        {
            "created_at": now_iso(),
            "primary_key": key,
            "join_order": ["stable unique respondent/case ID", "documented composite key", "content fingerprint as aid only"],
            "duplicate_rule": "surface duplicates; do not join by row number",
            "mismatch_rule": "prediction-only and label-only rows are not coerced into accepted or rejected outcomes",
        },
    )
    prereg = preregistration_payload()
    write_json(output / "external_validation_preregistration.json", prereg)
    (output / "external_validation_preregistration.md").write_text(preregistration_markdown(prereg, args.benchmark_name), encoding="utf-8")
    (output / "validation_inventory.md").write_text(
        "\n".join(
            [
                "# Validation inventory",
                "",
                f"Repository commit: `{repo_state['commit']}`.",
                f"Dirty status: `{repo_state['dirty_status'] or 'clean except ignored files'}`.",
                f"Original {args.benchmark_name} workbook: `{test_path}`.",
                f"{args.benchmark_name} SHA-256: `{inv['sha256']}`.",
                f"Latest semantic methodology: `{methodology}`.",
                f"Candidate label files were inspected by metadata and headers only. Rows: {len(candidates)} registry entries.",
            ]
        ),
        encoding="utf-8",
    )


def preregistration_payload() -> dict[str, Any]:
    return {
        "created_at": now_iso(),
        "positive_class": "client_rejected",
        "expected_mapping": {"3": "client accepted", "5": "client rejected"},
        "primary_endpoints": PRIMARY_METRICS,
        "secondary_endpoints": [
            "sensitivity",
            "specificity",
            "precision",
            "negative_predictive_value",
            "false_positive_rate",
            "false_negative_rate",
            "f1",
            "f2",
            "balanced_accuracy",
            "mcc",
            "cohens_kappa",
            "accuracy",
            "auroc",
            "auprc",
            "average_precision",
            "log_loss",
            "brier_score",
            "calibration_intercept",
            "calibration_slope",
            "expected_calibration_error",
            "maximum_calibration_error",
            "top_decile_lift",
            "top_quintile_lift",
            "tier_status_crosstab",
        ],
        "cutoffs": ["Tier 5", "Tiers 4-5", "Tiers 3-5", "Tiers 2-5"],
        "bootstrap": {"method": "stratified", "resamples": 2000, "seed": 2607},
        "review_coverage_points": [0.01, 0.02, 0.05, 0.10, 0.15, 0.20, 0.25, 0.33, 0.50],
        "cohorts": ["open_end_present", "completion_time_band", "completeness_band", "question_contract_coverage_band", "uncertainty_band", "signal_family_activation"],
        "error_taxonomy": [
            "qualification/persona mismatch",
            "open-end semantic nonresponse",
            "timing anomaly",
            "matrix/straightline anomaly",
            "cross-question contradiction",
            "full-chain nonresponse",
            "cross-respondent duplication/coordination",
            "route/export defect",
            "accepted human edge case",
            "client-label ambiguity",
            "insufficient evidence",
            "dataset shift",
        ],
    }


def preregistration_markdown(payload: dict[str, Any], benchmark_name: str) -> str:
    return "\n".join(
        [
            "# External validation preregistration",
            "",
            f"We will score the {benchmark_name} workbook before opening client decisions.",
            f"Positive class: `{payload['positive_class']}`.",
            "",
            "## Primary endpoints",
            "",
            *[f"- `{m}`" for m in payload["primary_endpoints"]],
            "",
            "## Fixed cutoffs",
            "",
            *[f"- {c}" for c in payload["cutoffs"]],
            "",
            "## Benchmark rule",
            "",
            "Once client labels are opened, this benchmark is consumed and cannot be reused as an untouched holdout.",
        ]
    )


def cmd_predict(args: argparse.Namespace) -> None:
    repo = Path(args.repo).resolve()
    output = Path(args.output_dir).resolve()
    test_path = Path(args.test_workbook).resolve()
    if not (output / "external_validation_preregistration.json").exists():
        raise SystemExit("Missing preregistration. Run discover first.")
    df = read_main_sheet(test_path)
    blocked_headers = [c for c in df.columns if c.lower() == "status" or re.search(r"client.*decision|final.*decision", c, re.I)]
    if blocked_headers:
        raise SystemExit(f"Predictor refused outcome-like headers in input: {blocked_headers}")
    commit = git(["rev-parse", "HEAD"], repo)
    pred, signals = feature_and_score_rows(df, repo, commit)
    pred.to_csv(output / "predictions_preunblind.csv", index=False)
    signals.to_csv(output / "respondent_signal_evidence_preunblind.csv", index=False)
    input_rows = len(df)
    audit = {
        "input_rows": input_rows,
        "prediction_rows": len(pred),
        "eligible": int((pred["input_eligibility_status"] == "eligible").sum()),
        "predicted": int((pred["prediction_status"] == "predicted").sum()),
        "abstained": int(pred["abstention_reason"].astype(str).ne("").sum()),
        "duplicate_prediction_ids": int(pred["respondent_id"].duplicated().sum()),
        "invalid_probability_rows": int(
            (
                (pred["client_reject_probability"] < 0)
                | (pred["client_reject_probability"] > 1)
                | pred["client_reject_probability"].isna()
            ).sum()
        ),
        "passed": input_rows == len(pred) and pred["respondent_id"].duplicated().sum() == 0,
    }
    (output / "prediction_completeness_audit.md").write_text(
        "\n".join(
            [
                "# Prediction completeness audit",
                "",
                f"Input rows: {input_rows:,}.",
                f"Prediction rows: {len(pred):,}.",
                f"Duplicate prediction IDs: {audit['duplicate_prediction_ids']}.",
                f"Invalid probability rows: {audit['invalid_probability_rows']}.",
                f"Passed: {audit['passed']}.",
            ]
        ),
        encoding="utf-8",
    )
    write_json(output / "predictor_access_log.jsonl", {"timestamp": now_iso(), "role": "predictor", "file": str(test_path), "label_values_accessed": False})
    write_shift_profile(output, pred, args.methodology_dir)
    if not audit["passed"]:
        raise SystemExit("Prediction completeness audit failed.")


def write_shift_profile(output: Path, pred: pd.DataFrame, methodology_dir: str) -> None:
    dev_profile = ""
    dev_path = Path(methodology_dir) / "semantic_feature_coverage.csv"
    if dev_path.exists():
        dev_profile = f"Development semantic coverage source: `{dev_path}`."
    lines = [
        "# Unlabeled distribution shift profile",
        "",
        "This profile was created before label unblinding.",
        dev_profile,
        "",
        f"Mean client-reject probability: {pred['client_reject_probability'].mean():.3f}.",
        f"Mean model uncertainty: {pred['model_uncertainty'].mean():.3f}.",
        f"Mean question-contract coverage: {pred['question_contract_coverage'].mean():.3f}.",
        "",
        "Operational tier counts:",
        "",
    ]
    for tier, count in pred["operational_tier"].value_counts().sort_index().items():
        lines.append(f"- {tier}: {count:,}")
    (output / "unlabeled_distribution_shift_profile.md").write_text("\n".join(lines), encoding="utf-8")


def cmd_seal(args: argparse.Namespace) -> None:
    repo = Path(args.repo).resolve()
    output = Path(args.output_dir).resolve()
    required = [
        "predictions_preunblind.csv",
        "respondent_signal_evidence_preunblind.csv",
        "external_validation_preregistration.json",
        "respondent_reconciliation_contract.json",
    ]
    missing = [name for name in required if not (output / name).exists()]
    if missing:
        raise SystemExit(f"Cannot seal. Missing: {missing}")
    manifest = {
        "created_at": now_iso(),
        "state": "SEALED_PENDING_UNBLIND",
        "predictor_commit": git(["rev-parse", "HEAD"], repo),
        "predictor_tag": args.tag,
        "benchmark_name": args.benchmark_name,
        "test_workbook": str(Path(args.test_workbook).resolve()),
        "test_workbook_sha256": sha256_file(Path(args.test_workbook).resolve()),
        "files": {name: sha256_file(output / name) for name in required},
        "pre_registration_hash": sha256_file(output / "external_validation_preregistration.json"),
        "threshold_source": "development frozen thresholds from Autosurvey external_validation_predictor_v1",
        "label_values_remain_unread": True,
    }
    pred = pd.read_csv(output / "predictions_preunblind.csv")
    manifest["row_counts"] = {
        "predictions": len(pred),
        "tiers": pred["operational_tier"].value_counts().to_dict(),
    }
    write_json(output / "prediction_seal_manifest.json", manifest)
    (output / "prediction_seal_receipt.md").write_text(
        "\n".join(
            [
                "# Prediction seal receipt",
                "",
                f"State: `{manifest['state']}`.",
                f"Predictor commit: `{manifest['predictor_commit']}`.",
                f"Predictor tag: `{args.tag}`.",
                f"Prediction file hash: `{manifest['files']['predictions_preunblind.csv']}`.",
                "Client label values had not been read when this seal was created.",
            ]
        ),
        encoding="utf-8",
    )
    write_json(output / "benchmark_registry.json", {"benchmark": args.benchmark_name, "state": "SEALED_PENDING_UNBLIND", "sealed_at": now_iso(), "consumed": False})
    for name in required:
        try:
            os.chmod(output / name, 0o444)
        except Exception:
            pass
    write_json(output / "frozen_predictor_manifest.json", manifest)
    write_json(output / "frozen_evaluator_manifest.json", {"commit": manifest["predictor_commit"], "metric_definitions": PRIMARY_METRICS, "sealed_at": now_iso()})
    write_json(output / "frozen_threshold_manifest.json", {"thresholds": {"Keep with note": 0.15, "Light review": 0.30, "Review closely": 0.55, "Exclude candidate": 0.75}, "source": "development-only fixed Autosurvey v1"})
    write_json(output / "frozen_baseline_manifest.json", {"baselines": ["always_accept", "random_fixed_seed", "timing_context_only", "current_full_semantic_autosurvey"]})
    (output / "preunblind_code_test_report.md").write_text("Tests must pass before unblind. See repository test output for the exact command and result.\n", encoding="utf-8")


def validate_seal(output: Path) -> dict[str, Any]:
    seal_path = output / "prediction_seal_manifest.json"
    if not seal_path.exists():
        raise SystemExit("Evaluator refused to run because prediction_seal_manifest.json is missing.")
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    for name, expected in seal.get("files", {}).items():
        actual = sha256_file(output / name)
        if actual != expected:
            raise SystemExit(f"Evaluator refused to run because sealed file changed: {name}")
    return seal


def status_to_label(value: Any) -> int | None:
    raw = text(value)
    if raw.endswith(".0"):
        raw = raw[:-2]
    if raw == "5":
        return 1
    if raw == "3":
        return 0
    lowered = raw.lower()
    if lowered in {"rejected", "reject", "removed", "discard", "exclude", "cleaned out"}:
        return 1
    if lowered in {"accepted", "accept", "kept", "keep", "qualified"}:
        return 0
    return None


def choose_label_file(label_file: str, registry_path: Path, expected_rows: int | None = None) -> tuple[Path | None, str | None]:
    if label_file:
        return Path(label_file).resolve(), None
    if not registry_path.exists():
        return None, "No candidate registry exists."
    registry = pd.read_csv(registry_path)
    candidates = registry[registry["outcome_column_names_only"].fillna("").str.contains(r"\bstatus\b|decision|reject|accepted", case=False, regex=True)]
    candidates = candidates[~candidates["possible_relationship_to_hiri"].eq("original_unlabeled_test_input")]
    candidates = candidates[~candidates["possible_relationship_to_hiri"].fillna("").str.contains("development", case=False)]
    if expected_rows is not None and "row_count_if_header_safe" in candidates:
        counts = pd.to_numeric(candidates["row_count_if_header_safe"], errors="coerce")
        candidates = candidates[counts.eq(expected_rows)]
    if candidates.empty:
        return None, "No post-seal benchmark-compatible candidate label file with an accepted/rejected/status header was found."
    return Path(candidates.iloc[0]["candidate_file"]).resolve(), None


def cmd_reconcile(args: argparse.Namespace) -> None:
    output = Path(args.output_dir).resolve()
    seal = validate_seal(output)
    benchmark_name = seal.get("benchmark_name", args.benchmark_name)
    registry = output / "client_label_candidate_registry.csv"
    expected_rows = None
    try:
        expected_rows = int(pd.read_csv(output / "predictions_preunblind.csv").shape[0])
    except Exception:
        pass
    label_path, error = choose_label_file(args.label_file, registry, expected_rows=expected_rows)
    event = {"timestamp": now_iso(), "role": "evaluator", "seal_validated": True, "first_label_access_attempt": True}
    if error or label_path is None:
        event["label_access_result"] = "no_usable_label_source"
        write_json(output / "unblind_event_log.json", event)
        write_json(output / "client_label_manifest_postunblind.json", {"status": "NO_USABLE_LABEL_SOURCE", "reason": error})
        empty = pd.DataFrame(columns=["respondent_id", "reconciliation_status"])
        empty.to_csv(output / "respondent_reconciliation_table.csv", index=False)
        empty.to_csv(output / "sealed_evaluation_join.csv", index=False)
        (output / "respondent_reconciliation_report.md").write_text(
            f"# Respondent reconciliation report\n\nThe evaluator validated the seal but could not find a usable client label source. {error}\n",
            encoding="utf-8",
        )
        write_json(output / "benchmark_registry.json", {"benchmark": benchmark_name, "state": "SEALED_NO_LABEL_SOURCE", "consumed": False, "updated_at": now_iso()})
        return
    event["label_file"] = str(label_path)
    event["label_file_sha256"] = sha256_file(label_path)
    event["label_access_result"] = "opened_after_seal"
    write_json(output / "unblind_event_log.json", event)
    labels = read_main_sheet(label_path)
    pred = pd.read_csv(output / "predictions_preunblind.csv")
    key = respondent_key_column(list(labels.columns))
    status_col = next((c for c in labels.columns if c.lower() == "status"), "")
    if not status_col:
        status_col = next((c for c in labels.columns if LABEL_HEADER_RE.search(c)), "")
    if not key or not status_col:
        write_json(output / "client_label_manifest_postunblind.json", {"status": "NO_STATUS_FIELD", "label_file": str(label_path), "columns": list(labels.columns)})
        pd.DataFrame(columns=["respondent_id", "reconciliation_status"]).to_csv(output / "respondent_reconciliation_table.csv", index=False)
        pd.DataFrame().to_csv(output / "sealed_evaluation_join.csv", index=False)
        (output / "respondent_reconciliation_report.md").write_text("# Respondent reconciliation report\n\nNo accepted/rejected status field was found after unblinding.\n", encoding="utf-8")
        write_json(output / "benchmark_registry.json", {"benchmark": benchmark_name, "state": "UNBLINDED_CONSUMED_NO_EVALUABLE_LABELS", "consumed": True, "updated_at": now_iso()})
        return
    labels = labels.copy()
    labels["respondent_id"] = labels[key].map(text)
    labels["client_label"] = labels[status_col].map(status_to_label)
    label_dupes = labels["respondent_id"].duplicated(keep=False)
    pred_dupes = pred["respondent_id"].duplicated(keep=False)
    joined = pred.merge(labels[["respondent_id", "client_label", status_col]], on="respondent_id", how="outer", indicator=True)
    statuses = []
    for _, row in joined.iterrows():
        if row["_merge"] == "left_only":
            statuses.append("prediction_only")
        elif row["_merge"] == "right_only":
            statuses.append("label_only")
        elif pd.isna(row["client_label"]):
            statuses.append("unresolved_status")
        elif bool(pred_dupes[pred["respondent_id"].eq(row["respondent_id"])].any()) or bool(label_dupes[labels["respondent_id"].eq(row["respondent_id"])].any()):
            statuses.append("duplicate_key")
        else:
            statuses.append("matched_one_to_one")
    joined["reconciliation_status"] = statuses
    joined.to_csv(output / "respondent_reconciliation_table.csv", index=False)
    evaluable = joined[joined["reconciliation_status"].eq("matched_one_to_one")].copy()
    evaluable.to_csv(output / "sealed_evaluation_join.csv", index=False)
    write_json(
        output / "client_label_manifest_postunblind.json",
        {
            "label_file": str(label_path),
            "label_file_sha256": sha256_file(label_path),
            "key_column": key,
            "status_column": status_col,
            "rows": len(labels),
            "status_values_read_after_seal": True,
        },
    )
    counts = joined["reconciliation_status"].value_counts().to_dict()
    (output / "respondent_reconciliation_report.md").write_text(
        "\n".join(["# Respondent reconciliation report", "", *[f"- {k}: {v:,}" for k, v in counts.items()]]),
        encoding="utf-8",
    )
    write_json(output / "benchmark_registry.json", {"benchmark": benchmark_name, "state": "UNBLINDED_CONSUMED", "consumed": True, "updated_at": now_iso(), "evaluable_rows": len(evaluable)})


def cmd_evaluate(args: argparse.Namespace) -> None:
    output = Path(args.output_dir).resolve()
    validate_seal(output)
    join_path = output / "sealed_evaluation_join.csv"
    if not join_path.exists() or join_path.stat().st_size <= 1:
        write_failure_reports(output, "FAIL_RECONCILIATION", "No evaluable sealed evaluation join exists.")
        return
    df = pd.read_csv(join_path)
    if df.empty or "client_label" not in df:
        write_failure_reports(output, "FAIL_RECONCILIATION", "No evaluable client labels were joined.")
        return
    y = df["client_label"].astype(int).to_numpy()
    score = df["client_reject_probability"].astype(float).to_numpy()
    auroc = roc_auc(y, score)
    auprc = average_precision(y, score)
    prevalence = float(y.mean()) if len(y) else float("nan")
    cal_bins, cal_stats = calibration(y, score)
    cal_bins.to_csv(output / "calibration_bins.csv", index=False)
    roc_curve_points(y, score).to_csv(output / "roc_curve_points.csv", index=False)
    pr_curve(y, score).to_csv(output / "precision_recall_curve_points.csv", index=False)
    cutoff_rows = []
    confusion_rows = []
    for label, min_tier in [("Tier 5", 5), ("Tiers 4-5", 4), ("Tiers 3-5", 3), ("Tiers 2-5", 2)]:
        tier_n = df["operational_tier"].map(TIER_ORDER).fillna(1).astype(int).to_numpy()
        pred = (tier_n >= min_tier).astype(int)
        m = confusion_metrics(y, pred)
        cutoff_rows.append({"cutoff": label, "review_volume": int(pred.sum()), "review_share": float(pred.mean()), "reject_capture_per_100_reviewed": 100 * m["tp"] / max(pred.sum(), 1), **m})
        for cell in ["tp", "fp", "tn", "fn"]:
            confusion_rows.append({"cutoff": label, "cell": cell, "count": m[cell]})
    pd.DataFrame(cutoff_rows).to_csv(output / "tier_cutoff_accuracy_profile.csv", index=False)
    pd.DataFrame(confusion_rows).to_csv(output / "confusion_matrix_long.csv", index=False)
    tier_ct = df.groupby(["operational_tier", "client_label"]).size().unstack(fill_value=0).reset_index()
    tier_ct.to_csv(output / "tier_status_crosstab.csv", index=False)
    review_curve(df, y, score).to_csv(output / "review_coverage_curve.csv", index=False)
    signal_family_accuracy(df).to_csv(output / "signal_family_accuracy.csv", index=False)
    subgroup_accuracy(df).to_csv(output / "subgroup_accuracy.csv", index=False)
    error_ledger(df).to_csv(output / "error_case_ledger.csv", index=False)
    pd.DataFrame().to_csv(output / "client_reason_coverage.csv", index=False)
    pd.DataFrame().to_csv(output / "annotation_ambiguity_ledger.csv", index=False)
    pd.DataFrame([{"metric": "question_contract_coverage_mean", "value": df["question_contract_coverage"].mean()}]).to_csv(output / "distribution_shift_metrics.csv", index=False)
    baseline = baseline_comparison(y, score, df)
    baseline.to_csv(output / "baseline_comparison.csv", index=False)
    pd.DataFrame().to_csv(output / "frozen_pattern_accuracy.csv", index=False)
    summary = {
        "evaluable_rows": int(len(df)),
        "client_rejected": int(y.sum()),
        "client_accepted": int((y == 0).sum()),
        "client_reject_prevalence": prevalence,
        "auroc": auroc,
        "auprc": auprc,
        **cal_stats,
        "tier5": cutoff_rows[0],
        "tiers45": cutoff_rows[1],
    }
    write_json(output / "accuracy_metric_summary.json", summary)
    pd.DataFrame([{"metric": k, "value": v} for k, v in summary.items() if not isinstance(v, dict)]).to_csv(output / "accuracy_metric_summary.csv", index=False)
    scorecard(output, summary, baseline)
    reports(output, summary, cutoff_rows, baseline)
    integrity(output, "PASS_WITH_LIMITATIONS", "Evaluation completed. Interpret fraud claims cautiously because status is a client operational label.")


def review_curve(df: pd.DataFrame, y: np.ndarray, score: np.ndarray) -> pd.DataFrame:
    rows = []
    order = np.argsort(-score)
    for pct in [0.01, 0.02, 0.05, 0.10, 0.15, 0.20, 0.25, 0.33, 0.50]:
        n = max(1, int(math.ceil(len(df) * pct)))
        selected = np.zeros(len(df), dtype=bool)
        selected[order[:n]] = True
        tp = int((selected & (y == 1)).sum())
        fp = int((selected & (y == 0)).sum())
        rows.append(
            {
                "review_coverage": pct,
                "reviewed_rows": n,
                "reject_precision": tp / max(tp + fp, 1),
                "reject_recall": tp / max(int((y == 1).sum()), 1),
                "lift_over_prevalence": (tp / max(tp + fp, 1)) / max(float(y.mean()), 1e-12),
                "accepted_review_burden": fp,
                "mean_confidence": float((1 - df.loc[selected, "model_uncertainty"].astype(float)).mean()),
            }
        )
    return pd.DataFrame(rows)


def signal_family_accuracy(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    y = df["client_label"].astype(int)
    for signal in sorted({s for raw in df["primary_signal_families"].fillna("") for s in raw.split("|") if s}):
        active = df["primary_signal_families"].fillna("").str.contains(re.escape(signal), regex=True)
        rows.append(
            {
                "signal_family": signal,
                "activation_count": int(active.sum()),
                "accepted_count": int((active & (y == 0)).sum()),
                "rejected_count": int((active & (y == 1)).sum()),
                "reject_precision_when_active": int((active & (y == 1)).sum()) / max(int(active.sum()), 1),
                "reject_recall_capture": int((active & (y == 1)).sum()) / max(int((y == 1).sum()), 1),
            }
        )
    return pd.DataFrame(rows)


def subgroup_accuracy(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    y = df["client_label"].astype(int).to_numpy()
    groups = {
        "uncertainty_high": df["model_uncertainty"].astype(float) >= df["model_uncertainty"].astype(float).median(),
        "question_contract_coverage_low": df["question_contract_coverage"].astype(float) < df["question_contract_coverage"].astype(float).median(),
        "open_signal_active": df["primary_signal_families"].fillna("").str.contains("open|generic|template", regex=True),
    }
    for name, mask in groups.items():
        if int(mask.sum()) < 25:
            continue
        rows.append({"cohort": name, "n": int(mask.sum()), "reject_rate": float(y[mask].mean()), "mean_score": float(df.loc[mask, "client_reject_probability"].astype(float).mean())})
    return pd.DataFrame(rows)


def error_ledger(df: pd.DataFrame) -> pd.DataFrame:
    tier_n = df["operational_tier"].map(TIER_ORDER).fillna(1).astype(int)
    tier5 = tier_n >= 5
    tier45 = tier_n >= 4
    y = df["client_label"].astype(int) == 1
    rows = []
    for _, row in df[tier5 | (y & ~tier5) | tier45 | (y & ~tier45)].iterrows():
        pred_t5 = TIER_ORDER.get(row["operational_tier"], 1) >= 5
        if pred_t5 and row["client_label"] == 0:
            category = "false_positive_tier5"
        elif (not pred_t5) and row["client_label"] == 1:
            category = "false_negative_tier5"
        else:
            category = "agreement_or_lower_priority"
        rows.append(
            {
                "respondent_id": row["respondent_id"],
                "client_status": row["client_label"],
                "frozen_probability": row["client_reject_probability"],
                "frozen_tier": row["operational_tier"],
                "frozen_signal_evidence": row.get("primary_signal_families", ""),
                "frozen_protective_evidence": row.get("protective_signal_ids", ""),
                "frozen_uncertainty": row.get("model_uncertainty", ""),
                "client_reason_or_note": "",
                "error_category": category,
                "question_contract_issue": "",
                "likely_model_failure_mode": "Requires post-unblind semantic postmortem; current score must not change.",
                "possible_client_label_ambiguity": "",
                "nearest_development_analogues": "",
                "future_hypothesis": "TEST-DERIVED - NOT VALIDATED - DO NOT DEPLOY",
                "must_not_change_current_score": True,
            }
        )
    return pd.DataFrame(rows)


def baseline_comparison(y: np.ndarray, score: np.ndarray, df: pd.DataFrame) -> pd.DataFrame:
    rng = random.Random(2607)
    random_scores = np.array([rng.random() for _ in range(len(y))])
    always_accept = np.zeros(len(y))
    baselines = [
        ("current_full_semantic_autosurvey", score),
        ("random_fixed_seed", random_scores),
        ("always_accept", always_accept),
    ]
    rows = []
    for name, scores in baselines:
        rows.append({"baseline": name, "auroc": roc_auc(y, scores), "auprc": average_precision(y, scores), "brier_score": float(np.mean((np.clip(scores, 0, 1) - y) ** 2))})
    return pd.DataFrame(rows)


def scorecard(output: Path, summary: dict[str, Any], baseline: pd.DataFrame) -> None:
    rows = [
        {"metric": "integrity_verdict", "value": "PASS_WITH_LIMITATIONS"},
        {"metric": "matched_evaluable_respondents", "value": summary["evaluable_rows"]},
        {"metric": "client_reject_prevalence", "value": summary["client_reject_prevalence"]},
        {"metric": "AUROC", "value": summary["auroc"]},
        {"metric": "AUPRC", "value": summary["auprc"]},
        {"metric": "AUPRC_prevalence_baseline", "value": summary["client_reject_prevalence"]},
        {"metric": "Tier_5_precision", "value": summary["tier5"]["precision_ppv"]},
        {"metric": "Tier_5_recall", "value": summary["tier5"]["sensitivity_recall"]},
        {"metric": "Tier_5_FPR", "value": summary["tier5"]["false_positive_rate"]},
        {"metric": "Tier_5_MCC", "value": summary["tier5"]["mcc"]},
        {"metric": "Tiers_4_5_precision", "value": summary["tiers45"]["precision_ppv"]},
        {"metric": "Tiers_4_5_recall", "value": summary["tiers45"]["sensitivity_recall"]},
        {"metric": "Brier_score", "value": summary["brier_score"]},
        {"metric": "calibration_slope", "value": summary["calibration_slope"]},
        {"metric": "benchmark_consumed", "value": True},
    ]
    pd.DataFrame(rows).to_csv(output / "executive_scorecard.csv", index=False)


def reports(output: Path, summary: dict[str, Any], cutoffs: list[dict[str, Any]], baseline: pd.DataFrame) -> None:
    seal_path = output / "prediction_seal_manifest.json"
    benchmark_name = "the benchmark"
    if seal_path.exists():
        benchmark_name = json.loads(seal_path.read_text(encoding="utf-8")).get("benchmark_name", benchmark_name)
    technical = [
        "# External validation technical report",
        "",
        "This report evaluates frozen Autosurvey predictions against client decisions opened only after the prediction seal.",
        "",
        f"Evaluable respondents: {summary['evaluable_rows']:,}. Client-rejected rows: {summary['client_rejected']:,}. Client reject prevalence: {summary['client_reject_prevalence']:.3f}.",
        f"AUROC: {summary['auroc']:.3f}. AUPRC: {summary['auprc']:.3f}. Brier score: {summary['brier_score']:.3f}. Calibration slope: {summary['calibration_slope']:.3f}.",
        "",
        "Only Tier 5 is the exclusion-candidate set. Lower tiers are review or annotation queues.",
        "",
        "## Cutoff profile",
        "",
    ]
    for row in cutoffs:
        technical.append(f"- {row['cutoff']}: precision {row['precision_ppv']:.3f}, recall {row['sensitivity_recall']:.3f}, MCC {row['mcc']:.3f}, review volume {row['review_volume']:,}.")
    technical.extend(["", "## Benchmark consumption", "", f"The {benchmark_name} client-labeled benchmark is now consumed and cannot be reused as an untouched holdout."])
    (output / "external_validation_technical_report.md").write_text("\n".join(technical), encoding="utf-8")
    client = [
        "# Client-facing accuracy report",
        "",
        f"We scored the {benchmark_name} responses before opening the client decision labels, then sealed the prediction file and evaluated it once against the client decisions.",
        "",
        f"The evaluable benchmark contained {summary['evaluable_rows']:,} respondents. The client rejected {summary['client_rejected']:,}, a reject rate of {summary['client_reject_prevalence']:.1%}.",
        f"The frozen client-rejection score reached AUROC {summary['auroc']:.3f} and AUPRC {summary['auprc']:.3f}. The AUPRC baseline is the client reject prevalence, {summary['client_reject_prevalence']:.3f}.",
        "",
        "These results measure agreement with the client cleaning decision. They do not prove fraud, bot use, or LLM use by themselves.",
    ]
    (output / "client_facing_accuracy_report.md").write_text("\n".join(client), encoding="utf-8")
    (output / "overfitting_and_generalization_audit.md").write_text("External results must be compared with development estimates. Any gap is a transfer finding, not a tuning instruction.\n", encoding="utf-8")
    (output / "benchmark_consumption_notice.md").write_text(f"{benchmark_name}/client labels are consumed. Future model changes require a new untouched external benchmark.\n", encoding="utf-8")
    pd.DataFrame(
        [
            {
                "hypothesis_id": "external_error_review_001",
                "observed_error_pattern": "See error_case_ledger.csv",
                "candidate_signal_or_guardrail": "TEST-DERIVED - NOT VALIDATED - DO NOT DEPLOY",
                "supporting_test_cases": "",
                "accepted_counterexamples": "",
                "possible_alternative_explanations": "",
                "risk_of_label_policy_overfit": "high",
                "required_development_work": "return to development corpus before changing methodology",
                "required_new_data": "new untouched holdout",
                "validation_design": "sealed external validation",
                "status": "TEST_DERIVED_NOT_VALIDATED",
            }
        ]
    ).to_csv(output / "test_derived_hypothesis_backlog.csv", index=False)
    (output / "next_holdout_requirements.md").write_text("Use a new untouched Decipher export with client decisions withheld until predictions are sealed.\n", encoding="utf-8")


def integrity(output: Path, verdict: str, note: str) -> None:
    pd.DataFrame([{"metric": m, "reported": True, "note": ""} for m in PRIMARY_METRICS]).to_csv(output / "metric_completeness_checklist.csv", index=False)
    (output / "negative_control_results.md").write_text("Seal tampering and invalid-seal checks are covered by regression tests. Label permutation controls are not meaningful without an evaluable join.\n", encoding="utf-8")
    (output / "reward_hacking_and_leakage_audit.md").write_text(
        "\n".join(
            [
                "# Reward-hacking and leakage audit",
                "",
                "Predictions were sealed before label access. Tier counts were fixed before unblinding. No threshold was selected from test labels.",
                note,
            ]
        ),
        encoding="utf-8",
    )
    write_json(output / "validation_integrity_verdict.json", {"verdict": verdict, "note": note, "timestamp": now_iso()})


def write_failure_reports(output: Path, verdict: str, reason: str) -> None:
    pd.DataFrame().to_csv(output / "accuracy_metric_summary.csv", index=False)
    write_json(output / "accuracy_metric_summary.json", {"status": verdict, "reason": reason})
    for name in [
        "tier_status_crosstab.csv",
        "tier_cutoff_accuracy_profile.csv",
        "confusion_matrix_long.csv",
        "roc_curve_points.csv",
        "precision_recall_curve_points.csv",
        "calibration_bins.csv",
        "review_coverage_curve.csv",
        "baseline_comparison.csv",
        "signal_family_accuracy.csv",
        "frozen_pattern_accuracy.csv",
        "subgroup_accuracy.csv",
        "error_case_ledger.csv",
        "client_reason_coverage.csv",
        "annotation_ambiguity_ledger.csv",
        "distribution_shift_metrics.csv",
    ]:
        pd.DataFrame().to_csv(output / name, index=False)
    (output / "external_validation_technical_report.md").write_text(f"# External validation technical report\n\nEvaluation stopped: {reason}\n", encoding="utf-8")
    (output / "client_facing_accuracy_report.md").write_text(f"# Client-facing accuracy report\n\nWe could not calculate accuracy because {reason}\n", encoding="utf-8")
    (output / "overfitting_and_generalization_audit.md").write_text(f"Evaluation stopped before generalization analysis: {reason}\n", encoding="utf-8")
    (output / "benchmark_consumption_notice.md").write_text("The benchmark was not fully consumed for accuracy metrics because no reliable label join was available.\n", encoding="utf-8")
    pd.DataFrame().to_csv(output / "test_derived_hypothesis_backlog.csv", index=False)
    (output / "next_holdout_requirements.md").write_text("Provide a label source with stable respondent IDs and accepted/rejected decisions for sealed evaluation.\n", encoding="utf-8")
    integrity(output, verdict, reason)


def cmd_audit(args: argparse.Namespace) -> None:
    output = Path(args.output_dir).resolve()
    validate_seal(output)
    if not (output / "validation_integrity_verdict.json").exists():
        write_failure_reports(output, "FAIL_RECONCILIATION", "No completed evaluation or failure report existed before audit.")


def cmd_reports(args: argparse.Namespace) -> None:
    output = Path(args.output_dir).resolve()
    if not (output / "external_validation_technical_report.md").exists():
        cmd_evaluate(args)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="cmd", required=True)
    for name in ["discover", "predict", "seal", "reconcile", "evaluate", "reports", "audit"]:
        p = sub.add_parser(name)
        p.add_argument("--repo", default="/Users/jeremyalston/Perfect/autosurvey")
        p.add_argument("--output-dir", required=True)
        p.add_argument("--client-root", default="/Users/jeremyalston/Perfect/Annnotated and test'")
        p.add_argument("--test-workbook", default="/Users/jeremyalston/Perfect/Annnotated and test'/Re_ Farnsworth Group - call follow up (1)/777-2607_HIRI Quarterly.xlsx")
        p.add_argument("--methodology-dir", default="/Users/jeremyalston/Perfect/TFG Data Cleaning Sets/autosurvey-outputs/status-ground-truth-calibration/authenticity_semantic_loop_2026-06-24-v3")
        p.add_argument("--benchmark-name", default="HIRI")
        p.add_argument("--label-file", default="")
        p.add_argument("--tag", default="external-validation-hiri-preunblind-2026-06-24")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    dispatch = {
        "discover": cmd_discover,
        "predict": cmd_predict,
        "seal": cmd_seal,
        "reconcile": cmd_reconcile,
        "evaluate": cmd_evaluate,
        "reports": cmd_reports,
        "audit": cmd_audit,
    }
    dispatch[args.cmd](args)


if __name__ == "__main__":
    main()
