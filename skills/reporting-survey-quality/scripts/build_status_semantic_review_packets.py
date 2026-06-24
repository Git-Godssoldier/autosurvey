#!/usr/bin/env python3
"""Build chunked semantic reading packets for TFG status-labeled rows."""

from __future__ import annotations

import argparse
import csv
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


TECHNICAL_FIELD = re.compile(
    r"(^record$|^uuid$|^rid$|^id$|^date$|^markers$|^status$|^qc$|^rd_|rd_|gettoken|termflags|"
    r"scrutinyflags|clientflags|redem_|_rd_review|rd_review|rd_search|qtime|start|end|duration|ip|"
    r"source|supplier|vendor|langassess|validclient|conditions|^vlist$|^vos$|^vbrowser$|^vmobile|"
    r"^list$|useragent|^url$|^session$|^bhf$|^sfh$|^supname$|possiblechannels|^noanswer)",
    re.I,
)
METADATA_FIELD = re.compile(
    r"^(record|uuid|rid|date|markers|status|qc|qtime|source|supplier|vendor|TERMFLAGS|SCRUTINYFLAGS|CLIENTFLAGS|redem_|RD_Search|RD_Review)",
    re.I,
)
EMPTY_VALUES = {"", "nan", "none", "null"}


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalized_status(value: object) -> str:
    raw = text(value)
    if raw.endswith(".0"):
        raw = raw[:-2]
    return raw


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Annotated .xlsx, .zip, or directory.")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--sheet", default="A1")
    parser.add_argument("--id-column", default="uuid")
    parser.add_argument("--rows-per-packet", type=int, default=25)
    parser.add_argument(
        "--max-field-chars",
        type=int,
        default=0,
        help="Optional per-field truncation. 0 means do not truncate.",
    )
    return parser.parse_args()


def workbook_entries(path: Path) -> list[tuple[str, bytes]]:
    path = path.expanduser().resolve()
    if path.is_dir():
        entries: list[tuple[str, bytes]] = []
        for child in sorted(path.rglob("*")):
            if child.suffix.lower() in {".xlsx", ".zip"}:
                entries.extend(workbook_entries(child))
        return entries
    if path.suffix.lower() == ".xlsx":
        return [(path.name, path.read_bytes())]
    if path.suffix.lower() == ".zip":
        entries = []
        with ZipFile(path) as zf:
            for name in zf.namelist():
                lower = name.lower()
                if lower.endswith(".xlsx"):
                    entries.append((name, zf.read(name)))
                elif lower.endswith(".zip"):
                    with ZipFile(BytesIO(zf.read(name))) as nested:
                        for nested_name in nested.namelist():
                            if nested_name.lower().endswith(".xlsx"):
                                entries.append((f"{name}::{nested_name}", nested.read(nested_name)))
        return entries
    raise SystemExit(f"Unsupported input type: {path}")


def prompt_map(workbook) -> dict[str, str]:
    if "Datamap" not in workbook.sheetnames:
        return {}
    ws = workbook["Datamap"]
    prompts: dict[str, str] = {}
    pattern = re.compile(r"^\[([^\]]+)\]:\s*(.*)$")
    for row in ws.iter_rows(values_only=True):
        first = text(row[0] if row else "")
        match = pattern.match(first)
        if match:
            field, prompt = match.groups()
            prompts[field.strip()] = prompt.strip()
            continue
        joined = " | ".join(text(value) for value in row[:4] if text(value))
        match = re.search(r"\(([^,)]+)(?:,[^)]+)?\)\s+(.+)", joined)
        if match:
            field, prompt = match.groups()
            prompts.setdefault(field.strip(), prompt.strip())
    return prompts


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-")[:120] or "workbook"


def clipped(value: str, max_chars: int) -> str:
    if max_chars <= 0 or len(value) <= max_chars:
        return value
    return value[:max_chars].rstrip() + " [truncated]"


def semantic_rows(entry_name: str, data: bytes, args: argparse.Namespace) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    if args.sheet not in workbook.sheetnames:
        sheet = workbook.sheetnames[0]
    else:
        sheet = args.sheet
    ws = workbook[sheet]
    headers = [text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: idx for idx, header in enumerate(headers) if header}
    if "status" not in index:
        return [], {"dataset": entry_name, "rows": ws.max_row - 1, "has_status": False}
    id_column = args.id_column if args.id_column in index else next(
        (candidate for candidate in ["uuid", "record", "RID"] if candidate in index),
        "",
    )
    prompts = prompt_map(workbook)
    answer_columns = [
        header
        for header in headers
        if header
        and header in index
        and not TECHNICAL_FIELD.search(header)
    ]
    metadata_columns = [header for header in headers if header and header in index and METADATA_FIELD.search(header)]
    rows: list[dict[str, object]] = []
    for source_row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        status = normalized_status(values[index["status"]] if index["status"] < len(values) else "")
        if status not in {"3", "5"}:
            continue
        key = text(values[index[id_column]]) if id_column and index[id_column] < len(values) else str(source_row_number)
        chain: list[tuple[str, str, str]] = []
        metadata: list[tuple[str, str]] = []
        for header in metadata_columns:
            idx = index[header]
            value = text(values[idx] if idx < len(values) else "")
            if value.lower() in EMPTY_VALUES:
                continue
            metadata.append((header, clipped(value, args.max_field_chars)))
        for header in answer_columns:
            idx = index[header]
            value = text(values[idx] if idx < len(values) else "")
            if value.lower() in EMPTY_VALUES:
                continue
            prompt = prompts.get(header, "")
            chain.append((header, prompt, clipped(value, args.max_field_chars)))
        rows.append(
            {
                "dataset": entry_name,
                "source_row_number": source_row_number,
                "respondent_key": key,
                "status": status,
                "decision": "accepted" if status == "3" else "rejected",
                "metadata": metadata,
                "chain": chain,
            }
        )
    summary = {
        "dataset": entry_name,
        "rows": ws.max_row - 1,
        "has_status": True,
        "semantic_rows": len(rows),
        "answer_columns": len(answer_columns),
        "id_column": id_column,
    }
    return rows, summary


def write_packet(path: Path, packet_rows: list[dict[str, object]], packet_number: int, total_packets: int) -> None:
    lines = [
        f"# Semantic review packet {packet_number} of {total_packets}",
        "",
        "Read every respondent below. For rejected rows, identify the actual authenticity signal.",
        "For accepted rows, identify guardrails that protect real respondents from over-flagging.",
        "Do not rely on the status label alone. Use the full answer chain.",
        "",
    ]
    for row in packet_rows:
        lines.extend(
            [
                "---",
                "",
                f"## {row['dataset']} | source row {row['source_row_number']} | {row['respondent_key']}",
                "",
        f"Status: {row['status']} ({row['decision']})",
        "",
                "Metadata and quality context:",
                "",
            ]
        )
        metadata = row["metadata"]
        if not metadata:
            lines.append("- No metadata fields were available.")
        for field, value in metadata:
            lines.append(f"- {field}: {value}")
        lines.extend(
            [
                "",
                "Full respondent answer chain:",
                "",
            ]
        )
        chain = row["chain"]
        if not chain:
            lines.append("- No nontechnical answer fields were available.")
        for field, prompt, value in chain:
            label = f"{field}"
            if prompt:
                label += f" | {prompt}"
            lines.append(f"- {label}: {value}")
        lines.extend(
            [
                "",
                "Agent reading notes to write after review:",
                "",
                "- Authenticity signal or guardrail:",
                "- Evidence from this chain:",
                "- Counterevidence or benign explanation:",
                "- How this should affect future scoring:",
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def protocol(index_rows: list[dict[str, object]], summaries: list[dict[str, object]]) -> str:
    total_rows = sum(int(row["rows"]) for row in index_rows)
    rejected = sum(int(row["rejected_rows"]) for row in index_rows)
    accepted = sum(int(row["accepted_rows"]) for row in index_rows)
    lines = [
        "# Status semantic reading protocol",
        "",
        "This folder exists so the agent reads rows semantically instead of relying on parsed flags.",
        "",
        f"Packets: {len(index_rows)}",
        f"Rows in packets: {total_rows}",
        f"Accepted status 3 rows: {accepted}",
        f"Rejected status 5 rows: {rejected}",
        "",
        "## Rules",
        "",
        "1. Read each packet in order.",
        "2. For every status 5 row, write the actual fabricated-response, bot-like, LLM-assisted, inattentive, contradictory, or unauthentic signal if one is present.",
        "3. For every status 3 row, write the accepted-row guardrail if the row contains a pattern that a weaker system might over-flag.",
        "4. Promote no signal until it has been checked against accepted counterexamples.",
        "5. After each packet, write a packet note in `semantic_packet_notes/` with discoveries, guardrails, uncertain rows, and scoring changes.",
        "6. After all packets, write a synthesis that explains the client problem shape and the signals to test on the blinded dataset.",
        "",
        "## Packet index",
        "",
    ]
    for row in index_rows:
        lines.append(
            f"- `{row['packet_path']}`: {row['rows']} rows, "
            f"{row['rejected_rows']} rejected, {row['accepted_rows']} accepted."
        )
    lines.extend(["", "## Workbook summaries", ""])
    for summary in summaries:
        if summary.get("has_status"):
            lines.append(
                f"- {summary['dataset']}: {summary['semantic_rows']} labeled rows and "
                f"{summary['answer_columns']} answer columns."
            )
        else:
            lines.append(f"- {summary['dataset']}: no status column. Treat as blinded or unlabeled.")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir.expanduser().resolve()
    packet_dir = output_dir / "semantic_review_packets"
    packet_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "semantic_packet_notes").mkdir(parents=True, exist_ok=True)

    packet_index: list[dict[str, object]] = []
    summaries: list[dict[str, object]] = []
    packet_number = 0
    pending_packets: list[tuple[Path, list[dict[str, object]]]] = []

    for entry_name, data in workbook_entries(args.input):
        rows, summary = semantic_rows(entry_name, data, args)
        summaries.append(summary)
        if not rows:
            continue
        base = safe_name(entry_name)
        for start in range(0, len(rows), args.rows_per_packet):
            packet_number += 1
            packet_rows = rows[start : start + args.rows_per_packet]
            path = packet_dir / f"{packet_number:04d}-{base}.md"
            pending_packets.append((path, packet_rows))
            accepted = sum(1 for row in packet_rows if row["status"] == "3")
            rejected = sum(1 for row in packet_rows if row["status"] == "5")
            packet_index.append(
                {
                    "packet_number": packet_number,
                    "packet_path": str(path.relative_to(output_dir)),
                    "dataset": entry_name,
                    "rows": len(packet_rows),
                    "accepted_rows": accepted,
                    "rejected_rows": rejected,
                    "first_source_row": packet_rows[0]["source_row_number"],
                    "last_source_row": packet_rows[-1]["source_row_number"],
                }
            )

    total_packets = len(pending_packets)
    for number, (path, rows) in enumerate(pending_packets, start=1):
        write_packet(path, rows, number, total_packets)

    with (output_dir / "semantic_review_packet_index.csv").open("w", newline="", encoding="utf-8") as handle:
        fieldnames = [
            "packet_number",
            "packet_path",
            "dataset",
            "rows",
            "accepted_rows",
            "rejected_rows",
            "first_source_row",
            "last_source_row",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(packet_index)

    (output_dir / "status_semantic_reading_protocol.md").write_text(protocol(packet_index, summaries), encoding="utf-8")
    print(output_dir / "status_semantic_reading_protocol.md")


if __name__ == "__main__":
    main()
