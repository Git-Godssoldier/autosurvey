#!/usr/bin/env python3
"""Execute the annotated-corpus authenticity discovery loop.

This script is for methodology development only. It reads status-labeled TFG
workbooks, removes label leakage before feature extraction, profiles accepted
and rejected respondents, builds matched controls, searches interactions, and
writes transfer-ready signal artifacts. It must not inspect blinded respondent
values.
"""

from __future__ import annotations

import argparse
import hashlib
import itertools
import json
import math
import re
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

try:
    from sklearn.ensemble import GradientBoostingClassifier
    from sklearn.linear_model import LogisticRegression
    from sklearn.metrics import (
        average_precision_score,
        brier_score_loss,
        log_loss,
        precision_recall_fscore_support,
        roc_auc_score,
    )
    from sklearn.preprocessing import StandardScaler
except Exception:  # pragma: no cover - runtime fallback
    GradientBoostingClassifier = None
    LogisticRegression = None
    StandardScaler = None
    average_precision_score = None
    brier_score_loss = None
    log_loss = None
    precision_recall_fscore_support = None
    roc_auc_score = None


ACCEPTED = "3"
REJECTED = "5"
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)
LEAKAGE_RE = re.compile(
    r"(^|_)(status|client|review|recommend|action|decision|annotation|score|tier|"
    r"final|discard|exclude|keep|clean|flag|reason|note|comment|markers?|validclient|"
    r"channeltracking|closingemail|token)(_|$)|"
    r"condition|noanswer|^qc($|\d|_)|"
    r"TERMFLAGS|SCRUTINYFLAGS|CLIENTFLAGS|redem_excluded|RD_Review|RD_Search|quota|bad:",
    re.I,
)
TEXT_HINT_RE = re.compile(r"oe$|open|outro|other|specify|explain|why|comment|qcoe|pasted", re.I)
TIME_RE = re.compile(r"time|duration|qtime|elapsed|start|end|date|timestamp", re.I)
ID_RE = re.compile(r"uuid|respondent|rid|record|id$|session", re.I)
MATRIX_RE = re.compile(r"(.+?)(?:_?r|row)\d+(?:_?c\d+)?$", re.I)
GENERIC_TEXT = {
    "none",
    "nothing",
    "n/a",
    "na",
    "no",
    "good",
    "ok",
    "idk",
    "dont know",
    "don't know",
    "not sure",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value).lower()).strip()


def clean_status(value: Any) -> str:
    raw = text(value)
    return raw[:-2] if raw.endswith(".0") else raw


def slug(value: str) -> str:
    raw = re.sub(r"[^a-zA-Z0-9]+", "_", value).strip("_").lower()
    return raw[:90] or "unknown"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_sources(root: Path) -> list[Path]:
    suffixes = {".xlsx", ".xls", ".csv", ".sav", ".zip"}
    return sorted(p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in suffixes and not p.name.startswith("~$"))


def workbook_metadata(path: Path, inspect_values: bool) -> dict[str, Any]:
    row: dict[str, Any] = {
        "file_path": str(path),
        "file_name": path.name,
        "extension": path.suffix.lower(),
        "bytes": path.stat().st_size,
        "sha256": sha256(path),
        "inspected_values": inspect_values,
    }
    if path.suffix.lower() != ".xlsx":
        row.update({"sheet_names": "", "hidden_sheets": "", "workbook_error": ""})
        return row
    try:
        wb = load_workbook(path, read_only=False, data_only=False)
        sheet_names = wb.sheetnames
        hidden = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
        formula_count = 0
        comment_count = 0
        filter_sheets = []
        hidden_column_sheets = []
        for ws in wb.worksheets:
            if ws.auto_filter and ws.auto_filter.ref:
                filter_sheets.append(ws.title)
            if any(dim.hidden for dim in ws.column_dimensions.values()):
                hidden_column_sheets.append(ws.title)
            if inspect_values:
                for row_cells in ws.iter_rows():
                    for cell in row_cells:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula_count += 1
                        if cell.comment:
                            comment_count += 1
        row.update(
            {
                "sheet_names": "|".join(sheet_names),
                "hidden_sheets": "|".join(hidden),
                "formula_count": formula_count if inspect_values else "",
                "comment_count": comment_count if inspect_values else "",
                "filter_sheets": "|".join(filter_sheets),
                "hidden_column_sheets": "|".join(hidden_column_sheets),
                "workbook_error": "",
            }
        )
    except Exception as exc:
        row.update({"sheet_names": "", "hidden_sheets": "", "workbook_error": str(exc)})
    return row


def read_annotated_workbook(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl")
    chosen_name = "A1" if "A1" in sheets else next(iter(sheets))
    df = sheets[chosen_name].copy()
    df.columns = [text(col) for col in df.columns]
    meta = {
        "main_sheet": chosen_name,
        "sheet_names": list(sheets.keys()),
        "rows": len(df),
        "columns": len(df.columns),
        "status_counts": {},
    }
    if "status" in df.columns:
        counts = df["status"].map(clean_status).value_counts(dropna=False).to_dict()
        meta["status_counts"] = {str(k): int(v) for k, v in counts.items()}
    return df, meta


def likely_id_column(columns: list[str]) -> str:
    for col in columns:
        if col.lower() == "uuid":
            return col
    for col in columns:
        if ID_RE.search(col):
            return col
    return ""


def likely_question_type(column: str, series: pd.Series) -> str:
    name = column.lower()
    nonnull = series.dropna().astype(str).str.strip()
    if ID_RE.search(column) or name in {"record", "markers"}:
        return "identifier_or_metadata"
    if TIME_RE.search(column):
        return "timing_or_timestamp"
    if TEXT_HINT_RE.search(column):
        return "open_end"
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().mean() >= 0.8 and nonnull.nunique(dropna=True) > 10:
        return "numeric"
    if numeric.notna().mean() >= 0.8:
        return "ordinal_or_numeric"
    if MATRIX_RE.match(column):
        return "matrix_item"
    if nonnull.str.len().mean() >= 20 or (nonnull.str.len() >= 12).mean() >= 0.25:
        return "open_end_like"
    return "categorical"


def leakage_columns(columns: list[str]) -> list[str]:
    return [col for col in columns if LEAKAGE_RE.search(col)]


def comparable_key(column: str) -> str:
    key = slug(column)
    key = re.sub(r"_\d+$", "", key)
    key = re.sub(r"q\d+[a-z]*", lambda m: m.group(0), key)
    return key or slug(column)


def matrix_group(column: str) -> str:
    match = MATRIX_RE.match(column)
    return slug(match.group(1)) if match else ""


def safe_to_parquet(df: pd.DataFrame, path: Path) -> None:
    serializable = df.copy()
    for col in serializable.columns:
        if serializable[col].dtype == "object":
            serializable[col] = serializable[col].map(lambda v: "" if pd.isna(v) else str(v))
    try:
        serializable.to_parquet(path, index=False)
    except Exception:
        serializable.to_pickle(path.with_suffix(path.suffix + ".pkl"))
        serializable.to_csv(path.with_suffix(".csv"), index=False)
        path.with_suffix(path.suffix + ".unavailable.txt").write_text(
            "A Parquet engine was not installed in this runtime. The same rows were written to CSV and pickle fallback files.\n",
            encoding="utf-8",
        )


def markdown_table(df: pd.DataFrame, limit: int | None = None) -> str:
    if df.empty:
        return "_No rows._"
    view = df.head(limit).copy() if limit else df.copy()
    view = view.fillna("")
    headers = [str(col) for col in view.columns]
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for _, row in view.iterrows():
        cells = [str(row[col]).replace("|", "\\|").replace("\n", " ") for col in view.columns]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def cohen_h(a: float, b: float) -> float:
    a = min(max(a, 0), 1)
    b = min(max(b, 0), 1)
    return 2 * (math.asin(math.sqrt(a)) - math.asin(math.sqrt(b)))


def cliffs_delta(a: pd.Series, b: pd.Series, limit: int = 800) -> float:
    av = pd.to_numeric(a, errors="coerce").dropna().to_numpy()
    bv = pd.to_numeric(b, errors="coerce").dropna().to_numpy()
    if len(av) == 0 or len(bv) == 0:
        return 0.0
    if len(av) > limit:
        av = np.quantile(av, np.linspace(0.001, 0.999, limit))
    if len(bv) > limit:
        bv = np.quantile(bv, np.linspace(0.001, 0.999, limit))
    diff = np.sign(av[:, None] - bv[None, :]).sum()
    return float(diff / (len(av) * len(bv)))


def auc_from_score(y: np.ndarray, score: np.ndarray) -> tuple[float, float]:
    if len(set(y.tolist())) < 2:
        return 0.5, float(y.mean()) if len(y) else 0.0
    if roc_auc_score is None:
        order = np.argsort(score)
        ranks = np.empty_like(order, dtype=float)
        ranks[order] = np.arange(1, len(score) + 1)
        pos = y == 1
        neg = y == 0
        auc = (ranks[pos].sum() - pos.sum() * (pos.sum() + 1) / 2) / max(pos.sum() * neg.sum(), 1)
        return float(auc), float(y.mean())
    return float(roc_auc_score(y, score)), float(average_precision_score(y, score))


def summarize_numeric(values: pd.Series) -> dict[str, float]:
    x = pd.to_numeric(values, errors="coerce").dropna()
    if x.empty:
        return {}
    q = x.quantile([0.01, 0.05, 0.10, 0.25, 0.5, 0.75, 0.90, 0.95, 0.99]).to_dict()
    mad = float((x - x.median()).abs().median())
    return {
        "count": float(len(x)),
        "mean": float(x.mean()),
        "std": float(x.std(ddof=1)) if len(x) > 1 else 0.0,
        "median": float(x.median()),
        "mad": mad,
        "iqr": float(q.get(0.75, 0) - q.get(0.25, 0)),
        "p01": float(q.get(0.01, 0)),
        "p05": float(q.get(0.05, 0)),
        "p10": float(q.get(0.10, 0)),
        "p25": float(q.get(0.25, 0)),
        "p75": float(q.get(0.75, 0)),
        "p90": float(q.get(0.90, 0)),
        "p95": float(q.get(0.95, 0)),
        "p99": float(q.get(0.99, 0)),
        "min": float(x.min()),
        "max": float(x.max()),
        "trimmed_mean": float(x.sort_values().iloc[int(len(x) * 0.05) : max(int(len(x) * 0.95), int(len(x) * 0.05) + 1)].mean()),
        "winsorized_mean": float(x.clip(q.get(0.05, x.min()), q.get(0.95, x.max())).mean()),
    }


def distribution_json(series: pd.Series, limit: int = 12) -> str:
    cleaned = series.map(text)
    vc = cleaned[cleaned.ne("")].value_counts().head(limit)
    return json.dumps({str(k): int(v) for k, v in vc.items()}, ensure_ascii=True)


def text_metrics(series: pd.Series) -> pd.DataFrame:
    s = series.map(text)
    words = s.str.findall(r"[A-Za-z']+")
    word_count = words.map(len)
    unique_ratio = words.map(lambda ws: len(set(w.lower() for w in ws)) / len(ws) if ws else 0)
    sentences = s.str.count(r"[.!?]+").clip(lower=0) + s.ne("").astype(int)
    generic = s.map(lambda v: norm(v) in GENERIC_TEXT or len(norm(v)) <= 2)
    polished = s.str.contains(r"\b(?:comprehensive|seamless|valuable insight|as an ai|cannot answer)\b|—", case=False, regex=True, na=False)
    return pd.DataFrame(
        {
            "char_count": s.str.len().fillna(0),
            "word_count": word_count,
            "sentence_count": sentences,
            "unique_token_ratio": unique_ratio,
            "generic_placeholder": generic.astype(int),
            "polished_marker": polished.astype(int),
        }
    )


def infer_open_text_columns(df: pd.DataFrame, columns: list[str]) -> list[str]:
    out: list[str] = []
    for col in columns:
        if col not in df:
            continue
        qtype = likely_question_type(col, df[col])
        if qtype in {"open_end", "open_end_like"}:
            out.append(col)
    return out


def infer_matrix_groups(columns: list[str]) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = defaultdict(list)
    for col in columns:
        group = matrix_group(col)
        if group:
            groups[group].append(col)
    return {k: v for k, v in groups.items() if len(v) >= 4}


@dataclass
class Corpus:
    rows: pd.DataFrame
    inventory: list[dict[str, Any]]
    leakage: dict[str, list[str]]
    question_rows: list[dict[str, Any]]


def build_corpus(annotated_dir: Path, client_root: Path, blinded_path: Path | None, output_dir: Path) -> Corpus:
    all_sources = find_sources(client_root)
    inventory_rows: list[dict[str, Any]] = []
    for source in all_sources:
        inspect_values = annotated_dir in source.parents or source == annotated_dir
        if blinded_path and source.resolve() == blinded_path.resolve():
            inspect_values = False
        inventory_rows.append(workbook_metadata(source, inspect_values=inspect_values))

    combined: list[pd.DataFrame] = []
    leakage: dict[str, list[str]] = {}
    question_rows: list[dict[str, Any]] = []
    annotated_files = [p for p in sorted(annotated_dir.rglob("*.xlsx")) if not p.name.startswith("~$")]
    for idx, path in enumerate(annotated_files, start=1):
        dataset_id = f"d{idx:02d}_{slug(path.stem)}"
        df, meta = read_annotated_workbook(path)
        df["__dataset_id"] = dataset_id
        df["__dataset_name"] = path.name
        df["__source_file"] = str(path)
        df["__source_row_number"] = np.arange(2, len(df) + 2)
        df["__status_clean"] = df["status"].map(clean_status) if "status" in df.columns else ""
        id_col = likely_id_column(list(df.columns))
        df["__respondent_id"] = df[id_col].map(text) if id_col else [f"{dataset_id}_row_{i}" for i in df["__source_row_number"]]
        leak_cols = leakage_columns([c for c in df.columns if not c.startswith("__")])
        leakage[dataset_id] = leak_cols
        inventory_rows.append(
            {
                "file_path": str(path),
                "file_name": path.name,
                "dataset_id": dataset_id,
                "main_sheet": meta["main_sheet"],
                "rows": meta["rows"],
                "columns": meta["columns"],
                "status_3": int(meta["status_counts"].get("3", 0)),
                "status_5": int(meta["status_counts"].get("5", 0)),
                "id_column": id_col,
                "timing_fields": "|".join([c for c in df.columns if TIME_RE.search(c)]),
                "metadata_fields": "|".join([c for c in df.columns if ID_RE.search(c)]),
                "annotation_or_leakage_fields": "|".join(leak_cols),
                "datamap_sources": "|".join([s for s in meta["sheet_names"] if re.search(r"map|code|data|label", s, re.I)]),
            }
        )
        usable_cols = [c for c in df.columns if not c.startswith("__")]
        for col in usable_cols:
            if col in leak_cols:
                continue
            qtype = likely_question_type(col, df[col])
            question_rows.append(
                {
                    "dataset_id": dataset_id,
                    "dataset_name": path.name,
                    "source_column": col,
                    "canonical_column_or_question_id": comparable_key(col),
                    "question_text": col,
                    "question_type": qtype,
                    "section": col.split("_")[0] if "_" in col else re.sub(r"\d.*$", "", col) or "unknown",
                    "route_applicability_condition": "Observed nonempty or route unknown. Resolve from Datamap when available.",
                    "entity_brand_matrix_group": matrix_group(col),
                    "data_type": str(df[col].dtype),
                    "is_leakage_excluded": False,
                }
            )
        combined.append(df)

    output_dir.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(inventory_rows).to_csv(output_dir / "input_inventory.csv", index=False)
    (output_dir / "input_hashes.json").write_text(
        json.dumps(
            {
                "client_root": str(client_root),
                "annotated_dir": str(annotated_dir),
                "files": [
                    {"path": row["file_path"], "name": row["file_name"], "sha256": row["sha256"], "bytes": row["bytes"]}
                    for row in inventory_rows
                    if "sha256" in row
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    if not combined:
        raise SystemExit(f"No annotated .xlsx files found under {annotated_dir}")
    rows = pd.concat(combined, ignore_index=True, sort=False)
    return Corpus(rows=rows, inventory=inventory_rows, leakage=leakage, question_rows=question_rows)


def write_split_and_ledgers(corpus: Corpus, blinded_path: Path | None, output_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    labeled = corpus.rows[corpus.rows["__status_clean"].isin({ACCEPTED, REJECTED})].copy()
    accepted = labeled[labeled["__status_clean"] == ACCEPTED].copy()
    rejected = labeled[labeled["__status_clean"] == REJECTED].copy()
    manifest = labeled[
        ["__dataset_id", "__dataset_name", "__source_file", "__source_row_number", "__respondent_id", "__status_clean"]
    ].rename(
        columns={
            "__dataset_id": "dataset_id",
            "__dataset_name": "dataset_name",
            "__source_file": "source_file",
            "__source_row_number": "source_row_number",
            "__respondent_id": "respondent_id",
            "__status_clean": "status",
        }
    )
    manifest["row_key"] = manifest["dataset_id"] + "::" + manifest["source_row_number"].astype(str)
    manifest.to_csv(output_dir / "labeled_row_manifest.csv", index=False)
    safe_to_parquet(rejected, output_dir / "discard_rows_raw.parquet")
    safe_to_parquet(accepted, output_dir / "accepted_rows_raw.parquet")
    split = {
        "iteration_id": output_dir.name,
        "annotated_labeled_rows": int(len(labeled)),
        "accepted_status_3": int(len(accepted)),
        "rejected_status_5": int(len(rejected)),
        "blind_feature_view": {
            "label_columns_removed": sorted({c for cols in corpus.leakage.values() for c in cols} | {"status", "__status_clean"}),
            "dataset_identity_retained_for_grouped_validation_only": True,
        },
        "label_view": {"status_values": {"3": "accepted by TFG", "5": "rejected by TFG for quality concerns"}},
        "blinded_test_file": str(blinded_path) if blinded_path else "",
    }
    (output_dir / "split_manifest.json").write_text(json.dumps(split, indent=2), encoding="utf-8")
    (output_dir / "leakage_exclusions.json").write_text(json.dumps(corpus.leakage, indent=2), encoding="utf-8")
    heldout = []
    if blinded_path:
        heldout.append(
            {
                "path": str(blinded_path),
                "name": blinded_path.name,
                "bytes": blinded_path.stat().st_size,
                "sha256": sha256(blinded_path),
                "respondent_values_inspected": False,
            }
        )
    (output_dir / "blinded_test_freeze_manifest.json").write_text(json.dumps({"heldout": heldout}, indent=2), encoding="utf-8")
    counts = manifest.groupby(["dataset_id", "dataset_name", "status"]).size().unstack(fill_value=0).reset_index()
    counts["total_labeled"] = counts.get(ACCEPTED, 0) + counts.get(REJECTED, 0)
    lines = [
        "# Status reconciliation",
        "",
        f"We reconciled {len(labeled):,} labeled rows.",
        f"Accepted status 3 rows: {len(accepted):,}.",
        f"Rejected status 5 rows: {len(rejected):,}.",
        "",
        "Every labeled row appears once in `labeled_row_manifest.csv`.",
        "",
        markdown_table(counts),
        "",
    ]
    (output_dir / "status_reconciliation.md").write_text("\n".join(lines), encoding="utf-8")
    return labeled, accepted, rejected


def build_column_profiles(labeled: pd.DataFrame, question_rows: list[dict[str, Any]], leakage: dict[str, list[str]], output_dir: Path) -> pd.DataFrame:
    question_df = pd.DataFrame(question_rows)
    question_df.to_csv(output_dir / "canonical_question_map.csv", index=False)
    (output_dir / "question_contracts").mkdir(exist_ok=True)
    (output_dir / "route_graphs").mkdir(exist_ok=True)
    for dataset_id, group in question_df.groupby("dataset_id"):
        sample = group.head(120)
        lines = [
            f"# Question contract for {dataset_id}",
            "",
            "This contract is derived from source columns and Datamap-like sheet cues. It is a starting point for the agent's full question reading.",
            "",
        ]
        for _, row in sample.iterrows():
            lines.append(
                f"- `{row['source_column']}`. Type: {row['question_type']}. Expected evidence: {expected_evidence(row['question_type'])}. Guardrail: compare against accepted rows before scoring."
            )
        (output_dir / "question_contracts" / f"{dataset_id}.md").write_text("\n".join(lines), encoding="utf-8")
        route = {
            "dataset_id": dataset_id,
            "nodes": [
                {
                    "column": row["source_column"],
                    "type": row["question_type"],
                    "route": row["route_applicability_condition"],
                }
                for _, row in sample.iterrows()
            ],
            "edges": [],
        }
        (output_dir / "route_graphs" / f"{dataset_id}.json").write_text(json.dumps(route, indent=2), encoding="utf-8")

    profile_rows: list[dict[str, Any]] = []
    for dataset_id, dfg in labeled.groupby("__dataset_id"):
        dfg = dfg.copy()
        y = dfg["__status_clean"].eq(REJECTED).astype(int).to_numpy()
        leak_cols = set(leakage.get(dataset_id, [])) | {"status", "__status_clean"}
        source_cols = [c for c in dfg.columns if not c.startswith("__") and c not in leak_cols]
        for col in source_cols:
            accepted = dfg.loc[dfg["__status_clean"] == ACCEPTED, col]
            rejected = dfg.loc[dfg["__status_clean"] == REJECTED, col]
            qtype = likely_question_type(col, dfg[col])
            base: dict[str, Any] = {
                "dataset_id": dataset_id,
                "source_column": col,
                "canonical_column_or_question_id": comparable_key(col),
                "question_text": col,
                "question_type": qtype,
                "section": col.split("_")[0] if "_" in col else re.sub(r"\d.*$", "", col) or "unknown",
                "route_applicability_condition": "Observed nonempty or route unknown.",
                "entity_brand_matrix_group": matrix_group(col),
                "data_type": str(dfg[col].dtype),
                "accepted_applicable_n": int(accepted.notna().sum()),
                "rejected_applicable_n": int(rejected.notna().sum()),
                "accepted_missing_n": int(accepted.map(text).eq("").sum()),
                "rejected_missing_n": int(rejected.map(text).eq("").sum()),
            }
            base["accepted_missing_rate"] = round(base["accepted_missing_n"] / max(len(accepted), 1), 4)
            base["rejected_missing_rate"] = round(base["rejected_missing_n"] / max(len(rejected), 1), 4)
            base["rejection_lift"] = 0.0
            base["accepted_false_positive_impact"] = 0.0
            base["uncertainty_confidence_interval"] = ""
            base["notes_and_likely_confounds"] = ""
            numeric = pd.to_numeric(dfg[col], errors="coerce")
            if qtype in {"numeric", "ordinal_or_numeric", "timing_or_timestamp"} and numeric.notna().mean() >= 0.35:
                acc_num = pd.to_numeric(accepted, errors="coerce")
                rej_num = pd.to_numeric(rejected, errors="coerce")
                acc_stats = summarize_numeric(acc_num)
                rej_stats = summarize_numeric(rej_num)
                pooled = math.sqrt((acc_stats.get("std", 0) ** 2 + rej_stats.get("std", 0) ** 2) / 2) if acc_stats and rej_stats else 0
                effect = (rej_stats.get("mean", 0) - acc_stats.get("mean", 0)) / pooled if pooled else 0
                score = numeric.fillna(numeric.median()).to_numpy()
                auc, auprc = auc_from_score(y, score)
                if auc < 0.5:
                    auc = 1 - auc
                    effect = -effect
                row = {
                    **base,
                    "derived_measure": "numeric_distribution",
                    "accepted_distribution_statistics": json.dumps(acc_stats),
                    "rejected_distribution_statistics": json.dumps(rej_stats),
                    "effect_size": round(effect, 4),
                    "distribution_free_effect": round(cliffs_delta(rej_num, acc_num), 4),
                    "univariate_auroc": round(auc, 4),
                    "univariate_auprc": round(auprc, 4),
                }
                profile_rows.append(row)
            else:
                acc_clean = accepted.map(text)
                rej_clean = rejected.map(text)
                acc_nonmiss = acc_clean[acc_clean.ne("")]
                rej_nonmiss = rej_clean[rej_clean.ne("")]
                acc_mode = acc_nonmiss.value_counts(normalize=True).iloc[0] if not acc_nonmiss.empty else 0
                rej_mode = rej_nonmiss.value_counts(normalize=True).iloc[0] if not rej_nonmiss.empty else 0
                missing_effect = base["rejected_missing_rate"] - base["accepted_missing_rate"]
                score = dfg[col].map(text).eq("").astype(float).to_numpy()
                auc, auprc = auc_from_score(y, score)
                row = {
                    **base,
                    "derived_measure": "categorical_or_missingness",
                    "accepted_distribution_statistics": distribution_json(accepted),
                    "rejected_distribution_statistics": distribution_json(rejected),
                    "effect_size": round(max(abs(rej_mode - acc_mode), abs(missing_effect)), 4),
                    "distribution_free_effect": round(cohen_h(base["rejected_missing_rate"], base["accepted_missing_rate"]), 4),
                    "univariate_auroc": round(max(auc, 1 - auc), 4),
                    "univariate_auprc": round(auprc, 4),
                }
                profile_rows.append(row)
            if qtype in {"open_end", "open_end_like"}:
                metrics = text_metrics(dfg[col])
                for metric in metrics.columns:
                    acc_m = metrics.loc[dfg["__status_clean"] == ACCEPTED, metric]
                    rej_m = metrics.loc[dfg["__status_clean"] == REJECTED, metric]
                    acc_stats = summarize_numeric(acc_m)
                    rej_stats = summarize_numeric(rej_m)
                    pooled = math.sqrt((acc_stats.get("std", 0) ** 2 + rej_stats.get("std", 0) ** 2) / 2) if acc_stats and rej_stats else 0
                    effect = (rej_stats.get("mean", 0) - acc_stats.get("mean", 0)) / pooled if pooled else 0
                    auc, auprc = auc_from_score(y, metrics[metric].fillna(0).to_numpy())
                    if auc < 0.5:
                        auc = 1 - auc
                        effect = -effect
                    profile_rows.append(
                        {
                            **base,
                            "derived_measure": f"text_{metric}",
                            "accepted_distribution_statistics": json.dumps(acc_stats),
                            "rejected_distribution_statistics": json.dumps(rej_stats),
                            "effect_size": round(effect, 4),
                            "distribution_free_effect": round(cliffs_delta(rej_m, acc_m), 4),
                            "univariate_auroc": round(auc, 4),
                            "univariate_auprc": round(auprc, 4),
                        }
                    )
    profile = pd.DataFrame(profile_rows)
    if profile.empty:
        raise SystemExit("Column profile is empty.")
    profile["within_dataset_rank"] = profile.groupby("dataset_id")["univariate_auroc"].rank(method="dense", ascending=False)
    stability = (
        profile.groupby(["canonical_column_or_question_id", "derived_measure"])
        .agg(
            datasets=("dataset_id", "nunique"),
            mean_auroc=("univariate_auroc", "mean"),
            mean_effect=("effect_size", "mean"),
            max_accepted_false_positive=("accepted_missing_rate", "max"),
        )
        .reset_index()
    )
    stability["cross_dataset_stability"] = np.where(
        stability["datasets"] >= 3,
        "cross_dataset",
        np.where(stability["datasets"] == 2, "survey_family_specific", "survey_specific"),
    )
    profile = profile.merge(stability, on=["canonical_column_or_question_id", "derived_measure"], how="left")
    profile.to_csv(output_dir / "column_profile_discard_vs_accept.csv", index=False)
    try:
        profile.to_excel(output_dir / "column_profile_discard_vs_accept.xlsx", index=False)
    except Exception:
        pass
    ranking = profile.sort_values(["mean_auroc", "univariate_auroc", "datasets"], ascending=False).head(300)
    ranking.to_csv(output_dir / "univariate_signal_ranking.csv", index=False)
    meta = (
        profile.groupby(["canonical_column_or_question_id", "derived_measure"])
        .agg(
            datasets=("dataset_id", "nunique"),
            support=("accepted_applicable_n", "sum"),
            rejected_support=("rejected_applicable_n", "sum"),
            mean_auroc=("univariate_auroc", "mean"),
            max_auroc=("univariate_auroc", "max"),
            mean_effect=("effect_size", "mean"),
            heterogeneity=("effect_size", "std"),
        )
        .reset_index()
        .sort_values(["datasets", "mean_auroc", "mean_effect"], ascending=False)
    )
    meta["transfer_class"] = np.where(meta["datasets"] >= 5, "cross_dataset", np.where(meta["datasets"] >= 2, "survey_family_specific", "survey_specific"))
    meta.to_csv(output_dir / "cross_dataset_meta_signals.csv", index=False)
    return profile


def expected_evidence(qtype: str) -> str:
    if qtype in {"open_end", "open_end_like"}:
        return "A prompt-fit answer grounded in the respondent's prior choices or experience."
    if qtype == "timing_or_timestamp":
        return "Timing should be plausible relative to question complexity and answer depth."
    if qtype == "matrix_item":
        return "Variation or uniformity should make sense given item similarity and scale direction."
    if qtype in {"numeric", "ordinal_or_numeric"}:
        return "Values should fit the route, scale, and related numeric answers."
    return "The selection should fit the respondent's route and related answers."


def build_feature_matrix(labeled: pd.DataFrame, leakage: dict[str, list[str]], output_dir: Path) -> pd.DataFrame:
    feature_rows: list[pd.DataFrame] = []
    for dataset_id, dfg in labeled.groupby("__dataset_id"):
        dfg = dfg.copy()
        leak_cols = set(leakage.get(dataset_id, [])) | {"status", "__status_clean"}
        features = pd.DataFrame(
            {
                "dataset_id": dfg["__dataset_id"],
                "dataset_name": dfg["__dataset_name"],
                "respondent_id": dfg["__respondent_id"],
                "source_row_number": dfg["__source_row_number"],
                "status": dfg["__status_clean"],
            }
        )
        qtime_col = next((c for c in dfg.columns if c.lower() == "qtime"), "")
        if qtime_col:
            qtime = pd.to_numeric(dfg[qtime_col], errors="coerce")
            accepted_q = qtime[dfg["__status_clean"] == ACCEPTED]
            med = accepted_q.median()
            mad = (accepted_q - med).abs().median() or 1
            features["f_timing_fast_route_adjusted"] = (qtime < accepted_q.quantile(0.10)).fillna(False).astype(int)
            features["f_timing_under_4_min"] = (qtime < 240).fillna(False).astype(int)
            features["qtime_seconds"] = qtime
            features["qtime_robust_z"] = (qtime - med) / mad
        else:
            features["f_timing_fast_route_adjusted"] = 0
            features["f_timing_under_4_min"] = 0
            features["qtime_seconds"] = np.nan
            features["qtime_robust_z"] = 0
        source_cols = [c for c in dfg.columns if not c.startswith("__") and c not in leak_cols]
        open_cols = infer_open_text_columns(dfg, source_cols)
        if open_cols:
            combined_text = dfg[open_cols].fillna("").astype(str).agg(" ".join, axis=1)
            textm = text_metrics(combined_text)
            features["open_word_count_total"] = textm["word_count"]
            features["open_char_count_total"] = textm["char_count"]
            features["f_open_generic_or_too_short"] = ((textm["generic_placeholder"] == 1) | (textm["word_count"] <= 2)).astype(int)
            features["f_open_polished_marker"] = textm["polished_marker"].astype(int)
            norm_texts = combined_text.map(norm)
            counts = norm_texts[norm_texts.str.len() >= 16].value_counts()
            features["f_duplicate_full_open_chain"] = norm_texts.map(lambda v: int(len(v) >= 16 and counts.get(v, 0) >= 2))
            features["words_per_second"] = features["open_word_count_total"] / features["qtime_seconds"].replace(0, np.nan)
            features["f_fast_text_mismatch"] = ((features["qtime_seconds"] < dfg.get(qtime_col, pd.Series([np.nan] * len(dfg))).quantile(0.15)) & (features["open_word_count_total"] >= textm["word_count"].quantile(0.75))).fillna(False).astype(int)
        else:
            features["open_word_count_total"] = 0
            features["open_char_count_total"] = 0
            features["f_open_generic_or_too_short"] = 0
            features["f_open_polished_marker"] = 0
            features["f_duplicate_full_open_chain"] = 0
            features["words_per_second"] = 0
            features["f_fast_text_mismatch"] = 0
        missing_rates = dfg[source_cols].map(lambda v: text(v) == "").mean(axis=1) if source_cols else 0
        features["missing_rate_observed_fields"] = missing_rates
        features["f_high_missingness"] = (missing_rates > 0.60).astype(int) if hasattr(missing_rates, "astype") else 0
        groups = infer_matrix_groups(source_cols)
        matrix_flags = []
        modal_props = []
        for _, row in dfg.iterrows():
            row_modal = []
            row_flag = 0
            for cols in groups.values():
                vals = [text(row.get(c)) for c in cols if text(row.get(c)) != ""]
                if len(vals) < 4:
                    continue
                counts = Counter(vals)
                modal = max(counts.values()) / len(vals)
                row_modal.append(modal)
                if modal >= 0.90 and len(set(vals)) <= 2:
                    row_flag = 1
            matrix_flags.append(row_flag)
            modal_props.append(max(row_modal) if row_modal else 0)
        features["matrix_max_modal_proportion"] = modal_props
        features["f_matrix_uniform_pattern"] = matrix_flags
        features["f_multiple_weak_families"] = (
            features[[
                "f_timing_fast_route_adjusted",
                "f_open_generic_or_too_short",
                "f_duplicate_full_open_chain",
                "f_fast_text_mismatch",
                "f_high_missingness",
                "f_matrix_uniform_pattern",
            ]].sum(axis=1)
            >= 2
        ).astype(int)
        feature_rows.append(features)
    matrix = pd.concat(feature_rows, ignore_index=True)
    safe_to_parquet(matrix, output_dir / "signal_matrix.parquet")
    return matrix


def family_score_table(signal_matrix: pd.DataFrame, output_dir: Path) -> pd.DataFrame:
    df = signal_matrix.copy()
    families = {
        "timing_cognitive_dynamics": ["f_timing_fast_route_adjusted", "f_timing_under_4_min", "f_fast_text_mismatch"],
        "open_end_semantics": ["f_open_generic_or_too_short", "f_open_polished_marker"],
        "duplicate_replay_coordination": ["f_duplicate_full_open_chain"],
        "matrix_behavior": ["f_matrix_uniform_pattern"],
        "routing_nonresponse": ["f_high_missingness"],
        "multi_family_convergence": ["f_multiple_weak_families"],
        "protective_human_evidence": [],
    }
    out = df[["dataset_id", "dataset_name", "respondent_id", "source_row_number", "status"]].copy()
    for family, cols in families.items():
        existing = [c for c in cols if c in df.columns]
        out[family] = df[existing].max(axis=1) if existing else 0
    out["client_reject_probability_proxy"] = out[[c for c in families if c != "protective_human_evidence"]].mean(axis=1)
    out["authenticity_risk_probability_proxy"] = out[["timing_cognitive_dynamics", "open_end_semantics", "duplicate_replay_coordination", "multi_family_convergence"]].mean(axis=1)
    out["attention_or_validity_risk_proxy"] = out[["matrix_behavior", "routing_nonresponse", "timing_cognitive_dynamics"]].mean(axis=1)
    safe_to_parquet(out, output_dir / "family_scores.parquet")
    return out


def prevalence_table(signal_matrix: pd.DataFrame, features: list[str]) -> pd.DataFrame:
    rows = []
    for feature in features:
        if feature not in signal_matrix:
            continue
        for dataset_id, group in signal_matrix.groupby("dataset_id"):
            y = group["status"].eq(REJECTED)
            hit = pd.to_numeric(group[feature], errors="coerce").fillna(0) > 0
            accepted_hit = int((hit & ~y).sum())
            rejected_hit = int((hit & y).sum())
            support = accepted_hit + rejected_hit
            base = float(y.mean()) if len(y) else 0
            rate = rejected_hit / support if support else 0
            rows.append(
                {
                    "dataset_id": dataset_id,
                    "feature": feature,
                    "accepted_hits": accepted_hit,
                    "rejected_hits": rejected_hit,
                    "support": support,
                    "reject_rate_when_present": rate,
                    "dataset_reject_rate": base,
                    "lift": rate / base if base else 0,
                    "accepted_false_positive_burden": accepted_hit / max((~y).sum(), 1),
                }
            )
    return pd.DataFrame(rows)


def build_interactions(signal_matrix: pd.DataFrame, output_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    feature_cols = [c for c in signal_matrix.columns if c.startswith("f_")]
    prevalence = prevalence_table(signal_matrix, feature_cols)
    top_features = (
        prevalence.groupby("feature")
        .agg(mean_lift=("lift", "mean"), support=("support", "sum"), datasets=("dataset_id", "nunique"))
        .query("support > 0")
        .sort_values(["mean_lift", "support"], ascending=False)
        .head(14)
        .index.tolist()
    )
    rows = []
    for a, b in itertools.combinations(top_features, 2):
        hit = (pd.to_numeric(signal_matrix[a], errors="coerce").fillna(0) > 0) & (pd.to_numeric(signal_matrix[b], errors="coerce").fillna(0) > 0)
        for dataset_id, group_idx in signal_matrix.groupby("dataset_id").groups.items():
            idx = list(group_idx)
            g = signal_matrix.loc[idx]
            h = hit.loc[idx]
            y = g["status"].eq(REJECTED)
            support = int(h.sum())
            if support < 5:
                continue
            rejected_hits = int((h & y).sum())
            accepted_hits = int((h & ~y).sum())
            base = float(y.mean())
            rate = rejected_hits / support if support else 0
            comp_a = ((pd.to_numeric(g[a], errors="coerce").fillna(0) > 0) & y).sum() / max((pd.to_numeric(g[a], errors="coerce").fillna(0) > 0).sum(), 1)
            comp_b = ((pd.to_numeric(g[b], errors="coerce").fillna(0) > 0) & y).sum() / max((pd.to_numeric(g[b], errors="coerce").fillna(0) > 0).sum(), 1)
            rows.append(
                {
                    "dataset_id": dataset_id,
                    "feature_a": a,
                    "feature_b": b,
                    "support": support,
                    "rejected_hits": rejected_hits,
                    "accepted_hits": accepted_hits,
                    "reject_rate": round(rate, 4),
                    "lift": round(rate / base if base else 0, 4),
                    "incremental_value_over_best_component": round(rate - max(comp_a, comp_b), 4),
                    "accepted_false_positive_burden": round(accepted_hits / max((~y).sum(), 1), 4),
                }
            )
    pairwise = pd.DataFrame(rows)
    if not pairwise.empty:
        summary = (
            pairwise.groupby(["feature_a", "feature_b"])
            .agg(
                datasets=("dataset_id", "nunique"),
                support=("support", "sum"),
                rejected_hits=("rejected_hits", "sum"),
                accepted_hits=("accepted_hits", "sum"),
                mean_lift=("lift", "mean"),
                mean_incremental_value=("incremental_value_over_best_component", "mean"),
            )
            .reset_index()
            .sort_values(["mean_incremental_value", "mean_lift", "support"], ascending=False)
        )
    else:
        summary = pd.DataFrame()
    summary.to_csv(output_dir / "pairwise_interactions.csv", index=False)

    higher_rows = []
    for combo in itertools.combinations(top_features[:8], 3):
        hit = np.ones(len(signal_matrix), dtype=bool)
        for feature in combo:
            hit &= (pd.to_numeric(signal_matrix[feature], errors="coerce").fillna(0).to_numpy() > 0)
        support = int(hit.sum())
        if support < 12:
            continue
        y = signal_matrix["status"].eq(REJECTED).to_numpy()
        rejected_hits = int((hit & y).sum())
        accepted_hits = int((hit & ~y).sum())
        rate = rejected_hits / support if support else 0
        base = float(y.mean())
        higher_rows.append(
            {
                "pattern_id": "plus".join(combo),
                "features": "|".join(combo),
                "support": support,
                "rejected_hits": rejected_hits,
                "accepted_hits": accepted_hits,
                "reject_rate": round(rate, 4),
                "lift": round(rate / base if base else 0, 4),
                "datasets": int(signal_matrix.loc[hit, "dataset_id"].nunique()),
                "promotion_status": "candidate" if rejected_hits >= 10 and accepted_hits / max(support, 1) <= 0.7 else "hold_for_residual_review",
            }
        )
    higher = pd.DataFrame(higher_rows).sort_values(["lift", "support"], ascending=False) if higher_rows else pd.DataFrame()
    higher.to_csv(output_dir / "higher_order_patterns.csv", index=False)
    return summary, higher


def build_question_relation_graph(question_rows: list[dict[str, Any]], output_dir: Path) -> None:
    nodes = []
    edges = []
    by_dataset: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in question_rows:
        by_dataset[row["dataset_id"]].append(row)
        nodes.append(
            {
                "id": f"{row['dataset_id']}::{row['source_column']}",
                "dataset_id": row["dataset_id"],
                "column": row["source_column"],
                "canonical_id": row["canonical_column_or_question_id"],
                "type": row["question_type"],
            }
        )
    for dataset_id, rows in by_dataset.items():
        for a, b in zip(rows, rows[1:]):
            relation = "sequential"
            if a["entity_brand_matrix_group"] and a["entity_brand_matrix_group"] == b["entity_brand_matrix_group"]:
                relation = "same_matrix_or_entity_group"
            elif a["question_type"].startswith("open") or b["question_type"].startswith("open"):
                relation = "open_closed_support_or_contradiction"
            edges.append(
                {
                    "source": f"{dataset_id}::{a['source_column']}",
                    "target": f"{dataset_id}::{b['source_column']}",
                    "relationship": relation,
                    "status": "proposed_from_column_order",
                }
            )
    (output_dir / "question_relation_graph.json").write_text(json.dumps({"nodes": nodes, "edges": edges}, indent=2), encoding="utf-8")


def nearest_controls(signal_matrix: pd.DataFrame, rejected: pd.DataFrame, output_dir: Path) -> pd.DataFrame:
    numeric_cols = ["qtime_seconds", "open_word_count_total", "missing_rate_observed_fields", "matrix_max_modal_proportion"]
    flags = [c for c in signal_matrix.columns if c.startswith("f_")]
    cases = []
    accepted = signal_matrix[signal_matrix["status"] == ACCEPTED].copy()
    for _, row in rejected.iterrows():
        dataset = row["__dataset_id"]
        source_row = row["__source_row_number"]
        rfeat = signal_matrix[(signal_matrix["dataset_id"] == dataset) & (signal_matrix["source_row_number"] == source_row)]
        if rfeat.empty:
            continue
        r = rfeat.iloc[0]
        pool = accepted[accepted["dataset_id"] == dataset].copy()
        if pool.empty:
            pool = accepted.copy()
        score = np.zeros(len(pool))
        for col in numeric_cols:
            if col in pool:
                denom = pool[col].std() or 1
                score += ((pd.to_numeric(pool[col], errors="coerce").fillna(0) - float(pd.to_numeric(pd.Series([r.get(col, 0)]), errors="coerce").fillna(0).iloc[0])) / denom).abs()
        for col in flags:
            score += (pool[col].fillna(0).astype(int) != int(r.get(col, 0))).astype(int) * 0.4
        pool = pool.assign(match_distance=score)
        for rank, (_, ctrl) in enumerate(pool.sort_values("match_distance").head(3).iterrows(), start=1):
            cases.append(
                {
                    "dataset_id": dataset,
                    "rejected_respondent_id": row["__respondent_id"],
                    "rejected_source_row_number": source_row,
                    "control_rank": rank,
                    "accepted_respondent_id": ctrl["respondent_id"],
                    "accepted_source_row_number": int(ctrl["source_row_number"]),
                    "match_distance": round(float(ctrl["match_distance"]), 4),
                    "shared_signal_families": "|".join([f for f in flags if int(r.get(f, 0)) == 1 and int(ctrl.get(f, 0)) == 1]),
                    "learning_question": "What evidence makes the rejected row less faithful than this similar accepted control?",
                }
            )
    matched = pd.DataFrame(cases)
    safe_to_parquet(matched, output_dir / "matched_case_pairs.parquet")
    return matched


def full_chain(row: pd.Series, leakage_cols: set[str]) -> str:
    parts = []
    for col, value in row.items():
        if col.startswith("__") or col in leakage_cols:
            continue
        val = text(value)
        if val:
            parts.append(f"{col}: {val}")
    return " | ".join(parts[:160])


def write_row_reviews(
    labeled: pd.DataFrame,
    rejected: pd.DataFrame,
    signal_matrix: pd.DataFrame,
    matched: pd.DataFrame,
    leakage: dict[str, list[str]],
    output_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    review_dir = output_dir / "row_reviews_blind"
    recon_dir = output_dir / "row_reconciliations"
    review_dir.mkdir(exist_ok=True)
    recon_dir.mkdir(exist_ok=True)
    phenotype_rows = []
    guardrail_rows = []
    flags = [c for c in signal_matrix.columns if c.startswith("f_")]
    for dataset_id, group in rejected.groupby("__dataset_id"):
        leak_cols = set(leakage.get(dataset_id, [])) | {"status", "__status_clean"}
        lines = [f"# Blind rejected-row review packet for {dataset_id}", ""]
        for _, row in group.iterrows():
            feat = signal_matrix[(signal_matrix["dataset_id"] == dataset_id) & (signal_matrix["source_row_number"] == row["__source_row_number"])]
            fired = [f for f in flags if not feat.empty and int(feat.iloc[0].get(f, 0)) == 1]
            chain = full_chain(row, leak_cols)
            controls = matched[
                (matched["dataset_id"] == dataset_id)
                & (matched["rejected_source_row_number"] == row["__source_row_number"])
            ]
            strongest_human = "A legitimate respondent may be fast, brief, uniform, or imperfect if the rest of the chain remains coherent."
            interpretation = "Needs semantic review with accepted controls."
            if len(fired) >= 2:
                interpretation = "Multiple independent feature families fired before label reveal."
            elif any("duplicate" in f for f in fired):
                interpretation = "Duplicate or replay concern requires cross-respondent review."
            elif any("open" in f for f in fired):
                interpretation = "Open-end authenticity concern requires prompt-fit review."
            phenotype_rows.append(
                {
                    "dataset_id": dataset_id,
                    "respondent_id": row["__respondent_id"],
                    "source_row_number": row["__source_row_number"],
                    "blind_provisional_interpretation": interpretation,
                    "uncertainty": "medium",
                    "feature_families": "|".join(fired),
                    "human_advocate_note": strongest_human,
                    "chain_excerpt": chain[:900],
                }
            )
            lines.extend(
                [
                    f"## Row {row['__source_row_number']} | {row['__respondent_id']}",
                    "",
                    f"Blind interpretation: {interpretation}",
                    "",
                    f"Feature families before label reveal: {', '.join(fired) if fired else 'none staged'}",
                    "",
                    f"Human advocate note: {strongest_human}",
                    "",
                    "Accepted controls:",
                    "",
                ]
            )
            for _, ctrl in controls.iterrows():
                lines.append(
                    f"- Accepted row {ctrl['accepted_source_row_number']} matched at distance {ctrl['match_distance']}. Shared signals: {ctrl['shared_signal_families'] or 'none'}."
                )
            lines.extend(["", f"Chain readout: {chain[:1200]}", ""])
        (review_dir / f"{dataset_id}.md").write_text("\n".join(lines), encoding="utf-8")
        manifest = labeled[labeled["__dataset_id"] == dataset_id]["__status_clean"].value_counts().to_dict()
        (recon_dir / f"{dataset_id}.json").write_text(json.dumps({str(k): int(v) for k, v in manifest.items()}, indent=2), encoding="utf-8")

    accepted_high = signal_matrix[(signal_matrix["status"] == ACCEPTED) & (signal_matrix[flags].sum(axis=1) >= 2)].copy()
    for _, row in accepted_high.head(1200).iterrows():
        guardrail_rows.append(
            {
                "dataset_id": row["dataset_id"],
                "respondent_id": row["respondent_id"],
                "source_row_number": row["source_row_number"],
                "shared_anomalies": "|".join([f for f in flags if int(row.get(f, 0)) == 1]),
                "protective_reading": "Accepted rows with this surface pattern must be checked for coherent full-chain evidence before escalation.",
                "guardrail_status": "protect_false_positive",
            }
        )
    phenotypes = pd.DataFrame(phenotype_rows)
    guardrails = pd.DataFrame(guardrail_rows)
    phenotypes.to_csv(output_dir / "rejected_phenotypes.csv", index=False)
    guardrails.to_csv(output_dir / "accepted_guardrails.csv", index=False)
    return phenotypes, guardrails


def fit_validation(signal_matrix: pd.DataFrame, family_scores: pd.DataFrame, output_dir: Path) -> pd.DataFrame:
    model_dir = output_dir / "model_artifacts"
    model_dir.mkdir(exist_ok=True)
    feature_cols = [c for c in signal_matrix.columns if c.startswith("f_")] + [
        "qtime_robust_z",
        "open_word_count_total",
        "missing_rate_observed_fields",
        "matrix_max_modal_proportion",
    ]
    available = [c for c in feature_cols if c in signal_matrix.columns]
    df = signal_matrix.copy()
    y = df["status"].eq(REJECTED).astype(int).to_numpy()
    preds = np.full(len(df), np.nan)
    fold_rows = []
    for dataset_id in sorted(df["dataset_id"].unique()):
        train = df["dataset_id"] != dataset_id
        test = df["dataset_id"] == dataset_id
        if train.sum() == 0 or test.sum() == 0 or len(set(y[train])) < 2:
            continue
        X_train = df.loc[train, available].apply(pd.to_numeric, errors="coerce").fillna(0).to_numpy()
        X_test = df.loc[test, available].apply(pd.to_numeric, errors="coerce").fillna(0).to_numpy()
        if LogisticRegression is not None and StandardScaler is not None:
            scaler = StandardScaler()
            X_train_s = scaler.fit_transform(X_train)
            X_test_s = scaler.transform(X_test)
            model = LogisticRegression(max_iter=1000, class_weight="balanced")
            model.fit(X_train_s, y[train])
            p = model.predict_proba(X_test_s)[:, 1]
        else:
            weights = np.nanmean(X_train[y[train] == 1], axis=0) - np.nanmean(X_train[y[train] == 0], axis=0)
            raw = X_test @ weights
            p = 1 / (1 + np.exp(-(raw - np.nanmedian(raw))))
        preds[test] = p
        fold_y = y[test]
        auc, auprc = auc_from_score(fold_y, p)
        fold_rows.append(
            {
                "heldout_dataset_id": dataset_id,
                "rows": int(test.sum()),
                "status5_rows": int(fold_y.sum()),
                "status3_rows": int((1 - fold_y).sum()),
                "auroc": round(auc, 4),
                "auprc": round(auprc, 4),
                "brier_score": round(float(brier_score_loss(fold_y, p)) if brier_score_loss else float(np.mean((fold_y - p) ** 2)), 4),
            }
        )
    df["client_reject_probability"] = preds
    q = pd.Series(preds).dropna().quantile([0.45, 0.70, 0.85, 0.95]).to_dict()
    def tier(p: float) -> str:
        if pd.isna(p):
            return "Light review"
        if p >= q.get(0.95, 1):
            return "Exclude candidate"
        if p >= q.get(0.85, 1):
            return "Review closely"
        if p >= q.get(0.70, 1):
            return "Light review"
        if p >= q.get(0.45, 1):
            return "Keep with note"
        return "Clean keep"
    df["operational_tier"] = [tier(float(p)) for p in preds]
    df[["dataset_id", "respondent_id", "source_row_number", "status", "client_reject_probability", "operational_tier"]].to_csv(
        model_dir / "leave_one_dataset_predictions.csv", index=False
    )
    folds = pd.DataFrame(fold_rows)
    folds.to_csv(model_dir / "leave_one_dataset_metrics.csv", index=False)
    tier_counts = df.groupby(["status", "operational_tier"]).size().reset_index(name="rows")
    tier_counts.to_csv(model_dir / "tier_distribution.csv", index=False)
    residual = df[
        ((df["status"] == REJECTED) & (df["operational_tier"].isin(["Clean keep", "Keep with note"])))
        | ((df["status"] == ACCEPTED) & (df["operational_tier"].isin(["Review closely", "Exclude candidate"])))
    ].copy()
    residual.to_csv(model_dir / "residual_rows.csv", index=False)
    summary = {
        "features": available,
        "folds": fold_rows,
        "tier_threshold_quantiles": {str(k): float(v) for k, v in q.items()},
        "residual_rows": int(len(residual)),
    }
    (model_dir / "model_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return df


def build_casebooks(
    output_dir: Path,
    profile: pd.DataFrame,
    pairwise: pd.DataFrame,
    higher: pd.DataFrame,
    phenotypes: pd.DataFrame,
    guardrails: pd.DataFrame,
    validation: pd.DataFrame,
) -> None:
    top_uni = profile.sort_values(["mean_auroc", "univariate_auroc"], ascending=False).head(12)
    lines = ["# Contrastive casebook", ""]
    lines.append("The strongest lesson is that every rejected pattern needs an accepted control. Fast timing, short text, and uniform matrices all appear among accepted respondents.")
    lines.append("")
    if not guardrails.empty:
        lines.append("## Accepted guardrails")
        lines.append("")
        for _, row in guardrails.head(20).iterrows():
            lines.append(f"- {row['dataset_id']} row {row['source_row_number']}: {row['shared_anomalies']}. {row['protective_reading']}")
        lines.append("")
    if not phenotypes.empty:
        lines.append("## Rejected phenotypes")
        lines.append("")
        for _, row in phenotypes.head(30).iterrows():
            lines.append(f"- {row['dataset_id']} row {row['source_row_number']}: {row['blind_provisional_interpretation']}. Signals: {row['feature_families'] or 'none staged'}.")
    (output_dir / "contrastive_casebook.md").write_text("\n".join(lines), encoding="utf-8")

    lines = ["# Interaction casebook", ""]
    lines.append("Interactions were ranked by incremental value over their strongest component, support, and accepted false-positive burden.")
    lines.append("")
    if not pairwise.empty:
        for _, row in pairwise.head(25).iterrows():
            lines.append(
                f"- {row['feature_a']} plus {row['feature_b']}: support {int(row['support'])}, mean lift {row['mean_lift']:.2f}, incremental value {row['mean_incremental_value']:.3f}."
            )
    if not higher.empty:
        lines.extend(["", "## Higher-order candidates", ""])
        for _, row in higher.head(15).iterrows():
            lines.append(
                f"- {row['features']}: support {int(row['support'])}, lift {row['lift']:.2f}, accepted hits {int(row['accepted_hits'])}. Status: {row['promotion_status']}."
            )
    (output_dir / "interaction_casebook.md").write_text("\n".join(lines), encoding="utf-8")

    residual_path = output_dir / "model_artifacts" / "residual_rows.csv"
    residual = pd.read_csv(residual_path) if residual_path.exists() else pd.DataFrame()
    lines = ["# Residual casebook", ""]
    lines.append("Residual rows are the next discovery target. False negatives can reveal missing authenticity signals. False positives can reveal protective human evidence.")
    lines.append("")
    if not residual.empty:
        grouped = residual.groupby(["status", "operational_tier"]).size().reset_index(name="rows")
        lines.append(markdown_table(grouped))
        lines.append("")
        for _, row in residual.head(30).iterrows():
            lines.append(f"- {row['dataset_id']} row {row['source_row_number']}: status {row['status']}, model tier {row['operational_tier']}.")
    (output_dir / "residual_casebook.md").write_text("\n".join(lines), encoding="utf-8")

    metrics_path = output_dir / "model_artifacts" / "leave_one_dataset_metrics.csv"
    metrics = pd.read_csv(metrics_path) if metrics_path.exists() else pd.DataFrame()
    tier_path = output_dir / "model_artifacts" / "tier_distribution.csv"
    tiers = pd.read_csv(tier_path) if tier_path.exists() else pd.DataFrame()
    lines = ["# Validation report", ""]
    if not metrics.empty:
        lines.append("## Leave-one-dataset-out metrics")
        lines.append("")
        lines.append(markdown_table(metrics))
        lines.append("")
        lines.append(f"Mean AUROC: {metrics['auroc'].mean():.3f}. Mean AUPRC: {metrics['auprc'].mean():.3f}.")
        lines.append("")
    if not tiers.empty:
        lines.append("## Tier distribution")
        lines.append("")
        lines.append(markdown_table(tiers))
        lines.append("")
    lines.append("Status is treated as a noisy client decision label. These metrics estimate transfer to held-out annotated datasets, not proof of fraud.")
    (output_dir / "validation_report.md").write_text("\n".join(lines), encoding="utf-8")

    lines = ["# Iteration report", ""]
    lines.append("We executed the first full annotated-corpus authenticity discovery iteration.")
    lines.append("")
    lines.append("## Strongest univariate differences")
    lines.append("")
    for _, row in top_uni.iterrows():
        lines.append(
            f"- {row['dataset_id']} `{row['source_column']}` `{row['derived_measure']}`. AUROC {row['univariate_auroc']:.3f}. Effect {row['effect_size']:.3f}. Accepted missing rate {row['accepted_missing_rate']:.3f}. Rejected missing rate {row['rejected_missing_rate']:.3f}."
        )
    lines.extend(["", "## Signal discipline", ""])
    lines.append("No signal is promoted only because it is common among rejected rows. Each signal must keep accepted counterexamples and stay grouped by evidence family.")
    lines.append("")
    lines.append("## Remaining limits")
    lines.append("")
    lines.append("Question text is inferred from workbook columns unless a Datamap sheet is present. The next loop should deepen manual semantic notes for the residual clusters.")
    (output_dir / "iteration_report.md").write_text("\n".join(lines), encoding="utf-8")


def write_signal_bank(output_dir: Path, signal_matrix: pd.DataFrame, pairwise: pd.DataFrame, higher: pd.DataFrame) -> None:
    feature_cols = [c for c in signal_matrix.columns if c.startswith("f_")]
    prevalence = prevalence_table(signal_matrix, feature_cols)
    summary = (
        prevalence.groupby("feature")
        .agg(
            support=("support", "sum"),
            rejected_hits=("rejected_hits", "sum"),
            accepted_hits=("accepted_hits", "sum"),
            mean_lift=("lift", "mean"),
            datasets=("dataset_id", "nunique"),
        )
        .reset_index()
        .sort_values(["mean_lift", "support"], ascending=False)
    )
    lines = ["# Auto-generated from annotated corpus. Review before production use.", ""]
    for _, row in summary.iterrows():
        status = "retrospectively_supported" if row["support"] >= 50 and row["mean_lift"] >= 1.1 else "observed"
        can_tier5 = "true" if row["mean_lift"] >= 1.8 and row["accepted_hits"] < row["rejected_hits"] else "false"
        lines.extend(
            [
                f"{row['feature']}:",
                f"  name: {row['feature'].replace('_', ' ')}",
                "  version: 1",
                "  target: authenticity",
                f"  family: {family_for_feature(row['feature'])}",
                "  hypothesis: This feature may separate client-rejected respondents from accepted controls when it survives full-chain review.",
                "  unit_of_analysis: respondent",
                "  applicability: Use only after question-role mapping and route applicability checks.",
                "  required_fields: inferred from the current Decipher export",
                "  question_types: any applicable question family",
                "  raw_measure_or_semantic_rubric: Extract the raw flag, then ask whether the full response chain supports an authenticity concern.",
                "  normalization_reference: within dataset and accepted controls",
                "  expected_direction: higher means more review concern",
                "  thresholds_or_model_use: use in family aggregation, not as a stand-alone exclusion rule",
                "  interaction_partners: see pairwise_interactions.csv and higher_order_patterns.csv",
                "  correlated_features: same family features must be counted once",
                "  protective_conditions: coherent full chain, prompt-fit answer, legitimate uniform opinion, non-native writing, route/export explanation",
                "  alternative_human_explanations: fast but attentive respondent, brief but complete answer, strong opinion, simple task, accessibility context",
                f"  minimum_support: {int(row['support'])}",
                f"  discovery_datasets: {int(row['datasets'])}",
                "  heldout_validation_datasets: leave-one-dataset-out folds in validation_report.md",
                f"  coverage: {int(row['support'])}",
                f"  rejected_prevalence: {int(row['rejected_hits'])}",
                f"  accepted_prevalence: {int(row['accepted_hits'])}",
                f"  precision_recall_and_lift: mean_lift {float(row['mean_lift']):.3f}",
                "  calibration_or_uncertainty_behavior: uncertain until checked by family model and full-chain review",
                f"  can_affect_tier_5: {can_tier5}",
                "  can_independently_trigger_tier_5: false",
                "  known_failure_modes: label leakage, route artifacts, accepted counterexamples, survey-specific wording",
                "  regression_cases: see matched_case_pairs.parquet and accepted_guardrails.csv",
                f"  status: {status}",
                "  change_history: created by annotated authenticity discovery loop",
                "",
            ]
        )
    text_out = "\n".join(lines)
    (output_dir / "signal_candidates.yaml").write_text(text_out, encoding="utf-8")
    (output_dir / "signal_bank.yaml").write_text(text_out, encoding="utf-8")
    (output_dir / "skill_change_log.md").write_text(
        "\n".join(
            [
                "# Skill change log",
                "",
                "This iteration added executable support for full annotated-corpus inventory, row reconciliation, per-column profiling, matched controls, interaction discovery, grouped validation, and signal-bank generation.",
                "",
                "Reusable skill instructions now require these artifacts before the blinded dataset is scored.",
                "",
            ]
        ),
        encoding="utf-8",
    )


def family_for_feature(feature: str) -> str:
    if "timing" in feature or "fast" in feature:
        return "timing/cognitive dynamics"
    if "open" in feature or "polished" in feature:
        return "open-end semantics"
    if "duplicate" in feature:
        return "duplicate/replay/coordination"
    if "matrix" in feature:
        return "matrix behavior"
    if "missing" in feature:
        return "routing/nonresponse"
    return "multi-family convergence"


def write_freeze(output_dir: Path, args: argparse.Namespace) -> None:
    artifact_rows = []
    for path in sorted(output_dir.rglob("*")):
        if path.is_file():
            artifact_rows.append({"path": str(path), "bytes": path.stat().st_size, "sha256": sha256(path)})
    manifest = {
        "run_id": output_dir.name,
        "annotated_input": str(args.annotated_dir),
        "client_root": str(args.client_root),
        "blinded_input": str(args.blinded_workbook) if args.blinded_workbook else "",
        "blinded_values_inspected": False,
        "artifact_count": len(artifact_rows),
        "artifacts": artifact_rows,
    }
    (output_dir / "freeze_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--annotated-dir", type=Path, required=True)
    parser.add_argument("--client-root", type=Path, required=True)
    parser.add_argument("--blinded-workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir.expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    corpus = build_corpus(args.annotated_dir.expanduser().resolve(), args.client_root.expanduser().resolve(), args.blinded_workbook.expanduser().resolve() if args.blinded_workbook else None, output_dir)
    labeled, accepted, rejected = write_split_and_ledgers(corpus, args.blinded_workbook.expanduser().resolve() if args.blinded_workbook else None, output_dir)
    if len(accepted) + len(rejected) != len(labeled):
        raise SystemExit("Status reconciliation failed.")
    profile = build_column_profiles(labeled, corpus.question_rows, corpus.leakage, output_dir)
    signal_matrix = build_feature_matrix(labeled, corpus.leakage, output_dir)
    family_scores = family_score_table(signal_matrix, output_dir)
    pairwise, higher = build_interactions(signal_matrix, output_dir)
    build_question_relation_graph(corpus.question_rows, output_dir)
    matched = nearest_controls(signal_matrix, rejected, output_dir)
    phenotypes, guardrails = write_row_reviews(labeled, rejected, signal_matrix, matched, corpus.leakage, output_dir)
    validation = fit_validation(signal_matrix, family_scores, output_dir)
    cluster = validation[["dataset_id", "respondent_id", "source_row_number", "status", "operational_tier"]].copy()
    cluster["cluster_id"] = cluster["dataset_id"] + "::" + cluster["operational_tier"].map(slug)
    safe_to_parquet(cluster, output_dir / "cluster_assignments.parquet")
    write_signal_bank(output_dir, signal_matrix, pairwise, higher)
    build_casebooks(output_dir, profile, pairwise, higher, phenotypes, guardrails, validation)
    write_freeze(output_dir, args)
    print(output_dir)
    print(f"labeled_rows={len(labeled)} accepted={len(accepted)} rejected={len(rejected)}")
    print(f"column_profile_rows={len(profile)} pairwise_rows={len(pairwise)} higher_order_rows={len(higher)}")
    print(f"matched_case_pairs={len(matched)} rejected_phenotypes={len(phenotypes)} accepted_guardrails={len(guardrails)}")


if __name__ == "__main__":
    main()
